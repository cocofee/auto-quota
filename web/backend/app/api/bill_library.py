"""
编清单 API

上传工程量Excel（算量导出/手工表格等）→ 自动匹配12位清单编码 → 下载标准工程量清单。

两种模式（由 MATCH_BACKEND 环境变量控制）：
  - local：本地直接执行（开发环境，有清单库数据）
  - remote：转发到本地匹配服务（懒猫部署，清单库在用户电脑上）

两个接口：
  - POST /bill-compiler/preview   上传Excel + 选清单版本，返回编码匹配预览
  - POST /bill-compiler/execute   上传Excel + 选清单版本，直接返回编好的Excel文件
"""

import asyncio
import tempfile
from pathlib import Path

from fastapi import APIRouter, Depends, File, UploadFile, HTTPException, Form
from fastapi.responses import FileResponse, Response
from loguru import logger

from app.auth.deps import get_current_user
from app.config import UPLOAD_MAX_MB
from app.models.user import User
from src.excel_compat import ExcelFormatInfo, validate_excel_upload
from src.output_writer import safe_excel_text

router = APIRouter()


def _sanitize_client_filename(filename: str | None, default: str = "upload.xlsx") -> str:
    raw = (filename or "").replace("\\", "/").split("/")[-1]
    raw = raw.replace("\x00", "").replace("\r", "").replace("\n", "").strip().strip(". ")
    return raw or default


async def _read_and_validate_excel(file: UploadFile, label: str) -> tuple[bytes, str, ExcelFormatInfo]:
    """读取上传文件并做格式/大小校验"""
    filename = _sanitize_client_filename(file.filename)
    max_bytes = UPLOAD_MAX_MB * 1024 * 1024
    content = bytearray()
    while chunk := await file.read(1024 * 1024):
        content.extend(chunk)
        if len(content) > max_bytes:
            raise HTTPException(status_code=400, detail=f"{label}大小超过 {UPLOAD_MAX_MB}MB 限制")
    try:
        info = validate_excel_upload(filename, bytes(content[:8]))
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return bytes(content), filename, info


async def _save_upload(content: bytes, suffix: str, prefix: str) -> str:
    """把已校验文件保存到临时目录，返回临时文件路径"""
    with tempfile.NamedTemporaryFile(
        mode="wb", suffix=suffix, delete=False, prefix=prefix,
    ) as tmp:
        tmp.write(content)
        return tmp.name


# ============================================================
# 本地模式：直接调用 BillReader + compile_items
# ============================================================

def _do_compile_local(file_path: str, bill_version: str) -> dict:
    """本地执行编清单，返回预览结果"""
    from src.bill_reader import BillReader
    from src.bill_compiler import compile_items

    reader = BillReader()
    items = reader.read_file(file_path)
    if not items:
        raise HTTPException(status_code=400, detail="未从Excel中读取到清单项，请检查文件格式。")

    compiled = compile_items(items, bill_version=bill_version)

    result_items = []
    matched_count = 0
    for i, item in enumerate(compiled):
        original_code = item.get("code", "").strip()
        bill_match = item.get("bill_match")

        if bill_match and bill_match.get("code_12"):
            code = bill_match["code_12"]
            code_source = "matched"
            matched_count += 1
        elif original_code and len(original_code) >= 9:
            code = original_code
            code_source = "original"
            matched_count += 1
        else:
            code = original_code
            code_source = "unmatched"

        result_items.append({
            "index": i + 1,
            "name": item.get("name", ""),
            "description": item.get("description", ""),
            "unit": item.get("unit", ""),
            "quantity": item.get("quantity", None),
            "bill_code": safe_excel_text(code),
            "bill_code_source": code_source,
            "matched_name": safe_excel_text(bill_match.get("name", "") if bill_match else ""),
            "standard_name": safe_excel_text(item.get("standard_name", "")),
            "standard_unit": safe_excel_text(item.get("standard_unit", "")),
            "compiled_features": safe_excel_text(item.get("compiled_features", "")),
            "sheet_name": safe_excel_text(item.get("sheet_name", "")),
            "section": safe_excel_text(item.get("section", "")),
        })

    return {
        "total": len(result_items),
        "matched": matched_count,
        "unmatched": len(result_items) - matched_count,
        "bill_version": bill_version,
        "items": result_items,
    }


def _do_export_local(file_path: str, bill_version: str) -> str:
    """本地执行编清单并生成结果Excel，返回文件路径"""
    import openpyxl
    from src.bill_reader import BillReader
    from src.bill_compiler import compile_items
    from src.bill_feature_builder import group_items_by_section

    reader = BillReader()
    items = reader.read_file(file_path)
    if not items:
        raise HTTPException(status_code=400, detail="未从Excel中读取到清单项，请检查文件格式。")

    compiled = compile_items(items, bill_version=bill_version)

    # 按系统类型+计算项目分组，插入分部标题行
    grouped = group_items_by_section(compiled)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "工程量清单"

    headers = ["序号", "项目编码", "项目名称", "项目特征描述", "计量单位", "工程量"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = openpyxl.styles.Font(bold=True)

    # 分部标题样式
    title_font = openpyxl.styles.Font(bold=True, size=11)

    row_num = 2
    for entry in grouped:
        if entry.get("_is_title"):
            # 分部标题行：只在"项目名称"列写标题
            ws.cell(row=row_num, column=3, value=safe_excel_text(entry["_title"])).font = title_font
            row_num += 1
            continue

        # 正常清单项
        item = entry
        original_code = item.get("code", "").strip()
        bill_match = item.get("bill_match")

        if bill_match and bill_match.get("code_12"):
            code = bill_match["code_12"]
        elif original_code and len(original_code) >= 9:
            code = original_code
        else:
            code = original_code

        display_name = item.get("standard_name") or item.get("name", "")
        display_features = item.get("compiled_features") or item.get("description", "")
        display_unit = item.get("standard_unit") or item.get("unit", "")

        ws.cell(row=row_num, column=1, value=item.get("_seq", ""))
        ws.cell(row=row_num, column=2, value=safe_excel_text(code))
        ws.cell(row=row_num, column=3, value=safe_excel_text(display_name))
        ws.cell(row=row_num, column=4, value=safe_excel_text(display_features))
        ws.cell(row=row_num, column=5, value=safe_excel_text(display_unit))
        ws.cell(row=row_num, column=6, value=item.get("quantity", ""))
        row_num += 1

    col_widths = [8, 18, 20, 50, 10, 12]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    out_dir = Path(tempfile.gettempdir()) / "bill_compiler"
    out_dir.mkdir(exist_ok=True)
    orig_name = Path(file_path).stem
    out_path = out_dir / f"{orig_name}_工程量清单.xlsx"
    wb.save(str(out_path))
    wb.close()

    return str(out_path)


# ============================================================
# 远程模式：转发请求到本地匹配服务（local_match_server.py）
# ============================================================

async def _forward_to_local_service(endpoint: str, file_content: bytes,
                                     filename: str, bill_version: str) -> dict:
    """转发编清单请求到本地匹配服务，返回JSON结果"""
    import httpx
    from app.config import LOCAL_MATCH_URL, LOCAL_MATCH_API_KEY

    if not LOCAL_MATCH_URL:
        raise HTTPException(
            status_code=503,
            detail="编清单需要本地匹配服务，但未配置 LOCAL_MATCH_URL。\n"
                   "请确保本地电脑运行了 local_match_server.py 并配置了正确地址。",
        )

    url = f"{LOCAL_MATCH_URL}/compile-bill/{endpoint}"
    logger.info(f"编清单转发到本地服务: {url}")

    async with httpx.AsyncClient(timeout=120) as client:
        try:
            resp = await client.post(
                url,
                files={"file": (filename, file_content)},
                data={"bill_version": bill_version},
                headers={"X-API-Key": LOCAL_MATCH_API_KEY or ""},
            )
        except httpx.ConnectError:
            raise HTTPException(
                status_code=503,
                detail="无法连接本地匹配服务，请确认：\n"
                       "1. 本地电脑已运行 local_match_server.py\n"
                       "2. LOCAL_MATCH_URL 配置正确\n"
                       "3. 防火墙允许9100端口",
            )

    if resp.status_code != 200:
        detail = resp.text[:500] if resp.text else f"HTTP {resp.status_code}"
        raise HTTPException(status_code=resp.status_code, detail=f"本地服务返回错误: {detail}")

    return resp


# ============================================================
# API 接口
# ============================================================

@router.post("/bill-compiler/preview")
async def preview_compile(
    file: UploadFile = File(description="工程量Excel文件（算量导出/手工表格）"),
    bill_version: str = Form(default="2024", description="清单版本: 2024 或 2013"),
    user: User = Depends(get_current_user),
):
    """预览编清单结果"""
    del user

    if bill_version not in ("2024", "2013"):
        raise HTTPException(status_code=400, detail=f"不支持的清单版本: {bill_version}")

    from app.config import MATCH_BACKEND
    content, filename, info = await _read_and_validate_excel(file, "工程量文件")

    if MATCH_BACKEND == "remote":
        # 远程模式：转发到本地匹配服务
        resp = await _forward_to_local_service("preview", content, filename, bill_version)
        return resp.json()
    else:
        # 本地模式：直接执行
        tmp_path = await _save_upload(content, info.normalized_suffix, "bill_")
        try:
            result = await asyncio.to_thread(_do_compile_local, tmp_path, bill_version)
            return result
        except HTTPException:
            raise
        except Exception as e:
            logger.error(f"编清单预览失败: {e}")
            raise HTTPException(status_code=500, detail=f"编清单失败: {e}")
        finally:
            try:
                Path(tmp_path).unlink(missing_ok=True)
            except Exception:
                pass


@router.post("/bill-compiler/execute")
async def execute_compile(
    file: UploadFile = File(description="工程量Excel文件（算量导出/手工表格）"),
    bill_version: str = Form(default="2024", description="清单版本: 2024 或 2013"),
    user: User = Depends(get_current_user),
):
    """执行编清单，返回结果Excel文件下载"""
    del user

    if bill_version not in ("2024", "2013"):
        raise HTTPException(status_code=400, detail=f"不支持的清单版本: {bill_version}")

    from app.config import MATCH_BACKEND
    content, filename, info = await _read_and_validate_excel(file, "工程量文件")

    if MATCH_BACKEND == "remote":
        # 远程模式：转发到本地匹配服务，直接返回Excel二进制
        resp = await _forward_to_local_service("execute", content, filename, bill_version)

        # 从本地服务的响应头获取文件名
        orig_name = Path(filename or "工程量").stem
        download_name = f"{orig_name}_工程量清单.xlsx"

        return Response(
            content=resp.content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{download_name}"'},
        )
    else:
        # 本地模式：直接执行
        tmp_path = await _save_upload(content, info.normalized_suffix, "bill_")
        try:
            result_path = await asyncio.to_thread(_do_export_local, tmp_path, bill_version)
            orig_name = Path(filename or "工程量").stem
            download_name = f"{orig_name}_工程量清单.xlsx"
            return FileResponse(
                path=result_path,
                filename=download_name,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        except HTTPException:
            raise
        except Exception as e:
            logger.error(f"编清单导出失败: {e}")
            raise HTTPException(status_code=500, detail=f"编清单失败: {e}")
        finally:
            try:
                Path(tmp_path).unlink(missing_ok=True)
            except Exception:
                pass
