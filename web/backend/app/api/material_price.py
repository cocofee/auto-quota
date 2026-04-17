"""
智能填主材 API

提供主材价格查询和回填功能：
  - GET  /provinces          获取有价格数据的省份列表
  - GET  /cities             获取指定省份的城市列表
  - GET  /periods            获取指定省份/城市的期次列表
  - POST /parse              上传Excel，解析出主材行（本地处理，不转发）
  - POST /lookup             批量查价（根据省份/城市/期次）
  - POST /contribute         用户提交手填价格（存入候选层）
  - POST /export             把价格写回原Excel的主材行单价列，返回下载

远程模式（MATCH_BACKEND=remote）下，provinces/cities/periods/lookup/contribute
转发到本地匹配服务（local_match_server.py），因为价格库在本地电脑上。
parse/export 始终在本地处理（只需要操作Excel，不需要价格库）。
"""

import asyncio
import tempfile
import re
import uuid as _uuid_mod
from pathlib import Path
from typing import Optional
from urllib.parse import quote

from fastapi import APIRouter, File, UploadFile, HTTPException, Query
from fastapi.responses import FileResponse
from loguru import logger

from app.config import MATCH_BACKEND, LOCAL_MATCH_URL, LOCAL_MATCH_API_KEY
from app.services.local_http import local_match_async_client

router = APIRouter()

# 主材库路径（本地模式使用）
_MATERIAL_DB_PATH = Path(__file__).parent.parent.parent.parent.parent / "db" / "common" / "material.db"

# 上传文件缓存（parse后保留，export时用）
# key=file_key(uuid), value={"path": 文件路径, "name": 原始文件名}
_uploaded_files: dict[str, dict] = {}


def _is_remote() -> bool:
    """是否使用远程模式（价格库在本地电脑上）"""
    return MATCH_BACKEND == "remote" and bool(LOCAL_MATCH_URL)


def _get_db():
    """获取MaterialDB实例（延迟导入，避免启动时报错）"""
    from src.material_db import MaterialDB
    return MaterialDB(str(_MATERIAL_DB_PATH))


def _load_material_price_cache() -> dict:
    cache_path = Path(__file__).resolve().parents[4] / "data" / "material_prices.json"
    if not cache_path.exists():
        return {}
    try:
        import json as _json
        data = _json.loads(cache_path.read_text(encoding="utf-8"))
        return data if isinstance(data, dict) else {}
    except Exception:
        return {}


def _build_gldjc_search_url(name: str = "", spec: str = "", keyword: str = "", province_code: str = "1") -> str:
    search_keyword = str(keyword or "").strip()
    material_name = str(name or "").strip()
    material_spec = str(spec or "").strip()

    if not search_keyword:
        if material_spec and material_spec not in material_name:
            search_keyword = f"{material_name} {material_spec}".strip()
        else:
            search_keyword = material_name

    if not search_keyword:
        return ""

    return f"https://www.gldjc.com/scj/so.html?keyword={quote(search_keyword, safe='')}&l={province_code}"


def _build_lookup_label(
    *,
    name: str = "",
    spec: str = "",
    unit: str = "",
    price: object = None,
    source: str = "",
    match_label: str = "",
) -> str | None:
    explicit_label = str(match_label or "").strip()
    if explicit_label:
        return explicit_label

    material_text = " ".join(part for part in [str(name or "").strip(), str(spec or "").strip()] if part).strip()
    parts: list[str] = []

    if str(source or "").strip():
        parts.append(str(source).strip())
    if material_text:
        parts.append(material_text)
    if str(unit or "").strip():
        parts.append(str(unit).strip())

    try:
        if price is not None and str(price).strip() != "":
            parts.append(f"{float(price):.2f}")
    except (TypeError, ValueError):
        pass

    return " | ".join(parts) if parts else None


def _extract_material_family_token(text: str) -> str:
    value = str(text or "").strip()
    if not value:
        return ""

    patterns = [
        r"HDPE",
        r"PE-?RT",
        r"PERT",
        r"PEX",
        r"PP-?R",
        r"PPR",
        r"U-?PVC",
        r"PVC-?U",
        r"CPVC",
        r"FRPP",
        r"ABS",
        r"\bPE\b",
        r"\bPVC\b",
        r"\bPP\b",
        r"\bPB\b",
    ]
    for pattern in patterns:
        matched = re.search(pattern, value, flags=re.IGNORECASE)
        if matched:
            return re.sub(r"\s+", "", matched.group(0)).upper()
    return ""


def _normalize_material_family_token(token: str) -> str:
    value = str(token or "").strip().upper()
    if value in {"PP-R", "PPR"}:
        return "PPR"
    if value in {"U-PVC", "UPVC", "PVC-U"}:
        return "UPVC"
    if value in {"PE-RT", "PERT"}:
        return "PERT"
    return value


def _build_lookup_name_variants(name: str, object_type: str = "") -> list[str]:
    """为材料名称生成查询变体（用于数据库 fuzzy 命中）。

    之前只在 object_type == "pipe" 时生成 family token 变体，导致
    "PPR 给水管件" 这种 pipe_fitting 行拿不到 "PPR管件"/"PPR弯头" 变体，
    数据库侧命中率偏低。这里把条件放宽到 pipe/pipe_fitting/material 三类，
    并为管件额外追加 "{family}管件/{family}弯头/{family}三通" 之类常见写法。
    """
    base_name = str(name or "").strip()
    if not base_name:
        return []

    variants: list[str] = []

    def _append(candidate: str):
        value = str(candidate or "").strip()
        if value and value not in variants:
            variants.append(value)

    _append(base_name)

    family_token = _normalize_material_family_token(_extract_material_family_token(base_name))
    # 把条件从 "仅 pipe" 放宽到 pipe / pipe_fitting / material / 空
    eligible_types = {"pipe", "pipe_fitting", "material", ""}
    if family_token and object_type in eligible_types:
        family_pattern = re.escape(family_token).replace(r"\-", "-?")
        stripped = re.sub(family_pattern, "", base_name, count=1, flags=re.IGNORECASE).strip(" -_/")
        stripped = re.sub(r"\s+", "", stripped)
        if stripped:
            _append(f"{family_token}{stripped}")
            _append(f"{stripped}{family_token}")

        # 管道类变体（保留原逻辑）
        if object_type in {"pipe", "material", ""}:
            _append(f"{family_token}管")
            if "给水" in base_name:
                _append(f"{family_token}给水管")
            if "排水" in base_name:
                _append(f"{family_token}排水管")
            if any(token in base_name for token in ("滴灌", "滴水", "灌溉")):
                _append(f"{family_token}滴水管")
                _append(f"{family_token}灌溉管")

        # 管件类变体（新增）
        if object_type in {"pipe_fitting", "material", ""}:
            _append(f"{family_token}管件")
            # 根据原始名里的具体管件类型衍生
            for fitting_kw in ("弯头", "三通", "四通", "异径", "大小头", "接头", "法兰"):
                if fitting_kw in base_name:
                    _append(f"{family_token}{fitting_kw}")

    return variants


def _format_lookup_reason(tags: list[str], note: str) -> str:
    clean_tags = [str(tag or "").strip() for tag in tags if str(tag or "").strip()]
    clean_note = str(note or "").strip()
    if clean_tags and clean_note:
        return f"{' / '.join(clean_tags)}: {clean_note}"
    if clean_note:
        return clean_note
    if clean_tags:
        return " / ".join(clean_tags)
    return ""


def _build_lookup_selection_context(
    *,
    requested_name: str,
    requested_spec: str,
    requested_object_type: str,
    requested_critical_spec: str,
    matched_name: str = "",
    matched_spec: str = "",
    matched_object_type: str = "",
    source: str = "",
    region: str = "",
    variant_name: str = "",
) -> dict[str, object]:
    tags: list[str] = []
    notes: list[str] = []

    if source:
        tags.append(source)
    if region:
        tags.append(region)

    req_spec = str(requested_spec or "").strip()
    hit_spec = str(matched_spec or "").strip()
    if req_spec and hit_spec:
        if re.sub(r"\s+", "", req_spec).upper() == re.sub(r"\s+", "", hit_spec).upper():
            tags.append("规格命中")
        else:
            tags.append("规格近似命中")
            notes.append(f"候选规格={hit_spec}")

    if requested_critical_spec:
        tags.append("关键规格参与")
        notes.append(f"关键规格={requested_critical_spec}")

    request_type = str(requested_object_type or "").strip()
    hit_type = str(matched_object_type or "").strip()
    if request_type and hit_type:
        if request_type == hit_type:
            tags.append("对象一致")
        else:
            tags.append("对象近似")
            notes.append(f"候选对象={hit_type}")

    req_name = str(requested_name or "").strip()
    hit_name = str(matched_name or "").strip()
    if req_name and hit_name and req_name != hit_name:
        tags.append("名称改写命中")
        notes.append(f"候选名称={hit_name}")

    if variant_name and variant_name != req_name:
        notes.append(f"查询变体={variant_name}")

    review_hint = "；".join(notes)
    return {
        "selected_region": region or "",
        "reason_tags": tags,
        "selected_reason": _format_lookup_reason(tags, review_hint),
        "review_hint": review_hint,
    }


def _material_db_ready() -> bool:
    import sqlite3

    if not _MATERIAL_DB_PATH.exists() or _MATERIAL_DB_PATH.stat().st_size <= 0:
        return False

    try:
        conn = sqlite3.connect(str(_MATERIAL_DB_PATH))
        try:
            row = conn.execute(
                "SELECT 1 FROM sqlite_master WHERE type='table' AND name='price_fact'"
            ).fetchone()
            return row is not None
        finally:
            conn.close()
    except Exception:
        return False


async def _remote_get(path: str, params: dict = None) -> dict:
    """转发GET请求到本地匹配服务"""
    import httpx
    url = f"{LOCAL_MATCH_URL.rstrip('/')}{path}"
    headers = {"X-API-Key": LOCAL_MATCH_API_KEY}
    try:
        async with local_match_async_client(timeout=30.0) as client:
            resp = await client.get(url, headers=headers, params=params)
        if resp.status_code == 200:
            return resp.json()
        logger.warning(f"远程主材查询返回 {resp.status_code}: {resp.text[:200]}")
        return {}
    except Exception as e:
        logger.error(f"远程主材查询失败: [{type(e).__name__}] {e} | url={url}")
        raise HTTPException(502, f"连接本地匹配服务失败: {e}")


async def _remote_post(path: str, payload: dict) -> dict:
    """转发POST请求到本地匹配服务"""
    import httpx
    url = f"{LOCAL_MATCH_URL.rstrip('/')}{path}"
    headers = {"X-API-Key": LOCAL_MATCH_API_KEY}
    try:
        async with local_match_async_client(timeout=60.0) as client:
            resp = await client.post(url, headers=headers, json=payload)
        if resp.status_code == 200:
            return resp.json()
        logger.warning(f"远程主材操作返回 {resp.status_code}: {resp.text[:200]}")
        return {}
    except Exception as e:
        logger.error(f"远程主材操作失败: [{type(e).__name__}] {e} | url={url}")
        raise HTTPException(502, f"连接本地匹配服务失败: {e}")


def _material_hint_family(name: str) -> str:
    """判定材料归属族（合并版：pipe/valve/device，原本两份定义的并集）。"""
    value = _normalize_material_hint(name)
    pipe_names = {"复合管", "钢管", "塑料管", "塑料给水管", "塑料排水管", "管材", "管件"}
    valve_names = {"阀门", "螺纹阀门", "法兰阀门", "减压器", "减压阀", "螺纹减压阀", "法兰减压阀", "过滤器"}
    if value in {_normalize_material_hint(x) for x in pipe_names}:
        return "pipe"
    if value in {_normalize_material_hint(x) for x in valve_names}:
        return "valve"
    if any(token in str(name or "") for token in ("地漏", "盆", "洁具", "卫生器具", "器具")):
        return "device"
    return ""


def _is_compatible_material_hint(material_name: str, candidate_name: str) -> bool:
    """判定候选名称与目标材料是否属于兼容品类（合并版）。"""
    text = str(candidate_name or "").strip()
    if not text:
        return False

    family = _material_hint_family(material_name)
    if not family:
        return True

    if family == "pipe":
        if any(token in text for token in ("刷", "漆", "涂", "标识", "油")):
            return False
        # 塑料族 token 本身就可代表管材（如候选只写 "PPR"/"PVC"）
        if _looks_like_material_family_token(text):
            return True
        return any(token in text for token in ("管", "管材", "管件"))

    if family == "valve":
        return any(token in text for token in ("阀", "过滤器", "减压"))

    if family == "device":
        return any(token in text for token in ("地漏", "盆", "洁具", "器", "卫生"))

    return True


# ============================================================
# 1. 获取省份列表
# ============================================================

@router.get("/material-price/provinces")
async def get_provinces():
    """返回有价格数据的省份列表（按数据量降序）"""
    if _is_remote():
        return await _remote_get("/material-price/provinces")

    if not _material_db_ready():
        cache = _load_material_price_cache()
        count = len(cache)
        if count > 0:
            return {"provinces": [{"name": "全国", "count": count}], "fallback": "json_cache"}
        return {"provinces": []}

    import sqlite3
    conn = sqlite3.connect(str(_MATERIAL_DB_PATH))
    try:
        rows = conn.execute(
            """SELECT province, COUNT(*) as cnt
               FROM price_fact
               WHERE province != '' AND province != '全国'
               GROUP BY province
               ORDER BY cnt DESC"""
        ).fetchall()
        return {
            "provinces": [{"name": r[0], "count": r[1]} for r in rows]
        }
    finally:
        conn.close()


# ============================================================
# 2. 获取城市列表
# ============================================================

@router.get("/material-price/cities")
async def get_cities(province: str = Query(..., description="省份名称")):
    """返回指定省份下有价格数据的城市列表"""
    if _is_remote():
        return await _remote_get("/material-price/cities", {"province": province})

    if not _material_db_ready():
        return {"cities": []}

    import sqlite3
    conn = sqlite3.connect(str(_MATERIAL_DB_PATH))
    try:
        rows = conn.execute(
            """SELECT city, COUNT(*) as cnt
               FROM price_fact
               WHERE province = ? AND city != ''
               GROUP BY city
               ORDER BY cnt DESC""",
            (province,)
        ).fetchall()
        return {
            "cities": [{"name": r[0], "count": r[1]} for r in rows]
        }
    finally:
        conn.close()


# ============================================================
# 3. 获取期次列表
# ============================================================

@router.get("/material-price/periods")
async def get_periods(
    province: str = Query(..., description="省份名称"),
    city: str = Query("", description="城市名称（可选）"),
):
    """返回指定省份/城市的信息价期次列表（按时间倒序）"""
    if _is_remote():
        params = {"province": province}
        if city:
            params["city"] = city
        return await _remote_get("/material-price/periods", params)

    if not _material_db_ready():
        cache = _load_material_price_cache()
        latest = ""
        for item in cache.values():
            if isinstance(item, dict):
                date_val = str(item.get("query_date") or "").strip()
                if date_val and date_val > latest:
                    latest = date_val
        if latest:
            return {"periods": [{"start": latest, "end": latest, "count": len(cache), "label": latest}]}
        return {"periods": []}

    import sqlite3
    conn = sqlite3.connect(str(_MATERIAL_DB_PATH))
    try:
        # 构建查询条件
        conditions = ["province = ?", "period_start != ''"]
        params = [province]
        if city:
            conditions.append("city = ?")
            params.append(city)

        where = " AND ".join(conditions)
        rows = conn.execute(
            f"""SELECT period_start, period_end, COUNT(*) as cnt
                FROM price_fact
                WHERE {where}
                GROUP BY period_start, period_end
                ORDER BY period_end DESC
                LIMIT 24""",
            params
        ).fetchall()
        return {
            "periods": [
                {
                    "start": r[0],
                    "end": r[1],
                    "count": r[2],
                    # 生成显示标签，如"2026年2月"
                    "label": _format_period_label(r[0], r[1]),
                }
                for r in rows
            ]
        }
    finally:
        conn.close()


def _format_period_label(start: str, end: str) -> str:
    """把期次日期格式化成中文标签，如 '2026-02-01' → '2026年2月'"""
    try:
        parts = start.split("-")
        year, month = int(parts[0]), int(parts[1])
        return f"{year}年{month}月"
    except (IndexError, ValueError):
        return f"{start} ~ {end}"


# ============================================================
# 4. 解析Excel中的主材行
# ============================================================

@router.post("/material-price/parse")
async def parse_materials(file: UploadFile = File(...)):
    """上传Excel，识别出主材行，返回主材列表

    文件会保留在临时目录，export时用file_key找回来写入价格。
    """
    # 校验文件类型
    filename = file.filename or ""
    if not filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(400, "请上传Excel文件（.xlsx/.xls）")

    # 保存到临时文件（不删除，留给export用）
    content = await file.read()
    suffix = Path(filename).suffix
    file_key = str(_uuid_mod.uuid4())
    with tempfile.NamedTemporaryFile(
        mode="wb", suffix=suffix, delete=False, prefix="material_"
    ) as tmp:
        tmp.write(content)
        tmp_path = tmp.name

    # 记录文件路径和原始文件名
    _uploaded_files[file_key] = {"path": tmp_path, "name": Path(filename).stem}

    # 在线程池中解析（CPU密集型）
    try:
        result = await asyncio.to_thread(_do_parse, tmp_path)
        materials = result["materials"]
        return {
            "materials": materials,
            "all_rows": result["all_rows"],
            "is_mixed": result["is_mixed"],
            "count": len(materials),
            "file_key": file_key,  # 前端保存，export时回传
        }
    except Exception as e:
        logger.error(f"解析主材失败: {e}")
        # 解析失败才删文件
        Path(tmp_path).unlink(missing_ok=True)
        _uploaded_files.pop(file_key, None)
        raise HTTPException(500, f"解析失败: {e}")


def _do_parse(excel_path: str) -> dict:
    """解析Excel，返回主材行 + 所有行（用于层级预览）

    返回:
    {
        "materials": [...],   # 只含主材行（查价/导出用）
        "all_rows": [...],    # 所有行含类型标记（前端层级展示用）
        "is_mixed": bool,     # 是否混合表（分部分项格式）
    }
    """
    import openpyxl

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    materials = []
    all_rows = []
    is_mixed = False

    for ws in wb.worksheets:
        result = _parse_sheet(ws)
        materials.extend(result["materials"])
        all_rows.extend(result["all_rows"])
        if result["is_mixed"]:
            is_mixed = True

    wb.close()
    return {"materials": materials, "all_rows": all_rows, "is_mixed": is_mixed}


def _parse_sheet(ws) -> dict:
    """解析单个sheet，返回主材行 + 所有行（带类型标记）

    行类型：
    - "section": 分部标题行（如"给水工程"、"新风系统"）
    - "bill": 清单行（有项目编码，如030701003005）
    - "quota": 定额行（编码像C7-1-20）
    - "material": 主材行（编码为"主"等）

    对于纯材料表，all_rows == materials（没有层级）。
    """
    materials = []
    all_rows = []

    # 找表头行
    header_row = None
    col_map = {}
    is_pure_material_table = False
    for row_idx in range(1, min(20, ws.max_row + 1)):
        for col_idx in range(1, min(15, ws.max_column + 1)):
            val = str(ws.cell(row=row_idx, column=col_idx).value or "").strip()
            if val in ("名称", "项目名称", "材料名称"):
                header_row = row_idx
                break
        if header_row:
            break

    if not header_row:
        return {"materials": [], "all_rows": [], "is_mixed": False}

    # 建立列映射——扫描表头行及下一行（处理双行表头）
    _PRICE_KEYWORDS = ("单价", "综合单价", "不含税单价", "含税单价",
                       "市场价", "信息价", "除税单价", "除税信息价")
    # 也检测"序号"列（用于判断清单行，但不作为code列）
    seq_col = None
    for scan_row in range(header_row, min(header_row + 2, ws.max_row + 1)):
        for col_idx in range(1, min(15, ws.max_column + 1)):
            val = str(ws.cell(row=scan_row, column=col_idx).value or "").strip()
            if not val:
                continue
            if val in ("编码", "编号", "材料编码", "项目编码"):
                col_map.setdefault("code", col_idx)
            elif val in ("名称", "项目名称", "材料名称"):
                col_map.setdefault("name", col_idx)
            elif val in ("规格", "规格型号", "规格、型号等特殊要求"):
                col_map.setdefault("spec", col_idx)
            elif val in ("单位", "计量单位"):
                col_map.setdefault("unit", col_idx)
            elif val in ("数量", "工程量", "消耗量", "用量"):
                col_map.setdefault("qty", col_idx)
            elif val in _PRICE_KEYWORDS:
                col_map.setdefault("price", col_idx)
            if val in ("序号",):
                seq_col = col_idx
            # 纯材料表特征
            if val in ("材料名称", "材料编码", "市场价", "信息价",
                        "除税单价", "除税信息价", "消耗量"):
                is_pure_material_table = True

    sheet_name = ws.title or ""
    if any(kw in sheet_name for kw in ("材料", "主材", "材价", "物资")):
        is_pure_material_table = True

    if "name" not in col_map:
        return {"materials": [], "all_rows": [], "is_mixed": False}

    # 确定价格列位置
    price_col = col_map.get("price")
    if price_col is None and "qty" in col_map:
        price_col = col_map["qty"] + 1

    # 也检测"项目特征描述"列（用于清单行展示）
    # 任务导出的结果表常见双行表头/合并表头，描述列不一定落在第一行。
    desc_col = None
    for scan_row in range(header_row, min(header_row + 2, ws.max_row + 1)):
        for col_idx in range(1, min(20, ws.max_column + 1)):
            val = str(ws.cell(row=scan_row, column=col_idx).value or "").strip()
            if val in ("项目特征描述", "项目特征", "特征描述", "描述"):
                desc_col = col_idx
                break
        if desc_col is not None:
            break

    # 遍历数据行
    current_bill: dict | None = None
    for row_idx in range(header_row + 1, ws.max_row + 1):
        code_val = ""
        if "code" in col_map:
            code_val = str(ws.cell(row=row_idx, column=col_map["code"]).value or "").strip()

        name_val = str(ws.cell(row=row_idx, column=col_map["name"]).value or "").strip()
        if not name_val:
            continue

        unit_val = ""
        if "unit" in col_map:
            unit_val = str(ws.cell(row=row_idx, column=col_map["unit"]).value or "").strip()

        qty_val = None
        if "qty" in col_map:
            raw = ws.cell(row=row_idx, column=col_map["qty"]).value
            if raw is not None:
                try:
                    qty_val = float(raw)
                except (ValueError, TypeError):
                    pass

        spec_val = ""
        if "spec" in col_map:
            spec_val = str(ws.cell(row=row_idx, column=col_map["spec"]).value or "").strip()

        price_val = None
        if price_col is not None:
            raw = ws.cell(row=row_idx, column=price_col).value
            if raw is not None:
                try:
                    price_val = float(raw)
                except (ValueError, TypeError):
                    pass

        # 纯材料表：所有行都当主材
        if is_pure_material_table:
            normalized = _build_normalized_material_fields(name_val, spec_val)
            mat = {
                "type": "material",
                "row": row_idx,
                "sheet": ws.title,
                "header_row": header_row,
                "code": code_val,
                "name": name_val,
                "name_col": col_map.get("name"),
                "spec": spec_val,
                "spec_col": col_map.get("spec"),
                "unit": unit_val,
                "qty": qty_val,
                "existing_price": price_val,
                "price_col": price_col,
                "lookup_price": None,
                "lookup_source": None,
                "lookup_url": None,
                "lookup_label": None,
                "suggested_name": normalized.get("suggested_name", ""),
                "suggested_spec": normalized.get("suggested_spec", ""),
                "normalized_name": normalized.get("normalized_name", ""),
                "normalized_spec": normalized.get("normalized_spec", ""),
                "critical_spec_text": normalized.get("critical_spec_text", ""),
                "normalized_query_text": normalized.get("normalized_query_text", ""),
                "object_type": normalized.get("object_type", ""),
                "family": normalized.get("family", ""),
                "normalization_confidence": normalized.get("normalization_confidence", ""),
                "connection_hint": normalized.get("connection_hint", ""),
                "material_hint": normalized.get("material_hint", ""),
                "desc_type_hint": normalized.get("desc_type_hint", ""),
            }
            materials.append(mat)
            all_rows.append(mat)
            continue

        # 混合表：判断行类型
        seq_val = ""
        if seq_col is not None:
            seq_val = str(ws.cell(row=row_idx, column=seq_col).value or "").strip()

        row_type = _classify_row(code_val, name_val, seq_val)

        if row_type == "material":
            normalized = _build_normalized_material_fields(
                name_val,
                spec_val,
                current_bill.get("name", "") if current_bill else "",
                current_bill.get("desc", "") if current_bill else "",
            )
            suggested_name = normalized.get("suggested_name", "")
            suggested_spec = normalized.get("suggested_spec", "")
            if current_bill:
                suggested_name = normalized.get("suggested_name", "")
                suggested_spec = normalized.get("suggested_spec", "")
            mat = {
                "type": "material",
                "row": row_idx,
                "sheet": ws.title,
                "header_row": header_row,
                "code": code_val,
                "name": name_val,
                "desc": current_bill.get("desc", "") if current_bill else "",
                "name_col": col_map.get("name"),
                "suggested_name": suggested_name,
                "spec": spec_val,
                "spec_col": col_map.get("spec"),
                "suggested_spec": suggested_spec,
                "unit": unit_val,
                "qty": qty_val,
                "existing_price": price_val,
                "price_col": price_col,
                "lookup_price": None,
                "lookup_source": None,
                "lookup_url": None,
                "lookup_label": None,
                "normalized_name": normalized.get("normalized_name", ""),
                "normalized_spec": normalized.get("normalized_spec", ""),
                "critical_spec_text": normalized.get("critical_spec_text", ""),
                "normalized_query_text": normalized.get("normalized_query_text", ""),
                "object_type": normalized.get("object_type", ""),
                "family": normalized.get("family", ""),
                "normalization_confidence": normalized.get("normalization_confidence", ""),
                "connection_hint": normalized.get("connection_hint", ""),
                "material_hint": normalized.get("material_hint", ""),
                "desc_type_hint": normalized.get("desc_type_hint", ""),
            }
            materials.append(mat)
            all_rows.append(mat)
        elif row_type == "bill":
            # 清单行：读取项目特征描述
            desc_val = ""
            if desc_col is not None:
                desc_val = str(ws.cell(row=row_idx, column=desc_col).value or "").strip()
            current_bill = {
                "name": name_val,
                "desc": desc_val,
            }
            all_rows.append({
                "type": "bill",
                "row": row_idx,
                "sheet": ws.title,
                "code": code_val,
                "name": name_val,
                "desc": desc_val,
                "unit": unit_val,
                "qty": qty_val,
            })
        elif row_type == "quota":
            all_rows.append({
                "type": "quota",
                "row": row_idx,
                "sheet": ws.title,
                "code": code_val,
                "name": name_val,
                "desc": current_bill.get("desc", "") if current_bill else "",
                "unit": unit_val,
                "qty": qty_val,
            })
        elif row_type == "section":
            current_bill = None
            all_rows.append({
                "type": "section",
                "row": row_idx,
                "sheet": ws.title,
                "name": name_val,
            })

    return {
        "materials": materials,
        "all_rows": all_rows,
        "is_mixed": not is_pure_material_table,
    }


def _classify_row(code: str, name: str, seq: str) -> str:
    """判断分部分项混合表中一行的类型

    返回: "bill"(清单行) / "quota"(定额行) / "material"(主材行) / "section"(分部标题)
    """
    c = code.strip() if code else ""

    # 主材行判断（优先，因为B列="主"最明确）
    if _is_material_row(c, name):
        return "material"

    # 清单行：项目编码通常就是 9-15 位纯数字。
    # 不能强依赖"序号"表头，因为不少导出表首列为空表头，但数据区仍然有序号。
    if re.fullmatch(r"\d{9,15}", c):
        return "bill"

    # 定额行：编码像定额编号（如C10-2-123、A-9-63、SC20）
    if c and re.match(r"^[A-Za-z]{1,4}\d", c):
        return "quota"

    # 分部标题行：没有编码、没有序号、没有单位，只有名称
    if not c and not seq:
        return "section"

    return "section"  # 无法识别的行当标题处理


def _is_material_row(code: str, name: str) -> bool:
    """判断是否为主材行

    识别规则（和 bill_reader.py 保持一致）：
    1. 编码以 ZCGL 开头（广联达主材编码）
    2. 编码包含 Z@ 或 @
    3. 编码为"主"字
    4. 编码为"补充主材"开头
    5. 7-8位纯数字编码（广联达材料编码）
    6. 名称包含典型材料关键词且不是定额行
    """
    c = code.strip() if code else ""

    # 编码特征判断
    if c.upper().startswith("ZCGL"):
        return True
    if c.upper().startswith("ZC"):
        return True
    if re.fullmatch(r"Z[0-9A-Z]+", c.upper()):
        return True
    if "Z@" in c or "@" in c:
        return True
    if c == "主":
        return True
    if c.startswith("补充主材"):
        return True
    if re.fullmatch(r"\d{7,8}", c):
        return True

    # 名称特征判断（补充：没有编码但名称像材料的行）
    if not c and name:
        material_keywords = [
            "钢管", "管材", "管件", "阀门", "水表", "电缆", "电线",
            "灯具", "开关", "插座", "配电箱", "桥架", "线槽",
            "PPR", "PVC", "PE管", "铜管", "不锈钢管",
            "水泥", "砂浆", "砂石", "混凝土", "钢筋",
            "防水卷材", "保温材料", "涂料", "油漆",
            "风机", "风口", "消火栓", "喷头", "烟感", "温感",
        ]
        for kw in material_keywords:
            if kw in name:
                return True

    return False


def _suggest_material_from_bill_context(material_name: str, bill_name: str, desc: str) -> tuple[str, str]:
    info = _extract_material_from_desc(desc)
    candidate_name = info["name"]
    candidate_spec = info["spec"]
    candidate_type = info["type"]
    candidate_material = info["material"]
    candidate_connection = info["connection"]
    candidate_model = str(info.get("model") or "").strip()
    raw_material_name = str(material_name or "").strip()
    cleaned_material_name = _strip_inline_spec_from_name(raw_material_name) or raw_material_name
    bill_candidate_name = _normalize_bill_candidate_name(bill_name)
    fallback_spec = candidate_spec or candidate_model or _extract_inline_material_spec(raw_material_name)
    if not fallback_spec and bill_candidate_name != str(bill_name or "").strip():
        fallback_spec = _extract_inline_material_spec(bill_name)
    if not candidate_name and not candidate_spec and not candidate_type and not _is_viable_material_candidate(bill_candidate_name):
        return "", fallback_spec

    reference_name = cleaned_material_name or raw_material_name
    normalized_material = _normalize_material_hint(raw_material_name)
    normalized_bill = _normalize_material_hint(bill_candidate_name or bill_name)
    normalized_candidate = _normalize_material_hint(candidate_name)
    material_name_is_spec_only = _is_spec_only_material_name(raw_material_name)
    material_name_is_generic = _is_effective_generic_material_name(raw_material_name)

    suggested_name = ""
    if _should_use_conduit_material_name(reference_name, bill_candidate_name or bill_name, candidate_material):
        suggested_name = candidate_material.strip()
    elif _should_use_bill_type_for_material(reference_name) and candidate_type:
        suggested_name = candidate_type
        if _should_prefix_connection(candidate_type) and candidate_connection and candidate_connection not in suggested_name:
            conn_prefix = _connection_prefix(candidate_connection)
            if conn_prefix and conn_prefix not in suggested_name:
                suggested_name = f"{conn_prefix}{suggested_name}"
        if _should_prefix_material(candidate_type) and candidate_material and candidate_material not in suggested_name:
            suggested_name = f"{candidate_material}{suggested_name}"
    elif _should_use_pipe_fitting_type(reference_name, candidate_type, candidate_name):
        suggested_name = _compose_pipe_fitting_name(
            reference_name,
            candidate_type or candidate_name,
            candidate_material,
        )
    elif _should_prefix_specific_pipe_fitting(reference_name, bill_candidate_name or bill_name, candidate_type, candidate_material):
        suggested_name = _compose_specific_pipe_fitting_name(candidate_type, candidate_material)
    elif (
        candidate_name
        and not _looks_like_bare_material_token(candidate_name)
        and _looks_like_installation_item_name(reference_name)
        and _is_compatible_material_hint(reference_name, candidate_name)
    ):
        suggested_name = candidate_name
    elif (
        candidate_name
        and _should_prefix_material_family(reference_name, candidate_name)
        and (not _looks_like_bare_material_token(candidate_name) or material_name_is_generic)
    ):
        suggested_name = f"{candidate_name}{cleaned_material_name.strip()}"
    elif _should_use_specific_equipment_name(reference_name, bill_candidate_name or bill_name, candidate_type, candidate_name):
        suggested_name = _clean_equipment_candidate_name(candidate_type or candidate_name)
    elif _should_merge_device_name(reference_name, bill_candidate_name or bill_name, candidate_type, candidate_material):
        device_name = candidate_type.strip() or bill_candidate_name.strip() or bill_name.strip()
        if candidate_material and candidate_material not in device_name:
            suggested_name = f"{candidate_material}{device_name}"
        else:
            suggested_name = device_name
    elif (
        candidate_type
        and re.search(r"(PP-?R|PPR|PVC-?U|U-?PVC|CPVC|HDPE|PE-?RT|PERT|PEX|FRPP|PB|ABS)", candidate_type, re.IGNORECASE)
        and not _looks_like_bare_material_token(candidate_type)
        and _is_compatible_material_hint(reference_name, candidate_type)
    ):
        suggested_name = candidate_type.strip()
    elif (
        candidate_name
        and re.search(r"(PP-?R|PPR|PVC-?U|U-?PVC|CPVC|HDPE|PE-?RT|PERT|PEX|FRPP|PB|ABS)", candidate_name, re.IGNORECASE)
        and not _looks_like_bare_material_token(candidate_name)
        and _is_compatible_material_hint(reference_name, candidate_name)
    ):
        suggested_name = candidate_name
    elif (
        (material_name_is_spec_only or material_name_is_generic or _looks_like_installation_item_name(reference_name))
        and _is_viable_material_candidate(bill_candidate_name)
        and _is_compatible_material_hint(reference_name or bill_candidate_name, bill_candidate_name)
    ):
        suggested_name = bill_candidate_name.strip()
    elif (
        material_name_is_generic
        and candidate_name
        and re.search(r"(PP-?R|PPR|PVC-?U|U-?PVC|CPVC|HDPE|PE-?RT|PERT|PEX|FRPP|PB|ABS)", candidate_name, re.IGNORECASE)
        and _is_compatible_material_hint(reference_name, candidate_name)
    ):
        suggested_name = candidate_name
    elif (
        material_name_is_generic
        and bill_candidate_name
        and _is_viable_material_candidate(bill_candidate_name)
        and _is_compatible_material_hint(reference_name, bill_candidate_name)
    ):
        suggested_name = bill_candidate_name.strip()
    elif (
        normalized_material
        and normalized_bill
        and normalized_material == normalized_bill
        and candidate_name
        and not _looks_like_bare_material_token(candidate_name)
    ):
        suggested_name = candidate_name
    elif (
        normalized_material
        and normalized_candidate
        and normalized_material == normalized_candidate
        and candidate_name
        and not _looks_like_bare_material_token(candidate_name)
    ):
        suggested_name = candidate_name

    if suggested_name and _normalize_material_hint(suggested_name) == normalized_material:
        suggested_name = ""
    if suggested_name and _is_generic_pipe_fitting_name(reference_name) and _looks_like_bare_material_token(suggested_name):
        suggested_name = ""
    if material_name_is_spec_only and raw_material_name and fallback_spec:
        raw_spec = raw_material_name.strip()
        if len(raw_spec) > len(fallback_spec):
            fallback_spec = raw_spec

    if (
        suggested_name
        and candidate_model
        and "配电箱" in suggested_name
        and candidate_model not in fallback_spec
    ):
        fallback_spec = candidate_model

    return suggested_name, fallback_spec


def _extract_material_from_desc(desc: str) -> dict[str, str]:
    """从清单项目特征描述里提取 name/spec/type/material/connection。

    合并版：
      - 保留第二版的"1." → 换行修复（规避 OCR 把多行特征挤到一行）；
      - 保留第二版的"型号规格"额外 key 和从规格反向拆名的能力；
      - 保留第一版更全的"连接形式"候选 key（接口形式/焊接方法等）。
    """
    if not desc:
        return {"name": "", "spec": "", "type": "", "material": "", "connection": "", "model": ""}

    pairs: dict[str, str] = {}
    normalized_desc = re.sub(r"(?<!\S)(\d+)\.\s*", r"\n\1.", str(desc or ""))
    for raw_line in normalized_desc.splitlines():
        line = re.sub(r"^\s*\d+[\.、\s]*", "", raw_line.strip())
        if not line:
            continue
        parts = re.split(r"[:：]", line, maxsplit=1)
        if len(parts) != 2:
            continue
        key = _normalize_desc_key(parts[0].strip())
        value = parts[1].strip()
        if key and value:
            pairs[key] = value

    combined = ""
    for key in ("材质规格", "材质及规格", "材质型号", "规格压力等级"):
        if pairs.get(key):
            combined = pairs[key]
            break

    candidate_type = ""
    for key in ("类型", "类别", "名称"):
        if pairs.get(key):
            candidate_type = pairs[key].strip()
            break

    candidate_material = ""
    for key in ("材质", "材质要求"):
        if pairs.get(key):
            candidate_material = pairs[key].strip()
            break

    candidate_connection = ""
    for key in ("连接形式", "连接方式", "接口形式", "接口方式", "焊接方法", "焊接方式"):
        if pairs.get(key):
            candidate_connection = pairs[key].strip()
            break

    name = ""
    spec = ""
    if combined:
        name, spec = _split_material_and_spec(combined)

    if not name and candidate_material:
        name = candidate_material

    model = ""
    if not spec:
        for key in ("型号规格", "规格型号", "规格", "型号", "公称直径", "规格压力等级"):
            raw_spec = pairs.get(key, "").strip()
            if not raw_spec:
                continue
            maybe_name, maybe_spec = _split_material_and_spec(raw_spec)
            if maybe_spec:
                if maybe_name and not candidate_type:
                    candidate_type = maybe_name
                if maybe_name and (not name or _looks_like_bare_material_token(name)):
                    name = maybe_name
                cleaned_spec = _clean_material_spec(maybe_spec)
                if _is_placeholder_material_spec(cleaned_spec):
                    continue
                spec = cleaned_spec
                break
            cleaned_raw_spec = _clean_material_spec(raw_spec)
            if _is_placeholder_material_spec(cleaned_raw_spec):
                continue
            spec = cleaned_raw_spec
            if key == "型号" and cleaned_raw_spec:
                model = cleaned_raw_spec
            break
    elif pairs.get("型号"):
        cleaned_model = _clean_material_spec(pairs.get("型号", "").strip())
        if cleaned_model and not _is_placeholder_material_spec(cleaned_model):
            model = cleaned_model

    return {
        "name": name.strip(),
        "spec": _clean_material_spec(spec),
        "type": candidate_type,
        "material": candidate_material,
        "connection": candidate_connection,
        "model": model.strip(),
    }


def _split_material_and_spec(text: str) -> tuple[str, str]:
    value = str(text or "").strip()
    if not value:
        return "", ""

    patterns = [
        r"^(?P<name>.+?)(?P<spec>(?:DN|De)\d+[^\u4e00-\u9fff]*)$",
        r"^(?P<name>.+?)\s+(?P<spec>(?:DN|De)\d+[^\u4e00-\u9fff]*)$",
        r"^(?P<name>.+?)(?P<spec>DN\d+[A-Za-z0-9\-\./]*(?:\s+[A-Za-z0-9\-\./]+)*)$",
        r"^(?P<name>.+?)\s+(?P<spec>DN\d+[A-Za-z0-9\-\./]*(?:\s+[A-Za-z0-9\-\./]+)*)$",
        r"^(?P<name>.+?)(?P<spec>De\d+[A-Za-z0-9\-\./]*(?:\s+[A-Za-z0-9\-\./]+)*)$",
        r"^(?P<name>.+?)\s+(?P<spec>De\d+[A-Za-z0-9\-\./]*(?:\s+[A-Za-z0-9\-\./]+)*)$",
        r"^(?P<name>.+?)\s+(?P<spec>\d+(?:\.\d+)?(?:mm|mm2|㎡|m2))$",
        r"^(?P<name>.+?)\s+(?P<spec>\d+(?:\*\d+){1,3})$",
    ]
    for pattern in patterns:
        matched = re.match(pattern, value, flags=re.IGNORECASE)
        if matched:
            return matched.group("name").strip(), matched.group("spec").strip()
    return value, ""


def _clean_material_spec(spec: str) -> str:
    value = str(spec or "").strip()
    if not value:
        return ""

    matched = re.search(r"(DN\d+(?:[\*xX×]DN?\d+){1,3})", value, flags=re.IGNORECASE)
    if matched:
        return matched.group(1).replace(" ", "")

    matched = re.search(r"(DN\d+[A-Za-z0-9\-\./]*)", value, flags=re.IGNORECASE)
    if matched:
        return matched.group(1).replace(" ", "")

    matched = re.search(r"(De\d+[A-Za-z0-9\-\./]*(?:\s+[A-Za-z0-9\-\./]+)*)", value, flags=re.IGNORECASE)
    if matched:
        return matched.group(1).strip()

    return value


def _is_placeholder_material_spec(spec: str) -> bool:
    value = re.sub(r"\s+", "", str(spec or "")).lower()
    if not value:
        return True
    placeholder_tokens = (
        "综合考虑",
        "详见图纸",
        "详见设计",
        "按设计要求",
        "按图纸",
        "见图纸",
        "详设计图纸",
        "详施工图",
        "未注明",
        "未详",
        "见设计说明",
        "按系统图",
    )
    return any(token in value for token in placeholder_tokens)


def _normalize_material_hint(text: str) -> str:
    return re.sub(r"[\s\-\(\)（）]", "", str(text or "")).lower()


def _is_generic_material_name(name: str) -> bool:
    """是否为泛名（需要借助清单项目特征描述补充具体材料）。

    合并版：
      - 取两版通用名的并集（管材族 + 阀门族 + 电气/消防族）；
      - 遇到 "法兰阀门/螺纹减压阀" 这种组合词，交由 _should_use_bill_type_for_material 兜底。
    """
    if _should_use_bill_type_for_material(name):
        return True
    generic_names = {
        # 管道族
        "复合管", "管材", "管件", "钢管", "塑料管", "塑料给水管", "塑料排水管",
        # 阀门/过滤/减压族
        "阀门", "螺纹阀门", "法兰阀门", "减压器", "减压阀", "螺纹减压阀",
        "法兰减压阀", "过滤器", "螺纹过滤器", "法兰过滤器",
        # 电气/消防/暖通族（这些常靠项目特征描述补全）
        "桥架", "电缆", "电线", "配电箱", "灯具", "喷头", "风口", "水表",
        "法兰",
    }
    return _normalize_material_hint(name) in {_normalize_material_hint(x) for x in generic_names}


def _should_use_bill_type_for_material(name: str) -> bool:
    """是否应从清单行的"类型/类别"字段取具体材料名（合并版，带 regex 兜底）。"""
    generic_valve_names = {
        "阀门", "螺纹阀门", "减压器", "减压阀",
        "螺纹减压阀", "法兰阀门", "法兰减压阀", "过滤器",
    }
    normalized = _normalize_material_hint(name)
    if normalized in {_normalize_material_hint(x) for x in generic_valve_names}:
        return True

    # 诸如 "法兰阀门"、"螺纹减压阀"、"PPR 阀门" 之类"前缀+阀/减/过滤"的组合
    raw = re.sub(r"[\s\-\(\)\uFF08\uFF09]", "", str(name or ""))
    prefix_tokens = [
        "ppr", "pe", "pex", "pert", "pvc", "upvc", "cpvc", "hdpe", "frpp", "pp", "abs",
        "塑料", "金属", "钢制",
        "法兰", "螺纹", "沟槽", "焊接", "热熔", "熔接", "电熔", "承插",
    ]
    suffix_tokens = ["阀门", "过滤器", "减压阀", "减压器"]
    prefix_pattern = "|".join(re.escape(x) for x in prefix_tokens)
    suffix_pattern = "|".join(re.escape(x) for x in suffix_tokens)
    return bool(re.fullmatch(rf"(?i)(?:{prefix_pattern})+(?:{suffix_pattern})", raw))


def _connection_prefix(connection: str) -> str:
    """从连接形式文本（如"法兰连接"）提取前缀词。"""
    text = str(connection or "")
    if "法兰" in text:
        return "法兰"
    if "螺纹" in text:
        return "螺纹"
    if "沟槽" in text:
        return "沟槽"
    if "焊接" in text:
        return "焊接"
    return ""


def _normalize_desc_key(text: str) -> str:
    return re.sub(r"[\s、，,：:/（）\(\)\-]", "", str(text or ""))


def _is_specific_lookup_name(name: str) -> bool:
    value = str(name or "").strip()
    return bool(value) and not _is_generic_material_name(value)


def _should_prefix_connection(candidate_type: str) -> bool:
    value = _normalize_material_hint(candidate_type)
    return value in {_normalize_material_hint(x) for x in {"过滤器", "阀门", "减压阀", "减压器"}}


def _should_prefix_material(candidate_type: str) -> bool:
    value = str(candidate_type or "").strip()
    return "阀" in value or "减压器" in value


# ============================================================
# 4b. 从任务结果中提取主材行
# ============================================================
# _normalize_desc_key / _material_hint_family 已在文件顶部定义，此处不再重复。


def _looks_like_installation_item_name(name: str) -> bool:
    """判定是否为"安装项"（例如"管道敷设"），用于避免把这类行当材料去查价。

    注意：之前文件里还有另一个弱化版本的同名函数（只判定 4 个关键词），
    会覆盖这份更严谨的实现，导致塑料管的安装项被误判。已合并保留此版本。
    """
    text = str(name or "").strip()
    if not text:
        return False
    install_tokens = ("安装", "敷设", "铺设", "组装", "调试", "管线", "组成")
    if not any(token in text for token in install_tokens):
        return False
    if _material_hint_family(text):
        return False
    if re.search(r"(HDPE|PE-?RT|PE|PPR|PP-?R|UPVC|U-PVC|PVC|CPVC|镀锌|钢塑复合|不锈钢|铸铁)", text, re.IGNORECASE):
        return False
    return True


def _spec_contains_exact(result_spec: str, target_spec: str) -> bool:
    """判断 target_spec 是否作为"整词"出现在 result_spec 中。

    先做一次规格归一化（×/x/X/* → X，φ/Φ → PHI），避免
    "DN25×2.75" 和 "DN25x2.75" 被当成两个不同规格。
    """

    def _normalize(text: str) -> str:
        value = re.sub(r"\s+", "", str(text or "").upper())
        value = value.replace("×", "X").replace("*", "X")
        value = value.replace("Φ", "PHI").replace("φ", "PHI")
        return value

    result_text = _normalize(result_spec)
    target_text = _normalize(target_spec)
    if not result_text or not target_text:
        return False

    pattern = re.escape(target_text)
    if target_text[0].isdigit():
        pattern = rf"(?<!\d){pattern}"
    if target_text[-1].isdigit():
        pattern = rf"{pattern}(?!\d)"
    return re.search(pattern, result_text) is not None


# 单位兼容组（与 gldjc_price.UNIT_COMPAT_GROUPS 保持一致）
# 注意：t 和 kg 不能放一组（差1000倍），m 和 m² 也不能放一组
_UNIT_COMPAT_GROUPS = [
    {"kg", "千克", "公斤"},
    {"t", "吨"},
    {"个", "只", "套", "台", "件", "组", "块"},
    {"m", "米"},
    {"m²", "㎡", "平方米", "m2"},
    {"m³", "立方米", "m3"},
    {"根", "条", "支"},
    {"桶", "瓶"},
    {"卷", "盘"},
    {"副", "付", "对"},
]


def check_unit_compatible(unit_a: str, unit_b: str) -> bool:
    """检查两个单位是否兼容（同组内视为兼容）。

    本文件的 _is_usable_gldjc_cache_entry 会调用它判定广材网缓存能否使用；
    之前此函数未定义，导致有 matched_unit 的缓存行直接抛 NameError，
    等同于整条广材网缓存兜底路径不可用。
    """
    if not unit_a or not unit_b:
        return False
    a = str(unit_a).strip().lower()
    b = str(unit_b).strip().lower()
    if a == b:
        return True
    for group in _UNIT_COMPAT_GROUPS:
        lower_group = {u.lower() for u in group}
        if a in lower_group and b in lower_group:
            return True
    return False


def _is_convertible_price_unit(target_unit: str, source_unit: str, name: str = "", spec: str = "") -> bool:
    if not target_unit or not source_unit:
        return False
    if check_unit_compatible(target_unit, source_unit):
        return True

    import sys

    tools_dir = str(Path(__file__).resolve().parents[4] / "tools")
    if tools_dir not in sys.path:
        sys.path.insert(0, tools_dir)

    try:
        from gldjc_price import _try_convert_price
    except Exception:
        return False

    return _try_convert_price(1.0, source_unit, target_unit, name, spec) is not None


def _is_usable_gldjc_cache_entry(name: str, spec: str, unit: str, cached: dict) -> bool:
    if not isinstance(cached, dict):
        return False
    if not cached.get("price_with_tax"):
        return False
    if _looks_like_installation_item_name(name):
        return False

    matched_unit = str(cached.get("matched_unit") or "").strip()
    if unit and matched_unit and not _is_convertible_price_unit(unit, matched_unit, name, spec):
        return False

    matched_spec = str(cached.get("matched_spec") or "").strip()
    if spec:
        if not matched_spec:
            return False
        if not _spec_contains_exact(matched_spec, spec):
            return False

    return True


# _is_compatible_material_hint 和 _extract_material_from_desc 的第二版已并入顶部定义，移除重复。


def _extract_critical_spec_text(raw_spec: str, main_spec: str = "") -> str:
    text = str(raw_spec or "").strip()
    if not text:
        return ""

    main_text = str(main_spec or "").strip()
    normalized = text
    if main_text:
        normalized = re.sub(re.escape(main_text), "", normalized, count=1, flags=re.IGNORECASE)

    normalized = normalized.replace("*", " ").replace("×", " ").replace("X", " ")
    normalized = re.sub(r"^[\s,，;；:/：\-\|]+", "", normalized)
    normalized = re.sub(r"[\s,，;；:/：\-\|]+$", "", normalized)
    normalized = re.sub(r"\s+", " ", normalized).strip()

    if not normalized:
        return ""

    critical_tokens = []
    seen: set[str] = set()
    patterns = [
        r"(?:P\s*=\s*)?\d+(?:\.\d+)?\s*MPa",
        r"PN\s*\d+(?:\.\d+)?",
        r"S\d+(?:\.\d+)?",
        r"SDR\s*\d+(?:\.\d+)?",
        r"SN\s*\d+(?:\.\d+)?",
        r"PE100",
        r"PE80",
        r"HDPE",
        r"UPVC|U-PVC|PVC-U|PVC",
        r"PPR|PP-R|PE-RT|PERT|PE",
    ]
    for pattern in patterns:
        for matched in re.findall(pattern, normalized, flags=re.IGNORECASE):
            token = re.sub(r"\s+", "", str(matched)).strip()
            if token and token.lower() not in seen:
                seen.add(token.lower())
                critical_tokens.append(token)

    if critical_tokens:
        return " ".join(critical_tokens)

    if len(normalized) <= 40 and not re.search(r"[\u4e00-\u9fff]", normalized):
        return normalized
    return ""


def _extract_raw_spec_from_desc(desc: str) -> str:
    if not desc:
        return ""

    pairs: dict[str, str] = {}
    normalized_desc = re.sub(r"(?<!\S)(\d+)\.\s*", r"\n\1.", str(desc or ""))
    for raw_line in normalized_desc.splitlines():
        line = re.sub(r"^\s*\d+[\.、\s]*", "", raw_line.strip())
        if not line:
            continue
        parts = re.split(r"[:：]", line, maxsplit=1)
        if len(parts) != 2:
            continue
        key = _normalize_desc_key(parts[0].strip())
        value = parts[1].strip()
        if key and value:
            pairs[key] = value

    for key in ("材质规格", "材质及规格", "材质型号", "规格压力等级", "型号规格", "规格型号", "规格", "型号", "公称直径"):
        value = str(pairs.get(key) or "").strip()
        if value:
            return value
    return ""


def _infer_material_object_type(name: str, bill_name: str = "", candidate_type: str = "") -> str:
    primary_text = str(name or "").strip()
    text = " ".join(part for part in [primary_text, str(candidate_type or "").strip(), str(bill_name or "").strip()] if part)
    if not text:
        return ""
    if _looks_like_installation_item_name(primary_text):
        return "installation_item"
    if any(token in text for token in ("阀", "过滤器", "减压器", "减压阀", "止回")):
        return "valve"
    if any(token in text for token in ("管件", "弯头", "三通", "四通", "异径", "接头", "传力接头", "补偿器", "法兰盘")):
        return "pipe_fitting"
    if primary_text in {"法兰", "松套法兰", "活套法兰", "盲板法兰"}:
        return "pipe_fitting"
    if "管" in text:
        return "pipe"
    if any(token in text for token in ("泵", "机组", "设备", "水箱", "风机")):
        return "equipment"
    if any(token in text for token in ("地漏", "洁具", "器", "箱", "表")):
        return "device"
    return "material"


def _infer_material_family(object_type: str, normalized_name: str) -> str:
    if object_type:
        return object_type
    family = _material_hint_family(normalized_name)
    if family == "pipe":
        return "pipe"
    if family == "valve":
        return "valve"
    if family == "device":
        return "device"
    return "material"


def _should_keep_original_name_for_object_type_conflict(material_name: str, suggested_name: str) -> bool:
    original_type = _infer_material_object_type(material_name)
    suggested_type = _infer_material_object_type(suggested_name)
    strong_types = {"pipe", "pipe_fitting", "valve", "equipment", "device"}
    if original_type not in strong_types:
        return False
    if not suggested_type or suggested_type in {original_type, "material", "installation_item"}:
        return False
    return suggested_type != original_type


def _build_normalized_material_fields(
    material_name: str,
    material_spec: str,
    bill_name: str = "",
    desc: str = "",
) -> dict[str, str]:
    info = _extract_material_from_desc(desc)
    suggested_name, suggested_spec = _suggest_material_from_bill_context(material_name, bill_name, desc)
    if suggested_name and _should_keep_original_name_for_object_type_conflict(material_name, suggested_name):
        suggested_name = ""

    normalized_name = (suggested_name or material_name or "").strip()
    normalized_spec = (suggested_spec or _clean_material_spec(material_spec) or "").strip()

    raw_spec_candidates = [
        material_spec,
        _extract_raw_spec_from_desc(desc),
        str(info.get("spec") or "").strip(),
    ]
    raw_spec_text = next((item.strip() for item in raw_spec_candidates if str(item or "").strip()), "")
    critical_spec_text = _extract_critical_spec_text(raw_spec_text, normalized_spec)

    object_type = _infer_material_object_type(
        normalized_name or material_name,
        bill_name=bill_name,
        candidate_type=str(info.get("type") or "").strip(),
    )
    family = _infer_material_family(object_type, normalized_name or material_name)

    query_parts = [normalized_name, normalized_spec, critical_spec_text]
    normalized_query_text = " ".join(part for part in query_parts if str(part or "").strip()).strip()

    confidence = "low"
    if normalized_name and normalized_spec and critical_spec_text:
        confidence = "high"
    elif normalized_name and normalized_spec:
        confidence = "medium"
    elif normalized_name:
        confidence = "low"

    return {
        "suggested_name": suggested_name,
        "suggested_spec": suggested_spec,
        "normalized_name": normalized_name,
        "normalized_spec": normalized_spec,
        "critical_spec_text": critical_spec_text,
        "normalized_query_text": normalized_query_text,
        "object_type": object_type,
        "family": family,
        "normalization_confidence": confidence,
        "connection_hint": str(info.get("connection") or "").strip(),
        "material_hint": str(info.get("material") or "").strip(),
        "desc_type_hint": str(info.get("type") or "").strip(),
    }


def _extract_inline_material_spec(text: str) -> str:
    value = str(text or "").strip()
    if not value:
        return ""

    patterns = [
        r"(DN\d+(?:[\*xX×]DN?\d+){1,3})",
        r"(DN\d+[A-Za-z0-9,\-\./]*)",
        r"(De\d+[A-Za-z0-9,\-\./]*(?:\s+[A-Za-z0-9,\-\./]+)*)",
        r"(\d+(?:\.\d+)?(?:mm|mm2|㎡|m2))",
    ]
    for pattern in patterns:
        matched = re.search(pattern, value, flags=re.IGNORECASE)
        if matched:
            return matched.group(1).strip()
    return ""


def _strip_inline_spec_from_name(text: str) -> str:
    value = str(text or "").strip()
    if not value:
        return ""

    value = re.sub(
        r"[（(]\s*(?:热熔|螺纹|法兰|沟槽|焊接|承插|胶粘|电熔|卡箍)[^)）]*[)）]",
        " ",
        value,
        flags=re.IGNORECASE,
    )
    spec = _extract_inline_material_spec(value)
    if spec:
        value = re.sub(re.escape(spec), " ", value, count=1, flags=re.IGNORECASE)

    cleanup_patterns = [
        r"(?:,|，)?\s*PN\s*\d+(?:\.\d+)?",
        r"(?:,|，)?\s*P\s*=\s*\d+(?:\.\d+)?\s*MPa?",
        r"(?:,|，)?\s*\d+(?:\.\d+)?\s*MPa",
        r"(?:,|，)?\s*S\d+(?:\.\d+)?",
        r"(?:,|，)?\s*SDR\s*\d+(?:\.\d+)?",
        r"(?:,|，)?\s*SN\s*\d+(?:\.\d+)?",
    ]
    for pattern in cleanup_patterns:
        value = re.sub(pattern, " ", value, flags=re.IGNORECASE)

    value = re.sub(r"^[\s,，;；:/]+|[\s,，;；:/]+$", "", value)
    value = re.sub(r"\s+", " ", value).strip()
    return value


def _is_spec_only_material_name(name: str) -> bool:
    raw = str(name or "").strip()
    if not raw:
        return False
    cleaned = _strip_inline_spec_from_name(raw)
    if not cleaned:
        return True
    return (
        not re.search(r"[\u4e00-\u9fffA-Za-z]", cleaned)
        and bool(re.search(r"\d", raw))
    )


def _is_effective_generic_material_name(name: str) -> bool:
    raw = str(name or "").strip()
    if not raw:
        return False
    cleaned = _strip_inline_spec_from_name(raw) or raw
    return _is_generic_material_name(cleaned)


def _normalize_bill_candidate_name(name: str) -> str:
    value = str(name or "").strip()
    if not value:
        return ""
    value = re.sub(r"[（(][^)）]*[)）]", " ", value)
    value = _strip_inline_spec_from_name(value)
    value = re.sub(r"\b(?:室内|室外|户内|户外|地上|地下)\b", " ", value)
    value = re.sub(r"\s+", " ", value).strip()
    return value


def _is_viable_material_candidate(name: str) -> bool:
    value = str(name or "").strip()
    if not value:
        return False
    if _is_spec_only_material_name(value):
        return False
    if _looks_like_installation_item_name(value):
        return False
    if _is_effective_generic_material_name(value):
        return False
    return True


# _is_generic_material_name / _should_use_bill_type_for_material /
# _connection_prefix / _is_specific_lookup_name / _should_prefix_connection /
# _should_prefix_material / _looks_like_installation_item_name
# 的重复定义已清理，统一使用文件上方的唯一实现。


def _looks_like_bare_material_token(name: str) -> bool:
    value = _normalize_material_hint(name)
    bare_tokens = {
        "pe", "pvc", "upvc", "cpvc", "ppr", "hdpe", "pert", "pex", "frpp", "pp", "abs", "pb",
        "塑料", "钢", "钢材", "钢制", "不锈钢", "铜", "黄铜", "铸铁", "球墨铸铁",
    }
    return value in {_normalize_material_hint(x) for x in bare_tokens}


def _looks_like_material_family_token(name: str) -> bool:
    value = _normalize_material_hint(name)
    families = {"upvc", "pvcu", "pvc", "cpvc", "ppr", "pprct", "pe", "hdpe", "pert", "pex", "pb", "abs", "frpp", "pp"}
    return value in families


def _should_prefix_material_family(material_name: str, candidate_name: str) -> bool:
    if not _looks_like_material_family_token(candidate_name):
        return False
    material_text = str(material_name or "").strip()
    if not any(token in material_text for token in ("管", "管件")):
        return False
    return _normalize_material_hint(candidate_name) not in _normalize_material_hint(material_text)


def _is_conduit_accessory_name(name: str) -> bool:
    text = str(name or "").strip()
    if not text:
        return False
    accessory_tokens = (
        "锁紧螺母",
        "管接头",
        "护口",
        "接线盒",
        "接线箱",
        "线盒",
        "拉线盒",
        "过线盒",
        "分线盒",
        "86盒",
    )
    return any(token in text for token in accessory_tokens)


def _should_use_conduit_material_name(material_name: str, bill_name: str, candidate_material: str) -> bool:
    material_text = str(material_name or "").strip()
    bill_text = str(bill_name or "").strip()
    candidate = str(candidate_material or "").strip()
    if not candidate:
        return False
    if _is_conduit_accessory_name(material_text):
        return False
    if any(token in candidate for token in ("综合考虑", "详见")):
        return False
    conduit_tokens = ("配管", "线管", "导管", "电管", "穿线管", "钢管公称口径", "薄壁钢管")
    text = f"{material_text} {bill_text}"
    if not any(token in text for token in conduit_tokens):
        return False
    return True


def _should_merge_device_name(material_name: str, bill_name: str, candidate_type: str, candidate_material: str) -> bool:
    device_name = str(candidate_type or bill_name or "").strip()
    if not device_name:
        return False
    if not any(token in device_name for token in ("地漏", "盆", "洁具", "器", "卫生")):
        return False
    material_text = str(material_name or "").strip()
    if not material_text:
        return False
    if any(token in material_text for token in ("地漏", "盆", "洁具", "器", "卫生")):
        return False
    if candidate_material and material_text == candidate_material.strip():
        return True
    return any(token in material_text for token in ("不锈钢", "黄铜", "陶瓷", "塑料", "铸铁", "木质", "+"))


def _clean_equipment_candidate_name(name: str) -> str:
    value = str(name or "").strip()
    if not value:
        return ""
    value = re.sub(r"[（(][^）)]*(?:甲供|乙供|自带|设备甲供|主材甲供)[^）)]*[）)]", "", value, flags=re.IGNORECASE)
    value = re.sub(r"(设备甲供|主材甲供|甲供材|甲供|乙供|设备自带|自带)", "", value, flags=re.IGNORECASE)
    value = re.sub(r"^[\s,，:：;/；]+|[\s,，:：;/；]+$", "", value)
    value = re.sub(r"\s+", " ", value).strip()
    return value


def _should_use_specific_equipment_name(material_name: str, bill_name: str, candidate_type: str, candidate_name: str) -> bool:
    reference_name = str(material_name or bill_name or "").strip()
    candidate = _clean_equipment_candidate_name(candidate_type or candidate_name)
    if not reference_name or not candidate:
        return False
    equipment_tokens = ("泵", "风机", "机组", "设备", "水箱", "换热器", "冷水机", "空调", "配电箱", "配电柜", "控制箱", "控制柜")
    if not any(token in candidate for token in equipment_tokens):
        return False
    if _normalize_material_hint(candidate) == _normalize_material_hint(reference_name):
        return False
    generic_reference_names = {"泵", "水泵", "离心泵", "离心式泵", "风机", "机组", "设备", "水箱"}
    normalized_reference = _normalize_material_hint(reference_name)
    normalized_generic_refs = {_normalize_material_hint(x) for x in generic_reference_names}
    return normalized_reference in normalized_generic_refs or len(candidate) > len(reference_name)


def _is_generic_pipe_fitting_name(name: str) -> bool:
    text = str(name or "").strip()
    if "管件" not in text:
        return False
    specific_tokens = ("弯头", "三通", "四通", "异径", "盲板", "接头", "补偿器", "伸缩节", "止流器")
    return not any(token in text for token in specific_tokens)


def _should_use_pipe_fitting_type(material_name: str, candidate_type: str, candidate_name: str) -> bool:
    if not _is_generic_pipe_fitting_name(material_name):
        return False
    explicit_type = str(candidate_type or "").strip()
    if explicit_type:
        return _is_specific_pipe_fitting_type(explicit_type) or _is_specific_lookup_name(explicit_type)
    return _is_specific_pipe_fitting_type(candidate_name)


def _is_specific_pipe_fitting_type(text: str) -> bool:
    value = str(text or "").strip()
    if not value:
        return False
    specific_tokens = (
        "弯头", "三通", "四通", "异径", "大小头", "盲板", "法兰", "接头",
        "传力接头", "补偿器", "伸缩节", "止流器", "短管",
    )
    return any(token in value for token in specific_tokens)


def _compose_pipe_fitting_name(material_name: str, fitting_type: str, candidate_material: str) -> str:
    fitting = str(fitting_type or "").strip()
    if not fitting:
        return ""
    if re.search(r"[A-Za-z0-9]", fitting):
        return fitting
    if any(token in fitting for token in ("传力接头", "补偿器", "伸缩节", "止流器")):
        return fitting

    material_prefix = ""
    material_text = str(candidate_material or "").strip()
    if material_text and material_text not in fitting:
        material_prefix = material_text

    qualifier = ""
    for token in ("对焊", "承插", "法兰", "螺纹", "沟槽", "焊接"):
        if token in str(material_name or "") and token not in fitting:
            qualifier = token
            break

    return f"{material_prefix}{qualifier}{fitting}".strip()


def _should_prefix_specific_pipe_fitting(
    material_name: str,
    bill_name: str,
    candidate_type: str,
    candidate_material: str,
) -> bool:
    fitting = str(candidate_type or "").strip()
    material_text = str(candidate_material or "").strip()
    if not fitting or not material_text:
        return False
    if not _is_specific_pipe_fitting_type(fitting):
        return False
    if material_text in fitting:
        return False

    normalized_fitting = _normalize_material_hint(fitting)
    normalized_material_name = _normalize_material_hint(material_name)
    normalized_bill_name = _normalize_material_hint(bill_name)
    if normalized_material_name == normalized_fitting:
        return True
    if normalized_bill_name == normalized_fitting:
        return True
    if fitting and fitting in str(bill_name or ""):
        return True
    return "综合" in str(material_name or "")


def _compose_specific_pipe_fitting_name(candidate_type: str, candidate_material: str) -> str:
    fitting = str(candidate_type or "").strip()
    material_text = str(candidate_material or "").strip()
    if not fitting:
        return ""
    if not material_text or material_text in fitting:
        return fitting
    return f"{material_text}{fitting}"


@router.get("/material-price/from-task/{task_id}")
async def parse_from_task(task_id: str):
    """从已完成的套定额任务中提取主材行

    读取任务的output Excel（带主材版本），解析出主材行。
    """
    import uuid as _uuid
    from sqlalchemy import select
    from app.database import async_session
    from app.models.task import Task

    # 查任务
    try:
        task_uuid = _uuid.UUID(task_id)
    except ValueError:
        raise HTTPException(400, "无效的任务ID")

    async with async_session() as session:
        result = await session.execute(
            select(Task).where(Task.id == task_uuid)
        )
        task = result.scalar_one_or_none()

    if not task:
        raise HTTPException(404, "任务不存在")
    if task.status != "completed":
        raise HTTPException(400, "任务尚未完成")
    if not task.output_path or not Path(task.output_path).exists():
        raise HTTPException(404, "输出文件不存在")

    # 解析output Excel中的主材行
    try:
        result = await asyncio.to_thread(_do_parse, task.output_path)
        materials = result["materials"]
        # 用任务的output_path作为file_key（不需要复制，直接指向）
        file_key = f"task-{task_id}"
        _uploaded_files[file_key] = {
            "path": task.output_path,
            "name": Path(task.original_filename or "").stem or f"task_{task_id[:8]}",
        }
        return {
            "materials": materials,
            "all_rows": result["all_rows"],
            "is_mixed": result["is_mixed"],
            "count": len(materials),
            "task_name": task.original_filename or "",
            "file_key": file_key,
        }
    except Exception as e:
        logger.error(f"从任务提取主材失败: {e}")
        raise HTTPException(500, f"提取失败: {e}")


# ============================================================
# 5. 批量查价
# ============================================================

@router.post("/material-price/lookup")
async def lookup_prices(body: dict):
    """根据主材列表和地区信息，批量查价

    请求体:
    {
        "materials": [{"name": "镀锌钢管", "spec": "DN25", "unit": "m"}, ...],
        "province": "湖北",
        "city": "武汉",           // 可选
        "period_end": "2026-02-28",  // 可选，指定期次
        "price_type": "info"      // 可选：info=信息价, market=市场价；普通查价默认只查信息价
    }
    """
    materials = body.get("materials", [])
    province = body.get("province", "")
    city = body.get("city", "")
    period_end = body.get("period_end", "")
    price_type = body.get("price_type", "info")

    if not materials:
        raise HTTPException(400, "materials 不能为空")
    if not province:
        raise HTTPException(400, "province 不能为空")

    # 远程模式：转发到本地匹配服务
    if _is_remote():
        result = await _remote_post("/material-price/lookup", {
            "materials": materials,
            "province": province,
            "city": city,
            "period_end": period_end,
            "price_type": price_type,
        })
        return result or {"results": [], "stats": {"total": 0, "found": 0, "not_found": 0}}

    # 本地模式：直接查价
    results = await asyncio.to_thread(
        _do_lookup, materials, province, city, period_end, price_type
    )
    # 统计
    found = sum(1 for r in results if r.get("lookup_price") is not None)
    return {
        "results": results,
        "stats": {
            "total": len(results),
            "found": found,
            "not_found": len(results) - found,
        }
    }


def _do_lookup(materials: list[dict], province: str, city: str,
               period_end: str, price_type: str = "info") -> list[dict]:
    """批量查价核心逻辑

    price_type: info=只查信息价, market=只查市场价
    普通查价不再使用本地市场价或广材网市场价缓存兜底。
    """
    db = _get_db() if _material_db_ready() else None

    # 价格类型映射到 source_type 过滤条件
    source_filter = None
    if price_type in {"all", "info"}:
        source_filter = "government"  # 信息价
    elif price_type == "market":
        source_filter = "market"      # 市场价

    results = []
    for mat in materials:
        name = mat.get("name", "").strip()
        spec = mat.get("spec", "").strip()
        unit = mat.get("unit", "").strip()
        object_type = str(mat.get("object_type") or "").strip() or _infer_material_object_type(name)

        if not name:
            results.append({**mat, "lookup_price": None, "lookup_source": "名称为空"})
            continue

        # 用 MaterialDB 的查价方法
        kwargs = dict(
            province=province,
            city=city,
            period_end=period_end,
            spec=spec,
            target_unit=unit,
            object_type=object_type,
        )
        if source_filter:
            kwargs["source_type"] = source_filter
        name_is_spec_only = (
            bool(name)
            and not re.search(r"[\u4e00-\u9fff]", name)
            and bool(re.search(r"\d", name))
            and re.sub(r"(DN|DE|MM|CM|Φ|PHI|X|×|\*|/|\.|-|\d|\s)+", "", name.upper()) == ""
        )
        if name_is_spec_only:
            results.append({
                **mat,
                "lookup_price": None,
                "lookup_source": "名称异常",
                "lookup_url": None,
                "lookup_label": None,
            })
            continue

        price_info = None
        if db is not None:
            price_info = db.search_price_by_name(name, **kwargs)

        if price_info:
            results.append({
                **mat,
                "lookup_price": price_info["price"],
                "lookup_source": price_info.get("source", "价格库"),
                "lookup_url": None,
                "lookup_label": _build_lookup_label(
                    name=name,
                    spec=spec,
                    unit=price_info.get("unit") or unit,
                    price=price_info.get("price"),
                    source=price_info.get("source", "价格库"),
                ),
            })
        else:
            # 尝试从名称中提取规格再查一次
            m = re.search(r'[Dd][Nn]\s*\d+|De\s*\d+|Φ\s*\d+|\d+mm', name)
            if m:
                extracted_spec = m.group(0).replace(" ", "")
                short_name = name[:m.start()].strip()
                if short_name:
                    kwargs2 = dict(
                        province=province,
                        city=city,
                        period_end=period_end,
                        spec=extracted_spec,
                        target_unit=unit,
                        object_type=object_type or _infer_material_object_type(short_name),
                    )
                    if source_filter:
                        kwargs2["source_type"] = source_filter
                    price_info2 = db.search_price_by_name(short_name, **kwargs2) if db is not None else None
                    if price_info2:
                        results.append({
                            **mat,
                            "lookup_price": price_info2["price"],
                            "lookup_source": price_info2.get("source", "价格库"),
                            "lookup_url": None,
                            "lookup_label": _build_lookup_label(
                                name=short_name,
                                spec=extracted_spec,
                                unit=price_info2.get("unit") or unit,
                                price=price_info2.get("price"),
                                source=price_info2.get("source", "价格库"),
                            ),
                        })
                        continue

            results.append({**mat, "lookup_price": None, "lookup_source": "未查到", "lookup_url": None, "lookup_label": None})

    return results


# ============================================================
# 6. 用户贡献价格（存入候选层）
# ============================================================

@router.post("/material-price/contribute")
async def contribute_price(body: dict):
    """用户手填价格，存入价格库候选层

    请求体:
    {
        "items": [
            {
                "name": "镀锌钢管",
                "spec": "DN25",
                "unit": "m",
                "price": 18.5,
                "province": "湖北",
                "city": "武汉"
            },
            ...
        ]
    }
    """
    items = body.get("items", [])
    if not items:
        raise HTTPException(400, "items 不能为空")

    # 远程模式：转发到本地匹配服务
    if _is_remote():
        result = await _remote_post("/material-price/contribute", {"items": items})
        return result or {"saved": 0, "message": "转发失败"}

    # 本地模式：直接写入
    saved = await asyncio.to_thread(_do_contribute, items)
    return {"saved": saved, "message": f"已保存 {saved} 条价格，感谢贡献！"}


def _do_contribute(items: list[dict]) -> int:
    """保存用户贡献的价格到候选层"""
    db = _get_db()
    saved = 0

    for item in items:
        name = item.get("name", "").strip()
        spec = item.get("spec", "").strip()
        unit = item.get("unit", "").strip()
        price = item.get("price")
        province = item.get("province", "").strip()
        city = item.get("city", "").strip()

        if not name or price is None:
            continue

        try:
            price_val = float(price)
        except (ValueError, TypeError):
            continue

        # 异常值检测：价格为负数或超过100万，视为异常
        if price_val <= 0 or price_val > 1_000_000:
            logger.warning(f"用户贡献异常价格: {name} {spec} = {price_val}，已跳过")
            continue

        # 先确保材料主数据存在
        material_id = db.add_material(name, spec=spec, unit=unit)

        # 存入价格库（候选层：authority_level='reference'）
        db.add_price(
            material_id=material_id,
            price_incl_tax=price_val,
            source_type="user_contribute",
            province=province,
            city=city,
            unit=unit,
            authority_level="reference",
            source_doc="用户手填",
            dedup=True,
        )
        saved += 1

    return saved


# ============================================================
# 7. 导出：把价格写回原Excel的主材行单价列
# ============================================================

@router.post("/material-price/export")
async def export_with_prices(body: dict):
    """把查到的价格写回到原Excel的主材行单价列，返回修改后的Excel下载

    请求体:
    {
        "file_key": "xxx",  // parse返回的file_key
        "materials": [      // 带价格的主材列表（前端合并后的最终结果）
            {
                "row": 10,
                "sheet": "分部分项...",
                "price_col": 7,
                "final_price": 18.5  // 最终价格（手填优先，否则查价，否则原有）
            },
            ...
        ]
    }
    """
    file_key = body.get("file_key", "")
    materials = body.get("materials", [])

    if not file_key:
        raise HTTPException(400, "file_key 不能为空")

    # 找到原文件
    file_info = _uploaded_files.get(file_key)
    if not file_info or not Path(file_info["path"]).exists():
        raise HTTPException(404, "原始文件不存在或已过期，请重新上传")

    source_path = file_info["path"]
    original_name = file_info.get("name", "output")  # 原始文件名（不含后缀）

    # 复制一份到临时文件（不改原文件）
    import shutil
    suffix = Path(source_path).suffix
    with tempfile.NamedTemporaryFile(
        mode="wb", suffix=suffix, delete=False, prefix="material_export_"
    ) as tmp:
        tmp_path = tmp.name
    shutil.copy2(source_path, tmp_path)

    # 在线程池中写入价格
    try:
        written = await asyncio.to_thread(_do_write_material_updates, tmp_path, materials)
        logger.info(f"智能填主材导出：写入 {written} 个价格")

        # 生成下载文件名（用原始文件名，不是临时文件名）
        download_name = f"{original_name}_已填价{suffix}"

        return FileResponse(
            path=tmp_path,
            filename=download_name,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            # FileResponse发送完毕后自动删除临时文件
            background=None,
        )
    except Exception as e:
        Path(tmp_path).unlink(missing_ok=True)
        logger.error(f"写入价格失败: {e}")
        raise HTTPException(500, f"导出失败: {e}")


def _do_write_prices(excel_path: str, materials: list[dict]) -> int:
    """把价格写入Excel的主材行单价列

    按 sheet名+行号+列号 定位每个主材行，写入 final_price。
    """
    import openpyxl

    wb = openpyxl.load_workbook(excel_path)
    written = 0

    # 按sheet分组
    sheet_materials: dict[str, list[dict]] = {}
    for mat in materials:
        sheet_name = mat.get("sheet", "")
        if sheet_name not in sheet_materials:
            sheet_materials[sheet_name] = []
        sheet_materials[sheet_name].append(mat)

    for sheet_name, mats in sheet_materials.items():
        if sheet_name not in wb.sheetnames:
            logger.warning(f"sheet [{sheet_name}] 不存在，跳过 {len(mats)} 条")
            continue

        ws = wb[sheet_name]
        for mat in mats:
            row = mat.get("row")
            price_col = mat.get("price_col")
            final_price = mat.get("final_price")

            if row is None or price_col is None or final_price is None:
                continue

            try:
                price_val = float(final_price)
            except (ValueError, TypeError):
                continue

            # 写入单价
            cell = ws.cell(row=row, column=price_col, value=round(price_val, 2))
            cell.number_format = '0.00'
            written += 1

    wb.save(excel_path)
    wb.close()
    return written


def _shift_column_for_link_insertions(col, anchor_cols: list[int]) -> int | None:
    if col is None:
        return None
    try:
        col_idx = int(col)
    except (TypeError, ValueError):
        return None
    return col_idx + sum(1 for anchor in anchor_cols if anchor < col_idx)


def _resolve_link_column(anchor_col, anchor_cols: list[int]) -> int | None:
    if anchor_col is None:
        return None
    try:
        anchor_idx = int(anchor_col)
    except (TypeError, ValueError):
        return None
    return anchor_idx + 1 + sum(1 for anchor in anchor_cols if anchor < anchor_idx)


def _sanitize_export_lookup_label(label: str, critical_spec_text: str = "") -> str:
    text = str(label or "").strip()
    if not text:
        return ""

    critical = str(critical_spec_text or "").strip()
    if critical:
        patterns = [
            rf"\s*\|\s*关键规格[:：]?\s*{re.escape(critical)}",
            rf"\s*关键规格[:：]?\s*{re.escape(critical)}",
            rf"\s*\|\s*{re.escape(critical)}(?=\s*(?:\||$))",
        ]
        for pattern in patterns:
            text = re.sub(pattern, "", text, flags=re.IGNORECASE)

    text = re.sub(r"\s*\|\s*\|\s*", " | ", text)
    text = re.sub(r"^\s*\|\s*", "", text)
    text = re.sub(r"\s*\|\s*$", "", text)
    return text.strip()


def _do_write_material_updates(excel_path: str, materials: list[dict]) -> int:
    """Write edited material names and prices back into the reviewed Excel."""
    import openpyxl
    from openpyxl.styles import Font

    wb = openpyxl.load_workbook(excel_path)
    written = 0
    link_font = Font(color="0563C1", underline="single")

    sheet_materials: dict[str, list[dict]] = {}
    for mat in materials:
        sheet_name = mat.get("sheet", "")
        if sheet_name not in sheet_materials:
            sheet_materials[sheet_name] = []
        sheet_materials[sheet_name].append(mat)

    for sheet_name, mats in sheet_materials.items():
        if sheet_name not in wb.sheetnames:
            logger.warning(f"sheet [{sheet_name}] not found, skipped {len(mats)} material rows")
            continue

        ws = wb[sheet_name]
        link_col = None
        link_header_row = 1
        for mat in mats:
            lookup_url = str(mat.get("lookup_url") or "").strip()
            if not lookup_url:
                continue
            header_row = mat.get("header_row")
            try:
                link_header_row = int(header_row) if header_row is not None else 1
            except (TypeError, ValueError):
                link_header_row = 1
            link_col = ws.max_column + 1
            ws.cell(row=link_header_row, column=link_col, value="广材网链接")
            break

        for mat in mats:
            row = mat.get("row")
            if row is None:
                continue

            row_written = False

            name_col = mat.get("name_col")
            final_name = str(mat.get("final_name") or "").strip()
            spec_col = mat.get("spec_col")
            final_spec = str(mat.get("final_spec") or "").strip()
            write_name = final_name
            if spec_col is None and final_spec and final_spec not in final_name:
                write_name = f"{final_name} {final_spec}".strip()

            if name_col is not None and write_name:
                ws.cell(row=row, column=name_col, value=write_name)
                row_written = True

            if spec_col is not None:
                ws.cell(row=row, column=spec_col, value=final_spec)
                row_written = True

            lookup_url = str(mat.get("lookup_url") or "").strip()
            if lookup_url and link_col is not None:
                link_text = _sanitize_export_lookup_label(
                    str(mat.get("lookup_label") or lookup_url).strip() or lookup_url,
                    str(mat.get("critical_spec_text") or "").strip(),
                ) or lookup_url
                link_cell = ws.cell(row=row, column=link_col, value=link_text)
                link_cell.hyperlink = lookup_url
                link_cell.font = link_font
                row_written = True

            price_col = mat.get("price_col")
            final_price = mat.get("final_price")
            if price_col is not None and final_price is not None:
                try:
                    price_val = float(final_price)
                except (ValueError, TypeError):
                    price_val = None

                if price_val is not None:
                    cell = ws.cell(row=row, column=price_col, value=round(price_val, 2))
                    cell.number_format = '0.00'
                    row_written = True

            if row_written:
                written += 1

    wb.save(excel_path)
    wb.close()
    return written


# ============================================================
# 8. 广材网实时查价（管理员专用）
# ============================================================

@router.post("/material-price/gldjc-lookup")
async def gldjc_lookup(body: dict):
    """管理员专用：对DB查不到的材料，实时爬广材网查价

    请求体:
    {
        "materials": [{"name": "镀锌钢管 DN25", "unit": "m", "spec": "DN25"}, ...],
        "cookie": "token=bearer xxx"  // 广材网登录Cookie
    }

    返回: { results: [...], total, found }
    """
    materials = body.get("materials", [])
    cookie = body.get("cookie", "").strip()
    province = body.get("province", "").strip()
    city = body.get("city", "").strip()
    period_end = body.get("period_end", "").strip()

    if not cookie:
        raise HTTPException(400, "请输入广材网Cookie")
    if not materials:
        raise HTTPException(400, "材料列表为空")

    # 远程模式：转发到本地匹配服务（广材网Cookie绑定IP，必须从本机发请求）
    if _is_remote():
        import httpx
        url = f"{LOCAL_MATCH_URL.rstrip('/')}/material-price/gldjc-lookup"
        headers = {"X-API-Key": LOCAL_MATCH_API_KEY}
        try:
            async with local_match_async_client(timeout=600.0) as client:
                resp = await client.post(url, headers=headers, json={
                    "materials": materials,
                    "cookie": cookie,
                    "province": province,
                    "city": city,
                    "period_end": period_end,
                })
            if resp.status_code == 200:
                return resp.json()
            logger.warning(f"远程广材网查价返回 {resp.status_code}: {resp.text[:200]}")
            raise HTTPException(resp.status_code, resp.json().get("detail", "远程查价失败"))
        except httpx.HTTPError as e:
            logger.error(f"远程广材网查价失败: {e}")
            raise HTTPException(502, f"连接本地匹配服务失败: {e}")

    # 本地模式：直接在线程池中执行
    results = await asyncio.to_thread(
        _do_gldjc_lookup, materials, cookie, province, city, period_end
    )

    found = sum(1 for r in results if r.get("gldjc_price"))
    return {
        "results": results,
        "total": len(results),
        "found": found,
    }


@router.post("/material-price/gldjc-cookie-verify")
async def gldjc_cookie_verify(body: dict):
    """快速验证广材网Cookie是否有效。"""
    cookie = body.get("cookie", "").strip()
    province = body.get("province", "").strip()
    city = body.get("city", "").strip()

    if not cookie:
        raise HTTPException(400, "请输入广材网Cookie")

    if _is_remote():
        import httpx

        url = f"{LOCAL_MATCH_URL.rstrip('/')}/material-price/gldjc-cookie-verify"
        headers = {"X-API-Key": LOCAL_MATCH_API_KEY}
        try:
            async with local_match_async_client(timeout=60.0) as client:
                resp = await client.post(url, headers=headers, json={
                    "cookie": cookie,
                    "province": province,
                    "city": city,
                })
            if resp.status_code == 200:
                return resp.json()
            logger.warning(f"远程广材网Cookie校验返回 {resp.status_code}: {resp.text[:200]}")
            detail = "远程Cookie校验失败"
            try:
                detail = resp.json().get("detail", detail)
            except Exception:
                if resp.text.strip():
                    detail = resp.text.strip()[:200]
            raise HTTPException(resp.status_code, detail)
        except httpx.HTTPError as e:
            logger.error(f"远程广材网Cookie校验失败: {e}")
            raise HTTPException(502, f"连接本地匹配服务失败: {e}")

    from pathlib import Path
    import requests as _requests
    import sys

    tools_dir = str(Path(__file__).resolve().parents[4] / "tools")
    if tools_dir not in sys.path:
        sys.path.insert(0, tools_dir)

    from gldjc_price import SEARCH_URL, _get_headers, resolve_gldjc_area_code

    verify_keyword = "焊接钢管 DN80"
    session = _requests.Session()
    for part in re.split(r";\s*", cookie):
        if "=" in part:
            key, value = part.split("=", 1)
            session.cookies.set(key.strip(), value.strip())

    region_plans: list[tuple[str, str]] = []
    province_code = resolve_gldjc_area_code(province=province, city="")
    if province_code and province_code != "1":
        region_plans.append((province or "省级", province_code))
    region_plans.append(("全国", "1"))

    seen_codes: set[str] = set()
    checks: list[dict] = []
    for scope_label, area_code in region_plans:
        if area_code in seen_codes:
            continue
        seen_codes.add(area_code)
        try:
            resp = session.get(
                SEARCH_URL,
                params={"keyword": verify_keyword, "l": area_code},
                headers=_get_headers(),
                timeout=20,
            )
            resp.raise_for_status()
        except _requests.RequestException as e:
            return {
                "ok": False,
                "status": "limited",
                "message": f"请求广材网失败: {e}",
                "keyword": verify_keyword,
                "scope": scope_label,
                "area_code": area_code,
                "url": _build_gldjc_search_url(keyword=verify_keyword, province_code=area_code),
                "checks": checks,
            }

        html = resp.text or ""
        result_count = len(re.findall(r'class="price-block"', html))
        lower = html.lower()
        if "请登录" in html or "登录后" in html or ("login" in lower and "token" not in lower):
            status = "invalid"
            message = "Cookie已失效，请重新复制完整Cookie"
        elif any(flag in html for flag in ("访问过于频繁", "请求过于频繁", "安全验证", "行为验证", "验证码")):
            status = "limited"
            message = "广材网当前触发限制，请稍后重试或更换Cookie"
        elif result_count > 0:
            status = "valid"
            message = "Cookie有效"
        else:
            status = "limited"
            message = "未命中测试结果，疑似Cookie受限或页面异常"

        check = {
            "scope": scope_label,
            "area_code": area_code,
            "result_count": result_count,
            "status": status,
            "url": _build_gldjc_search_url(keyword=verify_keyword, province_code=area_code),
        }
        checks.append(check)

        if status in {"valid", "invalid"}:
            return {
                "ok": status == "valid",
                "status": status,
                "message": message,
                "keyword": verify_keyword,
                "scope": scope_label,
                "area_code": area_code,
                "url": check["url"],
                "checks": checks,
            }

    last_check = checks[-1] if checks else {
        "scope": province or "全国",
        "area_code": province_code or "1",
        "url": _build_gldjc_search_url(keyword=verify_keyword, province_code=province_code or "1"),
    }
    return {
        "ok": False,
        "status": "limited",
        "message": "未命中测试结果，疑似Cookie受限或页面异常",
        "keyword": verify_keyword,
        "scope": last_check["scope"],
        "area_code": last_check["area_code"],
        "url": last_check["url"],
        "checks": checks,
    }


def _do_gldjc_lookup(
    materials: list[dict],
    cookie: str,
    province: str = "",
    city: str = "",
    period_end: str = "",
) -> list[dict]:
    """实时调用广材网搜索+打分+缓存（同步，在线程池中运行）

    防封策略：
    1. 单次上限30条（前端也应校验）
    2. 随机间隔5~8秒（模拟人工浏览）
    3. 检测登录失效/被封，立即停止
    4. 每10条保存一次缓存（中途中断不丢数据）
    5. 搜索结果去重（同关键词只搜一次）
    """
    import sys
    import time
    import random

    # 安全上限：单次最多30条实时搜索
    MAX_BATCH = 30

    # 把 tools/ 加入路径，复用 gldjc_price.py 的核心函数
    tools_dir = str(Path(__file__).resolve().parents[4] / "tools")
    if tools_dir not in sys.path:
        sys.path.insert(0, tools_dir)

    from gldjc_price import (
        parse_material, search_material_web, filter_and_score,
        get_representative_price_result, build_match_label, determine_confidence,
        build_approximate_price_candidates,
        GldjcCookieInvalidError,
        load_cache, save_cache, update_cache, check_cache, build_region_search_plans,
    )
    import requests as _requests
    from datetime import datetime

    # 创建带Cookie的session
    session = _requests.Session()
    for part in re.split(r";\s*", cookie):
        if "=" in part:
            key, value = part.split("=", 1)
            session.cookies.set(key.strip(), value.strip())

    cache = load_cache()
    results = []

    # 本次搜索去重：同关键词只搜一次
    search_dedup: dict[str, list] = {}

    # 计数：实际发起的网络请求数（不含缓存命中）
    net_requests = 0
    blocked = False  # 是否被封/登录失效
    blocked_reason = "已暂停（疑似被限制）"

    for i, mat in enumerate(materials):
        name = mat.get("name", "").strip()
        unit = mat.get("unit", "").strip()
        spec = mat.get("spec", "").strip()
        lookup_url = _build_gldjc_search_url(name=name, spec=spec)

        if not name:
            results.append({**mat, "gldjc_price": None, "gldjc_source": "名称为空", "gldjc_url": None})
            continue

        # 实时查价默认绕过旧缓存，避免历史误匹配结果被重复复用

        # 安全检查：网络请求超过上限，剩余全部跳过
        if net_requests >= MAX_BATCH:
            results.append({
                **mat, "gldjc_price": None,
                "gldjc_source": f"已达单次上限{MAX_BATCH}条，请分批查询",
                "gldjc_url": lookup_url,
            })
            continue

        # 被封检测：之前发现异常则不再请求
        if blocked:
            results.append({**mat, "gldjc_price": None, "gldjc_source": blocked_reason, "gldjc_url": lookup_url})
            continue

        # 实时搜索广材网
        parsed = parse_material(name, spec)
        base_name = parsed["base_name"]
        specs = parsed["specs"]
        search_plans = build_region_search_plans(parsed["search_keywords"], province=province, city=city)
        link_keyword = parsed["search_keywords"][0] if parsed["search_keywords"] else (f"{base_name} {specs[0]}".strip() if specs else base_name)
        link_plan = search_plans[0] if search_plans else {"keyword": link_keyword, "area_code": "1", "scope": "全国"}
        lookup_url = _build_gldjc_search_url(
            name=name,
            spec=spec,
            keyword=link_plan["keyword"],
            province_code=link_plan["area_code"],
        )

        all_results = []
        searched_keyword = ""
        searched_scope = ""
        searched_area_code = link_plan["area_code"]
        for plan in search_plans:
            kw = plan["keyword"]
            area_code = plan["area_code"]
            scope = plan["scope"]
            dedup_key = f"{area_code}|{kw}"
            # 搜索去重：同关键词复用结果
            if dedup_key in search_dedup:
                all_results = search_dedup[dedup_key]
                searched_keyword = kw
                searched_scope = scope
                searched_area_code = area_code
                if all_results:
                    break
                continue

            # 随机间隔（模拟人工浏览，5~8秒）
            if net_requests > 0:
                delay = random.uniform(5, 8)
                time.sleep(delay)

            try:
                web_results = search_material_web(session, kw, area_code)
            except GldjcCookieInvalidError as exc:
                logger.warning(f"广材网 Cookie 失效，停止后续查询: {exc}")
                blocked = True
                blocked_reason = str(exc)
                break
            net_requests += 1
            searched_keyword = kw
            searched_scope = scope
            searched_area_code = area_code
            search_dedup[dedup_key] = web_results

            # 检测被封/登录失效（搜索函数返回空列表可能是正常无结果，
            # 但如果连续3次都空，大概率有问题）
            if not web_results and net_requests >= 3:
                recent_empty = sum(
                    1 for r in list(search_dedup.values())[-3:]
                    if not r
                )
                if recent_empty >= 3:
                    logger.warning("广材网连续3次搜索无结果，疑似Cookie失效或被限制，停止查询")
                    blocked = True

            if web_results:
                all_results = web_results
                break

            # 关键词降级间隔（较短，因为主间隔已经够长）
            time.sleep(random.uniform(1, 2))

        if blocked and not all_results:
            results.append({**mat, "gldjc_price": None, "gldjc_source": blocked_reason, "gldjc_url": lookup_url, "gldjc_label": None})
            continue

        if not all_results:
            # 搜不到，缓存"未匹配"避免重复搜
            update_cache(cache, name, unit, {
                "price_with_tax": None, "confidence": "低",
                "match_status": "未匹配", "source": "广材网",
                "query_date": datetime.now().strftime("%Y-%m-%d"),
                "result_count": 0,
                "searched_keyword": searched_keyword,
                "gldjc_url": None,
                "gldjc_scope": province or "全国",
            }, spec=spec, region=province or "全国")
            results.append({**mat, "gldjc_price": None, "gldjc_source": "未查到", "gldjc_url": None, "gldjc_label": None})
            continue

        # 打分过滤
        scored = filter_and_score(all_results, unit, specs, base_name, request_name=name, request_spec=spec)
        if not scored:
            scored = filter_and_score(
                all_results,
                unit,
                specs,
                base_name,
                request_name=name,
                request_spec=spec,
                allow_relaxed_spec=True,
            )
        confidence = determine_confidence(scored, unit, specs)
        approximate_candidates = []
        price_source = scored
        if not price_source:
            approximate_candidates = build_approximate_price_candidates(
                all_results,
                unit,
                request_name=name,
                request_spec=spec,
            )
            price_source = approximate_candidates
            confidence = "低"
        used_approximate = bool(approximate_candidates)
        selected_result = get_representative_price_result(price_source)
        exact_detail_url = str(
            (selected_result or {}).get("detail_url")
            or (selected_result or {}).get("url")
            or ""
        ).strip()
        selected_price = None
        if selected_result:
            try:
                selected_price = round(float(selected_result.get("market_price")), 2)
            except (TypeError, ValueError):
                selected_price = None
        match_label = build_match_label(
            selected_result,
            fallback_text=(searched_keyword or name),
        )
        if used_approximate:
            match_status = "近似匹配"
            if match_label:
                match_label = f"近似价 | {match_label}"
        else:
            match_status = "精确匹配" if confidence == "高" else ("模糊匹配" if confidence == "中" else "低置信度")
        scope_label = searched_scope or province or "全国"
        lookup_url = exact_detail_url or _build_gldjc_search_url(
            name=name,
            spec=spec,
            keyword=searched_keyword or link_keyword,
            province_code=searched_area_code or "1",
        )

        if selected_price:
            update_cache(cache, name, unit, {
                "price_with_tax": selected_price,
                "price_without_tax": round(selected_price / 1.13, 2),
                "confidence": confidence,
                "match_status": match_status,
                "source": "广材网",
                "query_date": datetime.now().strftime("%Y-%m-%d"),
                "result_count": len(price_source),
                "searched_keyword": searched_keyword,
                "gldjc_url": lookup_url,
                "match_label": match_label,
                "gldjc_scope": scope_label,
            }, spec=spec, region=scope_label)
            results.append({
                **mat,
                "gldjc_price": selected_price,
                "gldjc_source": f"{'广材网近似价' if used_approximate else '广材网市场价'}({scope_label})",
                "gldjc_url": lookup_url,
                "gldjc_label": match_label,
            })
        else:
            update_cache(cache, name, unit, {
                "price_with_tax": selected_price,
                "price_without_tax": round(selected_price / 1.13, 2) if selected_price else None,
                "confidence": "低",
                "match_status": "低置信度",
                "source": "广材网",
                "query_date": datetime.now().strftime("%Y-%m-%d"),
                "result_count": len(price_source),
                "searched_keyword": searched_keyword,
                "gldjc_url": None,
                "match_label": None,
                "gldjc_scope": scope_label,
            }, spec=spec, region=scope_label)
            results.append({
                **mat,
                "gldjc_price": None,
                "gldjc_source": "未查到",
                "gldjc_url": None,
                "gldjc_label": None,
            })

        # 每10条保存一次缓存（防中途中断丢数据）
        if net_requests % 10 == 0:
            save_cache(cache)

    # 最终保存缓存
    save_cache(cache)
    logger.info(f"广材网查价完成：{len(materials)}条材料，{net_requests}次网络请求，"
                f"{'被限制提前停止' if blocked else '正常完成'}")
    return results
