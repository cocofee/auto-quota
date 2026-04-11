"""
Unified file intake API.
"""

from __future__ import annotations

import asyncio
import uuid
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from pathlib import Path

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from loguru import logger

import config as quota_config
from app.auth.deps import get_current_user
from app.config import UPLOAD_DIR, UPLOAD_MAX_MB
from app.models.task import Task
from app.models.user import User
from app.services.match_service import save_upload_file
from app.tasks.match_task import execute_match
from app.schemas.file_intake import (
    FileClassifyRequest,
    FileClassifyResponse,
    FileIntakeResponse,
    FileManualReviewConfirmRequest,
    FileManualReviewConfirmResponse,
    FileParseRequest,
    FileParseResponse,
    FileRouteRequest,
    FileRouteResponse,
)
from app.text_utils import normalize_client_filename
from src.experience_db import ExperienceDB
from src.file_intake_db import FileIntakeDB
from src.price_reference_db import PriceReferenceDB

router = APIRouter()


_LEARNING_REGION_ROUTING = {
    "BJ": {
        "default": "北京市建设工程施工消耗量标准(2024)",
    },
    "FJ": {
        "建筑": "福建省房屋建筑与装饰工程预算定额(2017)",
        "装饰": "福建省房屋建筑与装饰工程预算定额(2017)",
        "市政": "福建省市政工程预算定额(2017)",
        "园林": "福建省园林绿化工程预算定额(2017)",
        "绿化": "福建省园林绿化工程预算定额(2017)",
        "安装": "福建省通用安装工程预算定额(2017)",
        "default": "福建省房屋建筑与装饰工程预算定额(2017)",
    },
    "ZJ": {
        "建筑": "浙江省房屋建筑与装饰工程预算定额(2018)",
        "装饰": "浙江省房屋建筑与装饰工程预算定额(2018)",
        "市政": "浙江省市政工程预算定额(2018)",
        "园林": "浙江省园林绿化及仿古建筑工程预算定额(2018)",
        "绿化": "浙江省园林绿化及仿古建筑工程预算定额(2018)",
        "安装": "浙江省通用安装工程预算定额(2018)",
        "default": "浙江省房屋建筑与装饰工程预算定额(2018)",
    },
    "JS": {
        "建筑": "江苏省建筑与装饰工程计价定额(2014)",
        "装饰": "江苏省建筑与装饰工程计价定额(2014)",
        "市政": "江苏省市政工程计价定额(2014)",
        "安装": "江苏省安装工程计价定额(2014)",
        "default": "江苏省建筑与装饰工程计价定额(2014)",
    },
}


@dataclass
class IngestResult:
    file_id: str | None
    ingest_intent: str
    evidence_level: str
    written_learning: int
    written_price_reference: int
    skipped: int
    warnings: list[str]
    errors: list[str]
    route_targets: list[str]

    def to_dict(self) -> dict:
        return asdict(self)


def _utc_from_ts(value) -> datetime:
    ts = float(value or 0)
    return datetime.fromtimestamp(ts, tz=timezone.utc)


def _to_response(record: dict) -> FileIntakeResponse:
    return FileIntakeResponse(
        file_id=record["file_id"],
        filename=record["filename"],
        status=record["status"],
        file_type=record.get("file_type") or "",
        source_hint=record.get("source_hint") or "",
        province=record.get("province") or "",
        project_name=record.get("project_name") or "",
        project_stage=record.get("project_stage") or "",
        current_stage=record.get("current_stage") or "",
        next_action=record.get("next_action") or "",
        receipt_summary=record.get("receipt_summary") or {},
        failure_type=record.get("failure_type") or "",
        failure_stage=record.get("failure_stage") or "",
        needs_manual_review=bool(record.get("needs_manual_review")),
        manual_review_reason=record.get("manual_review_reason") or "",
        classify_result=record.get("classify_result") or {},
        parse_summary=record.get("parse_summary") or {},
        route_result=record.get("route_result") or {},
        error_message=record.get("error_message") or "",
        created_at=_utc_from_ts(record.get("created_at")),
        updated_at=_utc_from_ts(record.get("updated_at")),
    )


def _load_headers_from_excel(path: Path) -> list[str]:
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    headers: list[str] = []
    try:
        for ws in wb.worksheets[:2]:
            for row in ws.iter_rows(min_row=1, max_row=8, values_only=True):
                for cell in row:
                    text = str(cell or "").strip()
                    if text and len(text) <= 30:
                        headers.append(text.replace("\n", "").replace(" ", ""))
    finally:
        wb.close()
    return headers


def _classify_record(record: dict) -> tuple[str, dict]:
    path = Path(record["stored_path"])
    ext = (record.get("file_ext") or path.suffix).lower()
    filename = str(record.get("filename") or path.name or "")
    signals: list[str] = []
    file_type = "other"
    confidence = 0.25

    if ext in {".xlsx", ".xlsm", ".xltx", ".xltm", ".xls"} and path.exists():
        headers = _load_headers_from_excel(path)
        quote_terms = {"品牌", "型号", "规格型号", "设备名称", "材料名称", "单价", "安装调测费", "安装费", "报价", "报价清单"}
        boq_terms = {
            "清单项目", "项目名称", "项目特征", "项目特征描述", "综合单价", "定额编号", "定额名称",
            "工程量清单", "分部分项", "计量单位", "合价", "项目编码", "招标工程量清单", "清单计价"
        }
        task_terms = {"项目编码", "项目名称", "项目特征描述", "计量单位", "工程量", "任务分配"}

        quote_hits = [term for term in quote_terms if any(term in header for header in headers) or term in filename]
        boq_hits = [term for term in boq_terms if any(term in header for header in headers) or term in filename]
        task_hits = [term for term in task_terms if any(term in header for header in headers) or term in filename]

        if len(boq_hits) >= 2:
            file_type = "priced_bill_file"
            confidence = min(0.62 + 0.06 * len(boq_hits), 0.97)
            signals = boq_hits
        elif len(quote_hits) >= 2:
            file_type = "historical_quote_file"
            confidence = min(0.6 + 0.06 * len(quote_hits), 0.95)
            signals = quote_hits
        elif len(task_hits) >= 2:
            file_type = "quota_task_file"
            confidence = min(0.55 + 0.08 * len(task_hits), 0.9)
            signals = task_hits
    elif ext == ".xml" and path.exists():
        content = path.read_text(encoding="utf-8", errors="ignore")[:6000]
        if any(token in content for token in ("分部分项综合单价分析表", "DEZM", "QDXM", "定额编号", "清单项目", "综合单价")):
            file_type = "priced_bill_file"
            confidence = 0.9
            signals = ["XML", "定额编号", "综合单价分析"]

    result = {
        "file_type": file_type,
        "confidence": round(confidence, 3),
        "signals": signals,
    }
    return file_type, result


def _parse_record(record: dict) -> dict:
    path = Path(record["stored_path"])
    file_type = record.get("file_type") or "other"
    summary = {"warnings": []}

    if not path.exists():
        raise FileNotFoundError(f"file not found: {path}")

    if path.suffix.lower() in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        import openpyxl

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            non_empty_rows = 0
            for ws in wb.worksheets[:3]:
                for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 1000), values_only=True):
                    if any(cell not in (None, "") for cell in row):
                        non_empty_rows += 1
        finally:
            wb.close()

        if file_type == "priced_bill_file":
            summary.update({"bill_items": max(non_empty_rows - 1, 0), "quote_items": 0})
        elif file_type == "historical_quote_file":
            summary.update({"quote_items": max(non_empty_rows - 1, 0), "bill_items": 0})
        elif file_type == "quota_task_file":
            summary.update({"task_rows": max(non_empty_rows - 1, 0)})
        else:
            summary.update({"rows": non_empty_rows})
    else:
        summary.update({"bytes": path.stat().st_size})

    return summary


def _default_route_targets(file_type: str) -> list[str]:
    if file_type == "quota_task_file":
        return ["task_pipeline"]
    if file_type == "historical_quote_file":
        return ["price_reference_quote"]
    if file_type == "priced_bill_file":
        return ["price_reference_boq", "learning_pipeline"]
    return ["manual_review"]


def _evidence_to_source(evidence_level: str, *, source_context: dict | None = None) -> str:
    source_context = source_context or {}
    mapped = {
        "user_corrected": "user_correction",
        "openclaw_approved": "openclaw_approved",
        "completed_project": "completed_project",
        "reviewed_import": "reviewed_import",
    }
    if evidence_level in mapped:
        return mapped[evidence_level]
    preferred = str(source_context.get("source") or "").strip()
    if preferred:
        return preferred
    return "project_import"


def _resolve_learning_province(raw_province: str | None, *, specialty: str = "") -> str:
    province = str(raw_province or "").strip()
    if not province:
        return ""
    try:
        return quota_config.resolve_province(province, interactive=False)
    except Exception:
        pass

    routing = _LEARNING_REGION_ROUTING.get(province.upper())
    if not routing:
        return province

    specialty_text = str(specialty or "").strip()
    for keyword, short_name in routing.items():
        if keyword == "default":
            continue
        if keyword in specialty_text:
            return short_name

    default_name = routing.get("default", "")
    if default_name:
        return default_name
    return province


def _normalize_learning_record(record: dict, *, source_context: dict | None = None) -> dict:
    source_context = source_context or {}
    quota_ids = record.get("quota_ids")
    if quota_ids is None:
        quota_code = record.get("quota_code")
        if isinstance(quota_code, list):
            quota_ids = quota_code
        elif quota_code:
            quota_ids = [quota_code]
        else:
            quota_ids = []
    quota_names = record.get("quota_names")
    if quota_names is None:
        quota_name = record.get("quota_name")
        if isinstance(quota_name, list):
            quota_names = quota_name
        elif quota_name:
            quota_names = [quota_name]
        else:
            quota_names = []
    materials = record.get("materials")
    if materials is None:
        materials = record.get("materials_json") or []
    bill_text = str(record.get("bill_text") or record.get("raw_name") or "").strip()
    bill_name = str(record.get("bill_name") or record.get("display_name") or record.get("raw_name") or "").strip()
    project_name = str(record.get("project_name") or source_context.get("project_name") or "").strip()
    province = str(record.get("province") or source_context.get("province") or "").strip()
    return {
        "bill_text": bill_text,
        "bill_name": bill_name or None,
        "bill_code": str(record.get("bill_code") or record.get("code") or "").strip() or None,
        "bill_unit": str(record.get("bill_unit") or record.get("unit") or "").strip() or None,
        "quota_ids": [str(item).strip() for item in (quota_ids or []) if str(item).strip()],
        "quota_names": [str(item).strip() for item in (quota_names or []) if str(item).strip()],
        "materials": materials if isinstance(materials, list) else [],
        "specialty": str(record.get("specialty") or source_context.get("specialty") or "").strip() or None,
        "project_name": project_name or None,
        "province": province or None,
        "confidence": record.get("confidence", source_context.get("confidence", 80)),
        "feature_text": str(record.get("feature_text") or "").strip() or None,
        "install_method": str(record.get("install_method") or "").strip() or None,
        "parse_status": str(record.get("parse_status") or source_context.get("parse_status") or "").strip(),
        "notes": str(record.get("notes") or "").strip() or None,
    }


def _normalize_price_record(record: dict, *, source_context: dict | None = None) -> dict:
    source_context = source_context or {}
    tags = record.get("tags") or []
    if not isinstance(tags, list):
        tags = [str(tags)]
    business_type = str(record.get("business_type") or source_context.get("business_type") or "").strip().lower()
    raw_name = str(
        record.get("raw_name")
        or record.get("item_name_raw")
        or record.get("boq_name_raw")
        or record.get("bill_text")
        or ""
    ).strip()
    normalized_name = str(
        record.get("normalized_name")
        or record.get("item_name_normalized")
        or record.get("boq_name_normalized")
        or raw_name
    ).strip()
    materials = record.get("materials_json")
    if materials is None:
        materials = record.get("materials") or []
    normalized = {
        "record_type": "boq" if (business_type in {"priced_bill_file", "composite", "mixed"} or "composite" in tags) else "quote",
        "raw_name": raw_name,
        "normalized_name": normalized_name,
        "specialty": str(record.get("specialty") or source_context.get("specialty") or "").strip(),
        "unit": str(record.get("unit") or "").strip(),
        "brand": str(record.get("brand") or "").strip(),
        "model": str(record.get("model") or "").strip(),
        "spec": str(record.get("spec") or record.get("规格") or "").strip(),
        "system_name": str(record.get("system_name") or "").strip(),
        "region": str(record.get("region") or record.get("province") or source_context.get("province") or "").strip(),
        "source_date": str(record.get("source_date") or "").strip(),
        "project_name": str(record.get("project_name") or source_context.get("project_name") or "").strip(),
        "project_stage": str(record.get("project_stage") or source_context.get("project_stage") or "").strip(),
        "project_id": str(record.get("project_id") or "").strip(),
        "remarks": str(record.get("remarks") or record.get("notes") or "").strip(),
        "materials_json": materials if isinstance(materials, list) else [],
        "materials_signature": str(record.get("materials_signature") or "").strip(),
        "source_type": str(record.get("source_type") or source_context.get("source") or "parsed_file").strip(),
        "source_record_id": record.get("source_record_id"),
        "tags": tags,
    }
    if normalized["record_type"] == "quote":
        normalized.update(
            {
                "item_name_raw": raw_name,
                "item_name_normalized": normalized_name,
                "unit_price": record.get("unit_price"),
                "install_price": record.get("install_price"),
                "combined_unit_price": record.get("combined_unit_price") or record.get("price_value"),
            }
        )
    else:
        normalized.update(
            {
                "boq_code": str(record.get("boq_code") or record.get("bill_code") or "").strip(),
                "boq_name_raw": raw_name,
                "boq_name_normalized": normalized_name,
                "feature_text": str(record.get("feature_text") or "").strip(),
                "work_content": str(record.get("work_content") or "").strip(),
                "item_features_structured": record.get("item_features_structured"),
                "quantity": record.get("quantity"),
                "composite_unit_price": record.get("composite_unit_price") or record.get("unit_price") or record.get("price_value"),
                "quota_code": str(record.get("quota_code") or "").strip(),
                "quota_name": str(record.get("quota_name") or "").strip(),
                "quota_group": str(record.get("quota_group") or "").strip(),
                "labor_cost": record.get("labor_cost"),
                "material_cost": record.get("material_cost"),
                "machine_cost": record.get("machine_cost"),
                "management_fee": record.get("management_fee"),
                "profit": record.get("profit"),
                "measure_fee": record.get("measure_fee"),
                "other_fee": record.get("other_fee"),
                "tax": record.get("tax"),
                "currency": str(record.get("currency") or "CNY").strip(),
                "source_sheet": str(record.get("source_sheet") or "").strip(),
                "source_row_no": record.get("source_row_no"),
                "bill_text": str(record.get("bill_text") or raw_name).strip(),
                "subsystem_name": str(record.get("subsystem_name") or "").strip(),
            }
        )
    return normalized


def ingest(
    *,
    file_id: str | None = None,
    records: list[dict] | None = None,
    ingest_intent: str,
    evidence_level: str,
    business_type: str = "unknown",
    actor: str = "",
    source_context: dict | None = None,
    batch_mode: bool = False,
) -> IngestResult:
    source_context = source_context or {}
    warnings: list[str] = []
    errors: list[str] = []
    route_targets: list[str] = []
    written_learning = 0
    written_price_reference = 0
    skipped = 0

    file_record = None
    if file_id:
        file_record = FileIntakeDB().get_file(file_id)
        if not file_record:
            raise ValueError(f"file not found: {file_id}")

    if file_record:
        merged_context = {
            **file_record.get("source_context", {}),
            **source_context,
            "project_name": source_context.get("project_name") or file_record.get("project_name") or "",
            "province": source_context.get("province") or file_record.get("province") or "",
            "parse_status": source_context.get("parse_status") or "parsed",
        }
    else:
        merged_context = dict(source_context)

    if ingest_intent in {"learning", "dual_use"}:
        route_targets.append("learning_pipeline")
        if not records:
            warnings.append("learning ingest received no structured records")
        merged_learning_province = _resolve_learning_province(
            merged_context.get("province"),
            specialty=str(merged_context.get("specialty") or ""),
        )
        experience_db = ExperienceDB(province=merged_learning_province or None)
        source = _evidence_to_source(evidence_level, source_context=merged_context)
        normalized_learning_records: list[dict] = []
        for raw_record in records or []:
            normalized = _normalize_learning_record(raw_record, source_context=merged_context)
            if not normalized["bill_text"] or not normalized["quota_ids"]:
                skipped += 1
                warnings.append("learning record skipped: missing bill_text or quota_ids")
                continue
            learning_province = _resolve_learning_province(
                normalized["province"],
                specialty=normalized["specialty"] or str(merged_context.get("specialty") or ""),
            )
            normalized_learning_records.append(
                {
                    "bill_text": normalized["bill_text"],
                    "quota_ids": normalized["quota_ids"],
                    "quota_names": normalized["quota_names"],
                    "materials": normalized["materials"],
                    "bill_name": normalized["bill_name"],
                    "bill_code": normalized["bill_code"],
                    "bill_unit": normalized["bill_unit"],
                    "source": source,
                    "confidence": normalized["confidence"],
                    "province": learning_province or normalized["province"],
                    "project_name": normalized["project_name"],
                    "notes": normalized["notes"],
                    "specialty": normalized["specialty"],
                    "feature_text": normalized["feature_text"],
                    "install_method": normalized["install_method"],
                    "parse_status": normalized["parse_status"],
                }
            )
        if batch_mode and normalized_learning_records:
            bulk_result = experience_db.bulk_add_experiences(
                normalized_learning_records,
                skip_vector=True,
                skip_fts=True,
            )
            written_learning += int(bulk_result.get("written") or 0)
            skipped += int(bulk_result.get("rejected") or 0)
            if bulk_result.get("rejected"):
                warnings.append(f"learning batch rejected: {bulk_result['rejected']}")
        else:
            for normalized in normalized_learning_records:
                record_id = experience_db.add_experience(
                    bill_text=normalized["bill_text"],
                    quota_ids=normalized["quota_ids"],
                    quota_names=normalized["quota_names"],
                    materials=normalized["materials"],
                    bill_name=normalized["bill_name"],
                    bill_code=normalized["bill_code"],
                    bill_unit=normalized["bill_unit"],
                    source=normalized["source"],
                    confidence=normalized["confidence"],
                    province=normalized["province"],
                    project_name=normalized["project_name"],
                    notes=normalized["notes"],
                    specialty=normalized["specialty"],
                    feature_text=normalized["feature_text"],
                    install_method=normalized["install_method"],
                    parse_status=normalized["parse_status"],
                    skip_vector=True,
                    skip_fts=False,
                )
                if record_id > 0:
                    written_learning += 1
                else:
                    skipped += 1
                    warnings.append(f"learning record rejected: {normalized['bill_text'][:80]}")
        if batch_mode and written_learning:
            warnings.append("learning batch_mode enabled: per-record FTS/vector updates skipped; rebuild indexes after batch")

    if ingest_intent in {"price_reference", "dual_use"}:
        route_targets.append("price_reference")
        price_db = PriceReferenceDB()
        quote_items: list[dict] = []
        boq_items: list[dict] = []
        for raw_record in records or []:
            normalized = _normalize_price_record(raw_record, source_context=merged_context)
            if not normalized["raw_name"]:
                skipped += 1
                warnings.append("price reference record skipped: missing raw_name")
                continue
            if normalized["record_type"] == "quote":
                if all(normalized.get(field) in (None, "") for field in ("unit_price", "install_price", "combined_unit_price")):
                    skipped += 1
                    warnings.append(f"quote price record skipped: missing price fields for {normalized['raw_name'][:80]}")
                    continue
                quote_items.append(normalized)
            else:
                if normalized.get("composite_unit_price") in (None, ""):
                    skipped += 1
                    warnings.append(f"boq price record skipped: missing composite_unit_price for {normalized['raw_name'][:80]}")
                    continue
                boq_items.append(normalized)

        def _create_price_document(target_type: str) -> int:
            suffix = "quote" if target_type == "historical_quote_file" else "boq"
            base_file_id = file_record["file_id"] if file_record else f"ingest::{suffix}::{uuid.uuid4().hex[:16]}"
            if file_record:
                doc_file_id = f"{base_file_id}::{suffix}" if business_type == "mixed" else base_file_id
            else:
                doc_file_id = base_file_id
            return price_db.create_document_from_file(
                file_id=doc_file_id,
                document_type=target_type,
                project_name=merged_context.get("project_name") or (file_record.get("project_name") if file_record else "") or "",
                project_stage=merged_context.get("project_stage") or (file_record.get("project_stage") if file_record else "") or "",
                specialty=merged_context.get("specialty") or "",
                region=merged_context.get("province") or (file_record.get("province") if file_record else "") or "",
                source_file_name=(file_record.get("filename") if file_record else "") or "record_ingest.json",
                source_file_path=(file_record.get("stored_path") if file_record else "") or "",
                status="parsed",
                parse_status=merged_context.get("parse_status") or "parsed",
            )

        if quote_items:
            quote_document_id = _create_price_document("historical_quote_file")
            written_price_reference += price_db.replace_quote_items(quote_document_id, quote_items)
        if boq_items:
            boq_document_id = _create_price_document("priced_bill_file")
            written_price_reference += price_db.replace_boq_items(boq_document_id, boq_items)
        if not records:
            warnings.append("price_reference ingest received no structured records")

    if ingest_intent == "task_match":
        route_targets.append("task_pipeline")

    result = IngestResult(
        file_id=file_id,
        ingest_intent=ingest_intent,
        evidence_level=evidence_level,
        written_learning=written_learning,
        written_price_reference=written_price_reference,
        skipped=skipped,
        warnings=warnings,
        errors=errors,
        route_targets=route_targets,
    )

    if file_record:
        FileIntakeDB().update_route(file_id, route_result={"ingest_result": result.to_dict()})

    return result


async def _save_upload(
    *,
    file: UploadFile,
    province: str = "",
    project_name: str = "",
    project_stage: str = "",
    source_hint: str = "",
    actor: User,
) -> FileIntakeResponse:
    filename = normalize_client_filename(file.filename, "upload.bin")
    if not filename or filename in {".", ".."}:
        raise HTTPException(status_code=400, detail="文件名非法")

    max_size = UPLOAD_MAX_MB * 1024 * 1024
    db = FileIntakeDB()
    temp_id = f"fi_{uuid.uuid4().hex[:16]}"
    upload_dir = UPLOAD_DIR / "file_intake" / temp_id
    upload_dir.mkdir(parents=True, exist_ok=True)
    save_path = upload_dir / filename

    size = 0
    try:
        with open(save_path, "wb") as fh:
            while chunk := await file.read(8192):
                size += len(chunk)
                if size > max_size:
                    raise HTTPException(status_code=400, detail=f"文件过大（超过{UPLOAD_MAX_MB}MB）")
                fh.write(chunk)
    except Exception:
        save_path.unlink(missing_ok=True)
        raise

    record = db.create_file(
        filename=filename,
        stored_path=str(save_path),
        mime_type=file.content_type or "",
        file_ext=save_path.suffix.lower(),
        file_size=size,
        source_hint=source_hint,
        province=province,
        project_name=project_name,
        project_stage=project_stage,
        created_by=getattr(actor, "email", "") or getattr(actor, "nickname", "") or str(actor.id),
        actor=getattr(actor, "email", "") or getattr(actor, "nickname", "") or str(actor.id),
    )
    return _to_response(record)


async def _classify_file(file_id: str, req: FileClassifyRequest) -> FileClassifyResponse:
    db = FileIntakeDB()
    record = db.get_file(file_id)
    if not record:
        raise HTTPException(status_code=404, detail="file not found")
    if record.get("status") == "classified" and not req.force and record.get("classify_result"):
        result = record["classify_result"]
        return FileClassifyResponse(
            file_id=file_id,
            status=record["status"],
            file_type=result.get("file_type", record.get("file_type") or ""),
            confidence=float(result.get("confidence") or 0.0),
            signals=result.get("signals") or [],
        )

    try:
        file_type, result = await asyncio.to_thread(_classify_record, record)
        confidence = float(result.get("confidence") or 0.0)
        if file_type == "other" or confidence < 0.55:
            db.update_failure(
                file_id,
                error_message=f"classification confidence too low: {confidence}",
                failure_type="manual_review",
                failure_stage="classify-file",
                needs_manual_review=True,
                manual_review_reason="low_confidence_classification",
            )
            updated = db.get_file(file_id) or record
            return FileClassifyResponse(
                file_id=file_id,
                status=updated.get("status") or "waiting_human",
                file_type=file_type,
                confidence=confidence,
                signals=result.get("signals") or [],
            )

        db.update_classify(file_id, file_type=file_type, classify_result=result)
    except Exception as e:
        logger.error(f"file classify failed: {e}")
        db.update_failure(
            file_id,
            error_message=str(e),
            failure_type="hard_fail",
            failure_stage="classify-file",
        )
        raise HTTPException(status_code=500, detail=f"classify failed: {e}")

    return FileClassifyResponse(
        file_id=file_id,
        status="classified",
        file_type=file_type,
        confidence=float(result.get("confidence") or 0.0),
        signals=result.get("signals") or [],
    )


async def _parse_file(file_id: str, req: FileParseRequest) -> FileParseResponse:
    db = FileIntakeDB()
    record = db.get_file(file_id)
    if not record:
        raise HTTPException(status_code=404, detail="file not found")
    if not record.get("file_type"):
        await _classify_file(file_id, FileClassifyRequest(force=False))
        record = db.get_file(file_id) or record

    if record.get("status") == "parsed" and not req.force and record.get("parse_summary"):
        return FileParseResponse(
            file_id=file_id,
            status=record["status"],
            file_type=record.get("file_type") or "",
            parse_summary=record.get("parse_summary") or {},
        )

    try:
        summary = await asyncio.to_thread(_parse_record, record)
        db.update_parse(file_id, status="parsed", parse_summary=summary)
    except Exception as e:
        logger.error(f"file parse failed: {e}")
        db.update_failure(
            file_id,
            error_message=str(e),
            failure_type="hard_fail",
            failure_stage="parse-file",
        )
        raise HTTPException(status_code=500, detail=f"parse failed: {e}")

    return FileParseResponse(
        file_id=file_id,
        status="parsed",
        file_type=record.get("file_type") or "",
        parse_summary=summary,
    )


async def _route_file(file_id: str, req: FileRouteRequest, *, user: User) -> FileRouteResponse:
    intake_db = FileIntakeDB()
    record = intake_db.get_file(file_id)
    if not record:
        raise HTTPException(status_code=404, detail="file not found")
    if not record.get("file_type"):
        await _classify_file(file_id, FileClassifyRequest(force=False))
        record = intake_db.get_file(file_id) or record
    if not record.get("parse_summary"):
        await _parse_file(file_id, FileParseRequest(force=False))
        record = intake_db.get_file(file_id) or record

    targets = req.route_targets or _default_route_targets(record.get("file_type") or "other")
    target_results: list[dict] = []
    price_db = PriceReferenceDB()

    for target in targets:
        if target in {"price_reference_quote", "price_reference_boq"}:
            document_type = (
                "historical_quote_file" if target == "price_reference_quote" else "priced_bill_file"
            )
            document_id = await asyncio.to_thread(
                price_db.create_document_from_file,
                file_id=file_id,
                document_type=document_type,
                project_name=record.get("project_name") or "",
                project_stage=record.get("project_stage") or "",
                specialty=(record.get("classify_result") or {}).get("specialty", ""),
                region=record.get("province") or "",
                source_file_name=record.get("filename") or "",
                status="parsed",
            )
            target_results.append(
                {"target": target, "status": "ok", "document_id": document_id}
            )
        elif target == "learning_pipeline":
            ingest_result = ingest(
                file_id=file_id,
                records=[],
                ingest_intent="learning",
                evidence_level="raw_import",
                business_type=record.get("file_type") or "unknown",
                actor=record.get("actor") or record.get("created_by") or "",
                source_context={
                    "project_name": record.get("project_name") or "",
                    "province": record.get("province") or "",
                    "project_stage": record.get("project_stage") or "",
                    "parse_status": "parsed",
                },
            )
            target_results.append(
                {
                    "target": target,
                    "status": "ok",
                    "written_learning": ingest_result.written_learning,
                    "warnings": ingest_result.warnings,
                }
            )
        elif target == "task_pipeline":
            if not req.auto_create_task:
                target_results.append(
                    {"target": target, "status": "pending", "message": "set auto_create_task=true to create a task"}
                )
                continue

            province = str(record.get("province") or "").strip()
            if not province:
                target_results.append(
                    {"target": target, "status": "failed", "message": "province is required for task creation"}
                )
                continue

            task_id = uuid.uuid4()
            original_name = str(record.get("filename") or "input.xlsx").strip() or "input.xlsx"
            suffix = Path(original_name).suffix or Path(str(record.get("stored_path") or "")).suffix or ".xlsx"
            temp_upload_name = f"input{suffix if suffix.lower() in {'.xlsx', '.xls'} else '.xlsx'}"
            temp_source = Path(record.get("stored_path") or "")

            if not temp_source.exists():
                target_results.append(
                    {"target": target, "status": "failed", "message": "stored file not found"}
                )
                continue

            class _UploadShim:
                def __init__(self, path: Path, filename: str):
                    self.filename = filename
                    self.content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    self.file = open(path, "rb")

            upload = _UploadShim(temp_source, temp_upload_name)
            try:
                saved_path = await asyncio.to_thread(save_upload_file, upload, task_id)
            finally:
                upload.file.close()

            task = Task(
                id=task_id,
                user_id=user.id,
                name=Path(original_name).stem,
                file_path=str(saved_path),
                original_filename=original_name,
                mode="search",
                province=province,
                sheet=None,
                limit_count=None,
                use_experience=True,
                agent_llm=None,
                status="pending",
                progress=0,
                progress_message="等待执行...",
            )

            from app.database import async_session
            async with async_session() as db:
                db.add(task)
                await db.commit()
                try:
                    celery_result = execute_match.delay(
                        task_id=str(task_id),
                        file_path=str(saved_path),
                        params={
                            "mode": "search",
                            "province": province,
                            "sheet": None,
                            "limit": None,
                            "agent_llm": None,
                            "no_experience": False,
                        },
                    )
                    task.celery_task_id = celery_result.id
                    await db.commit()
                    target_results.append(
                        {
                            "target": target,
                            "status": "ok",
                            "task_id": str(task_id),
                            "celery_task_id": celery_result.id,
                            "message": "task created",
                        }
                    )
                except Exception as e:
                    task.status = "failed"
                    task.error_message = f"任务入队失败: {e}"
                    await db.commit()
                    target_results.append(
                        {"target": target, "status": "failed", "task_id": str(task_id), "message": str(e)}
                    )
        else:
            target_results.append({"target": target, "status": "pending"})

    route_result = {"targets": target_results}
    intake_db.update_route(file_id, route_result=route_result)
    return FileRouteResponse(file_id=file_id, status="routed", targets=target_results)


@router.post("/upload", response_model=FileIntakeResponse, status_code=201)
async def upload_file(
    file: UploadFile = File(...),
    province: str = Form(default=""),
    project_name: str = Form(default=""),
    project_stage: str = Form(default=""),
    source_hint: str = Form(default=""),
    user: User = Depends(get_current_user),
):
    return await _save_upload(
        file=file,
        province=province,
        project_name=project_name,
        project_stage=project_stage,
        source_hint=source_hint,
        actor=user,
    )


@router.get("/{file_id}", response_model=FileIntakeResponse)
async def get_file(file_id: str, user: User = Depends(get_current_user)):
    _ = user
    record = FileIntakeDB().get_file(file_id)
    if not record:
        raise HTTPException(status_code=404, detail="file not found")
    return _to_response(record)


@router.post("/{file_id}/classify", response_model=FileClassifyResponse)
async def classify_file(
    file_id: str,
    req: FileClassifyRequest,
    user: User = Depends(get_current_user),
):
    _ = user
    return await _classify_file(file_id, req)


@router.post("/{file_id}/parse", response_model=FileParseResponse)
async def parse_file(
    file_id: str,
    req: FileParseRequest,
    user: User = Depends(get_current_user),
):
    _ = user
    return await _parse_file(file_id, req)


@router.post("/{file_id}/route", response_model=FileRouteResponse)
async def route_file(
    file_id: str,
    req: FileRouteRequest,
    user: User = Depends(get_current_user),
):
    return await _route_file(file_id, req, user=user)


@router.post("/{file_id}/manual-review/confirm", response_model=FileManualReviewConfirmResponse)
async def confirm_manual_review(
    file_id: str,
    req: FileManualReviewConfirmRequest,
    user: User = Depends(get_current_user),
):
    _ = user
    db = FileIntakeDB()
    record = db.get_file(file_id)
    if not record:
        raise HTTPException(status_code=404, detail="file not found")
    if record.get("status") != "waiting_human":
        raise HTTPException(status_code=400, detail="file is not waiting human review")

    file_type = (req.file_type or record.get("file_type") or "").strip()
    if not file_type:
        raise HTTPException(status_code=400, detail="file_type is required for manual confirm")

    continue_from = (req.continue_from or "parse").strip().lower()
    if continue_from not in {"parse", "route"}:
        raise HTTPException(status_code=400, detail="continue_from must be parse or route")

    updated = db.confirm_manual_review(file_id, file_type=file_type, continue_from=continue_from)
    if continue_from == "parse":
        await _parse_file(file_id, FileParseRequest(force=True))
        updated = db.get_file(file_id) or updated
        message = "manual review confirmed, parse resumed"
    else:
        route_req = FileRouteRequest(
            route_targets=req.route_targets,
            auto_create_task=req.auto_create_task,
        )
        await _route_file(file_id, route_req, user=user)
        updated = db.get_file(file_id) or updated
        message = "manual review confirmed, route resumed"

    return FileManualReviewConfirmResponse(
        file_id=file_id,
        status=updated.get("status") or "",
        current_stage=updated.get("current_stage") or "",
        next_action=updated.get("next_action") or "",
        file_type=updated.get("file_type") or "",
        message=message,
    )
