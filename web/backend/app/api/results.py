"""
匹配结果 API

提供匹配结果的查看、纠正、批量确认和Excel导出。

路由挂载在 /api 前缀下:
    GET    /api/tasks/{id}/results              — 结果列表
    GET    /api/tasks/{id}/results/{result_id}  — 单条结果详情
    PUT    /api/tasks/{id}/results/{result_id}  — 纠正结果
    POST   /api/tasks/{id}/results/confirm      — 批量确认
    GET    /api/tasks/{id}/export               — 导出Excel（原始匹配结果）
    GET    /api/tasks/{id}/export-final         — 导出Excel（含纠正，实时生成）
"""

import asyncio
import json
import re
import uuid
from collections import defaultdict
from pathlib import Path

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import FileResponse
from loguru import logger
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db
from app.models.result import MatchResult
from app.models.user import User
from app.auth.deps import get_current_user
from app.schemas.result import (
    MatchResultResponse, ResultListResponse,
    CorrectResultRequest, ConfirmResultsRequest,
)
from app.api.shared import get_user_task, store_experience, store_experience_batch, flag_disputed_experience
from app.services.match_service import get_task_output_dir
from app.text_utils import normalize_client_filename, repair_mojibake_data
from src.ranking_feedback_db import (
    RankingFeedbackDB,
    infer_misrank_primary_factor,
    normalize_dimension_scores,
)

router = APIRouter()

# 置信度分档阈值（必须与 config.py CONFIDENCE_GREEN/YELLOW 和前端 experience.ts 保持一致）
# 修改时三处同步：config.py:585-586 / experience.ts:12-13 / 此处
_GREEN_THRESHOLD = 90
_YELLOW_THRESHOLD = 75


def _read_result_value(result, key: str, default=None):
    if isinstance(result, dict):
        return result.get(key, default)
    return getattr(result, key, default)


def _effective_confidence(result) -> int:
    if isinstance(result, (int, float)):
        try:
            return max(0, min(100, int(result)))
        except (TypeError, ValueError):
            return 0
    value = _read_result_value(result, "confidence_score", None)
    if value is None:
        value = _read_result_value(result, "confidence", 0)
    try:
        return max(0, min(100, int(value or 0)))
    except (TypeError, ValueError):
        return 0


def _resolve_light_status(result) -> str:
    light_status = str(_read_result_value(result, "light_status", "") or "").strip().lower()
    if light_status in {"green", "yellow", "red"}:
        return light_status

    confidence = _effective_confidence(result)
    if confidence >= _GREEN_THRESHOLD:
        return "green"
    if confidence >= _YELLOW_THRESHOLD:
        return "yellow"
    return "red"


def _is_confirmable_result(result) -> bool:
    return _resolve_light_status(result) != "red"


def _task_download_stem(original_filename: str | None) -> str:
    return Path(normalize_client_filename(original_filename, "result.xlsx")).stem


def _build_rebuilt_results_from_db(items: list[MatchResult]) -> list[dict]:
    rebuilt_results = []
    for item in items:
        quotas = item.corrected_quotas or item.quotas or []
        rebuilt_results.append({
            "bill_item": {
                "code": item.bill_code or "",
                "name": item.bill_name or "",
                "description": item.bill_description or "",
                "unit": item.bill_unit or "",
                "quantity": item.bill_quantity,
                "sheet_name": item.sheet_name or "",
                "section": item.section or "",
                "specialty": item.specialty or "",
            },
            "quotas": quotas,
            "confidence": 95 if item.corrected_quotas else item.confidence,
            "explanation": item.explanation or "",
            "match_source": "corrected" if item.corrected_quotas else (item.match_source or ""),
        })
    return rebuilt_results


def _normalize_locator_text(value) -> str:
    return str(value or "").strip()


def _normalize_locator_quantity(value) -> str:
    if value in (None, ""):
        return ""
    try:
        number = float(value)
    except (TypeError, ValueError):
        return _normalize_locator_text(value)
    if number.is_integer():
        return str(int(number))
    return f"{number:.6f}".rstrip("0").rstrip(".")


def _build_locator_key(*, sheet_name, code, name, unit, quantity) -> tuple[str, str, str, str, str]:
    return (
        _normalize_locator_text(sheet_name),
        _normalize_locator_text(code),
        _normalize_locator_text(name),
        _normalize_locator_text(unit),
        _normalize_locator_quantity(quantity),
    )


def _build_db_locator_key(item: MatchResult, *, with_code: bool) -> tuple[str, str, str, str, str]:
    return _build_locator_key(
        sheet_name=item.sheet_name,
        code=item.bill_code if with_code else "",
        name=item.bill_name,
        unit=item.bill_unit,
        quantity=item.bill_quantity,
    )


def _build_source_locator_key(item: dict, *, with_code: bool) -> tuple[str, str, str, str, str]:
    return _build_locator_key(
        sheet_name=item.get("sheet_name"),
        code=item.get("code") if with_code else "",
        name=item.get("name"),
        unit=item.get("unit"),
        quantity=item.get("quantity"),
    )


def _pick_unused_source_item(candidates: list[dict], used_rows: set[int]) -> dict | None:
    for source_item in candidates:
        source_row = source_item.get("source_row")
        if isinstance(source_row, int) and source_row not in used_rows:
            return source_item
    return None


def _apply_source_locators(rebuilt_results: list[dict], items: list[MatchResult], source_items: list[dict]) -> int:
    source_by_primary: dict[tuple[str, str, str, str, str], list[dict]] = defaultdict(list)
    source_by_fallback: dict[tuple[str, str, str, str, str], list[dict]] = defaultdict(list)
    for source_item in source_items:
        if not isinstance(source_item, dict):
            continue
        source_by_primary[_build_source_locator_key(source_item, with_code=True)].append(source_item)
        source_by_fallback[_build_source_locator_key(source_item, with_code=False)].append(source_item)

    matched_count = 0
    used_rows: set[int] = set()
    for db_item, rebuilt in zip(items, rebuilt_results):
        source_item = _pick_unused_source_item(
            source_by_primary[_build_db_locator_key(db_item, with_code=True)],
            used_rows,
        )
        if source_item is None:
            source_item = _pick_unused_source_item(
                source_by_fallback[_build_db_locator_key(db_item, with_code=False)],
                used_rows,
            )
        if source_item is None:
            continue

        bill_item = rebuilt.setdefault("bill_item", {})
        for field in ("source_row", "sheet_bill_seq", "sheet_name", "section", "specialty"):
            value = source_item.get(field)
            if value not in (None, "") and bill_item.get(field) in (None, ""):
                bill_item[field] = value

        source_row = source_item.get("source_row")
        if isinstance(source_row, int):
            used_rows.add(source_row)
        matched_count += 1

    return matched_count


def _build_rebuilt_results_from_source_file(task, items: list[MatchResult]) -> list[dict] | None:
    candidate_paths: list[Path] = []
    for raw_path in (getattr(task, "file_path", None), getattr(task, "output_path", None)):
        text = str(raw_path or "").strip()
        if not text:
            continue
        path = Path(text)
        if path.exists() and path not in candidate_paths:
            candidate_paths.append(path)

    if not candidate_paths:
        return None

    last_error = None
    for source_path in candidate_paths:
        try:
            from src.bill_reader import BillReader

            source_items = BillReader().read_file(str(source_path), sheet_name=getattr(task, "sheet", None))
        except Exception as exc:
            last_error = exc
            logger.warning(f"reload source bill items failed: {source_path} -> {exc}")
            continue

        if not source_items:
            continue

        rebuilt_results = _build_rebuilt_results_from_db(items)
        matched_count = _apply_source_locators(rebuilt_results, items, source_items)
        if matched_count > 0:
            if matched_count != len(rebuilt_results):
                logger.warning(
                    "source file locator recovery partially matched: "
                    f"matched={matched_count} db={len(rebuilt_results)} task={task.id}"
                )
            return rebuilt_results

    if last_error is not None:
        logger.warning(f"source file locator recovery unavailable for task={task.id}: {last_error}")
    return None


def _build_rebuilt_results_from_json(task, items: list[MatchResult]) -> list[dict] | None:
    json_path = Path(str(getattr(task, "json_output_path", "") or "").strip())
    if not json_path.exists():
        return None

    try:
        payload = json.loads(json_path.read_text(encoding="utf-8"))
    except Exception as exc:
        logger.warning(f"load json_output_path failed: {json_path} -> {exc}")
        return None

    source_results = payload.get("results")
    if not isinstance(source_results, list) or not source_results:
        logger.warning(f"json_output_path has no results list: {json_path}")
        return None

    items_by_index = {int(item.index): item for item in items}
    rebuilt_results: list[dict] = []

    for idx, original in enumerate(source_results):
        if not isinstance(original, dict):
            continue
        result = dict(original)
        db_item = items_by_index.get(idx)
        if db_item is not None:
            result["quotas"] = db_item.corrected_quotas or db_item.quotas or []
            result["confidence"] = 95 if db_item.corrected_quotas else db_item.confidence
            if db_item.explanation:
                result["explanation"] = db_item.explanation
            result["match_source"] = (
                "corrected" if db_item.corrected_quotas else (db_item.match_source or result.get("match_source", ""))
            )
        rebuilt_results.append(result)

    if not rebuilt_results:
        return None

    if len(rebuilt_results) != len(items):
        logger.warning(
            "json_output_path result count mismatch: "
            f"json={len(rebuilt_results)} db={len(items)} task={task.id}"
        )
    return rebuilt_results


def _to_result_response(match_result: MatchResult) -> MatchResultResponse:
    payload = {
        field_name: _read_result_value(match_result, field_name, None)
        for field_name in MatchResultResponse.model_fields
        if field_name not in {"knowledge_evidence", "knowledge_basis", "knowledge_summary", "trace"}
    }
    for field_name in (
        "bill_code",
        "bill_name",
        "bill_description",
        "bill_unit",
        "specialty",
        "sheet_name",
        "section",
        "review_risk",
        "light_status",
        "match_source",
        "explanation",
        "review_status",
        "review_note",
        "openclaw_review_note",
        "openclaw_review_actor",
        "openclaw_retry_query",
        "openclaw_review_confirmed_by",
    ):
        payload[field_name] = str(payload.get(field_name) or "")
    for field_name in ("index", "confidence", "confidence_score", "candidates_count"):
        value = payload.get(field_name)
        try:
            payload[field_name] = int(value or 0)
        except (TypeError, ValueError):
            payload[field_name] = 0
    payload["is_measure_item"] = bool(payload.get("is_measure_item", False))
    payload["openclaw_review_status"] = (
        str(payload.get("openclaw_review_status") or "").strip().lower() or "pending"
    )
    if payload["openclaw_review_status"] not in {"pending", "reviewed", "applied", "rejected"}:
        payload["openclaw_review_status"] = "pending"
    payload["openclaw_review_confirm_status"] = (
        str(payload.get("openclaw_review_confirm_status") or "").strip().lower() or "pending"
    )
    if payload["openclaw_review_confirm_status"] not in {"pending", "approved", "rejected"}:
        payload["openclaw_review_confirm_status"] = "pending"
    for field_name, allowed_values in {
        "openclaw_decision_type": {
            "agree",
            "override_within_candidates",
            "retry_search_then_select",
            "candidate_pool_insufficient",
            "abstain",
        },
        "openclaw_error_stage": {"retriever", "ranker", "arbiter", "final_validator", "unknown"},
        "openclaw_error_type": {
            "wrong_family",
            "wrong_param",
            "wrong_book",
            "synonym_gap",
            "low_confidence_override",
            "missing_candidate",
            "unknown",
        },
    }.items():
        value = str(payload.get(field_name) or "").strip().lower()
        payload[field_name] = value if value in allowed_values else None
    for field_name in (
        "quotas",
        "corrected_quotas",
        "openclaw_suggested_quotas",
        "alternatives",
    ):
        value = payload.get(field_name)
        if value is not None and not isinstance(value, list):
            payload[field_name] = None
    reason_codes = payload.get("openclaw_reason_codes")
    if isinstance(reason_codes, list):
        payload["openclaw_reason_codes"] = [
            str(item).strip()
            for item in reason_codes
            if str(item or "").strip()
        ] or None
    else:
        payload["openclaw_reason_codes"] = None
    for field_name in ("openclaw_review_payload", "human_feedback_payload"):
        value = payload.get(field_name)
        if value is not None and not isinstance(value, dict):
            payload[field_name] = None
    payload["knowledge_evidence"] = _extract_knowledge_evidence(match_result)
    payload["knowledge_basis"] = _extract_knowledge_basis(match_result)
    payload["knowledge_summary"] = _extract_knowledge_summary(match_result)
    trace = _read_result_value(match_result, "trace", None)
    payload["trace"] = trace if isinstance(trace, dict) else None
    repaired = repair_mojibake_data(payload, preserve_newlines=True)
    return MatchResultResponse.model_validate(repaired)


def _actor_identity(user: User | None) -> str:
    if user is None:
        return ""
    return (
        str(getattr(user, "email", "") or "").strip()
        or str(getattr(user, "nickname", "") or "").strip()
        or str(getattr(user, "id", "") or "").strip()
    )


def _compose_query_text(match_result: MatchResult) -> str:
    parts = [
        str(match_result.bill_name or "").strip(),
        str(match_result.bill_description or "").strip(),
    ]
    return " | ".join(part for part in parts if part)


def _quota_bundle_key(quotas: list[dict] | None) -> str:
    quota_ids = [str((item or {}).get("quota_id", "") or "").strip() for item in (quotas or [])]
    quota_ids = [item for item in quota_ids if item]
    if not quota_ids:
        return ""
    return f"quota_bundle:{'|'.join(quota_ids)}"


def _quota_db_id(quotas: list[dict] | None) -> int | None:
    for item in quotas or []:
        value = _read_result_value(item, "db_id", None)
        try:
            if value not in (None, ""):
                return int(value)
        except (TypeError, ValueError):
            continue
    return None


def _first_quota_id(quotas: list[dict] | None) -> str:
    for item in quotas or []:
        quota_id = str(_read_result_value(item, "quota_id", "") or "").strip()
        if quota_id:
            return quota_id
    return ""


def _trace_steps(match_result: MatchResult) -> list[dict]:
    trace = _read_result_value(match_result, "trace", None)
    if not isinstance(trace, dict):
        return []
    steps = trace.get("steps")
    if not isinstance(steps, list):
        return []
    return [step for step in steps if isinstance(step, dict)]


def _extract_knowledge_evidence(match_result: MatchResult) -> dict | None:
    for step in reversed(_trace_steps(match_result)):
        value = step.get("knowledge_evidence")
        if isinstance(value, dict):
            has_payload = any(value.get(key) for key in ("reference_cases", "quota_rules", "quota_explanations", "method_cards"))
            if has_payload:
                return value

    reference_case_ids: list[str] = []
    rule_ids: list[str] = []
    method_card_ids: list[str] = []
    quota_rule_ids: list[str] = []
    quota_explanation_ids: list[str] = []
    for step in _trace_steps(match_result):
        for item in step.get("reference_case_ids", []) or []:
            text = str(item or "").strip()
            if text and text not in reference_case_ids:
                reference_case_ids.append(text)
        for item in step.get("rule_context_ids", []) or []:
            text = str(item or "").strip()
            if text and text not in rule_ids:
                rule_ids.append(text)
        for item in step.get("quota_rule_ids", []) or []:
            text = str(item or "").strip()
            if text and text not in quota_rule_ids:
                quota_rule_ids.append(text)
        for item in step.get("quota_explanation_ids", []) or []:
            text = str(item or "").strip()
            if text and text not in quota_explanation_ids:
                quota_explanation_ids.append(text)
        for item in step.get("method_card_ids", []) or []:
            text = str(item or "").strip()
            if text and text not in method_card_ids:
                method_card_ids.append(text)

    if not any([reference_case_ids, rule_ids, method_card_ids, quota_rule_ids, quota_explanation_ids]):
        return None

    return {
        "reference_cases": [{"record_id": item} for item in reference_case_ids],
        "quota_rules": [{"id": item} for item in quota_rule_ids or rule_ids],
        "quota_explanations": [{"id": item} for item in quota_explanation_ids],
        "method_cards": [{"id": item} for item in method_card_ids],
    }


def _extract_latest_trace_dict(match_result: MatchResult, key: str) -> dict | None:
    for step in reversed(_trace_steps(match_result)):
        value = step.get(key)
        if isinstance(value, dict) and value:
            return value
    return None


def _extract_knowledge_basis(match_result: MatchResult) -> dict | None:
    explicit = _extract_latest_trace_dict(match_result, "knowledge_basis")
    if explicit:
        return explicit

    reference_case_ids: list[str] = []
    rule_ids: list[str] = []
    method_card_ids: list[str] = []
    for step in _trace_steps(match_result):
        for item in step.get("reference_case_ids", []) or []:
            text = str(item or "").strip()
            if text and text not in reference_case_ids:
                reference_case_ids.append(text)
        for item in (step.get("quota_rule_ids", []) or []) + (step.get("rule_context_ids", []) or []):
            text = str(item or "").strip()
            if text and text not in rule_ids:
                rule_ids.append(text)
        for item in step.get("method_card_ids", []) or []:
            text = str(item or "").strip()
            if text and text not in method_card_ids:
                method_card_ids.append(text)

    if not any([reference_case_ids, rule_ids, method_card_ids]):
        return None

    return {
        "reference_case_ids": reference_case_ids,
        "rule_ids": rule_ids,
        "method_card_ids": method_card_ids,
    }


def _extract_knowledge_summary(match_result: MatchResult) -> dict | None:
    explicit = _extract_latest_trace_dict(match_result, "knowledge_summary")
    if explicit:
        return explicit

    evidence = _extract_knowledge_evidence(match_result) or {}
    summary = {
        "reference_cases_count": len(evidence.get("reference_cases") or []),
        "quota_rules_count": len(evidence.get("quota_rules") or []),
        "quota_explanations_count": len(evidence.get("quota_explanations") or []),
        "method_cards_count": len(evidence.get("method_cards") or []),
    }
    if any(summary.values()):
        return summary
    return None


def _latest_selected_reasoning(match_result: MatchResult) -> dict:
    for step in reversed(_trace_steps(match_result)):
        value = step.get("selected_reasoning")
        if isinstance(value, dict):
            return value
    return {}


def _latest_trace_candidates(match_result: MatchResult) -> list[dict]:
    for step in reversed(_trace_steps(match_result)):
        value = step.get("candidates")
        if isinstance(value, list):
            return [item for item in value if isinstance(item, dict)]
    return []


def _candidate_scores_from_payload(payload: dict | None) -> dict:
    if not isinstance(payload, dict):
        return {}
    scores = normalize_dimension_scores(payload)
    if scores:
        return scores
    reasoning = payload.get("reasoning")
    if isinstance(reasoning, dict):
        return normalize_dimension_scores(reasoning)
    return {}


def _build_topk_snapshot(match_result: MatchResult) -> tuple[list[dict], dict[str, dict]]:
    snapshot: list[dict] = []
    dimension_scores: dict[str, dict] = {}
    seen_keys: set[str] = set()

    top1_quotas = match_result.quotas or []
    top1_key = _quota_bundle_key(top1_quotas)
    top1_quota = top1_quotas[0] if top1_quotas else {}
    top1_reasoning = _latest_selected_reasoning(match_result)
    if top1_key:
        snapshot.append(
            {
                "rank": 1,
                "candidate_key": top1_key,
                "quota_ids": [str(item.get("quota_id", "") or "").strip() for item in top1_quotas if item.get("quota_id")],
                "quota_names": [str(item.get("name", "") or "").strip() for item in top1_quotas if item.get("name")],
                "unit": str(top1_quota.get("unit", "") or "").strip(),
                "confidence": _effective_confidence(match_result),
                "source": str(top1_quota.get("source", "") or "").strip(),
                "db_id": _quota_db_id(top1_quotas),
            }
        )
        dimension_scores[top1_key] = _candidate_scores_from_payload(top1_reasoning or top1_quota)
        seen_keys.add(top1_key)

    next_rank = 2
    for alt in match_result.alternatives or []:
        quota_id = str(alt.get("quota_id", "") or "").strip()
        if not quota_id:
            continue
        candidate_key = f"quota:{quota_id}"
        if candidate_key in seen_keys:
            continue
        snapshot.append(
            {
                "rank": next_rank,
                "candidate_key": candidate_key,
                "quota_ids": [quota_id],
                "quota_names": [str(alt.get("name", "") or "").strip()],
                "unit": str(alt.get("unit", "") or "").strip(),
                "confidence": _effective_confidence(alt),
                "source": str(alt.get("source", "") or "").strip(),
                "reason": str(alt.get("reason", "") or "").strip(),
            }
        )
        dimension_scores[candidate_key] = _candidate_scores_from_payload(alt)
        seen_keys.add(candidate_key)
        next_rank += 1

    for candidate in _latest_trace_candidates(match_result):
        quota_id = str(candidate.get("quota_id", "") or "").strip()
        if not quota_id:
            continue
        candidate_key = f"quota:{quota_id}"
        if candidate_key in seen_keys:
            continue
        snapshot.append(
            {
                "rank": next_rank,
                "candidate_key": candidate_key,
                "quota_ids": [quota_id],
                "quota_names": [str(candidate.get("name", "") or "").strip()],
                "unit": str(candidate.get("unit", "") or "").strip(),
                "confidence": _effective_confidence(candidate),
                "source": str(candidate.get("source", "") or "").strip(),
            }
        )
        dimension_scores[candidate_key] = _candidate_scores_from_payload(candidate)
        seen_keys.add(candidate_key)
        next_rank += 1
        if next_rank > 10:
            break

    return snapshot, dimension_scores


def _selected_rank_from_snapshot(selected_key: str, snapshot: list[dict]) -> int | None:
    if not selected_key:
        return None
    for item in snapshot:
        if str(item.get("candidate_key", "") or "").strip() == selected_key:
            try:
                return int(item.get("rank"))
            except (TypeError, ValueError):
                return None
    return None


def _record_ranking_feedback(
    *,
    task,
    match_result: MatchResult,
    user: User | None,
    action: str,
    feedback_source: str,
    selected_quotas: list[dict] | None = None,
) -> None:
    selected_quotas = selected_quotas if selected_quotas is not None else (match_result.corrected_quotas or match_result.quotas or [])
    topk_snapshot, dimension_scores = _build_topk_snapshot(match_result)
    original_key = _quota_bundle_key(match_result.quotas or [])
    selected_key = _quota_bundle_key(selected_quotas)
    original_scores = dimension_scores.get(original_key, {})
    selected_scores = dimension_scores.get(selected_key, {})
    feedback_id = RankingFeedbackDB().record_feedback(
        task_id=str(task.id),
        result_id=str(match_result.id),
        province=str(task.province or "").strip(),
        query_text=_compose_query_text(match_result),
        selected_experience_id=_quota_db_id(selected_quotas),
        original_top1_experience_id=_quota_db_id(match_result.quotas or []),
        selected_candidate_key=selected_key,
        original_top1_candidate_key=original_key,
        original_rank_of_selected=_selected_rank_from_snapshot(selected_key, topk_snapshot),
        gate_bucket=_resolve_light_status(match_result),
        topk_snapshot=topk_snapshot,
        dimension_scores_json=dimension_scores,
        misrank_primary_factor=infer_misrank_primary_factor(original_scores, selected_scores),
        feedback_source=feedback_source,
        action=action,
        actor=_actor_identity(user),
    )
    logger.debug(f"ranking feedback recorded: id={feedback_id}, result={match_result.id}, action={action}")


async def _safe_record_ranking_feedback(**kwargs) -> None:
    try:
        await asyncio.to_thread(_record_ranking_feedback, **kwargs)
    except Exception as exc:
        logger.warning(f"ranking feedback record failed: {exc}")


def _compact_feedback_trace(trace: dict | None) -> dict:
    """提取经验回流需要的 trace 摘要。"""
    if not isinstance(trace, dict):
        return {}

    payload = {}
    for key in ("path", "final_source", "final_confidence"):
        value = trace.get(key)
        if value not in (None, "", [], {}):
            payload[key] = value

    steps_out = []
    for step in trace.get("steps", []) or []:
        if not isinstance(step, dict):
            continue
        item = {}
        for key in (
            "stage",
            "selected_quota",
            "selected_reasoning",
            "candidates_count",
            "candidates",
            "quota_ids",
            "confidence",
            "reason",
            "error_type",
            "error_reason",
            "final_source",
            "final_confidence",
            "final_validation",
            "final_review_correction",
            "reasoning_engaged",
            "reasoning_conflicts",
            "reasoning_decision",
            "reasoning_compare_points",
            "reference_cases_count",
            "reference_case_ids",
            "rules_context_count",
            "rule_context_ids",
            "quota_rules_count",
            "quota_rule_ids",
            "quota_explanations_count",
            "quota_explanation_ids",
            "method_cards_count",
            "method_card_ids",
            "method_card_categories",
            "knowledge_evidence",
            "knowledge_basis",
            "knowledge_summary",
            "unified_knowledge_meta",
            "query_route",
            "batch_context",
        ):
            value = step.get(key)
            if value not in (None, "", [], {}):
                item[key] = value
        if item:
            steps_out.append(item)

    if steps_out:
        payload["steps"] = steps_out[-6:]

    return payload


def _extract_feedback_meta(trace: dict | None) -> dict:
    """从 trace 中提取经验回流可直接消费的终检与仲裁摘要。"""
    if not isinstance(trace, dict):
        return {}

    final_validation = {}
    final_review_correction = {}
    reasoning_summary = {}
    query_route = {}
    batch_context = {}
    knowledge_evidence = {}
    knowledge_basis = {}
    knowledge_summary = {}

    for step in reversed(trace.get("steps", []) or []):
        if not isinstance(step, dict):
            continue
        if not final_validation and isinstance(step.get("final_validation"), dict):
            final_validation = step.get("final_validation") or {}
        if not final_review_correction and isinstance(step.get("final_review_correction"), dict):
            final_review_correction = step.get("final_review_correction") or {}
        if not reasoning_summary and (
            step.get("reasoning_engaged")
            or step.get("reasoning_conflicts")
            or step.get("reasoning_decision")
            or step.get("reasoning_compare_points")
        ):
            reasoning_summary = {
                "engaged": bool(step.get("reasoning_engaged")),
                "decision": step.get("reasoning_decision") or {},
                "conflict_summaries": step.get("reasoning_conflicts") or [],
                "compare_points": step.get("reasoning_compare_points") or [],
            }
        if not query_route and isinstance(step.get("query_route"), dict):
            query_route = step.get("query_route") or {}
        if not batch_context and isinstance(step.get("batch_context"), dict):
            batch_context = step.get("batch_context") or {}
        if not knowledge_evidence and isinstance(step.get("knowledge_evidence"), dict):
            knowledge_evidence = step.get("knowledge_evidence") or {}
        if not knowledge_basis and isinstance(step.get("knowledge_basis"), dict):
            knowledge_basis = step.get("knowledge_basis") or {}
        if not knowledge_summary and isinstance(step.get("knowledge_summary"), dict):
            knowledge_summary = step.get("knowledge_summary") or {}
        if final_validation and reasoning_summary and query_route and batch_context and knowledge_summary:
            break

    payload = {}
    if final_validation:
        payload["final_validation"] = final_validation
    if final_review_correction:
        payload["final_review_correction"] = final_review_correction
    if reasoning_summary:
        payload["reasoning_summary"] = reasoning_summary
    if query_route:
        payload["query_route"] = query_route
    if batch_context:
        payload["batch_context"] = batch_context
    if knowledge_evidence:
        payload["knowledge_evidence"] = knowledge_evidence
    if knowledge_basis:
        payload["knowledge_basis"] = knowledge_basis
    if knowledge_summary:
        payload["knowledge_summary"] = knowledge_summary
    return payload


def _build_feedback_payload(
    match_result,
    *,
    action: str,
    review_note: str = "",
    corrected_quotas: list[dict] | None = None,
) -> dict:
    """构造写入经验库的结构化回流快照。"""
    original_quotas = match_result.quotas or []
    chosen_quotas = corrected_quotas or match_result.corrected_quotas or original_quotas
    trace_payload = _compact_feedback_trace(match_result.trace)
    payload = {
        "action": action,
        "review_note": review_note or "",
        "match_source": match_result.match_source or "",
        "confidence": match_result.confidence or 0,
        "review_status": match_result.review_status or "",
        "bill_snapshot": {
            "name": match_result.bill_name or "",
            "description": match_result.bill_description or "",
            "unit": match_result.bill_unit or "",
            "specialty": match_result.specialty or "",
        },
        "original_quotas": original_quotas,
        "selected_quotas": chosen_quotas,
        "corrected_quotas": corrected_quotas or [],
        "alternatives": (match_result.alternatives or [])[:3],
        "trace": trace_payload,
    }
    payload.update(_extract_feedback_meta(trace_payload))
    return payload


async def _apply_confirm_result(
    *,
    match_result: MatchResult,
    task,
    review_note: str,
) -> None:
    match_result.review_status = "confirmed"
    match_result.review_note = review_note

    quotas_data = match_result.quotas
    if not quotas_data:
        return

    await store_experience(
        name=match_result.bill_name,
        desc=match_result.bill_description or "",
        quota_ids=[q["quota_id"] for q in quotas_data if q.get("quota_id")],
        quota_names=[q.get("name", "") for q in quotas_data],
        reason=f"API确认: {review_note or ''}",
        specialty=match_result.specialty or "",
        province=task.province,
        confirmed=True,
        feedback_payload=_build_feedback_payload(
            match_result,
            action="confirm",
            review_note=review_note or "",
        ),
    )


async def _apply_corrected_result(
    *,
    match_result: MatchResult,
    task,
    corrected_quotas: list[dict],
    review_note: str,
    reason_prefix: str,
) -> None:
    match_result.corrected_quotas = corrected_quotas
    match_result.review_status = "corrected"
    match_result.review_note = review_note

    await store_experience(
        name=match_result.bill_name,
        desc=match_result.bill_description or "",
        quota_ids=[q["quota_id"] for q in corrected_quotas if q.get("quota_id")],
        quota_names=[q.get("name", "") for q in corrected_quotas],
        reason=f"{reason_prefix}: {review_note or ''}",
        specialty=match_result.specialty or "",
        province=task.province,
        confirmed=False,
        feedback_payload=_build_feedback_payload(
            match_result,
            action="correct",
            review_note=review_note or "",
            corrected_quotas=corrected_quotas,
        ),
    )

    if match_result.match_source and "experience" in match_result.match_source:
        await flag_disputed_experience(
            bill_name=match_result.bill_name,
            province=task.province,
            reason=f"被纠正为 {[q['quota_id'] for q in corrected_quotas if q.get('quota_id')]}; {review_note or ''}",
        )


def _strip_material_rows(source_path: str, task_id: str) -> str:
    """去掉Excel中的主材行，返回处理后的文件路径

    主材行特征：A列为空，B列是材料编码格式（CL/ZCGL/含@/补充主材/纯数字7-8位/单字"主"）。
    """
    import openpyxl

    output_dir = get_task_output_dir(uuid.UUID(task_id))
    stripped_path = str(output_dir / "output_no_material.xlsx")

    # 如果已经生成过，直接返回（同一个任务的Excel不会变）
    if Path(stripped_path).exists():
        return stripped_path

    wb = openpyxl.load_workbook(source_path)
    for ws in wb.worksheets:
        # 从下往上删，避免行号偏移
        rows_to_delete = []
        for row_idx in range(1, ws.max_row + 1):
            a_val = ws.cell(row=row_idx, column=1).value
            b_val = ws.cell(row=row_idx, column=2).value
            # A列为空、B列有值 → 可能是主材行
            if (a_val is None or str(a_val).strip() == "") and b_val:
                b_str = str(b_val).strip()
                if _is_material_code_simple(b_str):
                    rows_to_delete.append(row_idx)

        for row_idx in reversed(rows_to_delete):
            ws.delete_rows(row_idx)

    wb.save(stripped_path)
    wb.close()
    return stripped_path


def _is_material_code_simple(code: str) -> bool:
    """判断是否为材料/主材编码（简化版，和 bill_reader._is_material_code 逻辑一致）"""
    if not code:
        return False
    # "主" 单字（兜底提取的主材行用这个标记）
    if code == "主":
        return True
    if re.match(r"^CL\d", code, re.IGNORECASE):
        return True
    if re.match(r"^ZCGL\d", code, re.IGNORECASE):
        return True
    if "Z@" in code or "@" in code:
        return True
    if code.startswith("补充主材"):
        return True
    # 纯数字7-8位（广联达材料编码）
    if re.fullmatch(r"\d{7,8}", code):
        return True
    return False


@router.get("/tasks/{task_id}/results", response_model=ResultListResponse)
async def list_results(
    task_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """获取任务的匹配结果列表

    返回所有匹配结果（按序号排序），附带置信度分布统计。
    """
    await get_user_task(task_id, user, db)

    # 查询所有结果
    result = await db.execute(
        select(MatchResult)
        .where(MatchResult.task_id == task_id)
        .order_by(MatchResult.index)
    )
    items = result.scalars().all()

    # 统计置信度分布 + 审核状态
    total = len(items)
    high_conf = sum(1 for r in items if _resolve_light_status(r) == "green")
    mid_conf = sum(1 for r in items if _resolve_light_status(r) == "yellow")
    low_conf = sum(1 for r in items if _resolve_light_status(r) == "red")
    no_match = sum(1 for r in items if not r.quotas)
    # 审核维度：已确认/已纠正/待审核
    confirmed = sum(1 for r in items if r.review_status == "confirmed")
    corrected = sum(1 for r in items if r.review_status == "corrected")
    pending = total - confirmed - corrected

    summary = {
        "total": total,
        "high_confidence": high_conf,
        "mid_confidence": mid_conf,
        "low_confidence": low_conf,
        "no_match": no_match,
        "confirmed": confirmed,    # 已确认条数
        "corrected": corrected,    # 已纠正条数
        "pending": pending,        # 待审核条数
    }

    return ResultListResponse(
        items=[_to_result_response(item) for item in items],
        total=total,
        summary=summary,
    )


@router.get("/tasks/{task_id}/results/{result_id}", response_model=MatchResultResponse)
async def get_result(
    task_id: uuid.UUID,
    result_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """获取单条匹配结果详情

    包含清单信息、匹配定额、置信度、匹配说明等。
    """
    await get_user_task(task_id, user, db)

    result = await db.execute(
        select(MatchResult).where(
            MatchResult.id == result_id,
            MatchResult.task_id == task_id,
        )
    )
    match_result = result.scalar_one_or_none()
    if not match_result:
        raise HTTPException(status_code=404, detail="结果不存在")
    return _to_result_response(match_result)


@router.put("/tasks/{task_id}/results/{result_id}", response_model=MatchResultResponse)
async def correct_result(
    task_id: uuid.UUID,
    result_id: uuid.UUID,
    req: CorrectResultRequest,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """纠正或确认匹配结果。"""
    task = await get_user_task(task_id, user, db)

    result = await db.execute(
        select(MatchResult).where(
            MatchResult.id == result_id,
            MatchResult.task_id == task_id,
        )
    )
    match_result = result.scalar_one_or_none()
    if not match_result:
        raise HTTPException(status_code=404, detail="结果不存在")

    if not req.corrected_quotas:
        review_status = req.review_status or "confirmed"
        if review_status == "confirmed":
            await _apply_confirm_result(
                match_result=match_result,
                task=task,
                review_note=req.review_note or "",
            )
            await db.flush()
            await _safe_record_ranking_feedback(
                task=task,
                match_result=match_result,
                user=user,
                action="confirm",
                feedback_source="web_result_confirm",
                selected_quotas=match_result.quotas or [],
            )
        else:
            match_result.review_status = review_status
            match_result.review_note = req.review_note
            await db.flush()
        return _to_result_response(match_result)

    corrected_quotas = [q.model_dump() for q in req.corrected_quotas]
    await _apply_corrected_result(
        match_result=match_result,
        task=task,
        corrected_quotas=corrected_quotas,
        review_note=req.review_note or "",
        reason_prefix="Web端纠正",
    )
    await db.flush()
    await _safe_record_ranking_feedback(
        task=task,
        match_result=match_result,
        user=user,
        action="correct",
        feedback_source="web_result_correct",
        selected_quotas=corrected_quotas,
    )

    return _to_result_response(match_result)


@router.post("/tasks/{task_id}/results/confirm")
async def confirm_results(
    task_id: uuid.UUID,
    req: ConfirmResultsRequest,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """批量确认匹配结果。"""
    task = await get_user_task(task_id, user, db)

    result = await db.execute(
        select(MatchResult).where(
            MatchResult.task_id == task_id,
            MatchResult.id.in_(req.result_ids),
        )
    )
    results = result.scalars().all()

    updated = 0
    skipped = 0
    skipped_low_conf = 0
    confirmed_records = []
    confirmed_results = []
    for r in results:
        if r.review_status == "corrected":
            skipped += 1
            continue
        if not _is_confirmable_result(r):
            skipped_low_conf += 1
            continue
        if r.review_status != "confirmed":
            r.review_status = "confirmed"
            updated += 1
            confirmed_results.append(r)
            quotas_data = r.corrected_quotas or r.quotas
            if quotas_data:
                confirmed_records.append({
                    "name": r.bill_name,
                    "desc": r.bill_description or "",
                    "quota_ids": [q["quota_id"] for q in quotas_data if q.get("quota_id")],
                    "quota_names": [q.get("name", "") for q in quotas_data],
                    "specialty": r.specialty or "",
                    "feedback_payload": _build_feedback_payload(
                        r,
                        action="confirm",
                        review_note="",
                    ),
                })

    await db.flush()

    if confirmed_records:
        await store_experience_batch(
            records=confirmed_records,
            province=task.province,
            reason="Web端确认",
            confirmed=True,
        )

    for r in confirmed_results:
        await _safe_record_ranking_feedback(
            task=task,
            match_result=r,
            user=user,
            action="confirm",
            feedback_source="web_batch_confirm",
            selected_quotas=r.corrected_quotas or r.quotas or [],
        )

    return {
        "confirmed": updated,
        "skipped_corrected": skipped,
        "skipped_low_confidence": skipped_low_conf,
        "total": len(results),
    }


@router.get("/tasks/{task_id}/export")
async def export_results(
    task_id: uuid.UUID,
    materials: bool = False,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """导出匹配结果Excel

    参数 materials：是否带主材行（默认不带，管理员可在前端勾选）。
    """
    task = await get_user_task(task_id, user, db)

    if task.status != "completed":
        raise HTTPException(status_code=400, detail="任务尚未完成，无法导出")

    if not task.output_path or not Path(task.output_path).exists():
        raise HTTPException(status_code=404, detail="输出文件不存在")

    # 构造下载文件名（原始文件名 + _定额匹配结果）
    download_name = _task_download_stem(task.original_filename) + "_定额匹配结果.xlsx"

    # 带主材：直接返回完整文件
    if materials:
        return FileResponse(
            path=task.output_path,
            filename=download_name,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    # 不带主材：去掉主材行后返回
    stripped_path = await asyncio.to_thread(
        _strip_material_rows, task.output_path, str(task_id)
    )
    return FileResponse(
        path=stripped_path,
        filename=download_name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@router.get("/tasks/{task_id}/export-final")
async def export_final(
    task_id: uuid.UUID,
    materials: bool = False,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """导出含纠正结果的Excel（实时从数据库生成）

    和 /export 的区别：
    - /export 返回匹配时生成的静态文件，不含后续纠正
    - /export-final 从数据库读最新结果（含纠正），重新生成Excel

    OpenClaw 确认+纠正完后调这个接口下载最终版。
    """
    task = await get_user_task(task_id, user, db)

    if task.status != "completed":
        raise HTTPException(status_code=400, detail="任务尚未完成，无法导出")

    # 从数据库读取所有结果
    result = await db.execute(
        select(MatchResult)
        .where(MatchResult.task_id == task_id)
        .order_by(MatchResult.index)
    )
    items = result.scalars().all()

    if not items:
        raise HTTPException(status_code=404, detail="没有匹配结果")

    # export-final 始终按最新结果重建，避免继续返回历史 output.xlsx 中已固化的错位内容
    rebuilt_results = _build_rebuilt_results_from_json(task, items)
    if rebuilt_results is None:
        rebuilt_results = _build_rebuilt_results_from_source_file(task, items)
    if rebuilt_results is None:
        rebuilt_results = _build_rebuilt_results_from_db(items)

    # ?????????OutputWriter ?????????
    original_file = None
    if task.file_path and Path(task.file_path).exists():
        original_file = task.file_path
    elif task.output_path and Path(task.output_path).exists():
        # ???????????????????? output.xlsx ??????
        original_file = task.output_path

    # 输出到临时文件
    output_dir = get_task_output_dir(uuid.UUID(str(task_id)))
    final_path = str(output_dir / "output_final.xlsx")

    # OutputWriter 是同步的，放到线程里跑
    def _generate():
        from src.output_writer import OutputWriter
        writer = OutputWriter()
        writer.write_results(rebuilt_results, final_path, original_file=original_file)

    try:
        await asyncio.to_thread(_generate)
    except Exception as e:
        logger.error(f"生成纠正后Excel失败: {e}")
        raise HTTPException(status_code=500, detail=f"生成Excel失败: {e}")

    download_name = _task_download_stem(task.original_filename) + "_最终结果.xlsx"
    export_path = final_path
    # 不带主材时去掉主材行
    if not materials:
        export_path = await asyncio.to_thread(
            _strip_material_rows, final_path, str(task_id) + "_final"
        )
    return FileResponse(
        path=export_path,
        filename=download_name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
