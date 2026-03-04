"""价格回填工具测试"""
import shutil
import uuid
from pathlib import Path

import openpyxl

# 添加项目根目录到path
import sys
sys.path.insert(0, str(Path(__file__).parent.parent))

from tools.price_backfill import (
    _detect_original_structure,
    _read_gld_prices,
    _build_mapping,
    backfill,
)


def _new_tmp_dir() -> Path:
    root = Path("output") / "test_tmp"
    root.mkdir(parents=True, exist_ok=True)
    d = root / f"backfill-{uuid.uuid4().hex}"
    d.mkdir(parents=True, exist_ok=False)
    return d


# ================================================================
# 测试1：甲方原始Excel结构检测
# ================================================================
def test_detect_original_structure():
    """能正确识别甲方原始Excel的表头和数据行"""
    tmp_dir = _new_tmp_dir()
    xlsx_path = tmp_dir / "original.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active

    # 模拟甲方清单格式
    ws.append(["分部分项工程项目清单"])
    ws.append([])
    ws.append(["序号", "项目内容", "计量单位", "工程数量",
               "综合单价（不含税）", "综合单价（含税）", "合价（不含税）", "合价（含税）"])
    ws.append([29, "电气火灾监控模块", "套", 2, 0, 0, 0, 0])
    ws.append([30, "配线WDZN-RYS-2*1.5", "米", 115.25, 0, 0, 0, 0])
    ws.append([31, "本专业小计", "元", None, None, None, 0, 0])
    try:
        wb.save(xlsx_path)
        wb.close()

        wb2 = openpyxl.load_workbook(str(xlsx_path))
        ws2 = wb2.active
        result = _detect_original_structure(ws2)
        wb2.close()

        assert result["header_row"] == 3
        assert "name" in result["col_map"]
        assert "unit_price" in result["col_map"] or "total_price" in result["col_map"]
        # 应该有2条数据（"本专业小计"被过滤）
        assert len(result["items"]) == 2
        assert result["items"][0]["name"] == "电气火灾监控模块"
        assert result["items"][1]["name"] == "配线WDZN-RYS-2*1.5"
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)


# ================================================================
# 测试2：广联达导出价格读取
# ================================================================
def test_read_gld_prices():
    """能正确读取广联达导出Excel的价格"""
    tmp_dir = _new_tmp_dir()
    xlsx_path = tmp_dir / "gld_export.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active

    # 模拟广联达导出格式
    ws.append(["序号", "项目编码", "项目名称", "项目特征描述",
               "计量单位", "工程量", "综合单价", "合价"])
    ws.append([1, "030901001001", "电气火灾监控模块", "消防设备", "套", 2, 1500.00, 3000.00])
    ws.append([None, "C9-3-45", "火灾监控模块安装", None, "套", 2, 1500.00, 3000.00])
    ws.append([2, "030901001002", "配线WDZN-RYS-2*1.5", "电线", "米", 115.25, 8.50, 979.63])
    try:
        wb.save(xlsx_path)
        wb.close()

        wb2 = openpyxl.load_workbook(str(xlsx_path))
        ws2 = wb2.active
        prices = _read_gld_prices(ws2)
        wb2.close()

        # 应该只读到2条清单行（跳过定额行）
        assert len(prices) == 2
        assert prices[0]["name"] == "电气火灾监控模块"
        assert prices[0]["unit_price"] == 1500.00
        assert prices[0]["total_price"] == 3000.00
        assert prices[1]["unit_price"] == 8.50
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)


# ================================================================
# 测试3：按序号映射
# ================================================================
def test_mapping_by_index():
    """按序号一一对应映射"""
    original_items = [
        {"row": 4, "index": "29", "name": "电气火灾监控模块"},
        {"row": 5, "index": "30", "name": "配线WDZN-RYS-2*1.5"},
    ]
    price_data = [
        {"index": "29", "name": "电气火灾监控模块", "unit_price": 1500.0, "total_price": 3000.0},
        {"index": "30", "name": "配线WDZN-RYS-2*1.5", "unit_price": 8.5, "total_price": 979.63},
    ]

    mapping = _build_mapping(original_items, price_data)
    assert len(mapping) == 2
    assert mapping[0]["match_method"] == "index"
    assert mapping[0]["unit_price"] == 1500.0
    assert mapping[1]["total_price"] == 979.63


# ================================================================
# 测试4：序号不匹配时用名称兜底
# ================================================================
def test_mapping_fallback_to_name():
    """序号不匹配时，用名称相似度兜底"""
    original_items = [
        {"row": 4, "index": "1", "name": "LED管型灯具"},
        {"row": 5, "index": "2", "name": "15寸LED圆盘灯"},
    ]
    # 广联达导出的序号不同
    price_data = [
        {"index": "101", "name": "LED管型灯具安装", "unit_price": 50.0, "total_price": 450.0},
        {"index": "102", "name": "15寸LED圆盘灯安装", "unit_price": 80.0, "total_price": 1120.0},
    ]

    mapping = _build_mapping(original_items, price_data)
    assert len(mapping) == 2
    # 应该用名称匹配兜底
    assert mapping[0]["match_method"].startswith("name")
    assert mapping[0]["unit_price"] == 50.0


# ================================================================
# 测试5：完整回填流程
# ================================================================
def test_backfill_end_to_end():
    """完整回填流程：原始文件→广联达文件→回填价格"""
    tmp_dir = _new_tmp_dir()

    # 创建甲方原始文件
    orig_path = tmp_dir / "甲方清单.xlsx"
    wb1 = openpyxl.Workbook()
    ws1 = wb1.active
    ws1.append(["序号", "项目内容", "计量单位", "工程数量", "综合单价", "合价"])
    ws1.append([1, "LED管型灯具", "个", 9, 0, 0])
    ws1.append([2, "电线WD-YJV3X2.5", "米", 3848, 0, 0])
    wb1.save(str(orig_path))
    wb1.close()

    # 创建广联达导出文件（带价格）
    gld_path = tmp_dir / "广联达导出.xlsx"
    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    ws2.append(["序号", "项目编码", "项目名称", "项目特征描述",
                "计量单位", "工程量", "综合单价", "合价"])
    ws2.append([1, None, "LED管型灯具", None, "个", 9, 45.50, 409.50])
    ws2.append([2, None, "电线WD-YJV3X2.5", None, "米", 3848, 5.20, 20009.60])
    wb2.save(str(gld_path))
    wb2.close()

    try:
        result_path, mapping = backfill(
            str(orig_path), str(gld_path), dry_run=False)

        assert result_path is not None
        assert Path(result_path).exists()
        assert "已回填" in result_path

        # 验证回填后的文件
        wb3 = openpyxl.load_workbook(result_path)
        ws3 = wb3.active
        # 第2行是第1条数据
        assert ws3.cell(row=2, column=5).value == 45.50   # 综合单价
        assert ws3.cell(row=2, column=6).value == 409.50   # 合价
        assert ws3.cell(row=3, column=5).value == 5.20
        assert ws3.cell(row=3, column=6).value == 20009.60
        wb3.close()
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)


# ================================================================
# 测试6：dry-run模式不写文件
# ================================================================
def test_dry_run_mode():
    """dry-run模式只预览，不生成文件"""
    tmp_dir = _new_tmp_dir()

    orig_path = tmp_dir / "原始.xlsx"
    wb1 = openpyxl.Workbook()
    ws1 = wb1.active
    ws1.append(["序号", "项目内容", "单价", "合价"])
    ws1.append([1, "开关", 0, 0])
    wb1.save(str(orig_path))
    wb1.close()

    gld_path = tmp_dir / "导出.xlsx"
    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    ws2.append(["序号", "项目名称", "综合单价", "合价"])
    ws2.append([1, "开关", 15.0, 15.0])
    wb2.save(str(gld_path))
    wb2.close()

    try:
        result_path, mapping = backfill(
            str(orig_path), str(gld_path), dry_run=True)

        assert result_path is None  # dry-run不生成文件
        assert mapping is not None
        assert len(mapping) == 1
        # 检查没有生成_已回填文件
        backfill_files = list(tmp_dir.glob("*已回填*"))
        assert len(backfill_files) == 0
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)
