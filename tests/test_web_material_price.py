# -*- coding: utf-8 -*-

import asyncio
from pathlib import Path
import sys

import openpyxl


BACKEND_ROOT = Path(__file__).resolve().parents[1] / "web" / "backend"
if str(BACKEND_ROOT) not in sys.path:
    sys.path.append(str(BACKEND_ROOT))

from app.api import material_price as material_price_api  # noqa: E402


def test_parse_sheet_exposes_material_name_and_spec_columns(tmp_path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "主材表"
    ws.append(["材料编码", "材料名称", "规格型号", "单位", "数量", "单价"])
    ws.append(["26010101", "镀锌钢管", "DN25", "m", 12, 18.5])

    result = material_price_api._parse_sheet(ws)

    assert len(result["materials"]) == 1
    material = result["materials"][0]
    assert material["name"] == "镀锌钢管"
    assert material["name_col"] == 2
    assert material["spec_col"] == 3
    assert material["price_col"] == 6


def test_write_material_updates_writes_name_spec_and_price(tmp_path: Path):
    file_path = tmp_path / "reviewed.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "主材表"
    ws.append(["材料编码", "材料名称", "规格型号", "单位", "数量", "单价"])
    ws.append(["26010101", "原主材名", "DN25", "m", 12, None])
    wb.save(file_path)
    wb.close()

    written = material_price_api._do_write_material_updates(
        str(file_path),
        [
            {
                "row": 2,
                "sheet": "主材表",
                "name_col": 2,
                "spec_col": 3,
                "final_spec": "DN32",
                "final_name": "镀锌钢管",
                "price_col": 6,
                "final_price": 18.5,
            }
        ],
    )

    assert written == 1

    wb2 = openpyxl.load_workbook(file_path, data_only=True)
    ws2 = wb2["主材表"]
    assert ws2.cell(row=2, column=2).value == "镀锌钢管"
    assert ws2.cell(row=2, column=3).value == "DN32"
    assert ws2.cell(row=2, column=6).value == 18.5
    wb2.close()


def test_z_prefix_rows_are_treated_as_material_targets():
    assert material_price_api._classify_row("031001007001", "复合管", "1") == "bill"
    assert material_price_api._classify_row("A10-1-366", "给排水管道", "") == "quota"
    assert material_price_api._classify_row("Z1728A01B01BY", "复合管", "") == "material"
    assert material_price_api._classify_row("Z1811A07B01BF", "给水室内钢塑复合管螺纹管件", "") == "material"


def test_parse_sheet_prefills_z_material_name_and_spec_from_bill_desc():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "安装工程"
    ws.append(["序号", "项目编码", "项目名称", "项目特征描述", "计量单位", "工程量", "综合单价"])
    ws.append(["1", "031001007001", "复合管", "1.安装部位:室内\n2.介质:给水（冷水）\n3.材质、规格:衬塑钢管 DN100", "m", 14.03, None])
    ws.append(["", "A10-1-366", "给排水管道 室内 钢塑复合管（螺纹连接） 公称直径（mm以内）100", "", "10m", 1.403, None])
    ws.append(["", "Z1728A01B01BY", "复合管", "", "m", 10.02, None])
    ws.append(["", "Z1811A07B01BF", "给水室内钢塑复合管螺纹管件", "", "个", 4.15, None])

    result = material_price_api._parse_sheet(ws)

    materials = [m for m in result["materials"] if m["code"].startswith("Z")]
    assert len(materials) == 2
    assert materials[0]["suggested_name"] == "衬塑钢管"
    assert materials[0]["suggested_spec"] == "DN100"
    assert materials[0]["normalized_name"] == "衬塑钢管"
    assert materials[0]["normalized_spec"] == "DN100"
    assert materials[0]["object_type"] == "pipe"
    assert materials[0]["normalization_confidence"] == "medium"
    assert materials[1]["name"] == "给水室内钢塑复合管螺纹管件"
    assert materials[1]["suggested_name"] == ""
    assert materials[1]["suggested_spec"] == "DN100"
    assert materials[1]["normalized_name"] == "给水室内钢塑复合管螺纹管件"
    assert materials[1]["normalized_spec"] == "DN100"


def test_build_normalized_material_fields_preserves_critical_spec():
    normalized = material_price_api._build_normalized_material_fields(
        "绿地灌溉管线安装",
        "",
        "绿地灌溉管线安装 De16",
        "1.材质:滴水管PE\n2.规格:De16*1.6MPa\n3.连接形式:热熔连接",
    )

    assert normalized["suggested_name"] == "滴水管PE"
    assert normalized["suggested_spec"] == "De16"
    assert normalized["normalized_name"] == "滴水管PE"
    assert normalized["normalized_spec"] == "De16"
    assert normalized["critical_spec_text"] == "1.6MPa"
    assert normalized["normalized_query_text"] == "滴水管PE De16 1.6MPa"
    assert normalized["object_type"] == "pipe"
    assert normalized["normalization_confidence"] == "high"


def test_build_normalized_material_fields_keeps_pipe_fitting_name_when_desc_suggests_pipe():
    normalized = material_price_api._build_normalized_material_fields(
        "\u7ed9\u6c34\u70ed\u7194\u7ba1\u4ef6",
        "",
        "\u7ed9\u6c34\u7ba1HDPE De32",
        "1.\u6750\u8d28:\u7ed9\u6c34\u7ba1HDPE\n2.\u89c4\u683c:De32*1.60MPa\n3.\u8fde\u63a5\u5f62\u5f0f:\u70ed\u7194\u8fde\u63a5",
    )

    assert normalized["suggested_name"] == ""
    assert normalized["suggested_spec"] == "De32"
    assert normalized["normalized_name"] == "\u7ed9\u6c34\u70ed\u7194\u7ba1\u4ef6"
    assert normalized["normalized_spec"] == "De32"
    assert normalized["object_type"] == "pipe_fitting"


def test_build_normalized_material_fields_prefers_specific_equipment_name_from_desc():
    normalized = material_price_api._build_normalized_material_fields(
        "\u79bb\u5fc3\u5f0f\u6cf5",
        "",
        "\u79bb\u5fc3\u5f0f\u6cf5",
        "1.\u540d\u79f0:\u79fb\u52a8\u5f0f\u6f5c\u6c34\u6392\u6c61\u6cf5\uff08\u8bbe\u5907\u7532\u4f9b\uff09\n2.\u578b\u53f7\u3001\u7535\u673a\u529f\u7387:Q=50m3/h,H=0.1MPa,N=5kw",
    )

    assert normalized["suggested_name"] == "\u79fb\u52a8\u5f0f\u6f5c\u6c34\u6392\u6c61\u6cf5"
    assert normalized["normalized_name"] == "\u79fb\u52a8\u5f0f\u6f5c\u6c34\u6392\u6c61\u6cf5"
    assert normalized["object_type"] == "equipment"


def test_extract_material_from_desc_supports_separate_material_and_spec_fields():
    info = material_price_api._extract_material_from_desc(
        "1.安装部位:室内\n2.介质:给水（冷水）\n3.材质:PPR管\n4.规格:De25 S5"
    )
    assert info["name"] == "PPR管"
    assert info["spec"] == "De25 S5"


def test_valve_rows_use_bill_type_for_generic_valve_names():
    suggested_name, suggested_spec = material_price_api._suggest_material_from_bill_context(
        "螺纹阀门",
        "减压器组成安装",
        "1.类型:可调式减压阀\n2.材质:黄铜\n3.规格:DN32\n4.连接形式:螺纹连接",
    )
    assert suggested_name == "黄铜可调式减压阀"
    assert suggested_spec == "DN32"


def test_non_valve_accessories_keep_name_but_inherit_spec():
    suggested_name, suggested_spec = material_price_api._suggest_material_from_bill_context(
        "螺纹Y型过滤器",
        "减压器组成安装",
        "1.类型:可调式减压阀\n2.材质:黄铜\n3.规格:DN32\n4.连接形式:螺纹连接",
    )
    assert suggested_name == ""
    assert suggested_spec == "DN32"


def test_non_generic_rows_do_not_get_renamed_from_bill_context():
    suggested_name, suggested_spec = material_price_api._suggest_material_from_bill_context(
        "法兰压盖",
        "镀锌钢管",
        "1.规格:DN100\n2.连接形式:法兰连接",
    )
    assert suggested_name == ""
    assert suggested_spec == "DN100"


def test_generic_pipe_prefers_specific_bill_name_and_spec():
    suggested_name, suggested_spec = material_price_api._suggest_material_from_bill_context(
        "钢管",
        "镀锌钢管",
        "1.安装部位:室外\n2.介质:雨水\n3.规格、压力等级:内外壁镀锌钢管DN100\n4.连接形式:沟槽连接",
    )
    assert suggested_name == "镀锌钢管"
    assert suggested_spec == "DN100"


def test_generic_filter_uses_connection_type_and_spec():
    suggested_name, suggested_spec = material_price_api._suggest_material_from_bill_context(
        "法兰阀门",
        "法兰阀门",
        "1.类型:过滤器\n2.规格、压力等级:DN100\n3.连接形式:法兰连接",
    )
    assert suggested_name == "法兰过滤器"
    assert suggested_spec == "DN100"


def test_upvc_pipe_rows_keep_material_family_and_spec():
    suggested_name, suggested_spec = material_price_api._suggest_material_from_bill_context(
        "塑料排水管",
        "套管",
        "1.介质:排水\n2.材质、规格:U-PVC De110\n3.连接形式:承插式胶粘剂粘接",
    )
    assert suggested_name == "U-PVC塑料排水管"
    assert suggested_spec == "De110"


def test_device_rows_merge_material_name_and_device_name():
    suggested_name, suggested_spec = material_price_api._suggest_material_from_bill_context(
        "不锈钢",
        "地漏",
        "1.名称:地漏\n2.材质:不锈钢\n3.型号、规格:DN50",
    )
    assert suggested_name == "不锈钢地漏"
    assert suggested_spec == "DN50"


def test_plastic_valve_rows_use_desc_type_when_name_is_generic_family():
    suggested_name, suggested_spec = material_price_api._suggest_material_from_bill_context(
        "塑料阀门",
        "塑料阀门 De50 PE",
        "1.类型:泄水阀\n2.规格:De50 PE\n3.连接形式:熔接",
    )
    assert suggested_name == "泄水阀"
    assert suggested_spec == "De50 PE"


def test_metal_valve_rows_use_desc_type_when_name_is_generic_family():
    suggested_name, suggested_spec = material_price_api._suggest_material_from_bill_context(
        "\u91d1\u5c5e\u9600\u95e8",
        "\u91d1\u5c5e\u9600\u95e8",
        "1.\u7c7b\u578b:\u95f8\u9600\n2.\u6750\u8d28:\u7403\u58a8\u94f8\u94c1\n3.\u89c4\u683c\u3001\u538b\u529b\u7b49\u7ea7:DN100 1.0MPa\n4.\u8fde\u63a5\u5f62\u5f0f:\u6cd5\u5170\u8fde\u63a5\n5.\u542b\u6cd5\u5170",
    )
    assert suggested_name == "\u7403\u58a8\u94f8\u94c1\u95f8\u9600"
    assert suggested_spec == "DN100"


def test_flanged_generic_valve_keeps_specific_valve_name_after_normalization():
    normalized = material_price_api._build_normalized_material_fields(
        "\u6cd5\u5170\u9600\u95e8",
        "",
        "\u91d1\u5c5e\u9600\u95e8",
        "1.\u7c7b\u578b:\u95f8\u9600\n2.\u6750\u8d28:\u7403\u58a8\u94f8\u94c1\n3.\u89c4\u683c\u3001\u538b\u529b\u7b49\u7ea7:DN100 1.0MPa\n4.\u8fde\u63a5\u5f62\u5f0f:\u6cd5\u5170\u8fde\u63a5\n5.\u542b\u6cd5\u5170",
    )
    assert normalized["suggested_name"] == "\u7403\u58a8\u94f8\u94c1\u95f8\u9600"
    assert normalized["normalized_name"] == "\u7403\u58a8\u94f8\u94c1\u95f8\u9600"
    assert normalized["normalized_spec"] == "DN100"
    assert normalized["object_type"] == "valve"


def test_installation_item_name_yields_material_name_from_desc():
    suggested_name, suggested_spec = material_price_api._suggest_material_from_bill_context(
        "绿地灌溉管线安装",
        "绿地灌溉管线安装 De16",
        "1.材质:滴水管PE\n2.规格:De16*1.6MPa\n3.连接形式:热熔连接\n4.压力试验及吹、洗设计要求:符合设计及相关规范\n5.安装部位:明敷\n6.介质:水",
    )
    assert suggested_name == "滴水管PE"
    assert suggested_spec == "De16"


def test_installation_item_name_does_not_override_to_bare_material_token():
    suggested_name, suggested_spec = material_price_api._suggest_material_from_bill_context(
        "绿地灌溉管线安装",
        "绿地灌溉管线安装 De16",
        "1.材质:PE\n2.规格:De16\n3.连接形式:热熔连接",
    )
    assert suggested_name == ""
    assert suggested_spec == "De16"


def test_spec_like_material_name_falls_back_to_specific_bill_name():
    suggested_name, suggested_spec = material_price_api._suggest_material_from_bill_context(
        "DN100*DN100*DN80",
        "不锈钢无缝三通",
        "1.规格:DN100*DN100*DN80\n2.介质:水\n3.连接形式:氩弧焊",
    )
    assert suggested_name == "不锈钢无缝三通"
    assert suggested_spec == "DN100*DN100*DN80"


def test_generic_material_name_prefers_specific_desc_name_over_generic_bill_name():
    suggested_name, suggested_spec = material_price_api._suggest_material_from_bill_context(
        "塑料给水管",
        "给排水管道 室内塑料给水管（热熔连接）",
        "1.名称:PP-R管\n2.规格、压力等级:dn25, PN16\n3.连接形式:热熔",
    )
    assert suggested_name == "PP-R管"
    assert suggested_spec == "dn25"


def test_spec_like_material_name_can_fall_back_to_bill_name_without_desc():
    suggested_name, suggested_spec = material_price_api._suggest_material_from_bill_context(
        "DN100*DN100*DN80",
        "不锈钢无缝三通",
        "",
    )
    assert suggested_name == "不锈钢无缝三通"
    assert suggested_spec == "DN100*DN100*DN80"


def test_generic_material_name_with_inline_spec_prefers_specific_desc_name():
    suggested_name, suggested_spec = material_price_api._suggest_material_from_bill_context(
        "塑料给水管 dn25",
        "给排水管道 室内塑料给水管（热熔连接）",
        "1.名称:PP-R管\n2.规格、压力等级:dn25, PN16\n3.连接形式:热熔",
    )
    assert suggested_name == "PP-R管"
    assert suggested_spec == "dn25"


def test_build_normalized_material_fields_uses_bill_fallback_when_raw_name_is_only_spec():
    normalized = material_price_api._build_normalized_material_fields(
        "DN100*DN100*DN80",
        "",
        "不锈钢无缝三通",
        "",
    )
    assert normalized["normalized_name"] == "不锈钢无缝三通"
    assert normalized["normalized_spec"] == "DN100*DN100*DN80"


def test_qualified_valve_name_keeps_original_material_qualifier():
    suggested_name, suggested_spec = material_price_api._suggest_material_from_bill_context(
        "黄铜减压阀",
        "减压器组成安装",
        "1.类型:可调式减压阀\n2.规格:DN32\n3.连接形式:螺纹连接",
    )
    assert suggested_name == ""
    assert suggested_spec == "DN32"


def test_write_material_updates_merges_spec_into_name_when_no_spec_column(tmp_path: Path):
    file_path = tmp_path / "reviewed-no-spec.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["编码", "名称", "单价"])
    ws.append(["M001", "原主材", None])
    wb.save(file_path)
    wb.close()

    written = material_price_api._do_write_material_updates(
        str(file_path),
        [
            {
                "row": 2,
                "sheet": "Sheet1",
                "name_col": 2,
                "final_name": "镀锌钢管",
                "spec_col": None,
                "final_spec": "DN32",
                "price_col": 3,
                "final_price": 18.5,
            }
        ],
    )

    assert written == 1

    wb2 = openpyxl.load_workbook(file_path, data_only=True)
    ws2 = wb2["Sheet1"]
    assert ws2.cell(row=2, column=2).value == "镀锌钢管 DN32"
    assert ws2.cell(row=2, column=3).value == 18.5
    wb2.close()


def test_write_material_updates_inserts_gldjc_link_after_spec_column(tmp_path: Path):
    file_path = tmp_path / "reviewed-link.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "主材表"
    ws.append(["材料编码", "材料名称", "规格型号", "单价"])
    ws.append(["26010101", "原主材名", "DN25", None])
    wb.save(file_path)
    wb.close()

    written = material_price_api._do_write_material_updates(
        str(file_path),
        [
            {
                "row": 2,
                "sheet": "主材表",
                "header_row": 1,
                "name_col": 2,
                "final_name": "镀锌钢管",
                "spec_col": 3,
                "final_spec": "DN32",
                "price_col": 4,
                "final_price": 18.5,
                "lookup_url": "https://www.gldjc.com/scj/so.html?keyword=%E9%95%80%E9%94%8C%E9%92%A2%E7%AE%A1%20DN32&l=1",
                "lookup_label": "贵盈 | 镀锌钢管 DN32 | m | 18.50",
            }
        ],
    )

    assert written == 1

    wb2 = openpyxl.load_workbook(file_path, data_only=False)
    ws2 = wb2["主材表"]
    assert ws2.cell(row=1, column=5).value == "广材网链接"
    assert ws2.cell(row=2, column=2).value == "镀锌钢管"
    assert ws2.cell(row=2, column=3).value == "DN32"
    assert ws2.cell(row=2, column=4).value == 18.5
    assert ws2.cell(row=2, column=5).value == "贵盈 | 镀锌钢管 DN32 | m | 18.50"
    assert ws2.cell(row=2, column=5).hyperlink.target == "https://www.gldjc.com/scj/so.html?keyword=%E9%95%80%E9%94%8C%E9%92%A2%E7%AE%A1%20DN32&l=1"
    wb2.close()


def test_write_material_updates_inserts_gldjc_link_after_name_when_no_spec_column(tmp_path: Path):
    file_path = tmp_path / "reviewed-link-no-spec.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["编码", "名称", "单价"])
    ws.append(["M001", "原主材", None])
    wb.save(file_path)
    wb.close()

    written = material_price_api._do_write_material_updates(
        str(file_path),
        [
            {
                "row": 2,
                "sheet": "Sheet1",
                "header_row": 1,
                "name_col": 2,
                "final_name": "镀锌钢管",
                "spec_col": None,
                "final_spec": "DN32",
                "price_col": 3,
                "final_price": 18.5,
                "lookup_url": "https://www.gldjc.com/scj/so.html?keyword=%E9%95%80%E9%94%8C%E9%92%A2%E7%AE%A1%20DN32&l=1",
                "lookup_label": "贵盈 | 镀锌钢管 DN32 | m | 18.50",
            }
        ],
    )

    assert written == 1

    wb2 = openpyxl.load_workbook(file_path, data_only=False)
    ws2 = wb2["Sheet1"]
    assert ws2.cell(row=1, column=4).value == "广材网链接"
    assert ws2.cell(row=2, column=2).value == "镀锌钢管 DN32"
    assert ws2.cell(row=2, column=3).value == 18.5
    assert ws2.cell(row=2, column=4).value == "贵盈 | 镀锌钢管 DN32 | m | 18.50"
    assert ws2.cell(row=2, column=4).hyperlink.target == "https://www.gldjc.com/scj/so.html?keyword=%E9%95%80%E9%94%8C%E9%92%A2%E7%AE%A1%20DN32&l=1"
    wb2.close()


def test_write_material_updates_strips_critical_spec_from_export_link_label(tmp_path: Path):
    file_path = tmp_path / "reviewed-link-sanitized.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "主材表"
    ws.append(["材料编码", "材料名称", "规格型号", "单价"])
    ws.append(["26010101", "原主材名", "De63", None])
    wb.save(file_path)
    wb.close()

    written = material_price_api._do_write_material_updates(
        str(file_path),
        [
            {
                "row": 2,
                "sheet": "主材表",
                "header_row": 1,
                "name_col": 2,
                "final_name": "给水管HDPE",
                "spec_col": 3,
                "final_spec": "De63",
                "price_col": 4,
                "final_price": 128.82,
                "lookup_url": "https://www.gldjc.com/scj/so.html?keyword=HDPE%20De63&l=440000",
                "lookup_label": "广东信息价 | 给水管HDPE De63 | 关键规格: 1.60MPa | m | 128.82",
                "critical_spec_text": "1.60MPa",
            }
        ],
    )

    assert written == 1

    wb2 = openpyxl.load_workbook(file_path, data_only=False)
    ws2 = wb2["主材表"]
    assert ws2.cell(row=2, column=5).value == "广东信息价 | 给水管HDPE De63 | m | 128.82"
    assert ws2.cell(row=2, column=5).hyperlink.target == "https://www.gldjc.com/scj/so.html?keyword=HDPE%20De63&l=440000"
    wb2.close()


def test_pipe_fitting_row_uses_specific_name_from_desc_name_field():
    suggested_name, suggested_spec = material_price_api._suggest_material_from_bill_context(
        "法兰铸铁管件",
        "法兰铸铁管件",
        "1.名称:VSSJAF-10型双法兰传力接头\n2.规格:DN700，P=1.0MPa",
    )
    assert suggested_name == "VSSJAF-10型双法兰传力接头"
    assert suggested_spec == "DN700"


def test_pipe_fitting_row_builds_specific_shape_from_spec_and_material():
    suggested_name, suggested_spec = material_price_api._suggest_material_from_bill_context(
        "碳钢对焊管件",
        "碳钢对焊管件",
        "1.材质:碳钢\n2.规格:三通 DN600\n3.焊接方法:电弧焊",
    )
    assert suggested_name == "碳钢对焊三通"
    assert suggested_spec == "DN600"


def test_pipe_fitting_row_supports_single_line_numbered_desc():
    suggested_name, suggested_spec = material_price_api._suggest_material_from_bill_context(
        "碳钢对焊管件",
        "碳钢对焊管件 DN600",
        "1.材质:碳钢 2.规格:三通 DN600 3.焊接方法:电弧焊",
    )
    assert suggested_name == "碳钢对焊三通"
    assert suggested_spec == "DN600"


def test_specific_pipe_fitting_name_gets_material_prefix():
    suggested_name, suggested_spec = material_price_api._suggest_material_from_bill_context(
        "90°弯头",
        "90°弯头 DN200",
        "1.材质:碳钢\n2.规格:90°弯头 DN200\n3.连接方式:电弧焊",
    )
    assert suggested_name == "碳钢90°弯头"
    assert suggested_spec == "DN200"


def test_composite_material_name_can_still_use_specific_pipe_fitting_type():
    suggested_name, suggested_spec = material_price_api._suggest_material_from_bill_context(
        "钢板(综合)",
        "90°弯头 DN200",
        "1.材质:碳钢\n2.规格:90°弯头 DN200\n3.连接方式:电弧焊",
    )
    assert suggested_name == "碳钢90°弯头"
    assert suggested_spec == "DN200"


def test_do_lookup_does_not_use_gldjc_market_cache(monkeypatch):
    monkeypatch.setattr(material_price_api, "_material_db_ready", lambda: False)
    monkeypatch.setattr(
        material_price_api,
        "_load_material_price_cache",
        lambda: {
            "泄水阀|DE63|个|广东": {
                "price_with_tax": 128.82,
                "searched_keyword": "泄水阀 De63",
                "gldjc_url": "https://www.gldjc.com/scj/so.html?keyword=%E6%B3%84%E6%B0%B4%E9%98%80%20De63&l=440000",
                "match_label": "品牌A | 泄水阀 De63 | 个 | 128.82",
            },
            "泄水阀|DE63|个|全国": {
                "price_with_tax": 3489.00,
                "searched_keyword": "泄水阀 De63",
                "gldjc_url": "https://www.gldjc.com/scj/so.html?keyword=%E6%B3%84%E6%B0%B4%E9%98%80%20De63&l=1",
                "match_label": "品牌B | 泄水阀 De63 | 个 | 3489.00",
            },
        },
    )

    results = material_price_api._do_lookup(
        [{"name": "泄水阀", "spec": "De63", "unit": "个"}],
        province="广东",
        city="",
        period_end="",
        price_type="all",
    )

    assert results[0]["lookup_price"] is None
    assert results[0]["lookup_source"] == "未查到"
    assert results[0]["lookup_label"] is None
    assert results[0]["lookup_url"] is None


def test_do_lookup_builds_summary_label_for_db_results(monkeypatch):
    class _FakeDB:
        def search_price_by_name(self, name: str, **kwargs):
            assert name == "镀锌钢管"
            assert kwargs["province"] == "广东"
            assert kwargs["spec"] == "DN32"
            assert kwargs["target_unit"] == "m"
            assert kwargs["object_type"] == "pipe"
            return {
                "price": 18.5,
                "unit": "m",
                "source": "广东信息价",
            }

    monkeypatch.setattr(material_price_api, "_material_db_ready", lambda: True)
    monkeypatch.setattr(material_price_api, "_get_db", lambda: _FakeDB())
    monkeypatch.setattr(material_price_api, "_load_material_price_cache", lambda: {})

    results = material_price_api._do_lookup(
        [{"name": "镀锌钢管", "spec": "DN32", "unit": "m"}],
        province="广东",
        city="",
        period_end="",
        price_type="all",
    )

    assert results[0]["lookup_price"] == 18.5
    assert results[0]["lookup_source"] == "广东信息价"
    assert results[0]["lookup_label"] == "广东信息价 | 镀锌钢管 DN32 | m | 18.50"
    assert results[0]["lookup_url"] is None


def test_do_lookup_passes_city_and_period_to_db(monkeypatch):
    class _FakeDB:
        def search_price_by_name(self, name: str, **kwargs):
            assert name == "焊接钢管"
            assert kwargs["province"] == "江西"
            assert kwargs["city"] == "九江"
            assert kwargs["period_end"] == "2025-08-31"
            assert kwargs["spec"] == "DN80"
            assert kwargs["target_unit"] == "m"
            return {
                "price": 28.04,
                "unit": "m",
                "source": "江西九江信息价(2025-08-31)",
            }

    monkeypatch.setattr(material_price_api, "_material_db_ready", lambda: True)
    monkeypatch.setattr(material_price_api, "_get_db", lambda: _FakeDB())
    monkeypatch.setattr(material_price_api, "_load_material_price_cache", lambda: {})

    results = material_price_api._do_lookup(
        [{"name": "焊接钢管", "spec": "DN80", "unit": "m"}],
        province="江西",
        city="九江",
        period_end="2025-08-31",
        price_type="info",
    )

    assert results[0]["lookup_price"] == 28.04
    assert results[0]["lookup_source"] == "江西九江信息价(2025-08-31)"


def test_gldjc_cookie_verify_forwards_to_local_match_service(monkeypatch):
    class _FakeResponse:
        status_code = 200

        def json(self):
            return {
                "ok": True,
                "status": "valid",
                "message": "Cookie有效",
                "keyword": "焊接钢管 DN80",
            }

    class _FakeClient:
        async def __aenter__(self):
            return self

        async def __aexit__(self, exc_type, exc, tb):
            return False

        async def post(self, url, headers=None, json=None):
            assert url.endswith("/material-price/gldjc-cookie-verify")
            assert json["cookie"] == "token=bearer test"
            assert json["province"] == "江西"
            assert json["city"] == "九江"
            return _FakeResponse()

    monkeypatch.setattr(material_price_api, "_is_remote", lambda: True)
    monkeypatch.setattr(material_price_api, "LOCAL_MATCH_URL", "http://match-service:9300")
    monkeypatch.setattr(material_price_api, "LOCAL_MATCH_API_KEY", "test-key")
    monkeypatch.setattr(material_price_api, "local_match_async_client", lambda timeout=60.0: _FakeClient())

    result = asyncio.run(
        material_price_api.gldjc_cookie_verify({
            "cookie": "token=bearer test",
            "province": "江西",
            "city": "九江",
        })
    )

    assert result["ok"] is True
    assert result["status"] == "valid"
