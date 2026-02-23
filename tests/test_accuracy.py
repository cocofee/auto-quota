"""
准确率测试：用带标准答案的Excel验证匹配精度。

注意：本文件是手工评测脚本，不是 pytest 单元测试。
已改为 main() 入口，避免 pytest 收集阶段执行耗时逻辑。
"""
from __future__ import annotations

import argparse
from pathlib import Path

import openpyxl

from src.text_parser import parser as text_parser
from src.hybrid_searcher import HybridSearcher
from src.param_validator import ParamValidator


def _is_bill_serial(value) -> bool:
    """识别清单序号，兼容 1 / 1.0 / "2.0" / "03"。"""
    if value is None:
        return False
    if isinstance(value, int):
        return value >= 0
    if isinstance(value, float):
        return value.is_integer() and value >= 0
    text = str(value).strip()
    if not text:
        return False
    if text.isdigit():
        return True
    if text.endswith(".0"):
        body = text[:-2].strip()
        return body.isdigit()
    return False


def load_test_cases(excel_path: str) -> list[tuple[dict, list[str]]]:
    """读取Excel，提取 (清单项, 正确定额列表)。"""
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    test_cases: list[tuple[dict, list[str]]] = []
    current_bill = None
    current_quotas: list[str] = []

    for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row_idx <= 2:
            continue

        a = row[0] if len(row) > 0 else None
        b = row[1] if len(row) > 1 else None
        c = row[2] if len(row) > 2 else None
        d = row[3] if len(row) > 3 else None
        e = row[4] if len(row) > 4 else None
        f = row[5] if len(row) > 5 else None

        if _is_bill_serial(a):
            if current_bill:
                test_cases.append((current_bill, current_quotas))

            current_bill = {
                "code": str(b).strip() if b else "",
                "name": str(c).strip() if c else "",
                "description": str(d).strip() if d else "",
                "unit": str(e).strip() if e else "",
                "quantity": f,
            }
            current_quotas = []
        elif b and str(b).strip().startswith("C"):
            quota_id = str(b).strip().split()[0].rstrip("换").strip()
            current_quotas.append(quota_id)

    if current_bill:
        test_cases.append((current_bill, current_quotas))

    wb.close()
    return test_cases


def evaluate_cases(test_cases: list[tuple[dict, list[str]]]) -> tuple[str, dict]:
    """执行匹配并返回文本报告与统计。"""
    searcher = HybridSearcher()
    validator = ParamValidator()

    lines = []
    lines.append("=" * 80)
    lines.append(f"准确率测试：{len(test_cases)} 条清单")
    lines.append("=" * 80)

    exact_match = 0
    partial_match = 0
    no_match = 0
    no_result = 0

    for i, (bill, correct_quotas) in enumerate(test_cases, 1):
        name = bill["name"]
        desc = bill["description"]

        search_query = text_parser.build_quota_query(name, desc)
        full_query = f"{name} {desc}".strip()

        candidates = searcher.search(search_query, top_k=10)
        if not candidates:
            no_result += 1
            lines.append(f"\n第{i:2d}条 [{name}] query=[{search_query}]")
            lines.append(f"  正确: {correct_quotas}")
            lines.append("  结果: 搜索无结果 ✗")
            continue

        validated = validator.validate_candidates(full_query, candidates)
        matched = [c for c in validated if c.get("param_match", True)]
        top = matched[0] if matched else (validated[0] if validated else None)

        if not top:
            no_result += 1
            lines.append(f"\n第{i:2d}条 [{name}] query=[{search_query}]")
            lines.append(f"  正确: {correct_quotas}")
            lines.append("  结果: 无匹配候选 ✗")
            continue

        system_quota_id = top.get("quota_id", "")
        correct_main = correct_quotas[0] if correct_quotas else ""

        if system_quota_id == correct_main:
            exact_match += 1
            mark = "✓"
        elif (system_quota_id and correct_main and
              system_quota_id.rsplit("-", 1)[0] == correct_main.rsplit("-", 1)[0]):
            partial_match += 1
            mark = "≈"
        else:
            no_match += 1
            mark = "✗"

        lines.append(f"\n第{i:2d}条 [{name}] query=[{search_query}]")
        lines.append(f"  正确: {correct_main:15s} | 系统: {system_quota_id:15s} {mark}")
        if mark != "✓":
            lines.append(f"  系统定额名: {top.get('name', '')[:50]}")
            lines.append(
                "  参数: "
                f"match={top.get('param_match')}, "
                f"score={top.get('param_score', 0):.2f}, "
                f"{top.get('param_detail', '')[:60]}"
            )

    total = len(test_cases)
    pct = lambda n: (n * 100 // total) if total else 0
    lines.append(f"\n{'=' * 80}")
    lines.append(f"测试结果统计（{total}条清单）:")
    lines.append(f"  精确匹配 ✓: {exact_match:3d} ({pct(exact_match)}%)")
    lines.append(f"  近似匹配 ≈: {partial_match:3d} ({pct(partial_match)}%)")
    lines.append(f"  不匹配   ✗: {no_match:3d} ({pct(no_match)}%)")
    lines.append(f"  无结果     : {no_result:3d} ({pct(no_result)}%)")
    lines.append(f"  准确率(精确+近似): {pct(exact_match + partial_match)}%")
    lines.append("=" * 80)

    summary = {
        "total": total,
        "exact": exact_match,
        "partial": partial_match,
        "no_match": no_match,
        "no_result": no_result,
        "accuracy": pct(exact_match + partial_match),
    }
    return "\n".join(lines), summary


def main() -> int:
    parser = argparse.ArgumentParser(description="准确率测试脚本（手工评测）")
    parser.add_argument(
        "--input",
        required=True,
        help="带标准答案的Excel路径",
    )
    parser.add_argument(
        "--output",
        default="accuracy_test_result.txt",
        help="输出报告路径",
    )
    args = parser.parse_args()

    input_path = Path(args.input)
    if not input_path.exists():
        print(f"输入文件不存在: {input_path}")
        return 1

    test_cases = load_test_cases(str(input_path))
    print(f"共提取 {len(test_cases)} 条测试用例")

    output_text, summary = evaluate_cases(test_cases)
    Path(args.output).write_text(output_text, encoding="utf-8")

    print(f"\n测试完成！结果写入 {args.output}")
    print(f"精确匹配: {summary['exact']}/{summary['total']} ({(summary['exact'] * 100 // max(summary['total'], 1))}%)")
    print(f"近似匹配: {summary['partial']}/{summary['total']}")
    print(f"准确率(精确+近似): {summary['accuracy']}%")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
