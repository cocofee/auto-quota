"""查看输入文件的所有Sheet名称和结构"""
import openpyxl

# 检查小栗AI文件的Sheet结构
file_path = r"C:\Users\Administrator\Desktop\6#配套楼-小栗AI自-北京小栗AI自动加定额202602071708.xlsx"
wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

lines = []
lines.append(f"文件: {file_path}")
lines.append(f"Sheet数量: {len(wb.sheetnames)}")
lines.append("")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    lines.append(f"=== Sheet: [{sheet_name}] ===")
    lines.append(f"  行数: {ws.max_row}, 列数: {ws.max_column}")

    # 读取前5行看看内容
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row_idx > 5:
            break
        vals = []
        for col_idx, val in enumerate(row):
            if val is not None:
                v = str(val).replace("\n", "\\n")[:30]
                vals.append(f"{chr(65+col_idx)}=[{v}]")
        if vals:
            lines.append(f"  行{row_idx}: {', '.join(vals[:6])}")
        else:
            lines.append(f"  行{row_idx}: (空行)")
    lines.append("")

wb.close()

output = "\n".join(lines)
with open("tests/debug_sheets.txt", "w", encoding="utf-8") as f:
    f.write(output)
print("Done. Output: tests/debug_sheets.txt")
