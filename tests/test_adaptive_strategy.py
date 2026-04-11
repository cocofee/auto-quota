from types import SimpleNamespace

from openpyxl import Workbook

import main as main_mod
from src.adaptive_strategy import AdaptiveStrategy, summarize_adaptive_strategy_metrics
from src.match_engine import _annotate_adaptive_strategies
from src.output_writer import OutputWriter


def test_select_fast_for_simple_item_with_high_hit_rate():
    strategy = AdaptiveStrategy(historical_hit_rates={"C10": 0.92})

    item = {
        "name": "steel_pipe",
        "specialty": "C10",
        "params": {
            "dn": "100",
            "material": "steel",
        },
    }

    assert strategy.select_strategy(item) == "fast"


def test_select_deep_for_complex_item():
    strategy = AdaptiveStrategy(historical_hit_rates={"C10": 0.9})

    item = {
        "name": (
            "complex fire pipe item (with supports accessories height notes) "
            "interface requirements insulation anticorrosion and commissioning"
        ),
        "specialty": "C10",
        "params": {
            "dn": "100",
            "material": "steel",
            "connection": "grooved",
            "pressure": "1.6MPa",
            "coating": "hot_dip",
            "usage": "fire",
        },
    }

    assert strategy.select_strategy(item) == "deep"
    assert strategy._compute_complexity(item) == 0.8


def test_select_standard_when_hit_rate_is_low_but_item_is_simple():
    strategy = AdaptiveStrategy(historical_hit_rates={"C4": 0.35})

    item = {
        "name": "distribution_box",
        "specialty": "C4",
        "params": {
            "model": "AL1",
            "voltage": "380V",
            "ampere": "63A",
        },
    }

    assert strategy.select_strategy(item) == "standard"


def test_select_deep_when_hit_rate_is_low_and_params_are_sparse():
    strategy = AdaptiveStrategy(historical_hit_rates={"C4": 0.12})

    item = {
        "name": "distribution_box",
        "specialty": "C4",
        "params": {},
    }

    assert strategy.select_strategy(item) == "deep"


def test_select_standard_for_midrange_case():
    strategy = AdaptiveStrategy(historical_hit_rates={"C9": 0.7})

    item = {
        "name": "rubber_joint",
        "specialty": "C9",
        "params": {
            "dn": "150",
        },
    }

    decision = strategy.evaluate(item)
    assert decision["strategy"] == "standard"
    assert decision["complexity"] == 0.0
    assert decision["param_completeness"] == 1 / 3


def test_supports_object_style_items_and_tracker_fallback():
    class FakeTracker:
        def get_knowledge_hit_report(self, days: int = 7) -> dict:
            assert days == 14
            return {
                "layer_metrics": [
                    {"layer": "ExperienceDB", "hit_rate": 84.0},
                ],
            }

    strategy = AdaptiveStrategy(accuracy_tracker=FakeTracker(), tracker_days=14)
    item = SimpleNamespace(
        name="fan_coil",
        specialty="C7",
        params={"power": "2.2kW", "air_volume": "1200m3/h"},
    )

    assert strategy.select_strategy(item) == "fast"


def test_summarize_adaptive_strategy_metrics():
    summary = summarize_adaptive_strategy_metrics([
        {
            "bill_item": {"adaptive_strategy": "fast"},
            "quotas": [{"quota_id": "Q1"}],
            "performance": {"total": 0.6},
        },
        {
            "bill_item": {"adaptive_strategy": "deep"},
            "quotas": [],
            "performance": {"total": 1.8},
        },
        {
            "bill_item": {"adaptive_strategy": "deep"},
            "quotas": [{"quota_id": "Q2"}],
            "performance": {"total": 1.2},
        },
        {
            "bill_item": {},
            "quotas": [],
        },
    ])

    assert summary["total"] == 4
    assert summary["with_strategy_count"] == 3
    assert summary["missing_strategy_count"] == 1
    assert summary["overall_avg_time_sec"] == 1.2
    assert summary["distribution"]["fast"]["count"] == 1
    assert summary["distribution"]["fast"]["avg_time_sec"] == 0.6
    assert summary["distribution"]["deep"]["count"] == 2
    assert summary["distribution"]["deep"]["matched_rate"] == 50.0
    assert summary["distribution"]["unknown"]["count"] == 1


def test_build_run_stats_includes_adaptive_strategy_summary():
    results = [
        {
            "bill_item": {"adaptive_strategy": "fast"},
            "quotas": [{"quota_id": "Q1"}],
            "confidence": 95,
            "match_source": "experience_exact",
            "performance": {"total": 0.4},
        },
        {
            "bill_item": {"adaptive_strategy": "standard", "_review_rejected": True},
            "quotas": [],
            "confidence": 0,
            "match_source": "search",
            "performance": {"total": 1.0},
        },
    ]

    stats = main_mod._build_run_stats(results, elapsed=3.5)

    assert stats["total"] == 2
    assert stats["exp_hits"] == 1
    assert stats["review_rejected"] == 1
    assert stats["adaptive_strategy"]["distribution"]["fast"]["count"] == 1
    assert stats["adaptive_strategy"]["distribution"]["standard"]["avg_time_sec"] == 1.0


def test_output_writer_stats_sheet_includes_adaptive_strategy_rows():
    wb = Workbook()
    ws = wb.active
    writer = OutputWriter()

    writer._write_stats_sheet(ws, [
        {
            "bill_item": {"adaptive_strategy": "fast"},
            "quotas": [{"quota_id": "Q1"}],
            "confidence": 95,
            "performance": {"total": 0.5},
        },
        {
            "bill_item": {"adaptive_strategy": "deep"},
            "quotas": [],
            "confidence": 0,
            "performance": {"total": 1.7},
        },
    ])

    values = [ws.cell(row=i, column=1).value for i in range(1, ws.max_row + 1)]
    assert "自适应策略" in values
    assert "自适应策略-fast" in values
    assert "自适应策略-deep" in values


def test_annotate_adaptive_strategies_assigns_item_strategy():
    items = [
        {
            "name": "steel_pipe",
            "specialty": "C10",
            "params": {
                "dn": "100",
                "material": "steel",
            },
        },
            {
                "name": (
                "complex fire pipe item (with supports accessories height notes) "
                "interface requirements insulation anticorrosion and commissioning"
            ),
            "specialty": "C10",
            "params": {
                "dn": "100",
                "material": "steel",
                "connection": "grooved",
                "pressure": "1.6MPa",
                "coating": "hot_dip",
                "usage": "fire",
            },
        },
    ]

    counts = _annotate_adaptive_strategies(
        items,
        selector=AdaptiveStrategy(historical_hit_rates={"C10": 0.92}),
    )

    assert items[0]["adaptive_strategy"] == "fast"
    assert items[1]["adaptive_strategy"] == "deep"
    assert counts["fast"] == 1
    assert counts["deep"] == 1
