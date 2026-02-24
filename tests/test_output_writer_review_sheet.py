import openpyxl

from src.output_writer import OutputWriter


def test_review_sheet_includes_high_confidence_no_match_item():
    wb = openpyxl.Workbook()
    ws = wb.active
    writer = OutputWriter()

    results = [
        {
            "bill_item": {"name": "DN25钢管", "description": "无匹配场景"},
            "confidence": 95,
            "quotas": [],
            "match_source": "agent",
            "explanation": "无候选",
            "alternatives": [],
        }
    ]

    writer._write_review_sheet(ws, results)

    assert ws.cell(row=2, column=1).value == 1
    assert ws.cell(row=2, column=2).value == "DN25钢管"
    assert ws.cell(row=2, column=6).value == "—"


def test_review_sheet_includes_high_confidence_agent_fallback_item():
    wb = openpyxl.Workbook()
    ws = wb.active
    writer = OutputWriter()

    results = [
        {
            "bill_item": {"name": "DN32钢管", "description": "降级来源场景"},
            "confidence": 92,
            "quotas": [{"quota_id": "C10-1-1", "name": "钢管安装"}],
            "match_source": "agent_fallback",
            "explanation": "降级结果",
            "alternatives": [],
        }
    ]

    writer._write_review_sheet(ws, results)

    assert ws.cell(row=2, column=1).value == 1
    assert ws.cell(row=2, column=4).value == "C10-1-1"
    assert ws.cell(row=2, column=6).value.startswith("★★★")
