import openpyxl

from src.output_writer import OutputWriter


def test_detect_bill_layout_for_shifted_unit_qty_columns():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["序号", "项目名称", "项目特征描述", "单位", "工程量", "不含税单价"])

    writer = OutputWriter()
    layout = writer._detect_bill_layout(ws, header_row=1)

    assert layout["unit_col"] == 4
    assert layout["qty_col"] == 5


def test_process_bill_sheet_writes_quota_unit_qty_to_detected_columns():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "通风空调工程量清单"

    ws.append(["序号", "项目名称", "项目特征描述", "单位", "工程量", "不含税单价", "不含税合价"])
    ws.append([1, "70℃防火阀 FD", "1、名称:70℃防火阀", "个", 7, None, None])

    results = [{
        "bill_item": {
            "sheet_name": ws.title,
            "sheet_bill_seq": 1,
            "name": "70℃防火阀 FD",
            "unit": "个",
            "quantity": 7,
        },
        "quotas": [{
            "quota_id": "C7-7-1",
            "name": "防火阀安装",
            "unit": "个",
        }],
        "confidence": 90,
        "explanation": "匹配成功",
        "alternatives": [],
    }]

    writer = OutputWriter()
    writer._process_bill_sheet(ws, results)

    quota_row = 3
    assert ws.cell(row=quota_row, column=2).value == "C7-7-1"
    assert ws.cell(row=quota_row, column=3).value == "防火阀安装"
    assert ws.cell(row=quota_row, column=4).value == "个"
    assert ws.cell(row=quota_row, column=5).value in (7, 7.0)
