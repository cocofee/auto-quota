"""Benchmark 回归测试。

验证：
  1. benchmark 配置文件存在且格式正确
  2. 基线文件存在且包含有效指标
  3. JSON试卷目录有足够试卷
  4. 指标字段结构完整（不做实际运行，实际运行用 run_benchmark.py）
"""
from __future__ import annotations

import json
from pathlib import Path

import pytest

# 文件路径
CONFIG_PATH = Path("tests/benchmark_config.json")
BASELINE_PATH = Path("tests/benchmark_baseline.json")
PAPERS_DIR = Path("tests/benchmark_papers")

# JSON试卷必须包含的字段
REQUIRED_JSON_METRICS = ["total", "correct", "hit_rate"]

# Excel数据集必须包含的字段
REQUIRED_EXCEL_METRICS = [
    "total", "green_rate", "yellow_rate", "red_rate",
    "exp_hit_rate", "avg_time_sec",
]


class TestBenchmarkConfig:
    """benchmark 配置文件格式验证"""

    def test_config_exists(self):
        """配置文件存在"""
        assert CONFIG_PATH.exists(), f"配置文件不存在: {CONFIG_PATH}"

    def test_config_valid_json(self):
        """配置文件是有效JSON"""
        config = json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
        assert "datasets" in config, "缺少 datasets 字段"

    def test_config_has_excel_datasets(self):
        """配置文件至少定义1个Excel数据集"""
        config = json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
        datasets = config["datasets"]
        assert len(datasets) >= 1, f"Excel数据集不足1组，当前{len(datasets)}组"

    def test_json_papers_sufficient(self):
        """JSON试卷目录至少有10个省份试卷"""
        papers = list(PAPERS_DIR.glob("*.json"))
        # 排除以_开头的元数据文件
        papers = [p for p in papers if not p.name.startswith("_")]
        assert len(papers) >= 10, f"JSON试卷不足10份，当前{len(papers)}份"

    def test_config_dataset_fields(self):
        """每个Excel数据集都有必要字段"""
        config = json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
        for name, ds in config["datasets"].items():
            assert "path" in ds, f"{name} 缺少 path 字段"
            assert "expected_items_range" in ds, f"{name} 缺少 expected_items_range 字段"
            assert "province" in ds, f"{name} 缺少 province 字段"
            rng = ds["expected_items_range"]
            assert len(rng) == 2, f"{name} 的 expected_items_range 应为 [min, max]"
            assert rng[0] <= rng[1], f"{name} 的 expected_items_range 范围无效"

    def test_config_has_tolerance(self):
        """配置文件定义了回归容差"""
        config = json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
        tol = config.get("regression_tolerance", {})
        assert "green_rate" in tol, "缺少 green_rate 容差"
        assert "red_rate" in tol, "缺少 red_rate 容差"


class TestBenchmarkBaseline:
    """基线文件格式验证"""

    def test_baseline_exists(self):
        """基线文件存在"""
        assert BASELINE_PATH.exists(), (
            f"基线文件不存在: {BASELINE_PATH}\n"
            "请先运行: python tools/run_benchmark.py --save"
        )

    def test_baseline_valid_json(self):
        """基线文件是有效JSON，包含新版字段"""
        if not BASELINE_PATH.exists():
            pytest.skip("基线文件不存在")
        baseline = json.loads(BASELINE_PATH.read_text(encoding="utf-8"))
        assert "version" in baseline, "缺少 version 字段"
        assert "date" in baseline, "缺少 date 字段"
        assert "json_papers" in baseline, "缺少 json_papers 字段"
        assert "excel_datasets" in baseline, "缺少 excel_datasets 字段"

    def test_baseline_has_datasets(self):
        """基线包含JSON试卷和Excel数据集的指标"""
        if not BASELINE_PATH.exists():
            pytest.skip("基线文件不存在")
        baseline = json.loads(BASELINE_PATH.read_text(encoding="utf-8"))
        json_papers = baseline.get("json_papers", {})
        excel_datasets = baseline.get("excel_datasets", {})
        assert len(json_papers) >= 1, "基线中没有JSON试卷指标"
        assert len(excel_datasets) >= 1, "基线中没有Excel数据集指标"

    def test_baseline_json_papers_complete(self):
        """基线中每个JSON试卷包含必要指标"""
        if not BASELINE_PATH.exists():
            pytest.skip("基线文件不存在")
        baseline = json.loads(BASELINE_PATH.read_text(encoding="utf-8"))
        for name, metrics in baseline.get("json_papers", {}).items():
            for field in REQUIRED_JSON_METRICS:
                assert field in metrics, (
                    f"JSON试卷 {name} 缺少指标字段: {field}"
                )

    def test_baseline_excel_metrics_complete(self):
        """基线中每个Excel数据集包含完整指标"""
        if not BASELINE_PATH.exists():
            pytest.skip("基线文件不存在")
        baseline = json.loads(BASELINE_PATH.read_text(encoding="utf-8"))
        for name, metrics in baseline.get("excel_datasets", {}).items():
            for field in REQUIRED_EXCEL_METRICS:
                assert field in metrics, (
                    f"Excel数据集 {name} 缺少指标字段: {field}"
                )

    def test_baseline_rates_valid(self):
        """基线中Excel数据集的比率值在 [0, 1] 范围内"""
        if not BASELINE_PATH.exists():
            pytest.skip("基线文件不存在")
        baseline = json.loads(BASELINE_PATH.read_text(encoding="utf-8"))
        rate_fields = ["green_rate", "yellow_rate", "red_rate",
                       "exp_hit_rate"]
        for name, metrics in baseline.get("excel_datasets", {}).items():
            for field in rate_fields:
                val = metrics.get(field, 0)
                assert 0 <= val <= 1, (
                    f"数据集 {name} 的 {field}={val} 超出 [0,1] 范围"
                )

    def test_baseline_total_positive(self):
        """基线中的总条数 > 0"""
        if not BASELINE_PATH.exists():
            pytest.skip("基线文件不存在")
        baseline = json.loads(BASELINE_PATH.read_text(encoding="utf-8"))
        for name, metrics in baseline.get("json_papers", {}).items():
            assert metrics.get("total", 0) > 0, (
                f"JSON试卷 {name} 的 total={metrics.get('total')} 应该 > 0"
            )


class TestDirtyDataFixture:
    """B4脏数据样本文件验证"""

    def test_dirty_data_exists(self):
        """脏数据样本Excel存在"""
        path = Path("tests/fixtures/dirty_data_sample.xlsx")
        assert path.exists(), (
            f"脏数据样本不存在: {path}\n"
            "请运行: python tests/fixtures/gen_dirty_data.py"
        )

    def test_dirty_data_readable(self):
        """脏数据样本可以被 openpyxl 正常读取"""
        path = Path("tests/fixtures/dirty_data_sample.xlsx")
        if not path.exists():
            pytest.skip("脏数据样本不存在")

        import openpyxl
        wb = openpyxl.load_workbook(str(path), read_only=True)
        ws = wb[wb.sheetnames[0]]

        row_count = sum(1 for _ in ws.iter_rows())
        assert row_count >= 17, f"脏数据样本行数不足: {row_count}行（预期至少17行）"
        wb.close()
