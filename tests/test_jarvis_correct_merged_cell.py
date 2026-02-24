from __future__ import annotations

import shutil
from pathlib import Path
from uuid import uuid4

import openpyxl

from tools.jarvis_correct import correct_excel


def _new_tmp_dir() -> Path:
    root = Path("output") / "test_tmp"
    root.mkdir(parents=True, exist_ok=True)
    d = root / f"jarvis-correct-{uuid4().hex}"
    d.mkdir(parents=True, exist_ok=False)
    return d


def _build_sheet(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active

    # Row 2: bill row, Row 3: quota row
    ws.cell(row=2, column=1, value=1)
    ws.cell(row=2, column=2, value="OLD-ID")
    ws.cell(row=2, column=3, value="OLD-NAME")
    ws.cell(row=2, column=10, value="OLD-MARK")
    ws.cell(row=2, column=11, value="OLD-REASON")

    wb.save(path)


def test_correct_excel_skips_merged_slave_cells():
    tmp_dir = _new_tmp_dir()
    try:
        input_path = tmp_dir / "input.xlsx"
        output_path = tmp_dir / "output.xlsx"
        _build_sheet(input_path)

        wb = openpyxl.load_workbook(input_path)
        ws = wb.active
        # Target cells for seq=1 are row3 col2/3 and row2 col10/11.
        # Make all those targets merged slave cells.
        ws.merge_cells("B2:B3")
        ws.merge_cells("C2:C3")
        ws.merge_cells("J1:J2")
        ws.merge_cells("K1:K2")
        wb.save(input_path)

        result = correct_excel(
            str(input_path),
            [{"seq": 1, "quota_id": "NEW-ID", "quota_name": "NEW-NAME"}],
            str(output_path),
        )
        assert Path(result).exists()

        out_wb = openpyxl.load_workbook(output_path)
        out_ws = out_wb.active
        # Master cells keep original value because merged slave targets are skipped.
        assert out_ws.cell(row=2, column=2).value == "OLD-ID"
        assert out_ws.cell(row=2, column=3).value == "OLD-NAME"
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)


def test_correct_excel_updates_normal_cells():
    tmp_dir = _new_tmp_dir()
    try:
        input_path = tmp_dir / "input_normal.xlsx"
        output_path = tmp_dir / "output_normal.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=2, column=1, value=1)
        ws.cell(row=3, column=2, value="OLD-ID")
        ws.cell(row=3, column=3, value="OLD-NAME")
        ws.cell(row=2, column=10, value="OLD-MARK")
        ws.cell(row=2, column=11, value="OLD-REASON")
        wb.save(input_path)

        correct_excel(
            str(input_path),
            [{"seq": 1, "quota_id": "NEW-ID", "quota_name": "NEW-NAME"}],
            str(output_path),
        )

        out_wb = openpyxl.load_workbook(output_path)
        out_ws = out_wb.active
        assert out_ws.cell(row=3, column=2).value == "NEW-ID"
        assert out_ws.cell(row=3, column=3).value == "NEW-NAME"
        assert out_ws.cell(row=2, column=10).value == "★★★已审核"
        assert "OLD-ID" in out_ws.cell(row=2, column=11).value
        assert "NEW-ID" in out_ws.cell(row=2, column=11).value
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)


def test_correct_excel_locates_by_sheet_and_sheet_bill_seq():
    tmp_dir = _new_tmp_dir()
    try:
        input_path = tmp_dir / "multi_sheet.xlsx"
        output_path = tmp_dir / "multi_sheet_out.xlsx"

        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "S1"
        ws2 = wb.create_sheet("S2")

        # Both sheets have local serial "1", but should be corrected independently.
        ws1.cell(row=2, column=1, value=1)
        ws1.cell(row=3, column=2, value="OLD-S1-ID")
        ws1.cell(row=3, column=3, value="OLD-S1-NAME")
        ws1.cell(row=2, column=10, value="OLD-S1-MARK")
        ws1.cell(row=2, column=11, value="OLD-S1-REASON")

        ws2.cell(row=2, column=1, value=1)
        ws2.cell(row=3, column=2, value="OLD-S2-ID")
        ws2.cell(row=3, column=3, value="OLD-S2-NAME")
        ws2.cell(row=2, column=10, value="OLD-S2-MARK")
        ws2.cell(row=2, column=11, value="OLD-S2-REASON")

        wb.save(input_path)

        corrections = [
            {
                "seq": 1,
                "quota_id": "NEW-S1-ID",
                "quota_name": "NEW-S1-NAME",
                "sheet_name": "S1",
                "sheet_bill_seq": 1,
            },
            {
                "seq": 2,
                "quota_id": "NEW-S2-ID",
                "quota_name": "NEW-S2-NAME",
                "sheet_name": "S2",
                "sheet_bill_seq": 1,
            },
        ]
        correct_excel(str(input_path), corrections, str(output_path))

        out_wb = openpyxl.load_workbook(output_path)
        out_s1 = out_wb["S1"]
        out_s2 = out_wb["S2"]
        assert out_s1.cell(row=3, column=2).value == "NEW-S1-ID"
        assert out_s2.cell(row=3, column=2).value == "NEW-S2-ID"
        assert out_s1.cell(row=2, column=10).value == "★★★已审核"
        assert out_s2.cell(row=2, column=10).value == "★★★已审核"
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)


def test_correct_excel_sheet_locator_works_when_active_sheet_has_no_serial():
    tmp_dir = _new_tmp_dir()
    try:
        input_path = tmp_dir / "cover_first.xlsx"
        output_path = tmp_dir / "cover_first_out.xlsx"

        wb = openpyxl.Workbook()
        ws_cover = wb.active
        ws_cover.title = "Cover"
        ws_cover.cell(row=1, column=1, value="封面")

        ws_data = wb.create_sheet("Data")
        ws_data.cell(row=2, column=1, value=1)
        ws_data.cell(row=3, column=2, value="OLD-ID")
        ws_data.cell(row=3, column=3, value="OLD-NAME")
        ws_data.cell(row=2, column=10, value="OLD-MARK")
        ws_data.cell(row=2, column=11, value="OLD-REASON")
        wb.save(input_path)

        corrections = [{
            "seq": 1,
            "quota_id": "NEW-ID",
            "quota_name": "NEW-NAME",
            "sheet_name": "Data",
            "sheet_bill_seq": 1,
        }]
        correct_excel(str(input_path), corrections, str(output_path))

        out_wb = openpyxl.load_workbook(output_path)
        out_ws = out_wb["Data"]
        assert out_ws.cell(row=3, column=2).value == "NEW-ID"
        assert out_ws.cell(row=3, column=3).value == "NEW-NAME"
        assert out_ws.cell(row=2, column=10).value == "★★★已审核"
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)


def test_correct_excel_legacy_mode_accepts_text_serial_float():
    tmp_dir = _new_tmp_dir()
    try:
        input_path = tmp_dir / "legacy_text_serial.xlsx"
        output_path = tmp_dir / "legacy_text_serial_out.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=2, column=1, value="1.0")
        ws.cell(row=3, column=2, value="OLD-ID")
        ws.cell(row=3, column=3, value="OLD-NAME")
        ws.cell(row=2, column=10, value="OLD-MARK")
        ws.cell(row=2, column=11, value="OLD-REASON")
        wb.save(input_path)

        correct_excel(
            str(input_path),
            [{"seq": 1, "quota_id": "NEW-ID", "quota_name": "NEW-NAME"}],
            str(output_path),
        )

        out_wb = openpyxl.load_workbook(output_path)
        out_ws = out_wb.active
        assert out_ws.cell(row=3, column=2).value == "NEW-ID"
        assert out_ws.cell(row=3, column=3).value == "NEW-NAME"
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)
