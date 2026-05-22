"""Regression tests for merged-cell-safe Excel output writing."""

from __future__ import annotations

import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import PatternFill

from src.output_writer import OutputWriter, _safe_write_cell


class TestSafeWriteCell:
    def test_normal_cell_write(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        cell = _safe_write_cell(ws, 1, 1, "test")

        assert cell is not None
        assert cell.value == "test"

    def test_normal_cell_no_value(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="original")

        cell = _safe_write_cell(ws, 1, 1)

        assert cell is not None
        assert cell.value == "original"

    def test_merged_cell_returns_none(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.merge_cells("A1:A3")

        cell_a2 = ws.cell(row=2, column=1)
        assert isinstance(cell_a2, MergedCell)

        result = _safe_write_cell(ws, 2, 1, "should-skip")

        assert result is None

    def test_merged_cell_master_still_writable(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.merge_cells("A1:A3")

        cell = _safe_write_cell(ws, 1, 1, "master")

        assert cell is not None
        assert cell.value == "master"


class TestWriteBillExtraInfoWithMergedCells:
    def _create_ws_with_merged_extra_cols(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        headers = [
            "序号", "编码", "名称", "特征", "单位", "工程量",
            "", "", "", "推荐度", "说明", "备选1", "备选2", "备选3", "主材",
        ]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        ws.cell(row=2, column=1, value=1)
        ws.cell(row=2, column=3, value="DN25镀锌钢管")

        ws.cell(row=3, column=1, value=2)
        ws.cell(row=3, column=3, value="DN32镀锌钢管")
        ws.merge_cells(start_row=3, start_column=10, end_row=4, end_column=10)
        ws.merge_cells(start_row=3, start_column=12, end_row=3, end_column=14)
        return wb, ws

    def test_write_extra_info_no_merge_normal(self):
        _, ws = self._create_ws_with_merged_extra_cols()
        writer = OutputWriter.__new__(OutputWriter)
        result = {
            "confidence": 85,
            "quotas": [{"quota_id": "C10-1-1", "name": "管道安装"}],
            "explanation": "匹配成功",
            "alternatives": [
                {"quota_id": "C10-1-2", "name": "备选A"},
                {"quota_id": "C10-1-3", "name": "备选B"},
            ],
            "materials": [{"name": "镀锌钢管", "spec": "DN25"}],
        }

        writer._write_bill_extra_info(ws, 2, result)

        assert ws.cell(row=2, column=10).value is not None

    def test_write_extra_info_with_merged_cells_no_crash(self):
        _, ws = self._create_ws_with_merged_extra_cols()
        writer = OutputWriter.__new__(OutputWriter)
        result = {
            "confidence": 90,
            "quotas": [{"quota_id": "C10-2-1", "name": "管道安装"}],
            "explanation": "匹配成功",
            "alternatives": [
                {"quota_id": "C10-2-2", "name": "备选A"},
                {"quota_id": "C10-2-3", "name": "备选B"},
                {"quota_id": "C10-2-4", "name": "备选C"},
            ],
            "materials": [],
        }

        writer._write_bill_extra_info(ws, 4, result)


class TestWriteAlternativeCellsWithMergedCells:
    def test_alternatives_on_merged_cols_no_crash(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.merge_cells(start_row=2, start_column=12, end_row=2, end_column=14)

        alternatives = [
            {"quota_id": "C4-1-1", "name": "定额A"},
            {"quota_id": "C4-1-2", "name": "定额B"},
        ]

        OutputWriter._write_alternative_cells(ws, 2, start_col=12, alternatives=alternatives)

        assert ws.cell(row=2, column=12).value is not None


class TestWriteNoMatchRowWithMergedCells:
    def test_no_match_row_with_merged_cells_no_crash(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.merge_cells(start_row=2, start_column=3, end_row=3, end_column=3)

        OutputWriter._write_no_match_row(ws, 3, "无匹配结果", max_col=5)


class TestApplyRowStyleWithMergedCells:
    def test_apply_style_with_merged_cells_no_crash(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.merge_cells("B2:D2")

        OutputWriter._apply_row_style(ws, 2, start_col=1, end_col=5, wrap_cols={3, 4})


class TestSetHeaderCellWithMergedCells:
    def test_set_header_on_normal_cell(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        fill = PatternFill(start_color="4472C4", fill_type="solid")

        cell = OutputWriter._set_header_cell(ws, 1, 1, "表头", fill)

        assert cell is not None
        assert cell.value == "表头"

    def test_set_header_on_merged_cell_returns_none(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.merge_cells("J1:K1")
        fill = PatternFill(start_color="4472C4", fill_type="solid")

        result = OutputWriter._set_header_cell(ws, 1, 11, "备选", fill)

        assert result is None

    def test_add_extra_headers_with_merged_header_row(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.merge_cells("J1:L1")
        writer = OutputWriter.__new__(OutputWriter)

        writer._add_extra_headers(ws, header_row=1)


class TestApplyPostFormatWithMergedCells:
    def test_post_format_skips_merged_cells(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="序号")
        ws.cell(row=1, column=3, value="名称")
        ws.cell(row=2, column=1, value=1)
        ws.cell(row=2, column=3, value="DN25钢管")
        ws.merge_cells("C2:D2")

        writer = OutputWriter.__new__(OutputWriter)

        writer._apply_post_format(ws, header_row=1)


def test_write_results_preserve_structure_skips_merged_extra_cells(tmp_path):
    """Exercise the real preserve-structure export path against merged extras."""
    source_path = tmp_path / "merged_source.xlsx"
    output_path = tmp_path / "merged_output.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fire"
    ws.append(["序号", "编码", "名称", "项目特征", "单位", "工程量"])
    ws.append([1, "030109001001", "镀锌钢管安装", "", "m", 12])
    ws.merge_cells("J1:J2")
    ws.merge_cells("L1:N1")
    wb.save(source_path)

    writer = OutputWriter()
    results = [{
        "bill_item": {
            "sheet_name": "Fire",
            "source_row": 2,
            "sheet_bill_seq": 1,
            "code": "030109001001",
            "name": "镀锌钢管安装",
            "unit": "m",
            "quantity": 12,
        },
        "quotas": [{
            "quota_id": "C10-1-1",
            "name": "管道安装",
            "unit": "m",
        }],
        "confidence": 90,
        "explanation": "匹配成功",
        "alternatives": [
            {"quota_id": "C10-1-2", "name": "备选A"},
            {"quota_id": "C10-1-3", "name": "备选B"},
        ],
    }]

    writer.write_results(results, str(output_path), original_file=str(source_path))

    out_wb = openpyxl.load_workbook(output_path)
    try:
        out_ws = out_wb["Fire"]
        assert "J1:J2" in {str(rng) for rng in out_ws.merged_cells.ranges}
        assert out_ws.cell(row=3, column=2).value == "C10-1-1"
        assert out_ws.cell(row=3, column=3).value == "管道安装"
    finally:
        out_wb.close()


class TestReviewSheetSelection:
    def test_review_sheet_includes_high_confidence_no_match_item(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        writer = OutputWriter()

        results = [{
            "bill_item": {"name": "DN25钢管", "description": "无匹配场景"},
            "confidence": 95,
            "quotas": [],
            "match_source": "agent",
            "explanation": "无候选",
            "alternatives": [],
        }]

        writer._write_review_sheet(ws, results)

        assert ws.cell(row=2, column=1).value == 1
        assert ws.cell(row=2, column=2).value == "DN25钢管"
        assert ws.cell(row=2, column=6).value

    def test_review_sheet_includes_high_confidence_agent_fallback_item(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        writer = OutputWriter()

        results = [{
            "bill_item": {"name": "DN32钢管", "description": "降级来源场景"},
            "confidence": 92,
            "quotas": [{"quota_id": "C10-1-1", "name": "钢管安装"}],
            "match_source": "agent_fallback",
            "explanation": "降级结果",
            "alternatives": [],
        }]

        writer._write_review_sheet(ws, results)

        assert ws.cell(row=2, column=1).value == 1
        assert ws.cell(row=2, column=4).value == "C10-1-1"
        assert ws.cell(row=2, column=6).value

    def test_review_sheet_includes_high_confidence_search_item_forced_manual_review(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        writer = OutputWriter()

        results = [{
            "bill_item": {"name": "DN40钢管", "description": "统一排序命中但需终审"},
            "confidence": 91,
            "quotas": [{"quota_id": "C10-2-1", "name": "钢管安装"}],
            "match_source": "search",
            "require_final_review": True,
            "reasoning_decision": {
                "reason": "arbitrated_small_gap",
                "require_final_review": True,
            },
            "explanation": "统一排序改写后需人工复核",
            "alternatives": [],
        }]

        writer._write_review_sheet(ws, results)

        assert ws.cell(row=2, column=1).value == 1
        assert ws.cell(row=2, column=4).value == "C10-2-1"
        assert ws.cell(row=2, column=6).value
