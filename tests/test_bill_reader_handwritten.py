"""手填清单格式识别测试

测试系统对非标准表头（如"项目内容""子目名称"）和带序号分节标题的处理能力。
"""
import shutil
import uuid
from pathlib import Path

import openpyxl

from src.bill_reader import BillReader


def _new_tmp_dir() -> Path:
    root = Path("output") / "test_tmp"
    root.mkdir(parents=True, exist_ok=True)
    d = root / f"handwritten-{uuid.uuid4().hex}"
    d.mkdir(parents=True, exist_ok=False)
    return d


# ================================================================
# 测试1：非标准表头"项目内容"能正确映射到name
# ================================================================
def test_detects_project_content_header():
    """表头用"项目内容"代替"项目名称"，系统能正确识别"""
    tmp_dir = _new_tmp_dir()
    xlsx_path = tmp_dir / "project_content.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "消防电工程"

    # 模拟截图中的格式：序号 | 项目内容 | 计量单位 | 工程数量
    ws.append(["序号", "项 目 内 容", "计量单位", "工程数量"])
    ws.append([29, "电气火灾监控模块", "套", 2])
    ws.append([30, "配线WDZN-RYS-2*1.5", "米", 115.25])
    ws.append([31, "液位探测装置", "台", 2])
    try:
        wb.save(xlsx_path)
        wb.close()

        items = BillReader().read_excel(str(xlsx_path))
        assert len(items) == 3, f"期望3条，实际{len(items)}条"
        assert items[0]["name"] == "电气火灾监控模块"
        assert items[1]["name"] == "配线WDZN-RYS-2*1.5"
        assert items[2]["unit"] == "台"
        assert items[2]["quantity"] == 2.0
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)


# ================================================================
# 测试2："子目名称""子目编码""子目特征"表头能正确映射
# ================================================================
def test_detects_sub_item_headers():
    """表头用"子目编码""子目名称""子目特征描述"，系统能正确识别"""
    tmp_dir = _new_tmp_dir()
    xlsx_path = tmp_dir / "sub_item.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active

    ws.append(["序号", "子目编码", "子目名称", "子目特征描述", "计量单位", "工程量"])
    ws.append([1, None, "LED管型灯具", "220v 22w吸顶安装", "个", 9])
    ws.append([2, None, "15寸LED圆盘灯", "220v 13w吸顶安装", "个", 14])
    try:
        wb.save(xlsx_path)
        wb.close()

        items = BillReader().read_excel(str(xlsx_path))
        assert len(items) == 2, f"期望2条，实际{len(items)}条"
        assert items[0]["name"] == "LED管型灯具"
        assert items[0]["description"] == "220v 22w吸顶安装"
        assert items[1]["quantity"] == 14.0
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)


# ================================================================
# 测试3："工作内容及范围"映射到description
# ================================================================
def test_detects_work_scope_as_description():
    """表头含"工作内容及范围"，映射到description字段"""
    tmp_dir = _new_tmp_dir()
    xlsx_path = tmp_dir / "work_scope.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active

    ws.append(["序号", "项目名称", "项目特征", "工作内容及范围", "单位", "工程量"])
    ws.append([104, "消火栓镀锌钢管", "1.规格:DN32", "管道安装及试压", "m", 10.72])
    try:
        wb.save(xlsx_path)
        wb.close()

        items = BillReader().read_excel(str(xlsx_path))
        assert len(items) == 1
        assert items[0]["name"] == "消火栓镀锌钢管"
        # 当"项目特征"和"工作内容及范围"同时存在时，
        # 后扫描到的列会覆盖description映射，实际取到"工作内容及范围"列的值
        assert items[0].get("description") == "管道安装及试压"
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)


# ================================================================
# 测试4：有序号的分节标题行被正确识别为标题（不当作数据行）
# ================================================================
def test_section_header_with_serial_number():
    """手填清单中"4.0 | 二、一层照明"这种带序号的标题行应被过滤"""
    tmp_dir = _new_tmp_dir()
    xlsx_path = tmp_dir / "section_with_seq.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active

    # 模拟照明工程量清单的格式
    ws.append(["序号", "子目编码", "子目名称", "子目特征描述", "计量单位", "工程量"])
    ws.append([1, None, "LED管型灯具", "220v 22w吸顶安装", "个", 9])
    ws.append([2, None, "15寸LED圆盘灯", "220v 13w吸顶安装", "个", 14])
    ws.append([3, None, "6寸防水防尘灯", None, "个", 16])
    # 这是分节标题行（有序号但是标题）
    ws.append([4, None, "二、一层羽毛球网球夹层照明", None, None, None])
    ws.append([5, None, "LED管型灯具", "220v 22w吸顶安装", "个", 9])
    ws.append([6, None, "12寸LED圆盘灯", "220v 11w 吸顶安装", "个", 66])
    try:
        wb.save(xlsx_path)
        wb.close()

        items = BillReader().read_excel(str(xlsx_path))
        # 应该有5条数据（3+2），标题行"二、一层羽毛球..."被过滤
        assert len(items) == 5, f"期望5条，实际{len(items)}条: {[i['name'] for i in items]}"
        names = [i["name"] for i in items]
        assert "二、一层羽毛球网球夹层照明" not in names
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)


# ================================================================
# 测试5：多种中文序号分节标题都能识别
# ================================================================
def test_various_section_title_formats():
    """各种分节标题格式都能被正确识别"""
    tmp_dir = _new_tmp_dir()
    xlsx_path = tmp_dir / "various_sections.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active

    ws.append(["序号", "项目内容", "计量单位", "工程数量"])
    # 正常数据行
    ws.append([1, "电气火灾监控模块", "套", 2])
    # 各种标题格式（都应被过滤）
    ws.append([2, "一、消防报警系统", None, None])
    ws.append([3, "消防联动报警控制器", "台", 1])
    ws.append([4, "三、应急照明系统", None, None])
    ws.append([5, "应急照明控制主机", "台", 1])
    ws.append([6, "五.电线电缆", None, None])
    ws.append([7, "配线ZR-BV-2*2.5", "米", 1315.91])
    try:
        wb.save(xlsx_path)
        wb.close()

        items = BillReader().read_excel(str(xlsx_path))
        # 应该有4条数据，3条标题行被过滤
        names = [i["name"] for i in items]
        assert len(items) == 4, f"期望4条，实际{len(items)}条: {names}"
        assert "一、消防报警系统" not in names
        assert "三、应急照明系统" not in names
        assert "五.电线电缆" not in names
        assert "配线ZR-BV-2*2.5" in names
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)


# ================================================================
# 测试6：标准清单格式不受影响（回归测试）
# ================================================================
def test_standard_format_still_works():
    """确保改动不影响标准清单格式的识别"""
    tmp_dir = _new_tmp_dir()
    xlsx_path = tmp_dir / "standard.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active

    ws.append(["序号", "项目编码", "项目名称", "项目特征描述", "计量单位", "工程量"])
    ws.append([1, "030901001001", "室内给水管道", "1.材质:PPR\n2.规格:DN25", "m", 120])
    ws.append([None, None, "给水系统", None, None, None])  # 分部标题（无序号）
    ws.append([2, "030901001002", "室内给水管道", "1.材质:PPR\n2.规格:DN32", "m", 85])
    try:
        wb.save(xlsx_path)
        wb.close()

        items = BillReader().read_excel(str(xlsx_path))
        assert len(items) == 2, f"期望2条，实际{len(items)}条"
        assert items[0]["code"] == "030901001001"
        assert items[0]["name"] == "室内给水管道"
        assert items[1]["code"] == "030901001002"
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)
