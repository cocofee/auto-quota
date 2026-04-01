import sys
from pathlib import Path
from types import SimpleNamespace

import openpyxl

from src.output_writer import OutputWriter


BACKEND_ROOT = Path(__file__).resolve().parents[1] / "web" / "backend"
if str(BACKEND_ROOT) not in sys.path:
    sys.path.append(str(BACKEND_ROOT))

from app.api import results as results_api  # noqa: E402


def _build_source_workbook(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fire"
    ws.append(["序号", "编码", "名称", "项目特征", "单位", "工程量"])
    ws.append([1, "030109001001", "泵1", "", "台", 2])
    ws.append([2, "030109001002", "泵2", "", "台", 2])
    ws.append([3, "030109001003", "泵3", "", "台", 2])
    ws.append([4, "030404017001", "配电箱", "", "台", 17])
    ws.append([5, "030404032001", "端子箱", "", "台", 17])
    wb.save(path)


def _build_duplicate_code_workbook(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fire"
    ws.append(["序号", "编码", "名称", "项目特征", "单位", "工程量"])
    ws.append([1, "030109001001", "泵1", "", "台", 2])
    ws.append([2, "030408002001", "控制电缆A", "", "m", 119.3])
    ws.append([3, "030408002001", "控制电缆B", "", "m", 60])
    wb.save(path)


def test_build_rebuilt_results_from_source_file_recovers_source_row_for_subset(tmp_path):
    source_path = tmp_path / "input.xlsx"
    _build_source_workbook(source_path)

    task = SimpleNamespace(
        id="task-export-rebuild",
        file_path=str(source_path),
        output_path=None,
        sheet=None,
    )
    items = [
        SimpleNamespace(
            index=0,
            corrected_quotas=None,
            quotas=[{"quota_id": "4-4-16", "name": "端子箱安装 户内", "unit": "台"}],
            confidence=88,
            explanation="ok",
            match_source="search",
            bill_code="030404032001",
            bill_name="端子箱",
            bill_description="",
            bill_unit="台",
            bill_quantity=17,
            sheet_name="Fire",
            section="",
            specialty="C4",
        ),
    ]

    rebuilt = results_api._build_rebuilt_results_from_source_file(task, items)

    assert rebuilt is not None
    assert rebuilt[0]["bill_item"]["source_row"] == 6
    assert rebuilt[0]["bill_item"]["sheet_bill_seq"] == 5
    assert rebuilt[0]["bill_item"]["sheet_name"] == "Fire"


def test_output_writer_ignores_stale_subset_relative_sheet_bill_seq(tmp_path):
    source_path = tmp_path / "input.xlsx"
    output_path = tmp_path / "output.xlsx"
    _build_source_workbook(source_path)

    writer = OutputWriter()
    subset_results = [{
        "bill_item": {
            "sheet_name": "Fire",
            "sheet_bill_seq": 1,
            "code": "030404032001",
            "name": "端子箱",
            "unit": "台",
            "quantity": 17,
        },
        "quotas": [{
            "quota_id": "4-4-16",
            "name": "端子箱安装 户内",
            "unit": "台",
        }],
        "confidence": 88,
        "explanation": "subset old mapping",
        "alternatives": [],
    }]

    writer.write_results(subset_results, str(output_path), original_file=str(source_path))

    ws = openpyxl.load_workbook(output_path)["Fire"]
    assert ws.cell(row=2, column=3).value == "泵1"
    assert ws.cell(row=3, column=3).value == "泵2"
    assert ws.cell(row=4, column=3).value == "泵3"
    assert ws.cell(row=5, column=3).value == "配电箱"
    assert ws.cell(row=6, column=3).value == "端子箱"
    assert ws.cell(row=7, column=2).value == "4-4-16"
    assert ws.cell(row=7, column=3).value == "端子箱安装 户内"


def test_output_writer_locator_matching_handles_duplicate_codes_without_source_row(tmp_path):
    source_path = tmp_path / "input_dup.xlsx"
    output_path = tmp_path / "output_dup.xlsx"
    _build_duplicate_code_workbook(source_path)

    writer = OutputWriter()
    subset_results = [
        {
            "bill_item": {
                "sheet_name": "Fire",
                "sheet_bill_seq": 1,
                "code": "030408002001",
                "name": "控制电缆A",
                "unit": "m",
                "quantity": 119.3,
            },
            "quotas": [{
                "quota_id": "4-9-382",
                "name": "控制缆终端头≤24芯",
                "unit": "m",
            }],
            "confidence": 88,
            "explanation": "duplicate code A",
            "alternatives": [],
        },
        {
            "bill_item": {
                "sheet_name": "Fire",
                "sheet_bill_seq": 2,
                "code": "030408002001",
                "name": "控制电缆B",
                "unit": "m",
                "quantity": 60,
            },
            "quotas": [{
                "quota_id": "4-9-388",
                "name": "矿物绝缘控制电缆 终端头≤7芯",
                "unit": "m",
            }],
            "confidence": 87,
            "explanation": "duplicate code B",
            "alternatives": [],
        },
    ]

    writer.write_results(subset_results, str(output_path), original_file=str(source_path))

    ws = openpyxl.load_workbook(output_path)["Fire"]
    assert ws.cell(row=3, column=3).value == "控制电缆A"
    assert ws.cell(row=4, column=2).value == "4-9-382"
    assert ws.cell(row=5, column=3).value == "控制电缆B"
    assert ws.cell(row=6, column=2).value == "4-9-388"
