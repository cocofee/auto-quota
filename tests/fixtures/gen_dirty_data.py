"""
生成 B4 脏数据样本 Excel。

运行方式：python tests/fixtures/gen_dirty_data.py
输出：tests/fixtures/dirty_data_sample.xlsx
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Font

OUTPUT_PATH = Path(__file__).parent / "dirty_data_sample.xlsx"

# 20条脏数据清单项
# 格式：(序号, 编码, 名称, 特征描述, 单位, 工程量)
DIRTY_ITEMS = [
    # === 第1组：缺字段（5条）===
    (1, "030101001001", "", "De25,热熔连接", "m", 120),
    # 名称为空
    (2, "030101002001", "给水管道安装", "", "m", 85),
    # 描述为空
    (3, "030101003001", "排水管道安装", "De110,粘接", None, 60),
    # 单位为空
    (4, "", "镀锌钢管安装", "DN25,丝接", "m", 45),
    # 编码为空
    (5, "030101005001", "阀门安装", "DN25,截止阀", "m", None),
    # 工程量为空

    # === 第2组：单位异常（5条）===
    (6, "030201001001", "配电箱安装", "XL-21型,落地式", "套", 3),
    # "套"应为"台"
    (7, "030202001001", "灯具安装", "LED吸顶灯,圆形", "棵", 50),
    # "棵"应为"套"
    (8, "030203001001", "电缆敷设", "YJV-3×120+2×70", "m", 350),
    # 正确单位（对照组）
    (9, "030204001001", "风管制作安装", "镀锌钢板,δ=0.75mm", "m", 200),
    # "m"应为"m2"
    (10, "030205001001", "消火栓安装", "室内消火栓,DN65", "个", 24),
    # "个"应为"组"

    # === 第3组：描述噪声（5条）===
    (11, "030301001001", "给水管道  安装  ", "  De25 , 热熔连接  ,PPR管  ", "m", 100),
    # 多余空格
    (12, "030302001001", "排水管道安装\n", "De110<br>粘接,UPVC管", "m", 80),
    # 换行符和HTML标签
    (13, "030303001001", "镀锌钢管安装（DN25）（丝接）", "***重要***DN25,丝接,镀锌钢管", "m", 65),
    # 全角括号和特殊标记
    (14, "030304001001", "电缆~敷设~", "YJV-3*120+2*70,沿桥架敷设", "m", 280),
    # 波浪号噪声，星号替代乘号
    (15, "030305001001", "配管（SC20）", "配管SC20，暗敷,从配电箱至灯位", "m", 150),
    # 全角逗号混用

    # === 第4组：名称歧义（5条）===
    (16, "030401001001", "管道", "DN25", "m", 90),
    # 不知道是给水还是排水还是消防
    (17, "030402001001", "灯", "10W", "套", 30),
    # 不知道是什么灯
    (18, "030403001001", "阀", "DN50", "个", 12),
    # 不知道是什么阀
    (19, "030404001001", "线", "BV-2.5mm2", "m", 500),
    # 不知道是配线还是电缆
    (20, "030405001001", "安装", "不锈钢,DN25", "m", 40),
    # 不知道安装什么
]


def generate():
    """生成脏数据样本 Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "脏数据测试"

    # 表头（模仿广联达清单格式）
    headers = ["序号", "项目编码", "项目名称", "项目特征", "计量单位", "工程量"]
    header_font = Font(bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font

    # 第2行留空（模拟广联达第2行是子表头）
    ws.cell(row=2, column=1, value="")

    # 写入数据
    for i, item in enumerate(DIRTY_ITEMS):
        row = i + 3  # 从第3行开始（跳过2行表头）
        seq, code, name, desc, unit, qty = item
        ws.cell(row=row, column=1, value=seq)
        ws.cell(row=row, column=2, value=code)
        ws.cell(row=row, column=3, value=name)
        ws.cell(row=row, column=4, value=desc)
        ws.cell(row=row, column=5, value=unit)
        ws.cell(row=row, column=6, value=qty)

    # 调整列宽
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 10

    wb.save(str(OUTPUT_PATH))
    print(f"已生成脏数据样本: {OUTPUT_PATH}")
    print(f"共 {len(DIRTY_ITEMS)} 条测试清单项")
    print("  第1组(1-5): 缺字段")
    print("  第2组(6-10): 单位异常")
    print("  第3组(11-15): 描述噪声")
    print("  第4组(16-20): 名称歧义")


if __name__ == "__main__":
    generate()
