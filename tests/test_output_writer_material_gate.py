import openpyxl
from src.output_writer import OutputWriter, _resolve_output_materials


def _project_name_from_filename(source_file_name: str) -> str:
    if not source_file_name:
        return ""
    if "20室外配套灌溉工程" in source_file_name:
        return "20室外配套灌溉工程"
    if "其他项目" in source_file_name:
        return "其他项目"
    return ""


def _make_result(*, source_materials=None, materials=None, source_file_name=""):
    project_name = _project_name_from_filename(source_file_name)
    bill_item = {
        "unit": "m",
        "quantity": 12,
        "description": "名称:球墨铸铁 DN100 1.0MPa",
        "source_file_name": source_file_name,
        "source_file_stem": project_name,
        "source_file_title": project_name,
        "project_name": project_name,
    }
    if source_materials is not None:
        bill_item["source_materials"] = source_materials
    return {
        "bill_item": bill_item,
        "quotas": [{"quota_id": "A1-1", "name": "给水管道安装"}],
        "materials": materials or [],
    }


def test_resolve_output_materials_defers_target_source_material_before_review():
    result = _make_result(
        source_file_name="[安徽]20室外配套灌溉工程_wx_zip(2)_定额匹配结果.xlsx",
        source_materials=[{"name": "球墨铸铁", "spec": "DN100 1.0MPa", "unit": "m", "qty": 12}],
    )

    assert _resolve_output_materials(result) == []


def test_get_material_text_hides_target_material_from_review_sheet():
    result = _make_result(
        source_file_name="[安徽]20室外配套灌溉工程_wx_zip(2)_定额匹配结果.xlsx",
        materials=[{"name": "球墨铸铁", "spec": "DN100 1.0MPa", "unit": "m"}],
    )

    assert OutputWriter._get_material_text(result) == ""


def test_output_material_gate_does_not_affect_other_projects():
    result = _make_result(
        source_file_name="[安徽]其他项目_wx_zip(2)_定额匹配结果.xlsx",
        source_materials=[{"name": "球墨铸铁", "spec": "DN100 1.0MPa", "unit": "m", "qty": 12}],
    )

    materials = _resolve_output_materials(result)
    assert len(materials) == 1
    assert materials[0]["name"] == "球墨铸铁"

def test_get_material_text_hides_target_source_material_summary():
    result = _make_result(
        source_file_name="[安徽]20室外配套灌溉工程_wx_zip(2)_定额匹配结果.xlsx",
        source_materials=[{"name": "球墨铸铁", "spec": "DN100 1.0MPa", "unit": "m", "qty": 12}],
    )

    assert OutputWriter._get_material_text(result) == ""


def test_output_writer_hides_material_text_by_default():
    result = _make_result(
        source_materials=[{"name": "球墨铸铁", "spec": "DN100 1.0MPa", "unit": "m", "qty": 12}],
    )

    writer = OutputWriter()

    assert writer._get_material_text(result, include_materials=writer.include_materials) == ""


def test_output_writer_does_not_emit_material_rows_by_default():
    wb = openpyxl.Workbook()
    ws = wb.active
    writer = OutputWriter()
    result = _make_result(
        source_materials=[{"name": "球墨铸铁", "spec": "DN100 1.0MPa", "unit": "m", "qty": 12}],
    )

    next_row = writer._write_quota_rows(ws, 2, result, bill_unit="m", bill_qty=12, max_col=15)

    assert next_row == 3
    assert ws.cell(row=2, column=2).value == "A1-1"
    assert ws.cell(row=3, column=2).value in (None, "")


def test_output_writer_can_emit_material_rows_when_explicitly_enabled():
    wb = openpyxl.Workbook()
    ws = wb.active
    writer = OutputWriter(include_materials=True)
    result = _make_result(
        source_materials=[{"name": "球墨铸铁", "spec": "DN100 1.0MPa", "unit": "m", "qty": 12}],
    )

    next_row = writer._write_quota_rows(ws, 2, result, bill_unit="m", bill_qty=12, max_col=15)

    assert next_row == 4
    assert ws.cell(row=2, column=2).value == "A1-1"
    assert ws.cell(row=3, column=2).value == "主"
