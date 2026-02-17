"""
大模型对比测试脚本
功能：用同一份电气清单（89条），分别调DeepSeek-V3和Qwen-Plus，
     对比两个模型的匹配结果，并和Opus 4.6基准对比。
用法：python tools/llm_compare_test.py
"""

import json
import re
import sqlite3
import sys
import time
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# 把项目根目录加入路径，以便导入config
PROJECT_ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from dotenv import load_dotenv
load_dotenv(PROJECT_ROOT / ".env")
import config

# ============================================================
# 配置
# ============================================================
BILL_FILE = r"C:\Users\Administrator\Desktop\北京大学新燕园校区电教厅改造工程.xlsx"
SHEET_NAME = "4.6 分部分项工程项目清单计价表【电气工程】"
QUOTA_DB = str(config.get_quota_db_path())  # 定额数据库路径
OUTPUT_FILE = r"C:\Users\Administrator\Desktop\大模型对比测试.xlsx"

# ============================================================
# Opus 4.6 基准结果（已经在上一步生成的匹配结果）
# 格式：序号 → 定额编号列表
# ============================================================
OPUS_BASELINE = {
    1: ["C4-4-31"], 2: ["C4-4-30"], 3: ["C4-4-30"], 4: ["C4-4-30"],
    5: ["C4-4-31"], 6: ["C4-4-31"], 7: ["C4-4-31"], 8: ["C4-4-31"],
    9: ["C4-4-31"], 10: ["C4-4-31"], 11: ["C4-4-31"], 12: ["C4-4-31"],
    13: ["C4-4-32"],
    14: ["C4-8-14"], 15: ["C4-8-14"], 16: ["C4-8-13"], 17: ["C4-8-13"],
    18: ["C4-8-13"], 19: ["C4-8-11"], 20: ["C4-8-11"], 21: ["C4-8-11"],
    22: ["C4-8-11"], 23: ["C4-8-11"], 24: ["C4-8-11"],
    25: ["C4-8-242"], 26: ["C4-8-242"], 27: ["C4-8-242"], 28: ["C4-8-242"],
    29: ["C4-8-241"], 30: ["C4-8-241"], 31: ["C4-8-236"], 32: ["C4-8-241"],
    33: ["C4-8-237"], 34: ["C4-8-241"],
    35: ["C4-11-282"], 36: ["C4-11-282"], 37: ["C4-11-283"],
    38: ["C4-11-121"], 39: ["C4-11-122"], 40: ["C4-11-123"],
    41: ["C4-11-25"], 42: ["C4-11-27"], 43: ["C4-11-29"], 44: ["C4-11-31"],
    45: ["C4-11-249"], 46: ["C4-11-249"],
    47: ["C4-13-1", "C4-13-2"],
    48: ["C12-2-61"],
    49: ["C4-12-21"], 50: ["C4-12-21"], 51: ["C4-12-21"],
    52: ["C4-12-21"], 53: ["C4-12-21"], 54: ["C4-12-9"], 55: ["C4-12-9"],
    56: ["C4-12-67"], 57: ["C4-12-67"], 58: ["C4-12-67"],
    59: ["C4-12-65"], 60: ["C4-12-65"], 61: ["C4-12-67"],
    62: ["C4-12-21"], 63: ["C4-12-67"],
    64: ["C4-12-21"], 65: ["C4-12-21"], 66: ["C4-12-21"], 67: ["C4-12-21"],
    68: ["C4-4-103"], 69: ["C4-4-108"], 70: ["C4-4-103"],
    71: ["C4-4-122"], 72: ["C4-4-103"],
    73: ["C4-11-382"], 76: ["C4-11-382"], 77: ["C4-11-384"],
    74: ["C4-4-31"], 75: ["C4-4-30"],
    78: ["C4-9-11"], 79: ["C4-9-68"], 80: ["C4-9-79"],
    81: ["C4-6-2"], 82: ["C4-6-1"], 83: ["C4-6-1"],
    84: ["C4-6-1"], 85: ["C4-6-1"], 86: ["C4-6-4"],
    87: ["C4-8-493"], 88: ["C4-8-494"], 89: ["C4-8-495"],
}

# ============================================================
# 清单项按类别分组，以及每个类别对应的定额查询SQL
# "category_name": (序号范围, SQL查询条件)
# ============================================================
CATEGORY_QUERIES = {
    "配电箱": {
        "indices": list(range(1, 14)),  # 1-13
        "sql": """
            SELECT quota_id, name, unit FROM quotas
            WHERE quota_id LIKE 'C4-4-%'
            AND (name LIKE '%配电箱%' OR name LIKE '%开关箱%'
                 OR name LIKE '%动力箱%' OR name LIKE '%控制箱%')
            ORDER BY quota_id
        """,
    },
    "电力电缆敷设": {
        "indices": list(range(14, 25)),  # 14-24
        "sql": """
            SELECT quota_id, name, unit FROM quotas
            WHERE quota_id LIKE 'C4-8-%'
            AND name LIKE '%电缆%敷设%'
            AND name NOT LIKE '%控制电缆%'
            AND name NOT LIKE '%控制缆%'
            AND CAST(REPLACE(quota_id, 'C4-8-', '') AS INTEGER) <= 30
            ORDER BY CAST(REPLACE(quota_id, 'C4-8-', '') AS INTEGER)
        """,
    },
    "电缆头": {
        "indices": list(range(25, 35)),  # 25-34
        "sql": """
            SELECT quota_id, name, unit FROM quotas
            WHERE quota_id LIKE 'C4-8-%'
            AND name LIKE '%非铠装%电缆%终端头%'
            ORDER BY quota_id
        """,
    },
    "配线": {
        "indices": list(range(35, 38)),  # 35-37
        "sql": """
            SELECT quota_id, name, unit FROM quotas
            WHERE quota_id LIKE 'C4-11-%'
            AND (name LIKE '%穿%线%' OR name LIKE '%配线%')
            ORDER BY quota_id
        """,
    },
    "配管": {
        "indices": list(range(38, 45)),  # 38-44
        "sql": """
            SELECT quota_id, name, unit FROM quotas
            WHERE quota_id LIKE 'C4-11-%'
            AND (name LIKE '%紧定%' OR name LIKE '%扣压%'
                 OR name LIKE '%焊接钢管%')
            AND CAST(REPLACE(quota_id, 'C4-11-', '') AS INTEGER) <= 50
            ORDER BY CAST(REPLACE(quota_id, 'C4-11-', '') AS INTEGER)
        """,
    },
    "桥架": {
        "indices": [45, 46],
        "sql": """
            SELECT quota_id, name, unit FROM quotas
            WHERE quota_id LIKE 'C4-11-%'
            AND name LIKE '%桥架%'
            ORDER BY quota_id
        """,
    },
    "支架": {
        "indices": [47],
        "sql": """
            SELECT quota_id, name, unit FROM quotas
            WHERE quota_id LIKE 'C4-13-%'
            ORDER BY quota_id
        """,
    },
    "刷油": {
        "indices": [48],
        "sql": """
            SELECT quota_id, name, unit FROM quotas
            WHERE quota_id LIKE 'C12-2-%'
            AND name LIKE '%金属%刷油%'
            ORDER BY quota_id
        """,
    },
    "灯具": {
        "indices": list(range(49, 68)),  # 49-67
        "sql": """
            SELECT quota_id, name, unit FROM quotas
            WHERE quota_id LIKE 'C4-12-%'
            AND (name LIKE '%灯%' OR name LIKE '%照明%' OR name LIKE '%标志%')
            AND CAST(REPLACE(quota_id, 'C4-12-', '') AS INTEGER) <= 80
            ORDER BY CAST(REPLACE(quota_id, 'C4-12-', '') AS INTEGER)
        """,
    },
    "开关插座": {
        "indices": [68, 69, 70, 71, 72],
        "sql": """
            SELECT quota_id, name, unit FROM quotas
            WHERE quota_id LIKE 'C4-4-%'
            AND (name LIKE '%开关%' OR name LIKE '%插座%' OR name LIKE '%按钮%')
            ORDER BY quota_id
        """,
    },
    "接线盒": {
        "indices": [73, 76, 77],
        "sql": """
            SELECT quota_id, name, unit FROM quotas
            WHERE quota_id LIKE 'C4-11-%'
            AND (name LIKE '%接线盒%' OR name LIKE '%过路盒%' OR name LIKE '%分线盒%')
            ORDER BY quota_id
        """,
    },
    "控制器": {
        "indices": [74, 75],
        "sql": """
            SELECT quota_id, name, unit FROM quotas
            WHERE (quota_id LIKE 'C4-4-%' OR quota_id LIKE 'C5-%')
            AND (name LIKE '%控制%' OR name LIKE '%智能%' OR name LIKE '%配电箱%')
            AND quota_id LIKE 'C4-4-%'
            ORDER BY quota_id
            LIMIT 30
        """,
    },
    "防雷接地": {
        "indices": [78, 79, 80],
        "sql": """
            SELECT quota_id, name, unit FROM quotas
            WHERE quota_id LIKE 'C4-9-%'
            AND (name LIKE '%接地%' OR name LIKE '%防雷%' OR name LIKE '%引下线%'
                 OR name LIKE '%接闪%' OR name LIKE '%避雷%')
            ORDER BY CAST(REPLACE(quota_id, 'C4-9-', '') AS INTEGER)
        """,
    },
    "电动机": {
        "indices": list(range(81, 87)),  # 81-86
        "sql": """
            SELECT quota_id, name, unit FROM quotas
            WHERE quota_id LIKE 'C4-6-%'
            ORDER BY quota_id
        """,
    },
    "防火封堵": {
        "indices": [87, 88, 89],
        "sql": """
            SELECT quota_id, name, unit FROM quotas
            WHERE quota_id LIKE 'C4-8-%'
            AND (name LIKE '%防火%' OR name LIKE '%封堵%' OR name LIKE '%堵洞%')
            ORDER BY quota_id
        """,
    },
}


def read_bill_items(file_path, sheet_name):
    """读取清单项目"""
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    items = {}  # 用序号作为key
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 3:
            continue
        a_val = row[0] if len(row) > 0 else None
        if a_val is None:
            continue
        try:
            idx = int(a_val)
        except (ValueError, TypeError):
            continue
        items[idx] = {
            "index": idx,
            "code": str(row[1] or "").strip() if len(row) > 1 else "",
            "name": str(row[2] or "").strip() if len(row) > 2 else "",
            "description": str(row[3] or "").strip() if len(row) > 3 else "",
            "unit": str(row[4] or "").strip() if len(row) > 4 else "",
            "quantity": row[5] if len(row) > 5 else None,
        }
    wb.close()
    return items


def get_category_quotas(db_path):
    """按类别从定额库查询候选定额"""
    conn = sqlite3.connect(db_path)
    category_quotas = {}
    for cat_name, cat_info in CATEGORY_QUERIES.items():
        rows = conn.execute(cat_info["sql"]).fetchall()
        quota_list = []
        for r in rows:
            quota_list.append({
                "quota_id": r[0],
                "name": r[1],
                "unit": r[2] or "",
            })
        category_quotas[cat_name] = quota_list
        print(f"  {cat_name}: {len(quota_list)} 条候选定额")
    conn.close()
    return category_quotas


def get_item_category(idx):
    """根据序号查找所属类别"""
    for cat_name, cat_info in CATEGORY_QUERIES.items():
        if idx in cat_info["indices"]:
            return cat_name
    return None


def build_prompt(bill_item, quota_list):
    """构建发给大模型的Prompt"""
    # 定额候选列表文本
    quota_lines = []
    for i, q in enumerate(quota_list, 1):
        quota_lines.append(f"{i}. [{q['quota_id']}] {q['name']} | 单位:{q['unit']}")
    quota_text = "\n".join(quota_lines)

    prompt = f"""你是经验丰富的北京造价师，精通北京2024版安装定额。
根据清单项信息，从定额列表中选择最合适的定额。

## 清单项
- 项目名称：{bill_item['name']}
- 项目特征描述：{bill_item['description']}
- 计量单位：{bill_item['unit']}
- 工程量：{bill_item['quantity']}

## 可选定额列表
{quota_text}

## 匹配要求
1. 从列表中选1条最合适的主定额
2. "以内"表示该定额适用于不超过指定参数的项目
3. 注意回路数、截面积、管径、功率等参数取档（向上取最近的档位）
4. 如果需要多条定额（如支架制作+安装），都列出来

## 输出格式
严格按以下JSON格式回答（不要输出其他内容）：
```json
{{"quota_ids": ["C4-4-31"], "reason": "8回路配电箱墙上明装，取≤8回路档", "confidence": 90}}
```
如果需要多条定额：
```json
{{"quota_ids": ["C4-13-1", "C4-13-2"], "reason": "支架需要制作+安装两条定额", "confidence": 85}}
```"""
    return prompt


def call_llm(client, model, prompt, timeout=30):
    """调用大模型API（OpenAI兼容接口）"""
    try:
        response = client.chat.completions.create(
            model=model,
            messages=[{"role": "user", "content": prompt}],
            temperature=0.1,
            max_tokens=500,
            timeout=timeout,
        )
        return response.choices[0].message.content
    except Exception as e:
        return f"ERROR: {e}"


def parse_response(text):
    """解析大模型返回的JSON"""
    if text.startswith("ERROR:"):
        return {"quota_ids": [], "reason": text, "confidence": 0}

    # 尝试从返回文本中提取JSON
    # 大模型可能返回 ```json ... ``` 包裹的JSON
    json_match = re.search(r'```json\s*(\{.*?\})\s*```', text, re.DOTALL)
    if json_match:
        text = json_match.group(1)
    else:
        # 尝试直接找JSON对象
        json_match = re.search(r'\{[^{}]*"quota_ids"[^{}]*\}', text, re.DOTALL)
        if json_match:
            text = json_match.group(0)

    try:
        result = json.loads(text)
        quota_ids = result.get("quota_ids", [])
        # 有的模型可能返回 quota_id（单数）而不是 quota_ids（复数）
        if not quota_ids and "quota_id" in result:
            quota_ids = [result["quota_id"]]
        return {
            "quota_ids": quota_ids,
            "reason": result.get("reason", ""),
            "confidence": result.get("confidence", 0),
        }
    except (json.JSONDecodeError, TypeError):
        # JSON解析失败，尝试提取定额编号
        ids = re.findall(r'C\d+-\d+-\d+', text)
        return {
            "quota_ids": ids[:3],  # 最多取3个
            "reason": text[:200],
            "confidence": 0,
        }


def compare_quota_ids(pred_ids, baseline_ids):
    """比较预测的定额编号和基准是否一致
    返回: "完全一致" / "主定额一致" / "不一致"
    """
    if not pred_ids or not baseline_ids:
        return "不一致"

    # 完全一致（集合相同）
    if set(pred_ids) == set(baseline_ids):
        return "完全一致"

    # 主定额一致（第一个定额相同）
    if pred_ids[0] == baseline_ids[0]:
        return "主定额一致"

    return "不一致"


def write_output_excel(items, results, output_path):
    """写入对比结果Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "对比结果"

    # 样式
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, size=10, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    green_fill = PatternFill("solid", fgColor="C6EFCE")
    yellow_fill = PatternFill("solid", fgColor="FFEB9C")
    red_fill = PatternFill("solid", fgColor="FFC7CE")
    wrap = Alignment(wrap_text=True, vertical="top")

    # 表头
    headers = [
        "序号", "项目名称", "项目特征(摘要)",
        "Opus4.6\n定额编号",
        "DeepSeek\n定额编号", "DeepSeek\n匹配说明",
        "Qwen\n定额编号", "Qwen\n匹配说明",
        "DeepSeek\nvs Opus", "Qwen\nvs Opus",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 列宽
    widths = [5, 16, 30, 14, 14, 30, 14, 30, 12, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # 数据行
    row = 2
    for idx in sorted(results.keys()):
        r = results[idx]
        item = items.get(idx, {})
        desc = item.get("description", "")
        if len(desc) > 80:
            desc = desc[:80] + "..."

        opus_ids = OPUS_BASELINE.get(idx, [])
        opus_str = "\n".join(opus_ids)

        ds_ids = r.get("deepseek", {}).get("quota_ids", [])
        ds_str = "\n".join(ds_ids) if ds_ids else "(无结果)"
        ds_reason = r.get("deepseek", {}).get("reason", "")

        qw_ids = r.get("qwen", {}).get("quota_ids", [])
        qw_str = "\n".join(qw_ids) if qw_ids else "(无结果)"
        qw_reason = r.get("qwen", {}).get("reason", "")

        ds_cmp = compare_quota_ids(ds_ids, opus_ids)
        qw_cmp = compare_quota_ids(qw_ids, opus_ids)

        ws.cell(row=row, column=1, value=idx).border = thin_border
        ws.cell(row=row, column=2, value=item.get("name", "")).border = thin_border

        c3 = ws.cell(row=row, column=3, value=desc)
        c3.border = thin_border
        c3.alignment = wrap

        c4 = ws.cell(row=row, column=4, value=opus_str)
        c4.border = thin_border
        c4.alignment = wrap

        c5 = ws.cell(row=row, column=5, value=ds_str)
        c5.border = thin_border
        c5.alignment = wrap

        c6 = ws.cell(row=row, column=6, value=ds_reason)
        c6.border = thin_border
        c6.alignment = wrap

        c7 = ws.cell(row=row, column=7, value=qw_str)
        c7.border = thin_border
        c7.alignment = wrap

        c8 = ws.cell(row=row, column=8, value=qw_reason)
        c8.border = thin_border
        c8.alignment = wrap

        # DeepSeek对比结果 - 颜色
        c9 = ws.cell(row=row, column=9, value=ds_cmp)
        c9.border = thin_border
        c9.alignment = Alignment(horizontal="center", vertical="center")
        if "一致" in ds_cmp:
            c9.fill = green_fill
        else:
            c9.fill = red_fill

        # Qwen对比结果 - 颜色
        c10 = ws.cell(row=row, column=10, value=qw_cmp)
        c10.border = thin_border
        c10.alignment = Alignment(horizontal="center", vertical="center")
        if "一致" in qw_cmp:
            c10.fill = green_fill
        else:
            c10.fill = red_fill

        row += 1

    ws.freeze_panes = "A2"

    # ============================================================
    # 统计Sheet
    # ============================================================
    ws2 = wb.create_sheet("统计")
    stats_header_font = Font(bold=True, size=11)

    # 计算统计
    total = len(results)
    ds_full = sum(1 for r in results.values()
                  if compare_quota_ids(r.get("deepseek", {}).get("quota_ids", []),
                                       OPUS_BASELINE.get(list(results.keys())[0], [])) != "x"
                  and "一致" in compare_quota_ids(
                      r.get("deepseek", {}).get("quota_ids", []),
                      OPUS_BASELINE.get(r.get("_idx", 0), [])))
    # 重新计算
    ds_match = 0
    ds_main = 0
    qw_match = 0
    qw_main = 0
    ds_total_time = 0
    qw_total_time = 0

    for idx, r in results.items():
        opus_ids = OPUS_BASELINE.get(idx, [])
        ds_ids = r.get("deepseek", {}).get("quota_ids", [])
        qw_ids = r.get("qwen", {}).get("quota_ids", [])

        ds_cmp = compare_quota_ids(ds_ids, opus_ids)
        if ds_cmp == "完全一致":
            ds_match += 1
        elif ds_cmp == "主定额一致":
            ds_main += 1

        qw_cmp = compare_quota_ids(qw_ids, opus_ids)
        if qw_cmp == "完全一致":
            qw_match += 1
        elif qw_cmp == "主定额一致":
            qw_main += 1

        ds_total_time += r.get("deepseek_time", 0)
        qw_total_time += r.get("qwen_time", 0)

    # 写统计表
    ws2.cell(row=1, column=1, value="指标").font = stats_header_font
    ws2.cell(row=1, column=2, value="DeepSeek-V3").font = stats_header_font
    ws2.cell(row=1, column=3, value="Qwen-Plus").font = stats_header_font

    stats_data = [
        ("总清单条数", total, total),
        ("完全一致", ds_match, qw_match),
        ("主定额一致", ds_main, qw_main),
        ("一致率(含主定额)", f"{(ds_match + ds_main) * 100 / max(total, 1):.1f}%",
         f"{(qw_match + qw_main) * 100 / max(total, 1):.1f}%"),
        ("完全一致率", f"{ds_match * 100 / max(total, 1):.1f}%",
         f"{qw_match * 100 / max(total, 1):.1f}%"),
        ("不一致", total - ds_match - ds_main, total - qw_match - qw_main),
        ("总耗时(秒)", f"{ds_total_time:.1f}", f"{qw_total_time:.1f}"),
        ("平均每条(秒)", f"{ds_total_time / max(total, 1):.2f}",
         f"{qw_total_time / max(total, 1):.2f}"),
    ]

    for i, (label, ds_val, qw_val) in enumerate(stats_data, 2):
        ws2.cell(row=i, column=1, value=label)
        ws2.cell(row=i, column=2, value=ds_val)
        ws2.cell(row=i, column=3, value=qw_val)

    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 15
    ws2.column_dimensions["C"].width = 15

    # 说明
    ws2.cell(row=12, column=1, value="说明：").font = stats_header_font
    ws2.cell(row=13, column=1, value="完全一致 = 所有定额编号完全相同")
    ws2.cell(row=14, column=1, value="主定额一致 = 第一条（主）定额编号相同，关联定额可能不同")
    ws2.cell(row=15, column=1, value="一致率是和 Opus 4.6 基准对比的结果，不代表绝对准确率")
    ws2.cell(row=16, column=1, value="Opus 4.6 基准本身也可能有错，最终需要人工审核确认")

    wb.save(output_path)


def main():
    print("=" * 60)
    print("大模型对比测试: DeepSeek-V3 vs Qwen-Plus")
    print("=" * 60)

    # 检查API Key
    has_deepseek = bool(config.DEEPSEEK_API_KEY)
    has_qwen = bool(config.QWEN_API_KEY)

    if not has_deepseek and not has_qwen:
        print("\n错误：没有配置任何API Key！")
        print("请在 .env 文件中填入至少一个：")
        print("  DEEPSEEK_API_KEY=sk-xxx    (去 platform.deepseek.com 注册)")
        print("  QWEN_API_KEY=sk-xxx        (去 dashscope.console.aliyun.com 注册)")
        return

    print(f"\nAPI状态:")
    print(f"  DeepSeek: {'已配置' if has_deepseek else '未配置(跳过)'}")
    print(f"  Qwen:     {'已配置' if has_qwen else '未配置(跳过)'}")

    # 初始化API客户端
    ds_client = None
    qw_client = None

    if has_deepseek:
        from openai import OpenAI
        ds_client = OpenAI(
            api_key=config.DEEPSEEK_API_KEY,
            base_url=config.DEEPSEEK_BASE_URL,
        )

    if has_qwen:
        from openai import OpenAI
        qw_client = OpenAI(
            api_key=config.QWEN_API_KEY,
            base_url=config.QWEN_BASE_URL,
        )

    # 读取清单
    print(f"\n读取清单: {BILL_FILE}")
    items = read_bill_items(BILL_FILE, SHEET_NAME)
    print(f"  共 {len(items)} 条清单项")

    # 查询定额候选
    print(f"\n查询定额候选（按类别分组）:")
    category_quotas = get_category_quotas(QUOTA_DB)

    # 逐条匹配
    print(f"\n开始匹配...")
    results = {}
    total = len(items)

    for count, idx in enumerate(sorted(items.keys()), 1):
        item = items[idx]
        category = get_item_category(idx)

        if category is None:
            print(f"  [{count}/{total}] #{idx} {item['name']} - 未找到类别，跳过")
            continue

        quotas = category_quotas.get(category, [])
        if not quotas:
            print(f"  [{count}/{total}] #{idx} {item['name']} - {category}无候选定额，跳过")
            continue

        prompt = build_prompt(item, quotas)

        result = {"_idx": idx}

        # 调DeepSeek
        if ds_client:
            t0 = time.time()
            ds_text = call_llm(ds_client, config.DEEPSEEK_MODEL, prompt)
            ds_time = time.time() - t0
            result["deepseek"] = parse_response(ds_text)
            result["deepseek_time"] = ds_time
            ds_ids = result["deepseek"]["quota_ids"]
        else:
            result["deepseek"] = {"quota_ids": [], "reason": "未配置API Key", "confidence": 0}
            result["deepseek_time"] = 0
            ds_ids = []

        # 调Qwen
        if qw_client:
            t0 = time.time()
            qw_text = call_llm(qw_client, config.QWEN_MODEL, prompt)
            qw_time = time.time() - t0
            result["qwen"] = parse_response(qw_text)
            result["qwen_time"] = qw_time
            qw_ids = result["qwen"]["quota_ids"]
        else:
            result["qwen"] = {"quota_ids": [], "reason": "未配置API Key", "confidence": 0}
            result["qwen_time"] = 0
            qw_ids = []

        # 简单进度显示
        opus_ids = OPUS_BASELINE.get(idx, [])
        ds_cmp = compare_quota_ids(ds_ids, opus_ids) if ds_ids else "-"
        qw_cmp = compare_quota_ids(qw_ids, opus_ids) if qw_ids else "-"
        print(f"  [{count}/{total}] #{idx} {item['name'][:15]} | "
              f"DS:{','.join(ds_ids) if ds_ids else '-':15} ({ds_cmp}) | "
              f"QW:{','.join(qw_ids) if qw_ids else '-':15} ({qw_cmp})")

        results[idx] = result

    # 写入Excel
    print(f"\n写入结果: {OUTPUT_FILE}")
    write_output_excel(items, results, OUTPUT_FILE)

    # 打印总结
    print("\n" + "=" * 60)
    print("测试完成！")
    print("=" * 60)

    total_matched = len(results)
    if has_deepseek:
        ds_ok = sum(1 for r in results.values()
                    if "一致" in compare_quota_ids(
                        r.get("deepseek", {}).get("quota_ids", []),
                        OPUS_BASELINE.get(r["_idx"], [])))
        print(f"DeepSeek-V3: {ds_ok}/{total_matched} 与Opus一致 ({ds_ok*100/max(total_matched,1):.1f}%)")

    if has_qwen:
        qw_ok = sum(1 for r in results.values()
                    if "一致" in compare_quota_ids(
                        r.get("qwen", {}).get("quota_ids", []),
                        OPUS_BASELINE.get(r["_idx"], [])))
        print(f"Qwen-Plus:   {qw_ok}/{total_matched} 与Opus一致 ({qw_ok*100/max(total_matched,1):.1f}%)")

    print(f"\n详细结果请查看: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
