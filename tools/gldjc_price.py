"""
广材网材料价格查询工具 v2

改进内容（Phase 1）：
1. 名称拆解：把"热浸锌镀锌钢管 DN70"拆成品名+规格
2. 非材料项拦截：措施费、安装费等自动跳过
3. 结果过滤+打分：单位/规格/品名三维打分，低分不取价
4. 输出新增3列：匹配状态、置信度、广材网可点击链接
5. 搜索缓存去重：同名材料只搜一次
6. JSON主材库缓存：查完自动存，下次先查库

用法：
    python tools/gldjc_price.py "材料清单.xlsx" --cookie "token=bearer xxx"
    python tools/gldjc_price.py "材料清单.xlsx" --cookie "xxx" --no-cache  # 强制重查
"""

import re
import json
import sys
import time
import random
import argparse
import hashlib
from pathlib import Path
from datetime import datetime, timedelta
from urllib.parse import quote, urljoin

import requests
import openpyxl
from openpyxl.styles import PatternFill, Font
from lxml import html as lxml_html

# ========== 配置 ==========

# 广材网搜索页URL
SEARCH_URL = "https://www.gldjc.com/scj/so.html"

# 主材缓存文件路径
CACHE_FILE = Path(__file__).parent.parent / "data" / "material_prices.json"

# 缓存有效期（天）
CACHE_EXPIRE_DAYS = 30

# 请求头（模拟浏览器）
HEADERS = {
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8",
    "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
    "Accept-Encoding": "gzip, deflate, br",
    "Connection": "keep-alive",
    "Upgrade-Insecure-Requests": "1",
    "Sec-Fetch-Dest": "document",
    "Sec-Fetch-Mode": "navigate",
    "Sec-Fetch-Site": "same-origin",
    "Sec-Fetch-User": "?1",
    "Referer": "https://www.gldjc.com/",
}

# User-Agent池（随机选一个，避免固定UA被标记）
_UA_POOL = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0.0.0 Safari/537.36 Edg/130.0.0.0",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:133.0) Gecko/20100101 Firefox/133.0",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/129.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36 Edg/131.0.0.0",
]


def _get_headers() -> dict:
    """每次请求随机选一个UA，其余头不变"""
    h = dict(HEADERS)
    h["User-Agent"] = random.choice(_UA_POOL)
    return h

# 非材料项关键词（出现这些词的行自动跳过，不查价）
NON_MATERIAL_KEYWORDS = [
    "措施", "改造", "安装费", "调试", "拆除", "修复", "运输",
    "脚手架", "接入费", "制作", "保护措施", "试压", "冲洗",
    "检测", "验收", "税金", "利润", "管理费", "规费",
    "临时设施", "文明施工", "夜间施工", "冬雨季施工",
    "二次搬运", "已完工程保护", "大型机械进出场",
]

# 单位兼容组（同组内的单位视为兼容）
# 注意：t和kg绝对不能放一组！差1000倍，m和m²也不能放一组
UNIT_COMPAT_GROUPS = [
    {"kg", "千克", "公斤"},                    # 重量-千克级
    {"t", "吨"},                               # 重量-吨级（和kg不兼容！差1000倍）
    {"个", "只", "套", "台", "件", "组", "块"},# 计件（造价里个/台/套常互换）
    {"m", "米"},                               # 长度
    {"m²", "㎡", "平方米"},                    # 面积（和m不兼容！）
    {"m³", "立方米"},                          # 体积
    {"根", "条", "支"},                        # 条状物
    {"桶", "瓶"},                              # 容器
    {"卷", "盘"},                              # 卷状物
    {"副", "付", "对"},                        # 成对计量
]

_UNIT_ALIASES = {
    "吨": "t",
    "t": "t",
    "kg": "kg",
    "公斤": "kg",
    "千克": "kg",
    "米": "m",
    "m": "m",
    "平方米": "m2",
    "㎡": "m2",
    "m²": "m2",
    "立方米": "m3",
    "m³": "m3",
    "m3": "m3",
    "百米": "百米",
    "千米": "km",
    "km": "km",
    "条公里": "km",
}

# 镀锌钢管理论重量表（GB/T 3091，每米公斤数）
_PIPE_WEIGHT_PER_METER = {
    "DN15": 1.357, "DN20": 1.764, "DN25": 2.554,
    "DN32": 3.306, "DN40": 3.84,  "DN50": 5.33,
    "DN65": 7.09,  "DN70": 7.09,  "DN80": 8.47,
    "DN100": 12.15, "DN125": 15.04, "DN150": 19.26,
    "DN200": 30.97, "DN250": 42.56, "DN300": 54.90,
}
_GALVANIZED_PIPE_FACTOR = 1.06

# 名称中的修饰词（拆解时去掉，只保留核心品名）
# 注意：连接形式（法兰/螺纹/沟槽/承插/热熔/卡压/丝接）从这里拆出去了！
# 这些词决定价格差异，不能当噪声处理。
NOISE_PREFIXES = [
    # 镀锌/防腐类
    "热浸锌", "热浸镀锌", "热镀锌", "冷镀锌", "电镀锌",
    # 场所/位置
    "给水室外", "给水室内", "排水室外", "排水室内", "室外", "室内",
    "地下", "地上", "屋面", "楼层",
    # 品质/规格类
    "国标", "非标", "加厚", "普通", "优质", "标准", "一级", "二级",
    "A型", "B型", "C型", "I型", "II型",
    # 其他修饰
    "柔性", "刚性", "单壁", "双壁", "薄壁", "厚壁",
    "阻燃", "耐火", "低烟无卤",
]

# 连接形式（决定价格的关键字段，不当噪声删；参与打分）
CONNECTION_FORMS = [
    "丝接", "螺纹", "法兰", "卡压", "沟槽", "承插", "热熔", "电熔",
    "熔接", "焊接", "对焊", "承插焊", "卡箍",
]


def _extract_connection_form(text: str) -> str:
    """从名称/规格中识别连接形式，返回规范化名（为空表示未识别）。

    识别后用于：(1) 不从 base_name 里删掉；(2) 在打分时作硬约束/加分项。
    """
    if not text:
        return ""
    for form in CONNECTION_FORMS:
        if form in text:
            # 归一化同义连接形式
            if form in {"丝接", "螺纹"}:
                return "螺纹"
            if form in {"熔接", "热熔"}:
                return "热熔"
            if form in {"对焊", "承插焊", "焊接"}:
                return "焊接"
            if form in {"卡箍", "卡压"}:
                return "卡压"
            return form
    return ""

# 广材网品名映射（清单常见写法 → 广材网能搜到的关键词）
# 这张表解决"清单写A，广材网只认B"的问题
_GLDJC_NAME_MAP = {
    # 管材
    "衬塑PP-R钢管": "PPR管",
    "PP-R管": "PPR管",
    "PP-R给水管": "PPR给水管",
    "衬塑钢管": "衬塑钢管",
    "镀锌焊接钢管": "镀锌钢管",
    "镀锌无缝钢管": "镀锌钢管",
    "给水铸铁管": "铸铁管",
    "排水铸铁管": "铸铁排水管",
    "柔性铸铁管": "柔性铸铁排水管",
    "HDPE双壁波纹管": "HDPE波纹管",
    "HDPE排水管": "HDPE管",
    "UPVC排水管": "PVC排水管",
    "U-PVC排水管": "PVC排水管",
    "CPVC电力管": "CPVC管",
    "PE给水管": "PE管",
    "PE-RT地暖管": "PE-RT管",
    "铝塑复合管": "铝塑管",
    "薄壁不锈钢管": "不锈钢管",
    "焊接钢管": "焊接钢管",
    # 管件
    "弯头": "弯头",
    "三通": "三通",
    "异径管": "大小头",
    "变径": "大小头",
    # 阀门
    "截止阀": "截止阀",
    "闸阀": "闸阀",
    "蝶阀": "蝶阀",
    "止回阀": "止回阀",
    "球阀": "球阀",
    "减压阀": "减压阀",
    "平衡阀": "平衡阀",
    "过滤器": "过滤器",
    "倒流防止器": "倒流防止器",
    # 电气
    "电力电缆": "电力电缆",
    "控制电缆": "控制电缆",
    "BV电线": "BV线",
    "BV线": "BV线",
    "RVS双绞线": "RVS线",
    "金属线槽": "金属线槽",
    "桥架": "电缆桥架",
    "镀锌线管": "镀锌线管",
    "PVC线管": "PVC线管",
    "KBG管": "KBG管",
    "JDG管": "JDG管",
    "配电箱": "配电箱",
    "开关": "开关",
    "插座": "插座",
    # 消防
    "消防喷淋头": "喷淋头",
    "消火栓": "消火栓",
    "灭火器": "灭火器",
    "烟感探测器": "烟感",
    "温感探测器": "温感",
    # 暖通
    "风管": "镀锌风管",
    "镀锌钢板风管": "镀锌风管",
    "保温材料": "保温棉",
    "橡塑保温": "橡塑保温",
    "风口": "风口",
    "风阀": "风阀",
    # 五金/卫浴
    "蹲便器": "蹲便器",
    "坐便器": "坐便器",
    "洗脸盆": "洗脸盆",
    "水龙头": "水龙头",
    "地漏": "地漏",
    "角阀": "角阀",
}

# 规格提取正则（从名称中提取规格信息）
# 注意：顺序很重要，长模式放前面避免被短模式截断
SPEC_PATTERNS = [
    r'DN\d+',                  # 管径 DN25, DN100
    r'De\d+',                  # 外径 De25
    r'Φ\d+(?:mm)?',            # 直径 Φ10, Φ10mm
    r'\d+[×x\*]\d+(?:\.\d+)?(?:mm²?)?',  # 尺寸 100×100, 4x1.5, 2*1.5mm²（支持小数）
    r'\d+(?:\.\d+)?mm²',       # 截面 2.5mm², 50mm²（支持小数）
    r'\d+kV[A]?',              # 电压/容量 10kV, 630kVA
    r'\d+[AW]H?',              # 电流/功率/容量 100A, 500W, 24AH
    r'\d+V',                   # 电压 12V, 220V
    r'Q[0-9]+[A-Z]*',          # 牌号 Q235B
    r'HRB\d+',                 # 钢筋牌号 HRB400
]


# ========== 同义词加载 ==========

_engineering_synonyms: dict | None = None  # 懒加载缓存


def _load_engineering_synonyms() -> dict:
    """加载Jarvis工程同义词表（718条），懒加载只读一次"""
    global _engineering_synonyms
    if _engineering_synonyms is not None:
        return _engineering_synonyms
    syn_path = Path(__file__).parent.parent / "data" / "engineering_synonyms.json"
    if syn_path.exists():
        try:
            _engineering_synonyms = json.load(open(syn_path, encoding="utf-8"))
        except Exception:
            _engineering_synonyms = {}
    else:
        _engineering_synonyms = {}
    return _engineering_synonyms


def _get_gldjc_name(base_name: str, original: str) -> str | None:
    """从广材网品名映射表找到对应的广材网常用名

    先精确匹配base_name，再匹配original，再做包含匹配
    """
    # 精确匹配
    if base_name in _GLDJC_NAME_MAP:
        return _GLDJC_NAME_MAP[base_name]
    if original in _GLDJC_NAME_MAP:
        return _GLDJC_NAME_MAP[original]
    # 包含匹配（清单名包含映射表的key）
    for key, val in _GLDJC_NAME_MAP.items():
        if key in base_name or key in original:
            return val
    return None


def _get_synonym(base_name: str) -> str | None:
    """从Jarvis工程同义词表查一个别名

    同义词表格式: {"清单写法": ["定额写法1", ...]}
    这里取第一个定额写法作为搜索别名
    """
    syns = _load_engineering_synonyms()
    if not syns:
        return None
    # 精确匹配
    if base_name in syns:
        vals = syns[base_name]
        return vals[0] if vals else None
    # 包含匹配（清单名包含同义词key的核心部分）
    for key, vals in syns.items():
        if len(key) >= 3 and key in base_name and vals:
            return vals[0]
    return None


# ========== 名称拆解 ==========

def parse_material(name: str, spec_col: str = "") -> dict:
    """
    把材料名称拆成：基础品名 + 规格列表 + 搜索关键词

    例：
    "热浸锌镀锌钢管 DN70" → base_name="镀锌钢管", specs=["DN70"]
    "HDPE 三通 DN100"     → base_name="HDPE三通", specs=["DN100"]
    "T2纯紫铜 Φ10mm"      → base_name="紫铜", specs=["Φ10mm"]
    """
    original = name.strip()

    # 1. 从名称中提取规格
    specs = []
    clean_name = original
    for pattern in SPEC_PATTERNS:
        found = re.findall(pattern, clean_name, re.IGNORECASE)
        specs.extend(found)
        clean_name = re.sub(pattern, '', clean_name, flags=re.IGNORECASE)

    # 2. 从规格型号列也提取（如果有的话）
    if spec_col:
        for pattern in SPEC_PATTERNS:
            found = re.findall(pattern, spec_col, re.IGNORECASE)
            specs.extend(found)

    # 去重保序
    seen = set()
    unique_specs = []
    for s in specs:
        s_upper = s.upper()
        if s_upper not in seen:
            seen.add(s_upper)
            unique_specs.append(s)
    specs = unique_specs

    # 3. 去掉修饰词
    base_name = clean_name.strip()
    for noise in NOISE_PREFIXES:
        base_name = base_name.replace(noise, "")
    # 清理规格提取后残留的碎片（如"mm"、"-"、"."等）
    base_name = re.sub(r'[\s\-\.]+$', '', base_name)  # 去末尾残留符号
    base_name = re.sub(r'^[\s\-\.]+', '', base_name)  # 去开头残留符号
    base_name = re.sub(r'\s+', '', base_name).strip()  # 去多余空格

    # 4. 构建搜索关键词（分层降级，5层策略）
    # 第1优先：品名+核心规格（最精确）
    # 第2优先：纯品名（去掉规格的）
    # 第3优先：广材网品名映射+规格（清单写法→广材网写法）
    # 第4优先：广材网品名映射（纯映射名）
    # 第5优先：工程同义词+规格
    search_keywords = []

    if specs:
        search_keywords.append(f"{base_name} {specs[0]}")
    search_keywords.append(base_name)

    # 广材网品名映射：检查原名/品名是否有对应的广材网常用名
    mapped_name = _get_gldjc_name(base_name, original)
    if mapped_name and mapped_name != base_name:
        if specs:
            search_keywords.append(f"{mapped_name} {specs[0]}")
        search_keywords.append(mapped_name)

    # 工程同义词：从Jarvis的同义词表查一个别名
    synonym = _get_synonym(base_name)
    if synonym and synonym != base_name and synonym != mapped_name:
        if specs:
            search_keywords.append(f"{synonym} {specs[0]}")
        search_keywords.append(synonym)

    # 最后兜底：如果品名和原始名差异大，也加原始名
    if base_name != original and len(base_name) >= 2:
        search_keywords.append(original)

    # 去重保序
    seen_kw: set[str] = set()
    unique_kw: list[str] = []
    for kw in search_keywords:
        if kw not in seen_kw:
            seen_kw.add(kw)
            unique_kw.append(kw)
    search_keywords = unique_kw

    return {
        "original": original,
        "base_name": base_name,
        "specs": specs,
        "search_keywords": search_keywords,
    }


def is_non_material(name: str) -> bool:
    """
    判断是否为非材料项（措施费、安装费等不需要查价的项目）

    返回 True 表示应该跳过
    """
    for kw in NON_MATERIAL_KEYWORDS:
        if kw in name:
            return True
    return False


def _extract_dn(text: str) -> str | None:
    if not text:
        return None
    matched = re.search(r"[Dd][Nn]\s*(\d+)", text)
    if matched:
        return f"DN{matched.group(1)}"
    return None


def _normalize_unit(unit: str) -> str:
    raw = str(unit or "").strip().lower()
    if not raw:
        return ""
    raw = (
        raw.replace("／", "/")
        .replace(" ", "")
        .replace("（", "(")
        .replace("）", ")")
    )
    return _UNIT_ALIASES.get(raw, raw)


def _is_supported_steel_pipe(name: str, spec: str) -> bool:
    text = f"{name} {spec}"
    return (
        "钢管" in text
        or _extract_dn(text) is not None
        or re.search(r'(?:Φ|φ)?\s*\d+(?:\.\d+)?\s*[×xX\*]\s*\d+(?:\.\d+)?', text) is not None
    )


def _extract_pipe_outer_diameter_and_thickness(name: str, spec: str) -> tuple[float | None, float | None]:
    text = f"{name} {spec}"
    matched = re.search(
        r'(?:外径\s*)?(?:Φ|φ)?\s*(\d+(?:\.\d+)?)\s*(?:mm)?\s*[×xX\*]\s*(\d+(?:\.\d+)?)',
        text,
        flags=re.IGNORECASE,
    )
    if matched:
        return float(matched.group(1)), float(matched.group(2))
    return None, None


def _estimate_pipe_weight_kg_per_m(name: str, spec: str) -> float | None:
    text = f"{name} {spec}"
    if not _is_supported_steel_pipe(name, spec):
        return None

    outer_diameter, thickness = _extract_pipe_outer_diameter_and_thickness(name, spec)
    factor = _GALVANIZED_PIPE_FACTOR if "镀锌" in text else 1.0

    if outer_diameter and thickness and outer_diameter > thickness > 0:
        base_weight = (outer_diameter - thickness) * thickness * 0.02466
        return round(base_weight * factor, 4)

    dn = _extract_dn(text)
    if dn:
        if "无缝" in text:
            return None
        return _PIPE_WEIGHT_PER_METER.get(dn)

    return None


def _convert_ton_to_meter(ton_price: float, name: str, spec: str) -> float | None:
    weight = _estimate_pipe_weight_kg_per_m(name, spec)
    if not weight:
        return None
    return round(ton_price * weight / 1000, 2)


def _try_convert_price(price: float, from_unit: str, to_unit: str,
                       name: str = "", spec: str = "") -> float | None:
    fu = _normalize_unit(from_unit)
    tu = _normalize_unit(to_unit)

    if not fu or not tu:
        return None
    if fu == tu:
        return round(price, 2)
    if fu == "t" and tu == "m":
        return _convert_ton_to_meter(price, name, spec)
    if fu == "t" and tu == "kg":
        return round(price / 1000, 2)
    if fu == "百米" and tu == "m":
        return round(price / 100, 2)
    if fu == "km" and tu == "m":
        return round(price / 1000, 2)
    return None


def _normalize_material_hint(text: str) -> str:
    return re.sub(r"[\s\-\(\)（）/]", "", str(text or "")).lower()


def _detect_material_family(text: str) -> str:
    value = str(text or "").strip()
    if not value:
        return ""

    fitting_tokens = ("管件", "弯头", "三通", "四通", "异径", "大小头", "法兰", "接头", "补偿器", "伸缩节", "止流器", "短管")
    valve_tokens = ("阀", "过滤器", "减压")
    device_tokens = ("地漏", "洁具", "器具", "洗脸盆", "蹲便器", "坐便器")

    if any(token in value for token in fitting_tokens):
        return "fitting"
    if any(token in value for token in valve_tokens):
        return "valve"
    if any(token in value for token in device_tokens):
        return "device"
    if any(token in value for token in ("管", "管材")):
        return "pipe"
    return ""


def _looks_like_installation_item_name(text: str) -> bool:
    value = str(text or "").strip()
    if not value:
        return False
    install_tokens = ("安装", "敷设", "铺设", "组装", "调试", "管线", "组成")
    if not any(token in value for token in install_tokens):
        return False
    if _detect_material_family(value):
        return False
    if _extract_material_tokens(value):
        return False
    return True


def _extract_material_tokens(text: str) -> set[str]:
    """从名称/规格文本中识别材料 token 集合。

    注意：按"长度优先"匹配，命中后从 raw 中消耗该区段，避免
      "UPVC 排水管" 同时命中 UPVC 和 PVC 导致它被判为与普通 PVC 兼容。
    """
    raw = str(text or "").upper()
    # 按长度降序排列，长 token 优先；同组同义词共用一个归一化 key
    token_map = [
        # (原始 token, 归一化 key)
        ("球墨铸铁", "球墨铸铁"),
        ("钢塑复合", "钢塑复合"),
        ("PVC-U", "UPVC"),
        ("U-PVC", "UPVC"),
        ("UPVC", "UPVC"),
        ("PVC-C", "CPVC"),
        ("CPVC", "CPVC"),
        ("PE-RT", "PERT"),
        ("PERT", "PERT"),
        ("HDPE", "HDPE"),
        ("PP-R", "PPR"),
        ("PPR", "PPR"),
        ("不锈钢", "不锈钢"),
        ("镀锌", "镀锌"),
        ("铸铁", "铸铁"),
        ("衬塑", "衬塑"),
        ("涂塑", "涂塑"),
        ("黄铜", "黄铜"),
        ("无缝", "无缝"),
        ("焊接", "焊接"),
        ("PSP", "PSP"),
        ("PVC", "PVC"),  # 注意：放在 UPVC/CPVC/PVC-U/PVC-C 之后，避免误吞
        ("PE", "PE"),    # 放在 HDPE/PE-RT 之后
        ("铜", "铜"),
    ]
    found: set[str] = set()
    consumed = raw
    for token, key in token_map:
        if token in consumed:
            found.add(key)
            # 从 consumed 里抹掉已匹配的 token，避免后续短 token 重复命中
            consumed = consumed.replace(token, " " * len(token))

    # UPVC/CPVC 被记为单独的族，不应再留 "PVC" 作为同族（已在消耗步骤中避免）
    return found


def _extract_semantic_keywords(text: str) -> set[str]:
    raw = str(text or "")
    keyword_groups = (
        "热熔",
        "电熔",
        "承插",
        "法兰",
        "沟槽",
        "螺纹",
        "给水",
        "排水",
        "雨水",
        "三通",
        "四通",
        "弯头",
        "直接",
        "直通",
        "异径",
        "大小头",
        "套管",
        "地漏",
        "雨水斗",
        "闸阀",
        "蝶阀",
        "球阀",
        "止回阀",
        "截止阀",
        "过滤器",
        "管件",
    )
    return {token for token in keyword_groups if token in raw}


def _infer_pipe_material_signature(name: str) -> str:
    text = str(name or "").strip().upper()
    if not text:
        return ""
    if any(token in text for token in ("衬塑钢管", "钢塑复合", "PSP", "涂塑钢管", "涂覆钢管", "涂覆碳钢管", "衬塑", "涂塑", "内衬塑")):
        return "composite_steel_pipe"
    if "不锈钢" in text and "管" in text:
        return "stainless_steel_pipe"
    if "镀锌" in text:
        return "galvanized_steel_pipe"
    if "无缝" in text:
        return "seamless_steel_pipe"
    if "焊接" in text:
        return "welded_steel_pipe"
    if "钢管" in text:
        return "steel_pipe"
    return ""


def _pipe_signature_compatible(request_name: str, candidate_name: str) -> bool | None:
    request_signature = _infer_pipe_material_signature(request_name)
    candidate_signature = _infer_pipe_material_signature(candidate_name)
    if not request_signature or not candidate_signature:
        return None
    if request_signature == candidate_signature:
        return True
    if request_signature == "steel_pipe" and candidate_signature in {"galvanized_steel_pipe", "welded_steel_pipe", "seamless_steel_pipe"}:
        return True
    if candidate_signature == "steel_pipe" and request_signature in {"galvanized_steel_pipe", "welded_steel_pipe", "seamless_steel_pipe"}:
        return True
    return False


def _has_conflicting_scene(target_text: str, candidate_text: str) -> bool:
    target = str(target_text or "")
    candidate = str(candidate_text or "")
    scene_pairs = (
        ("给水", "排水"),
        ("排水", "给水"),
        ("雨水", "给水"),
        ("给水", "雨水"),
    )
    return any(left in target and right in candidate for left, right in scene_pairs)


def _has_accessory_tokens(candidate_text: str) -> bool:
    candidate = str(candidate_text or "")
    accessory_tokens = (
        "管卡", "卡箍", "吊卡", "支架", "托架", "套管", "胶水",
        "螺栓", "螺母", "垫片", "抱箍", "管夹", "码钉",
    )
    return any(token in candidate for token in accessory_tokens)


def _is_compatible_result(base_name: str, result_spec: str) -> bool:
    candidate = str(result_spec or "").strip()
    if not candidate:
        return False

    if _looks_like_installation_item_name(base_name):
        return False

    target_family = _detect_material_family(base_name)
    candidate_family = _detect_material_family(candidate)
    if target_family == "pipe" and candidate_family != "pipe":
        return False
    if target_family == "fitting" and candidate_family != "fitting":
        return False
    if target_family == "valve" and candidate_family != "valve":
        return False
    if target_family == "device" and candidate_family != "device":
        return False
    if target_family and candidate_family and target_family != candidate_family:
        return False

    if target_family == "pipe" and "管件" in candidate:
        return False
    if target_family in {"pipe", "fitting"} and _has_accessory_tokens(candidate):
        return False

    if "热熔管件" in str(base_name or "") and any(token in candidate for token in ("法兰管卡", "U型管卡", "管卡")):
        return False

    pipe_signature_ok = _pipe_signature_compatible(base_name, candidate)
    if pipe_signature_ok is False:
        return False

    target_tokens = _extract_material_tokens(base_name)
    candidate_tokens = _extract_material_tokens(candidate)
    if target_tokens and candidate_tokens and not (target_tokens & candidate_tokens):
        if pipe_signature_ok is not True:
            return False

    if _has_conflicting_scene(base_name, candidate):
        return False

    return True


# ========== 结果过滤与打分 ==========

def check_unit_compatible(unit_a: str, unit_b: str) -> bool:
    """检查两个单位是否兼容（在同一个兼容组内）"""
    if not unit_a or not unit_b:
        return False
    a = unit_a.strip().lower()
    b = unit_b.strip().lower()
    if a == b:
        return True
    for group in UNIT_COMPAT_GROUPS:
        lower_group = {u.lower() for u in group}
        if a in lower_group and b in lower_group:
            return True
    return False


def _normalize_spec_text(text: str) -> str:
    """归一化规格文本，统一尺寸分隔符和直径符号。

    ×/x/X/* → X，φ/Φ → PHI；空白全部去掉、字母转大写。
    这样 "DN25×2.75"、"DN25x2.75"、"DN25*2.75" 会被视为同一规格。
    """
    value = re.sub(r"\s+", "", str(text or "").upper())
    if not value:
        return ""
    value = value.replace("×", "X").replace("*", "X")
    value = value.replace("Φ", "PHI").replace("φ", "PHI")
    return value


def _spec_contains_exact(result_spec: str, target_spec: str) -> bool:
    result_text = _normalize_spec_text(result_spec)
    target_text = _normalize_spec_text(target_spec)
    if not result_text or not target_text:
        return False

    pattern = re.escape(target_text)
    if target_text[0].isdigit():
        pattern = rf"(?<!\d){pattern}"
    if target_text[-1].isdigit():
        pattern = rf"{pattern}(?!\d)"

    if re.search(pattern, result_text) is not None:
        return True

    # DN/De 同口径兼容，并支持“DN(mm):50 / 公称直径DN(mm):50”这类广材网写法
    matched = re.fullmatch(r"(DN|DE)(\d+(?:\.\d+)?)", target_text)
    if matched:
        prefix = matched.group(1)
        size = matched.group(2)
        alt_prefix = "DE" if prefix == "DN" else "DN"
        loose_patterns = [
            rf"{prefix}\D*{re.escape(size)}(?!\d)",
            rf"{alt_prefix}\D*{re.escape(size)}(?!\d)",
        ]
        for loose_pattern in loose_patterns:
            if re.search(loose_pattern, result_text) is not None:
                return True

    return False


def check_spec_match(result_spec: str, target_specs: list[str]) -> bool:
    """检查搜索结果的规格描述是否包含目标规格"""
    if not target_specs:
        return True  # 没有目标规格，不做过滤
    for ts in target_specs:
        if _spec_contains_exact(result_spec, ts):
            return True
    return False


def _extract_spec_numeric_values(text: str) -> list[float]:
    source = _normalize_spec_text(text)
    values: list[float] = []
    for matched in re.findall(r"(?:DN|DE|Φ|φ)\D*(\d+(?:\.\d+)?)", source, flags=re.IGNORECASE):
        try:
            values.append(float(matched))
        except ValueError:
            continue
    return values


def _score_relaxed_spec_match(result_spec: str, target_specs: list[str]) -> float | None:
    if not target_specs:
        return 0

    result_values = _extract_spec_numeric_values(result_spec)
    if not result_values:
        return 0

    best_score: float | None = None
    for target_spec in target_specs:
        target_values = _extract_spec_numeric_values(target_spec)
        if not target_values:
            continue
        for target_value in target_values:
            for result_value in result_values:
                diff = abs(result_value - target_value)
                ratio = diff / max(target_value, 1.0)
                candidate_score: float | None
                if diff == 0:
                    candidate_score = 25
                elif diff <= 1 or ratio <= 0.03:
                    candidate_score = 15
                elif diff <= 3 or ratio <= 0.08:
                    candidate_score = 8
                elif diff <= 5 and ratio <= 0.15:
                    candidate_score = 3
                else:
                    candidate_score = None

                if candidate_score is not None and (best_score is None or candidate_score > best_score):
                    best_score = candidate_score

    if best_score is not None:
        return best_score
    return None


def filter_and_score(results: list[dict], target_unit: str,
                     target_specs: list[str], base_name: str,
                     request_name: str = "", request_spec: str = "",
                     allow_relaxed_spec: bool = False) -> list[dict]:
    """
    对广材网搜索结果进行过滤和打分

    打分规则：
    - 单位兼容：+40分（最重要，单位不对差几个数量级）
    - 规格命中：+30分（规格不对差几倍到几十倍）
    - 品名关键词命中：+20分
    - 基础分：+10分（搜到了就有基础分）

    返回：打分后的结果列表（只保留>=40分的），按分数降序
    """
    scored = []
    request_name = request_name or base_name
    request_spec = request_spec or " ".join(target_specs)

    for original_result in results:
        r = dict(original_result)
        score = 10  # 基础分
        unit_match = False
        result_spec = r.get("spec", "")
        spec_match = check_spec_match(result_spec, target_specs)
        raw_unit = str(r.get("unit") or "").strip()
        raw_price = r.get("market_price")

        try:
            numeric_price = round(float(raw_price), 2)
        except (TypeError, ValueError):
            continue

        if not _is_compatible_result(base_name, result_spec):
            continue

        # 单位打分（最关键）
        if target_unit and raw_unit:
            if check_unit_compatible(target_unit, raw_unit):
                score += 40
                unit_match = True
                r["_raw_unit"] = raw_unit
                r["_raw_market_price"] = numeric_price
            else:
                converted_price = _try_convert_price(
                    numeric_price,
                    raw_unit,
                    target_unit,
                    request_name,
                    request_spec,
                )
                if converted_price is None:
                    # 单位明确不兼容，直接淘汰，避免数量级错误
                    continue
                score += 35
                unit_match = True
                r["_raw_unit"] = raw_unit
                r["_raw_market_price"] = numeric_price
                r["_converted_from_unit"] = raw_unit
                r["unit"] = target_unit
                r["market_price"] = converted_price

        # 有明确目标规格时，优先精确命中；拿不到时允许退到可解释的近似口径
        if target_specs and not spec_match:
            if not allow_relaxed_spec:
                continue
            relaxed_spec_score = _score_relaxed_spec_match(result_spec, target_specs)
            if relaxed_spec_score is None:
                continue
            score += relaxed_spec_score
            r["_relaxed_spec_match"] = True
        elif spec_match:
            score += 30

        # 品名关键词打分
        if base_name and len(base_name) >= 2:
            spec_text = result_spec
            # 检查品名关键字是否出现在结果的规格描述中
            name_chars = set(base_name)
            match_count = sum(1 for c in name_chars if c in spec_text)
            if match_count >= len(name_chars) * 0.5:
                score += 20

        target_material_tokens = _extract_material_tokens(base_name)
        result_material_tokens = _extract_material_tokens(result_spec)
        if target_material_tokens and result_material_tokens and (target_material_tokens & result_material_tokens):
            score += 20

        # 连接形式打分：法兰/螺纹/沟槽/承插/热熔/卡压决定价差，必须当硬约束
        # 参考源：request_name + 原始请求名 + 请求规格，候选源：result_spec
        target_conn = _extract_connection_form(f"{request_name} {base_name} {request_spec}")
        candidate_conn = _extract_connection_form(result_spec)
        if target_conn and candidate_conn:
            if target_conn == candidate_conn:
                score += 15
                r["_conn_match"] = True
            else:
                # 连接形式明确冲突 → 减 25 分，阀门/管件族会直接掉到 40 分门槛以下
                score -= 25
                r["_conn_mismatch"] = (target_conn, candidate_conn)

        if "_raw_market_price" not in r:
            r["_raw_market_price"] = numeric_price
        if "_raw_unit" not in r:
            r["_raw_unit"] = raw_unit
        if r.get("market_price") in (None, ""):
            r["market_price"] = numeric_price

        r["score"] = score
        r["_unit_match"] = unit_match
        r["_spec_match"] = spec_match
        scored.append(r)

    # 只保留>=40分的（至少单位要兼容）
    filtered = [r for r in scored if r["score"] >= 40]
    filtered.sort(key=lambda x: x["score"], reverse=True)
    return filtered


def build_approximate_price_candidates(results: list[dict], target_unit: str,
                                       request_name: str = "", request_spec: str = "") -> list[dict]:
    """Build looser fallback candidates when strict filtering returns nothing.

    This path still requires unit compatibility or a supported unit conversion,
    but it no longer rejects results only because the material subtype/spec text
    is imperfect. The caller should mark the output as an approximate price.
    """
    candidates = []
    request_family = _detect_material_family(request_name)
    request_tokens = _extract_material_tokens(request_name)
    request_keywords = _extract_semantic_keywords(request_name)

    for original_result in results:
        r = dict(original_result)
        raw_unit = str(r.get("unit") or "").strip()
        raw_price = r.get("market_price")
        result_spec = str(r.get("spec") or "").strip()
        candidate_family = _detect_material_family(result_spec)
        result_tokens = _extract_material_tokens(result_spec)
        result_keywords = _extract_semantic_keywords(result_spec)
        exact_spec_match = bool(request_spec and _spec_contains_exact(result_spec, request_spec))
        relaxed_spec_score = _score_relaxed_spec_match(result_spec, [request_spec]) if request_spec else None
        shared_material_tokens = request_tokens & result_tokens
        shared_keywords = request_keywords & result_keywords

        try:
            numeric_price = round(float(raw_price), 2)
        except (TypeError, ValueError):
            continue

        if request_family and candidate_family and request_family != candidate_family:
            continue
        if request_tokens and result_tokens and not shared_material_tokens:
            continue
        if request_family in {"fitting", "valve", "device"} and request_keywords and not shared_keywords:
            continue
        if request_spec and not exact_spec_match and relaxed_spec_score is None:
            continue
        if not shared_material_tokens and not shared_keywords and not exact_spec_match and relaxed_spec_score is None:
            continue

        score = 10.0
        if target_unit:
            if raw_unit and check_unit_compatible(target_unit, raw_unit):
                score += 40
                r["_unit_match"] = True
            else:
                converted_price = _try_convert_price(
                    numeric_price,
                    raw_unit,
                    target_unit,
                    request_name,
                    request_spec,
                )
                if converted_price is None:
                    continue
                score += 35
                r["_unit_match"] = True
                r["_converted_from_unit"] = raw_unit
                r["unit"] = target_unit
                r["market_price"] = converted_price

        if "_raw_market_price" not in r:
            r["_raw_market_price"] = numeric_price
        if "_raw_unit" not in r:
            r["_raw_unit"] = raw_unit
        if r.get("market_price") in (None, ""):
            r["market_price"] = numeric_price

        if exact_spec_match:
            score += 20
        elif relaxed_spec_score is not None:
            score += relaxed_spec_score

        if shared_material_tokens:
            score += min(20, 8 * len(shared_material_tokens))

        if shared_keywords:
            score += min(20, 8 * len(shared_keywords))

        if request_family and request_family == candidate_family:
            score += 10

        r["score"] = score
        r["_approximate_match"] = True
        candidates.append(r)

    candidates.sort(key=lambda item: float(item.get("score", 0) or 0), reverse=True)
    return candidates


def determine_confidence(scored_results: list[dict], target_unit: str,
                         target_specs: list[str]) -> str:
    """
    根据过滤后的结果判断置信度

    高：单位一致 + 规格命中 + 多条报价
    中：单位一致但规格未命中，或只有1-2条报价
    低：其他情况
    """
    if not scored_results:
        return "低"

    top = scored_results[0]
    count = len(scored_results)

    if target_unit and not top.get("_unit_match"):
        return "低"
    if target_specs and not top.get("_spec_match") and not top.get("_relaxed_spec_match"):
        return "低"

    # 连接形式冲突或规格只是松弛命中 → 置信度封顶"中"
    relaxed_or_conflict = top.get("_relaxed_spec_match") or top.get("_conn_mismatch")

    if not relaxed_or_conflict and top["score"] >= 80 and count >= 3:
        return "高"
    elif top["score"] >= 50 and count >= 2:
        return "中"
    else:
        return "低"


# ========== 广材网搜索 ==========

def search_material_web(session: requests.Session, keyword: str,
                        province_code: str = "1") -> list[dict]:
    """
    在广材网搜索材料，从SSR渲染的HTML中提取价格数据

    返回: [{"spec": "镀锌焊接钢管|DN25", "brand": "贵盈",
            "market_price": 3663.72, "unit": "t"}, ...]
    """
    params = {"keyword": keyword, "l": province_code}

    try:
        resp = session.get(SEARCH_URL, params=params, headers=_get_headers(), timeout=30)
        resp.raise_for_status()
    except Exception as e:
        print(f"  [错误] 搜索'{keyword}'失败: {e}")
        return []

    html = resp.text

    # 检测是否被踢出登录（Cookie失效）
    if "请登录" in html or "login" in html.lower() and "token" not in html.lower():
        print("\n  [警告] Cookie可能已失效，建议更换Cookie后重试")

    def _clean_text(value: str) -> str:
        return re.sub(r"\s+", " ", str(value or "")).strip()

    def _extract_first_text(node, xpath_expr: str) -> str:
        texts = node.xpath(xpath_expr)
        for text in texts:
            cleaned = _clean_text(text if isinstance(text, str) else getattr(text, "text_content", lambda: "")())
            if cleaned:
                return cleaned
        return ""

    def _extract_unit(card_node) -> str:
        unit_texts = card_node.xpath(
            ".//*[contains(@class,'colspan-cell') and contains(@class,'pure-text')]//text()"
        )
        for raw in unit_texts:
            candidate = _clean_text(raw)
            if not candidate:
                continue
            if re.fullmatch(r"[A-Za-z㎡m³m²tT个件台套只块组卷盘付副对根条支桶瓶千克公斤吨百米千米公里]+", candidate):
                return candidate
        return ""

    def _extract_detail_url(card_node) -> str:
        raw_candidates = card_node.xpath(
            "self::a[@href]/@href"
            " | .//a[@href]/@href"
            " | self::*[@data-href]/@data-href"
            " | .//*[@data-href]/@data-href"
            " | self::*[@data-url]/@data-url"
            " | .//*[@data-url]/@data-url"
        )
        normalized: list[str] = []
        for raw in raw_candidates:
            href = _clean_text(raw)
            if not href or href in {"#", "/"}:
                continue
            if href.lower().startswith("javascript:"):
                continue
            normalized.append(urljoin("https://www.gldjc.com/", href))

        if not normalized:
            return ""

        for href in normalized:
            if any(flag in href for flag in ("/info/", "/xunjia/", "/scj/", "detail")):
                return href
        return normalized[0]

    def _find_price_card(price_node):
        current = price_node
        best = None
        while current is not None:
            try:
                has_spec = bool(current.xpath(".//*[contains(@class,'m-detail-content')]"))
            except Exception:
                has_spec = False
            if has_spec:
                best = current
                has_price_block = bool(current.xpath(".//*[contains(@class,'price-block')]"))
                if has_price_block:
                    return current
            current = current.getparent()
        return best

    try:
        tree = lxml_html.fromstring(html)
        price_nodes = tree.xpath("//span[contains(@class,'change-point')]")
        results = []
        for price_node in price_nodes:
            price_text = _clean_text(price_node.text_content())
            if not re.fullmatch(r"\d+(?:\.\d+)?", price_text):
                continue
            try:
                price_val = float(price_text)
            except ValueError:
                continue

            card = _find_price_card(price_node)
            if card is None:
                continue

            spec = _extract_first_text(card, ".//*[contains(@class,'m-detail-content')]//text()")
            if not spec:
                continue

            brand = _extract_first_text(card, ".//*[contains(@class,'brand-box')]//text()")
            unit = _extract_unit(card)
            detail_url = _extract_detail_url(card)

            results.append({
                "keyword": keyword,
                "spec": spec,
                "brand": brand,
                "unit": unit,
                "market_price": price_val,
                "detail_url": detail_url,
            })

        if results:
            return results
    except Exception:
        pass

    # 兜底：保留旧正则逻辑，防止页面结构轻微变化时完全失效
    specs_raw = re.findall(r'class="m-detail-content"[^>]*>(.*?)</div>', html, re.DOTALL)
    specs = [re.sub(r'<[^>]+>', '', s).strip() for s in specs_raw]
    brands = re.findall(r'class="brand-box"[^>]*>\s*(\S+)\s*<div', html, re.DOTALL)
    price_matches = re.findall(
        r'class="price-block"[^>]*><span[^>]*class="change-point"[^>]*>([0-9.]+)</span>', html
    )
    units = re.findall(r'class="colspan-cell width-56 pure-text"[^>]*>\s*([^<\s]+)\s*</div>', html)

    results = []
    for i in range(len(price_matches)):
        try:
            price_val = float(price_matches[i])
        except ValueError:
            continue
        results.append({
            "keyword": keyword,
            "spec": specs[i] if i < len(specs) else "",
            "brand": brands[i] if i < len(brands) else "",
            "unit": units[i] if i < len(units) else "",
            "market_price": price_val,
        })

    return results


def get_median_price(prices: list[dict], price_field: str = "market_price") -> float | None:
    """取真实中位价对应的数值，不做均值计算。"""
    valid_prices = [p[price_field] for p in prices if p.get(price_field)]
    if not valid_prices:
        return None

    valid_prices.sort()
    return float(valid_prices[(len(valid_prices) - 1) // 2])


def get_top_price_result(prices: list[dict], price_field: str = "market_price") -> dict | None:
    """取排序后的首条结果，保证价格和搜索页可直接对应。"""
    if not prices:
        return None
    top = prices[0]
    if not top.get(price_field):
        return None
    return top


def _build_price_clusters(prices: list[dict], price_field: str = "market_price") -> list[list[dict]]:
    priced_items: list[tuple[float, dict]] = []
    for item in prices:
        raw_price = item.get(price_field)
        if raw_price in (None, ""):
            continue
        try:
            price_val = float(raw_price)
        except (TypeError, ValueError):
            continue
        priced_items.append((price_val, item))

    if not priced_items:
        return []

    priced_items.sort(key=lambda pair: pair[0])
    clusters: list[list[dict]] = []
    current_cluster: list[dict] = [priced_items[0][1]]
    current_prices: list[float] = [priced_items[0][0]]

    for price_val, item in priced_items[1:]:
        cluster_max = max(current_prices)
        cluster_min = min(current_prices)
        price_span_ratio = (price_val - cluster_max) / max(cluster_min, 1.0)
        price_gap = price_val - cluster_max

        if price_gap <= 5 or price_span_ratio <= 0.15:
            current_cluster.append(item)
            current_prices.append(price_val)
            continue

        clusters.append(current_cluster)
        current_cluster = [item]
        current_prices = [price_val]

    clusters.append(current_cluster)
    return clusters


def get_representative_price_result(prices: list[dict], price_field: str = "market_price") -> dict | None:
    """优先取直接报价中的低位密集合理簇，避免高价/换算价把结果带偏。"""
    if not prices:
        return None

    direct_prices = [item for item in prices if not item.get("_converted_from_unit")]
    pool = direct_prices or prices

    clusters = _build_price_clusters(pool, price_field=price_field)
    if not clusters:
        return get_top_price_result(pool, price_field=price_field)

    def _cluster_sort_key(cluster: list[dict]) -> tuple:
        cluster_prices = []
        cluster_score = 0.0
        for item in cluster:
            try:
                cluster_prices.append(float(item.get(price_field)))
            except (TypeError, ValueError):
                continue
            cluster_score = max(cluster_score, float(item.get("score", 0) or 0))
        min_price = min(cluster_prices) if cluster_prices else float("inf")
        return (-len(cluster), -cluster_score, min_price)

    best_cluster = sorted(clusters, key=_cluster_sort_key)[0]

    best = None
    best_key = None
    for item in best_cluster:
        raw_price = item.get(price_field)
        if raw_price in (None, ""):
            continue
        try:
            price_val = float(raw_price)
        except (TypeError, ValueError):
            continue
        key = (
            price_val,
            -float(item.get("score", 0) or 0),
        )
        if best_key is None or key < best_key:
            best = item
            best_key = key

    return best or get_top_price_result(pool, price_field=price_field)


def build_match_label(result: dict | None, fallback_text: str = "查看") -> str:
    if not isinstance(result, dict):
        return fallback_text

    brand = str(result.get("brand") or "").strip()
    spec = str(result.get("spec") or "").strip()
    unit = str(result.get("unit") or "").strip()
    price = result.get("market_price")

    parts = []
    if brand:
        parts.append(brand)
    if spec:
        parts.append(spec[:40])
    if unit:
        if result.get("_converted_from_unit"):
            parts.append(f"{unit}(由{result['_converted_from_unit']}换算)")
        else:
            parts.append(unit)
    if price not in (None, ""):
        try:
            parts.append(f"{float(price):.2f}")
        except (TypeError, ValueError):
            pass

    return " | ".join(parts) if parts else fallback_text


def build_gldjc_url(keyword: str, province_code: str = "1") -> str:
    """构建广材网搜索链接（用户可点击查看）"""
    encoded = quote(keyword, safe='')
    return f"https://www.gldjc.com/scj/so.html?keyword={encoded}&l={province_code}"


# ========== JSON主材缓存 ==========

def load_cache() -> dict:
    """加载主材价格缓存"""
    if CACHE_FILE.exists():
        try:
            with open(CACHE_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except (json.JSONDecodeError, IOError):
            return {}
    return {}


def save_cache(cache: dict):
    """保存主材价格缓存"""
    CACHE_FILE.parent.mkdir(parents=True, exist_ok=True)
    with open(CACHE_FILE, 'w', encoding='utf-8') as f:
        json.dump(cache, f, ensure_ascii=False, indent=2)


_PROVINCE_AREA_CODES = {
    "全国": "1",
    "北京": "110000",
    "天津": "120000",
    "河北": "130000",
    "山西": "140000",
    "内蒙古": "150000",
    "辽宁": "210000",
    "吉林": "220000",
    "黑龙江": "230000",
    "上海": "310000",
    "江苏": "320000",
    "浙江": "330000",
    "安徽": "340000",
    "福建": "350000",
    "江西": "360000",
    "山东": "370000",
    "河南": "410000",
    "湖北": "420000",
    "湖南": "430000",
    "广东": "440000",
    "广西": "450000",
    "海南": "460000",
    "重庆": "500000",
    "四川": "510000",
    "贵州": "520000",
    "云南": "530000",
    "西藏": "540000",
    "陕西": "610000",
    "甘肃": "620000",
    "青海": "630000",
    "宁夏": "640000",
    "新疆": "650000",
    "香港": "810000",
    "澳门": "820000",
    "台湾": "710000",
}


def _normalize_cache_spec(spec: str) -> str:
    """归一化缓存规格 key，避免同一规格因书写差异被存成多份缓存。

    做的事：
      - 去空白、转大写；
      - 尺寸分隔符 ×/x/X/* → 统一为 X；
      - 直径符号 φ/Φ → 统一为 PHI；
      - DN/DE 保持大写（由 upper() 已处理）。

    示例：
      "DN25×2.75"、"dn25x2.75"、"DN25*2.75"  →  "DN25X2.75"
      "Φ10mm"、"φ10MM"、"PHI10MM"             →  "PHI10MM"
    """
    text = re.sub(r"\s+", "", str(spec or "").strip()).upper()
    if not text:
        return ""
    # 尺寸分隔符统一
    text = text.replace("×", "X").replace("*", "X")
    # 直径符号统一
    text = text.replace("Φ", "PHI").replace("φ", "PHI")
    return text


def _normalize_cache_region(region: str) -> str:
    return re.sub(r"\s+", "", str(region or "").strip())


def _normalize_region_name(name: str) -> str:
    text = str(name or "").strip()
    if not text:
        return ""
    replacements = (
        "省", "市", "壮族自治区", "回族自治区", "维吾尔自治区",
        "自治区", "特别行政区", "地区", "盟",
    )
    for token in replacements:
        text = text.replace(token, "")
    return text.strip()


def resolve_gldjc_area_code(province: str = "", city: str = "") -> str:
    normalized_province = _normalize_region_name(province)
    if normalized_province in _PROVINCE_AREA_CODES:
        return _PROVINCE_AREA_CODES[normalized_province]
    return "1"


def build_region_search_plans(search_keywords: list[str], province: str = "", city: str = "") -> list[dict]:
    plans: list[dict] = []
    seen: set[tuple[str, str]] = set()

    province_code = resolve_gldjc_area_code(province=province, city="")
    province_name = _normalize_region_name(province)

    def _add(keyword: str, area_code: str, scope: str):
        key = (keyword, area_code)
        if not keyword or key in seen:
            return
        seen.add(key)
        plans.append({"keyword": keyword, "area_code": area_code, "scope": scope})

    if province_code != "1":
        for kw in search_keywords:
            _add(kw, province_code, province_name or "全国")

    for kw in search_keywords:
        _add(kw, "1", "全国")

    return plans


def get_cache_key(name: str, unit: str, spec: str = "", region: str = "") -> str:
    """生成缓存key（材料名+规格+单位+地区）。"""
    return f"{name.strip()}|{_normalize_cache_spec(spec)}|{unit.strip()}|{_normalize_cache_region(region)}"


def check_cache(cache: dict, name: str, unit: str, spec: str = "", region: str = "") -> dict | None:
    """
    查缓存：有且未过期返回缓存数据，否则返回None

    缓存结构：{
        "镀锌钢管 DN25|m": {
            "price_with_tax": 18.5,
            "price_without_tax": 16.37,
            "confidence": "高",
            "match_status": "精确匹配",
            "source": "广材网",
            "query_date": "2026-03-10",
            "result_count": 8,
            "match_detail": "搜到DN25镀锌钢管，单位m，8条报价"
        }
    }
    """
    key = get_cache_key(name, unit, spec, region=region)
    if key not in cache:
        return None

    entry = cache[key]
    query_date = entry.get("query_date", "")
    if not query_date:
        return None

    # 检查是否过期
    try:
        cached_time = datetime.strptime(query_date, "%Y-%m-%d")
        if datetime.now() - cached_time > timedelta(days=CACHE_EXPIRE_DAYS):
            return None  # 过期
    except ValueError:
        return None

    return entry


def update_cache(cache: dict, name: str, unit: str, data: dict, spec: str = "", region: str = ""):
    """更新缓存"""
    key = get_cache_key(name, unit, spec, region=region)
    cache[key] = data


# ========== Excel输出 ==========

# Excel背景色
FILL_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")   # 高置信度
FILL_YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # 需核对
FILL_RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")      # 搜不到
FILL_BLUE = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")    # 非材料项
FONT_LINK = Font(color="0563C1", underline="single")  # 超链接样式


def process_excel(input_path: str, cookie_str: str, output_path: str = None,
                  province_code: str = "1", delay: float = 3.0,
                  use_cache: bool = True):
    """
    读取Excel材料清单，逐个查询广材网价格，写回Excel

    v2改进：
    - 名称拆解+规格提取
    - 单位/规格过滤
    - 非材料项拦截
    - 置信度标注
    - 广材网链接
    - JSON缓存
    """
    input_file = Path(input_path)
    if not input_file.exists():
        print(f"文件不存在: {input_path}")
        return

    if not output_path:
        output_path = str(input_file.with_stem(input_file.stem + "_含价格"))

    # 读取Excel
    wb = openpyxl.load_workbook(input_path)
    ws = wb.active

    # 找到列索引（自动识别列名）
    header_row = 1
    col_map = {}
    for col_idx, cell in enumerate(ws[header_row], 1):
        val = str(cell.value or "").strip()
        if val in ("材料名称", "名称", "项目名称"):
            col_map["name"] = col_idx
        elif val in ("规格型号", "规格", "型号"):
            col_map["spec"] = col_idx
        elif val in ("单位", "计量单位"):
            col_map["unit"] = col_idx
        elif val in ("含税市场价", "市场价", "含税价"):
            col_map["price_with_tax"] = col_idx
        elif val in ("不含税市场价", "不含税价", "除税价"):
            col_map["price_without_tax"] = col_idx
        elif val == "税率":
            col_map["tax_rate"] = col_idx

    if "name" not in col_map:
        print("未找到'材料名称'列，请检查Excel表头")
        return

    # 新增输出列：匹配状态、置信度、广材网链接
    max_col = ws.max_column
    if "price_with_tax" not in col_map:
        col_map["price_with_tax"] = max_col + 1
        ws.cell(row=header_row, column=col_map["price_with_tax"], value="含税市场价")
        max_col += 1
    if "price_without_tax" not in col_map:
        col_map["price_without_tax"] = max_col + 1
        ws.cell(row=header_row, column=col_map["price_without_tax"], value="不含税市场价")
        max_col += 1

    # 新增3列
    col_map["match_status"] = max_col + 1
    ws.cell(row=header_row, column=col_map["match_status"], value="匹配状态")
    col_map["confidence"] = max_col + 2
    ws.cell(row=header_row, column=col_map["confidence"], value="置信度")
    col_map["gldjc_link"] = max_col + 3
    ws.cell(row=header_row, column=col_map["gldjc_link"], value="广材网链接")

    # 创建session，注入cookie
    session = requests.Session()
    for part in re.split(r";\s*", cookie_str):
        if "=" in part:
            key, value = part.split("=", 1)
            session.cookies.set(key.strip(), value.strip())

    # 加载缓存
    cache = load_cache() if use_cache else {}

    # 搜索结果缓存（本次运行内去重，同名材料只搜一次）
    search_cache = {}

    # 统计
    total = ws.max_row - 1
    stats = {"成功": 0, "缓存命中": 0, "未匹配": 0, "非材料项": 0, "需核对": 0, "已有价格": 0}

    print(f"开始查询 {total} 条材料价格（v2：带过滤+置信度+链接）")
    print(f"缓存：{'启用' if use_cache else '禁用'}，缓存条数：{len(cache)}")
    print(f"每次请求间隔 {delay} 秒（防止被封）")
    print()

    for row_idx in range(2, ws.max_row + 1):
        name = str(ws.cell(row=row_idx, column=col_map["name"]).value or "").strip()
        spec_col = str(ws.cell(row=row_idx, column=col_map.get("spec", col_map["name"])).value or "").strip()
        unit = str(ws.cell(row=row_idx, column=col_map.get("unit", col_map["name"])).value or "").strip()

        if not name:
            continue

        current = row_idx - 1
        display_name = name.encode('gbk', errors='replace').decode('gbk')

        # 跳过已经有价格的行
        existing_price = ws.cell(row=row_idx, column=col_map["price_with_tax"]).value
        if existing_price and str(existing_price).strip():
            print(f"  [{current}/{total}] {display_name} → 已有价格，跳过")
            stats["已有价格"] += 1
            continue

        # 非材料项拦截
        if is_non_material(name):
            print(f"  [{current}/{total}] {display_name} → 非材料项，跳过")
            ws.cell(row=row_idx, column=col_map["match_status"], value="非材料项")
            ws.cell(row=row_idx, column=col_map["confidence"], value="-")
            # 蓝色背景
            for col in [col_map["match_status"], col_map["confidence"]]:
                ws.cell(row=row_idx, column=col).fill = FILL_BLUE
            stats["非材料项"] += 1
            continue

        # 查缓存
        if use_cache:
            cached = (
                check_cache(cache, name, unit, spec_col, region=province_code)
                or check_cache(cache, name, unit, spec_col, region="全国")
                or check_cache(cache, name, unit, spec_col)
            )
            if cached:
                price_tax = cached.get("price_with_tax")
                price_notax = cached.get("price_without_tax")
                conf = cached.get("confidence", "中")
                status = cached.get("match_status", "缓存")

                if price_tax and conf != "低":
                    ws.cell(row=row_idx, column=col_map["price_with_tax"], value=price_tax)
                    ws.cell(row=row_idx, column=col_map["price_without_tax"], value=price_notax)
                    ws.cell(row=row_idx, column=col_map["match_status"], value=f"{status}(缓存)")
                    ws.cell(row=row_idx, column=col_map["confidence"], value=conf)
                    fill = FILL_GREEN if conf == "高" else FILL_YELLOW
                    for col in [col_map["price_with_tax"], col_map["price_without_tax"],
                                col_map["match_status"], col_map["confidence"]]:
                        ws.cell(row=row_idx, column=col).fill = fill
                    print(f"  [{current}/{total}] {display_name} → 缓存命中 {price_tax}元")
                    stats["缓存命中"] += 1
                    continue

        # 名称拆解
        parsed = parse_material(name, spec_col)
        base_name = parsed["base_name"]
        specs = parsed["specs"]

        # 广材网链接（不管能不能查到价格，都给链接）
        link_keyword = f"{base_name} {specs[0]}" if specs else base_name
        link_url = build_gldjc_url(link_keyword, province_code)

        # 分层搜索：先精确（品名+规格），再宽泛（纯品名）
        all_results = []
        searched_keyword = ""
        for kw in parsed["search_keywords"]:
            if kw in search_cache:
                # 本次运行内去重
                all_results = search_cache[kw]
                searched_keyword = kw
                break

            print(f"  [{current}/{total}] 搜索: {kw.encode('gbk', errors='replace').decode('gbk')}...",
                  end=" ", flush=True)

            results = search_material_web(session, kw, province_code)
            search_cache[kw] = results
            searched_keyword = kw

            if results:
                all_results = results
                break

            # 没搜到，降级到下一个关键词
            print("未找到，降级...", end=" ", flush=True)
            time.sleep(delay * 0.5 + random.uniform(0, 0.5))

        if not all_results:
            # 所有关键词都没搜到
            print("全部未找到")
            ws.cell(row=row_idx, column=col_map["match_status"], value="未匹配")
            ws.cell(row=row_idx, column=col_map["confidence"], value="低")
            # 写入链接
            link_cell = ws.cell(row=row_idx, column=col_map["gldjc_link"])
            link_cell.value = link_url
            link_cell.hyperlink = link_url
            link_cell.font = FONT_LINK
            # 红色背景
            for col in [col_map["match_status"], col_map["confidence"]]:
                ws.cell(row=row_idx, column=col).fill = FILL_RED
            stats["未匹配"] += 1
            # 缓存"未匹配"结果（避免下次再搜）
            update_cache(cache, name, unit, {
                "price_with_tax": None,
                "price_without_tax": None,
                "confidence": "低",
                "match_status": "未匹配",
                "source": "广材网",
                "query_date": datetime.now().strftime("%Y-%m-%d"),
                "result_count": 0,
                "searched_keyword": searched_keyword,
            }, spec=spec_col, region=province_code)
            time.sleep(delay + random.uniform(0, 1))
            continue

        # 过滤+打分：先严格匹配，失败再退到松弛规格匹配
        scored = filter_and_score(all_results, unit, specs, base_name,
                                  request_name=name, request_spec=spec_col)
        if not scored and specs:
            # 第一轮全被规格过滤掉，尝试放宽（DN25 也可匹配到 DN25x2.75 之类）
            scored = filter_and_score(all_results, unit, specs, base_name,
                                      request_name=name, request_spec=spec_col,
                                      allow_relaxed_spec=True)
            if scored:
                # 标记这些结果是松弛命中，在输出层提示用户核对
                for r in scored:
                    r["_relaxed_spec_match"] = True

        confidence = determine_confidence(scored, unit, specs)

        # 只允许使用通过过滤的结果取价；宁可留空，也不回填错误价格
        price_source = scored
        if not scored:
            confidence = "低"

        selected_result = get_representative_price_result(price_source)
        selected_price = None
        if selected_result:
            try:
                selected_price = round(float(selected_result.get("market_price")), 2)
            except (TypeError, ValueError):
                selected_price = None

        # 生成匹配说明
        if scored:
            top = scored[0]
            match_detail = f"搜到{top.get('spec', '')[:20]}，单位{top.get('unit', '?')}，{len(scored)}条报价"
        else:
            match_detail = f"搜到{len(all_results)}条但单位/规格不匹配"

        # 根据置信度决定是否填价
        match_status = ""
        if confidence == "高":
            match_status = "精确匹配"
        elif confidence == "中":
            match_status = "模糊匹配"
        else:
            match_status = "低置信度"

        # 只要广材网抓到了可用价格就填，低置信度也回填，避免留空
        if selected_price:
            tax_rate = 1.13  # 默认13%增值税
            # 如果Excel里有税率列，用Excel里的
            if "tax_rate" in col_map:
                tr = ws.cell(row=row_idx, column=col_map["tax_rate"]).value
                if tr and float(tr) > 0:
                    tax_rate = 1 + float(tr) / 100 if float(tr) > 1 else 1 + float(tr)

            price_notax = round(selected_price / tax_rate, 2)
            ws.cell(row=row_idx, column=col_map["price_with_tax"], value=selected_price)
            ws.cell(row=row_idx, column=col_map["price_without_tax"], value=price_notax)

            fill = FILL_GREEN if confidence == "高" else (FILL_YELLOW if confidence == "中" else FILL_RED)
            print(f"含税 {selected_price} 元（{confidence}，{len(price_source)}条报价）")
            stats["成功"] += 1

            # 更新缓存
            update_cache(cache, name, unit, {
                "price_with_tax": selected_price,
                "price_without_tax": price_notax,
                "confidence": confidence,
                "match_status": match_status,
                "source": "广材网",
                "query_date": datetime.now().strftime("%Y-%m-%d"),
                "result_count": len(price_source),
                "match_detail": match_detail,
                "searched_keyword": searched_keyword,
                "match_label": build_match_label(selected_result),
            }, spec=spec_col, region=province_code)
        else:
            # 完全没有价格才留空
            fill = FILL_RED
            print(f"低置信度，不填价（{match_detail}）")
            stats["需核对"] += 1

            update_cache(cache, name, unit, {
                "price_with_tax": selected_price,  # 存着但不写入Excel
                "price_without_tax": round(selected_price / 1.13, 2) if selected_price else None,
                "confidence": "低",
                "match_status": match_status,
                "source": "广材网",
                "query_date": datetime.now().strftime("%Y-%m-%d"),
                "result_count": len(price_source),
                "match_detail": match_detail,
                "searched_keyword": searched_keyword,
                "match_label": build_match_label(selected_result, fallback_text=match_detail),
            }, spec=spec_col, region=province_code)

        # 写匹配状态和置信度
        ws.cell(row=row_idx, column=col_map["match_status"], value=match_status)
        ws.cell(row=row_idx, column=col_map["confidence"], value=confidence)

        # 写广材网链接（所有行都给，方便核对）
        link_cell = ws.cell(row=row_idx, column=col_map["gldjc_link"])
        link_cell.value = link_url
        link_cell.hyperlink = link_url
        link_cell.font = FONT_LINK

        # 背景色
        for col in [col_map["match_status"], col_map["confidence"]]:
            ws.cell(row=row_idx, column=col).fill = fill
        if confidence in ("高", "中"):
            for col in [col_map["price_with_tax"], col_map["price_without_tax"]]:
                ws.cell(row=row_idx, column=col).fill = fill

        # 随机延迟（只在真正发了请求时才延迟）
        if searched_keyword not in search_cache or len(search_cache) <= 1:
            time.sleep(delay + random.uniform(0, 1))

        # 每10条自动保存
        done = sum(v for v in stats.values())
        if done % 10 == 0 and done > 0:
            wb.save(output_path)
            if use_cache:
                save_cache(cache)

    # 最终保存
    wb.save(output_path)
    if use_cache:
        save_cache(cache)

    print()
    print("=" * 50)
    print(f"查询完成！结果保存到: {output_path}")
    print(f"  成功（自动填价）：{stats['成功']} 条")
    print(f"  缓存命中：{stats['缓存命中']} 条")
    print(f"  需核对（低置信度，给了链接）：{stats['需核对']} 条")
    print(f"  未匹配（搜不到）：{stats['未匹配']} 条")
    print(f"  非材料项（自动跳过）：{stats['非材料项']} 条")
    print(f"  已有价格（跳过）：{stats['已有价格']} 条")
    print(f"  主材缓存：{len(cache)} 条")
    print()
    print("颜色说明：")
    print("  绿色 = 高置信度，可直接用")
    print("  黄色 = 中置信度，建议点链接核对")
    print("  红色 = 低置信度/未找到，请手动查价")
    print("  蓝色 = 非材料项，已跳过")


def main():
    parser = argparse.ArgumentParser(description="广材网材料价格查询工具 v2")
    parser.add_argument("excel", help="材料清单Excel文件路径")
    parser.add_argument("--cookie", required=True,
                        help="广材网cookie字符串（至少包含token=bearer xxx）")
    parser.add_argument("--output", "-o",
                        help="输出Excel路径（默认原文件名_含价格）")
    parser.add_argument("--province", default="1",
                        help="省份代码（1=全国，默认）")
    parser.add_argument("--delay", type=float, default=3.0,
                        help="请求间隔秒数（默认3秒）")
    parser.add_argument("--no-cache", action="store_true",
                        help="不使用缓存，强制重新查询广材网")

    args = parser.parse_args()
    process_excel(args.excel, args.cookie, args.output,
                  args.province, args.delay, not args.no_cache)


if __name__ == "__main__":
    main()
