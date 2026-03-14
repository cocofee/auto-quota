# -*- coding: utf-8 -*-
"""
清单扫描器 — 快速摸底一个文件/文件夹/压缩包里有多少清单、什么专业

用途：接活前快速评估工作量和报价
  - 支持 Excel(.xlsx/.xls)、ZIP压缩包、文件夹
  - 自动识别专业（安装/土建/市政/园林 + 细分册号）
  - 按名称+描述去重，得到实际需要套定额的条数

用法：
    python tools/bill_scanner.py "某项目.xlsx"
    python tools/bill_scanner.py "某项目.zip"
    python tools/bill_scanner.py "F:\\jarvis\\给排水"
    python tools/bill_scanner.py "F:\\jarvis\\给排水" --output report.xlsx
    python tools/bill_scanner.py "F:\\jarvis" --limit 10          # 只扫前10个文件
"""

import sys
import os
import argparse
import zipfile
import tempfile
import shutil
from pathlib import Path
from collections import Counter, defaultdict
from datetime import datetime

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))


# 大类映射：册号 → 大类名
MAJOR_CATEGORY = {
    "C1": "安装工程", "C2": "安装工程", "C3": "安装工程",
    "C4": "安装工程", "C5": "安装工程", "C6": "安装工程",
    "C7": "安装工程", "C8": "安装工程", "C9": "安装工程",
    "C10": "安装工程", "C11": "安装工程", "C12": "安装工程",
    "C13": "安装工程",
    "A": "土建工程",
    "D": "市政工程",
    "E": "园林绿化",
}

# 册号 → 细分专业名
BOOK_NAMES = {
    "C1": "机械设备", "C2": "热力设备", "C3": "静置设备",
    "C4": "电气", "C5": "智能化", "C6": "自动化仪表",
    "C7": "通风空调", "C8": "工业管道", "C9": "消防",
    "C10": "给排水采暖", "C11": "通信", "C12": "刷油防腐",
    "C13": "其他附属",
    "A": "房建装饰",
    "D": "市政",
    "E": "园林绿化",
}


def collect_excel_files(input_path: str, limit: int = 0) -> list[str]:
    """从输入路径收集所有Excel文件（支持文件/文件夹/ZIP）

    返回: Excel文件路径列表（ZIP时返回解压后的临时路径）
    """
    p = Path(input_path)
    temp_dirs = []  # 记录临时目录，调用方负责清理
    excel_files = []

    if p.is_file() and p.suffix.lower() in (".zip", ".rar", ".7z"):
        # 解压ZIP
        if p.suffix.lower() == ".zip":
            tmp = tempfile.mkdtemp(prefix="bill_scan_")
            temp_dirs.append(tmp)
            try:
                with zipfile.ZipFile(str(p), 'r') as zf:
                    zf.extractall(tmp)
                # 递归找Excel
                for root, dirs, files in os.walk(tmp):
                    for f in files:
                        if f.lower().endswith(('.xlsx', '.xls')) and not f.startswith('~'):
                            excel_files.append(os.path.join(root, f))
            except Exception as e:
                print(f"  解压失败: {e}")
        else:
            print(f"  暂不支持 {p.suffix} 格式，请先解压")
    elif p.is_file() and p.suffix.lower() in ('.xlsx', '.xls'):
        excel_files.append(str(p))
    elif p.is_dir():
        # 递归扫描文件夹
        for root, dirs, files in os.walk(str(p)):
            for f in sorted(files):
                if f.lower().endswith(('.xlsx', '.xls')) and not f.startswith('~'):
                    excel_files.append(os.path.join(root, f))
    else:
        print(f"  无法识别的路径: {input_path}")

    if limit > 0:
        excel_files = excel_files[:limit]

    return excel_files, temp_dirs


def _fallback_read_excel(file_path: str) -> list[tuple]:
    """宽松模式读取Excel — 兜底用，识别"名称/规格型号/材质/数量"等列头

    标准BillReader读不出来的文件走这个，比如检修材料表、设备清册等。
    只要有"名称"列就尝试提取。支持 .xls 和 .xlsx。

    返回: [(name, desc, sheet_name, code), ...]
    """
    import openpyxl

    # 可能的列名映射
    NAME_COLS = ["名称", "材料名称", "设备名称", "货物名称", "项目名称",
                 "品名", "物资名称", "材料", "清单名称",
                 "项目描述", "工程项目", "清单项目", "分项工程名称",
                 "物料名称", "主材名称", "产品名称"]
    DESC_COLS = ["规格型号", "规格", "型号", "规格型号及技术参数", "技术参数",
                 "材质", "项目特征", "特征描述", "说明", "备注",
                 "规格型号-类型", "工作内容",
                 "参数", "工程内容", "构造做法", "单位", "计量单位"]

    result_items = []
    fp = Path(file_path)

    # 判断文件格式：.xls 用 xlrd，.xlsx 用 openpyxl
    is_xls = False
    if fp.suffix.lower() == '.xls':
        is_xls = True
    else:
        # 有些 .xlsx 实际上是 .xls 格式（后缀骗人）
        try:
            with open(file_path, 'rb') as f:
                magic = f.read(8)
            if magic[:4] == b'\xd0\xcf\x11\xe0':  # OLE2 格式 = xls
                is_xls = True
        except Exception:
            pass

    if is_xls:
        return _fallback_read_xls(file_path, NAME_COLS, DESC_COLS)
    else:
        return _fallback_read_xlsx(file_path, NAME_COLS, DESC_COLS)


def _fallback_read_xls(file_path: str, NAME_COLS: list, DESC_COLS: list) -> list[tuple]:
    """用 BillReader 的 xls→xlsx 转换后再用 openpyxl 读"""
    import tempfile
    result_items = []

    try:
        # 借用 BillReader 的转换方法
        from src.bill_reader import BillReader
        reader = BillReader()
        temp_path = reader._convert_xls_to_xlsx(file_path)
        if temp_path:
            result_items = _fallback_read_xlsx(temp_path, NAME_COLS, DESC_COLS)
            try:
                os.remove(temp_path)
            except Exception:
                pass
    except Exception:
        pass

    return result_items


def _fallback_read_xlsx(file_path: str, NAME_COLS: list, DESC_COLS: list) -> list[tuple]:
    """用 openpyxl 读取 .xlsx 格式"""
    import openpyxl
    result_items = []

    try:
        wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
    except Exception:
        return []

    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # 在前15行里找表头
            header_row = None
            name_col = None
            desc_cols = []

            for i, row in enumerate(ws.iter_rows(max_row=50, values_only=False)):
                cells = [str(c.value).strip() if c.value else "" for c in row]

                for j, val in enumerate(cells):
                    if any(kw == val for kw in NAME_COLS):
                        header_row = i
                        name_col = j
                        break

                if name_col is not None:
                    for j, val in enumerate(cells):
                        if j != name_col and any(kw == val for kw in DESC_COLS):
                            desc_cols.append(j)
                    break

            if name_col is None:
                continue

            # 从表头下一行开始读数据
            for row in ws.iter_rows(min_row=header_row + 2, values_only=True):
                cells = list(row)
                if name_col >= len(cells):
                    continue
                name = str(cells[name_col]).strip() if cells[name_col] else ""
                if not name or name == "None" or name.isdigit():
                    continue
                # 过滤明显的非清单行（合计、小计、汇总等）
                if name in ("合计", "小计", "总计", "汇总", "序号", "编号"):
                    continue

                desc_parts = []
                for dc in desc_cols:
                    if dc < len(cells) and cells[dc]:
                        val = str(cells[dc]).strip()
                        if val and val != "None" and val not in ("——", "--"):
                            desc_parts.append(val)
                desc = "\n".join(desc_parts)

                result_items.append((name, desc, sheet_name, ""))
    finally:
        wb.close()

    return result_items


def scan_one_file(file_path: str, reader) -> dict:
    """扫描单个Excel文件，返回清单信息

    先用标准BillReader读，读不出来再用宽松模式兜底。

    返回: {
        "file": 文件名,
        "items": [(name, desc, sheet_name, code), ...],  # 原始清单项
        "source": "standard"(标准清单) / "fallback"(宽松识别) / None(空文件),
        "error": 错误信息（如果有）
    }
    """
    result_items = []
    source = None  # 数据来源标记

    # 第一步：标准模式（真正的工程量清单）
    try:
        items = reader.read_excel(file_path)
        for item in items:
            name = str(item.get("name", "")).strip()
            desc = str(item.get("description", "")).strip()
            sheet = str(item.get("section_title", "") or item.get("sheet_name", ""))
            code = str(item.get("code", "")).strip()
            if name and len(name) >= 2:
                result_items.append((name, desc, sheet, code))
        if result_items:
            source = "standard"
    except Exception:
        pass

    # 第二步：标准模式没读出来，用宽松模式兜底（材料表/设备表等）
    if not result_items:
        try:
            result_items = _fallback_read_excel(file_path)
            if result_items:
                source = "fallback"
        except Exception:
            pass

    return {
        "file": os.path.basename(file_path),
        "path": file_path,
        "items": result_items,
        "source": source,
        "error": None,
    }


def classify_items(items: list[tuple], classifier_fn) -> dict:
    """对清单项进行专业分类

    参数:
        items: [(name, desc, sheet, code), ...]
        classifier_fn: specialty_classifier.classify 函数

    返回: {
        "total": 总条数,
        "unique": 去重后条数,
        "by_major": {"安装工程": 120, "土建工程": 30, ...},  # 大类
        "by_book": {"C10": 50, "C4": 40, ...},               # 细分册号
        "by_book_detail": {"C10(给排水采暖)": 50, ...},       # 带名称
        "dedup_items": [(name, desc, book), ...],             # 去重后的条目
    }
    """
    # 去重：名称+描述完全一致的只算一条
    seen = set()
    unique_items = []
    for name, desc, sheet, code in items:
        key = (name, desc)
        if key not in seen:
            seen.add(key)
            unique_items.append((name, desc, sheet, code))

    # 分类
    by_major = Counter()   # 大类计数
    by_book = Counter()    # 册号计数
    dedup_items = []

    for name, desc, sheet, code in unique_items:
        try:
            result = classifier_fn(name, desc, section_title=sheet, bill_code=code)
            book = result.get("primary") or "未知"
        except Exception:
            book = "未知"

        major = MAJOR_CATEGORY.get(book, "其他")
        by_major[major] += 1
        by_book[book] += 1
        dedup_items.append((name, desc, book))

    # 带名称的册号统计
    by_book_detail = {}
    for book, count in by_book.items():
        bname = BOOK_NAMES.get(book, book)
        by_book_detail[f"{book}({bname})"] = count

    return {
        "total": len(items),
        "unique": len(unique_items),
        "by_major": dict(by_major),
        "by_book": dict(by_book),
        "by_book_detail": by_book_detail,
        "dedup_items": dedup_items,
    }


def print_report(input_path: str, file_results: list[dict], classification: dict):
    """打印终端报告"""
    print()
    print("=" * 60)
    print(f"  清单扫描报告")
    print(f"  扫描路径: {input_path}")
    print(f"  扫描时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)

    # 文件统计
    total_files = len(file_results)
    ok_files = sum(1 for r in file_results if r["items"])
    err_files = sum(1 for r in file_results if r["error"])
    empty_files = sum(1 for r in file_results if not r["items"] and not r["error"])

    print(f"\n  文件统计:")
    print(f"    扫描文件数:   {total_files}")
    print(f"    有效清单文件: {ok_files}")
    if err_files:
        print(f"    读取失败:     {err_files}")
    if empty_files:
        print(f"    无清单内容:   {empty_files}")

    # 清单统计
    print(f"\n  清单统计:")
    print(f"    清单总条数:   {classification['total']}")
    print(f"    去重后条数:   {classification['unique']}")
    if classification['total'] > 0:
        dup_rate = (1 - classification['unique'] / classification['total']) * 100
        print(f"    重复率:       {dup_rate:.1f}%")

    # 大类分布
    print(f"\n  专业大类:")
    for major, count in sorted(classification['by_major'].items(),
                                key=lambda x: -x[1]):
        pct = count / classification['unique'] * 100 if classification['unique'] else 0
        bar = "█" * int(pct / 3)
        print(f"    {major:10s}  {count:>5d}条 ({pct:4.1f}%)  {bar}")

    # 细分专业
    print(f"\n  细分专业:")
    for book_detail, count in sorted(classification['by_book_detail'].items(),
                                      key=lambda x: -x[1]):
        pct = count / classification['unique'] * 100 if classification['unique'] else 0
        print(f"    {book_detail:20s}  {count:>5d}条 ({pct:4.1f}%)")

    # 各文件概览（只显示有清单的文件，最多20个）
    ok_results = [r for r in file_results if r["items"]]
    if ok_results:
        print(f"\n  文件明细（前20个）:")
        for r in ok_results[:20]:
            print(f"    {r['file'][:40]:40s}  {len(r['items']):>5d}条")
        if len(ok_results) > 20:
            print(f"    ... 还有 {len(ok_results) - 20} 个文件")

    print()
    print("=" * 60)
    print(f"  结论: {classification['unique']}条去重清单，"
          f"{len(classification['by_major'])}个大类，"
          f"{len(classification['by_book'])}个细分专业")
    print("=" * 60)
    print()


def save_excel_report(output_path: str, file_results: list[dict],
                       classification: dict, input_path: str):
    """保存Excel报告"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        print("  需要 openpyxl 才能输出Excel报告")
        return

    wb = openpyxl.Workbook()

    # --- Sheet1: 汇总 ---
    ws = wb.active
    ws.title = "扫描汇总"
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF", size=11)

    ws.append(["清单扫描报告"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([f"扫描路径: {input_path}"])
    ws.append([f"扫描时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
    ws.append([])

    ws.append(["指标", "数值"])
    ws.append(["清单总条数", classification['total']])
    ws.append(["去重后条数", classification['unique']])
    dup_rate = (1 - classification['unique'] / classification['total']) * 100 if classification['total'] else 0
    ws.append(["重复率", f"{dup_rate:.1f}%"])
    ws.append([])

    ws.append(["专业大类", "条数", "占比"])
    for major, count in sorted(classification['by_major'].items(), key=lambda x: -x[1]):
        pct = count / classification['unique'] * 100 if classification['unique'] else 0
        ws.append([major, count, f"{pct:.1f}%"])
    ws.append([])

    ws.append(["细分专业", "条数", "占比"])
    for book_detail, count in sorted(classification['by_book_detail'].items(), key=lambda x: -x[1]):
        pct = count / classification['unique'] * 100 if classification['unique'] else 0
        ws.append([book_detail, count, f"{pct:.1f}%"])

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15

    # --- Sheet2: 去重清单明细 ---
    ws2 = wb.create_sheet("去重清单明细")
    headers = ["序号", "清单名称", "项目特征", "专业"]
    ws2.append(headers)
    for cell in ws2[1]:
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for i, (name, desc, book) in enumerate(classification['dedup_items'], 1):
        bname = BOOK_NAMES.get(book, book)
        ws2.append([i, name, desc[:200], f"{book}({bname})"])

    ws2.column_dimensions['A'].width = 8
    ws2.column_dimensions['B'].width = 35
    ws2.column_dimensions['C'].width = 50
    ws2.column_dimensions['D'].width = 18

    # --- Sheet3: 文件明细 ---
    ws3 = wb.create_sheet("文件明细")
    headers = ["序号", "文件名", "清单条数", "状态"]
    ws3.append(headers)
    for cell in ws3[1]:
        cell.font = header_font_white
        cell.fill = header_fill

    for i, r in enumerate(file_results, 1):
        status = "有效" if r["items"] else (f"错误: {r['error']}" if r["error"] else "无清单")
        ws3.append([i, r["file"], len(r["items"]), status])

    ws3.column_dimensions['A'].width = 8
    ws3.column_dimensions['B'].width = 45
    ws3.column_dimensions['C'].width = 12
    ws3.column_dimensions['D'].width = 30

    wb.save(output_path)
    print(f"  Excel报告已保存: {output_path}")


def main():
    parser = argparse.ArgumentParser(
        description="清单扫描器：快速摸底文件里的清单数量和专业分布"
    )
    parser.add_argument("input", help="Excel文件 / ZIP压缩包 / 文件夹路径")
    parser.add_argument("--output", "-o", help="输出Excel报告路径（可选）")
    parser.add_argument("--limit", type=int, default=0,
                        help="最多扫描多少个文件（0=不限，调试用）")
    args = parser.parse_args()

    input_path = args.input
    if not os.path.exists(input_path):
        print(f"路径不存在: {input_path}")
        return

    # 收集Excel文件
    print(f"正在收集文件...")
    excel_files, temp_dirs = collect_excel_files(input_path, limit=args.limit)
    if not excel_files:
        print("没有找到Excel文件")
        return
    print(f"  找到 {len(excel_files)} 个Excel文件")

    # 初始化读取器和分类器
    from src.bill_reader import BillReader
    from src.specialty_classifier import classify as classify_specialty

    reader = BillReader()

    # 逐文件扫描
    print(f"正在扫描清单...")
    file_results = []
    all_items = []  # 所有清单项汇总

    for i, fpath in enumerate(excel_files):
        if (i + 1) % 10 == 0 or i == 0:
            print(f"  进度: {i+1}/{len(excel_files)}")
        result = scan_one_file(fpath, reader)
        file_results.append(result)
        all_items.extend(result["items"])

    print(f"  扫描完成，共 {len(all_items)} 条原始清单")

    # 分类+去重
    print(f"正在分类和去重...")
    classification = classify_items(all_items, classify_specialty)

    # 打印报告
    print_report(input_path, file_results, classification)

    # 输出Excel报告
    if args.output:
        save_excel_report(args.output, file_results, classification, input_path)

    # 清理临时目录
    for td in temp_dirs:
        try:
            shutil.rmtree(td)
        except Exception:
            pass


if __name__ == "__main__":
    main()
