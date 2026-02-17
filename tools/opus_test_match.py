"""
Opus 4.6 直接匹配测试脚本
功能：读取北京大学电教厅电气工程清单，用预设的匹配规则生成定额匹配结果Excel
"""

import sqlite3
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from pathlib import Path

# ============================================================
# 配置
# ============================================================
BILL_FILE = r"C:\Users\Administrator\Desktop\北京大学新燕园校区电教厅改造工程.xlsx"
SHEET_NAME = "4.6 分部分项工程项目清单计价表【电气工程】"
QUOTA_DB = r"C:\Users\Administrator\Documents\trae_projects\auto-quota\db\provinces\北京2024\quota.db"
OUTPUT_FILE = r"C:\Users\Administrator\Desktop\Opus4.6匹配测试结果.xlsx"

# ============================================================
# 匹配规则（Opus 4.6 人工分析后的匹配决策）
# 格式：序号 → (定额编号列表, 置信度, 备注)
# ============================================================
MATCH_RULES = {
    # === 配电箱 (1-13) ===
    # 根据特征描述中的回路数选择定额
    # C4-4-30: ≤4回路, C4-4-31: ≤8回路, C4-4-32: ≤16回路, C4-4-33: ≤24回路
    1:  (["C4-4-31"], "高", "APE-2, 8回路, 墙上明装"),
    2:  (["C4-4-30"], "高", "APE-1, 3回路, 墙上明装"),
    3:  (["C4-4-30"], "高", "APE-1, 3回路, 墙上明装"),
    4:  (["C4-4-30"], "高", "APE-Z, 2回路, 墙上明装"),
    5:  (["C4-4-31"], "高", "APL-1~3, 6回路, 墙上明装"),
    6:  (["C4-4-31"], "高", "APL-1~3, 6回路, 墙上明装"),
    7:  (["C4-4-31"], "高", "APL-1~3, 6回路, 墙上明装"),
    8:  (["C4-4-31"], "高", "APL-4, 8回路, 墙上明装"),
    9:  (["C4-4-31"], "高", "APL-5, 6回路, 墙上明装"),
    10: (["C4-4-31"], "高", "APL-6, 6回路, 墙上明装"),
    11: (["C4-4-31"], "高", "APL-7, 5回路, 墙上明装"),
    12: (["C4-4-31"], "高", "APS, 5回路, 墙上明装"),
    13: (["C4-4-32"], "高", "APX, 10回路, 墙上明装"),

    # === 电力电缆 (14-24) ===
    # 沿桥架/墙面敷设，根据截面积选定额
    # C4-8-11: ≤6mm², C4-8-12: ≤10mm², C4-8-13: ≤16mm², C4-8-14: ≤25mm²
    # C4-8-15: ≤50mm², C4-8-16: ≤95mm², C4-8-17: ≤150mm²
    14: (["C4-8-14"], "高", "YJV-5×10, 截面10→取25档"),
    15: (["C4-8-14"], "高", "YJV-5×10, 截面10→取25档"),
    16: (["C4-8-13"], "高", "YJV-5×6, 截面6→取16档"),
    17: (["C4-8-13"], "高", "YJV-5×6, 截面6→取16档"),
    18: (["C4-8-13"], "高", "YJV-5×4, 截面4→取16档"),
    19: (["C4-8-11"], "高", "YJV-5×4, 截面4→取6档"),
    20: (["C4-8-11"], "高", "YJV-3×2.5, 截面2.5→取6档"),
    21: (["C4-8-11"], "高", "YJV-5×2.5, 截面2.5→取6档"),
    22: (["C4-8-11"], "高", "YJV-3×2.5+2×1.5, 截面2.5→取6档"),
    23: (["C4-8-11"], "高", "YJV-5×2.5, 截面2.5→取6档"),
    24: (["C4-8-11"], "高", "YJV-3×2.5+2×1.5, 截面2.5→取6档"),

    # === 电缆头 (25-34) ===
    # 非铠装电缆终端头，根据芯数×截面选定额
    # C4-8-231: 2芯≤6mm², C4-8-232: 2芯≤25, C4-8-233: 2芯≤70
    # C4-8-236: 3芯≤6mm², C4-8-237: 3芯≤25, C4-8-238: 3芯≤70
    # C4-8-241: 4芯≤6mm², C4-8-242: 4芯≤25, C4-8-243: 4芯≤70
    # 5芯用4芯+1芯的方式
    25: (["C4-8-242"], "高", "YJV-5×10→5芯25mm²以内，按4芯+1芯"),
    26: (["C4-8-242"], "高", "YJV-5×10→5芯25mm²以内"),
    27: (["C4-8-242"], "高", "YJV-5×6→5芯6mm²"),
    28: (["C4-8-242"], "高", "YJV-5×6→5芯6mm²"),
    29: (["C4-8-241"], "高", "YJV-5×4→5芯4mm²"),
    30: (["C4-8-241"], "高", "YJV-5×4→5芯4mm²"),
    31: (["C4-8-236"], "高", "YJV-3×2.5→3芯2.5mm²"),
    32: (["C4-8-241"], "高", "YJV-5×2.5→5芯2.5mm²"),
    33: (["C4-8-237"], "中", "YJV-3×2.5+2×1.5→按3芯取"),
    34: (["C4-8-241"], "高", "YJV-5×2.5→5芯2.5mm²"),

    # === 配线 (35-37) ===
    # 管内穿铜芯线，照明线路
    # C4-11-282: ≤2.5mm², C4-11-283: ≤4mm², C4-11-284: ≤6mm²
    35: (["C4-11-282"], "高", "BV-2.5mm²，照明线路"),
    36: (["C4-11-282"], "高", "BV-2.5mm²，照明线路"),
    37: (["C4-11-283"], "高", "BV-4mm²，照明线路"),

    # === 配管 (38-44) ===
    # JDG管→紧定式薄壁钢管，SC管→焊接钢管
    # C4-11-121: JDG DN20, C4-11-122: JDG DN25, C4-11-123: JDG DN32
    # C4-11-23~31: 焊接钢管明配/暗配
    38: (["C4-11-121"], "高", "JDG管 DN20 暗敷"),
    39: (["C4-11-122"], "高", "JDG管 DN25 暗敷"),
    40: (["C4-11-123"], "高", "JDG管 DN32 暗敷"),
    41: (["C4-11-25"], "高", "SC管 DN25 暗敷(焊接钢管)"),
    42: (["C4-11-27"], "高", "SC管 DN40 暗敷(焊接钢管)"),
    43: (["C4-11-29"], "高", "SC管 DN50 暗敷(焊接钢管)"),
    44: (["C4-11-31"], "高", "SC管 DN70 暗敷(焊接钢管)"),

    # === 桥架 (45-46) ===
    # C4-11-249: 钢制槽式桥架 ≤400mm宽
    45: (["C4-11-249"], "高", "槽式桥架 200×100"),
    46: (["C4-11-249"], "高", "槽式桥架 400×150"),

    # === 支吊架 (47) ===
    # C4-13-1: 支架制作, C4-13-2: 支架安装
    47: (["C4-13-1", "C4-13-2"], "高", "金属支架制作+安装"),

    # === 刷油 (48) ===
    # C12-2-61: 金属结构刷油
    48: (["C12-2-61"], "高", "金属结构刷油"),

    # === 灯具 (49-67) ===
    # C4-12-21: 嵌入式灯(方/圆), C4-12-9: 壁灯, C4-12-213~216: 嵌入式荧光灯
    # C4-12-65~69: 标志灯/诱导灯
    49: (["C4-12-21"], "高", "嵌入式LED面板灯 300×600"),
    50: (["C4-12-21"], "高", "嵌入式LED面板灯 600×600"),
    51: (["C4-12-21"], "高", "嵌入式LED面板灯 300×300"),
    52: (["C4-12-21"], "高", "嵌入式LED筒灯"),
    53: (["C4-12-21"], "高", "嵌入式LED筒灯"),
    54: (["C4-12-9"], "高", "LED壁灯"),
    55: (["C4-12-9"], "高", "LED壁灯"),
    56: (["C4-12-67"], "高", "壁式安全出口标志灯"),
    57: (["C4-12-67"], "高", "壁式单向疏散标志灯"),
    58: (["C4-12-67"], "高", "壁式双向疏散标志灯"),
    59: (["C4-12-65"], "高", "吊装双面标志灯"),
    60: (["C4-12-65"], "高", "吊装安全出口标志灯"),
    61: (["C4-12-67"], "高", "壁式楼层标志灯"),
    62: (["C4-12-21"], "中", "嵌入式应急照明灯"),
    63: (["C4-12-67"], "高", "壁式应急照明灯"),
    64: (["C4-12-21"], "高", "嵌入式LED格栅灯 600×600"),
    65: (["C4-12-21"], "高", "嵌入式LED格栅灯 300×1200"),
    66: (["C4-12-21"], "高", "嵌入式LED射灯"),
    67: (["C4-12-21"], "高", "嵌入式LED灯带"),

    # === 开关 (68-72) ===
    # C4-4-103: 单控单联暗开关, C4-4-104: 单控双联, C4-4-108: 双控单联
    # C4-4-122: 插座暗装
    68: (["C4-4-103"], "高", "单极开关 暗装 1位"),
    69: (["C4-4-108"], "高", "双控开关 暗装 1位"),
    70: (["C4-4-103"], "中", "紧急按钮 暗装"),
    71: (["C4-4-122"], "高", "五孔插座 暗装"),
    72: (["C4-4-103"], "中", "按钮开关 暗装"),

    # === 接线盒 (73, 76-77) ===
    # C4-11-382: 暗装接线盒 ≤100×100
    # C4-11-383: 暗装接线盒 ≤150×150
    73: (["C4-11-382"], "高", "明装接线盒"),
    76: (["C4-11-382"], "高", "暗装接线盒"),
    77: (["C4-11-384"], "中", "暗装过路盒 大尺寸"),

    # === 控制器 (74-75) ===
    # 智能照明控制器不太好找标准定额，用配电箱类或调试类
    74: (["C4-4-31"], "低", "智能照明控制器(可编程)→按小型配电箱"),
    75: (["C4-4-30"], "低", "智能照明控制面板→按小型配电箱"),

    # === 防雷接地 (78-80) ===
    # C4-9-11: 接地母线 ≤40×4, C4-9-68: 引下线 利用柱筋
    # C4-9-79: 接闪网/带
    78: (["C4-9-11"], "高", "接地母线 -40×4 明敷"),
    79: (["C4-9-68"], "高", "防雷引下线 利用柱内钢筋"),
    80: (["C4-9-79"], "高", "接闪带/网 圆钢Φ10"),

    # === 电动机 (81-86) ===
    # C4-6-1: ≤3kW, C4-6-2: ≤7.5kW, C4-6-3: ≤15kW, C4-6-4: ≤22kW
    81: (["C4-6-2"], "高", "排烟风机 5.5kW→取7.5档"),
    82: (["C4-6-1"], "高", "排烟风机 1.1kW→取3档"),
    83: (["C4-6-1"], "高", "排烟补风机 0.55kW→取3档"),
    84: (["C4-6-1"], "高", "排烟补风机 0.75kW→取3档"),
    85: (["C4-6-1"], "高", "空调末端风机 0.37kW→取3档"),
    86: (["C4-6-4"], "中", "变频冷水机组 较大功率→按≤22kW"),

    # === 防火封堵 (87-89) ===
    # C4-8-493: 防火堵洞 ≤0.1m², C4-8-494: ≤0.5m², C4-8-495: ≤1m²
    87: (["C4-8-493"], "高", "防火封堵 ≤0.1m²"),
    88: (["C4-8-494"], "高", "防火封堵 ≤0.5m²"),
    89: (["C4-8-495"], "高", "防火封堵 ≤1.0m²"),
}


def get_quota_names(db_path, quota_ids):
    """从定额数据库查询定额名称"""
    conn = sqlite3.connect(db_path)
    result = {}
    for qid in quota_ids:
        row = conn.execute(
            "SELECT quota_id, name FROM quotas WHERE quota_id = ?", (qid,)
        ).fetchone()
        if row:
            result[qid] = row[1]
        else:
            result[qid] = "(未找到)"
    conn.close()
    return result


def read_bill_items(file_path, sheet_name):
    """读取清单项目"""
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb[sheet_name]

    items = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 3:  # 跳过表头（前3行）
            continue
        a_val = row[0] if len(row) > 0 else None  # 序号
        if a_val is None:
            continue
        try:
            idx = int(a_val)
        except (ValueError, TypeError):
            continue

        items.append({
            "index": idx,
            "code": str(row[1] or "").strip() if len(row) > 1 else "",
            "name": str(row[2] or "").strip() if len(row) > 2 else "",
            "description": str(row[3] or "").strip() if len(row) > 3 else "",
            "unit": str(row[4] or "").strip() if len(row) > 4 else "",
            "quantity": row[5] if len(row) > 5 else None,
        })

    wb.close()
    return items


def main():
    print("读取清单数据...")
    items = read_bill_items(BILL_FILE, SHEET_NAME)
    print(f"  共 {len(items)} 条清单项")

    # 收集所有用到的定额编号
    all_quota_ids = set()
    for rule in MATCH_RULES.values():
        all_quota_ids.update(rule[0])

    print(f"查询定额名称（{len(all_quota_ids)} 条）...")
    quota_names = get_quota_names(QUOTA_DB, all_quota_ids)

    # 创建输出Excel
    print("生成Excel...")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Opus4.6匹配结果"

    # 样式定义
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    green_fill = PatternFill("solid", fgColor="C6EFCE")   # 高置信度
    yellow_fill = PatternFill("solid", fgColor="FFEB9C")  # 中置信度
    red_fill = PatternFill("solid", fgColor="FFC7CE")     # 低置信度
    wrap_alignment = Alignment(wrap_text=True, vertical="top")

    # 写表头
    headers = [
        "序号", "项目编码", "项目名称", "项目特征描述",
        "单位", "工程量",
        "定额编号", "定额名称", "置信度", "匹配说明"
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 设置列宽
    col_widths = [6, 16, 20, 40, 6, 10, 14, 35, 8, 30]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # 写数据行
    row_num = 2
    for item in items:
        idx = item["index"]
        rule = MATCH_RULES.get(idx)

        if rule is None:
            # 没有匹配规则的项目
            quota_ids_str = ""
            quota_names_str = ""
            confidence = "未匹配"
            note = ""
        else:
            quota_id_list, confidence, note = rule
            quota_ids_str = "\n".join(quota_id_list)
            names_list = [quota_names.get(qid, "(未找到)") for qid in quota_id_list]
            quota_names_str = "\n".join(names_list)

        ws.cell(row=row_num, column=1, value=idx).border = thin_border
        ws.cell(row=row_num, column=2, value=item["code"]).border = thin_border
        ws.cell(row=row_num, column=3, value=item["name"]).border = thin_border

        desc_cell = ws.cell(row=row_num, column=4, value=item["description"])
        desc_cell.border = thin_border
        desc_cell.alignment = wrap_alignment

        ws.cell(row=row_num, column=5, value=item["unit"]).border = thin_border
        ws.cell(row=row_num, column=6, value=item["quantity"]).border = thin_border

        qid_cell = ws.cell(row=row_num, column=7, value=quota_ids_str)
        qid_cell.border = thin_border
        qid_cell.alignment = wrap_alignment

        qname_cell = ws.cell(row=row_num, column=8, value=quota_names_str)
        qname_cell.border = thin_border
        qname_cell.alignment = wrap_alignment

        conf_cell = ws.cell(row=row_num, column=9, value=confidence)
        conf_cell.border = thin_border
        conf_cell.alignment = Alignment(horizontal="center", vertical="center")

        # 根据置信度设置颜色
        if confidence == "高":
            conf_cell.fill = green_fill
        elif confidence == "中":
            conf_cell.fill = yellow_fill
        elif confidence == "低":
            conf_cell.fill = red_fill

        note_cell = ws.cell(row=row_num, column=10, value=note)
        note_cell.border = thin_border
        note_cell.alignment = wrap_alignment

        row_num += 1

    # 冻结首行
    ws.freeze_panes = "A2"

    # 添加统计Sheet
    ws2 = wb.create_sheet("匹配统计")
    stats = {"高": 0, "中": 0, "低": 0, "未匹配": 0}
    for item in items:
        idx = item["index"]
        rule = MATCH_RULES.get(idx)
        if rule:
            stats[rule[1]] += 1
        else:
            stats["未匹配"] += 1

    ws2.cell(row=1, column=1, value="置信度").font = header_font
    ws2.cell(row=1, column=2, value="数量").font = header_font
    ws2.cell(row=1, column=3, value="占比").font = header_font

    total = len(items)
    for i, (level, count) in enumerate(stats.items(), 2):
        ws2.cell(row=i, column=1, value=level)
        ws2.cell(row=i, column=2, value=count)
        ws2.cell(row=i, column=3, value=f"{count * 100 / total:.1f}%")

    ws2.cell(row=6, column=1, value="总计").font = header_font
    ws2.cell(row=6, column=2, value=total).font = header_font

    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 10
    ws2.column_dimensions["C"].width = 10

    # 保存
    wb.save(OUTPUT_FILE)
    print(f"\n完成！结果已保存到: {OUTPUT_FILE}")
    print(f"  总计: {total} 条")
    print(f"  高置信度: {stats['高']} 条 ({stats['高'] * 100 / total:.1f}%)")
    print(f"  中置信度: {stats['中']} 条 ({stats['中'] * 100 / total:.1f}%)")
    print(f"  低置信度: {stats['低']} 条 ({stats['低'] * 100 / total:.1f}%)")
    print(f"  未匹配: {stats['未匹配']} 条")


if __name__ == "__main__":
    main()
