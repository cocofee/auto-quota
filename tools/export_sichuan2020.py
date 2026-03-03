# -*- coding: utf-8 -*-
"""
导出四川2020所有定额到Excel
输出目录: data/quota_data/四川/四川省2020序列定额/
"""
import sys
import os
import re
import openpyxl
import xml.etree.ElementTree as ET
from collections import defaultdict

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from tools.export_quota_excel import parse_index_file

# 配置
DB_BASE = r"D:\广联达\数据库SC30.0\四川\定额库\四川省2020序列定额\定额库"
OUTPUT_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                          "data", "quota_data", "四川", "四川省2020序列定额")


def simplify_name(full_name):
    """从完整定额名提取简称"""
    name = full_name
    name = re.sub(r'^四川省', '', name)
    name = re.sub(r'量清单计价定额.*$', '', name)
    name = re.sub(r'预算定额.*$', '', name)
    if not name.endswith('工程'):
        name += '工程'
    return name


def get_book_letter_map(xml_path):
    """从子目专业.xml获取册名映射 {'A': '机械设备安装工程', ...}"""
    if not os.path.isfile(xml_path):
        return {}
    tree = ET.parse(xml_path)
    root = tree.getroot()
    result = {}
    # 四川的册用字母编号 A,B,C,...,H,J,K,...（跳过I）
    letters = 'ABCDEFGHJKLMNPQRSTUVWXYZ'
    for i, record in enumerate(root.findall('Record')):
        desc = record.get('Description', '')
        # 去掉前面的字母前缀（如 "A 机械设备安装工程" → "机械设备安装工程"）
        clean = re.sub(r'^[A-Z]\s+', '', desc)
        if i < len(letters):
            result[letters[i]] = clean
    return result


def clean_sheet_name(name):
    """清理Excel Sheet名中的非法字符"""
    bad_chars = ['[', ']', ':', '*', '?', '/', '\\']
    for c in bad_chars:
        name = name.replace(c, '')
    return name[:31]  # Excel限制31字符


def clean_cell_value(text):
    """清理单元格内容中的非法字符（Excel不允许控制字符）"""
    if not text:
        return text
    # 去掉ASCII控制字符（0x00-0x1F，除了\t\n\r）
    return re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', text)


def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    total_files = 0
    total_records = 0

    for quota_name in sorted(os.listdir(DB_BASE)):
        quota_path = os.path.join(DB_BASE, quota_name)
        idx_path = os.path.join(quota_path, "数据", "子目索引.Index")
        xml_path = os.path.join(quota_path, "基础数据", "子目专业.xml")

        if not os.path.isfile(idx_path):
            continue

        # 1. 解析Index
        records = parse_index_file(idx_path)
        if not records:
            print(f"跳过（无记录）: {quota_name}")
            continue

        # 2. 获取册名
        book_names = get_book_letter_map(xml_path)

        # 3. 按第2个字母（册）分组
        books = defaultdict(list)
        for code, name in records:
            if len(code) >= 2:
                books[code[1]].append((code, name))
            else:
                books['_'].append((code, name))

        # 4. 生成Excel
        short_name = simplify_name(quota_name)
        output_path = os.path.join(OUTPUT_DIR, short_name + ".xlsx")

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        sheet_idx = 0
        entry_count = 0
        for book_letter in sorted(books.keys()):
            entries = books[book_letter]
            entry_count += len(entries)
            sheet_idx += 1

            # Sheet名称: "001_A 机械设备安装工程"
            bn = book_names.get(book_letter, '')
            if bn:
                sheet_title = f"{sheet_idx:03d}_{book_letter} {bn}"
            else:
                sheet_title = f"{sheet_idx:03d}_{book_letter}"
            sheet_title = clean_sheet_name(sheet_title)

            ws = wb.create_sheet(title=sheet_title)

            # 无表头，直接写数据（和已有数据一致）
            for row_idx, (code, name) in enumerate(entries, 1):
                ws.cell(row=row_idx, column=1, value=clean_cell_value(code))
                ws.cell(row=row_idx, column=2, value=clean_cell_value(name))

            # 列宽
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 60

        wb.save(output_path)
        print(f"  {short_name}.xlsx  ({len(books)}个Sheet, {entry_count}条)")
        total_files += 1
        total_records += entry_count

    print()
    print(f"完成: {total_files}个文件, {total_records}条记录")
    print(f"输出: {OUTPUT_DIR}")


if __name__ == "__main__":
    main()
