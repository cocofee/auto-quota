"""
造价文件收集工具

功能：扫描微信/企业微信/自定义目录，自动筛选计价相关文件，按专业分类收集。
支持增量模式：记住已处理过的文件，下次只看新增/修改的文件。

用法：
    python tools/collect_wechat_files.py                    # 增量收集微信（只看新文件）
    python tools/collect_wechat_files.py --full             # 全量重扫（忽略历史）
    python tools/collect_wechat_files.py --preview          # 只统计不复制
    python tools/collect_wechat_files.py --source wechat    # 只扫微信
    python tools/collect_wechat_files.py --source wxwork    # 只扫企业微信
    python tools/collect_wechat_files.py --dir "D:\\广联达临时文件"  # 扫自定义目录

目录结构：
    F:\\jarvis\\
      土建装饰\\      给排水\\      电气\\      消防\\
      通风空调\\      智能化\\      市政\\      园林景观\\
      钢结构幕墙\\    电力\\        综合\\
"""

import os
import sys
import re
import json
import shutil
import hashlib
import argparse
import tempfile
import zipfile
import atexit
from pathlib import Path
from datetime import datetime
from collections import defaultdict

# 模块级变量：记录所有解压临时目录，atexit 兜底清理（防止异常退出时残留）
_all_temp_dirs: list[str] = []


def _cleanup_temp_dirs():
    """进程退出时兜底清理临时目录（防止C盘被撑满、防止火绒误报exe）"""
    for td in _all_temp_dirs:
        try:
            shutil.rmtree(td, ignore_errors=True)
        except Exception:
            pass
    _all_temp_dirs.clear()


atexit.register(_cleanup_temp_dirs)

# 增量记录文件路径（记住已处理过的文件，避免重复扫描）
_HISTORY_FILE = os.path.join(os.path.dirname(__file__), ".collect_history.json")


def _load_history():
    """加载增量历史记录，格式: {文件路径: 修改时间戳}"""
    if os.path.exists(_HISTORY_FILE):
        try:
            with open(_HISTORY_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {}


def _save_history(history):
    """保存增量历史记录"""
    try:
        with open(_HISTORY_FILE, "w", encoding="utf-8") as f:
            json.dump(history, f, ensure_ascii=False)
    except Exception as e:
        print(f"  [警告] 保存增量记录失败: {e}")

# ============================================================
# 配置
# ============================================================

# 数据来源目录
SOURCES = {
    "wechat": r"F:\微信聊天\xwechat_files\summerfly2008_9de8",
    "wxwork": r"C:\Users\Administrator\Documents\WXWork",
}

DEFAULT_OUTPUT = r"F:\jarvis"

# 广联达及行业造价软件文件扩展名
SOFTWARE_EXTENSIONS = {
    # 广联达计价（核心）
    ".gbq6", ".gbq7", ".gbq9", ".gbq",  # 计价文件
    ".gpb6", ".gpb7", ".gpb",            # 投标报价
    ".zbqd",                              # 清单组价
    ".qdg4", ".qdg",                      # 清单文件
    ".gczjwj",                            # 广联达XML导出
    # 广联达行业软件（电力/市政等）
    ".gczj",                              # 工程造价
    ".gad",                               # 广联达安装
    ".zjxm",                              # 造价项目
    ".ysq",                               # 预算清
    ".ygl",                               # 预概算
    ".yfjz",                              # 预付结转
    ".bj23",                              # 北京2023
    ".sxzb4",                             # 陕西
    ".spw",                               # 广联达水利/电力
    ".bsj",                               # 博赛计价
    ".13jz",                              # 新点13计价
    # 其他造价软件
    ".e2d",                               # 新点计价
    ".pbq",                               # 品茗计价
    # 不收集的：.gqi4(概算索引) .gtj(图形算量) .gcl(算量) .gcfx(分析)
}

EXCEL_EXTENSIONS = {".xlsx", ".xls", ".xlsm"}
ARCHIVE_EXTENSIONS = {".zip", ".rar", ".7z"}

# 12位标准清单编码正则：0开头，12位数字（如031001008004）
BILL_CODE_PATTERN = re.compile(r'^0[1-9]\d{10}$')

# 非造价关键词（文件名包含这些的直接排除）
EXCLUDE_KEYWORDS = [
    "田径", "竞走", "越野赛", "越野跑", "马拉松", "激光跑", "接力",
    "精英组", "公开组", "甲组", "乙组", "成绩册", "号码簿",
    "运动会", "体育", "athlete", "certificate", "import-template",
    "芯片", "火腿", "善捐", "法会", "民谣", "五金", "直播计划",
    "报名表", "考勤", "工资表", "薪资",
    "GCPSetup", "Windows11", "WPS VBA", "Setup", "安装包",
]

# ============================================================
# 专业分类规则（核心）
# ============================================================

# 专业名称 → 关键词列表（按优先级排序，越靠前越优先匹配）
SPECIALTY_RULES = [
    # 先匹配具体专业（避免被"安装"这种大词吃掉）
    ("消防", [
        "消防", "喷淋", "喷头", "消火栓", "报警", "灭火",
        "气体灭火", "防排烟", "消防水",
    ]),
    ("通风空调", [
        "通风", "空调", "暖通", "风管", "风机", "冷冻",
        "新风", "排风", "送风", "制冷", "采暖", "供暖", "地暖",
    ]),
    ("智能化", [
        "智能化", "弱电", "安防", "监控", "综合布线", "楼宇对讲",
        "门禁", "广播", "网络", "信息化", "自控", "BA系统",
    ]),
    ("给排水", [
        "给排水", "给水", "排水", "管道安装", "水泵", "阀门",
        "卫生器具", "热水", "中水", "雨水", "污水",
        "管道", "水管", "镀锌钢管", "PPR",
    ]),
    ("电气", [
        "电气", "配电", "电缆", "照明", "灯具", "开关", "插座",
        "桥架", "配管穿线", "配电箱", "动力", "防雷接地", "防雷",
        "变配电", "电力安装",
    ]),
    ("电力", [
        "变电站", "变电所", "输电线路", "输电", "10kV", "10KV",
        "35kV", "35KV", "110kV", "220kV", "光伏", "风电",
        "升压站", "箱变", "变压器",
    ]),
    ("钢结构幕墙", [
        "钢结构", "幕墙", "玻璃幕墙", "石材幕墙", "铝板幕墙",
        "门窗工程", "门窗",
    ]),
    ("园林景观", [
        "园林", "绿化", "景观", "种植", "苗木", "草坪",
        "园建", "铺装",
    ]),
    ("市政", [
        "市政", "道路", "桥梁", "路面", "路基", "管网",
        "排水管网", "雨污分流", "交通", "隧道",
    ]),
    ("土建装饰", [
        "土建", "装饰", "装修", "混凝土", "钢筋", "模板",
        "砌体", "砌筑", "防水", "屋面", "保温", "涂料",
        "吊顶", "地面", "墙面", "抹灰", "油漆", "刷油",
        "基础", "地基", "桩基", "基坑", "土方",
        "拆除", "脚手架", "外墙",
    ]),
]

# Excel中识别为造价文件的列名组合（至少命中一个组合中>=2个关键词）
COST_COLUMN_GROUPS = [
    ["项目名称", "项目编码", "项目特征", "工程量", "计量单位", "综合单价"],
    ["定额编号", "定额名称", "子目", "人工费", "材料费", "机械费"],
    ["分部分项", "措施项目", "合价", "工程量清单"],
    ["工程名称", "综合单价", "合价", "工程量"],
    ["材料名称", "规格型号", "单价", "合价"],
    ["设备名称", "规格型号", "单价", "合价"],
    ["结算金额", "送审金额", "审定金额", "工程量"],
]


# ============================================================
# 工具函数
# ============================================================

def get_file_md5(filepath, chunk_size=8192):
    """计算文件MD5用于去重"""
    h = hashlib.md5()
    try:
        with open(filepath, "rb") as f:
            while True:
                chunk = f.read(chunk_size)
                if not chunk:
                    break
                h.update(chunk)
        return h.hexdigest()
    except Exception:
        return None


def is_excluded_by_name(filename):
    """文件名包含非造价关键词则排除"""
    name_lower = filename.lower()
    for kw in EXCLUDE_KEYWORDS:
        if kw.lower() in name_lower:
            return True
    return False


def detect_specialty(text):
    """
    从文本中识别专业类别。
    text 可以是文件名、Sheet名、或Excel内容。
    返回专业名称字符串，匹配不到返回None。
    """
    for specialty, keywords in SPECIALTY_RULES:
        for kw in keywords:
            if kw in text:
                return specialty
    return None


def detect_specialty_from_excel(filepath):
    """
    打开Excel，综合文件名+Sheet名+内容判断专业。
    返回: (is_cost_file, specialty)
      is_cost_file: 是否为造价文件
      specialty: 专业名称（"综合"表示无法判断）
    """
    ext = Path(filepath).suffix.lower()
    filename = Path(filepath).stem

    # 第1步：先看文件名能不能判断专业
    spec_from_name = detect_specialty(filename)

    # 第2步：打开Excel检查内容
    try:
        if ext == ".xls":
            is_cost, spec_from_content, sheet_specs = _analyze_xls(filepath)
        else:
            is_cost, spec_from_content, sheet_specs = _analyze_xlsx(filepath)
    except Exception:
        # 打不开的文件，只有文件名含强信号词才收
        strong_kw = ["清单", "预算", "定额", "计价", "造价", "概算", "结算"]
        if any(kw in filename for kw in strong_kw):
            return True, spec_from_name or "综合"
        return False, None

    if not is_cost:
        return False, None

    # 第3步：确定专业（优先级：Sheet名 > 文件名 > 内容）
    # 如果有多个Sheet各自是不同专业，说明是综合文件
    if len(sheet_specs) > 1:
        return True, "综合"

    # 单一专业
    spec = None
    if sheet_specs:
        spec = list(sheet_specs)[0]
    if not spec:
        spec = spec_from_name
    if not spec:
        spec = spec_from_content
    if not spec:
        spec = "综合"

    return True, spec


def _check_columns_match(text):
    """检查Excel内容是否命中列名组合（>=2个关键词）"""
    for group in COST_COLUMN_GROUPS:
        hit = sum(1 for kw in group if kw in text)
        if hit >= 2:
            return True
    return False


def _analyze_xlsx(filepath):
    """
    分析xlsx文件。
    返回: (is_cost, content_specialty, sheet_specialties_set)
    """
    import openpyxl
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    is_cost = False
    content_spec = None
    sheet_specs = set()

    try:
        for sheet_name in wb.sheetnames:
            # Sheet名本身判断专业
            s = detect_specialty(sheet_name)
            if s:
                sheet_specs.add(s)

            # Sheet名是造价专用的
            if sheet_name in ("分部分项", "措施项目", "清单", "汇总"):
                is_cost = True

            ws = wb[sheet_name]
            all_values = []
            for row_idx, row in enumerate(ws.iter_rows(max_row=15, values_only=True)):
                for cell in row:
                    if cell is not None:
                        all_values.append(str(cell).strip())

            joined = " ".join(all_values)

            # 检查是否为造价文件
            if _check_columns_match(joined):
                is_cost = True

            # 从内容中识别专业
            if not content_spec:
                content_spec = detect_specialty(joined)

    finally:
        wb.close()

    return is_cost, content_spec, sheet_specs


def _analyze_xls(filepath):
    """分析xls文件（同xlsx逻辑）"""
    import xlrd
    wb = xlrd.open_workbook(filepath, on_demand=True)
    is_cost = False
    content_spec = None
    sheet_specs = set()

    try:
        for sheet_name in wb.sheet_names():
            s = detect_specialty(sheet_name)
            if s:
                sheet_specs.add(s)

            if sheet_name in ("分部分项", "措施项目", "清单", "汇总"):
                is_cost = True

            ws = wb.sheet_by_name(sheet_name)
            all_values = []
            for row_idx in range(min(15, ws.nrows)):
                for col_idx in range(min(20, ws.ncols)):
                    val = ws.cell_value(row_idx, col_idx)
                    if val:
                        all_values.append(str(val).strip())

            joined = " ".join(all_values)

            if _check_columns_match(joined):
                is_cost = True

            if not content_spec:
                content_spec = detect_specialty(joined)

    finally:
        wb.release_resources()

    return is_cost, content_spec, sheet_specs


def has_standard_bill_codes(filepath):
    """检查Excel是否包含标准12位清单编码

    遍历所有Sheet，检查每行前5列，找到至少1个12位编码就算标准清单。
    只检查前200行（清单通常在前面，避免大文件太慢）。
    """
    ext = Path(filepath).suffix.lower()
    try:
        if ext == ".xls":
            return _check_bill_codes_xls(filepath)
        else:
            return _check_bill_codes_xlsx(filepath)
    except Exception:
        return False


def _check_bill_codes_xlsx(filepath):
    """检查xlsx文件中是否有12位清单编码"""
    import openpyxl
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    try:
        for ws in wb.worksheets:
            row_count = 0
            try:
                for row in ws.iter_rows(values_only=True):
                    row_count += 1
                    if row_count > 200:
                        break
                    for cell in row[:5]:
                        if cell is None:
                            continue
                        val = str(cell).strip()
                        if BILL_CODE_PATTERN.match(val):
                            return True
            except Exception:
                continue
    finally:
        wb.close()
    return False


def _check_bill_codes_xls(filepath):
    """检查xls文件中是否有12位清单编码"""
    import xlrd
    wb = xlrd.open_workbook(filepath, on_demand=True)
    try:
        for sheet_name in wb.sheet_names():
            ws = wb.sheet_by_name(sheet_name)
            for row_idx in range(min(200, ws.nrows)):
                for col_idx in range(min(5, ws.ncols)):
                    val = str(ws.cell_value(row_idx, col_idx)).strip()
                    if BILL_CODE_PATTERN.match(val):
                        return True
    finally:
        wb.release_resources()
    return False


def detect_specialty_for_software(filepath):
    """对广联达等软件文件，只能从文件名判断专业"""
    spec = detect_specialty(Path(filepath).stem)
    return spec or "综合"


# ============================================================
# 压缩包处理
# ============================================================

def extract_archive(archive_path, temp_dir):
    """解压压缩包，返回解压出的文件列表"""
    ext = Path(archive_path).suffix.lower()
    try:
        if ext == ".zip":
            return _extract_zip(archive_path, temp_dir)
        elif ext == ".7z":
            return _extract_7z(archive_path, temp_dir)
        elif ext == ".rar":
            return _extract_rar(archive_path, temp_dir)
    except Exception as e:
        print(f"  [警告] 解压失败 {Path(archive_path).name}: {e}")
    return []


def _extract_zip(archive_path, temp_dir):
    files = []
    with zipfile.ZipFile(archive_path, 'r') as zf:
        for info in zf.infolist():
            if info.is_dir():
                continue
            try:
                fname = info.filename.encode('cp437').decode('gbk')
            except (UnicodeDecodeError, UnicodeEncodeError):
                fname = info.filename

            safe_name = Path(fname).name
            if not safe_name:
                continue

            out_path = os.path.join(temp_dir, safe_name)
            counter = 1
            base, ext = os.path.splitext(out_path)
            while os.path.exists(out_path):
                out_path = f"{base}({counter}){ext}"
                counter += 1

            try:
                with zf.open(info) as src, open(out_path, 'wb') as dst:
                    shutil.copyfileobj(src, dst)
                files.append(out_path)
            except Exception:
                pass
    return files


def _extract_7z(archive_path, temp_dir):
    import py7zr
    with py7zr.SevenZipFile(archive_path, 'r') as z:
        z.extractall(path=temp_dir)
    files = []
    for root, dirs, fnames in os.walk(temp_dir):
        for fname in fnames:
            files.append(os.path.join(root, fname))
    return files


def _extract_rar(archive_path, temp_dir):
    import subprocess
    seven_zip = r"C:\Program Files\7-Zip\7z.exe"
    if not os.path.exists(seven_zip):
        return []
    subprocess.run(
        [seven_zip, "x", "-y", f"-o{temp_dir}", archive_path],
        capture_output=True, timeout=60
    )
    files = []
    for root, dirs, fnames in os.walk(temp_dir):
        for fname in fnames:
            files.append(os.path.join(root, fname))
    return files


# ============================================================
# 文件复制（去重+防重名）
# ============================================================

def safe_copy(src, dst_dir, source_tag, seen_md5s):
    """复制文件到目标目录，MD5去重，防重名。返回目标路径或None"""
    md5 = get_file_md5(src)
    if md5 and md5 in seen_md5s:
        return None
    if md5:
        seen_md5s.add(md5)

    stem = Path(src).stem
    ext = Path(src).suffix
    # Windows路径限制260字符，截断过长文件名
    max_name_len = 200 - len(dst_dir) - len(source_tag) - len(ext) - 5
    if max_name_len < 20:
        max_name_len = 20
    if len(stem) > max_name_len:
        stem = stem[:max_name_len]
    dst_name = f"{stem}_{source_tag}{ext}"
    dst_path = os.path.join(dst_dir, dst_name)

    counter = 2
    while os.path.exists(dst_path):
        dst_name = f"{stem}_{source_tag}({counter}){ext}"
        dst_path = os.path.join(dst_dir, dst_name)
        counter += 1

    os.makedirs(dst_dir, exist_ok=True)
    try:
        shutil.copy2(src, dst_path)
        return dst_path
    except Exception:
        return None


# ============================================================
# 主流程
# ============================================================

def collect_files(output_dir, sources_to_scan, preview=False, full=False,
                  extra_dirs=None):
    """主收集流程

    参数:
        output_dir: 输出目录（F:\\jarvis）
        sources_to_scan: 要扫描的来源列表（wechat/wxwork）
        preview: 预览模式（只统计不复制）
        full: 全量模式（忽略增量历史，从头扫）
        extra_dirs: 额外的自定义目录列表，如 ["D:\\广联达临时文件"]
    """

    print("=" * 60)
    print("  造价文件收集工具（按专业分类）")
    print("=" * 60)

    # 加载增量历史（记录已处理过的文件路径+修改时间）
    if full:
        history = {}
        print("  [全量模式] 忽略历史记录，从头扫描\n")
    else:
        history = _load_history()
        print(f"  [增量模式] 已有 {len(history)} 条历史记录，只扫新文件\n")

    if preview:
        print("  [预览模式] 只统计，不复制文件\n")

    # 统计: stats[专业][类型] = count
    stats = defaultdict(lambda: defaultdict(int))
    seen_md5s = set()

    # 收集结果: [(filepath, specialty, source_tag), ...]
    collected = []
    excluded_count = 0
    skipped_count = 0  # 增量跳过的文件数

    # ==== 第1步：扫描所有来源目录（增量：跳过已处理的文件） ====
    all_excel = []      # (filepath, tag)
    all_software = []   # (filepath, tag)
    all_archives = []   # (filepath, tag)

    for source_name, source_dir in SOURCES.items():
        if source_name not in sources_to_scan:
            continue
        if not os.path.exists(source_dir):
            print(f"[警告] 目录不存在: {source_dir}")
            continue

        tag = "wx" if source_name == "wechat" else "wxwork"
        print(f"\n[扫描] {tag}: {source_dir}")

        file_count = 0
        new_count = 0
        for root, dirs, files in os.walk(source_dir):
            for fname in files:
                filepath = os.path.join(root, fname)
                ext = Path(fname).suffix.lower()
                file_count += 1

                if file_count % 5000 == 0:
                    print(f"  已扫描 {file_count} 个文件（新增 {new_count}）...")

                # 增量判断：文件路径+修改时间都没变 → 跳过
                try:
                    mtime = os.path.getmtime(filepath)
                except OSError:
                    continue
                old_mtime = history.get(filepath)
                if old_mtime is not None and abs(mtime - old_mtime) < 1:
                    skipped_count += 1
                    continue

                # 记录这个文件（不管是不是造价文件都记，下次直接跳过）
                history[filepath] = mtime

                if is_excluded_by_name(fname):
                    continue

                new_count += 1
                if ext in SOFTWARE_EXTENSIONS:
                    all_software.append((filepath, tag))
                elif ext in EXCEL_EXTENSIONS:
                    all_excel.append((filepath, tag))
                elif ext in ARCHIVE_EXTENSIONS:
                    all_archives.append((filepath, tag))

        print(f"  扫描完成: 总{file_count} 个文件，新增 {new_count}，跳过 {file_count - new_count}")
        print(f"    新增 Excel: {len([x for x in all_excel if x[1]==tag])} | "
              f"软件: {len([x for x in all_software if x[1]==tag])} | "
              f"压缩包: {len([x for x in all_archives if x[1]==tag])}")

    # ==== 第1.5步：扫描自定义目录（和微信逻辑完全相同） ====
    if extra_dirs:
        for extra_dir in extra_dirs:
            if not os.path.exists(extra_dir):
                print(f"\n[警告] 目录不存在: {extra_dir}")
                continue

            tag = "local"  # 来源标签：本地目录
            print(f"\n[扫描] local: {extra_dir}")

            file_count = 0
            new_count = 0
            for root, dirs, files in os.walk(extra_dir):
                for fname in files:
                    filepath = os.path.join(root, fname)
                    ext = Path(fname).suffix.lower()
                    file_count += 1

                    if file_count % 5000 == 0:
                        print(f"  已扫描 {file_count} 个文件（新增 {new_count}）...")

                    # 增量判断
                    try:
                        mtime = os.path.getmtime(filepath)
                    except OSError:
                        continue
                    old_mtime = history.get(filepath)
                    if old_mtime is not None and abs(mtime - old_mtime) < 1:
                        skipped_count += 1
                        continue

                    history[filepath] = mtime

                    if is_excluded_by_name(fname):
                        continue

                    new_count += 1
                    if ext in SOFTWARE_EXTENSIONS:
                        all_software.append((filepath, tag))
                    elif ext in EXCEL_EXTENSIONS:
                        all_excel.append((filepath, tag))
                    elif ext in ARCHIVE_EXTENSIONS:
                        all_archives.append((filepath, tag))

            print(f"  扫描完成: 总{file_count} 个文件，新增 {new_count}，跳过 {file_count - new_count}")
            print(f"    新增 Excel: {len([x for x in all_excel if x[1]==tag])} | "
                  f"软件: {len([x for x in all_software if x[1]==tag])} | "
                  f"压缩包: {len([x for x in all_archives if x[1]==tag])}")

    # ==== 第2步：处理压缩包 ====
    _temp_dirs = []  # 记录所有临时目录，最后统一清理
    if all_archives:
        total_arc = len(all_archives)
        print(f"\n[解压] 处理 {total_arc} 个压缩包...")

        for idx, (archive_path, tag) in enumerate(all_archives):
            if (idx + 1) % 50 == 0:
                print(f"  解压 {idx+1}/{total_arc}...")

            temp_dir = tempfile.mkdtemp(prefix="jarvis_")
            _temp_dirs.append(temp_dir)
            _all_temp_dirs.append(temp_dir)  # atexit 兜底清理
            try:
                for fpath in extract_archive(archive_path, temp_dir):
                    ext = Path(fpath).suffix.lower()
                    if is_excluded_by_name(Path(fpath).name):
                        continue
                    if ext in SOFTWARE_EXTENSIONS:
                        all_software.append((fpath, f"{tag}_zip"))
                    elif ext in EXCEL_EXTENSIONS:
                        all_excel.append((fpath, f"{tag}_zip"))
            except Exception:
                pass

    # ==== 第3步：分析Excel文件（判断是否造价+识别专业+标准/非标分流） ====
    total = len(all_excel)
    print(f"\n[筛选] 分析 {total} 个Excel文件...")

    # 非标文件单独收集
    non_standard_collected = []
    standard_count = 0
    non_standard_count = 0

    for idx, (filepath, tag) in enumerate(all_excel):
        if (idx + 1) % 200 == 0:
            print(f"  分析 {idx+1}/{total}...")

        is_cost, specialty = detect_specialty_from_excel(filepath)
        if is_cost:
            # 进一步检查：有12位清单编码 → 标准清单，没有 → 非标文件
            if has_standard_bill_codes(filepath):
                collected.append((filepath, specialty, tag))
                stats[specialty]["excel"] += 1
                standard_count += 1
            else:
                non_standard_collected.append((filepath, specialty, tag))
                stats[specialty]["non_standard"] += 1
                non_standard_count += 1
        else:
            excluded_count += 1

    print(f"  标准清单（有12位编码）: {standard_count}")
    print(f"  非标文件（无12位编码）: {non_standard_count}")

    # ==== 第4步：处理软件文件（只从文件名判断专业） ====
    for filepath, tag in all_software:
        specialty = detect_specialty_for_software(filepath)
        collected.append((filepath, specialty, tag))
        stats[specialty]["software"] += 1

    # ==== 第5步：打印统计 ====
    print("\n" + "=" * 60)
    print("  按专业统计")
    print("=" * 60)

    # 固定专业顺序
    spec_order = [
        "土建装饰", "给排水", "电气", "消防", "通风空调",
        "智能化", "市政", "园林景观", "钢结构幕墙", "电力", "综合",
    ]

    total_excel = 0
    total_software = 0

    print(f"\n{'专业':<12} {'标准清单':>8} {'非标文件':>8} {'软件文件':>10} {'合计':>8}")
    print("-" * 50)
    for spec in spec_order:
        if spec in stats:
            e = stats[spec].get("excel", 0)
            ns = stats[spec].get("non_standard", 0)
            s = stats[spec].get("software", 0)
            total_excel += e
            total_software += s
            print(f"{spec:<12} {e:>8} {ns:>8} {s:>10} {e+ns+s:>8}")
    print("-" * 50)
    total_ns = sum(stats[sp].get("non_standard", 0) for sp in stats)
    print(f"{'合计':<12} {total_excel:>8} {total_ns:>8} {total_software:>10} {total_excel+total_ns+total_software:>8}")
    print(f"\n排除: {excluded_count} 个非造价文件")
    print(f"增量跳过: {skipped_count} 个已处理文件")

    if preview:
        print("\n[预览模式] 去掉 --preview 正式收集。")
        # 预览模式不保存增量记录，否则正式跑时会全部跳过
        # 清理解压临时目录
        for td in _temp_dirs:
            shutil.rmtree(td, ignore_errors=True)
        return stats

    # ==== 第6步：复制文件（标准清单 → jarvis/专业，非标 → jarvis/_非标文件/专业） ====
    print(f"\n[复制] 复制到 {output_dir} ...")
    copied = 0
    copied_ns = 0

    # 复制标准清单
    for filepath, specialty, tag in collected:
        dst_dir = os.path.join(output_dir, specialty)
        result = safe_copy(filepath, dst_dir, tag, seen_md5s)
        if result:
            copied += 1

    # 复制非标文件到 _非标文件 目录
    non_standard_dir = os.path.join(output_dir, "_非标文件")
    for filepath, specialty, tag in non_standard_collected:
        dst_dir = os.path.join(non_standard_dir, specialty)
        result = safe_copy(filepath, dst_dir, tag, seen_md5s)
        if result:
            copied_ns += 1

    print(f"\n[完成] 收集 {copied} 个标准清单 + {copied_ns} 个非标文件（去重后）")
    print(f"  输出目录: {output_dir}")
    print(f"  标准清单:")
    for spec in spec_order:
        spec_dir = os.path.join(output_dir, spec)
        if os.path.exists(spec_dir):
            count = len(os.listdir(spec_dir))
            print(f"    {spec}: {count} 个文件")
    if copied_ns > 0:
        print(f"  非标文件: {non_standard_dir}")
        for spec in spec_order:
            ns_dir = os.path.join(non_standard_dir, spec)
            if os.path.exists(ns_dir):
                count = len(os.listdir(ns_dir))
                print(f"    {spec}: {count} 个文件")

    # ==== 第7步：生成报告 ====
    try:
        _generate_report(output_dir, stats, spec_order, copied)
        print(f"  报告: {os.path.join(output_dir, '_收集报告.xlsx')}")
    except Exception as e:
        print(f"  [警告] 生成报告失败: {e}")

    # ==== 第8步：保存增量记录 ====
    _save_history(history)
    print(f"  增量记录已更新（共 {len(history)} 条）")

    # ==== 第9步：清理解压临时目录（防止C盘被撑满） ====
    if _temp_dirs:
        cleaned = 0
        for td in _temp_dirs:
            try:
                shutil.rmtree(td, ignore_errors=True)
                cleaned += 1
            except Exception:
                pass
        print(f"  已清理 {cleaned} 个临时目录")

    return stats


def _generate_report(output_dir, stats, spec_order, total_copied):
    """生成收集报告"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "按专业统计"

    headers = ["专业", "Excel数", "软件文件数", "合计"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    row = 2
    for spec in spec_order:
        if spec in stats:
            e = stats[spec].get("excel", 0)
            s = stats[spec].get("software", 0)
            ws.cell(row=row, column=1, value=spec)
            ws.cell(row=row, column=2, value=e)
            ws.cell(row=row, column=3, value=s)
            ws.cell(row=row, column=4, value=e + s)
            row += 1

    tf = Font(bold=True)
    ws.cell(row=row, column=1, value="合计（去重后）").font = tf
    ws.cell(row=row, column=4, value=total_copied).font = tf

    ws.column_dimensions["A"].width = 16
    for c in ["B", "C", "D"]:
        ws.column_dimensions[c].width = 14

    os.makedirs(output_dir, exist_ok=True)
    wb.save(os.path.join(output_dir, "_收集报告.xlsx"))


# ============================================================
# 入口
# ============================================================

def main():
    parser = argparse.ArgumentParser(description="造价文件收集工具（按专业分类）")
    parser.add_argument("--output", default=DEFAULT_OUTPUT, help="输出目录")
    parser.add_argument("--source", choices=["wechat", "wxwork", "all", "none"], default="all",
                        help="扫描来源（none=不扫微信，只扫--dir指定的目录）")
    parser.add_argument("--dir", action="append", dest="extra_dirs",
                        help="额外扫描的目录（可多次指定，如 --dir D:\\广联达临时文件）")
    parser.add_argument("--preview", action="store_true", help="预览模式")
    parser.add_argument("--full", action="store_true",
                        help="全量模式（忽略增量历史，从头扫描所有文件）")

    args = parser.parse_args()

    # 确定微信来源
    if args.source == "none":
        sources = []
    elif args.source == "all":
        sources = list(SOURCES.keys())
    else:
        sources = [args.source]

    collect_files(
        output_dir=args.output,
        sources_to_scan=sources,
        preview=args.preview,
        full=args.full,
        extra_dirs=args.extra_dirs,
    )


if __name__ == "__main__":
    main()
