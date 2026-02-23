"""
人工定额 vs 贾维斯定额 对比分析工具

用法：
    # 先跑贾维斯，再对比
    python main.py "带定额.xlsx" --mode agent --province "江西省房屋建筑..."
    python tools/eval_vs_human.py "带定额.xlsx" "output/匹配结果_xxx.xlsx" --province "江西省房屋建筑..."

    # 或者用批处理一键搞定
    对比分析.bat "带定额.xlsx"

功能：
    1. 从人工预算Excel中提取"清单→定额"对应关系（标准答案）
    2. 从贾维斯输出Excel中提取匹配结果
    3. 逐条对比，生成差异报告（控制台 + Excel）
"""

import argparse
import re
from pathlib import Path
from collections import OrderedDict

from loguru import logger
import config


def normalize_quota_code(code: str) -> str:
    """归一化定额编号，去掉换算/借用标记，方便对比

    1-42       → 1-42
    1-45 换    → 1-45
    借14-17 换 → 14-17
    补子目1    → 补子目1（保留，无法归一化）
    C4-4-31    → C4-4-31
    """
    if not code:
        return ""
    s = code.strip().replace(" ", "")
    # 去"换"后缀
    s = s.rstrip("换")
    # 去"借"前缀
    if s.startswith("借"):
        s = s[1:]
    return s


def parse_human_budget(excel_path: str) -> list[dict]:
    """解析人工预算Excel（广联达带定额导出格式）

    复用 import_reference 的解析逻辑，返回清单→定额对应列表。
    """
    from tools.import_reference import read_excel_pairs
    pairs = read_excel_pairs(excel_path)
    logger.info(f"人工预算解析完成: {len(pairs)} 条清单")
    return pairs


def parse_jarvis_output(excel_path: str) -> list[dict]:
    """解析贾维斯输出Excel

    贾维斯输出格式和人工预算类似（清单+定额交替排列），
    所以也可以用 read_excel_pairs 解析。

    但贾维斯输出有额外的列（推荐度、匹配说明、备选），需要适配。
    这里用简化版解析：只关心B列(编码/定额号)和C列(名称)。
    """
    import openpyxl

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    try:
        pairs = []
        for sheet_name in wb.sheetnames:
            # 跳过非清单sheet（如"待审核"、"统计汇总"）
            if any(kw in sheet_name for kw in ["待审核", "统计", "审核", "review"]):
                continue

            ws = wb[sheet_name]
            current_bill = None

            for row in ws.iter_rows(min_row=1, values_only=True):
                if not row or all(cell is None for cell in row):
                    continue

                cells = list(row) + [None] * max(0, 12 - len(row))
                col_a = str(cells[0] or "").strip()  # 序号
                col_b = str(cells[1] or "").strip()  # 编码/定额号
                col_c = str(cells[2] or "").strip()  # 名称
                col_d = str(cells[3] or "").strip()  # 描述/定额详情

                # 跳过表头
                if col_a in ("序号", "") and col_b in ("项目编码", "编码", ""):
                    if col_c in ("项目名称", "名称", ""):
                        continue

                # 清单行：B列是12位数字编码
                if re.match(r'^\d{9,12}$', col_b) and col_c:
                    if current_bill and current_bill["quotas"]:
                        pairs.append(current_bill)
                    current_bill = {
                        "bill_name": col_c,
                        "bill_code": col_b,
                        "bill_unit": str(cells[4] or "").strip(),
                        "bill_desc": col_d,
                        "sheet": sheet_name,
                        "quotas": [],
                    }
                    continue

                # 定额行：B列匹配定额编号格式
                if current_bill and col_b:
                    cleaned = col_b.replace(" ", "").rstrip("换")
                    if cleaned.startswith("借"):
                        cleaned = cleaned[1:]
                    is_quota = bool(re.match(r'^[A-Za-z]?\d{1,2}-\d+', cleaned)) or \
                               bool(re.match(r'^[A-Za-z]?\d{4,}', cleaned)) or \
                               col_b.startswith("补子目")
                    if is_quota:
                        current_bill["quotas"].append({
                            "code": col_b,
                            "name": col_c,
                        })

            # 最后一条
            if current_bill and current_bill["quotas"]:
                pairs.append(current_bill)

        logger.info(f"贾维斯输出解析完成: {len(pairs)} 条清单")
        return pairs
    finally:
        wb.close()


def compare(human_pairs: list[dict], jarvis_pairs: list[dict]) -> dict:
    """逐条对比人工和贾维斯的定额匹配结果

    返回:
        {
            "total_human": 人工清单总数,
            "total_jarvis": 贾维斯清单总数,
            "matched": 能配对的清单数,
            "agree": 完全一致数,
            "partial": 部分一致数,
            "disagree": 完全不同数,
            "human_only": 只有人工有的,
            "jarvis_only": 只有贾维斯有的,
            "details": [逐条详情...]
        }
    """
    # 按bill_code建索引
    human_map = OrderedDict()
    for p in human_pairs:
        code = p.get("bill_code", "")
        if code:
            human_map[code] = p

    jarvis_map = OrderedDict()
    for p in jarvis_pairs:
        code = p.get("bill_code", "")
        if code:
            jarvis_map[code] = p

    all_codes = list(OrderedDict.fromkeys(
        list(human_map.keys()) + list(jarvis_map.keys())
    ))

    details = []
    agree = partial = disagree = human_only_count = jarvis_only_count = 0

    for code in all_codes:
        h = human_map.get(code)
        j = jarvis_map.get(code)

        if h and not j:
            human_only_count += 1
            details.append({
                "bill_code": code,
                "bill_name": h["bill_name"],
                "sheet": h.get("sheet", ""),
                "status": "human_only",
                "human_quotas": [q["code"] for q in h["quotas"]],
                "human_names": [q["name"] for q in h["quotas"]],
                "jarvis_quotas": [],
                "jarvis_names": [],
            })
            continue

        if j and not h:
            jarvis_only_count += 1
            details.append({
                "bill_code": code,
                "bill_name": j["bill_name"],
                "sheet": j.get("sheet", ""),
                "status": "jarvis_only",
                "human_quotas": [],
                "human_names": [],
                "jarvis_quotas": [q["code"] for q in j["quotas"]],
                "jarvis_names": [q["name"] for q in j["quotas"]],
            })
            continue

        # 两边都有，对比定额编号
        h_codes = set(normalize_quota_code(q["code"]) for q in h["quotas"])
        j_codes = set(normalize_quota_code(q["code"]) for q in j["quotas"])

        # 去掉空值和补子目
        h_codes = {c for c in h_codes if c and not c.startswith("补子目")}
        j_codes = {c for c in j_codes if c and not c.startswith("补子目")}

        if h_codes == j_codes:
            status = "agree"
            agree += 1
        elif h_codes & j_codes:
            status = "partial"
            partial += 1
        else:
            status = "disagree"
            disagree += 1

        details.append({
            "bill_code": code,
            "bill_name": h["bill_name"],
            "sheet": h.get("sheet", ""),
            "status": status,
            "human_quotas": [q["code"] for q in h["quotas"]],
            "human_names": [q["name"] for q in h["quotas"]],
            "jarvis_quotas": [q["code"] for q in j["quotas"]],
            "jarvis_names": [q["name"] for q in j["quotas"]],
        })

    return {
        "total_human": len(human_pairs),
        "total_jarvis": len(jarvis_pairs),
        "matched": agree + partial + disagree,
        "agree": agree,
        "partial": partial,
        "disagree": disagree,
        "human_only": human_only_count,
        "jarvis_only": jarvis_only_count,
        "details": details,
    }


def print_report(result: dict):
    """打印控制台报告"""
    total = result["matched"] + result["human_only"] + result["jarvis_only"]
    matched = result["matched"]

    print("\n" + "=" * 60)
    print("对比分析报告")
    print("=" * 60)
    print(f"  人工清单: {result['total_human']} 条")
    print(f"  贾维斯:   {result['total_jarvis']} 条")
    print(f"  能配对:   {matched} 条")
    print()

    if matched > 0:
        print("一、总体对比")
        print(f"  完全一致:    {result['agree']:>4} 条 ({result['agree']/matched*100:.1f}%)")
        print(f"  部分一致:    {result['partial']:>4} 条 ({result['partial']/matched*100:.1f}%)")
        print(f"  完全不同:    {result['disagree']:>4} 条 ({result['disagree']/matched*100:.1f}%)")
    if result["human_only"]:
        print(f"  仅人工有:    {result['human_only']:>4} 条（贾维斯未匹配）")
    if result["jarvis_only"]:
        print(f"  仅贾维斯有:  {result['jarvis_only']:>4} 条")

    # 逐条差异明细（只打印不一致的）
    diffs = [d for d in result["details"] if d["status"] not in ("agree",)]
    if diffs:
        print(f"\n二、差异明细（{len(diffs)} 条）")
        print("-" * 60)

        # 按状态排序：disagree > partial > human_only > jarvis_only
        order = {"disagree": 0, "partial": 1, "human_only": 2, "jarvis_only": 3}
        diffs.sort(key=lambda d: order.get(d["status"], 9))

        for i, d in enumerate(diffs, 1):
            status_label = {
                "disagree": "完全不同",
                "partial": "部分一致",
                "human_only": "贾维斯漏配",
                "jarvis_only": "仅贾维斯",
            }.get(d["status"], d["status"])

            print(f"\n  [{status_label}] #{i} {d['bill_name']} ({d['bill_code']})")
            if d.get("sheet"):
                print(f"    Sheet: {d['sheet']}")
            if d["human_quotas"]:
                h_display = ", ".join(f"{c}({n})" for c, n in
                                      zip(d["human_quotas"], d["human_names"]))
                print(f"    人工: {h_display}")
            if d["jarvis_quotas"]:
                j_display = ", ".join(f"{c}({n})" for c, n in
                                      zip(d["jarvis_quotas"], d["jarvis_names"]))
                print(f"    贾维斯: {j_display}")

            if i >= 50:
                remaining = len(diffs) - 50
                if remaining > 0:
                    print(f"\n  ... 还有 {remaining} 条差异，详见输出Excel")
                break

    print("\n" + "=" * 60)


def write_excel_report(result: dict, output_path: str):
    """输出对比Excel报告"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "对比明细"

    # 表头
    headers = ["序号", "清单编码", "清单名称", "Sheet",
               "人工定额编号", "人工定额名称",
               "贾维斯定额编号", "贾维斯定额名称",
               "对比结果"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")

    # 颜色定义
    fills = {
        "agree": PatternFill("solid", fgColor="C6EFCE"),      # 绿色
        "partial": PatternFill("solid", fgColor="FFEB9C"),     # 黄色
        "disagree": PatternFill("solid", fgColor="FFC7CE"),    # 红色
        "human_only": PatternFill("solid", fgColor="E0E0E0"),  # 灰色
        "jarvis_only": PatternFill("solid", fgColor="D9E1F2"), # 蓝灰
    }

    status_labels = {
        "agree": "一致",
        "partial": "部分一致",
        "disagree": "不同",
        "human_only": "贾维斯漏配",
        "jarvis_only": "仅贾维斯",
    }

    # 按状态排序：不同 > 部分一致 > 漏配 > 仅贾维斯 > 一致
    order = {"disagree": 0, "partial": 1, "human_only": 2, "jarvis_only": 3, "agree": 4}
    sorted_details = sorted(result["details"], key=lambda d: order.get(d["status"], 9))

    for i, d in enumerate(sorted_details, 1):
        row = i + 1
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=d["bill_code"])
        ws.cell(row=row, column=3, value=d["bill_name"])
        ws.cell(row=row, column=4, value=d.get("sheet", ""))
        ws.cell(row=row, column=5, value=", ".join(d["human_quotas"]))
        ws.cell(row=row, column=6, value=", ".join(d["human_names"]))
        ws.cell(row=row, column=7, value=", ".join(d["jarvis_quotas"]))
        ws.cell(row=row, column=8, value=", ".join(d["jarvis_names"]))
        ws.cell(row=row, column=9, value=status_labels.get(d["status"], d["status"]))

        # 行着色
        fill = fills.get(d["status"])
        if fill:
            for col in range(1, 10):
                ws.cell(row=row, column=col).fill = fill

    # 列宽
    widths = [5, 14, 25, 20, 20, 35, 20, 35, 10]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + col)].width = w

    # 汇总sheet
    ws2 = wb.create_sheet("汇总")
    summary_data = [
        ("指标", "数量", "占比"),
        ("能配对的清单", result["matched"], ""),
        ("完全一致", result["agree"],
         f"{result['agree']/max(result['matched'],1)*100:.1f}%"),
        ("部分一致", result["partial"],
         f"{result['partial']/max(result['matched'],1)*100:.1f}%"),
        ("完全不同", result["disagree"],
         f"{result['disagree']/max(result['matched'],1)*100:.1f}%"),
        ("贾维斯漏配", result["human_only"], ""),
        ("仅贾维斯有", result["jarvis_only"], ""),
    ]
    for r, row_data in enumerate(summary_data, 1):
        for c, val in enumerate(row_data, 1):
            cell = ws2.cell(row=r, column=c, value=val)
            if r == 1:
                cell.font = Font(bold=True)

    wb.save(output_path)
    logger.info(f"对比Excel已保存: {output_path}")


def main():
    parser = argparse.ArgumentParser(description="人工定额 vs 贾维斯定额 对比分析")
    parser.add_argument("human_excel", help="人工预算Excel（带定额）")
    parser.add_argument("jarvis_excel", help="贾维斯输出Excel（匹配结果）")
    parser.add_argument("--province", help="省份名称")
    parser.add_argument("--output", help="对比报告Excel输出路径")
    args = parser.parse_args()

    if args.province:
        # Validate/normalize province input without mutating global runtime state.
        config.resolve_province(args.province)

    # 1. 解析两个文件
    human_pairs = parse_human_budget(args.human_excel)
    jarvis_pairs = parse_jarvis_output(args.jarvis_excel)

    # 2. 对比
    result = compare(human_pairs, jarvis_pairs)

    # 3. 输出报告
    print_report(result)

    # 4. 输出Excel
    output_path = args.output
    if not output_path:
        output_dir = config.OUTPUT_DIR
        output_dir.mkdir(parents=True, exist_ok=True)
        output_path = str(output_dir / "对比报告.xlsx")

    write_excel_report(result, output_path)
    print(f"\n对比Excel: {output_path}")


if __name__ == "__main__":
    main()
