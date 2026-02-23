# -*- coding: utf-8 -*-
"""
Jarvis纠正工具：将审核纠正直接写回匹配结果Excel。

用法：
    python tools/jarvis_correct.py <匹配结果Excel> <纠正JSON>

纠正JSON格式：
[
  {"seq": 25, "quota_id": "C4-8-234", "quota_name": "户内干包式电缆终端头 70mm²"},
  {"seq": 35, "quota_id": "C4-11-287", "quota_name": "管内穿铜芯线动力线路 截面2.5"}
]
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.cell.cell import MergedCell


def _safe_write_cell(
    ws,
    row: int,
    col: int,
    value,
    skipped: list[str],
    seq_for_log: int,
    field_name: str,
) -> bool:
    """Safely write a cell and skip merged slave cells."""
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        skipped.append(
            f"序号{seq_for_log} {field_name} 跳过：目标单元格({row},{col})是合并从属格"
        )
        return False
    cell.value = value
    return True


def _is_bill_serial(a_val) -> bool:
    """Check whether column-A value is a bill serial number."""
    if a_val is None:
        return False
    if isinstance(a_val, int):
        return a_val > 0
    if isinstance(a_val, float):
        return a_val > 0 and a_val.is_integer()
    text = str(a_val).strip()
    if not text:
        return False
    if text.isdigit():
        return True
    return bool(re.fullmatch(r"\d+\.0+", text))


def _find_bill_rows(ws) -> list[int]:
    """Find all bill rows in one sheet by A-column serial marker."""
    rows: list[int] = []
    for row in range(1, ws.max_row + 1):
        if _is_bill_serial(ws.cell(row=row, column=1).value):
            rows.append(row)
    return rows


def correct_excel(excel_path: str, corrections: list, output_path: str | None = None) -> str:
    """将纠正结果写回匹配结果Excel并输出新文件。"""
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"文件不存在: {excel_path}")

    wb = openpyxl.load_workbook(excel_path)
    ws_active = wb.active

    # 旧格式兜底映射（仅活动Sheet）：序号 -> 定额行
    legacy_seq_to_quota_row: dict[int, int] = {}
    for row in range(1, ws_active.max_row + 1):
        seq_val = ws_active.cell(row=row, column=1).value
        if seq_val is None:
            continue
        try:
            seq = int(seq_val)
        except (ValueError, TypeError):
            continue

        quota_row = row + 1
        if quota_row <= ws_active.max_row:
            legacy_seq_to_quota_row[seq] = quota_row

    if not legacy_seq_to_quota_row:
        raise RuntimeError("Excel中未找到有效序号行，请检查文件结构")

    applied = 0
    skipped: list[str] = []
    bill_rows_cache: dict[str, list[int]] = {}

    for corr in corrections:
        seq = corr.get("seq")
        quota_id = corr.get("quota_id", "")
        quota_name = corr.get("quota_name", "")

        if seq is None:
            skipped.append(f"缺少seq字段: {corr}")
            continue

        try:
            seq_int = int(seq)
        except (ValueError, TypeError):
            skipped.append(f"seq不是有效数字: {seq}")
            continue

        target_ws = None
        item_row = None
        quota_row = None

        sheet_name = str(corr.get("sheet_name", "") or "").strip()
        sheet_bill_seq_raw = corr.get("sheet_bill_seq")
        try:
            sheet_bill_seq = int(sheet_bill_seq_raw)
        except (ValueError, TypeError):
            sheet_bill_seq = None

        # 优先：按 sheet + sheet_bill_seq 精准定位（多Sheet安全）
        if sheet_name and sheet_bill_seq is not None and sheet_bill_seq > 0:
            if sheet_name in wb.sheetnames:
                target_ws = wb[sheet_name]
                if sheet_name not in bill_rows_cache:
                    bill_rows_cache[sheet_name] = _find_bill_rows(target_ws)
                bill_rows = bill_rows_cache[sheet_name]
                if 1 <= sheet_bill_seq <= len(bill_rows):
                    item_row = bill_rows[sheet_bill_seq - 1]
                    quota_row = item_row + 1
                else:
                    skipped.append(
                        f"序号{seq_int} 跳过：sheet '{sheet_name}' 中 sheet_bill_seq="
                        f"{sheet_bill_seq} 超出范围(1..{len(bill_rows)})"
                    )
                    continue
            else:
                skipped.append(f"序号{seq_int} 跳过：sheet不存在 '{sheet_name}'")
                continue
        else:
            # 回退：兼容旧格式（仅活动Sheet）
            quota_row = legacy_seq_to_quota_row.get(seq_int)
            if quota_row is None:
                skipped.append(f"序号{seq_int}在Excel中未找到")
                continue
            target_ws = ws_active
            item_row = quota_row - 1

        old_id = target_ws.cell(row=quota_row, column=2).value or ""

        write_ok = [
            _safe_write_cell(
                target_ws, quota_row, 2, quota_id, skipped, seq_int, "定额编号"
            ),
            _safe_write_cell(
                target_ws, quota_row, 3, quota_name, skipped, seq_int, "定额名称"
            ),
            _safe_write_cell(
                target_ws, item_row, 10, "★★★已审核", skipped, seq_int, "审核标记"
            ),
            _safe_write_cell(
                target_ws,
                item_row,
                11,
                f"Jarvis纠正: {old_id} → {quota_id}",
                skipped,
                seq_int,
                "纠正说明",
            ),
        ]

        # 定额编号（B列）是核心写入，必须成功才算纠正成功
        # 审核标记（J列）和说明（K列）是辅助信息，失败不影响判定
        if write_ok[0]:  # write_ok[0] = 定额编号写入是否成功
            applied += 1

    if not output_path:
        p = Path(excel_path)
        output_path = str(p.parent / f"{p.stem}_已审核{p.suffix}")

    wb.save(output_path)

    print(f"纠正完成: 共{len(corrections)}条，成功{applied}条")
    if skipped:
        print(f"跳过{len(skipped)}条")
        for msg in skipped:
            print(f"  - {msg}")
    print(f"输出文件: {output_path}")

    return output_path


def main() -> None:
    parser = argparse.ArgumentParser(description="Jarvis纠正工具")
    parser.add_argument("excel_path", help="匹配结果Excel路径")
    parser.add_argument("corrections_json", help="纠正JSON文件路径")
    parser.add_argument(
        "--output",
        help="输出Excel路径（默认在原文件名后追加_已审核）",
    )
    args = parser.parse_args()

    with open(args.corrections_json, "r", encoding="utf-8") as f:
        corrections = json.load(f)

    if not isinstance(corrections, list):
        print("错误: JSON必须是数组格式")
        sys.exit(1)

    correct_excel(args.excel_path, corrections, args.output)


if __name__ == "__main__":
    main()
