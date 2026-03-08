# -*- coding: utf-8 -*-
"""
提取已审核Excel中的人工审核项（推荐度含"待审"的条目）

用法：
    python tools/extract_manual_items.py "<已审核Excel路径>" [--sheet "待审核1"]

功能：
    从"待审核"sheet中筛选推荐度含"待审"的行，打印详细信息。
"""
import sys
import os
import argparse

# 确保项目根目录在 sys.path 中
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl


def extract_manual_items(excel_path, sheet_name=None):
    """从已审核Excel提取人工审核项

    参数:
        excel_path: 已审核Excel文件路径
        sheet_name: 指定sheet名称，默认自动查找最后一个"待审核"sheet

    返回:
        list[dict] — 每个dict包含人工审核项的详细信息
    """
    wb = openpyxl.load_workbook(excel_path, read_only=True)

    # 自动查找待审核sheet（优先选最后一个，即最近一次运行的结果）
    if sheet_name is None:
        candidates = [s for s in wb.sheetnames if "待审核" in s]
        if not candidates:
            print("错误: Excel中没有找到包含'待审核'的sheet")
            wb.close()
            return []
        sheet_name = candidates[-1]  # 取最后一个（最近一次运行）
        print(f"自动选择sheet: {sheet_name}")

    ws = wb[sheet_name]

    # 表头: 清单序号(0), 清单名称(1), 项目特征(2), 当前定额编号(3),
    #       当前定额名称(4), 推荐度(5), 问题说明(6),
    #       备选1(7), 备选2(8), 备选3(9), 主材(10)
    manual_items = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue  # 跳过表头
        rec = str(row[5]) if row[5] else ""
        if "待审" not in rec:
            continue  # 只提取人工审核项

        item = {
            "seq": row[0],               # 清单序号
            "name": row[1],              # 清单名称
            "description": row[2],       # 项目特征（完整描述）
            "quota_id": row[3],          # 当前定额编号
            "quota_name": row[4],        # 当前定额名称
            "confidence": rec,           # 推荐度（含置信度百分比）
            "error_reason": row[6],      # 问题说明（含错误类型和原因）
            "alt1": row[7],              # 备选定额1
            "alt2": row[8],              # 备选定额2
            "alt3": row[9],              # 备选定额3
            "material": row[10] if len(row) > 10 else None,  # 主材
        }
        manual_items.append(item)

    wb.close()
    return manual_items


def print_manual_items(items):
    """格式化打印人工审核项"""
    print(f"\n共 {len(items)} 条人工审核项")
    print("=" * 80)

    for idx, item in enumerate(items, 1):
        print(f"\n--- 第{idx}条 (清单序号: {item['seq']}) ---")
        print(f"  清单名称: {item['name']}")

        # 项目特征（截取前100字符，太长就省略）
        desc = str(item['description'] or '')
        if len(desc) > 100:
            desc = desc[:100] + "..."
        print(f"  项目特征: {desc}")

        print(f"  当前定额编号: {item['quota_id']}")
        print(f"  当前定额名称: {item['quota_name']}")
        print(f"  推荐度: {item['confidence']}")
        print(f"  问题说明: {item['error_reason']}")

        # 备选定额
        if item['alt1']:
            print(f"  备选1: {item['alt1']}")
        if item['alt2']:
            print(f"  备选2: {item['alt2']}")
        if item['alt3']:
            print(f"  备选3: {item['alt3']}")
        if item.get('material'):
            print(f"  主材: {item['material']}")

    print("\n" + "=" * 80)


def main():
    parser = argparse.ArgumentParser(description="提取已审核Excel中的人工审核项")
    parser.add_argument("excel_path", help="已审核Excel文件路径")
    parser.add_argument("--sheet", default=None, help="指定sheet名称（默认自动选最后一个待审核sheet）")
    args = parser.parse_args()

    if not os.path.exists(args.excel_path):
        print(f"错误: 文件不存在: {args.excel_path}")
        sys.exit(1)

    items = extract_manual_items(args.excel_path, args.sheet)
    if items:
        print_manual_items(items)
    else:
        print("没有找到人工审核项")


if __name__ == "__main__":
    main()
