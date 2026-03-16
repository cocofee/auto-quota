"""
批量匹配执行器

功能：读取 batch.db 中 scanned 状态的文件，批量跑纯搜索匹配，收集结果。

设计要点（按 Codex 5.3 审核建议）：
1. 复用现有 match_search_only() 接口，不改核心逻辑
2. 纯搜索模式（不用Agent/LLM），免费+快速
3. 按省份分组初始化搜索引擎（不同省用不同定额库）
4. 断点续跑：基于数据库状态，已matched的不重跑
5. 每个文件处理完立即更新数据库状态（原子性）
6. 列映射机制：让不同格式（work_list/equipment_list）都能跑

用法：
    python tools/batch_runner.py                           # 跑全部可处理文件
    python tools/batch_runner.py --format standard_bill    # 只跑标准格式
    python tools/batch_runner.py --province 广东           # 只跑某省
    python tools/batch_runner.py --specialty 消防          # 只跑某专业
    python tools/batch_runner.py --limit 100               # 只跑前100个文件
    python tools/batch_runner.py --sample 3                # 每个省份x专业各采样3个（均匀铺开）
    python tools/batch_runner.py --sample 3 --review       # 采样+Jarvis审核出诊断报告
"""

import os
import sys
import json
import time
import hashlib
import argparse
from pathlib import Path
from datetime import datetime
from collections import defaultdict

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

# 批量跑时抑制详细日志，只让进度输出显示
from loguru import logger
logger.remove()  # 去掉所有handler，完全静音

from tools.batch_scanner import DB_PATH, get_db, init_db, _upsert_file, ALGORITHM_VERSION
from src.bill_reader import BillReader
from src.match_engine import init_search_components, init_experience_db, match_search_only
import config

# ============================================================
# 列映射配置（让不同格式的Excel都能转成标准清单项）
# ============================================================

# 不同格式对应的列名变体（按优先级排列，第一个匹配到的就用）
COLUMN_MAPPINGS = {
    "standard_bill": {
        # 标准格式直接用 BillReader 读取，不需要额外映射
    },
    "work_list": {
        "name": ["工作量名称", "工作项目", "项目", "名称"],
        "description": ["主要内容及范围说明", "内容说明", "说明", "范围"],
        "unit": ["单位"],
        "quantity": ["工作量", "数量"],
    },
    "equipment_list": {
        "name": ["设备名称", "材料名称", "品名", "项目名称", "名称"],
        "description": ["规格型号", "规格", "型号", "参数", "技术参数"],
        "unit": ["单位"],
        "quantity": ["数量", "工程量"],
    },
}

# 结果输出目录
RESULTS_DIR = Path(__file__).resolve().parent.parent / "output" / "batch" / "results"


# ============================================================
# 文件读取（支持多种格式）
# ============================================================

def read_file_items(file_path: str, fmt: str, sheet_info: list) -> list[dict]:
    """读取文件中的清单项。

    参数:
        file_path: Excel文件路径
        fmt: 格式类型（standard_bill/work_list/equipment_list）
        sheet_info: Sheet信息列表（从scanner扫描结果中获取）

    返回:
        清单项列表，每项包含 name, description, unit, quantity 等字段
    """
    if fmt == "standard_bill":
        # 标准格式用 BillReader 直接读（它有完整的列识别逻辑）
        reader = BillReader()
        try:
            items = reader.read_excel(file_path)
            return items
        except Exception as e:
            raise ValueError(f"BillReader读取失败: {e}")

    # 非标准格式：用列映射手动读取
    import openpyxl

    mapping = COLUMN_MAPPINGS.get(fmt)
    if not mapping:
        raise ValueError(f"不支持的格式: {fmt}")

    # 找出有清单数据的Sheet
    bill_sheets = []
    if sheet_info:
        for si in sheet_info:
            if isinstance(si, dict) and si.get("has_bill_data"):
                bill_sheets.append(si["name"])

    fp = Path(file_path)
    # .xls 转换
    actual_path = file_path
    temp_path = None
    # 用统一的magic bytes检测，兼容后缀是.xlsx但实际是.xls的文件
    from src.bill_reader import is_xls_format
    if is_xls_format(file_path):
        from tools.batch_scanner import _convert_xls
        temp_path = _convert_xls(file_path)
        actual_path = temp_path

    try:
        wb = openpyxl.load_workbook(str(actual_path), read_only=True, data_only=True)
        all_items = []

        try:
            sheets_to_read = bill_sheets if bill_sheets else wb.sheetnames[:5]
            for sn in sheets_to_read:
                if sn not in wb.sheetnames:
                    continue
                ws = wb[sn]
                items = _read_sheet_with_mapping(ws, sn, mapping)
                all_items.extend(items)
        finally:
            wb.close()
    finally:
        if temp_path:
            try:
                Path(temp_path).unlink(missing_ok=True)
            except Exception:
                pass

    return all_items


def _read_sheet_with_mapping(ws, sheet_name: str, mapping: dict) -> list[dict]:
    """用列映射读取单个Sheet的数据。

    逻辑：
    1. 扫描前20行找表头
    2. 根据mapping匹配列位置
    3. 从表头下方读取数据行
    """
    # 先读前20行找表头
    header_rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        header_rows.append(row)
        if i >= 19:
            break

    if not header_rows:
        return []

    # 在前20行中找表头（匹配mapping中的列名）
    header_row_idx = None
    col_map = {}  # field_name → col_index

    for row_idx in range(len(header_rows)):
        row = header_rows[row_idx]
        cells = [str(c).strip() if c is not None else "" for c in row]

        # 尝试匹配每个字段
        temp_map = {}
        for field, aliases in mapping.items():
            for col_idx, cell_text in enumerate(cells):
                if not cell_text:
                    continue
                for alias in aliases:
                    if alias in cell_text:
                        temp_map[field] = col_idx
                        break
                if field in temp_map:
                    break

        # 至少匹配到 name 字段才算找到表头
        if "name" in temp_map:
            header_row_idx = row_idx
            col_map = temp_map
            break

    if header_row_idx is None:
        return []

    # 从表头下一行开始读数据（逐行迭代，不一次性全读进内存）
    items = []
    for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if row_idx <= header_row_idx:
            continue  # 跳过表头及之前的行
        cells = [str(c).strip() if c is not None else "" for c in row]

        name = cells[col_map["name"]] if "name" in col_map and col_map["name"] < len(cells) else ""
        if not name or len(name) < 2:
            continue  # 跳过空行和太短的行

        desc = cells[col_map["description"]] if "description" in col_map and col_map["description"] < len(cells) else ""
        unit = cells[col_map["unit"]] if "unit" in col_map and col_map["unit"] < len(cells) else ""
        qty = cells[col_map["quantity"]] if "quantity" in col_map and col_map["quantity"] < len(cells) else ""

        # 组装成标准清单项格式（兼容 match_search_only 的输入）
        item = {
            "index": len(items) + 1,
            "code": "",
            "name": name,
            "description": desc,
            "unit": unit,
            "quantity": qty,
            "search_text": f"{name} {desc}".strip(),
            "params": {},
            "sheet_name": sheet_name,
            "section": "",
        }
        items.append(item)

    return items


# ============================================================
# 核心批量匹配逻辑
# ============================================================

def run_batch(format_filter: str = None, province_filter: str = None,
              specialty_filter: str = None, limit: int = None,
              sample: int = None, review: bool = False,
              progress_callback=None):
    """批量匹配主函数。

    流程：
    1. 从数据库读取 scanned 状态的文件
    2. 按省份+定额库类型分组（不同专业用不同定额库）
    3. 逐文件匹配，结果存到 output/batch/results/{省份}/
    4. 更新数据库状态为 matched

    参数:
        progress_callback: 进度回调函数，签名 callback(current, total, file_name)
                          用于Web端实时显示匹配进度
    """
    init_db()
    RESULTS_DIR.mkdir(parents=True, exist_ok=True)

    # 第1步：查询可处理的文件
    conn = get_db()
    try:
        query = "SELECT * FROM file_registry WHERE status = 'scanned'"
        params = []

        if format_filter:
            query += " AND format = ?"
            params.append(format_filter)
        if province_filter:
            query += " AND province = ?"
            params.append(province_filter)
        if specialty_filter:
            query += " AND specialty = ?"
            params.append(specialty_filter)

        query += " ORDER BY province, specialty"

        # sample模式下不在SQL层limit，采样后再截断
        if limit and not sample:
            query += " LIMIT ?"
            params.append(limit)

        files = conn.execute(query, params).fetchall()
    finally:
        conn.close()

    if not files:
        print("没有可处理的文件（都已matched或无scanned状态的文件）。")
        return

    # 均匀采样（--sample N：每个省份×专业各取N个）
    if sample:
        files = _sample_files(files, sample)
        if not files:
            print("采样后无文件可处理。")
            return
        # sample模式下的limit：截断采样总数
        if limit and len(files) > limit:
            files = files[:limit]
            print(f"截断到前{limit}个文件")

    print(f"待处理: {len(files)}个文件")

    # 第2步：按省份+定额库类型分组
    # 同一省份不同专业可能需要不同定额库（如广东电气→安装库，广东土建→建筑库）
    by_library = {}  # key: (省份, 定额库类型关键词)
    for f in files:
        prov = f["province"] or "未知省份"
        specialty = f["specialty"] or ""
        lib_type = _SPECIALTY_TO_LIBRARY_TYPE.get(specialty, "通用安装")
        group_key = (prov, lib_type)
        if group_key not in by_library:
            by_library[group_key] = []
        by_library[group_key].append(f)

    # 第3步：逐组处理
    total_success = 0
    total_errors = 0
    total_items = 0

    for (prov, lib_type), prov_files in by_library.items():
        # 初始化该组的搜索引擎（按省份+专业选择正确的定额库）
        specialty_hint = (prov_files[0]["specialty"] or "") if prov_files else ""
        resolved_province = _resolve_province(prov, specialty=specialty_hint)
        if not resolved_province:
            print(f"  ⚠ {prov} 没有定额库，跳过{len(prov_files)}个文件")
            _mark_files_error(prov_files, f"省份「{prov}」没有定额库")
            total_errors += len(prov_files)
            continue

        try:
            searcher, validator = init_search_components(resolved_province)
            experience_db = init_experience_db(no_experience=False, province=resolved_province)
        except Exception as e:
            print(f"  ⚠ {prov} 初始化失败，跳过")
            _mark_files_error(prov_files, f"搜索引擎初始化失败: {e}")
            total_errors += len(prov_files)
            continue

        # 逐文件匹配
        for i, f in enumerate(prov_files):
            file_path = f["file_path"]
            file_name = f["file_name"]
            fmt = f["format"]
            sheet_info_str = f["sheet_info"]

            # 进度回调（给Web端用）
            global_idx = total_success + total_errors + i + 1
            if progress_callback:
                try:
                    progress_callback(current=global_idx, total=len(files),
                                      file_name=file_name)
                except Exception:
                    pass

            # 短文件名（最多30字符）
            short_name = file_name if len(file_name) <= 30 else file_name[:27] + '...'
            # 固定前缀宽度，让后面的数据对齐
            prefix = f"  [{prov}] [{i+1}/{len(prov_files)}] {short_name}"
            prefix_padded = prefix.ljust(55)  # 前缀固定55字符宽，不够补空格
            start_time = time.time()

            try:
                sheet_info = json.loads(sheet_info_str) if sheet_info_str else []
                items = read_file_items(file_path, fmt, sheet_info)
                if not items:
                    print(f"{prefix_padded} 跳过")
                    _mark_file_matched(f, [], 0)
                    total_success += 1
                    continue

                total_in_file = len(items)

                # 逐条匹配并显示进度
                results = []
                for item_idx, item in enumerate(items):
                    single_results = match_search_only(
                        bill_items=[item],
                        searcher=searcher,
                        validator=validator,
                        experience_db=experience_db,
                        province=resolved_province,
                    )
                    results.extend(single_results)

                    done = item_idx + 1
                    if done % 10 == 0 or done == total_in_file:
                        pct = done * 100 // total_in_file
                        elapsed_so_far = time.time() - start_time
                        print(f"\r{prefix_padded} {total_in_file:>4d}条 {elapsed_so_far:>4.0f}s  {pct:>3d}%", end="", flush=True)

                elapsed = time.time() - start_time

                # Jarvis审核（--review模式）
                review_data = None
                if review and results:
                    review_data = _review_results(results, resolved_province)

                _save_results(f, results, elapsed, review_data=review_data)
                _mark_file_matched(f, results, elapsed)

                total_items += len(results)
                total_success += 1

                confs = [r.get("confidence", 0) for r in results]
                avg_conf = sum(confs) / len(confs) if confs else 0

                if avg_conf >= 70:
                    conf_str = f"\033[32m均{avg_conf:.0f}%\033[0m"
                elif avg_conf >= 50:
                    conf_str = f"\033[33m均{avg_conf:.0f}%\033[0m"
                else:
                    conf_str = f"\033[31m均{avg_conf:.0f}%\033[0m"

                print(f"\r{prefix_padded} {total_in_file:>4d}条 {elapsed:>4.0f}s  {conf_str}     ")

            except Exception as e:
                print(f"\r{prefix_padded} \033[31m错误\033[0m                ")
                _mark_file_error(f, str(e))
                total_errors += 1

    # 汇总
    print(f"\n  完成: {total_success}个文件 {total_items}条清单" +
          (f" 错误{total_errors}个" if total_errors else ""))

    # 返回统计结果（供Web端/调用方使用）
    return {
        "success_files": total_success,
        "error_files": total_errors,
        "total_items": total_items,
    }


# ============================================================
# 省份解析（省份名+专业 → 定额库代码）
# ============================================================

# 专业 → 定额库类型关键词映射
# 用于把文件的专业分类（如"电气"）转成定额库搜索关键词
# 关键词要足够精确，避免匹配到"城市轨道交通(安装分册)"之类的冷门库
_SPECIALTY_TO_LIBRARY_TYPE = {
    "电气": "通用安装",
    "消防": "通用安装",
    "给排水": "通用安装",
    "通风空调": "通用安装",
    "智能化": "通用安装",
    "电力": "通用安装",
    "综合": "通用安装",        # 综合类大多是安装工程
    "钢结构幕墙": "通用安装",
    "土建装饰": "房屋建筑",    # 匹配"房屋建筑与装饰"
    "市政": "市政",
    "园林景观": "园林",
}


def _get_quota_count(province_dir: Path) -> int:
    """获取省份定额库的条目数，0表示空库或无效。"""
    import sqlite3
    db_path = province_dir / "quota.db"
    if not db_path.exists():
        return 0
    try:
        conn = sqlite3.connect(str(db_path))
        tables = [t[0] for t in conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table'").fetchall()]
        if "quotas" not in tables:
            conn.close()
            return 0
        cnt = conn.execute("SELECT COUNT(*) FROM quotas").fetchone()[0]
        conn.close()
        return cnt
    except Exception:
        return 0


def _resolve_province(province_name: str, specialty: str = None) -> str:
    """把省份名+专业转成定额库目录名。

    例如：
        ("广东", "电气") → "广东省通用安装工程综合定额(2018)"
        ("广东", "市政") → "广东省市政工程综合定额(2018)"
        ("北京", "电气") → "北京市建设工程施工消耗量标准(2024)"

    优先使用 config.resolve_province() 的多关键词匹配能力。
    多匹配时用 lib_type 关键词过滤；仍多个时选定额条数最多的库。
    """
    # 根据专业决定定额库类型关键词
    lib_type = _SPECIALTY_TO_LIBRARY_TYPE.get(specialty, "安装") if specialty else None

    provinces_dir = Path(__file__).resolve().parent.parent / "db" / "provinces"

    # 尝试用 config.resolve_province() 解析（更智能的匹配逻辑）
    try:
        import config as cfg
        # 先尝试"省份+类型"精确匹配（如"广东安装"）
        if lib_type:
            try:
                return cfg.resolve_province(f"{province_name}{lib_type}", interactive=False)
            except ValueError:
                pass  # 匹配失败，继续尝试

        # 再尝试纯省份名匹配
        try:
            return cfg.resolve_province(province_name, interactive=False)
        except ValueError as e:
            # 多匹配时，从错误消息解析候选列表，用 lib_type 过滤
            err_msg = str(e)
            if "匹配到多个省份" in err_msg and lib_type and provinces_dir.exists():
                # 找出所有匹配省份名的有效目录，按 lib_type 过滤
                candidates = []
                for d in provinces_dir.iterdir():
                    if d.is_dir() and province_name in d.name:
                        cnt = _get_quota_count(d)
                        if cnt > 0:
                            candidates.append((d.name, cnt))
                # 用 lib_type 关键词筛选（"安装"→含"安装"或"施工"的库）
                type_kws = [lib_type]
                if lib_type == "通用安装":
                    type_kws = ["安装", "施工"]  # 北京等综合库不叫"安装"
                elif lib_type == "房屋建筑":
                    type_kws = ["建筑", "房屋", "施工"]
                filtered = [(n, c) for n, c in candidates if any(kw in n for kw in type_kws)]
                if filtered:
                    # 多个匹配时选定额条数最多的（主库通常最大）
                    return max(filtered, key=lambda x: x[1])[0]
                elif candidates:
                    # lib_type过滤无结果，选最大的库
                    return max(candidates, key=lambda x: x[1])[0]

    except Exception:
        pass  # config模块异常，回退到简单匹配

    # 回退：简单的目录扫描（跳过空数据库，选最大的库）
    if not provinces_dir or not provinces_dir.exists():
        return None

    # 收集所有匹配的有效目录
    candidates = []

    # 优先匹配含 lib_type 关键词的目录
    if lib_type:
        type_kws = [lib_type]
        if lib_type == "通用安装":
            type_kws = ["安装", "施工"]
        elif lib_type == "房屋建筑":
            type_kws = ["建筑", "房屋", "施工"]
        for d in provinces_dir.iterdir():
            if d.is_dir() and province_name in d.name:
                if any(kw in d.name for kw in type_kws):
                    cnt = _get_quota_count(d)
                    if cnt > 0:
                        candidates.append((d.name, cnt))

    # 没找到则放宽到任意匹配
    if not candidates:
        for d in provinces_dir.iterdir():
            if d.is_dir() and d.name.startswith(province_name):
                cnt = _get_quota_count(d)
                if cnt > 0:
                    candidates.append((d.name, cnt))

    if candidates:
        # 选定额条数最多的（主库通常最大）
        return max(candidates, key=lambda x: x[1])[0]

    return None


# ============================================================
# 结果保存和状态更新
# ============================================================

def _sample_files(files: list, n: int) -> list:
    """均匀采样：每个省份×专业组合各取n个文件。

    目的：快速铺满所有省份，而不是死磕一个省跑完再换下一个。
    例如 --sample 3：广东电气3个、广东消防3个、浙江电气3个……
    """
    import random
    groups = defaultdict(list)
    for f in files:
        prov = f["province"] or "未识别"
        # 跳过未识别省份的文件（没有定额库，跑了也白跑）
        if prov == "未识别":
            continue
        spec = f["specialty"] or "未分类"
        groups[(prov, spec)].append(f)

    sampled = []
    for (prov, spec), group_files in sorted(groups.items()):
        # 每组随机取n个（不够n个就全取）
        pick = random.sample(group_files, min(n, len(group_files)))
        sampled.extend(pick)

    # 打印采样计划
    print(f"\n采样计划（每组{n}个）:")
    sample_groups = defaultdict(int)
    for f in sampled:
        sample_groups[(f["province"] or "未识别", f["specialty"] or "未分类")] += 1
    for (prov, spec), cnt in sorted(sample_groups.items()):
        print(f"  {prov:10s} × {spec:10s} → {cnt}个")
    print(f"  合计: {len(sampled)}个文件\n")

    return sampled


def _review_results(results: list, province: str, sibling_provinces: list = None) -> dict:
    """对匹配结果做Jarvis审核，返回诊断信息。

    复用 jarvis_auto_review 的检测逻辑（纯规则，不查大模型），
    给每条结果标注错误分类：[词][跨][档][冷][非][脏]。
    """
    try:
        from tools.jarvis_auto_review import _detect_phase
    except ImportError:
        # 审核模块不可用，返回空诊断
        return {"reviewed": False, "reason": "jarvis_auto_review不可用"}

    # _detect_phase 需要和 jarvis_pipeline 一样格式的 results
    # batch_runner 的 results 是 match_search_only 返回的原始格式，可以直接用
    try:
        detected_errors, measure_items, no_match_items, correct_count = _detect_phase(results)
    except Exception as e:
        return {"reviewed": False, "reason": f"审核异常: {e}"}

    # 统计错误分类
    error_types = defaultdict(int)
    error_details = []
    for err in detected_errors:
        # 猜测错误分类
        error_type = _classify_error(err, results)
        error_types[error_type] += 1
        error_details.append({
            "seq": err.get("seq", 0),
            "bill_name": err.get("bill_item", {}).get("name", "") if isinstance(err.get("bill_item"), dict) else str(err.get("bill_item", "")),
            "quota_id": err.get("quota_id", ""),
            "quota_name": err.get("quota_name", ""),
            "error": err.get("error", ""),
            "error_type": error_type,
            "confidence": err.get("confidence", 0),
        })

    return {
        "reviewed": True,
        "correct_count": correct_count,
        "error_count": len(detected_errors),
        "measure_count": len(measure_items),
        "no_match_count": len(no_match_items),
        "error_types": dict(error_types),
        "error_details": error_details[:50],  # 最多存50条细节（防文件太大）
    }


def _classify_error(err: dict, results: list) -> str:
    """根据错误信息猜测错误分类代号。

    返回: 词/跨/档/冷/非/脏
    """
    error_msg = str(err.get("error", "")).lower()
    bill_name = ""
    if isinstance(err.get("bill_item"), dict):
        bill_name = err["bill_item"].get("name", "")

    # 类别不匹配 → 跨库
    if "类别不匹配" in error_msg or "category" in error_msg:
        return "跨"
    # 参数偏差 → 档位
    if "参数" in error_msg or "dn" in error_msg or "规格" in error_msg:
        return "档"
    # 材质不匹配 → 同义词
    if "材质" in error_msg or "material" in error_msg:
        return "词"
    # 管道用途 → 同义词
    if "用途" in error_msg:
        return "词"
    # 连接方式 → 档位
    if "连接" in error_msg:
        return "档"
    # 措施项 → 非定额
    if "措施" in error_msg:
        return "非"
    # 默认归为同义词缺口（最常见的错误类型）
    return "词"


def _save_results(file_info, results: list, elapsed: float,
                   review_data: dict = None):
    """保存匹配结果到 JSON 文件。"""
    prov = file_info["province"] or "未知省份"
    prov_dir = RESULTS_DIR / prov
    prov_dir.mkdir(parents=True, exist_ok=True)

    # 文件名用原文件名（去掉扩展名）+ 路径hash后缀，避免同名文件覆盖
    base_name = Path(file_info["file_name"]).stem
    # 用文件完整路径的hash后4位区分同名文件
    path_hash = hashlib.md5(file_info["file_path"].encode()).hexdigest()[:4]
    result_path = prov_dir / f"{base_name}_{path_hash}.json"

    # 简化结果（只保留关键字段，减少文件大小）
    # 注意：match_engine 返回的结构是嵌套的：
    #   清单信息在 bill_item 对象内，定额信息在 quotas 数组内
    simplified = []
    for r in results:
        # 提取清单信息（嵌套在 bill_item 里）
        bill_item = r.get("bill_item") or {}
        if hasattr(bill_item, "name"):
            # bill_item 可能是对象（有 .name 属性）
            bill_name = getattr(bill_item, "name", "")
            bill_desc = getattr(bill_item, "description", "")
        else:
            # bill_item 可能是字典
            bill_name = bill_item.get("name", "")
            bill_desc = bill_item.get("description", "")

        # 提取主定额信息（嵌套在 quotas 数组第一个元素里）
        quotas = r.get("quotas") or []
        if quotas:
            main_quota = quotas[0] if isinstance(quotas[0], dict) else {}
            quota_id = main_quota.get("quota_id", "")
            quota_name = main_quota.get("name", "")
        else:
            quota_id = ""
            quota_name = ""

        simplified.append({
            "name": bill_name,
            "description": bill_desc,
            "matched_quota_id": quota_id,
            "matched_quota_name": quota_name,
            "confidence": r.get("confidence", 0),
            "match_source": r.get("match_source", ""),
        })

    data = {
        "file_path": file_info["file_path"],
        "province": prov,
        "specialty": file_info["specialty"],
        "format": file_info["format"],
        "total_items": len(results),
        "elapsed_seconds": round(elapsed, 1),
        "match_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "algo_version": ALGORITHM_VERSION,
        "review": review_data,  # Jarvis审核诊断（--review模式才有）
        "results": simplified,
    }

    with open(result_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def _mark_file_matched(file_info, results: list, elapsed: float):
    """更新数据库状态为 matched。"""
    conn = get_db()
    try:
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        conn.execute("""
            UPDATE file_registry
            SET status = 'matched',
                match_time = ?,
                algo_version = ?,
                updated_at = ?
            WHERE file_path = ?
        """, (now, ALGORITHM_VERSION, now, file_info["file_path"]))
        conn.commit()
    finally:
        conn.close()


def _mark_file_error(file_info, error_msg: str):
    """更新数据库状态为 error。"""
    conn = get_db()
    try:
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        conn.execute("""
            UPDATE file_registry
            SET status = 'error',
                error_msg = ?,
                updated_at = ?
            WHERE file_path = ?
        """, (error_msg, now, file_info["file_path"]))
        conn.commit()
    finally:
        conn.close()


def _mark_files_error(files: list, error_msg: str):
    """批量标记文件为 error。"""
    conn = get_db()
    try:
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        for f in files:
            conn.execute("""
                UPDATE file_registry
                SET status = 'error',
                    error_msg = ?,
                    updated_at = ?
                WHERE file_path = ?
            """, (error_msg, now, f["file_path"]))
        conn.commit()
    finally:
        conn.close()


# ============================================================
# 命令行入口
# ============================================================

def main():
    parser = argparse.ArgumentParser(description="批量匹配执行器")
    parser.add_argument("--format", choices=["standard_bill", "work_list", "equipment_list"],
                        help="只跑某种格式")
    parser.add_argument("--province", help="只跑某省（如 广东）")
    parser.add_argument("--specialty", help="只跑某专业（如 消防）")
    parser.add_argument("--limit", type=int, help="只跑前N个文件")
    parser.add_argument("--sample", type=int,
                        help="均匀采样：每个省份x专业各取N个文件")
    parser.add_argument("--review", action="store_true",
                        help="匹配后加Jarvis审核，生成诊断报告")

    args = parser.parse_args()

    run_batch(
        format_filter=args.format,
        province_filter=args.province,
        specialty_filter=args.specialty,
        limit=args.limit,
        sample=args.sample,
        review=args.review,
    )


if __name__ == "__main__":
    main()
