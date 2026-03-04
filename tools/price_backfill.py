"""
价格回填工具：把广联达组价结果回填到甲方原始清单

用法：
    python tools/price_backfill.py <甲方原始Excel> <广联达导出Excel>
    python tools/price_backfill.py <甲方原始Excel> <广联达导出Excel> --dry-run
    python tools/price_backfill.py <甲方原始Excel> <广联达导出Excel> --output 指定输出路径.xlsx

工作原理：
    1. 读取甲方原始Excel的结构（表头、序号、名称、价格列位置）
    2. 读取广联达导出Excel的价格数据（综合单价、合价）
    3. 按序号一一对应（序号不匹配时用名称模糊匹配兜底）
    4. 复制原文件 → 只往价格列写入数值 → 保存为 xxx_已回填.xlsx

输出：
    在原文件旁生成 xxx_已回填.xlsx（不修改原文件）
"""

import re
import sys
import shutil
import argparse
from pathlib import Path
from difflib import SequenceMatcher

import openpyxl
from openpyxl.cell.cell import MergedCell


# ================================================================
# 甲方原始Excel读取
# ================================================================

# 表头关键词 → 字段名映射
HEADER_PATTERNS = {
    "index": ["序号"],
    "name": [
        "项目名称", "名称", "清单名称", "项目内容",
        "子目名称", "项 目 内 容", "设备名称",
    ],
    "unit": [
        "计量单位", "单位",
    ],
    "quantity": [
        "工程量", "工程数量", "数量",
    ],
    "unit_price": [
        "综合单价", "单价", "含税单价",
        "不含税单价", "单价(元)", "单价（元）",
    ],
    "total_price": [
        "合价", "金额", "含税合价", "合计",
        "不含税合价", "合价(元)", "合价（元）",
        "综合合价",
    ],
}


def _detect_original_structure(ws):
    """检测甲方原始Excel的表头结构

    返回:
        {
            "header_row": int,        # 表头行号（1-based）
            "col_map": {              # 列映射（字段名 → 列号，1-based）
                "index": 1,
                "name": 2,
                "unit_price": 7,      # 可能为None（原表没有价格列）
                "total_price": 8,
            },
            "items": [                # 数据行列表
                {"row": 7, "index": "1", "name": "电气火灾监控模块"},
                ...
            ]
        }
    """
    # 扫描前20行找表头
    col_map = {}
    header_row = None

    for row_idx in range(1, min(21, ws.max_row + 1)):
        row_col_map = {}
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if isinstance(cell, MergedCell) or cell.value is None:
                continue
            text = str(cell.value).strip().replace("\n", "").replace(" ", "")
            if len(text) > 20:
                continue

            for field, patterns in HEADER_PATTERNS.items():
                for pattern in patterns:
                    clean_pattern = pattern.replace(" ", "")
                    if clean_pattern in text:
                        # 对于价格列，区分"单价"和"合价"
                        # 避免"综合单价"同时匹配"单价"和"合价"
                        if field == "total_price" and "单价" in text and "合" not in text:
                            continue
                        row_col_map[field] = col_idx
                        break

        # 至少有name列才算有效表头
        if "name" in row_col_map and len(row_col_map) >= 2:
            col_map = row_col_map
            header_row = row_idx
            break

    if not header_row:
        raise ValueError("未找到有效表头行（至少需要'名称'和'序号'/'单价'列之一）")

    # 读取数据行
    items = []
    skip_keywords = ["合计", "小计", "本专业小计", "说明", "注：", "本页"]
    for row_idx in range(header_row + 1, ws.max_row + 1):
        name_col = col_map.get("name")
        name_cell = ws.cell(row=row_idx, column=name_col)
        if isinstance(name_cell, MergedCell) or not name_cell.value:
            continue
        name = str(name_cell.value).strip()
        if not name:
            continue
        # 跳过汇总行
        if any(kw in name for kw in skip_keywords):
            continue
        # 跳过分节标题行（以中文数字序号开头）
        if re.match(r'^[一二三四五六七八九十]+[、.\s]', name):
            continue

        # 读取序号
        idx_val = ""
        if "index" in col_map:
            idx_cell = ws.cell(row=row_idx, column=col_map["index"])
            if not isinstance(idx_cell, MergedCell) and idx_cell.value is not None:
                idx_val = str(idx_cell.value).strip()

        # 读取单位
        unit_val = ""
        if "unit" in col_map:
            u_cell = ws.cell(row=row_idx, column=col_map["unit"])
            if not isinstance(u_cell, MergedCell) and u_cell.value is not None:
                unit_val = str(u_cell.value).strip()

        # 读取数量
        qty_val = None
        if "quantity" in col_map:
            q_cell = ws.cell(row=row_idx, column=col_map["quantity"])
            if not isinstance(q_cell, MergedCell) and q_cell.value is not None:
                try:
                    qty_val = float(q_cell.value)
                except (ValueError, TypeError):
                    pass

        items.append({
            "row": row_idx,
            "index": idx_val,
            "name": name,
            "unit": unit_val,
            "quantity": qty_val,
        })

    return {
        "header_row": header_row,
        "col_map": col_map,
        "items": items,
    }


# ================================================================
# 广联达导出Excel价格读取
# ================================================================

def _read_gld_prices(ws):
    """读取广联达导出Excel的价格数据

    广联达格式特征：
    - A列有数字序号 → 清单行
    - A列为空、B列有定额编号 → 定额行（跳过）

    返回:
        [
            {"index": "1", "name": "电气火灾监控模块", "unit_price": 123.45, "total_price": 246.90},
            ...
        ]
    """
    # 先找表头，检测价格列位置
    header_row = None
    name_col = None
    unit_price_col = None
    total_price_col = None
    index_col = None
    unit_col = None      # 计量单位列
    quantity_col = None   # 工程量列

    for row_idx in range(1, min(21, ws.max_row + 1)):
        found_name = False
        temp_map = {}
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if isinstance(cell, MergedCell) or cell.value is None:
                continue
            text = str(cell.value).strip().replace("\n", "").replace(" ", "")
            if len(text) > 20:
                continue

            if text in ["序号"]:
                temp_map["index"] = col_idx
            for kw in ["项目名称", "名称", "清单名称", "项目内容", "子目名称", "设备名称"]:
                if kw.replace(" ", "") in text:
                    temp_map["name"] = col_idx
                    found_name = True
                    break
            # 检测单位列
            for kw in ["计量单位", "单位"]:
                if kw in text and "单价" not in text:
                    temp_map["unit"] = col_idx
                    break
            # 检测数量列
            for kw in ["工程量", "工程数量", "数量"]:
                if kw in text:
                    temp_map["quantity"] = col_idx
                    break
            # 检测价格列
            if "综合单价" in text or (text == "单价" and "合" not in text):
                temp_map["unit_price"] = col_idx
            elif "合价" in text or "合计" in text:
                if "unit_price" not in temp_map or col_idx != temp_map.get("unit_price"):
                    temp_map["total_price"] = col_idx

        if found_name:
            header_row = row_idx
            index_col = temp_map.get("index")
            name_col = temp_map.get("name")
            unit_col = temp_map.get("unit")
            quantity_col = temp_map.get("quantity")
            unit_price_col = temp_map.get("unit_price")
            total_price_col = temp_map.get("total_price")
            break

    if not header_row or not name_col:
        raise ValueError("广联达导出文件中未找到有效表头")

    # 如果没找到明确的价格列，尝试用位置推断（通常在G/H列）
    if not unit_price_col and not total_price_col:
        # 尝试在name列之后找数字列
        for col_idx in range(name_col + 1, min(ws.max_column + 1, name_col + 6)):
            cell = ws.cell(row=header_row, column=col_idx)
            if cell.value and "单价" in str(cell.value):
                unit_price_col = col_idx
            elif cell.value and ("合价" in str(cell.value) or "金额" in str(cell.value)):
                total_price_col = col_idx

    # 读取清单行数据（只读A列有序号的行）
    prices = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        # 判断是否为清单行（A列/index列有数字序号）
        is_bill_row = False
        idx_val = ""

        if index_col:
            idx_cell = ws.cell(row=row_idx, column=index_col)
            if not isinstance(idx_cell, MergedCell) and idx_cell.value is not None:
                val = str(idx_cell.value).strip()
                # 序号通常是数字
                try:
                    float(val)
                    is_bill_row = True
                    idx_val = val
                except ValueError:
                    pass

        if not is_bill_row:
            continue

        # 读取名称
        name_cell = ws.cell(row=row_idx, column=name_col)
        name = ""
        if not isinstance(name_cell, MergedCell) and name_cell.value:
            name = str(name_cell.value).strip()
        if not name:
            continue

        # 跳过汇总行
        if any(kw in name for kw in ["合计", "小计", "本页"]):
            continue

        # 读取单位
        gld_unit = ""
        if unit_col:
            u_cell = ws.cell(row=row_idx, column=unit_col)
            if not isinstance(u_cell, MergedCell) and u_cell.value is not None:
                gld_unit = str(u_cell.value).strip()

        # 读取数量
        gld_qty = None
        if quantity_col:
            q_cell = ws.cell(row=row_idx, column=quantity_col)
            if not isinstance(q_cell, MergedCell) and q_cell.value is not None:
                try:
                    gld_qty = float(q_cell.value)
                except (ValueError, TypeError):
                    pass

        # 读取价格
        unit_price = None
        total_price = None
        if unit_price_col:
            up_cell = ws.cell(row=row_idx, column=unit_price_col)
            if not isinstance(up_cell, MergedCell) and up_cell.value is not None:
                try:
                    unit_price = float(up_cell.value)
                except (ValueError, TypeError):
                    pass
        if total_price_col:
            tp_cell = ws.cell(row=row_idx, column=total_price_col)
            if not isinstance(tp_cell, MergedCell) and tp_cell.value is not None:
                try:
                    total_price = float(tp_cell.value)
                except (ValueError, TypeError):
                    pass

        prices.append({
            "row": row_idx,
            "index": idx_val,
            "name": name,
            "unit": gld_unit,
            "quantity": gld_qty,
            "unit_price": unit_price,
            "total_price": total_price,
        })

    return prices


# ================================================================
# 映射匹配
# ================================================================

def _build_mapping(original_items, price_data):
    """建立原始行→价格的映射

    策略A：按序号一一对应（最常用）
    策略B：按名称相似度匹配（兜底）

    返回:
        [
            {
                "row": 7,                    # 原始Excel行号
                "original_name": "电气火灾监控模块",
                "matched_name": "电气火灾监控模块",
                "unit_price": 123.45,
                "total_price": 246.90,
                "match_method": "index",     # "index" 或 "name"
            },
            ...
        ]
    """
    mapping = []

    # 先尝试按序号对应
    price_by_index = {}
    for p in price_data:
        if p["index"]:
            # 序号可能是浮点数（如"1.0"），统一转整数字符串
            try:
                idx_key = str(int(float(p["index"])))
            except ValueError:
                idx_key = p["index"]
            price_by_index[idx_key] = p

    # 按名称建索引（兜底用）
    unmatched_prices = list(price_data)

    for item in original_items:
        matched = None
        method = ""

        # 策略A：按序号
        idx_key = ""
        if item["index"]:
            try:
                idx_key = str(int(float(item["index"])))
            except ValueError:
                idx_key = item["index"]
        if idx_key and idx_key in price_by_index:
            matched = price_by_index[idx_key]
            method = "index"

        # 策略B：按名称相似度（兜底）
        if not matched and unmatched_prices:
            best_score = 0
            best_match = None
            for p in unmatched_prices:
                score = SequenceMatcher(None, item["name"], p["name"]).ratio()
                if score > best_score:
                    best_score = score
                    best_match = p
            # 相似度>0.6才认为匹配成功
            if best_score >= 0.6 and best_match:
                matched = best_match
                method = f"name({best_score:.0%})"
                unmatched_prices.remove(best_match)

        if matched:
            # ---- 防错配校验 ----
            warnings = []
            # 1) 名称相似度检查（即使序号匹配，名称差太远也报警）
            name_score = SequenceMatcher(
                None, item["name"], matched["name"]).ratio()
            if name_score < 0.4:
                warnings.append(f"名称差异大({name_score:.0%})")

            # 2) 单位一致性检查
            orig_unit = item.get("unit", "")
            gld_unit = matched.get("unit", "")
            if orig_unit and gld_unit and orig_unit != gld_unit:
                warnings.append(f"单位不同({orig_unit}→{gld_unit})")

            # 3) 数量一致性检查（允许10%偏差）
            orig_qty = item.get("quantity")
            gld_qty = matched.get("quantity")
            if orig_qty and gld_qty:
                diff_ratio = abs(orig_qty - gld_qty) / max(orig_qty, 0.001)
                if diff_ratio > 0.1:
                    warnings.append(
                        f"数量偏差({orig_qty}→{gld_qty})")

            mapping.append({
                "row": item["row"],
                "original_name": item["name"],
                "matched_name": matched["name"],
                "matched_row": matched.get("row"),
                "matched_index": matched.get("index"),
                "unit_price": matched.get("unit_price"),
                "total_price": matched.get("total_price"),
                "match_method": method,
                "warnings": warnings,
            })
        else:
            mapping.append({
                "row": item["row"],
                "original_name": item["name"],
                "matched_name": None,
                "matched_row": None,
                "matched_index": None,
                "unit_price": None,
                "total_price": None,
                "match_method": "未匹配",
                "warnings": [],
            })

    return mapping


# ================================================================
# 价格写入
# ================================================================

def _write_prices(original_path, mapping, col_map, output_path=None):
    """把价格写入原始Excel的副本

    关键：不改变原表的任何格式，只往价格列写入数值
    """
    original_path = Path(original_path)

    if output_path:
        out_path = Path(output_path)
    else:
        stem = original_path.stem
        out_path = original_path.parent / f"{stem}_已回填{original_path.suffix}"

    # 复制原文件（保留所有格式）
    shutil.copy2(str(original_path), str(out_path))

    # 打开副本写入价格
    wb = openpyxl.load_workbook(str(out_path))
    ws = wb.active

    unit_price_col = col_map.get("unit_price")
    total_price_col = col_map.get("total_price")

    written = 0
    for m in mapping:
        if m["match_method"] == "未匹配":
            continue

        row = m["row"]

        if unit_price_col and m["unit_price"] is not None:
            cell = ws.cell(row=row, column=unit_price_col)
            if not isinstance(cell, MergedCell):
                cell.value = round(m["unit_price"], 2)

        if total_price_col and m["total_price"] is not None:
            cell = ws.cell(row=row, column=total_price_col)
            if not isinstance(cell, MergedCell):
                cell.value = round(m["total_price"], 2)

        written += 1

    wb.save(str(out_path))
    wb.close()

    return str(out_path), written


# ================================================================
# 主入口
# ================================================================

def backfill(original_path, gld_path, output_path=None, dry_run=False):
    """价格回填主函数

    参数:
        original_path: 甲方原始Excel路径
        gld_path: 广联达导出Excel路径（带价格）
        output_path: 输出路径（默认在原文件旁生成_已回填文件）
        dry_run: 只预览映射，不写文件

    返回:
        (输出文件路径, 映射统计)
    """
    original_path = Path(original_path)
    gld_path = Path(gld_path)

    if not original_path.exists():
        raise FileNotFoundError(f"甲方原始文件不存在: {original_path}")
    if not gld_path.exists():
        raise FileNotFoundError(f"广联达导出文件不存在: {gld_path}")

    # 1. 读取甲方原始Excel结构
    print(f"读取甲方原始清单: {original_path}")
    wb_orig = openpyxl.load_workbook(str(original_path))
    ws_orig = wb_orig.active
    orig_info = _detect_original_structure(ws_orig)
    wb_orig.close()
    print(f"  表头行: 第{orig_info['header_row']}行")
    print(f"  列映射: {orig_info['col_map']}")
    print(f"  数据行: {len(orig_info['items'])}条")

    # 检查是否有价格列
    has_price_col = bool(
        orig_info["col_map"].get("unit_price") or
        orig_info["col_map"].get("total_price")
    )
    if not has_price_col:
        print("  警告: 原始清单中未找到单价/合价列，无法回填价格")
        print("  提示: 请确认甲方清单中有'综合单价''合价'等列名")
        return None, None

    # 2. 读取广联达导出Excel的价格
    print(f"\n读取广联达导出文件: {gld_path}")
    wb_gld = openpyxl.load_workbook(str(gld_path))
    ws_gld = wb_gld.active
    price_data = _read_gld_prices(ws_gld)
    wb_gld.close()
    print(f"  清单行: {len(price_data)}条")

    # 3. 建立映射
    mapping = _build_mapping(orig_info["items"], price_data)

    # 统计
    matched_by_index = sum(1 for m in mapping if m["match_method"] == "index")
    matched_by_name = sum(1 for m in mapping if m["match_method"].startswith("name"))
    unmatched = sum(1 for m in mapping if m["match_method"] == "未匹配")

    print(f"\n映射结果: 总{len(mapping)}条, "
          f"序号匹配{matched_by_index}, 名称匹配{matched_by_name}, "
          f"未匹配{unmatched}")

    # 打印映射详情
    for m in mapping:
        status = "✓" if m["match_method"] != "未匹配" else "✗"
        price_info = ""
        if m["unit_price"] is not None:
            price_info = f" 单价={m['unit_price']:.2f}"
        if m["total_price"] is not None:
            price_info += f" 合价={m['total_price']:.2f}"
        print(f"  {status} 行{m['row']}: {m['original_name'][:20]}"
              f" → {(m['matched_name'] or '无')[:20]}"
              f" [{m['match_method']}]{price_info}")

    # 4. 写入价格
    if dry_run:
        print("\n[dry-run模式] 不写文件，仅预览")
        return None, mapping
    else:
        result_path, written = _write_prices(
            original_path, mapping, orig_info["col_map"], output_path)
        print(f"\n已回填 {written}/{len(mapping)} 条价格")
        print(f"输出文件: {result_path}")
        return result_path, mapping


# ================================================================
# 命令行入口
# ================================================================

def main():
    parser = argparse.ArgumentParser(
        description="价格回填工具：把广联达组价结果回填到甲方原始清单")
    parser.add_argument("original", help="甲方原始Excel路径")
    parser.add_argument("gld_export", help="广联达导出Excel路径（带价格）")
    parser.add_argument("--output", "-o", help="指定输出路径（默认在原文件旁生成）")
    parser.add_argument("--dry-run", action="store_true",
                        help="只预览映射结果，不写文件")

    args = parser.parse_args()

    try:
        backfill(args.original, args.gld_export,
                 output_path=args.output, dry_run=args.dry_run)
    except Exception as e:
        print(f"\n错误: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
