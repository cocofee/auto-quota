# -*- coding: utf-8 -*-
"""
Jarvis 诊断工具统一入口 — 同义词管理、错题分析、跨省验证

子命令：
  excel           诊断已审核Excel的人工项根因（原有功能）
  benchmark-fix   从benchmark错题提取同义词缺口并修复
  verify          跨省搜索验证一个关键词
  audit-coverage  同义词跨省覆盖率审计
  audit-static    同义词表静态分析
  ranking-report  排序错误深度分析
  bucket          离线诊断分桶（多维深度分析）

用法：
    # 诊断Excel（向后兼容，可省略excel子命令）
    python tools/jarvis_diagnose.py "output/xxx_已审核.xlsx" --province "上海建筑"
    python tools/jarvis_diagnose.py "output/xxx_已审核.xlsx" --province "上海建筑" --fix

    # 错题精补（从benchmark错题提取同义词缺口）
    python tools/jarvis_diagnose.py benchmark-fix
    python tools/jarvis_diagnose.py benchmark-fix --fix
    python tools/jarvis_diagnose.py benchmark-fix --min-freq 3 --fix

    # 跨省验证
    python tools/jarvis_diagnose.py verify "镀锌钢管"

    # 同义词审计
    python tools/jarvis_diagnose.py audit-coverage --quick
    python tools/jarvis_diagnose.py audit-static --fix

    # 排序分析
    python tools/jarvis_diagnose.py ranking-report

    # 离线诊断分桶
    python tools/jarvis_diagnose.py bucket
"""

import sys
import os
import json
import shutil
import argparse
import subprocess
from pathlib import Path
from collections import Counter

PROJECT_ROOT = Path(__file__).resolve().parent.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))


# 清单常见叫法 → 定额库可能的叫法（跑的越多积累越多）
SYNONYM_VARIANTS = {
    "凿槽": ["刨沟", "开槽", "剔槽"],
    "洗漱台": ["洗脸盆", "洗手盆", "台盆", "面盆"],
    "跷板开关": ["暗开关", "翘板开关", "墙壁开关"],
    "翘板开关": ["暗开关", "跷板开关", "墙壁开关"],
    "贴膜": ["玻璃膜", "隔热膜", "防晒膜"],
    "脚手架搭拆": ["脚手架", "搭拆费"],
    "塑胶地板": ["塑料地板", "PVC地板", "橡胶地板"],
    "洗手台": ["洗脸盆", "洗手盆", "台盆"],
    "机柜": ["通信机柜", "配线箱", "配线架"],
    "大屏": ["显示屏", "LED屏"],
    "光纤": ["光缆", "光纤电缆"],
}


def _extract_manual_from_excel(excel_path: str) -> list[dict]:
    """从已审核Excel的待审核sheet提取人工审核项"""
    import openpyxl

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    # 找所有待审核sheet（可能有多个，如"待审核"和"待审核1"）
    review_sheets = [wb[sn] for sn in wb.sheetnames if "待审核" in sn]
    if not review_sheets:
        wb.close()
        return []

    # 找关键列（用第一个sheet的表头）
    ws0 = review_sheets[0]
    headers = [ws0.cell(1, c).value for c in range(1, ws0.max_column + 1)]

    # 找关键列
    col_map = {}
    for i, h in enumerate(headers):
        if h is None:
            continue
        h = str(h).strip()
        if "清单序号" in h:
            col_map["seq"] = i
        elif "清单名称" in h:
            col_map["name"] = i
        elif "项目特征" in h:
            col_map["desc"] = i
        elif "当前定额编号" in h:
            col_map["quota_id"] = i
        elif "当前定额名称" in h:
            col_map["quota_name"] = i
        elif "推荐度" in h:
            col_map["recommend"] = i
        elif "问题说明" in h or "审核说明" in h:
            col_map["note"] = i

    if "recommend" not in col_map:
        wb.close()
        return []

    # 从所有待审核sheet提取（去重，按清单名+定额编号）
    # 除了"推荐"（高置信度）之外，其他都要诊断
    items = []
    seen = set()
    for ws in review_sheets:
        for row in range(2, ws.max_row + 1):
            rec = str(ws.cell(row, col_map["recommend"] + 1).value or "")
            # 跳过高置信度推荐项（大概率正确的不用诊断）
            if "推荐" in rec:
                continue
            # 跳过空行
            if not rec or rec == "None":
                continue

            name = str(ws.cell(row, col_map.get("name", 0) + 1).value or "").strip()
            qid = str(ws.cell(row, col_map.get("quota_id", 0) + 1).value or "").strip()
            key = f"{name}|{qid}"
            if key in seen:
                continue
            seen.add(key)

            item = {
                "seq": ws.cell(row, col_map.get("seq", 0) + 1).value,
                "name": name,
                "desc": str(ws.cell(row, col_map.get("desc", 0) + 1).value or "").strip(),
                "quota_id": qid,
                "quota_name": str(ws.cell(row, col_map.get("quota_name", 0) + 1).value or "").strip(),
                "recommend": rec,
                "note": str(ws.cell(row, col_map.get("note", 0) + 1).value or "").strip(),
            }
            items.append(item)

    wb.close()
    return items


def _search_all_libraries(keywords: list[str], province: str,
                          sibling_provinces: list[str] | None) -> list[tuple]:
    """在主库+所有兄弟库搜索，返回所有命中结果"""
    from src.quota_search import search_quota_db

    all_results = []
    # 搜主库
    results = search_quota_db(keywords, province=province) or []
    for r in results[:5]:
        all_results.append((province, r[0], r[1]))

    # 搜兄弟库
    for sib in (sibling_provinces or []):
        results = search_quota_db(keywords, province=sib) or []
        for r in results[:5]:
            all_results.append((sib, r[0], r[1]))

    return all_results


def _try_synonym_variants(name: str, province: str,
                          sibling_provinces: list[str] | None) -> dict | None:
    """尝试同义变体替换，看能否搜到定额

    返回: {"original": "凿槽", "variant": "刨沟", "hit": ("03-4-13-16", "砖墙刨沟", "上海安装")}
    """
    for keyword, variants in SYNONYM_VARIANTS.items():
        if keyword not in name:
            continue
        for variant in variants:
            # 用变体词搜索
            results = _search_all_libraries([variant], province, sibling_provinces)
            if results:
                return {
                    "original": keyword,
                    "variant": variant,
                    "hit_province": results[0][0],
                    "hit_id": results[0][1],
                    "hit_name": results[0][2],
                }
    return None


# ---- 自动修复相关函数（--fix 模式）----

# 跨省验证用的代表省份（建筑+安装各4个，覆盖南北东西）
# 同义词可能在建筑库也可能在安装库，两种都要搜
VERIFY_PROVINCES_RAW = [
    "北京2024",        # 北京（建筑+安装合一）
    "广东房屋",        # 广东建筑
    "广东通用安装",    # 广东安装
    "江西房屋",        # 江西建筑
    "江西安装",        # 江西安装
    "浙江房屋",        # 浙江建筑
    "浙江安装",        # 浙江安装
]


def _cross_province_verify(variant: str) -> tuple[int, list[str]]:
    """验证目标词在几个代表省能搜到

    按省去重统计（同省建筑+安装算1个省），>=2省命中即通过。
    返回: (命中省数, 命中省份列表)
    """
    from src.quota_search import search_quota_db
    from config import resolve_province

    # 解析省份全名（只在首次调用时解析，缓存在函数属性上）
    if not hasattr(_cross_province_verify, "_resolved"):
        resolved = []
        for raw in VERIFY_PROVINCES_RAW:
            try:
                full = resolve_province(raw)
                resolved.append((raw, full))
            except Exception:
                pass
        _cross_province_verify._resolved = resolved

    # 搜索并按省去重（"广东房屋"和"广东通用安装"都算广东）
    hit_provinces = set()  # 用省份简称去重
    hit_details = []
    for raw, full in _cross_province_verify._resolved:
        try:
            results = search_quota_db([variant], province=full)
            if results:
                # 提取省份简称（取前2个汉字，如"北京""广东""江西"）
                short = raw[:2]
                hit_provinces.add(short)
                hit_details.append(raw)
        except Exception:
            pass
    return len(hit_provinces), hit_details


def _write_synonyms(pairs: list[tuple[str, str]]) -> int:
    """安全写入同义词到 engineering_synonyms.json

    规则：
    - 已有的key不覆盖
    - 按key长度降序排序（保护伞机制）
    - 写入前先备份

    返回: 实际写入的同义词对数
    """
    syn_path = PROJECT_ROOT / "data" / "engineering_synonyms.json"
    bak_path = syn_path.with_suffix(".json.bak")

    # 读取现有同义词表
    with open(syn_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    # 备份（覆盖上次备份）
    shutil.copy2(syn_path, bak_path)

    added = 0
    for original, variant in pairs:
        if original in data:
            continue  # 已存在，不覆盖
        data[original] = [variant]
        added += 1

    if added == 0:
        return 0

    # _specialty_scope排序特殊处理：meta字段放最前，普通词按长度降序
    sorted_data = dict(sorted(
        data.items(),
        key=lambda x: (not x[0].startswith("_"), len(x[0])),
        reverse=True,
    ))

    with open(syn_path, "w", encoding="utf-8") as f:
        json.dump(sorted_data, f, ensure_ascii=False, indent=2)

    return added


def _run_benchmark_check() -> tuple[bool, str]:
    """跑 benchmark 对比基线，检查有没有退化

    返回: (是否通过, 摘要文本)
    """
    try:
        result = subprocess.run(
            [sys.executable, "tools/run_benchmark.py", "--profile", "full", "--compare"],
            capture_output=True, text=True, timeout=600,
            cwd=str(PROJECT_ROOT),
        )
        output = result.stdout + result.stderr

        # 从输出中提取命中率变化（格式如 "命中率: 38.9% → 39.2% (▲+0.3%)"）
        for line in output.split("\n"):
            if "命中率" in line or "hit_rate" in line:
                # 找到命中率行，检查是否退化
                if "▼" in line or "退化" in line or "下降" in line:
                    return False, line.strip()
                return True, line.strip()

        # 没找到命中率行，看退出码
        if result.returncode == 0:
            return True, "benchmark通过（未解析到具体命中率）"
        return False, f"benchmark异常（退出码{result.returncode}）"
    except subprocess.TimeoutExpired:
        return False, "benchmark超时（>10分钟）"
    except Exception as e:
        return False, f"benchmark执行失败: {e}"


def _rollback_synonyms() -> bool:
    """从备份回滚同义词表"""
    syn_path = PROJECT_ROOT / "data" / "engineering_synonyms.json"
    bak_path = syn_path.with_suffix(".json.bak")
    if bak_path.exists():
        shutil.copy2(bak_path, syn_path)
        return True
    return False


# ---- benchmark-fix：从错题提取同义词缺口 ----

# 通用词黑名单（太泛的词不能单独做同义词key，必须带修饰词）
_GENERIC_WORDS = {
    "钢管", "阀门", "套管", "管道", "电缆", "灯", "开关",
    "配电箱", "桥架", "风管", "水管", "线管", "管件", "弯头",
    "三通", "接头", "插座", "喷头", "水表", "电表",
}


def _diagnose_cause_from_detail(d: dict) -> str:
    """从benchmark详情条目重新推导错误根因（复用run_benchmark的判定逻辑）

    与run_benchmark._diagnose_cause一致：
    - 无结果 → no_result
    - 专业册不同 → wrong_book
    - 关键词有交集 → wrong_tier（同族不同档位）
    - 其余 → synonym_gap
    """
    import re as _re

    algo_name = d.get('algo_name', '')
    if not algo_name:
        return 'no_result'

    stored_first = d['stored_names'][0] if d.get('stored_names') else ''
    stored_keywords = set(stored_first.replace('(', ' ').replace(')', ' ').split())
    algo_keywords = set(algo_name.replace('(', ' ').replace(')', ' ').split())
    ignore = {'安装', '制作', '周长', 'mm', 'm2', '以内', '≤'}
    stored_keywords -= ignore
    algo_keywords -= ignore

    # 检查专业册是否一致
    def get_book(qid):
        if len(qid) >= 2 and qid[0] == 'C' and qid[1].isalpha():
            letter_map = {'A': 'C1', 'B': 'C2', 'C': 'C3', 'D': 'C4',
                          'E': 'C5', 'F': 'C6', 'G': 'C7', 'H': 'C8',
                          'I': 'C9', 'J': 'C10', 'K': 'C11', 'L': 'C12'}
            return letter_map.get(qid[1], '')
        m = _re.match(r'(C\d+)-', qid)
        if m:
            return m.group(1)
        m = _re.match(r'(\d+)-', qid)
        if m:
            return f'C{m.group(1)}'
        return ''

    stored_id = d['stored_ids'][0] if d.get('stored_ids') else ''
    algo_id = d.get('algo_id', '')
    if stored_id and algo_id:
        if get_book(stored_id) and get_book(algo_id) and get_book(stored_id) != get_book(algo_id):
            return 'wrong_book'

    # 同族判断（关键词有交集 → 同类不同档位）
    family_overlap = stored_keywords & algo_keywords
    if len(family_overlap) > 0:
        return 'wrong_tier'

    return 'synonym_gap'


def benchmark_fix(result_path: str = None, min_freq: int = 2,
                  do_fix: bool = False, skip_benchmark: bool = False,
                  oracle_filter: str = "not_in") -> dict:
    """从benchmark错题中提取同义词缺口并修复

    流程：读错题 → 筛synonym_gap → 提取核心名词对 → 频次排序 → 跨省验证 → 写入

    参数:
        result_path: _latest_result.json路径（默认自动查找最新）
        min_freq: 最小频次阈值（影响几道题才值得加）
        do_fix: 是否写入同义词表
        skip_benchmark: 跳过benchmark回归检查
        oracle_filter: "not_in"=仅召回缺口(P0), "all"=含排序问题(P0+P1)
    """
    from tools.synonym_miner import extract_core_nouns

    # 1. 读取最新benchmark结果
    if result_path is None:
        result_path = str(PROJECT_ROOT / "tests" / "benchmark_papers" / "_latest_result.json")
    rp = Path(result_path)
    if not rp.exists():
        print(f"错误：找不到 {rp}，请先跑 python tools/run_benchmark.py")
        return {"error": "file_not_found"}
    data = json.loads(rp.read_text(encoding='utf-8'))
    print(f"读取: {rp.name} ({data.get('run_time', '?')})")

    # 2. 加载现有同义词（用于排除已有的）
    syn_path = PROJECT_ROOT / "data" / "engineering_synonyms.json"
    existing_syns = json.loads(syn_path.read_text(encoding='utf-8'))

    # 3. 遍历所有错题，筛选synonym_gap
    gap_items = []
    total_wrong = 0
    for result in data['results']:
        province = result['province']
        for d in result['details']:
            if d['is_match']:
                continue
            total_wrong += 1

            # oracle过滤：P0只看召回缺口，all看全部
            if oracle_filter == "not_in" and d.get('oracle_in_candidates', True):
                continue

            # 重新推导错误根因
            cause = _diagnose_cause_from_detail(d)
            if cause != 'synonym_gap':
                continue

            gap_items.append({
                'province': province[:10],
                'bill_name': d['bill_name'],
                'stored_names': d.get('stored_names', []),
                'algo_name': d.get('algo_name', ''),
            })

    # 4. 提取核心名词对并计数
    pair_counter = Counter()  # (bill_core, quota_core) → count
    pair_examples = {}  # (bill_core, quota_core) → [examples]

    for item in gap_items:
        bill_core = extract_core_nouns(item['bill_name'])
        if not item['stored_names']:
            continue
        quota_core = extract_core_nouns(item['stored_names'][0])

        if not bill_core or not quota_core or bill_core == quota_core:
            continue
        if len(bill_core) < 3:  # key最少3字（防过泛化）
            continue
        if bill_core in _GENERIC_WORDS:  # 通用词黑名单
            continue
        if bill_core in existing_syns:  # 已有同义词
            continue

        pair = (bill_core, quota_core)
        pair_counter[pair] += 1
        pair_examples.setdefault(pair, []).append(item)

    # 5. 按频次排序输出
    candidates = []
    for pair, count in pair_counter.most_common():
        if count < min_freq:
            continue
        candidates.append({
            'bill_core': pair[0],
            'quota_core': pair[1],
            'count': count,
            'examples': pair_examples[pair][:3],
        })

    # 打印分析结果
    filter_label = "仅召回缺口(P0)" if oracle_filter == "not_in" else "全部错题"
    print(f"\n{'='*60}")
    print(f"Benchmark错题同义词分析 [{filter_label}]")
    print(f"{'='*60}")
    print(f"  总错题: {total_wrong}条")
    print(f"  筛选后synonym_gap: {len(gap_items)}条")
    print(f"  提取候选词对: {len(candidates)}对 (频次>={min_freq})")

    if not candidates:
        print("  无候选词对（可能需要降低 --min-freq 或用 --oracle all）")
        return {"total_wrong": total_wrong, "gap_items": len(gap_items),
                "candidates": 0, "verified": 0, "written": 0}

    print(f"\n候选同义词（按影响题数排序）：")
    for i, c in enumerate(candidates, 1):
        examples = [e['bill_name'][:15] for e in c['examples'][:2]]
        provinces = list(set(e['province'][:4] for e in c['examples']))
        print(f"  {i}. {c['bill_core']} → {c['quota_core']}  "
              f"({c['count']}题, {'+'.join(provinces[:3])}) "
              f"例: {', '.join(examples)}")

    if not do_fix:
        print(f"\n提示: 加 --fix 参数可跨省验证后写入同义词表")
        return {"total_wrong": total_wrong, "gap_items": len(gap_items),
                "candidates": len(candidates)}

    # 6. 跨省验证
    print(f"\n{'='*60}")
    print(f"跨省验证（{len(candidates)}对候选）")
    print(f"{'='*60}")

    verified_pairs = []
    for c in candidates:
        hit_count, hit_details = _cross_province_verify(c['quota_core'])
        status = "✓通过" if hit_count >= 2 else "✗跳过"
        detail_str = ", ".join(hit_details[:3]) if hit_details else "无命中"
        print(f"  {c['bill_core']}→{c['quota_core']}: "
              f"{hit_count}省命中 [{detail_str}] → {status}")
        if hit_count >= 2:
            verified_pairs.append((c['bill_core'], c['quota_core']))

    if not verified_pairs:
        print("\n  无候选通过跨省验证")
        return {"total_wrong": total_wrong, "gap_items": len(gap_items),
                "candidates": len(candidates), "verified": 0, "written": 0}

    # 7. 写入同义词表
    print(f"\n{'='*60}")
    print(f"写入同义词表（{len(verified_pairs)}对通过验证）")
    print(f"{'='*60}")

    written = _write_synonyms(verified_pairs)
    if written > 0:
        print(f"  已写入 {written} 对同义词到 engineering_synonyms.json")
        # 清除同义词缓存
        try:
            import src.query_builder as qb
            qb._SYNONYMS_CACHE = None
        except Exception:
            pass
    else:
        print("  所有同义词已存在，无需写入")
        return {"total_wrong": total_wrong, "gap_items": len(gap_items),
                "candidates": len(candidates), "verified": len(verified_pairs),
                "written": 0, "benchmark_ok": True}

    # 8. Benchmark回归检查
    if skip_benchmark:
        print("\n  跳过benchmark检查（--skip-benchmark）")
        return {"total_wrong": total_wrong, "gap_items": len(gap_items),
                "candidates": len(candidates), "verified": len(verified_pairs),
                "written": written, "benchmark_ok": True}

    print(f"\n{'='*60}")
    print("Benchmark回归检查")
    print(f"{'='*60}")

    benchmark_ok, benchmark_summary = _run_benchmark_check()
    print(f"  {benchmark_summary}")

    if not benchmark_ok:
        print("  命中率退化，回滚同义词表...")
        if _rollback_synonyms():
            print("  已回滚到修改前的同义词表")
        else:
            print("  警告：回滚失败，请手动检查")

    return {"total_wrong": total_wrong, "gap_items": len(gap_items),
            "candidates": len(candidates), "verified": len(verified_pairs),
            "written": written, "benchmark_ok": benchmark_ok}


def auto_fix_synonyms(synonym_gap_items: list[dict], skip_benchmark: bool = False) -> dict:
    """自动修复同义词缺口（诊断 → 跨省验证 → 写入 → benchmark检查）

    参数:
        synonym_gap_items: 诊断出的同义词缺口条目列表
        skip_benchmark: 跳过benchmark（调试用）

    返回: {"verified": 通过验证数, "written": 写入数, "benchmark_ok": bool, "details": [...]}
    """
    if not synonym_gap_items:
        return {"verified": 0, "written": 0, "benchmark_ok": True, "details": []}

    # 提取去重的同义词对（清单词→定额词）
    pairs: dict[str, str] = {}  # original → variant（去重）
    for item in synonym_gap_items:
        syn = item.get("synonym_gap", {})
        if syn and syn.get("original") and syn.get("variant"):
            orig = syn["original"]
            if orig not in pairs:
                pairs[orig] = syn["variant"]

    if not pairs:
        return {"verified": 0, "written": 0, "benchmark_ok": True, "details": []}

    # ---- 跨省验证 ----
    print()
    print("=" * 60)
    print(f"跨省验证同义词（{len(pairs)}对）")
    print("=" * 60)

    verified_pairs = []  # 通过验证的同义词对
    details = []

    for original, variant in pairs.items():
        hit_count, hit_details = _cross_province_verify(variant)
        status = "通过" if hit_count >= 2 else "跳过（覆盖率不足）"
        detail_str = ", ".join(hit_details) if hit_details else "无命中"
        print(f"  {original}→{variant}: {hit_count}省命中 [{detail_str}] → {status}")

        detail = {
            "original": original, "variant": variant,
            "hit_count": hit_count, "passed": hit_count >= 2,
        }
        details.append(detail)

        if hit_count >= 2:
            verified_pairs.append((original, variant))

    if not verified_pairs:
        print("\n  无同义词通过跨省验证，跳过写入")
        return {"verified": 0, "written": 0, "benchmark_ok": True, "details": details}

    # ---- 写入同义词表 ----
    print()
    print("=" * 60)
    print(f"写入同义词表（{len(verified_pairs)}对通过验证）")
    print("=" * 60)

    written = _write_synonyms(verified_pairs)
    if written > 0:
        print(f"  已写入 {written} 对同义词到 engineering_synonyms.json")
        # 清除同义词缓存（让下次搜索用新同义词）
        try:
            from src.query_builder import _SYNONYMS_CACHE
            import src.query_builder as qb
            qb._SYNONYMS_CACHE = None
        except Exception:
            pass
    else:
        print("  所有同义词已存在，无需写入")
        return {"verified": len(verified_pairs), "written": 0,
                "benchmark_ok": True, "details": details}

    # ---- Benchmark回归检查 ----
    if skip_benchmark:
        print("\n  跳过benchmark检查（--skip-benchmark）")
        return {"verified": len(verified_pairs), "written": written,
                "benchmark_ok": True, "details": details}

    print()
    print("=" * 60)
    print("Benchmark回归检查")
    print("=" * 60)

    benchmark_ok, benchmark_summary = _run_benchmark_check()
    print(f"  {benchmark_summary}")

    if not benchmark_ok:
        # 退化了，回滚！
        print("  命中率退化，回滚同义词表...")
        if _rollback_synonyms():
            print("  已回滚到修改前的同义词表")
        else:
            print("  警告：回滚失败，请手动检查 engineering_synonyms.json")
        return {"verified": len(verified_pairs), "written": written,
                "benchmark_ok": False, "details": details}

    print("  验证通过，同义词已生效")
    return {"verified": len(verified_pairs), "written": written,
            "benchmark_ok": True, "details": details}


def _diagnose_item(item: dict, province: str,
                   sibling_provinces: list[str] | None) -> dict:
    """诊断单条人工审核项的根因"""
    name = item["name"]
    result = {**item, "diagnosis": "", "detail": ""}

    # Step 1: 用清单名直接搜所有库
    hits = _search_all_libraries([name], province, sibling_provinces)
    if hits:
        # 有结果 → 排序偏差（搜到了但匹配阶段没排上来）
        result["diagnosis"] = "排序偏差"
        result["detail"] = f"候选: {hits[0][1]} {hits[0][2][:30]} ({hits[0][0][-6:]})"
        result["candidates"] = hits[:3]
        return result

    # Step 2: 尝试同义替换
    syn_hit = _try_synonym_variants(name, province, sibling_provinces)
    if syn_hit:
        result["diagnosis"] = "同义词缺口"
        result["detail"] = (
            f"清单叫\"{syn_hit['original']}\", "
            f"定额叫\"{syn_hit['variant']}\" "
            f"({syn_hit['hit_id']} {syn_hit['hit_name'][:25]})"
        )
        result["synonym_gap"] = syn_hit
        return result

    # Step 3: 无解
    result["diagnosis"] = "需人工"
    result["detail"] = "定额库中未找到对应项"
    return result


def diagnose(excel_path: str, province: str,
             sibling_provinces: list[str] | None = None,
             auto_fix: bool = False,
             skip_benchmark: bool = False) -> dict:
    """分析已审核Excel的人工项根因，可选自动修复同义词缺口

    参数:
        auto_fix: 自动修复模式（跨省验证→写入同义词→benchmark检查）
        skip_benchmark: 跳过benchmark回归检查（调试用）

    返回: {
        "total": 人工项总数,
        "synonym_gap": [同义词缺口条目],
        "ranking_miss": [排序偏差条目],
        "needs_manual": [需人工条目],
        "fix_result": 自动修复结果(仅auto_fix=True时),
    }
    """
    # 自动获取兄弟库
    if sibling_provinces is None and province:
        from config import get_sibling_provinces
        sibling_provinces = get_sibling_provinces(province)

    # 提取人工项
    items = _extract_manual_from_excel(excel_path)
    if not items:
        print("未找到人工审核项。")
        return {"total": 0, "synonym_gap": [], "ranking_miss": [],
                "needs_manual": [], "suggestions": []}

    # 逐条诊断
    synonym_gap = []
    ranking_miss = []
    needs_manual = []

    for item in items:
        result = _diagnose_item(item, province, sibling_provinces)
        if result["diagnosis"] == "同义词缺口":
            synonym_gap.append(result)
        elif result["diagnosis"] == "排序偏差":
            ranking_miss.append(result)
        else:
            needs_manual.append(result)

    # 输出报告
    _print_report(items, synonym_gap, ranking_miss, needs_manual, excel_path)

    # 保存JSON
    output_dir = Path(excel_path).parent
    stem = Path(excel_path).stem[:40]
    json_path = output_dir / f"diagnosis_{stem}.json"
    report = {
        "total": len(items),
        "synonym_gap": [{k: v for k, v in r.items() if k != "candidates"} for r in synonym_gap],
        "ranking_miss": [{k: v for k, v in r.items() if k != "candidates"} for r in ranking_miss],
        "needs_manual": needs_manual,
    }
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(report, f, ensure_ascii=False, indent=2)
    print(f"\n诊断JSON: {json_path}")

    # ---- 自动修复同义词缺口（--fix 模式）----
    if auto_fix and synonym_gap:
        fix_result = auto_fix_synonyms(synonym_gap, skip_benchmark=skip_benchmark)
        report["fix_result"] = fix_result

    # ---- 写入 Obsidian 诊断报告 ----
    try:
        _write_ob_report(stem, province, report, synonym_gap, ranking_miss, needs_manual)
    except Exception as e:
        print(f"  OB报告写入跳过: {e}")

    return report


def _write_ob_report(stem: str, province: str, report: dict,
                     synonym_gap: list, ranking_miss: list, needs_manual: list):
    """将诊断报告写入 Obsidian 笔记（自动生成，用户可直接查看）"""
    from datetime import datetime
    import re

    ob_dir = Path(r"D:\Obsidian\工程造价\系统更新\诊断报告")
    if not ob_dir.exists():
        return

    today = datetime.now().strftime("%Y-%m-%d")
    # 项目名从stem提取（去掉"匹配结果_"等前缀和时间戳后缀，只留项目名）
    project_name = stem
    for prefix in ["匹配结果_", "diagnosis_"]:
        project_name = project_name.replace(prefix, "")
    project_name = re.split(r"_\d{8}_", project_name)[0]
    project_name = re.split(r"[-—]工程量清单", project_name)[0]
    project_name = project_name[:20]
    ob_path = ob_dir / f"{today} {project_name}.md"

    # 自动提取标签（从清单名称中识别专业分类）
    tag_keywords = {
        "装饰": ["隔断", "吊顶", "天棚", "墙面", "地板", "地面", "涂料", "油漆", "踢脚", "装饰板", "玻璃", "门", "窗"],
        "给排水": ["管道", "阀门", "水管", "排水", "给水", "地漏", "小便器", "洗漱", "水龙头", "减压"],
        "电气": ["配电箱", "开关", "插座", "配管", "电缆", "照明", "灯"],
        "弱电": ["网络", "光纤", "光缆", "监控", "摄像", "信息插座", "大屏", "机柜"],
        "消防": ["消防", "喷淋", "报警", "灭火"],
        "结构": ["钢筋", "混凝土", "砌体", "基础", "防水"],
        "拆除": ["拆除", "铲除"],
        "措施": ["脚手架", "搬运", "保护费"],
    }
    all_names = " ".join(r.get("name", "") for r in synonym_gap + ranking_miss + needs_manual)
    tags = []
    for tag, keywords in tag_keywords.items():
        if any(kw in all_names for kw in keywords):
            tags.append(tag)
    if not tags:
        tags = ["综合"]

    lines = [
        f"---",
        f"topic: Jarvis诊断报告",
        f"project: {project_name}",
        f"province: {province}",
        f"date: {today}",
        f"tags: [{', '.join(tags)}]",
        f"---",
        f"# 诊断报告: {project_name}",
        f"",
        f"**省份**: {province}  ",
        f"**诊断条目**: {report['total']}条  ",
        f"**分布**: 同义词缺口{len(synonym_gap)} | 排序偏差{len(ranking_miss)} | 需人工{len(needs_manual)}",
        f"",
    ]

    # 同义词缺口
    if synonym_gap:
        lines.append(f"## 同义词缺口（{len(synonym_gap)}条）")
        lines.append(f"加同义词可立即修复。")
        lines.append(f"")
        lines.append(f"| 清单名称 | 清单叫法 | 定额叫法 | 命中定额 |")
        lines.append(f"|----------|----------|----------|----------|")
        seen_pairs = set()
        for r in synonym_gap:
            syn = r.get("synonym_gap", {})
            pair_key = f"{syn.get('original', '')}→{syn.get('variant', '')}"
            if pair_key in seen_pairs:
                continue
            seen_pairs.add(pair_key)
            lines.append(
                f"| {r['name'][:20]} | {syn.get('original', '')} "
                f"| {syn.get('variant', '')} "
                f"| {syn.get('hit_id', '')} {syn.get('hit_name', '')[:20]} |"
            )
        lines.append(f"")

        # 自动修复结果
        fix = report.get("fix_result")
        if fix:
            lines.append(f"### 自动修复结果")
            lines.append(f"")
            lines.append(f"| 项目 | 结果 |")
            lines.append(f"|------|------|")
            lines.append(f"| 通过验证 | {fix['verified']}对 |")
            lines.append(f"| 实际写入 | {fix['written']}对 |")
            lines.append(f"| Benchmark | {'通过' if fix['benchmark_ok'] else '退化（已回滚）'} |")
            lines.append(f"")

    # 排序偏差
    if ranking_miss:
        lines.append(f"## 排序偏差（{len(ranking_miss)}条）")
        lines.append(f"正确定额在候选中但没排第一，需要改进排序算法。")
        lines.append(f"")
        lines.append(f"| 清单名称 | 搜到的候选 | 置信度 |")
        lines.append(f"|----------|-----------|--------|")
        for r in ranking_miss[:15]:
            rec = r.get("recommend", "")[:10]
            detail = r.get("detail", "")[:35]
            lines.append(f"| {r['name'][:20]} | {detail} | {rec} |")
        if len(ranking_miss) > 15:
            lines.append(f"| ...还有{len(ranking_miss) - 15}条 | | |")
        lines.append(f"")

    # 需人工
    if needs_manual:
        lines.append(f"## 需人工（{len(needs_manual)}条）")
        lines.append(f"定额库中未找到对应项，需要人工判断。")
        lines.append(f"")
        lines.append(f"| 清单名称 | 项目特征 | 当前匹配 |")
        lines.append(f"|----------|----------|----------|")
        for r in needs_manual[:20]:
            desc = r.get("desc", "")[:30].replace("\n", " ")
            quota = f"{r.get('quota_id', '')} {r.get('quota_name', '')[:20]}" if r.get("quota_id") else "无匹配"
            lines.append(f"| {r['name'][:20]} | {desc} | {quota} |")
        if len(needs_manual) > 20:
            lines.append(f"| ...还有{len(needs_manual) - 20}条 | | |")
        lines.append(f"")

    with open(ob_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    print(f"  OB报告: {ob_path}")


def _print_report(items, synonym_gap, ranking_miss, needs_manual, excel_path):
    """打印格式化诊断报告"""
    stem = Path(excel_path).stem[:30]
    print()
    print("=" * 60)
    print(f"诊断报告: {stem}")
    print(f"  诊断条目: {len(items)}条 (同义词缺口{len(synonym_gap)} "
          f"排序偏差{len(ranking_miss)} 需人工{len(needs_manual)})")
    print("=" * 60)

    if synonym_gap:
        print(f"\n■ 同义词缺口 ({len(synonym_gap)}条) — 加同义词可立即修复")
        # 按同义词对分组
        gap_groups: dict[str, list] = {}
        for r in synonym_gap:
            syn = r.get("synonym_gap", {})
            key = f"{syn.get('original', '')}→{syn.get('variant', '')}"
            gap_groups.setdefault(key, []).append(r)
        for key, group in gap_groups.items():
            count = f" x{len(group)}" if len(group) > 1 else ""
            detail = group[0]["detail"]
            print(f"  {group[0]['name']}{count}: {detail}")

    if ranking_miss:
        print(f"\n■ 排序偏差 ({len(ranking_miss)}条) — 正确定额在候选中但没排第一")
        for r in ranking_miss:
            rec = r.get("recommend", "")
            print(f"  [{rec[:8]}] {r['name'][:20]}: {r['detail']}")

    if needs_manual:
        print(f"\n■ 需人工 ({len(needs_manual)}条) — 系统暂无法处理")
        for r in needs_manual:
            rec = r.get("recommend", "")
            desc_short = r.get("desc", "")[:40]
            print(f"  [{rec[:8]}] {r['name'][:25]}")
            print(f"    特征: {desc_short}")
            print(f"    当前: {r['quota_id']} {r['quota_name'][:30]}")

    # 建议
    if synonym_gap:
        pairs = set()
        for r in synonym_gap:
            syn = r.get("synonym_gap", {})
            if syn:
                pairs.add((syn["original"], syn["variant"]))
        affected = len(synonym_gap)
        print(f"\n建议: 加{len(pairs)}对同义词可消除{affected}条人工项")
        for orig, var in pairs:
            print(f"  {orig} → {var}")

    print("=" * 60)


def main():
    # 已知子命令列表
    SUBCOMMANDS = {"excel", "benchmark-fix", "verify",
                   "audit-coverage", "audit-static", "ranking-report", "bucket"}

    # 向后兼容：如果第一个参数不是子命令，当作excel模式（旧用法）
    if len(sys.argv) > 1 and sys.argv[1] not in SUBCOMMANDS and not sys.argv[1].startswith('-'):
        sys.argv.insert(1, "excel")

    parser = argparse.ArgumentParser(
        description="Jarvis 诊断工具统一入口",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    subparsers = parser.add_subparsers(dest="command", help="子命令")

    # ---- 子命令: excel（原有功能）----
    p_excel = subparsers.add_parser("excel", help="诊断已审核Excel的人工项根因")
    p_excel.add_argument("excel_path", help="已审核Excel文件路径")
    p_excel.add_argument("--province", required=True, help="主定额库名称")
    p_excel.add_argument("--fix", action="store_true",
                         help="自动修复：跨省验证→写入→benchmark回归")
    p_excel.add_argument("--skip-benchmark", action="store_true",
                         help="跳过benchmark回归检查")

    # ---- 子命令: benchmark-fix（错题精补）----
    p_bench = subparsers.add_parser("benchmark-fix",
                                    help="从benchmark错题提取同义词缺口并修复")
    p_bench.add_argument("--input", default=None,
                         help="结果JSON路径（默认最新 _latest_result.json）")
    p_bench.add_argument("--min-freq", type=int, default=2,
                         help="最小频次阈值，影响几道题才入候选（默认2）")
    p_bench.add_argument("--oracle", choices=["not_in", "all"], default="not_in",
                         help="P0仅召回缺口(not_in) / 含排序问题(all)")
    p_bench.add_argument("--fix", action="store_true",
                         help="跨省验证通过后写入同义词表")
    p_bench.add_argument("--skip-benchmark", action="store_true",
                         help="跳过benchmark回归检查")

    # ---- 子命令: verify（跨省搜索验证）----
    p_verify = subparsers.add_parser("verify", help="跨省搜索验证一个关键词")
    p_verify.add_argument("keyword", help="要验证的关键词")

    # ---- 子命令: audit-coverage（同义词覆盖率审计）----
    p_acov = subparsers.add_parser("audit-coverage", help="同义词跨省覆盖率审计")
    p_acov.add_argument("--fix", action="store_true", help="自动修复低覆盖率")
    p_acov.add_argument("--quick", action="store_true", help="只用8个代表省快速筛选")
    p_acov.add_argument("--keyword", type=str, help="只审计含指定关键词的同义词")
    p_acov.add_argument("--min-coverage", type=int, default=8,
                        help="覆盖率低于此值标记为需修复（默认8/24省）")

    # ---- 子命令: audit-static（同义词静态分析）----
    p_ast = subparsers.add_parser("audit-static", help="同义词表静态分析（自映射/冲突等）")
    p_ast.add_argument("--fix", action="store_true", help="生成修复建议JSON")

    # ---- 子命令: ranking-report（排序错误分析）----
    p_rank = subparsers.add_parser("ranking-report", help="排序错误深度分析报告")
    p_rank.add_argument("--input", default=None, help="结果JSON路径（默认最新）")

    # ---- 子命令: bucket（离线诊断分桶）----
    p_bucket = subparsers.add_parser("bucket", help="离线诊断分桶（多维深度分析）")
    p_bucket.add_argument("--input", default=None, help="结果JSON路径（默认最新）")

    args = parser.parse_args()

    if args.command is None:
        parser.print_help()
        sys.exit(0)

    # ---- 分发到各子命令 ----
    if args.command == "excel":
        _cmd_excel(args)
    elif args.command == "benchmark-fix":
        _cmd_benchmark_fix(args)
    elif args.command == "verify":
        _cmd_verify(args)
    elif args.command == "audit-coverage":
        _cmd_audit_coverage(args)
    elif args.command == "audit-static":
        _cmd_audit_static(args)
    elif args.command == "ranking-report":
        _cmd_ranking_report(args)
    elif args.command == "bucket":
        _cmd_bucket(args)


def _cmd_excel(args):
    """excel子命令：诊断已审核Excel（原有功能）"""
    if not os.path.exists(args.excel_path):
        print(f"错误：文件不存在 {args.excel_path}")
        sys.exit(1)
    from config import resolve_province
    try:
        province = resolve_province(args.province)
    except Exception as e:
        print(f"错误：省份解析失败 - {e}")
        sys.exit(1)
    diagnose(args.excel_path, province, auto_fix=args.fix,
             skip_benchmark=args.skip_benchmark)


def _cmd_benchmark_fix(args):
    """benchmark-fix子命令：从错题提取同义词缺口"""
    benchmark_fix(
        result_path=args.input,
        min_freq=args.min_freq,
        do_fix=args.fix,
        skip_benchmark=args.skip_benchmark,
        oracle_filter=args.oracle,
    )


def _cmd_verify(args):
    """verify子命令：跨省搜索验证（替代 cross_province_search.py）"""
    subprocess.run(
        [sys.executable, str(PROJECT_ROOT / "tools" / "cross_province_search.py"),
         args.keyword],
        cwd=str(PROJECT_ROOT),
    )


def _cmd_audit_coverage(args):
    """audit-coverage子命令：同义词跨省覆盖率审计（替代 audit_synonym_coverage.py）"""
    cmd = [sys.executable, str(PROJECT_ROOT / "tools" / "audit_synonym_coverage.py")]
    if args.fix:
        cmd.append("--fix")
    if args.quick:
        cmd.append("--quick")
    if args.keyword:
        cmd.extend(["--keyword", args.keyword])
    if args.min_coverage != 8:
        cmd.extend(["--min-coverage", str(args.min_coverage)])
    subprocess.run(cmd, cwd=str(PROJECT_ROOT))


def _cmd_audit_static(args):
    """audit-static子命令：同义词静态分析（替代 analyze_synonyms.py）"""
    cmd = [sys.executable, str(PROJECT_ROOT / "tools" / "analyze_synonyms.py")]
    if args.fix:
        cmd.append("--fix")
    subprocess.run(cmd, cwd=str(PROJECT_ROOT))


def _cmd_ranking_report(args):
    """ranking-report子命令：排序错误分析（替代 m1_ranking_analysis.py）"""
    cmd = [sys.executable, str(PROJECT_ROOT / "tools" / "m1_ranking_analysis.py")]
    subprocess.run(cmd, cwd=str(PROJECT_ROOT))


# ============================================================
# bucket 子命令 — 离线诊断分桶（多维深度分析）
# ============================================================

def _get_quota_family(quota_id: str) -> str:
    """提取定额家族前缀，如 C4-4-31 → C4-4, 10-1-5 → 10-1"""
    import re as _re
    # 标准化：先去掉可能的字母前缀编码
    parts = quota_id.split('-')
    if len(parts) >= 3:
        return f"{parts[0]}-{parts[1]}"
    elif len(parts) == 2:
        return parts[0]
    return quota_id


def _cmd_bucket(args):
    """bucket子命令：离线诊断分桶（多维深度分析）

    从 _latest_result.json 做5个切片分析，输出报告到 output/temp/diagnosis-bucket.md
    """
    from collections import Counter, defaultdict

    # 找结果文件
    result_path = args.input
    if not result_path:
        result_path = PROJECT_ROOT / "tests" / "benchmark_papers" / "_latest_result.json"
    result_path = Path(result_path)

    if not result_path.exists():
        print(f"错误：找不到 {result_path}，请先跑 python tools/run_benchmark.py")
        sys.exit(1)

    data = json.loads(result_path.read_text(encoding='utf-8'))

    # 收集所有错题+正确题
    correct_items = []   # 正确的题
    ranking_errors = []  # 排序问题（oracle在候选）
    recall_errors = []   # 召回问题（oracle不在候选）
    total = 0

    for r in data['results']:
        prov = r['province']
        for d in r.get('details', []):
            total += 1
            d['_province'] = prov
            if d.get('is_match'):
                correct_items.append(d)
            elif d.get('oracle_in_candidates'):
                ranking_errors.append(d)
            else:
                recall_errors.append(d)

    total_correct = len(correct_items)
    total_wrong = len(ranking_errors) + len(recall_errors)

    lines = []
    lines.append(f"# 诊断分桶报告")
    lines.append(f"")
    lines.append(f"数据来源: {result_path.name}")
    lines.append(f"生成时间: {__import__('datetime').datetime.now().strftime('%Y-%m-%d %H:%M')}")
    lines.append(f"")

    # === 概览 ===
    lines.append(f"## 概览")
    lines.append(f"")
    lines.append(f"| 指标 | 数值 |")
    lines.append(f"|------|------|")
    lines.append(f"| 总题数 | {total} |")
    lines.append(f"| 正确 | {total_correct} ({total_correct*100//total}%) |")
    lines.append(f"| 错误 | {total_wrong} ({total_wrong*100//total}%) |")
    lines.append(f"| P1排序问题 | {len(ranking_errors)} ({len(ranking_errors)*100//max(total_wrong,1)}%) — 搜到了但排序靠后 |")
    lines.append(f"| P0召回问题 | {len(recall_errors)} ({len(recall_errors)*100//max(total_wrong,1)}%) — 根本没搜出来 |")
    lines.append(f"")

    # === 切片1：排名位置直方图 ===
    lines.append(f"## 切片1：排名位置分布（P1排序问题 {len(ranking_errors)}条）")
    lines.append(f"")
    lines.append(f"> 正确答案在候选列表中排第几位？越靠前越容易通过LTR优化修复。")
    lines.append(f"")

    rank_buckets = {"Top2": 0, "Top3-5": 0, "Top6-10": 0, "Top10+": 0}
    for e in ranking_errors:
        cands = e.get('all_candidate_ids', [])
        stored_ids = e.get('stored_ids', [])
        # 找正确答案在候选中的位置
        rank_pos = None
        for sid in stored_ids:
            if sid in cands:
                rank_pos = cands.index(sid) + 1  # 1-indexed
                break

        if rank_pos is None:
            rank_buckets["Top10+"] += 1
        elif rank_pos <= 2:
            rank_buckets["Top2"] += 1
        elif rank_pos <= 5:
            rank_buckets["Top3-5"] += 1
        elif rank_pos <= 10:
            rank_buckets["Top6-10"] += 1
        else:
            rank_buckets["Top10+"] += 1

        e['_rank_pos'] = rank_pos  # 记录排名供后续使用

    lines.append(f"| 排名区间 | 数量 | 占比 | 修复难度 |")
    lines.append(f"|----------|------|------|----------|")
    difficulty = {"Top2": "低（LTR微调）", "Top3-5": "中（LTR+参数特征）",
                  "Top6-10": "高（需更强信号）", "Top10+": "极高（候选截断）"}
    for bucket, count in rank_buckets.items():
        pct = count * 100 // max(len(ranking_errors), 1)
        lines.append(f"| {bucket} | {count} | {pct}% | {difficulty[bucket]} |")
    lines.append(f"")

    # === 切片2：召回同家族率 ===
    lines.append(f"## 切片2：召回同家族分析（P0召回问题 {len(recall_errors)}条）")
    lines.append(f"")
    lines.append(f"> 正确答案没搜到，但候选里有没有「同家族」的定额？")
    lines.append(f"> 同家族 = 定额编号前缀相同（如都是C4-4-*）")
    lines.append(f"")

    family_in = []    # 同家族在候选
    family_out = []   # 同家族不在候选

    for e in recall_errors:
        cands = e.get('all_candidate_ids', [])
        stored_ids = e.get('stored_ids', [])
        # 提取正确答案的家族前缀
        stored_families = {_get_quota_family(sid) for sid in stored_ids if sid}
        # 检查候选里有没有同家族
        cand_families = {_get_quota_family(cid) for cid in cands if cid}
        has_family = bool(stored_families & cand_families)

        if has_family:
            family_in.append(e)
        else:
            family_out.append(e)

    lines.append(f"| 类型 | 数量 | 占比 | 含义 |")
    lines.append(f"|------|------|------|------|")
    lines.append(f"| 同家族在候选 | {len(family_in)} | "
                 f"{len(family_in)*100//max(len(recall_errors),1)}% | "
                 f"搜到了家族但漏了正确型号（参数路由问题） |")
    lines.append(f"| 同家族不在 | {len(family_out)} | "
                 f"{len(family_out)*100//max(len(recall_errors),1)}% | "
                 f"完全搜偏（同义词/品类路由问题） |")
    lines.append(f"")

    # === 切片3：6桶交叉分析 ===
    lines.append(f"## 切片3：根因×Oracle 6桶交叉分析")
    lines.append(f"")

    # 给每条错题打根因标签
    buckets_6 = defaultdict(list)  # (根因, oracle状态) → [items]

    for e in ranking_errors:
        cause = _diagnose_cause_from_detail(e)
        buckets_6[(cause, "排序")].append(e)

    for e in recall_errors:
        cause = _diagnose_cause_from_detail(e)
        buckets_6[(cause, "召回")].append(e)

    bucket_labels = {
        ("wrong_tier", "排序"): ("A1", "选错档位+排序靠后", "LTR参数特征强化"),
        ("wrong_tier", "召回"): ("A2", "选错档位+正确答案被截断", "扩大候选池或参数路由"),
        ("synonym_gap", "排序"): ("B1", "同义词缺口+排序靠后", "同义词扩展+LTR优化"),
        ("synonym_gap", "召回"): ("B2", "同义词缺口+完全搜不到", "同义词扩展（最高优先级）"),
        ("wrong_book", "排序"): ("C1", "搜偏专业+正确答案也在", "品类路由优化"),
        ("wrong_book", "召回"): ("C2", "搜偏专业+正确答案缺失", "品类路由+借用专业"),
        ("no_result", "排序"): ("D1", "无结果但oracle在", "异常，需排查"),
        ("no_result", "召回"): ("D2", "完全无结果", "索引覆盖或数据缺失"),
    }

    lines.append(f"| 桶 | 根因 | Oracle | 数量 | 修复方向 |")
    lines.append(f"|----|----|--------|------|---------|")

    for key in [("wrong_tier", "排序"), ("wrong_tier", "召回"),
                ("synonym_gap", "排序"), ("synonym_gap", "召回"),
                ("wrong_book", "排序"), ("wrong_book", "召回"),
                ("no_result", "排序"), ("no_result", "召回")]:
        items = buckets_6.get(key, [])
        if not items:
            continue
        label_id, desc, fix_dir = bucket_labels.get(key, ("?", "?", "?"))
        lines.append(f"| {label_id} | {key[0]} | {key[1]} | {len(items)} | {fix_dir} |")

    lines.append(f"")

    # 每桶附3条样例
    for key in [("wrong_tier", "排序"), ("wrong_tier", "召回"),
                ("synonym_gap", "排序"), ("synonym_gap", "召回"),
                ("wrong_book", "排序"), ("wrong_book", "召回")]:
        items = buckets_6.get(key, [])
        if not items:
            continue
        label_id, desc, _ = bucket_labels.get(key, ("?", "?", "?"))
        lines.append(f"### {label_id}: {desc}（{len(items)}条，示例3条）")
        lines.append(f"")
        lines.append(f"| 省份 | 清单名 | 系统选的 | 正确答案 | conf |")
        lines.append(f"|------|--------|---------|---------|------|")
        for e in items[:3]:
            prov = e['_province'][:8]
            bill = e.get('bill_name', '?')[:20]
            algo = e.get('algo_name', '?')[:20]
            stored = (e['stored_names'][0][:20] if e.get('stored_names') else '?')
            conf = e.get('confidence', 0)
            lines.append(f"| {prov} | {bill} | {algo} | {stored} | {conf} |")
        lines.append(f"")

    # === 切片4：置信度分布 ===
    lines.append(f"## 切片4：置信度分布（正确 vs 错误）")
    lines.append(f"")

    conf_ranges = [(0, 20), (20, 40), (40, 60), (60, 80), (80, 100)]
    conf_labels = ["0-20", "20-40", "40-60", "60-80", "80-100"]

    lines.append(f"| 区间 | 正确 | 错误 | 错误率 | 含义 |")
    lines.append(f"|------|------|------|--------|------|")

    for (lo, hi), label in zip(conf_ranges, conf_labels):
        correct_count = sum(1 for e in correct_items if lo <= e.get('confidence', 0) < hi)
        wrong_count = sum(1 for e in ranking_errors + recall_errors
                          if lo <= e.get('confidence', 0) < hi)
        total_in_range = correct_count + wrong_count
        err_rate = wrong_count * 100 // max(total_in_range, 1)
        meaning = ""
        if err_rate < 30:
            meaning = "系统有自信且靠谱"
        elif err_rate < 50:
            meaning = "半对半错，需校准"
        else:
            meaning = "高错误率，系统自知不足"
        lines.append(f"| {label} | {correct_count} | {wrong_count} | {err_rate}% | {meaning} |")
    lines.append(f"")

    # === 切片5：省份切片 ===
    lines.append(f"## 切片5：省份表现")
    lines.append(f"")

    # 按省份聚合
    prov_stats = defaultdict(lambda: {"total": 0, "correct": 0, "ranking": 0, "recall": 0})
    for e in correct_items:
        prov_stats[e['_province']]["total"] += 1
        prov_stats[e['_province']]["correct"] += 1
    for e in ranking_errors:
        prov_stats[e['_province']]["total"] += 1
        prov_stats[e['_province']]["ranking"] += 1
    for e in recall_errors:
        prov_stats[e['_province']]["total"] += 1
        prov_stats[e['_province']]["recall"] += 1

    lines.append(f"| 省份 | 总数 | 命中率 | 排序错 | 召回错 | 排序占比 |")
    lines.append(f"|------|------|--------|--------|--------|----------|")

    # 按命中率升序（最差的排前面）
    for prov, s in sorted(prov_stats.items(), key=lambda x: x[1]["correct"]/max(x[1]["total"],1)):
        hit = s["correct"] * 100 // max(s["total"], 1)
        wrong_total = s["ranking"] + s["recall"]
        rank_pct = s["ranking"] * 100 // max(wrong_total, 1) if wrong_total else 0
        lines.append(f"| {prov[:25]} | {s['total']} | {hit}% | {s['ranking']} | {s['recall']} | {rank_pct}% |")
    lines.append(f"")

    # === 结论 ===
    lines.append(f"## 修复优先级建议")
    lines.append(f"")

    # 统计各桶大小
    a1 = len(buckets_6.get(("wrong_tier", "排序"), []))
    a2 = len(buckets_6.get(("wrong_tier", "召回"), []))
    b1 = len(buckets_6.get(("synonym_gap", "排序"), []))
    b2 = len(buckets_6.get(("synonym_gap", "召回"), []))
    top2 = rank_buckets["Top2"]

    lines.append(f"1. **最高收益：A1桶LTR优化**（{a1}条选错档位+排序靠后）")
    lines.append(f"   - 其中Top2有{top2}条 → LTR微调即可修复")
    lines.append(f"2. **同义词扩展**：B2桶（{b2}条完全搜不到）+ B1桶（{b1}条搜到但排后）")
    lines.append(f"3. **参数路由**：A2桶（{a2}条选错档位+被截断）→ 扩大候选池或参数路由")
    lines.append(f"4. **召回同家族**：{len(family_in)}条搜到家族但漏正确型号 → 参数匹配增强")

    # 写文件
    out_path = PROJECT_ROOT / "output" / "temp" / "diagnosis-bucket.md"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text('\n'.join(lines), encoding='utf-8')

    # 终端输出摘要
    print(f"诊断分桶报告已生成: {out_path}")
    print(f"")
    print(f"概览: {total}条 正确{total_correct}({total_correct*100//total}%) 错误{total_wrong}")
    print(f"  P1排序: {len(ranking_errors)}条 | P0召回: {len(recall_errors)}条")
    print(f"")
    print(f"排名分布: Top2={rank_buckets['Top2']} Top3-5={rank_buckets['Top3-5']} "
          f"Top6-10={rank_buckets['Top6-10']} Top10+={rank_buckets['Top10+']}")
    print(f"召回同家族: 在={len(family_in)} 不在={len(family_out)}")
    print(f"")
    print(f"6桶: A1(档位+排序)={a1} A2(档位+召回)={a2} "
          f"B1(同义词+排序)={b1} B2(同义词+召回)={b2}")
    print(f"")
    print(f"详细报告: {out_path}")


if __name__ == "__main__":
    main()
