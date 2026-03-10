# -*- coding: utf-8 -*-
import sys
if sys.stdout.encoding and sys.stdout.encoding.lower() != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')
    sys.stderr.reconfigure(encoding='utf-8')
"""
价格库Excel批量导入工具

两层架构（Codex 5.4审核通过）：
- 公共层：文件路由、单元格清洗、单位标准化、写库审计
- 品类层：每个品类一个解析规则（表头行号、价格列名、单位等）

用法：
    # 扫描目录，看有哪些品类可导入
    python tools/import_price_excel.py --scan "E:/BaiduNetdiskDownload/..."

    # 导入单个文件
    python tools/import_price_excel.py --file "xxx.xlsx" --template pipe

    # 导入整个品类目录
    python tools/import_price_excel.py --dir "安装01.塑料管材管件价格" --template pipe

    # 查看导入统计
    python tools/import_price_excel.py --stats
"""

import os
import re
import json
import argparse
from pathlib import Path
from datetime import datetime

# ======== 单位标准化映射 ========
# Codex建议：先标准化再入库，避免"m"和"米"被当成不同材料
UNIT_NORM = {
    "米": "m", "M": "m",
    "千克": "kg", "公斤": "kg", "KG": "kg",
    "吨": "t", "T": "t",
    "个": "个", "只": "个", "套": "套", "台": "台",
    "根": "根", "条": "根", "支": "根",
    "片": "片", "块": "片",
    "副": "副", "付": "副", "对": "副",
    "卷": "卷", "盘": "卷",
    "桶": "桶", "瓶": "桶",
    "平方米": "m²", "㎡": "m²",
    "立方米": "m³",
    "组": "组",
}

def normalize_unit(unit: str) -> str:
    """单位标准化"""
    u = unit.strip()
    return UNIT_NORM.get(u, u)


# ======== 规格标准化 ========
def normalize_spec(spec: str) -> str:
    """规格标准化（去空格、统一分隔符）"""
    s = spec.strip()
    # De25 → DN25 不做！（塑料管用De，钢管用DN，不能混）
    # 去多余空格
    s = re.sub(r'\s+', ' ', s)
    return s


# ======== 公共导入骨架 ========

def create_import_batch(db, source_file: str, source_type: str,
                        parser_template: str, notes: str = "") -> int:
    """创建导入批次记录，返回batch_id"""
    conn = db._conn()
    try:
        cursor = conn.execute(
            """INSERT INTO import_batch
               (source_type, source_file, parser_template, notes)
               VALUES (?, ?, ?, ?)""",
            (source_type, source_file, parser_template, notes)
        )
        conn.commit()
        return cursor.lastrowid
    finally:
        conn.close()


def update_batch_count(db, batch_id: int, count: int):
    """更新批次导入条数"""
    conn = db._conn()
    try:
        conn.execute(
            "UPDATE import_batch SET record_count=? WHERE id=?",
            (count, batch_id)
        )
        conn.commit()
    finally:
        conn.close()


def import_records(db, records: list, batch_id: int,
                   source_doc: str = "") -> dict:
    """
    批量导入解析好的记录到主材库

    每条record格式：
    {
        "name": "PPR塑料冷水管",   # 材料名称（必填）
        "spec": "De25",            # 规格（可选）
        "unit": "m",               # 单位（可选）
        "brand": "联塑",            # 品牌（可选）
        "category": "管材",         # 大类（可选）
        "subcategory": "PPR管",     # 小类（可选）
        "price": 5.25,             # 含税单价（可选，没有就不建价格记录）
        "tax_rate": 0.13,          # 税率（默认13%）
        "price_unit": "m",         # 价格单位（可选，默认和材料单位一致）
    }

    返回：{"imported": 数, "skipped": 数, "errors": 数}
    """
    stats = {"imported": 0, "skipped": 0, "errors": 0}

    for rec in records:
        name = str(rec.get("name", "")).strip()
        if not name:
            stats["skipped"] += 1
            continue

        spec = normalize_spec(str(rec.get("spec", "")))
        unit = normalize_unit(str(rec.get("unit", "")))

        try:
            # 写入材料主数据（已存在则返回已有ID）
            material_id = db.add_material(
                name=name,
                spec=spec,
                unit=unit,
                category=rec.get("category", ""),
                subcategory=rec.get("subcategory", ""),
                brand=rec.get("brand", ""),
                material_type=rec.get("material_type", ""),
            )

            # 有价格才写价格记录
            price = rec.get("price")
            if price and float(price) > 0:
                tax_rate = float(rec.get("tax_rate", 0.13))
                price_unit = normalize_unit(rec.get("price_unit", unit))
                db.add_price(
                    material_id=material_id,
                    price_incl_tax=float(price),
                    source_type="enterprise_price_lib",
                    tax_rate=tax_rate,
                    unit=price_unit,
                    source_doc=source_doc,
                    batch_id=batch_id,
                    authority_level="reference",
                    usable_for_quote=0,  # 2023年旧价格，仅参考
                    price_date="2023-01-30",  # 会被品类模板覆盖
                )

            stats["imported"] += 1
        except Exception as e:
            print(f"  导入失败: {name} {spec} - {e}")
            stats["errors"] += 1

    return stats


# ======== 品类解析模板 ========

def parse_pipe_vertical(filepath: str) -> list:
    """
    解析管材/管件纵向清单

    典型格式（联塑）：
    序号 | 材料名称 | 品牌 | 型号规格 | 单位 | 数量 | 含税单价（元）| 含税金额
    1   | PPR塑料冷水管 De25 | 联塑 | D25×2.8 S4 | 米 | 3800 | 5.25 | ...
    """
    import openpyxl

    wb = openpyxl.load_workbook(str(filepath), data_only=True, read_only=True)
    records = []

    for ws in wb:
        # 找表头行（包含"材料名称"或"序号"+"单价"的行）
        header_row = None
        col_map = {}

        for i, row in enumerate(ws.iter_rows(max_row=10, values_only=False)):
            cells = {c.column: str(c.value or "").strip() for c in row if hasattr(c, 'column')}
            # 检查是否是表头
            text = " ".join(cells.values())
            if "材料名称" in text or ("序号" in text and "单价" in text):
                header_row = i + 1  # openpyxl 1-indexed
                # 建立列映射
                for col_idx, val in cells.items():
                    vl = val.lower()
                    if "名称" in val and "材料" in val:
                        col_map["name"] = col_idx
                    elif "品牌" in val:
                        col_map["brand"] = col_idx
                    elif "型号" in val or "规格" in val:
                        col_map["spec"] = col_idx
                    elif val == "单位" or "单位" in val:
                        col_map["unit"] = col_idx
                    elif "含税单价" in val or "单价" in val:
                        col_map["price"] = col_idx
                break

        if not header_row or "name" not in col_map:
            continue

        # 读数据行
        for row in ws.iter_rows(min_row=header_row + 1, values_only=False):
            cells = {c.column: c.value for c in row if hasattr(c, 'column')}
            name = str(cells.get(col_map["name"], "") or "").strip()
            if not name:
                continue

            # 从名称中提取规格（如"PPR塑料冷水管 De25"→名称+规格）
            spec = str(cells.get(col_map.get("spec"), "") or "").strip()
            unit = str(cells.get(col_map.get("unit"), "") or "").strip()
            brand = str(cells.get(col_map.get("brand"), "") or "").strip()
            price = cells.get(col_map.get("price"))

            # 名称里有规格信息时提取出来
            name_spec = _extract_spec_from_name(name)
            if name_spec and not spec:
                spec = name_spec

            records.append({
                "name": name,
                "spec": spec,
                "unit": unit,
                "brand": brand,
                "category": "管材",
                "price": price,
                "tax_rate": 0.13,
            })

    wb.close()
    return records


def parse_cable_matrix(filepath: str) -> list:
    """
    解析电线电缆交叉矩阵

    格式（武汉二厂）：
    行2: 型号 | BV | BLV | BVR | ZCNBV | WDZB-BYJ | ...
    行3: 规格 | 铜塑 | 铝塑 | 铜塑软 | 耐火铜塑 | 辐照线 | ...
    行4+: 截面  | 价格 | 价格 | 价格 | ...

    值：元/km，需要转换成元/m（÷1000）
    """
    import openpyxl

    wb = openpyxl.load_workbook(str(filepath), data_only=True, read_only=True)
    records = []

    for ws in wb:
        # 跳过非价格Sheet（如"调价通知"）
        rows = list(ws.iter_rows(max_row=6, values_only=True))
        if len(rows) < 4:
            continue

        # 找型号行和规格行
        type_row = None  # 型号行（BV/RVV/WDZB等）
        spec_row = None  # 描述行（铜塑/辐照线等）
        data_start = None  # 数据开始行

        for i, row in enumerate(rows):
            cells = [str(c or "").strip() for c in row]
            if cells and cells[0] in ("型号", ""):
                # 检查是否有电缆型号
                has_cable_type = any(
                    re.match(r'^[A-Z]', c) for c in cells[1:] if c
                )
                if has_cable_type:
                    type_row = i
                    continue
            if cells and cells[0] in ("规格", ""):
                if type_row is not None and i == type_row + 1:
                    spec_row = i
                    data_start = i + 1
                    break

        if type_row is None:
            continue

        # 读取型号名称
        type_cells = list(rows[type_row])
        desc_cells = list(rows[spec_row]) if spec_row is not None else [None] * len(type_cells)

        # 确定数据开始行（跳过空行）
        if data_start is None:
            data_start = type_row + 2

        # 检查单位（元/km还是元/m）
        is_per_km = False
        for row in rows[:3]:
            for c in row:
                if c and "km" in str(c).lower():
                    is_per_km = True
                    break

        # 读所有数据行
        all_rows = list(ws.iter_rows(values_only=True))
        for row in all_rows[data_start:]:
            cells = list(row)
            if not cells or not cells[0]:
                continue

            cross_section = str(cells[0]).strip()
            if not cross_section or cross_section in ("型号", "规格", ""):
                continue

            # 每列一个型号
            for col_idx in range(1, min(len(cells), len(type_cells))):
                cable_type = str(type_cells[col_idx] or "").strip()
                if not cable_type:
                    continue

                price_val = cells[col_idx]
                if not price_val:
                    continue
                try:
                    price = float(price_val)
                except (ValueError, TypeError):
                    continue

                if price <= 0:
                    continue

                # 元/km → 元/m
                if is_per_km:
                    price = round(price / 1000, 4)

                # 描述（辅助信息）
                desc = str(desc_cells[col_idx] or "").strip() if col_idx < len(desc_cells) else ""

                # 拼接材料名称
                name = cable_type
                spec = cross_section
                if not re.search(r'mm', spec, re.I):
                    spec = spec + "mm²" if "*" in spec or "×" in spec or "x" in spec else spec

                records.append({
                    "name": name,
                    "spec": spec,
                    "unit": "m",
                    "brand": "飞鹤",  # 武汉二厂品牌
                    "category": "电线电缆",
                    "subcategory": desc,
                    "price": round(price, 4),
                    "tax_rate": 0.13,
                    "price_unit": "m",
                })

    wb.close()
    return records


def parse_light_fixture(filepath: str) -> list:
    """
    解析灯具清单

    典型格式（欧普）：
    序号 | 灯具类型名称 | 产品型号 | 产品图片 | 尺寸说明 | 数量 | 含税单价 | ...
    """
    import openpyxl

    wb = openpyxl.load_workbook(str(filepath), data_only=True, read_only=True)
    records = []

    for ws in wb:
        header_row = None
        col_map = {}

        # 找表头
        for i, row in enumerate(ws.iter_rows(max_row=10, values_only=False)):
            cells = {c.column: str(c.value or "").strip() for c in row if hasattr(c, 'column')}
            text = " ".join(cells.values())
            # 灯具表头特征：含"名称"或"灯具类型"且含"单价"或"报价"
            if ("名称" in text or "灯具" in text) and ("单价" in text or "报价" in text or "金额" in text):
                header_row = i + 1
                for col_idx, val in cells.items():
                    if "名称" in val or "类型" in val:
                        col_map["name"] = col_idx
                    elif "型号" in val and "名称" not in val and "类型" not in val:
                        col_map["spec"] = col_idx
                        col_map["spec"] = col_idx
                    elif "尺寸" in val:
                        col_map["size"] = col_idx
                    elif "含税单价" in val or ("单价" in val and "不含税" not in val):
                        col_map["price"] = col_idx
                    elif "不含税" in val:
                        col_map["price_excl"] = col_idx
                    elif "单位" in val:
                        col_map["unit"] = col_idx
                break

        if not header_row or "name" not in col_map:
            continue
        # 至少要有名称和价格
        if "price" not in col_map and "price_excl" not in col_map:
            continue

        for row in ws.iter_rows(min_row=header_row + 1, values_only=False):
            cells = {c.column: c.value for c in row if hasattr(c, 'column')}
            name = str(cells.get(col_map["name"], "") or "").strip()
            if not name or len(name) < 2:
                continue

            spec = str(cells.get(col_map.get("spec"), "") or "").strip()
            size = str(cells.get(col_map.get("size"), "") or "").strip()
            unit = str(cells.get(col_map.get("unit"), "") or "个").strip()

            # 取含税单价优先
            price = cells.get(col_map.get("price"))
            tax_rate = 0.13
            if not price and "price_excl" in col_map:
                excl = cells.get(col_map["price_excl"])
                if excl:
                    try:
                        price = float(excl) * (1 + tax_rate)
                    except (ValueError, TypeError):
                        pass

            # 规格：型号+尺寸拼起来
            full_spec = " ".join(filter(None, [spec, size])).strip()

            records.append({
                "name": name,
                "spec": full_spec,
                "unit": unit,
                "brand": "欧普",  # 默认，实际按文件覆盖
                "category": "灯具",
                "price": price,
                "tax_rate": tax_rate,
            })

    wb.close()
    return records


# ======== 辅助函数 ========

def _extract_spec_from_name(name: str) -> str:
    """从材料名称中提取规格信息"""
    # DN25, De20, Φ10 等
    m = re.search(r'(DN\d+|De\d+|Φ\d+|D\d+)', name, re.I)
    if m:
        return m.group(1)
    return ""


# ======== 通用解析器 ========

def _open_workbook(filepath: str):
    """
    统一打开Excel文件，支持.xlsx和.xls

    返回: (sheets, format)
      sheets = [(sheet_name, header_finder_func, data_reader_func), ...]
      format = 'xlsx' | 'xls'
    """
    fpath = str(filepath)
    if fpath.endswith('.xls') and not fpath.endswith('.xlsx'):
        # .xls 用 xlrd
        import xlrd
        wb = xlrd.open_workbook(fpath)
        sheets = []
        for si in range(wb.nsheets):
            ws = wb.sheet_by_index(si)
            sheets.append((ws.name, ws.nrows, ws.ncols, ws))
        return wb, sheets, 'xls'
    else:
        # .xlsx 用 openpyxl
        import openpyxl
        wb = openpyxl.load_workbook(fpath, data_only=True, read_only=True)
        sheets = []
        for ws in wb:
            sheets.append((ws.title, None, None, ws))
        return wb, sheets, 'xlsx'


def _find_header_and_parse(ws, fmt: str, max_header_row: int = 15,
                           category: str = "", brand: str = "",
                           default_unit: str = "个",
                           skip_sheets: list = None) -> list:
    """
    在单个Sheet中自动查找表头行并解析数据

    表头识别规则：一行中同时包含"名称类关键词"和"价格类关键词"
    """
    records = []

    # 名称列关键词（优先级从高到低）
    name_keywords = ["材料名称", "商品名称", "产品名称", "元件名称",
                     "项目名称", "灯具类型", "清单名称", "设备", "名称"]
    # 价格列关键词
    price_keywords = ["含税单价", "含税战略", "战略单价含税", "综合单价",
                      "战略价", "商品单价", "单价（含", "单价"]
    # 排除的价格列（不含税优先级低）
    price_excl_keywords = ["不含税", "除税"]

    # 读取前max_header_row行找表头
    if fmt == 'xls':
        nrows = ws.nrows
        ncols = ws.ncols
        def get_row_cells(ri):
            return {ci + 1: ws.cell_value(ri, ci) for ci in range(ncols)}
    else:
        # openpyxl
        all_rows = []
        for i, row in enumerate(ws.iter_rows(max_row=max_header_row, values_only=False)):
            cells = {c.column: c.value for c in row if hasattr(c, 'column')}
            all_rows.append(cells)
        def get_row_cells(ri):
            return all_rows[ri] if ri < len(all_rows) else {}
        nrows = max_header_row  # 只用于表头搜索

    # 找表头行
    header_row_idx = None
    col_map = {}

    for ri in range(min(nrows, max_header_row)):
        cells = get_row_cells(ri)
        # 去掉多余空格（配电箱等表格"单   价"→"单价"）
        text_map = {ci: re.sub(r'\s+', '', str(v or "")).strip()
                    for ci, v in cells.items()}
        full_text = " ".join(text_map.values())

        # 必须同时有名称词和价格词
        has_name = any(kw in full_text for kw in name_keywords)
        has_price = any(kw in full_text for kw in price_keywords)

        if not (has_name and has_price):
            continue

        # 建立列映射
        temp_map = {}
        for ci, val in text_map.items():
            if not val:
                continue
            # 名称列
            if "name" not in temp_map:
                for kw in name_keywords:
                    if kw in val:
                        temp_map["name"] = ci
                        break
            # 规格/型号列
            if "spec" not in temp_map:
                if "型号" in val or "规格" in val:
                    # 避免和名称列重复
                    if ci != temp_map.get("name"):
                        temp_map["spec"] = ci
            # 单位列
            if "unit" not in temp_map:
                if val in ("单位", "计量单位") or "计量" in val:
                    temp_map["unit"] = ci
            # 品牌列
            if "brand" not in temp_map:
                if "品牌" in val or "产地" in val:
                    temp_map["brand"] = ci
            # 含税价格列（优先）
            if "price" not in temp_map:
                for kw in price_keywords:
                    if kw in val and not any(ek in val for ek in price_excl_keywords):
                        temp_map["price"] = ci
                        break
            # 不含税价格列（备选）
            if "price_excl" not in temp_map:
                if any(ek in val for ek in price_excl_keywords) and "单价" in val:
                    temp_map["price_excl"] = ci

        if "name" in temp_map and ("price" in temp_map or "price_excl" in temp_map):
            header_row_idx = ri
            col_map = temp_map
            break

    if header_row_idx is None or "name" not in col_map:
        return records

    # 读数据行
    if fmt == 'xls':
        for ri in range(header_row_idx + 1, ws.nrows):
            cells = {ci + 1: ws.cell_value(ri, ci) for ci in range(ncols)}
            rec = _extract_record(cells, col_map, category, brand, default_unit)
            if rec:
                records.append(rec)
    else:
        # openpyxl — 需要重新读
        for row in ws.iter_rows(min_row=header_row_idx + 2, values_only=False):
            cells = {c.column: c.value for c in row if hasattr(c, 'column')}
            rec = _extract_record(cells, col_map, category, brand, default_unit)
            if rec:
                records.append(rec)

    return records


def _extract_record(cells: dict, col_map: dict,
                    category: str, brand: str, default_unit: str) -> dict:
    """从一行单元格中提取记录"""
    name = str(cells.get(col_map["name"], "") or "").strip()
    # 去换行
    name = name.replace("\n", " ").replace("\r", "").strip()
    if not name or len(name) < 2:
        return None
    # 跳过小计/合计/说明行
    if name in ("合计", "小计", "总计", "备注", "说明") or name.startswith("合计"):
        return None

    spec = str(cells.get(col_map.get("spec"), "") or "").strip()
    spec = spec.replace("\n", " ").strip()[:200]  # 截断过长规格
    unit = str(cells.get(col_map.get("unit"), "") or default_unit).strip()
    rec_brand = str(cells.get(col_map.get("brand"), "") or brand).strip()

    # 取价格：优先含税，其次不含税×1.13
    price = None
    if "price" in col_map:
        pv = cells.get(col_map["price"])
        if pv is not None:
            try:
                price = float(pv)
            except (ValueError, TypeError):
                pass
    if price is None and "price_excl" in col_map:
        pv = cells.get(col_map["price_excl"])
        if pv is not None:
            try:
                price = float(pv) * 1.13
            except (ValueError, TypeError):
                pass

    if price is not None and price <= 0:
        price = None

    return {
        "name": name,
        "spec": spec,
        "unit": unit,
        "brand": rec_brand,
        "category": category,
        "price": round(price, 2) if price else None,
        "tax_rate": 0.13,
    }


def parse_generic(filepath: str, category: str = "", brand: str = "",
                  default_unit: str = "个",
                  skip_sheets: list = None) -> list:
    """
    通用解析器：自动识别表头，提取名称+规格+单位+价格

    支持.xlsx和.xls，跳过封面/说明等无数据Sheet
    """
    wb, sheets, fmt = _open_workbook(filepath)
    records = []

    skip_names = skip_sheets or []
    # 默认跳过的Sheet名称关键词
    skip_keywords = ["封面", "说明", "编制", "界面", "品牌表", "汇总",
                     "报价说明", "备注"]

    for sheet_name, _, _, ws in sheets:
        # 跳过指定Sheet
        if sheet_name in skip_names:
            continue
        if any(kw in sheet_name for kw in skip_keywords):
            continue

        recs = _find_header_and_parse(
            ws, fmt, category=category, brand=brand,
            default_unit=default_unit
        )
        records.extend(recs)

    if hasattr(wb, 'close'):
        wb.close()
    return records


# ======== 品类配置表 ========
# 每个品类的解析参数（通用解析器通过这些参数适配不同品类）
CATEGORY_CONFIGS = {
    "02": {"category": "厨房电器", "brand": "", "unit": "台",
           "name": "厨房电器报价清单"},
    "04": {"category": "浴霸", "brand": "", "unit": "台",
           "name": "浴霸清单"},
    "05": {"category": "中央空调", "brand": "", "unit": "台",
           "name": "中央空调报价"},
    "06": {"category": "可视对讲", "brand": "", "unit": "个",
           "name": "可视对讲设备"},
    "07": {"category": "卫浴洁具", "brand": "", "unit": "个",
           "name": "卫浴洁具"},
    "08": {"category": "新风系统", "brand": "", "unit": "台",
           "name": "新风系统设备"},
    "09": {"category": "空气源", "brand": "", "unit": "台",
           "name": "空气源热泵"},
    "10": {"category": "电梯", "brand": "", "unit": "台",
           "name": "电梯设备"},
    "11": {"category": "智能化", "brand": "", "unit": "套",
           "name": "弱电智能化"},
    "12": {"category": "水泵", "brand": "", "unit": "台",
           "name": "水泵"},
    "13": {"category": "太阳能", "brand": "", "unit": "套",
           "name": "太阳能设备"},
    "16": {"category": "卫浴五金", "brand": "", "unit": "个",
           "name": "卫浴五金"},
    "19": {"category": "配电箱", "brand": "", "unit": "台",
           "name": "配电箱"},
    "29": {"category": "抗震支架", "brand": "", "unit": "套",
           "name": "抗震支架"},
    "30": {"category": "发电机", "brand": "", "unit": "套",
           "name": "柴油发电机"},
    "31": {"category": "供水设备", "brand": "", "unit": "",
           "name": "供水工程主材"},
    "32": {"category": "充电桩", "brand": "", "unit": "套",
           "name": "充电桩"},
    "33": {"category": "泛光照明", "brand": "", "unit": "套",
           "name": "泛光照明"},
    "34": {"category": "消防", "brand": "", "unit": "",
           "name": "消防工程"},
    "35": {"category": "水电安装", "brand": "", "unit": "",
           "name": "水电安装"},
    "38": {"category": "供配电", "brand": "", "unit": "",
           "name": "供配电工程"},
}


# ======== 模板注册表 ========
TEMPLATES = {
    "pipe": {
        "name": "管材/管件纵向清单",
        "parser": parse_pipe_vertical,
        "categories": ["安装01"],
    },
    "cable": {
        "name": "电线电缆交叉矩阵",
        "parser": parse_cable_matrix,
        "categories": ["安装24"],
    },
    "light": {
        "name": "灯具清单",
        "parser": parse_light_fixture,
        "categories": ["安装03"],
    },
    "generic": {
        "name": "通用解析器（自动识别表头）",
        "parser": parse_generic,
        "categories": list(CATEGORY_CONFIGS.keys()),
    },
}


# ======== 命令行入口 ========

def cmd_import_file(filepath: str, template: str, dry_run: bool = False):
    """导入单个文件"""
    if template not in TEMPLATES:
        print(f"未知模板: {template}，可选: {list(TEMPLATES.keys())}")
        return

    tmpl = TEMPLATES[template]
    print(f"解析中: {filepath}")
    print(f"模板: {tmpl['name']}")

    records = tmpl["parser"](filepath)
    print(f"解析到 {len(records)} 条记录")

    if not records:
        print("无数据，跳过")
        return

    # 预览前5条
    print("\n前5条预览：")
    for r in records[:5]:
        price_str = f"¥{r.get('price', '?')}" if r.get('price') else "无价格"
        print(f"  {r['name']} | {r.get('spec','')} | {r.get('unit','')} | "
              f"{r.get('brand','')} | {price_str}")

    if dry_run:
        print(f"\n[试运行] 共{len(records)}条，不写库")
        return records

    # 写库
    from src.material_db import MaterialDB
    db = MaterialDB()

    batch_id = create_import_batch(
        db, source_file=str(filepath),
        source_type="enterprise_price_lib",
        parser_template=template,
        notes=f"品类: {tmpl['name']}"
    )

    stats = import_records(db, records, batch_id, source_doc=Path(filepath).name)
    update_batch_count(db, batch_id, stats["imported"])

    print(f"\n导入完成 (batch #{batch_id}):")
    print(f"  成功: {stats['imported']}")
    print(f"  跳过: {stats['skipped']}")
    print(f"  错误: {stats['errors']}")

    return stats


def cmd_stats():
    """查看导入统计"""
    from src.material_db import MaterialDB
    db = MaterialDB()

    s = db.stats()
    print("主材库统计：")
    for k, v in s.items():
        print(f"  {k}: {v}")

    # 查看导入批次
    conn = db._conn()
    try:
        rows = conn.execute(
            "SELECT * FROM import_batch ORDER BY created_at DESC LIMIT 10"
        ).fetchall()
        if rows:
            print("\n最近导入批次：")
            for r in rows:
                print(f"  #{r['id']} {r['source_type']} "
                      f"{r['parser_template']} "
                      f"{r['record_count']}条 "
                      f"{r['created_at']}")
    finally:
        conn.close()


def main():
    parser = argparse.ArgumentParser(description="价格库Excel批量导入工具")
    parser.add_argument("--file", help="导入单个Excel文件")
    parser.add_argument("--template", "-t", help="解析模板(pipe/cable/light)")
    parser.add_argument("--dry-run", action="store_true", help="试运行，不写库")
    parser.add_argument("--stats", action="store_true", help="查看导入统计")
    parser.add_argument("--list-templates", action="store_true", help="列出可用模板")

    args = parser.parse_args()

    if args.list_templates:
        print("可用模板：")
        for k, v in TEMPLATES.items():
            print(f"  {k}: {v['name']} (适用: {v['categories']})")
    elif args.stats:
        cmd_stats()
    elif args.file:
        if not args.template:
            parser.error("--file 需要指定 --template")
        cmd_import_file(args.file, args.template, dry_run=args.dry_run)
    else:
        parser.print_help()


if __name__ == "__main__":
    main()
