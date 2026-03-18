"""
预算数据导入工具
功能：
1. 读取做好的预算Excel文件（广联达/造价Home等导出，清单+定额交替排列格式）
2. 解析清单→定额对应关系
3. 一次导入，自动往两个库都存：
   - 经验库：带定额编号，同省份项目可直接匹配
   - 通用知识库：定额名称模式，跨省份通用

使用方法：
    # 导入一个做好的预算文件（默认北京）
    python tools/import_reference.py 预算文件.xlsx

    # 指定省份和项目名
    python tools/import_reference.py 预算文件.xlsx --province 北京 --project 丰台安置房

    # 查看解析结果但不导入（调试用）
    python tools/import_reference.py 预算文件.xlsx --dry-run

Excel格式说明（广联达云计价导出格式）：
- 清单行：有序号，编码为12位数字，有项目特征描述
- 定额行：无序号，编码为 C4-4-31 或 C10-1-5 格式，无项目特征描述
- 章节标题行：只有"项目名称"列有值，其余为空
"""

import argparse
import re
import sys
from pathlib import Path

# 把项目根目录加入路径，这样才能导入 src/ 下的模块
PROJECT_ROOT = Path(__file__).parent.parent

from loguru import logger
from src.text_parser import normalize_bill_text
from src.bill_reader import (
    _is_material_code as _shared_is_material_code,
    _is_quota_code as _shared_is_quota_code,
)
from db.sqlite import connect as _db_connect
import config


# 缓存：避免同一省份的定额编号集合被重复加载
_quota_id_sets_cache: dict[str, set[str]] = {}


def _quota_ids_exist_in_province(quota_ids: list[str], province: str) -> bool:
    """检查定额编号是否存在于指定省份的定额库中

    用于多定额库导入时的路由判断：定额编号属于哪个省的库，就导入到哪个省。
    只要有一个编号命中，就认为属于该省（一条清单的多个定额通常属于同一省）。

    参数:
        quota_ids: 定额编号列表（如 ["C10-1-5", "C10-1-6"]）
        province: 省份/定额库名称

    返回:
        True 表示至少有一个编号在该省的定额库中存在
    """
    global _quota_id_sets_cache

    if not quota_ids or not province:
        return False

    # 从缓存加载，避免重复读库
    if province not in _quota_id_sets_cache:
        quota_db_path = config.get_quota_db_path(province)
        if not quota_db_path.exists():
            _quota_id_sets_cache[province] = set()
        else:
            try:
                conn = _db_connect(quota_db_path)
                try:
                    rows = conn.execute("SELECT quota_id FROM quotas").fetchall()
                    _quota_id_sets_cache[province] = {row[0] for row in rows}
                finally:
                    conn.close()
            except Exception as e:
                logger.warning(f"加载{province}定额库失败: {e}")
                _quota_id_sets_cache[province] = set()

    id_set = _quota_id_sets_cache[province]
    if not id_set:
        return False

    # 清洗编号后再查（和 experience_db._validate_quota_ids 的清洗逻辑一致）
    for qid in quota_ids:
        qid_clean = qid.strip().replace(" ", "")
        qid_clean = re.sub(r'换$', '', qid_clean)
        if qid_clean.startswith("借"):
            qid_clean = qid_clean[1:]
        qid_clean = re.sub(r'\*[\d.]+$', '', qid_clean)
        qid_clean = qid_clean.strip()
        if qid_clean in id_set:
            return True

    return False


def read_excel_pairs(excel_path: str) -> list[dict]:
    """
    读取造价Home导出的Excel，解析清单→定额对应关系

    返回:
        [
            {
                "bill_name": "钢筋混凝土蓄水池",
                "bill_desc": "项目特征描述...",
                "bill_code": "010507006001",
                "bill_unit": "m³",
                "bill_pattern": "钢筋混凝土蓄水池 项目特征...",  # 用于通用知识库
                "quotas": [
                    {"code": "5-325", "name": "混凝土蓄水池 C30"},
                    {"code": "5-90",  "name": "模板安装"},
                ]
            },
            ...
        ]
    """
    import openpyxl

    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    try:
        # 遍历所有Sheet（造价Home导出的Excel可能有多个工作簿）
        pairs = []
        sheet_names = wb.sheetnames
        logger.info(f"Excel包含 {len(sheet_names)} 个工作簿: {sheet_names[:10]}{'...' if len(sheet_names) > 10 else ''}")

        for sheet_name in sheet_names:
            ws = wb[sheet_name]
            current_bill = None  # 当前正在处理的清单项
            current_section = ""  # 当前分部标题（如"给排水工程"）
            sheet_pairs = 0  # 当前Sheet解析出的清单数

            for row in ws.iter_rows(min_row=1, values_only=True):
                # 跳过空行
                if not row or all(cell is None for cell in row):
                    continue

                # 确保至少有9列（序号/编码/名称/特征/空列/单位/工程量/单价/合价）
                cells = list(row) + [None] * (9 - len(row)) if len(row) < 9 else list(row)

                col_a = str(cells[0] or "").strip()  # 序号
                col_b = str(cells[1] or "").strip()  # 项目编码 / 定额编号
                col_c = str(cells[2] or "").strip()  # 项目名称 / 定额名称
                col_d = str(cells[3] or "").strip()  # 项目特征描述
                # cells[4] 通常是项目特征的合并空列
                col_unit = str(cells[5] or "").strip()  # 计量单位（第6列）
                # cells[6] 是工程量
                col_price = cells[7]  # 综合单价/主材单价（第8列），保留原始数值

                # 跳过表头行（"序号"、"项目编码"等文字）
                if col_a in ("序号", "") and col_b in ("项目编码", "编码", ""):
                    if col_c in ("项目名称", "名称", ""):
                        continue

                # 判断行类型
                row_type = _classify_row(col_a, col_b, col_c, col_d)

                if row_type == "bill":
                    # 清单行：保存之前的清单（如果有），开始新的清单
                    if current_bill and current_bill["quotas"]:
                        pairs.append(current_bill)
                        sheet_pairs += 1

                    # 构建清单模式文本（使用共享的规范化函数，确保和匹配时格式一致）
                    bill_pattern = normalize_bill_text(col_c, col_d)

                    current_bill = {
                        "bill_name": col_c,
                        "bill_desc": col_d,
                        "bill_code": col_b,
                        "bill_unit": col_unit,
                        "bill_pattern": bill_pattern,
                        "section": current_section,  # 继承当前分部标题
                        "quotas": [],
                    }

                elif row_type == "quota" and current_bill is not None:
                    # 定额行：挂到当前清单下（带空的materials列表，后续主材行填充）
                    current_bill["quotas"].append({
                        "code": col_b,
                        "name": col_c,
                        "materials": [],  # 主材列表，由后续material行填充
                    })

                elif row_type == "material" and current_bill is not None:
                    # 主材行：挂到当前清单的最后一条定额下
                    if current_bill["quotas"]:
                        # 解析主材单价（cells[7]，可能为空或非数字）
                        mat_price = None
                        try:
                            if col_price is not None:
                                mat_price = round(float(col_price), 2)
                        except (ValueError, TypeError):
                            pass

                        current_bill["quotas"][-1]["materials"].append({
                            "code": col_b,
                            "name": col_c,
                            "unit": col_unit,
                            "price": mat_price,  # 主材单价（元），可能为None
                        })

                elif row_type == "other" and col_c and len(col_c) >= 2:
                    # 可能是分部/章节标题行（如"给排水工程"、"电气安装"）
                    # 保存下来供后续清单项继承
                    current_section = col_c

            # 保存当前Sheet最后一个清单
            if current_bill and current_bill["quotas"]:
                pairs.append(current_bill)
                sheet_pairs += 1

            if sheet_pairs > 0:
                logger.debug(f"  工作簿 '{sheet_name}': {sheet_pairs} 条清单")

        return pairs
    finally:
        wb.close()


def _classify_row(col_a: str, col_b: str, col_c: str, col_d: str) -> str:
    """
    判断一行是清单行、定额行、主材行还是标题行

    判断顺序很重要（优先级从高到低）：
    1. 清单行：9-12位数字编码 + 有项目特征描述（col_d非空）
    2. 主材行：含Z@或纯数字7位+，无项目特征描述（先排除主材，再判定额）
    3. 定额行：编码格式为 X-XXX 或 字母+数字（如D00003）
    4. 标题行/其他：不符合以上规则
    """
    has_serial = bool(re.match(r'^\d+$', col_a))  # 序号是纯数字
    has_bill_code = bool(re.match(r'^\d{9,12}$', col_b))  # 9-12位数字编码
    has_desc = bool(col_d)  # 有项目特征描述

    # ① 清单行：有12位编码+有项目特征描述，或有序号+名称+描述
    # 关键：必须有 has_desc，否则12位主材编码（如080801013001）会误判
    if has_bill_code and col_c and has_desc:
        return "bill"
    if has_serial and col_c and has_desc:
        return "bill"

    # ② 主材行（必须在定额之前判断，否则材料编码里的@会被后续规则误处理）
    # 直接复用 bill_reader 的主材编码规则，保持“导入带定额清单”和主读取链路一致。
    # 例如：01190031、26010101Z@2、SZFWJC001@2、CL17033110-1@1、补充主材005
    # 注意：有序号的行是清单项（如 [7] 080801013001 防鼠板），不是主材
    if col_c and not col_d and not has_serial:
        is_material_code = _shared_is_material_code(col_b)
        if is_material_code:
            return "material"

    # ③ 定额行：编码格式为 X-XXX 或 字母开头+数字
    # 常见定额编号格式：5-325, 8-2947, D00003, 1-790, 5-92换, AD0003换
    # 江西等省份特殊格式：1-45 换（带空格）、借14-17 换（借用其他册）、补子目1
    cleaned_code = col_b.replace(" ", "").rstrip("换")  # 去空格和"换"后缀
    if cleaned_code.startswith("借"):
        cleaned_code = cleaned_code[1:]  # 去"借"前缀

    quota_id_patterns = (
        # 字母/字母+数字前缀 + 多段数字（如 C10-1-5, A-1-1, SC1-1-1, GY-1）
        r'^[A-Za-z]{1,3}\d{0,2}(-\d+)+$',
        # 纯数字前缀 + 多段数字（如 99-1-1, 2003-1-1）
        r'^\d{1,4}(-\d+)+$',
        # 字母开头+连续数字（如 D00003, B010, C00187@1）
        r'^[A-Za-z]\d{3,}',
    )
    is_quota_code = any(re.match(pat, cleaned_code) for pat in quota_id_patterns) or bool(re.match(
        r'^补子目', col_b  # 补充子目（如"补子目1"）
    ))

    if is_quota_code and col_c:
        return "quota"

    # ④ 兜底：没有项目特征描述的12位编码清单（少数Excel格式描述在别的列）
    if has_bill_code and col_c and has_serial:
        return "bill"

    return "other"


def _classify_row_v2(col_a: str, col_b: str, col_c: str, col_d: str) -> str:
    """判断一行是清单、定额、主材还是其他。"""
    has_serial = bool(re.fullmatch(r"\d+", col_a))
    has_bill_code = bool(re.fullmatch(r"\d{9,12}", col_b))
    has_desc = bool(col_d)

    if has_bill_code and col_c and has_desc:
        return "bill"
    if has_serial and col_c and has_desc:
        return "bill"

    if col_c and not col_d and not has_serial and _shared_is_material_code(col_b):
        return "material"

    normalized_quota_code = col_b.replace(" ", "").strip()
    if normalized_quota_code.endswith("换"):
        normalized_quota_code = normalized_quota_code[:-1]
    if normalized_quota_code.startswith("借"):
        normalized_quota_code = normalized_quota_code[1:]

    if col_c and (
        _shared_is_quota_code(normalized_quota_code)
        or normalized_quota_code.startswith("补子目")
    ):
        return "quota"

    if has_bill_code and col_c and has_serial:
        return "bill"

    return "other"


_classify_row = _classify_row_v2


def convert_to_kb_records(pairs: list[dict]) -> list[dict]:
    """
    将清单→定额对转换为通用知识库的导入格式

    关键：只保留定额名称（不保留编号），因为编号是省份专属的
    同时通过 specialty_classifier 自动判断每条记录的专业册号
    """
    from src.specialty_classifier import classify as classify_specialty, parse_section_title

    records = []

    for pair in pairs:
        if not isinstance(pair, dict):
            continue
        bill_pattern = pair.get("bill_pattern", "")
        if not isinstance(bill_pattern, str) or not bill_pattern.strip():
            continue

        # 主定额名称模式列表
        quota_patterns = []
        quotas = pair.get("quotas", [])
        if not isinstance(quotas, list):
            quotas = []
        for q in quotas:
            if not isinstance(q, dict):
                continue
            name = str(q.get("name", "")).strip()
            if name:
                quota_patterns.append(name)

        if not quota_patterns:
            continue

        # 判断专业：优先用分部标题，其次用关键词匹配
        section = pair.get("section", "")
        specialty = None
        if section:
            specialty = parse_section_title(section)
        if not specialty:
            classification = classify_specialty(pair.get("bill_name", ""), pair.get("bill_desc", ""))
            specialty = classification.get("primary")

        records.append({
            "bill_pattern": bill_pattern,
            "quota_patterns": quota_patterns,
            "associated_patterns": [],  # 暂时为空，后续可从定额关系中提取
            "param_hints": {},          # 暂时为空，后续可从text_parser提取
            "specialty": specialty,     # 专业册号（如"C10"）
        })

    return records


def import_to_experience(pairs: list[dict], project_name: str,
                         province: str = None, all_provinces: list[str] = None,
                         source: str = "project_import",
                         skip_vector: bool = False):
    """
    将清单→定额对导入经验库（带定额编号，同省份可直接匹配）

    支持多定额库：按顺序尝试每个定额库，定额编号在哪个库能校验通过就存到哪个。

    参数:
        pairs: read_excel_pairs() 返回的清单→定额对列表
        project_name: 项目名称（用于标记来源）
        province: 主定额库（单省份模式，兼容旧调用）
        all_provinces: 所有选中的定额库列表（多省份模式，优先使用）

    返回:
        {"added": 新增数, "skipped": 跳过数}
    """
    from src.experience_db import ExperienceDB

    # 构建省份尝试列表（主省份在前）
    provinces_to_try = []
    if all_provinces:
        provinces_to_try = list(all_provinces)
    elif province:
        provinces_to_try = [province]

    exp_db = ExperienceDB()
    inserted = 0
    matched_existing = 0
    duplicate_hits = 0
    written = 0
    skipped = 0
    seen_record_ids = set()

    for pair in pairs:
        if not isinstance(pair, dict):
            skipped += 1
            logger.warning(f"经验库导入跳过: 记录结构非法（期望dict，实际{type(pair).__name__}）")
            continue
        bill_text = pair.get("bill_pattern", "")  # 清单名称+特征描述
        if not isinstance(bill_text, str) or not bill_text.strip():
            skipped += 1
            logger.warning(f"经验库导入跳过: 清单'{pair.get('bill_name', '')[:40]}' 缺少有效bill_pattern")
            continue
        quotas = pair.get("quotas", [])
        if not isinstance(quotas, list):
            skipped += 1
            logger.warning(f"经验库导入跳过: 清单'{pair.get('bill_name', '')[:40]}' quotas结构非法")
            continue
        quota_ids = [q.get("code", "") for q in quotas if isinstance(q, dict) and q.get("code")]
        quota_names = [q.get("name", "") for q in quotas if isinstance(q, dict) and q.get("code")]

        # 提取主材信息：每条定额下的materials合并成一个列表
        # 格式：[{"quota_code": "4-14-379", "name": "单联单控开关", "unit": "只", "price": 4.57}, ...]
        materials = []
        for q in quotas:
            if not isinstance(q, dict) or not q.get("code"):
                continue
            for m in q.get("materials", []):
                if isinstance(m, dict) and m.get("name"):
                    mat_entry = {
                        "quota_code": q["code"],
                        "name": m["name"],
                        "code": m.get("code", ""),
                        "unit": m.get("unit", ""),
                    }
                    if m.get("price") is not None:
                        mat_entry["price"] = m["price"]
                    materials.append(mat_entry)

        if not quota_ids:
            skipped += 1
            continue

        try:
            # 多定额库模式：按顺序尝试每个定额库，定额编号在哪个库就存到哪个
            record_id = -1
            for try_province in provinces_to_try:
                # 预检查：定额编号是否存在于该省的定额库
                # 没有这个检查的话，add_experience() 对缺失编号只发警告不报错，
                # 导致循环总是停在第一个省，把所有数据都导入到错误的库
                if len(provinces_to_try) > 1 and not _quota_ids_exist_in_province(quota_ids, try_province):
                    continue  # 编号不属于这个省，跳过试下一个

                existing = exp_db._find_exact_match(
                    bill_text,
                    try_province,
                    authority_only=False,
                )
                existing_record_id = int(existing["id"]) if existing else None
                record_id = exp_db.add_experience(
                    bill_text=bill_text,
                    quota_ids=quota_ids,
                    quota_names=quota_names,
                    materials=materials,
                    bill_name=pair.get("bill_name"),     # 补传：清单名称
                    bill_code=pair.get("bill_code"),     # 补传：清单编码
                    bill_unit=pair.get("bill_unit"),     # 补传：计量单位
                    confidence=90,
                    source=source,
                    project_name=project_name,
                    province=try_province,
                    skip_vector=skip_vector,
                )
                if record_id > 0:
                    break  # 导入成功，不再尝试下一个定额库

            if record_id > 0:
                written += 1
                if existing_record_id is None:
                    inserted += 1
                    seen_record_ids.add(record_id)
                elif record_id in seen_record_ids:
                    duplicate_hits += 1
                else:
                    matched_existing += 1
                    seen_record_ids.add(record_id)
            else:
                skipped += 1
        except Exception as e:
            logger.warning(
                f"经验库导入失败并跳过: bill='{pair.get('bill_name', '')[:40]}', "
                f"code='{pair.get('bill_code', '')}', error={e}"
            )
            skipped += 1

    return {
        "inserted": inserted,
        "matched_existing": matched_existing,
        "duplicate_hits": duplicate_hits,
        "written": written,
        "skipped": skipped,
    }


def _select_quota_db() -> str:
    """交互式选择定额库版本

    列出 db/provinces/ 下所有已导入的定额库，让用户选择。
    经验导入必须绑定具体的定额库版本，才能：
    1. 校验定额编号是否存在
    2. 判断经验是否过期（定额库更新后 stale 检测）

    返回:
        完整的省份定额版本名称（如 "北京市建设工程施工消耗量标准(2024)"）
    """
    from db.sqlite import connect as _db_connect
    db_provinces = config.list_db_provinces()

    if not db_provinces:
        print("\n  [错误] 没有找到已导入的定额库")
        print(f"  请先运行 python tools/import_all.py 导入定额数据")
        sys.exit(1)

    if len(db_provinces) == 1:
        selected = db_provinces[0]
        print(f"\n  只有1个定额库，自动选择: {selected}")
        return selected

    # 列出所有已导入的定额库（含定额条数，帮助用户区分）
    print("\n  请选择该预算文件使用的定额库:")
    print()
    for i, p in enumerate(db_provinces, 1):
        db_path = config.get_quota_db_path(p)
        db_info = ""
        if db_path.exists():
            try:
                conn = _db_connect(db_path)
                try:
                    count = conn.execute("SELECT COUNT(*) FROM quotas").fetchone()[0]
                finally:
                    conn.close()
                db_info = f"{count}条定额"
            except Exception:
                db_info = "数据库已存在"
        print(f"    [{i}] {p}  ({db_info})")

    print()
    while True:
        try:
            choice = input(f"  输入编号 [1-{len(db_provinces)}]: ").strip()
            idx = int(choice) - 1
            if 0 <= idx < len(db_provinces):
                selected = db_provinces[idx]
                print(f"  → 已选择: {selected}")
                return selected
            print(f"  编号超出范围，请输入 1-{len(db_provinces)}")
        except ValueError:
            print(f"  请输入数字编号")
        except EOFError:
            # 非交互环境（如管道/脚本调用），无法读取输入，直接退出
            print("\n  [错误] 非交互环境无法选择定额库，请用 --province 参数指定")
            sys.exit(1)


def main():
    parser = argparse.ArgumentParser(
        description="预算数据导入工具 - 从做好的预算Excel导入经验（一次导入，经验库+通用知识库两边都存）",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
使用示例:
  # 导入预算文件（交互式选择定额库版本）
  python tools/import_reference.py 预算文件.xlsx

  # 指定定额库版本（跳过交互选择）
  python tools/import_reference.py 预算文件.xlsx --province "北京市建设工程施工消耗量标准(2024)"

  # 指定项目名称
  python tools/import_reference.py 预算文件.xlsx --project 丰台安置房

  # 查看解析结果但不导入（调试用）
  python tools/import_reference.py 预算文件.xlsx --dry-run
        """,
    )
    parser.add_argument("input_file", help="带定额的预算Excel文件（广联达/造价Home等导出）")
    parser.add_argument("--province", default=None,
                        help="定额库版本全称（如 '北京市建设工程施工消耗量标准(2024)'）。"
                             "不指定则交互式选择。")
    parser.add_argument("--aux-provinces", default=None,
                        help="辅助定额库（逗号分隔）。安装+土建混合项目时，"
                             "定额编号在主定额库校验不过会自动尝试辅助定额库。")
    parser.add_argument("--project", default=None, help="项目名称（默认用文件名）")
    parser.add_argument("--dry-run", action="store_true", help="只解析不导入（调试用）")
    parser.add_argument("--trust", action="store_true",
                        help="信任模式：数据进权威层（默认进候选层，需人工确认后晋升）")
    parser.add_argument("--no-analyze", action="store_true",
                        help="导入后不自动分析生成方法卡片（默认会自动分析）")

    args = parser.parse_args()

    # 验证文件
    input_path = Path(args.input_file)
    if not input_path.exists():
        logger.error(f"文件不存在: {input_path}")
        sys.exit(1)

    # 确定定额库版本：指定了就解析匹配，没指定就交互选择
    if args.province:
        try:
            province = config.resolve_province(args.province)
        except ValueError as e:
            logger.error(f"省份解析失败: {e}")
            sys.exit(1)
        # 校验该省份是否有定额库（导入场景允许库不存在，只是校验会宽松些）
        db_path = config.get_quota_db_path(province)
        if not db_path.exists():
            logger.warning(f"主定额库尚未导入: {province}（导入的记录将不做定额编号校验）")
    else:
        province = _select_quota_db()

    # 解析辅助定额库（安装+土建混合项目时使用）
    aux_provinces = []
    if args.aux_provinces:
        for ap in args.aux_provinces.split(","):
            ap = ap.strip()
            if not ap:
                continue
            try:
                resolved = config.resolve_province(ap)
                if resolved != province:  # 不重复加主定额库
                    aux_provinces.append(resolved)
            except ValueError:
                logger.warning(f"辅助定额库解析失败，跳过: {ap}")

    # 构建完整的省份列表（主定额库在前）
    all_provinces = [province] + aux_provinces

    project_name = args.project or input_path.stem

    # 第1步：解析Excel
    logger.info(f"解析文件: {input_path}")
    pairs = read_excel_pairs(str(input_path))
    logger.info(f"解析完成: {len(pairs)}条清单→定额对应关系")

    if not pairs:
        logger.warning("未找到有效的清单→定额数据，请检查文件格式")
        sys.exit(1)

    # 打印解析结果摘要
    total_quotas = sum(
        len(p.get("quotas", []))
        for p in pairs
        if isinstance(p, dict) and isinstance(p.get("quotas", []), list)
    )
    # 统计主材：遍历所有清单→定额→主材
    total_materials = 0
    bills_with_materials = 0
    for p in pairs:
        if not isinstance(p, dict):
            continue
        has_mat = False
        for q in p.get("quotas", []):
            if isinstance(q, dict):
                mat_count = len(q.get("materials", []))
                total_materials += mat_count
                if mat_count > 0:
                    has_mat = True
        if has_mat:
            bills_with_materials += 1

    logger.info(f"  清单项: {len(pairs)}条")
    logger.info(f"  定额项: {total_quotas}条")
    logger.info(f"  主材项: {total_materials}条（{bills_with_materials}条清单含主材）")
    avg_per_bill = (total_quotas / len(pairs)) if pairs else 0
    logger.info(f"  平均每条清单: {avg_per_bill:.1f}条定额")

    # 打印前5条示例
    logger.info("--- 示例（前5条）---")
    for i, pair in enumerate(pairs[:5]):
        quotas = pair.get("quotas", []) if isinstance(pair, dict) else []
        if not isinstance(quotas, list):
            quotas = []
        quota_names = [str(q.get("name", ""))[:30] for q in quotas if isinstance(q, dict) and q.get("name")]
        if not quota_names:
            quota_names = ["(无有效定额名称)"]
        bill_name = pair.get("bill_name", "") if isinstance(pair, dict) else ""
        logger.info(f"  [{i+1}] {str(bill_name)[:40]}")
        logger.info(f"      → {', '.join(quota_names)}")

    if args.dry_run:
        logger.info("--- dry-run模式，不导入 ---")
        for i, pair in enumerate(pairs):
            bill_name = pair.get("bill_name", "") if isinstance(pair, dict) else ""
            bill_desc = pair.get("bill_desc", "") if isinstance(pair, dict) else ""
            bill_code = pair.get("bill_code", "") if isinstance(pair, dict) else ""
            bill_pattern = pair.get("bill_pattern", "") if isinstance(pair, dict) else ""
            quotas = pair.get("quotas", []) if isinstance(pair, dict) else []
            if not isinstance(quotas, list):
                quotas = []
            print(f"\n--- 第{i+1}条 ---")
            print(f"  清单: {bill_name}")
            if bill_desc:
                print(f"  特征: {str(bill_desc)[:100]}")
            print(f"  编码: {bill_code}")
            print(f"  模式: {str(bill_pattern)[:80]}")
            for q in quotas:
                if not isinstance(q, dict):
                    continue
                print(f"  定额: {q.get('code', '')} → {q.get('name', '')}")
        return

    # 第2步：导入经验库（带定额编号，自动路由到对应的定额库）
    # 已做好的预算是造价人员审过的，默认进权威层；加 --trust 不改变行为（本身就是可信数据）
    # 注意：校验不通过的记录会被 add_experience 自动标记为 project_import_suspect → 候选层
    source = "project_import"
    layer_hint = "authority（权威层，可直通匹配）"
    logger.info(f"导入经验库... 数据将进入 {layer_hint}")
    if len(all_provinces) > 1:
        logger.info(f"  多定额库模式: {', '.join(p[:20] for p in all_provinces)}")
    exp_stats = import_to_experience(pairs, project_name, all_provinces=all_provinces,
                                     source=source)
    logger.info(
        f"  经验库: 新增{exp_stats['inserted']}条, "
        f"命中已有{exp_stats['matched_existing']}条, "
        f"重复命中{exp_stats['duplicate_hits']}条, "
        f"跳过{exp_stats['skipped']}条"
    )

    # 第3步：导入通用知识库（定额名称模式，跨省份通用）
    logger.info("导入通用知识库...")
    kb_records = convert_to_kb_records(pairs)
    from src.universal_kb import UniversalKB
    kb = UniversalKB()

    kb_stats = kb.batch_import(
        kb_records,
        source_province=province,
        source_project=project_name,
    )
    logger.info(f"  通用知识库: 新增{kb_stats['added']}条, 合并{kb_stats['merged']}条")

    # 打印总结
    logger.info("=" * 50)
    logger.info("导入完成")
    logger.info(f"  项目: {project_name}")
    logger.info(f"  定额库: {', '.join(p[:25] for p in all_provinces)}")
    logger.info(f"  清单项: {len(pairs)}条")
    logger.info(f"  定额项: {total_quotas}条")
    logger.info(f"  主材项: {total_materials}条（{bills_with_materials}条清单含主材）")
    logger.info(
        f"  经验库: 新增{exp_stats['inserted']}条、"
        f"命中已有{exp_stats['matched_existing']}条、"
        f"重复命中{exp_stats['duplicate_hits']}条（带定额编号，同省直接用）"
    )
    logger.info(f"  通用知识库: +{kb_stats['added']}条（定额名称模式，跨省通用）")
    logger.info(f"  数据层级: {layer_hint}")
    logger.info("=" * 50)

    # 第4步（可选）：自动分析生成方法卡片（每个定额库都跑一遍）
    if not args.no_analyze:
        try:
            from tools.gen_method_cards import incremental_generate
            logger.info("自动分析：检查是否有新模式可以提炼方法卡片...")
            total_generated = 0
            for p in all_provinces:
                card_stats = incremental_generate(province=p, min_samples=5)
                total_generated += card_stats["generated"]
            if total_generated > 0:
                logger.info(f"  新生成 {total_generated} 张方法卡片")
            else:
                logger.info("  暂无新模式需要生成方法卡片（样本不足或已有卡片）")
        except Exception as e:
            logger.debug(f"方法卡片自动分析跳过（不影响导入结果）: {e}")


if __name__ == "__main__":
    main()
