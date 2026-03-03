"""
从广联达数据库导出定额子目到Excel

功能：读取广联达GBQ数据库目录，解析子目索引文件，
按"一册一个Excel文件、一章一个Sheet"的结构导出。

用法:
    python tools/export_quota_excel.py                     # 导出所有已安装省份
    python tools/export_quota_excel.py --province 北京     # 只导出北京
    python tools/export_quota_excel.py --list              # 列出所有可导出的省份和定额版本
"""

import os
import sys
import re
import argparse
from collections import defaultdict

# 确保能导入项目模块
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    print("错误: 需要 openpyxl 库。运行 pip install openpyxl 安装。")
    sys.exit(1)


# ===== 广联达数据库路径 =====
# GBQ的定额数据存放在这两个目录下（30.0=通用版, SC30.0=行业版）
GBQ_DB_DIRS = [
    r"D:\广联达\数据库30.0",
    r"D:\广联达\数据库SC30.0",
]

# 导出目录
OUTPUT_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "output", "quota_export")

# 项目已有的定额Excel目录（用来补充单位信息）
PROJECT_QUOTA_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "data", "quota_data")


def parse_index_file(index_path):
    """
    解析广联达 子目索引.Index 文件

    文件格式（逐条记录）:
      4字节: 记录ID (uint32, 小端序)
      1字节: 编码长度
      N字节: 编码 (通常ASCII, 部分省份含中文如江苏"补充-A001")
      1字节: 名称长度
      M字节: 名称 (GBK编码)
      2字节: 未知（可能是节点ID）
      2字节: 分类标识
      1字节: base64长度（固定为12）
      12字节: base64编码的加密数据

    返回: [(编码, 名称), ...]
    """
    with open(index_path, "rb") as f:
        data = f.read()

    records = []
    pos = 0
    while pos < len(data) - 10:
        try:
            # 记录ID (4字节)
            pos += 4
            # 编码长度 (1字节)
            code_len = data[pos]
            pos += 1
            # 编码
            # 用GBK解码（GBK兼容ASCII，且支持江苏等含中文编码的省份）
            code = data[pos:pos + code_len].decode("gbk", errors="replace")
            pos += code_len
            # 名称长度 (1字节)
            name_len = data[pos]
            pos += 1
            # 名称 (GBK)
            name = data[pos:pos + name_len].decode("gbk", errors="replace")
            pos += name_len
            # 跳过尾部17字节 (2+2+1+12)
            pos += 17

            records.append((code, name))
        except Exception:
            break

    return records


def parse_trade_xml(xml_path):
    """
    解析 子目专业.xml，获取册的名称映射

    返回: {"C": "通用安装工程", "A": "房屋建筑与装饰工程", ...}
    或对于福建格式: {100000000: "第一册 机械设备安装工程", ...}
    """
    import xml.etree.ElementTree as ET
    tree = ET.parse(xml_path)
    root = tree.getroot()

    trade_map = {}
    for record in root.findall("Record"):
        desc = record.get("Description", "")
        trade_id = record.get("TradeID", "")
        trade_map[trade_id] = desc

    return trade_map


def guess_code_format(records):
    """
    判断编码格式:
      - "letter": C10-1-1 形式（北京等）
      - "numeric": 30101001 形式（福建等）
    """
    if not records:
        return "unknown"

    sample = records[0][0]
    if re.match(r"^[A-Z]", sample):
        return "letter"
    elif re.match(r"^\d{8}$", sample):
        return "numeric"
    return "unknown"


def parse_code_letter(code):
    """
    解析字母格式编码: C10-1-1 -> (大类="C", 册=10, 章=1, 序号=1)
    """
    m = re.match(r"^([A-Z])(\d+)-(\d+)-(\d+)$", code)
    if not m:
        return None
    return {
        "category": m.group(1),
        "book": int(m.group(2)),
        "chapter": int(m.group(3)),
        "seq": int(m.group(4)),
    }


def parse_code_numeric(code):
    """
    解析数字格式编码: 30101001 -> (前缀=3, 册=01, 章=01, 序号=001)
    编码规则: 1位前缀 + 2位册号 + 2位章号 + 3位序号
    """
    if len(code) != 8:
        return None
    return {
        "category": "C",  # 数字格式默认归安装类
        "book": int(code[1:3]),
        "chapter": int(code[3:5]),
        "seq": int(code[5:8]),
    }


def load_existing_units(province_name, quota_name):
    """
    从项目已有的定额Excel中加载单位映射
    返回: {编码: 单位} 字典
    """
    code_to_unit = {}

    # 在 data/quota_data/ 下查找匹配的省份目录
    province_dir = os.path.join(PROJECT_QUOTA_DIR, province_name)
    if not os.path.isdir(province_dir):
        return code_to_unit

    # 遍历省份下的定额版本目录
    for version_dir in os.listdir(province_dir):
        version_path = os.path.join(province_dir, version_dir)
        if not os.path.isdir(version_path):
            continue

        # 读取所有Excel文件
        for fname in os.listdir(version_path):
            if not fname.endswith(".xlsx"):
                continue
            fpath = os.path.join(version_path, fname)
            try:
                wb = openpyxl.load_workbook(fpath, read_only=True)
                for ws in wb:
                    for row in ws.rows:
                        vals = [cell.value for cell in row]
                        if vals and vals[0]:
                            code = str(vals[0]).strip()
                            unit = str(vals[2]).strip() if len(vals) > 2 and vals[2] else ""
                            code_to_unit[code] = unit
                wb.close()
            except Exception as e:
                print(f"  警告: 读取 {fpath} 失败: {e}")

    return code_to_unit


def numeric_code_to_letter(code, book_prefix="C"):
    """
    数字编码转字母编码: 30101001 -> C1-1-1
    """
    parsed = parse_code_numeric(code)
    if not parsed:
        return code
    return f"{book_prefix}{parsed['book']}-{parsed['chapter']}-{parsed['seq']}"


def get_book_name(book_num, trade_map, code_format):
    """
    获取册的显示名称
    """
    if code_format == "numeric":
        # 福建格式: TradeID = 册号 * 100000000
        trade_id = str(book_num * 100000000)
        if trade_id in trade_map:
            return trade_map[trade_id]

    return f"第{book_num}册"


def get_chapter_name_from_entries(entries):
    """
    从章节的第一条子目名称推断章节名称
    取第一条名称的前几个关键词作为章名
    """
    if not entries:
        return "未知"

    first_name = entries[0][1]  # (code, name, unit) 中的 name
    # 取名称中第一个空格前的部分（通常是设备/工艺类型）
    parts = first_name.split()
    if parts:
        hint = parts[0]
        # 截取合理长度
        if len(hint) > 15:
            hint = hint[:15]
        return hint
    return first_name[:10]


def create_excel_for_book(book_num, book_name, chapters_data, output_path, code_format):
    """
    为一册创建Excel文件

    chapters_data: {章号: [(编码, 名称, 单位), ...]}
    """
    wb = openpyxl.Workbook()
    # 删除默认Sheet
    wb.remove(wb.active)

    # 样式定义
    header_font = Font(name="微软雅黑", size=11, bold=True)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    content_font = Font(name="微软雅黑", size=10)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for chapter_num in sorted(chapters_data.keys()):
        entries = chapters_data[chapter_num]

        # Sheet名称: "01_起重设备安装" 格式
        chapter_hint = get_chapter_name_from_entries(entries)
        # 清理Sheet名中的非法字符（Excel不允许 [ ] : * ? / \）
        chapter_hint = re.sub(r'[\[\]:*?/\\]', '', chapter_hint)
        sheet_name = f"{chapter_num:02d}_{chapter_hint}"
        # Excel Sheet名最长31字符
        if len(sheet_name) > 31:
            sheet_name = sheet_name[:31]

        ws = wb.create_sheet(title=sheet_name)

        # 表头
        headers = ["编码", "名称", "单位"]
        col_widths = [15, 60, 8]

        for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

        # 数据行
        for row_idx, (code, name, unit) in enumerate(entries, 2):
            ws.cell(row=row_idx, column=1, value=code).font = content_font
            ws.cell(row=row_idx, column=2, value=name).font = content_font
            ws.cell(row=row_idx, column=3, value=unit).font = content_font

            for col_idx in range(1, 4):
                ws.cell(row=row_idx, column=col_idx).border = thin_border

            # 编码列居中
            ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="center")
            # 单位列居中
            ws.cell(row=row_idx, column=3).alignment = Alignment(horizontal="center")

        # 冻结首行
        ws.freeze_panes = "A2"

    # 保存
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    return len(chapters_data)


def scan_quota_databases():
    """
    扫描所有广联达数据库目录，返回可导出的定额列表

    返回: [
        {
            "province": "北京",
            "series": "北京2024施工消耗量标准",
            "quota_name": "北京市建设工程施工消耗量标准(2024)",
            "index_path": "...",
            "xml_path": "...",
            "db_source": "30.0" 或 "SC30.0",
        },
        ...
    ]
    """
    results = []

    for db_dir in GBQ_DB_DIRS:
        if not os.path.isdir(db_dir):
            continue

        db_source = os.path.basename(db_dir)

        for province in os.listdir(db_dir):
            province_path = os.path.join(db_dir, province)
            quota_lib = os.path.join(province_path, "定额库")
            if not os.path.isdir(quota_lib):
                continue

            for series in os.listdir(quota_lib):
                series_path = os.path.join(quota_lib, series)
                if not os.path.isdir(series_path):
                    continue

                # 找定额库子目录
                inner_lib = os.path.join(series_path, "定额库")
                if not os.path.isdir(inner_lib):
                    continue

                for quota_name in os.listdir(inner_lib):
                    quota_path = os.path.join(inner_lib, quota_name)
                    index_path = os.path.join(quota_path, "数据", "子目索引.Index")
                    xml_path = os.path.join(quota_path, "基础数据", "子目专业.xml")

                    if os.path.isfile(index_path):
                        results.append({
                            "province": province,
                            "series": series,
                            "quota_name": quota_name,
                            "index_path": index_path,
                            "xml_path": xml_path if os.path.isfile(xml_path) else None,
                            "db_source": db_source,
                        })

    return results


def export_one_quota(quota_info, output_base_dir):
    """
    导出一套定额到Excel

    返回: (成功导出的文件数, 总记录数)
    """
    province = quota_info["province"]
    quota_name = quota_info["quota_name"]
    index_path = quota_info["index_path"]
    xml_path = quota_info["xml_path"]
    db_source = quota_info["db_source"]

    print(f"\n{'='*60}")
    print(f"导出: {province} / {quota_name}")
    print(f"来源: {db_source}")
    print(f"{'='*60}")

    # 1. 解析Index文件
    records = parse_index_file(index_path)
    if not records:
        print(f"  警告: 未解析到任何记录")
        return 0, 0

    code_format = guess_code_format(records)
    print(f"  记录数: {len(records)}, 编码格式: {code_format}")

    # 2. 解析册名称
    trade_map = {}
    if xml_path:
        trade_map = parse_trade_xml(xml_path)

    # 3. 加载已有单位映射
    code_to_unit = load_existing_units(province, quota_name)
    if code_to_unit:
        print(f"  已有单位映射: {len(code_to_unit)} 条")

    # 4. 按大类 -> 册 -> 章 分组
    # category_books[大类][册号][章号] = [(编码, 名称, 单位), ...]
    category_books = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))

    unknown_unit_count = 0
    for code, name in records:
        if code_format == "letter":
            parsed = parse_code_letter(code)
        elif code_format == "numeric":
            parsed = parse_code_numeric(code)
        else:
            continue

        if not parsed:
            continue

        # 查找单位
        unit = ""
        if code in code_to_unit:
            unit = code_to_unit[code]
        elif code_format == "numeric":
            # 数字格式转字母格式再查
            letter_code = numeric_code_to_letter(code)
            unit = code_to_unit.get(letter_code, "")

        if not unit:
            unknown_unit_count += 1

        category = parsed["category"]
        book = parsed["book"]
        chapter = parsed["chapter"]
        category_books[category][book][chapter].append((code, name, unit))

    if unknown_unit_count > 0:
        print(f"  注意: {unknown_unit_count} 条缺少单位信息")

    # 5. 为每个大类的每册生成Excel
    # 输出目录: output/quota_export/{省份}_{定额名}/
    safe_name = quota_name.replace("(", "（").replace(")", "）")
    output_dir = os.path.join(output_base_dir, f"{province}_{safe_name}")

    total_files = 0
    total_records = 0

    for category in sorted(category_books.keys()):
        books = category_books[category]
        for book_num in sorted(books.keys()):
            chapters = books[book_num]

            # 获取册名称
            book_name = get_book_name(book_num, trade_map, code_format)

            # 文件名: "C01_机械设备安装工程.xlsx"
            file_name = f"{category}{book_num:02d}_{book_name}.xlsx"
            # 清理文件名中的非法字符
            file_name = re.sub(r'[<>:"/\\|?*]', '_', file_name)
            output_path = os.path.join(output_dir, file_name)

            # 统计
            entry_count = sum(len(entries) for entries in chapters.values())

            # 生成Excel
            sheet_count = create_excel_for_book(
                book_num, book_name, chapters, output_path, code_format
            )

            print(f"  {file_name} ({sheet_count}个Sheet, {entry_count}条)")
            total_files += 1
            total_records += entry_count

    print(f"\n  导出完成: {total_files}个文件, {total_records}条记录")
    print(f"  输出目录: {output_dir}")

    return total_files, total_records


def main():
    parser = argparse.ArgumentParser(description="从广联达数据库导出定额子目到Excel")
    parser.add_argument("--list", action="store_true", help="列出所有可导出的省份和定额版本")
    parser.add_argument("--province", type=str, help="只导出指定省份（如: 北京）")
    parser.add_argument("--quota", type=str, help="只导出名称包含此关键词的定额")
    parser.add_argument("--output", type=str, default=OUTPUT_DIR, help="输出目录")
    args = parser.parse_args()

    # 扫描可用的定额库
    print("扫描广联达数据库目录...")
    quotas = scan_quota_databases()

    if not quotas:
        print("错误: 未找到任何广联达定额数据库")
        print(f"检查路径: {GBQ_DB_DIRS}")
        sys.exit(1)

    print(f"找到 {len(quotas)} 套定额\n")

    # 列出模式
    if args.list:
        current_province = None
        for q in sorted(quotas, key=lambda x: (x["province"], x["quota_name"])):
            if q["province"] != current_province:
                current_province = q["province"]
                print(f"\n{current_province} ({q['db_source']}):")
            print(f"  - {q['quota_name']}")
        return

    # 过滤
    if args.province:
        quotas = [q for q in quotas if args.province in q["province"]]
        if not quotas:
            print(f"未找到省份 '{args.province}' 的定额")
            sys.exit(1)

    if args.quota:
        quotas = [q for q in quotas if args.quota in q["quota_name"]]
        if not quotas:
            print(f"未找到包含 '{args.quota}' 的定额")
            sys.exit(1)

    # 导出
    grand_total_files = 0
    grand_total_records = 0

    for q in sorted(quotas, key=lambda x: (x["province"], x["quota_name"])):
        files, records = export_one_quota(q, args.output)
        grand_total_files += files
        grand_total_records += records

    print(f"\n{'='*60}")
    print(f"全部完成: {grand_total_files}个Excel文件, {grand_total_records}条记录")
    print(f"输出目录: {args.output}")


if __name__ == "__main__":
    main()
