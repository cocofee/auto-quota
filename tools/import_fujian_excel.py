#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Import Fujian 2024 material prices from the bundled Excel file."""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from src.material_db import MaterialDB
from tools.import_price_excel import create_import_batch, update_batch_count, normalize_spec, normalize_unit
from tools.pdf_profiles.base_profile import guess_category

PROVINCE = "\u798f\u5efa"
SOURCE_TYPE = "official_info"
PERIOD_START = "2024-01-01"
PERIOD_END = "2024-12-31"
DEFAULT_SOURCE = PROJECT_ROOT / "data" / "pdf_info_price" / "fujian" / "fujian_2024_materials.xlsx"


def parse_workbook(filepath: Path) -> list[dict]:
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows = ws.iter_rows(min_row=4, values_only=True)
        records: list[dict] = []
        for row in rows:
            seq, material_code, name, spec, unit, price_excl, price_incl, note = row[:8]
            if not name:
                continue
            price = float(price_incl) if price_incl not in (None, "") else 0.0
            records.append({
                "material_code": str(material_code or "").strip(),
                "name": str(name).strip(),
                "spec": normalize_spec(str(spec or "")),
                "unit": normalize_unit(str(unit or "")),
                "price_incl_tax": price,
                "price_excl_tax": float(price_excl) if price_excl not in (None, "") else None,
                "note": str(note or "").strip(),
                "category": guess_category(str(name).strip()),
            })
        return records
    finally:
        wb.close()


def import_records(filepath: Path, dry_run: bool = False) -> dict:
    records = parse_workbook(filepath)
    stats = {"imported": 0, "skipped": 0, "errors": 0}

    if dry_run:
        print(f"Parsed {len(records)} record(s) from {filepath.name}")
        for rec in records[:10]:
            print(rec)
        return stats

    db = MaterialDB()
    batch_id = create_import_batch(
        db,
        source_file=str(filepath),
        source_type=SOURCE_TYPE,
        parser_template="fujian_excel",
        notes="Fujian 2024 Excel import",
        province=PROVINCE,
    )

    for rec in records:
        try:
            name = rec["name"]
            if not name:
                stats["skipped"] += 1
                continue
            material_id = db.add_material(
                name=name,
                spec=rec["spec"],
                unit=rec["unit"],
                category=rec["category"],
            )
            if rec["price_incl_tax"] > 0:
                db.add_price(
                    material_id=material_id,
                    price_incl_tax=rec["price_incl_tax"],
                    source_type=SOURCE_TYPE,
                    province=PROVINCE,
                    city="",
                    tax_rate=0.13,
                    period_start=PERIOD_START,
                    period_end=PERIOD_END,
                    source_doc=filepath.name,
                    batch_id=batch_id,
                    authority_level="official",
                    usable_for_quote=0,
                    unit=rec["unit"],
                    dedup=True,
                )
            stats["imported"] += 1
        except Exception as exc:
            print(f"import failed: {rec.get('name', '?')} - {exc}")
            stats["errors"] += 1

    update_batch_count(db, batch_id, stats["imported"])
    print(f"batch #{batch_id}: {stats}")
    return stats


def main() -> int:
    parser = argparse.ArgumentParser(description="Import Fujian 2024 material prices from local Excel")
    parser.add_argument("--file", default=str(DEFAULT_SOURCE), help="Excel file path")
    parser.add_argument("--dry-run", action="store_true", help="Preview only")
    args = parser.parse_args()

    filepath = Path(args.file)
    if not filepath.exists():
        print(f"file not found: {filepath}")
        return 1

    import_records(filepath, dry_run=args.dry_run)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
