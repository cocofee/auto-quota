# -*- coding: utf-8 -*-
"""
批量导出缺失省份的安装定额到 data/quota_data/
支持多种编码格式，自动检测并分册导出

输出格式：和四川2020一致（无表头，2列：编码|名称）
"""
import os
import sys
import re
import xml.etree.ElementTree as ET
from collections import defaultdict

sys.stdout.reconfigure(encoding='utf-8')
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

try:
    import openpyxl
except ImportError:
    print("错误: 需要 openpyxl，运行 pip install openpyxl")
    sys.exit(1)

from tools.export_quota_excel import parse_index_file

# 项目根目录
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(PROJECT_ROOT, "data", "quota_data")


# ===== 要导出的定额配置 =====
# 每项: (省份, 定额版本目录名, 广联达数据库基路径, 序列定额目录名)
TARGETS = [
    {
        "province": "上海",
        "version_name": "上海市安装工程预算定额(2016)",
        "db_base": r"D:\广联达\数据库30.0\上海\定额库\上海2016预算定额\定额库",
        "quota_names": ["上海市安装工程预算定额(2016)"],
    },
    {
        "province": "深圳",
        "version_name": "深圳市安装工程消耗量标准(2025)",
        "db_base": r"D:\广联达\数据库30.0\深圳\定额库\深圳2024序列定额\定额库",
        "quota_names": ["深圳市安装工程消耗量标准(2025)"],
    },
    {
        "province": "福建",
        "version_name": "福建省通用安装工程预算定额(2017)",
        "db_base": r"D:\广联达\数据库30.0\福建\定额库\福建2017序列定额\定额库",
        "quota_names": ["福建省通用安装工程预算定额(2017)"],
    },
    {
        "province": "云南",
        "version_name": "云南省通用安装工程计价标准(2020)",
        "db_base": r"D:\广联达\数据库30.0\云南\定额库\云南省2020序列定额\定额库",
        "quota_names": ["云南省通用安装工程计价标准(2020)"],
    },
    {
        "province": "内蒙古",
        "version_name": "内蒙古通用安装工程预算定额(2017)",
        "db_base": r"D:\广联达\数据库30.0\内蒙古\定额库\内蒙古2017序列定额\定额库",
        "quota_names": ["内蒙古通用安装工程预算定额(2017)"],
    },
    {
        "province": "甘肃",
        "version_name": "甘肃省安装工程预算定额(2013)",
        "db_base": r"D:\广联达\数据库30.0\甘肃\定额库\甘肃2013序列预算定额\定额库",
        "quota_names": ["甘肃省安装工程预算定额(2013)"],
    },
]


def get_book_names_from_xml(xml_path):
    """从子目专业.xml获取册名映射
    返回: {序号(从0开始): 册名}
    """
    if not os.path.isfile(xml_path):
        return {}
    try:
        tree = ET.parse(xml_path)
        root = tree.getroot()
        result = {}
        for i, record in enumerate(root.findall('Record')):
            desc = record.get('Description', '')
            # 有的带字母前缀如 "A 机械设备安装工程"，去掉
            clean = re.sub(r'^[A-Z]\s+', '', desc)
            result[i] = clean
        return result
    except Exception as e:
        print(f"  警告: XML解析失败 {xml_path}: {e}")
        return {}


def detect_book_number_auto(code):
    """从编码自动提取册号（逐条判断，不依赖全局格式检测）

    各格式的册号提取规则:
    - C10-1-1 → 册10（北京字母格式）
    - C15000100 → 册15（深圳字母+数字）
    - 030101-1 → 册01（深圳6位数字+连字符，第3-4位是册）
    - 30101001 → 册01（福建8位纯数字，第2-3位是册）
    - 03-13-1-1 → 册13（上海4级连字符，第2段是册）
    - 2-1-1 → 需要上下文判断（云南：第2段是册；通用：第1段是册）
    - 1-1 → 册1（内蒙古/甘肃2级连字符，第1段是册）
    """
    # 字母开头 + 数字-数字-数字 (北京: C10-1-1)
    m = re.match(r'^[A-Z](\d+)-(\d+)-(\d+)$', code)
    if m:
        return int(m.group(1))

    # 字母+7-8位数字 (深圳: C15000100)
    m = re.match(r'^[A-Z](\d{2})\d{4,6}$', code)
    if m:
        return int(m.group(1))

    # 6位数字+连字符 (深圳: 030101-1 → 册=第3-4位)
    m = re.match(r'^\d{2}(\d{2})\d{2}-\d+$', code)
    if m:
        return int(m.group(1))

    # 8位纯数字 (福建: 30101001 → 册=第2-3位)
    if re.match(r'^\d{8}$', code):
        return int(code[1:3])

    # 4级连字符 (上海: 03-13-1-1 → 册=第2段)
    m = re.match(r'^(\d+)-(\d+)-(\d+)-(\d+)$', code)
    if m:
        return int(m.group(2))

    # 2级连字符 (内蒙古/甘肃: 1-1, 10-2045 → 册=第1段)
    m = re.match(r'^(\d+)-(\d+)$', code)
    if m:
        return int(m.group(1))

    # 3级连字符 — 默认第1段是册（云南需要特殊处理，见下面的wrapper）
    m = re.match(r'^(\d+)-(\d+)-(\d+)$', code)
    if m:
        return int(m.group(1))

    # 兜底
    return 0


def detect_3level_prefix(records):
    """检测3级连字符编码是否有固定前缀（如云南的 2-X-Y）

    如果所有3级连字符编码的第1段都相同，说明第1段是大类前缀而非册号，
    真正的册号在第2段。

    返回: True（有固定前缀）/ False（没有）
    """
    three_level_first = set()
    count = 0
    for code, _ in records[:200]:  # 采样200条
        m = re.match(r'^(\d+)-(\d+)-(\d+)$', code)
        if m:
            three_level_first.add(m.group(1))
            count += 1
    # 如果采样到足够多的3级编码且第1段全部相同 → 有固定前缀
    return count >= 10 and len(three_level_first) == 1


def make_book_detector(records):
    """根据数据特征返回一个册号提取函数

    对于云南等有固定前缀的3级连字符格式，返回取第2段的提取函数；
    其他情况直接用 detect_book_number_auto。
    """
    has_prefix = detect_3level_prefix(records)

    if has_prefix:
        def detector(code):
            m = re.match(r'^(\d+)-(\d+)-(\d+)$', code)
            if m:
                return int(m.group(2))  # 第2段是册号
            return detect_book_number_auto(code)
        return detector, "3level_prefix"
    else:
        return detect_book_number_auto, "auto"


def clean_cell_value(text):
    """清理单元格内容中Excel不允许的控制字符"""
    if not text:
        return text
    return re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', text)


def clean_sheet_name(name):
    """清理Sheet名中Excel不允许的字符"""
    bad_chars = ['[', ']', ':', '*', '?', '/', '\\']
    for c in bad_chars:
        name = name.replace(c, '')
    return name[:31]


def export_single_quota(target):
    """导出一个省份的安装定额"""
    province = target["province"]
    version_name = target["version_name"]
    db_base = target["db_base"]
    quota_names = target["quota_names"]

    output_dir = os.path.join(DATA_DIR, province, version_name)
    os.makedirs(output_dir, exist_ok=True)

    print(f"\n{'='*60}")
    print(f"导出: {province} / {version_name}")
    print(f"{'='*60}")

    total_records = 0

    for quota_name in quota_names:
        quota_path = os.path.join(db_base, quota_name)
        idx_path = os.path.join(quota_path, "数据", "子目索引.Index")
        xml_path = os.path.join(quota_path, "基础数据", "子目专业.xml")

        if not os.path.isfile(idx_path):
            print(f"  错误: Index文件不存在 - {idx_path}")
            continue

        # 1. 解析Index
        records = parse_index_file(idx_path)
        if not records:
            print(f"  警告: 无记录 - {quota_name}")
            continue
        print(f"  解析 {quota_name}: {len(records)}条")

        # 2. 获取册名
        book_names = get_book_names_from_xml(xml_path)
        if book_names:
            print(f"  册名映射: {len(book_names)}册")
            for idx, name in sorted(book_names.items()):
                print(f"    [{idx}] {name}")

        # 3. 检测编码格式并按册号分组
        get_book, fmt_name = make_book_detector(records)
        print(f"  编码格式: {fmt_name}")

        books = defaultdict(list)
        for code, name in records:
            book_num = get_book(code)
            books[book_num].append((code, name))

        # 打印分组统计
        print(f"  分为{len(books)}册:")
        for bn in sorted(books.keys()):
            sample = books[bn][0][0]
            print(f"    册{bn}: {len(books[bn])}条 (首条: {sample})")

        # 4. 生成Excel — 每册一个文件
        for book_num in sorted(books.keys()):
            entries = books[book_num]

            # 册名：优先从XML取，册号和XML索引的对应关系
            # 大多数省份册号从1开始，XML索引从0开始
            bn_display = ""
            if book_num > 0 and (book_num - 1) in book_names:
                bn_display = book_names[book_num - 1]
            elif book_num in book_names:
                bn_display = book_names[book_num]

            # 文件名
            if bn_display:
                file_name = f"第{book_num:02d}册_{bn_display}.xlsx"
            else:
                file_name = f"第{book_num:02d}册.xlsx"

            # 清理文件名
            file_name = re.sub(r'[<>:"/\\|?*]', '_', file_name)
            output_path = os.path.join(output_dir, file_name)

            # 写Excel（无表头，2列：编码|名称）
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = clean_sheet_name(f"{book_num:02d}_{bn_display}" if bn_display else f"册{book_num:02d}")

            for row_idx, (code, name) in enumerate(entries, 1):
                ws.cell(row=row_idx, column=1, value=clean_cell_value(code))
                ws.cell(row=row_idx, column=2, value=clean_cell_value(name))

            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 60

            wb.save(output_path)
            total_records += len(entries)

        print(f"  输出: {output_dir}")

    print(f"  合计: {total_records}条")
    return total_records


def main():
    print("=" * 60)
    print("批量导出缺失省份的安装定额")
    print("=" * 60)

    grand_total = 0
    results = []

    for target in TARGETS:
        # 检查数据库路径是否存在
        if not os.path.isdir(target["db_base"]):
            print(f"\n跳过 {target['province']}: 数据库路径不存在 - {target['db_base']}")
            results.append((target["province"], 0, "路径不存在"))
            continue

        count = export_single_quota(target)
        grand_total += count
        results.append((target["province"], count, "成功"))

    print(f"\n{'='*60}")
    print(f"导出汇总")
    print(f"{'='*60}")
    for province, count, status in results:
        print(f"  {province}: {count}条 ({status})")
    print(f"  总计: {grand_total}条")


if __name__ == "__main__":
    main()
