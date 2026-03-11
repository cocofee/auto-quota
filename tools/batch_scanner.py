"""
批量文件扫描分类工具

功能：扫描 F:\\jarvis 目录，对每个Excel文件做格式深度分类，结果存入SQLite。

设计要点（按 Codex 5.3 审核意见 v2 修复）：
1. SQLite 替代 JSON（事务原子性、行级更新、索引查询）
2. UPSERT（INSERT...ON CONFLICT）替代 INSERT OR REPLACE（保留created_at）
3. 增量判定综合 mtime + algo_version + md5（不只看mtime）
4. matched状态：文件内容变了（mtime/md5变了）→ 降级为scanned重跑
5. status字段加CHECK约束、error_msg/skip_reason分开使用
6. 省份提取支持[省份]和【省份】两种括号
7. 文件元信息获取包在try内（防竞态删除文件导致崩溃）
8. Sheet数量限制防解压炸弹（最多读50个Sheet）

用法：
    python tools/batch_scanner.py "F:/jarvis"                    # 扫描全部（增量）
    python tools/batch_scanner.py "F:/jarvis" --specialty 电气    # 只扫某专业
    python tools/batch_scanner.py "F:/jarvis" --rescan            # 重新分类
    python tools/batch_scanner.py "F:/jarvis" --stats             # 只看统计
"""

import os
import sys
import re
import json
import argparse
import hashlib
from pathlib import Path
from datetime import datetime

# 把项目根目录加入搜索路径
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from db.sqlite import connect_init, connect

# ============================================================
# 常量配置
# ============================================================

# 数据库路径
DB_PATH = Path(__file__).resolve().parent.parent / "output" / "batch" / "batch.db"

# 数据库 schema 版本（升级字段时自增）
SCHEMA_VERSION = 2

# 算法版本：自动根据关键文件的内容指纹计算
# 任何模型文件或核心匹配代码发生变化，版本号自动改变，批量匹配会自动检测到需要重跑
# 不再需要手动修改版本号
_PROJECT_ROOT = Path(__file__).resolve().parent.parent

# 参与版本指纹计算的关键文件（模型+核心匹配代码）
_VERSION_FILES = [
    # 模型文件
    "data/ltr_model.txt",                   # LTR排序模型
    # 核心匹配代码（改了任何一个都可能影响匹配结果）
    "src/query_builder.py",                 # 搜索词构建
    "src/rule_validator.py",                # 规则匹配（参数提取+档位验证）
    "src/param_validator.py",               # 参数验证
    "src/bill_reader.py",                   # 清单读取（影响哪些行被识别为清单项）
    "src/bill_cleaner.py",                  # 清单清洗（名称修正+参数提取）
    "src/match_pipeline.py",               # 匹配流水线
]


def _compute_algorithm_version() -> str:
    """根据关键文件的内容计算算法版本指纹。

    把所有关键文件的内容拼在一起算MD5，取前8位作为版本号。
    任何文件有变化（哪怕改了一行代码），版本号就会变，
    批量扫描时会自动发现版本不匹配，标记文件需要重跑。

    如果某个文件不存在（比如模型还没训好），跳过不影响其他文件。
    """
    h = hashlib.md5()
    for rel_path in _VERSION_FILES:
        full_path = _PROJECT_ROOT / rel_path
        if not full_path.exists():
            continue
        try:
            # 代码文件读文本（忽略行尾空白差异），模型文件读二进制前1MB
            if rel_path.endswith(".py"):
                content = full_path.read_text(encoding="utf-8")
                h.update(content.encode("utf-8"))
            else:
                with open(full_path, "rb") as f:
                    h.update(f.read(1024 * 1024))  # 模型文件只读前1MB（够区分版本）
        except Exception:
            continue

    # 检查向量模型目录是否存在（目录名本身就包含版本信息）
    # 用 config.json 的内容作为向量模型的指纹
    vector_config = _PROJECT_ROOT / "models" / "qwen3-embedding-quota-v3" / "config.json"
    if vector_config.exists():
        try:
            h.update(vector_config.read_bytes())
        except Exception:
            pass

    fingerprint = h.hexdigest()[:8]
    return f"auto-{fingerprint}"


# 自动计算的算法版本（启动时算一次，整个进程内不变）
ALGORITHM_VERSION = _compute_algorithm_version()

# 支持的Excel后缀
EXCEL_EXTENSIONS = {".xlsx", ".xls", ".xlsm"}

# 非清单文件的排除关键词
NOT_BILL_KEYWORDS = [
    "合同", "台账", "工资", "考勤", "签证", "变更令", "会议纪要",
    "进度", "月报", "周报", "日报", "通知", "函件", "招标文件",
    "施组", "施工组织", "方案", "交底", "记录", "验收", "检测",
    "资料", "档案", "图纸", "设计说明", "勘察", "地勘",
]

# 格式分类规则：required_any 中任意一组全部命中即判定为该格式
FORMAT_RULES = [
    {
        "format": "standard_bill",  # 标准工程量清单
        "required_any": [
            ["项目名称", "项目特征"],
            ["项目名称", "工程量"],
            ["清单名称", "项目特征"],
            ["名称", "项目特征描述"],
            ["项目编码", "项目名称"],
        ],
    },
    {
        "format": "work_list",  # 工作量清单
        "required_any": [
            ["工作量名称", "工作量"],
            ["工作项目", "工作量"],
        ],
    },
    {
        "format": "equipment_list",  # 设备材料清单
        "required_any": [
            ["设备名称", "规格型号"],
            ["设备名称", "规格"],
            ["材料名称", "规格型号"],
            ["材料名称", "规格"],
            ["品名", "规格"],
        ],
    },
]

# 省份名称列表
PROVINCE_NAMES = [
    "北京", "天津", "河北", "山西", "内蒙古",
    "辽宁", "吉林", "黑龙江",
    "上海", "江苏", "浙江", "安徽", "福建", "江西", "山东",
    "河南", "湖北", "湖南", "广东", "广西", "海南",
    "重庆", "四川", "贵州", "云南", "西藏",
    "陕西", "甘肃", "青海", "宁夏", "新疆",
]

# 安全限制
MAX_FILE_SIZE = 50 * 1024 * 1024    # 50MB，超过则跳过
MAX_SHEETS_PER_FILE = 50            # 单文件最多读50个Sheet（防解压炸弹）


# ============================================================
# 数据库初始化
# ============================================================

def init_db():
    """初始化SQLite数据库，建表+索引。

    v2改进（按Codex 5.3审核）：
    - status 加 CHECK 约束，只允许5种合法状态
    - file_name 加 NOT NULL
    - 用 UPSERT 替代 INSERT OR REPLACE（见写入函数）
    """
    DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    conn = connect_init(DB_PATH)
    try:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS schema_info (
                key   TEXT PRIMARY KEY,
                value TEXT
            )
        """)
        conn.execute(
            "INSERT OR REPLACE INTO schema_info (key, value) VALUES (?, ?)",
            ("version", str(SCHEMA_VERSION))
        )

        # 核心表：加了CHECK约束和NOT NULL
        conn.execute("""
            CREATE TABLE IF NOT EXISTS file_registry (
                file_path       TEXT PRIMARY KEY,
                file_name       TEXT NOT NULL,
                file_size       INTEGER,
                file_mtime      TEXT,
                file_md5        TEXT,
                province        TEXT,
                specialty       TEXT,
                format          TEXT,
                status          TEXT NOT NULL DEFAULT 'pending'
                                CHECK(status IN ('pending','scanned','matched','skipped','error')),
                skip_reason     TEXT,
                error_msg       TEXT,
                sheet_info      TEXT,
                estimated_items INTEGER DEFAULT 0,
                algo_version    TEXT,
                scan_time       TEXT,
                match_time      TEXT,
                created_at      TEXT DEFAULT (datetime('now', 'localtime')),
                updated_at      TEXT DEFAULT (datetime('now', 'localtime'))
            )
        """)

        conn.execute("CREATE INDEX IF NOT EXISTS idx_status ON file_registry(status)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_province ON file_registry(province)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_specialty ON file_registry(specialty)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_format ON file_registry(format)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_md5 ON file_registry(file_md5)")

        conn.commit()
    finally:
        conn.close()


def get_db():
    """获取数据库连接"""
    return connect(DB_PATH, row_factory=True)


# ============================================================
# UPSERT 写入（替代 INSERT OR REPLACE，保留 created_at）
# ============================================================

def _upsert_file(conn, file_path: str, file_name: str, file_size: int,
                 file_mtime: str, file_md5: str, province: str,
                 specialty: str, fmt: str, status: str,
                 skip_reason: str, error_msg: str,
                 sheet_info: str, estimated_items: int,
                 algo_version: str, scan_time: str):
    """用 INSERT...ON CONFLICT DO UPDATE 写入，保留 created_at 不被覆盖。

    Codex P1 修复：INSERT OR REPLACE 会删旧行再插新行，导致 created_at/match_time 丢失。
    UPSERT 只更新指定字段，created_at 保留原值。
    """
    conn.execute("""
        INSERT INTO file_registry
            (file_path, file_name, file_size, file_mtime, file_md5,
             province, specialty, format, status, skip_reason, error_msg,
             sheet_info, estimated_items, algo_version, scan_time, updated_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, datetime('now','localtime'))
        ON CONFLICT(file_path) DO UPDATE SET
            file_name = excluded.file_name,
            file_size = excluded.file_size,
            file_mtime = excluded.file_mtime,
            file_md5 = excluded.file_md5,
            province = excluded.province,
            specialty = excluded.specialty,
            format = excluded.format,
            status = excluded.status,
            skip_reason = excluded.skip_reason,
            error_msg = excluded.error_msg,
            sheet_info = excluded.sheet_info,
            estimated_items = excluded.estimated_items,
            algo_version = excluded.algo_version,
            scan_time = excluded.scan_time,
            match_time = CASE WHEN excluded.status != 'matched' THEN NULL ELSE file_registry.match_time END,
            updated_at = datetime('now', 'localtime')
    """, (
        file_path, file_name, file_size, file_mtime, file_md5,
        province, specialty, fmt, status, skip_reason, error_msg,
        sheet_info, estimated_items, algo_version, scan_time,
    ))


# ============================================================
# 省份提取
# ============================================================

def extract_province(file_path: str) -> str:
    """从文件路径/文件名中提取省份。

    Codex P2 修复：同时支持 [省份] 和 【省份】 两种括号。
    """
    name = Path(file_path).stem

    # 第1步：[省份] 或 【省份】 标签
    for match in re.finditer(r'[\[【]([^\]】]+)[\]】]', name):
        tag = match.group(1)
        for p in PROVINCE_NAMES:
            if tag.startswith(p):
                return p

    # 第2步：文件名直接包含省份名
    for p in PROVINCE_NAMES:
        if p in name:
            return p

    return None


def extract_specialty(file_path: str) -> str:
    """从文件路径中提取专业（扫描根目录下的一级子目录名）。

    支持的根目录标识：jarvis（本地）、raw_files（容器内）
    例如：
      /app/raw_files/电气/xxx.xlsx → "电气"
      F:/jarvis/消防/xxx.xlsx → "消防"
    """
    _SCAN_ROOT_MARKERS = {"jarvis", "raw_files"}
    parts = Path(file_path).parts
    for i, part in enumerate(parts):
        if part.lower() in _SCAN_ROOT_MARKERS and i + 1 < len(parts):
            candidate = parts[i + 1]
            if not candidate.endswith(('.xlsx', '.xls', '.xlsm')):
                return candidate
    return "未分类"


# ============================================================
# Excel格式深度分类
# ============================================================

def classify_excel_format(file_path: str) -> dict:
    """打开Excel读取表头，判断文件格式类型。

    返回:
        {
            "format": "standard_bill" / "work_list" / "equipment_list" / "not_bill" / "unknown",
            "sheets": [...],
            "estimated_items": 45,
            "skip_reason": None,    # 跳过原因
            "error_msg": None,      # 错误信息（Codex P2: 和skip_reason分开）
        }
    """
    import openpyxl

    fp = Path(file_path)

    # 安全检查：文件大小
    try:
        file_size = fp.stat().st_size
    except OSError as e:
        return _error_result(f"无法获取文件信息: {e}")

    if file_size > MAX_FILE_SIZE:
        return _skip_result(f"文件过大（{file_size // 1024 // 1024}MB > 50MB限制）")

    # 文件名快速排除
    fname = fp.stem
    for kw in NOT_BILL_KEYWORDS:
        if kw in fname:
            return _skip_result(f"文件名含非清单关键词「{kw}」")

    # .xls 需要特殊处理
    actual_path = file_path
    temp_path = None
    # 用统一的magic bytes检测，兼容后缀是.xlsx但实际是.xls的文件
    from src.bill_reader import is_xls_format
    if is_xls_format(file_path):
        try:
            temp_path = _convert_xls(file_path)
            actual_path = temp_path
        except Exception as e:
            return _error_result(f".xls转换失败: {e}")

    try:
        wb = openpyxl.load_workbook(str(actual_path), read_only=True, data_only=True)
    except Exception as e:
        return _error_result(f"无法打开文件: {e}")

    sheets_info = []
    total_items = 0
    best_format = None

    try:
        # Codex P1: 限制Sheet数量，防解压炸弹
        sheet_names = wb.sheetnames[:MAX_SHEETS_PER_FILE]

        for sn in sheet_names:
            ws = wb[sn]
            header_texts = []
            data_rows = 0

            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= 20:
                    break
                for cell in row:
                    if cell is not None:
                        header_texts.append(str(cell).strip())

            try:
                data_rows = ws.max_row or 0
            except Exception:
                data_rows = 0

            sheet_format = _match_format(header_texts)
            has_bill = sheet_format is not None
            est = max(0, data_rows - 5) if has_bill else 0

            sheets_info.append({
                "name": sn,
                "rows": data_rows,
                "has_bill_data": has_bill,
                "format": sheet_format or "unknown",
            })

            if has_bill:
                total_items += est
                if best_format is None:
                    best_format = sheet_format
                elif sheet_format == "standard_bill":
                    best_format = "standard_bill"
    finally:
        wb.close()
        if temp_path:
            try:
                Path(temp_path).unlink(missing_ok=True)
            except Exception:
                pass

    if best_format is None:
        return {
            "format": "not_bill",
            "sheets": sheets_info,
            "estimated_items": 0,
            "skip_reason": "所有Sheet均未检测到清单表头",
            "error_msg": None,
        }

    return {
        "format": best_format,
        "sheets": sheets_info,
        "estimated_items": total_items,
        "skip_reason": None,
        "error_msg": None,
    }


def _skip_result(reason: str) -> dict:
    """生成跳过结果（skip_reason有值，error_msg为空）"""
    return {"format": "not_bill", "sheets": [], "estimated_items": 0,
            "skip_reason": reason, "error_msg": None}


def _error_result(msg: str) -> dict:
    """生成错误结果（error_msg有值，skip_reason为空）"""
    return {"format": "unknown", "sheets": [], "estimated_items": 0,
            "skip_reason": None, "error_msg": msg}


def _match_format(header_texts: list) -> str:
    """根据表头文本匹配格式规则。"""
    # 去掉换行/空格的版本，处理"项目\n名称"这种列名
    cleaned = [t.replace("\n", "").replace("\r", "").replace(" ", "")
               for t in header_texts]

    for rule in FORMAT_RULES:
        for group in rule["required_any"]:
            all_matched = True
            for kw in group:
                found = any(kw in t for t in header_texts)
                if not found:
                    kw_clean = kw.replace("\n", "").replace(" ", "")
                    found = any(kw_clean in t for t in cleaned)
                if not found:
                    all_matched = False
                    break
            if all_matched:
                return rule["format"]

    return None


def _convert_xls(file_path: str) -> str:
    """将 .xls 转为临时 .xlsx"""
    import xlrd
    import openpyxl
    import tempfile

    xls_wb = xlrd.open_workbook(str(file_path))
    temp_dir = Path(__file__).resolve().parent.parent / "output" / "temp"
    temp_dir.mkdir(parents=True, exist_ok=True)

    fd, temp_path = tempfile.mkstemp(suffix=".xlsx", prefix="scan_xls_", dir=str(temp_dir))
    os.close(fd)

    try:
        xlsx_wb = openpyxl.Workbook()
        xlsx_wb.remove(xlsx_wb.active)

        for sheet_idx in range(min(xls_wb.nsheets, MAX_SHEETS_PER_FILE)):
            xls_sheet = xls_wb.sheet_by_index(sheet_idx)
            xlsx_sheet = xlsx_wb.create_sheet(title=xls_sheet.name)
            max_rows = min(xls_sheet.nrows, 30)
            for row_idx in range(max_rows):
                for col_idx in range(xls_sheet.ncols):
                    cell = xls_sheet.cell(row_idx, col_idx)
                    value = cell.value
                    if cell.ctype == 3:
                        try:
                            value = xlrd.xldate_as_datetime(value, xls_wb.datemode)
                        except Exception:
                            pass
                    if value is not None and value != "":
                        xlsx_sheet.cell(row=row_idx + 1, column=col_idx + 1, value=value)
            xlsx_sheet._max_row = xls_sheet.nrows

        xlsx_wb.save(temp_path)
    except Exception:
        Path(temp_path).unlink(missing_ok=True)
        raise
    finally:
        xls_wb.release_resources()

    return temp_path


# ============================================================
# 文件MD5计算
# ============================================================

def compute_md5(file_path: str) -> str:
    """计算文件MD5（读前1MB+后1MB，兼顾速度和检测率）"""
    h = hashlib.md5()
    try:
        file_size = os.path.getsize(file_path)
        with open(file_path, "rb") as f:
            # 读前1MB
            h.update(f.read(1024 * 1024))
            # 如果文件大于2MB，还读后1MB
            if file_size > 2 * 1024 * 1024:
                f.seek(-1024 * 1024, 2)  # 从文件尾部往前1MB
                h.update(f.read(1024 * 1024))
        return h.hexdigest()
    except Exception:
        return None


# ============================================================
# 核心扫描逻辑
# ============================================================

def scan_directory(root_dir: str, specialty_filter: str = None,
                   rescan: bool = False):
    """扫描目录，对每个文件做格式分类并写入数据库。

    Codex P1 修复：
    - matched + 文件内容变了（mtime或md5变了）→ 降级重跑
    - 增量判定加入 algo_version 比对
    - 文件元信息获取包在try内
    """
    init_db()
    root = Path(root_dir)
    if not root.exists():
        print(f"错误: 目录不存在 {root_dir}")
        return

    # 第1步：一次遍历收集所有文件（Codex P2: 不重复遍历）
    print(f"扫描目录: {root_dir}")
    all_files = []
    non_excel_count = 0

    for dirpath, dirnames, filenames in os.walk(root):
        dirnames[:] = [d for d in dirnames if not d.startswith('.') and d != '__pycache__']
        for fname in filenames:
            ext = Path(fname).suffix.lower()
            if ext not in EXCEL_EXTENSIONS:
                non_excel_count += 1
                continue
            full_path = os.path.join(dirpath, fname)
            if specialty_filter:
                sp = extract_specialty(full_path)
                if sp != specialty_filter:
                    continue
            all_files.append(full_path)

    print(f"  找到 {len(all_files)} 个Excel文件 + {non_excel_count} 个非Excel文件")

    # 第2步：查数据库，确定哪些文件需要处理
    conn = get_db()
    try:
        existing = {}  # file_path → {status, file_mtime, file_md5, algo_version}
        for row in conn.execute(
            "SELECT file_path, status, file_mtime, file_md5, algo_version FROM file_registry"
        ):
            existing[row["file_path"]] = {
                "status": row["status"],
                "mtime": row["file_mtime"],
                "md5": row["file_md5"],
                "algo_version": row["algo_version"],
            }
    finally:
        conn.close()

    to_process = []
    skipped_count = 0

    for fp in all_files:
        fp_normalized = fp.replace("\\", "/")

        # Codex P1: 文件元信息获取包在try内（防竞态删除）
        try:
            mtime = datetime.fromtimestamp(os.path.getmtime(fp)).strftime("%Y-%m-%d %H:%M:%S")
        except OSError:
            continue  # 文件已被删除/不可访问，跳过

        if fp_normalized not in existing:
            # 全新文件
            to_process.append((fp, fp_normalized, mtime))
            continue

        old = existing[fp_normalized]

        if rescan and old["status"] != "matched":
            # --rescan：重新分类所有非matched的
            to_process.append((fp, fp_normalized, mtime))
            continue

        # Codex P1: 综合判定 mtime + md5 + algo_version
        # md5在这里提前算（只读前1MB，很快），用于检测mtime不变但内容变了的情况
        current_md5 = compute_md5(fp)
        mtime_same = (old["mtime"] == mtime)
        md5_same = (old["md5"] == current_md5) if (old["md5"] and current_md5) else mtime_same
        algo_same = (old["algo_version"] == ALGORITHM_VERSION)

        if mtime_same and md5_same and algo_same:
            # 文件没变 + 算法版本没变 → 跳过
            if old["status"] in ("scanned", "skipped", "matched"):
                skipped_count += 1
                continue

        # 文件变了 或 算法版本变了 或 之前出错 → 重新处理
        to_process.append((fp, fp_normalized, mtime))

    print(f"  需要处理: {len(to_process)} 个 | 跳过: {skipped_count} 个")

    if not to_process:
        print("没有新文件需要处理。")
        print_stats()
        return

    # 第3步：逐个分类并写入数据库
    success = 0
    errors = 0
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    for i, (fp, fp_normalized, mtime) in enumerate(to_process):
        if (i + 1) % 50 == 0 or i == 0:
            print(f"  处理中 [{i+1}/{len(to_process)}]...")

        # Codex P1: 所有文件操作包在try内
        try:
            province = extract_province(fp)
            specialty = extract_specialty(fp)
            file_size = os.path.getsize(fp)
            file_name = Path(fp).name
            md5 = compute_md5(fp)
        except OSError as e:
            errors += 1
            if errors <= 5:
                print(f"  ⚠ 文件不可访问: {Path(fp).name} → {e}")
            continue

        try:
            result = classify_excel_format(fp)
            fmt = result["format"]
            sheets_json = json.dumps(result["sheets"], ensure_ascii=False)
            estimated = result["estimated_items"]
            skip_reason = result.get("skip_reason")
            error_msg = result.get("error_msg")

            # 根据格式和错误信息确定状态
            if error_msg:
                status = "error"
            elif fmt == "not_bill":
                status = "skipped"
            else:
                status = "scanned"

            # UPSERT写入（保留created_at和match_time）
            conn = get_db()
            try:
                _upsert_file(
                    conn, fp_normalized, file_name, file_size, mtime, md5,
                    province, specialty, fmt, status,
                    skip_reason, error_msg,
                    sheets_json, estimated, ALGORITHM_VERSION, now,
                )
                conn.commit()
                success += 1
            finally:
                conn.close()

        except Exception as e:
            errors += 1
            # Codex P1: 错误写库也包在try内，不让写库失败中断全流程
            try:
                conn = get_db()
                try:
                    _upsert_file(
                        conn, fp_normalized, file_name, file_size, mtime, md5,
                        province, specialty, "unknown", "error",
                        None, str(e),  # skip_reason=None, error_msg=错误信息
                        "[]", 0, ALGORITHM_VERSION, now,
                    )
                    conn.commit()
                finally:
                    conn.close()
            except Exception:
                pass  # 写库也失败了，只能放弃这个文件

            if errors <= 5:
                print(f"  ⚠ 错误: {file_name} → {e}")

    print(f"\n扫描完成: 成功 {success} | 错误 {errors}")
    print_stats()


# ============================================================
# 统计输出
# ============================================================

def print_stats():
    """打印数据库统计"""
    if not DB_PATH.exists():
        print("数据库不存在，请先运行扫描。")
        return

    conn = get_db()
    try:
        total = conn.execute("SELECT COUNT(*) FROM file_registry").fetchone()[0]
        print(f"\n{'='*60}")
        print(f"文件登记册统计（共 {total} 条）")
        print(f"{'='*60}")

        print("\n按状态:")
        for row in conn.execute(
            "SELECT status, COUNT(*) as cnt FROM file_registry GROUP BY status ORDER BY cnt DESC"
        ):
            print(f"  {row['status']:12s} {row['cnt']:>5d}")

        print("\n按格式:")
        for row in conn.execute(
            "SELECT format, COUNT(*) as cnt FROM file_registry GROUP BY format ORDER BY cnt DESC"
        ):
            print(f"  {row['format']:20s} {row['cnt']:>5d}")

        print("\n按省份 (TOP 10):")
        for row in conn.execute("""
            SELECT COALESCE(province, '未识别') as prov, COUNT(*) as cnt
            FROM file_registry GROUP BY province ORDER BY cnt DESC LIMIT 10
        """):
            print(f"  {row['prov']:10s} {row['cnt']:>5d}")

        print("\n按专业:")
        for row in conn.execute(
            "SELECT specialty, COUNT(*) as cnt FROM file_registry GROUP BY specialty ORDER BY cnt DESC"
        ):
            print(f"  {row['specialty']:15s} {row['cnt']:>5d}")

        scannable = conn.execute(
            "SELECT COUNT(*) FROM file_registry WHERE status='scanned'"
        ).fetchone()[0]
        est_items = conn.execute(
            "SELECT SUM(estimated_items) FROM file_registry WHERE status='scanned'"
        ).fetchone()[0] or 0
        print(f"\n可匹配文件: {scannable} 个（约 {est_items} 条清单）")

    finally:
        conn.close()


# ============================================================
# 导出 JSON
# ============================================================

def export_json(output_path: str = None):
    """导出为JSON格式（给前端API或手动查看用）"""
    if not DB_PATH.exists():
        print("数据库不存在。")
        return

    if output_path is None:
        output_path = str(DB_PATH.parent / "file_registry.json")

    conn = get_db()
    try:
        rows = conn.execute("SELECT * FROM file_registry").fetchall()
        data = {
            "schema_version": SCHEMA_VERSION,
            "algorithm_version": ALGORITHM_VERSION,
            "export_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "total_files": len(rows),
            "files": {}
        }
        for row in rows:
            data["files"][row["file_path"]] = {
                "file_name": row["file_name"],
                "province": row["province"],
                "specialty": row["specialty"],
                "format": row["format"],
                "status": row["status"],
                "skip_reason": row["skip_reason"],
                "error_msg": row["error_msg"],
                "estimated_items": row["estimated_items"],
                "algo_version": row["algo_version"],
                "scan_time": row["scan_time"],
            }
    finally:
        conn.close()

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"已导出到 {output_path}")


# ============================================================
# 命令行入口
# ============================================================

def main():
    parser = argparse.ArgumentParser(description="批量文件扫描分类工具")
    parser.add_argument("root_dir", nargs="?", default=r"F:\jarvis",
                        help="扫描根目录（默认 F:\\jarvis）")
    parser.add_argument("--specialty", help="只扫描某个专业（如 电气）")
    parser.add_argument("--rescan", action="store_true",
                        help="重新分类（不重跑已matched的）")
    parser.add_argument("--stats", action="store_true",
                        help="只显示统计信息")
    parser.add_argument("--export-json", action="store_true",
                        help="导出为JSON格式")

    args = parser.parse_args()

    if args.stats:
        print_stats()
        return

    if args.export_json:
        export_json()
        return

    scan_directory(
        root_dir=args.root_dir,
        specialty_filter=args.specialty,
        rescan=args.rescan,
    )


if __name__ == "__main__":
    main()
