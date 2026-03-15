"""
清单数据提取工具（五步法第1步）

功能：扫描 F:\\jarvis 所有专业目录，从标准清单Excel中提取结构化数据存入SQLite。
自带识别能力：有12位清单编码的提取数据，非标文件跳过并记录。

提取字段：
  文件级：file_path, file_name, specialty, province, bill_count, status
  清单级：bill_code, bill_name, description, unit, quantity, section, sheet_name

用法：
    python tools/extract_bill_data.py                     # 增量提取全部
    python tools/extract_bill_data.py --specialty 给排水   # 只提取某专业
    python tools/extract_bill_data.py --stats              # 查看统计
    python tools/extract_bill_data.py --rescan             # 全量重扫
"""

import os
import re
import sys
import argparse
import warnings
from pathlib import Path
from datetime import datetime
from collections import defaultdict

# 把项目根目录加入搜索路径
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from db.sqlite import connect_init, connect

# ============================================================
# 常量配置
# ============================================================

JARVIS_DIR = Path("F:/jarvis")
DB_PATH = Path(__file__).resolve().parent.parent / "data" / "bill_library.db"

# 12位清单编码正则：0开头，12位数字
BILL_CODE_RE = re.compile(r'^0[1-9]\d{10}$')

# Excel扩展名
EXCEL_EXTENSIONS = {".xlsx", ".xls", ".xlsm"}

# 广联达等非Excel造价文件（跳过，不是垃圾但读不了）
GLD_EXTENSIONS = {
    ".gbq6", ".gbq7", ".gbq9", ".gbq",
    ".gpb6", ".gpb7", ".gpb",
    ".gad", ".qdg4", ".qdg", ".zbqd",
    ".13jz", ".spw", ".gczj", ".gczjwj",
    ".ysq", ".ygl", ".sxzb4", ".bj23",
    ".e2d", ".pbq", ".zjxm", ".yfjz", ".bsj",
}

# 省份列表（从文件名[省份]提取）
PROVINCE_NAMES = [
    "北京", "天津", "河北", "山西", "内蒙古",
    "辽宁", "吉林", "黑龙江",
    "上海", "江苏", "浙江", "安徽", "福建", "江西", "山东",
    "河南", "湖北", "湖南", "广东", "广西", "海南",
    "重庆", "四川", "贵州", "云南", "西藏",
    "陕西", "甘肃", "青海", "宁夏", "新疆",
]

# 需要跳过的行（包含这些文字的行不是清单数据）
SKIP_KEYWORDS = ["本页小计", "合计", "价税合计", "页 共"]

# 屏蔽 xlrd 的公式警告
warnings.filterwarnings("ignore", message=".*formula.*")


# ============================================================
# 数据库
# ============================================================

def init_db():
    """初始化SQLite数据库"""
    DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    conn = connect_init(DB_PATH)
    try:
        # 文件级记录
        conn.execute("""
            CREATE TABLE IF NOT EXISTS files (
                file_path    TEXT PRIMARY KEY,
                file_name    TEXT NOT NULL,
                specialty    TEXT,
                province     TEXT,
                bill_count   INTEGER DEFAULT 0,
                status       TEXT NOT NULL DEFAULT 'pending'
                             CHECK(status IN ('standard','non_standard','error','skipped')),
                error_msg    TEXT,
                scan_time    TEXT
            )
        """)

        # 清单项级记录
        conn.execute("""
            CREATE TABLE IF NOT EXISTS bill_items (
                id           INTEGER PRIMARY KEY AUTOINCREMENT,
                file_path    TEXT NOT NULL,
                sheet_name   TEXT,
                section      TEXT,
                bill_code    TEXT NOT NULL,
                bill_name    TEXT NOT NULL,
                description  TEXT,
                unit         TEXT,
                quantity     REAL,
                FOREIGN KEY (file_path) REFERENCES files(file_path)
            )
        """)

        conn.execute("CREATE INDEX IF NOT EXISTS idx_bill_code ON bill_items(bill_code)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_bill_name ON bill_items(bill_name)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_file_path ON bill_items(file_path)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_files_status ON files(status)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_files_specialty ON files(specialty)")

        conn.commit()
    finally:
        conn.close()


# ============================================================
# 工具函数
# ============================================================

def get_scan_dirs(filter_specialty=None):
    """获取要扫描的专业目录列表"""
    if not JARVIS_DIR.exists():
        return []
    dirs = sorted([
        d.name for d in JARVIS_DIR.iterdir()
        if d.is_dir() and not d.name.startswith("_")
    ])
    if filter_specialty:
        dirs = [d for d in dirs if d == filter_specialty]
    return dirs


def extract_province(filename):
    """从文件名提取省份（支持[省份]和【省份】两种格式）"""
    # 先尝试方括号
    m = re.search(r'[[\[【]([^]\]】]+)[]\]】]', filename)
    if m:
        text = m.group(1)
        for prov in PROVINCE_NAMES:
            if prov in text:
                return prov
        # [未知] 等不是省份
        return None

    # 没有方括号，从文件名本身找省份
    for prov in PROVINCE_NAMES:
        if prov in filename:
            return prov
    return None


def _cell_str(val):
    """把单元格值转为字符串，None→空字符串"""
    if val is None:
        return ""
    return str(val).strip()


def _is_skip_row(row_texts):
    """判断是否应该跳过这一行（表头、小计等）"""
    joined = " ".join(row_texts)
    for kw in SKIP_KEYWORDS:
        if kw in joined:
            return True
    # 表头行（包含"序号"和"项目编码"）
    if "序号" in joined and "项目编码" in joined:
        return True
    if "综合单价" in joined and "合价" in joined:
        return True
    return False


# ============================================================
# Excel解析：动态检测列位置 + 提取清单数据
# ============================================================

def _detect_columns(rows_iter, max_scan=50):
    """从Sheet的前N行中检测列位置

    找包含"项目编码"的行，确定各列的index。
    返回: (col_map, remaining_rows) 或 (None, [])
      col_map = {"code": int, "name": int, "desc": int, "unit": int, "qty": int}
    """
    buffered_rows = []

    for i, row in enumerate(rows_iter):
        if i >= max_scan:
            break
        texts = [_cell_str(c) for c in row]
        joined = " ".join(texts)

        # 找表头行
        if "项目编码" in joined:
            col_map = {}
            for col_idx, t in enumerate(texts):
                if "项目编码" in t:
                    col_map["code"] = col_idx
                elif "项目名称" in t or "清单名称" in t:
                    col_map["name"] = col_idx
                elif "项目特征" in t:
                    col_map["desc"] = col_idx
                elif "计量" in t and "单位" in t:
                    col_map["unit"] = col_idx
                elif t == "工程量":
                    col_map["qty"] = col_idx

            # 至少要有编码列和名称列
            if "code" in col_map and "name" in col_map:
                # 描述列如果没检测到，默认在名称列后面
                if "desc" not in col_map:
                    col_map["desc"] = col_map["name"] + 1
                if "unit" not in col_map:
                    col_map["unit"] = col_map.get("desc", col_map["name"] + 1) + 1
                if "qty" not in col_map:
                    col_map["qty"] = col_map["unit"] + 1
                return col_map, buffered_rows

        buffered_rows.append(row)

    return None, []


def _safe_float(val):
    """安全转换为浮点数"""
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def _extract_sheet_xlsx(ws, sheet_name):
    """从一个xlsx/xlsm Sheet中提取清单数据"""
    items = []
    rows_iter = ws.iter_rows(values_only=True)

    # 检测列位置
    col_map, _ = _detect_columns(rows_iter)
    if col_map is None:
        return items

    current_section = ""

    # 继续读取剩余行（iter_rows是生成器，已跳过表头）
    for row in rows_iter:
        # 确保行有足够的列
        cells = list(row) + [None] * 10  # 补齐防越界

        code_val = _cell_str(cells[col_map["code"]])
        name_val = _cell_str(cells[col_map["name"]])
        desc_val = _cell_str(cells[col_map.get("desc", col_map["name"] + 1)])
        unit_val = _cell_str(cells[col_map.get("unit", 0)])
        qty_val = cells[col_map.get("qty", 0)]

        row_texts = [_cell_str(c) for c in cells[:8]]

        # 跳过特殊行
        if _is_skip_row(row_texts):
            # 但如果是重复的表头行，需要重新检测列（有些文件每页表头不同）
            continue

        # 有12位编码 → 清单数据行
        if BILL_CODE_RE.match(code_val):
            items.append({
                "sheet_name": sheet_name,
                "section": current_section,
                "bill_code": code_val,
                "bill_name": name_val,
                "description": desc_val,
                "unit": unit_val,
                "quantity": _safe_float(qty_val),
            })
        elif name_val and not code_val:
            # 无编码但有名称 → 可能是章节标题（如"通风系统"、"给排水"）
            # 排除空行和数字行
            if len(name_val) >= 2 and not name_val.replace(".", "").isdigit():
                # 排除金额数字行
                if not re.match(r'^[\d,.]+$', name_val):
                    current_section = name_val

    return items


def _extract_sheet_xls(ws, sheet_name):
    """从一个xls Sheet中提取清单数据"""
    items = []

    # 构造行迭代器
    def row_iter():
        for row_idx in range(ws.nrows):
            yield [ws.cell_value(row_idx, c) for c in range(ws.ncols)]

    rows = row_iter()
    col_map, _ = _detect_columns(rows)
    if col_map is None:
        return items

    current_section = ""

    for row in rows:
        cells = list(row) + [None] * 10

        code_val = _cell_str(cells[col_map["code"]])
        name_val = _cell_str(cells[col_map["name"]])
        desc_val = _cell_str(cells[col_map.get("desc", col_map["name"] + 1)])
        unit_val = _cell_str(cells[col_map.get("unit", 0)])
        qty_val = cells[col_map.get("qty", 0)]

        row_texts = [_cell_str(c) for c in cells[:8]]

        if _is_skip_row(row_texts):
            continue

        if BILL_CODE_RE.match(code_val):
            items.append({
                "sheet_name": sheet_name,
                "section": current_section,
                "bill_code": code_val,
                "bill_name": name_val,
                "description": desc_val,
                "unit": unit_val,
                "quantity": _safe_float(qty_val),
            })
        elif name_val and not code_val:
            if len(name_val) >= 2 and not name_val.replace(".", "").isdigit():
                if not re.match(r'^[\d,.]+$', name_val):
                    current_section = name_val

    return items


# ============================================================
# 文件级处理
# ============================================================

def extract_file(filepath):
    """处理单个Excel文件，返回提取的清单项列表

    返回: (status, items, error_msg)
      status: 'standard' / 'non_standard' / 'error'
      items: 清单项字典列表
      error_msg: 出错时的错误信息
    """
    ext = Path(filepath).suffix.lower()
    all_items = []

    try:
        if ext == ".xls":
            import xlrd
            wb = xlrd.open_workbook(filepath, on_demand=True)
            try:
                for sheet_name in wb.sheet_names()[:50]:  # 最多50个Sheet
                    ws = wb.sheet_by_name(sheet_name)
                    items = _extract_sheet_xls(ws, sheet_name)
                    all_items.extend(items)
            finally:
                wb.release_resources()
        else:
            import openpyxl
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            try:
                for ws in wb.worksheets[:50]:
                    items = _extract_sheet_xlsx(ws, ws.title)
                    all_items.extend(items)
            finally:
                wb.close()

    except Exception as e:
        return "error", [], str(e)[:200]

    if all_items:
        return "standard", all_items, None
    else:
        return "non_standard", [], None


# ============================================================
# 主流程
# ============================================================

def run_extract(filter_specialty=None, rescan=False):
    """主提取流程"""
    init_db()
    conn = connect(DB_PATH)

    try:
        scan_dirs = get_scan_dirs(filter_specialty)
        if not scan_dirs:
            print("没有找到可扫描的目录")
            return

        # 加载已处理的文件（增量模式）
        if not rescan:
            existing = set(
                row[0] for row in
                conn.execute("SELECT file_path FROM files").fetchall()
            )
        else:
            # 全量重扫：清空数据
            conn.execute("DELETE FROM bill_items")
            conn.execute("DELETE FROM files")
            conn.commit()
            existing = set()

        print("=" * 60)
        print("清单数据提取（五步法第1步）")
        print("=" * 60)
        print(f"扫描目录: {', '.join(scan_dirs)}")
        if not rescan:
            print(f"增量模式: 已有 {len(existing)} 个文件记录，只处理新文件")
        else:
            print("全量模式: 从头扫描所有文件")
        print()

        # 统计
        stats = defaultdict(int)
        total_items = 0
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        for specialty in scan_dirs:
            dir_path = JARVIS_DIR / specialty
            if not dir_path.exists():
                continue

            # 收集所有Excel文件
            excel_files = []
            for root, dirs, files in os.walk(dir_path):
                for filename in files:
                    filepath = os.path.join(root, filename)
                    ext = Path(filename).suffix.lower()

                    # 跳过非Excel文件
                    if ext not in EXCEL_EXTENSIONS:
                        if ext in GLD_EXTENSIONS:
                            stats["skipped_gld"] += 1
                        continue

                    # 跳过临时文件
                    if filename.startswith('~') or filename.startswith('.'):
                        continue

                    # 增量：跳过已处理的
                    if filepath in existing:
                        stats["skipped_existing"] += 1
                        continue

                    excel_files.append((filepath, filename))

            if not excel_files:
                print(f"  {specialty}: 没有新文件")
                continue

            print(f"  {specialty}: 处理 {len(excel_files)} 个文件...", end="", flush=True)
            spec_standard = 0
            spec_items = 0

            for filepath, filename in excel_files:
                province = extract_province(filename)
                status, items, error_msg = extract_file(filepath)

                # 写入文件记录
                conn.execute("""
                    INSERT OR REPLACE INTO files
                    (file_path, file_name, specialty, province, bill_count, status, error_msg, scan_time)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """, (filepath, filename, specialty, province, len(items), status, error_msg, now))

                # 写入清单项
                if items:
                    conn.executemany("""
                        INSERT INTO bill_items
                        (file_path, sheet_name, section, bill_code, bill_name, description, unit, quantity)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    """, [
                        (filepath, it["sheet_name"], it["section"], it["bill_code"],
                         it["bill_name"], it["description"], it["unit"], it["quantity"])
                        for it in items
                    ])

                stats[status] += 1
                if status == "standard":
                    spec_standard += 1
                    spec_items += len(items)
                    total_items += len(items)

            # 每个专业处理完提交一次
            conn.commit()
            print(f" 标准{spec_standard}个/{spec_items}条")

        # 打印汇总
        print()
        print("=" * 60)
        total_new = stats["standard"] + stats["non_standard"] + stats["error"]
        print(f"提取完成 — 新处理 {total_new} 个文件")
        print(f"  标准清单: {stats['standard']} 个文件，{total_items} 条清单项")
        print(f"  非标文件: {stats['non_standard']} 个（跳过）")
        print(f"  打开失败: {stats['error']} 个")
        if stats["skipped_existing"]:
            print(f"  增量跳过: {stats['skipped_existing']} 个（已处理过）")
        if stats["skipped_gld"]:
            print(f"  广联达文件: {stats['skipped_gld']} 个（跳过）")
        print(f"  数据库: {DB_PATH}")

    finally:
        conn.close()


def show_stats():
    """显示数据库统计"""
    if not DB_PATH.exists():
        print("数据库不存在，请先运行提取")
        return

    conn = connect(DB_PATH, row_factory=True)
    try:
        print("=" * 60)
        print("清单数据库统计")
        print("=" * 60)

        # 文件统计
        print("\n文件统计:")
        for row in conn.execute(
            "SELECT status, COUNT(*) as cnt FROM files GROUP BY status ORDER BY cnt DESC"
        ):
            label = {"standard": "标准清单", "non_standard": "非标文件",
                     "error": "打开失败", "skipped": "跳过"}.get(row["status"], row["status"])
            print(f"  {label}: {row['cnt']}")

        # 清单项总数
        total = conn.execute("SELECT COUNT(*) FROM bill_items").fetchone()[0]
        unique_codes = conn.execute("SELECT COUNT(DISTINCT bill_code) FROM bill_items").fetchone()[0]
        unique_names = conn.execute("SELECT COUNT(DISTINCT bill_name) FROM bill_items").fetchone()[0]
        unique_descs = conn.execute(
            "SELECT COUNT(DISTINCT description) FROM bill_items WHERE description != ''"
        ).fetchone()[0]
        print(f"\n清单项统计:")
        print(f"  总条数: {total}")
        print(f"  不同编码数: {unique_codes}")
        print(f"  不同名称数: {unique_names}")
        print(f"  不同描述数: {unique_descs}")

        # 按专业统计
        print(f"\n按专业:")
        for row in conn.execute("""
            SELECT f.specialty, COUNT(DISTINCT f.file_path) as files,
                   COUNT(b.id) as items
            FROM files f
            LEFT JOIN bill_items b ON f.file_path = b.file_path
            WHERE f.status = 'standard'
            GROUP BY f.specialty
            ORDER BY items DESC
        """):
            print(f"  {row['specialty']}: {row['files']}个文件 / {row['items']}条")

        # 按省份统计
        print(f"\n按省份（前10）:")
        for row in conn.execute("""
            SELECT f.province, COUNT(DISTINCT f.file_path) as files,
                   COUNT(b.id) as items
            FROM files f
            LEFT JOIN bill_items b ON f.file_path = b.file_path
            WHERE f.status = 'standard' AND f.province IS NOT NULL
            GROUP BY f.province
            ORDER BY items DESC
            LIMIT 10
        """):
            print(f"  {row['province']}: {row['files']}个文件 / {row['items']}条")

        # 最常见的清单项
        print(f"\n最常见的清单项（前10）:")
        for row in conn.execute("""
            SELECT bill_code, bill_name, COUNT(*) as cnt,
                   COUNT(DISTINCT description) as desc_cnt
            FROM bill_items
            GROUP BY bill_code, bill_name
            ORDER BY cnt DESC
            LIMIT 10
        """):
            print(f"  {row['bill_code']} {row['bill_name']}: "
                  f"{row['cnt']}次出现, {row['desc_cnt']}种描述")

    finally:
        conn.close()


# ============================================================
# 入口
# ============================================================

def main():
    parser = argparse.ArgumentParser(description="清单数据提取（五步法第1步）")
    parser.add_argument("--specialty", help="只提取某个专业目录")
    parser.add_argument("--stats", action="store_true", help="查看统计")
    parser.add_argument("--rescan", action="store_true", help="全量重扫（清空重来）")

    args = parser.parse_args()

    if args.stats:
        show_stats()
    else:
        run_extract(
            filter_specialty=args.specialty,
            rescan=args.rescan,
        )


if __name__ == "__main__":
    main()
