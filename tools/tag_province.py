"""
省份标注工具

功能：扫描 F:\\jarvis\\ 下已收集的文件，自动识别省份，在文件名前加省份标签。

用法：
    python tools/tag_province.py                # 正式标注
    python tools/tag_province.py --preview      # 只统计不改名

效果：
    原文件: 某项目给排水工程_wx.xlsx
    改名后: [北京]某项目给排水工程_wx.xlsx

识别逻辑：
    1. 从文件名中找城市/省份名
    2. 打开Excel读前几行找工程地点、项目名称中的地名
    3. 匹配不到的标为[未知]
"""

import os
import re
import sys
import argparse
from pathlib import Path
from collections import defaultdict

# ============================================================
# 城市→省份映射表（地级市+直辖市+省名）
# ============================================================

# 直辖市
CITY_TO_PROVINCE = {}

# 先填充省份名本身（省名→省名简称）
PROVINCE_NAMES = {
    "北京": "北京", "天津": "天津", "上海": "上海", "重庆": "重庆",
    "河北": "河北", "山西": "山西", "辽宁": "辽宁", "吉林": "吉林",
    "黑龙江": "黑龙江", "江苏": "江苏", "浙江": "浙江", "安徽": "安徽",
    "福建": "福建", "江西": "江西", "山东": "山东", "河南": "河南",
    "湖北": "湖北", "湖南": "湖南", "广东": "广东", "海南": "海南",
    "四川": "四川", "贵州": "贵州", "云南": "云南", "陕西": "陕西",
    "甘肃": "甘肃", "青海": "青海", "台湾": "台湾",
    "内蒙古": "内蒙古", "广西": "广西", "西藏": "西藏",
    "宁夏": "宁夏", "新疆": "新疆",
}

# 省名本身也作为关键词
for name, province in PROVINCE_NAMES.items():
    CITY_TO_PROVINCE[name] = province

# 主要城市→省份映射（覆盖全国地级市+重要县级市）
_CITY_DATA = {
    # 北京
    "北京": "北京", "朝阳": "北京", "海淀": "北京", "丰台": "北京",
    "通州": "北京", "大兴": "北京", "昌平": "北京", "顺义": "北京",
    "房山": "北京", "密云": "北京", "怀柔": "北京", "延庆": "北京",
    # 天津
    "天津": "天津", "滨海": "天津", "塘沽": "天津", "武清": "天津",
    # 上海
    "上海": "上海", "浦东": "上海", "闵行": "上海", "松江": "上海",
    "嘉定": "上海", "青浦": "上海", "奉贤": "上海", "金山": "上海",
    "临港": "上海", "崇明": "上海",
    # 重庆
    "重庆": "重庆", "渝北": "重庆", "万州": "重庆", "涪陵": "重庆",
    # 河北
    "石家庄": "河北", "唐山": "河北", "秦皇岛": "河北", "邯郸": "河北",
    "邢台": "河北", "保定": "河北", "张家口": "河北", "承德": "河北",
    "沧州": "河北", "廊坊": "河北", "衡水": "河北", "雄安": "河北",
    # 山西
    "太原": "山西", "大同": "山西", "阳泉": "山西", "长治": "山西",
    "晋城": "山西", "朔州": "山西", "晋中": "山西", "运城": "山西",
    "忻州": "山西", "临汾": "山西", "吕梁": "山西",
    # 辽宁
    "沈阳": "辽宁", "大连": "辽宁", "鞍山": "辽宁", "抚顺": "辽宁",
    "本溪": "辽宁", "丹东": "辽宁", "锦州": "辽宁", "营口": "辽宁",
    "阜新": "辽宁", "辽阳": "辽宁", "盘锦": "辽宁", "铁岭": "辽宁",
    "朝阳": "辽宁", "葫芦岛": "辽宁",
    # 吉林
    "长春": "吉林", "四平": "吉林", "辽源": "吉林", "通化": "吉林",
    "白山": "吉林", "松原": "吉林", "白城": "吉林", "延边": "吉林",
    # 黑龙江
    "哈尔滨": "黑龙江", "齐齐哈尔": "黑龙江", "牡丹江": "黑龙江",
    "佳木斯": "黑龙江", "大庆": "黑龙江", "鸡西": "黑龙江",
    "鹤岗": "黑龙江", "双鸭山": "黑龙江", "伊春": "黑龙江",
    "七台河": "黑龙江", "黑河": "黑龙江", "绥化": "黑龙江",
    # 江苏
    "南京": "江苏", "无锡": "江苏", "徐州": "江苏", "常州": "江苏",
    "苏州": "江苏", "南通": "江苏", "连云港": "江苏", "淮安": "江苏",
    "盐城": "江苏", "扬州": "江苏", "镇江": "江苏", "泰州": "江苏",
    "宿迁": "江苏", "昆山": "江苏", "江阴": "江苏", "常熟": "江苏",
    "张家港": "江苏", "太仓": "江苏", "海安": "江苏",
    # 浙江
    "杭州": "浙江", "宁波": "浙江", "温州": "浙江", "嘉兴": "浙江",
    "湖州": "浙江", "绍兴": "浙江", "金华": "浙江", "衢州": "浙江",
    "舟山": "浙江", "台州": "浙江", "丽水": "浙江", "义乌": "浙江",
    # 安徽
    "合肥": "安徽", "芜湖": "安徽", "蚌埠": "安徽", "淮南": "安徽",
    "马鞍山": "安徽", "淮北": "安徽", "铜陵": "安徽", "安庆": "安徽",
    "黄山": "安徽", "滁州": "安徽", "阜阳": "安徽", "宿州": "安徽",
    "六安": "安徽", "亳州": "安徽", "池州": "安徽", "宣城": "安徽",
    # 福建
    "福州": "福建", "厦门": "福建", "莆田": "福建", "三明": "福建",
    "泉州": "福建", "漳州": "福建", "南平": "福建", "龙岩": "福建",
    "宁德": "福建",
    # 江西
    "南昌": "江西", "景德镇": "江西", "萍乡": "江西", "九江": "江西",
    "新余": "江西", "鹰潭": "江西", "赣州": "江西", "吉安": "江西",
    "宜春": "江西", "抚州": "江西", "上饶": "江西",
    # 山东
    "济南": "山东", "青岛": "山东", "淄博": "山东", "枣庄": "山东",
    "东营": "山东", "烟台": "山东", "潍坊": "山东", "济宁": "山东",
    "泰安": "山东", "威海": "山东", "日照": "山东", "临沂": "山东",
    "德州": "山东", "聊城": "山东", "滨州": "山东", "菏泽": "山东",
    # 河南
    "郑州": "河南", "开封": "河南", "洛阳": "河南", "平顶山": "河南",
    "安阳": "河南", "鹤壁": "河南", "新乡": "河南", "焦作": "河南",
    "濮阳": "河南", "许昌": "河南", "漯河": "河南", "三门峡": "河南",
    "南阳": "河南", "商丘": "河南", "信阳": "河南", "周口": "河南",
    "驻马店": "河南",
    # 湖北
    "武汉": "湖北", "黄石": "湖北", "十堰": "湖北", "宜昌": "湖北",
    "襄阳": "湖北", "鄂州": "湖北", "荆门": "湖北", "孝感": "湖北",
    "荆州": "湖北", "黄冈": "湖北", "咸宁": "湖北", "随州": "湖北",
    "恩施": "湖北",
    # 湖南
    "长沙": "湖南", "株洲": "湖南", "湘潭": "湖南", "衡阳": "湖南",
    "邵阳": "湖南", "岳阳": "湖南", "常德": "湖南", "张家界": "湖南",
    "益阳": "湖南", "郴州": "湖南", "永州": "湖南", "怀化": "湖南",
    "娄底": "湖南",
    # 广东
    "广州": "广东", "韶关": "广东", "深圳": "广东", "珠海": "广东",
    "汕头": "广东", "佛山": "广东", "江门": "广东", "湛江": "广东",
    "茂名": "广东", "肇庆": "广东", "惠州": "广东", "梅州": "广东",
    "汕尾": "广东", "河源": "广东", "阳江": "广东", "清远": "广东",
    "东莞": "广东", "中山": "广东", "潮州": "广东", "揭阳": "广东",
    "云浮": "广东", "南山": "广东", "笔村": "广东", "番禺": "广东",
    "白云": "广东", "黄埔": "广东", "花都": "广东",
    # 海南
    "海口": "海南", "三亚": "海南", "儋州": "海南", "琼海": "海南",
    # 四川
    "成都": "四川", "自贡": "四川", "攀枝花": "四川", "泸州": "四川",
    "德阳": "四川", "绵阳": "四川", "广元": "四川", "遂宁": "四川",
    "内江": "四川", "乐山": "四川", "南充": "四川", "眉山": "四川",
    "宜宾": "四川", "广安": "四川", "达州": "四川", "雅安": "四川",
    "巴中": "四川", "资阳": "四川", "阿坝": "四川", "甘孜": "四川",
    "凉山": "四川", "甘洛": "四川",
    # 贵州
    "贵阳": "贵州", "六盘水": "贵州", "遵义": "贵州", "安顺": "贵州",
    "毕节": "贵州", "铜仁": "贵州", "黔西南": "贵州", "黔东南": "贵州",
    "黔南": "贵州",
    # 云南
    "昆明": "云南", "曲靖": "云南", "玉溪": "云南", "保山": "云南",
    "昭通": "云南", "丽江": "云南", "普洱": "云南", "临沧": "云南",
    "大理": "云南", "红河": "云南", "文山": "云南", "西双版纳": "云南",
    # 陕西
    "西安": "陕西", "铜川": "陕西", "宝鸡": "陕西", "咸阳": "陕西",
    "渭南": "陕西", "延安": "陕西", "汉中": "陕西", "榆林": "陕西",
    "安康": "陕西", "商洛": "陕西", "韦曲": "陕西",
    # 甘肃
    "兰州": "甘肃", "嘉峪关": "甘肃", "金昌": "甘肃", "白银": "甘肃",
    "天水": "甘肃", "武威": "甘肃", "张掖": "甘肃", "平凉": "甘肃",
    "酒泉": "甘肃", "庆阳": "甘肃", "定西": "甘肃", "陇南": "甘肃",
    "临夏": "甘肃", "甘南": "甘肃", "玉门": "甘肃", "临洮": "甘肃",
    # 青海
    "西宁": "青海", "海东": "青海", "海北": "青海", "海西": "青海",
    # 广西
    "南宁": "广西", "柳州": "广西", "桂林": "广西", "梧州": "广西",
    "北海": "广西", "防城港": "广西", "钦州": "广西", "贵港": "广西",
    "玉林": "广西", "百色": "广西", "贺州": "广西", "河池": "广西",
    "来宾": "广西", "崇左": "广西",
    # 内蒙古
    "呼和浩特": "内蒙古", "包头": "内蒙古", "乌海": "内蒙古",
    "赤峰": "内蒙古", "通辽": "内蒙古", "鄂尔多斯": "内蒙古",
    "呼伦贝尔": "内蒙古", "巴彦淖尔": "内蒙古", "乌兰察布": "内蒙古",
    "阿拉善": "内蒙古",
    # 西藏
    "拉萨": "西藏", "日喀则": "西藏", "昌都": "西藏", "林芝": "西藏",
    # 宁夏
    "银川": "宁夏", "石嘴山": "宁夏", "吴忠": "宁夏", "固原": "宁夏",
    "中卫": "宁夏",
    # 新疆
    "乌鲁木齐": "新疆", "克拉玛依": "新疆", "吐鲁番": "新疆",
    "哈密": "新疆", "昌吉": "新疆", "博尔塔拉": "新疆",
    "巴音郭楞": "新疆", "阿克苏": "新疆", "克孜勒苏": "新疆",
    "喀什": "新疆", "和田": "新疆", "伊犁": "新疆", "塔城": "新疆",
    "阿勒泰": "新疆", "石河子": "新疆", "乌恰": "新疆",
    # 香港澳门
    "香港": "香港", "澳门": "澳门",
}

CITY_TO_PROVINCE.update(_CITY_DATA)

# 按城市名长度降序排列（优先匹配长名，避免"南"匹配到"南京"前被"南宁"吃掉）
SORTED_CITIES = sorted(CITY_TO_PROVINCE.keys(), key=len, reverse=True)


# ============================================================
# 省份识别
# ============================================================

def detect_province_from_text(text):
    """从文本中识别省份，返回省份名或None"""
    for city in SORTED_CITIES:
        if city in text:
            province = CITY_TO_PROVINCE[city]
            # 过滤太短的城市名匹配到无关文本（如"海"匹配到"上海"）
            # 要求城市名至少2个字
            if len(city) >= 2:
                return province
    return None


def detect_province_from_binary(filepath):
    """从二进制文件（GBQ/GPB等广联达文件）中搜索城市名，识别省份"""
    try:
        with open(filepath, "rb") as f:
            # 读前50KB和后50KB（项目信息常在头尾）
            head = f.read(50000)
            f.seek(0, 2)  # 到文件末尾
            size = f.tell()
            if size > 50000:
                f.seek(max(0, size - 50000))
                tail = f.read()
            else:
                tail = b""
            data = head + tail

        # 在二进制中搜索城市名的GBK编码（广联达文件多用GBK）
        for city in SORTED_CITIES:
            if len(city) < 2:
                continue
            try:
                encoded = city.encode("gbk")
                if encoded in data:
                    return CITY_TO_PROVINCE[city]
            except UnicodeEncodeError:
                continue
        return None
    except Exception:
        return None


def detect_province_from_excel(filepath):
    """打开Excel，从工程名称/项目名称等字段识别省份"""
    ext = Path(filepath).suffix.lower()
    try:
        if ext == ".xls":
            return _detect_from_xls(filepath)
        else:
            return _detect_from_xlsx(filepath)
    except Exception:
        return None


def _detect_from_xlsx(filepath):
    """从xlsx文件内容识别省份"""
    import openpyxl
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    try:
        for sheet_name in wb.sheetnames:
            # Sheet名也可能有线索
            p = detect_province_from_text(sheet_name)
            if p:
                return p

            ws = wb[sheet_name]
            # 读前20行，找地名
            for row_idx, row in enumerate(ws.iter_rows(max_row=20, values_only=True)):
                for cell in row:
                    if cell is not None:
                        text = str(cell).strip()
                        if len(text) > 2 and len(text) < 200:
                            p = detect_province_from_text(text)
                            if p:
                                return p
        return None
    finally:
        wb.close()


def _detect_from_xls(filepath):
    """从xls文件内容识别省份"""
    import xlrd
    wb = xlrd.open_workbook(filepath, on_demand=True)
    try:
        for sheet_name in wb.sheet_names():
            p = detect_province_from_text(sheet_name)
            if p:
                return p

            ws = wb.sheet_by_name(sheet_name)
            for row_idx in range(min(20, ws.nrows)):
                for col_idx in range(min(20, ws.ncols)):
                    val = ws.cell_value(row_idx, col_idx)
                    if val:
                        text = str(val).strip()
                        if len(text) > 2 and len(text) < 200:
                            p = detect_province_from_text(text)
                            if p:
                                return p
        return None
    finally:
        wb.release_resources()


def tag_file(filepath, preview=False, retag_unknown=False):
    """
    给单个文件标注省份。
    retag_unknown: 如果为True，对已标[未知]的文件重新尝试识别
    返回: (province, new_path) 或 (province, None) 如果preview
    """
    fname = Path(filepath).name
    stem = Path(filepath).stem
    ext = Path(filepath).suffix
    parent = Path(filepath).parent

    # 已经标过的检查
    if fname.startswith("[") and "]" in fname[:8]:
        existing = fname[1:fname.index("]")]
        valid_tags = set(PROVINCE_NAMES.values()) | {"未知", "香港", "澳门"}
        if existing in valid_tags:
            # 如果是[未知]且开启了retag，去掉[未知]前缀重新识别
            if existing == "未知" and retag_unknown:
                original_name = fname[fname.index("]") + 1:]
                # 用原始文件名重新走识别流程
                filepath_for_detect = filepath
                fname = original_name
                stem = Path(original_name).stem
                ext = Path(original_name).suffix
            else:
                return existing, None
        # 不是合法省份（如"审核"、"2"等原始文件名带方括号），当作未标注处理
        # 不需要去掉原始的方括号内容，直接在最前面加省份标签

    # 第1步：从文件名识别
    province = detect_province_from_text(stem)

    # 第2步：从Excel内容识别
    if not province and ext.lower() in (".xlsx", ".xls", ".xlsm"):
        province = detect_province_from_excel(filepath)

    # 第3步：从软件文件（GBQ/GPB等）的二进制内容识别
    SOFTWARE_EXTS = {
        ".gbq6", ".gbq7", ".gbq9", ".gbq",
        ".gpb6", ".gpb7", ".gpb",
        ".zbqd", ".qdg4", ".qdg",
        ".gczjwj", ".gczj", ".gad", ".zjxm",
        ".ysq", ".ygl", ".yfjz",
        ".bj23", ".sxzb4", ".spw", ".bsj", ".13jz",
        ".e2d", ".pbq",
    }
    if not province and ext.lower() in SOFTWARE_EXTS:
        province = detect_province_from_binary(filepath)

    if not province:
        province = "未知"

    # 构造新文件名
    original_fname = Path(filepath).name  # 磁盘上的实际文件名
    if retag_unknown and original_fname.startswith("[未知]"):
        if province != "未知":
            # 从[未知]成功识别到省份，替换前缀
            new_name = f"[{province}]" + original_fname[len("[未知]"):]
        else:
            # 还是未知，不改
            return province, None
    else:
        new_name = f"[{province}]{fname}"

    new_path = os.path.join(str(parent), new_name)

    if preview:
        return province, None

    # 重命名
    try:
        os.rename(filepath, new_path)
        return province, new_path
    except Exception as e:
        return province, None


# ============================================================
# 主流程
# ============================================================

def tag_all(target_dir, preview=False, retag_unknown=False):
    """扫描目录下所有文件，标注省份"""
    print("=" * 60)
    print("  省份标注工具")
    print("=" * 60)
    mode_info = []
    if preview:
        mode_info.append("预览模式")
    if retag_unknown:
        mode_info.append("重新识别[未知]")
    if mode_info:
        print(f"  [{' | '.join(mode_info)}]\n")

    stats = defaultdict(int)  # province → count
    total = 0
    tagged = 0
    retag_count = 0  # 从[未知]成功识别的数量

    # 遍历所有专业子目录
    for entry in sorted(os.listdir(target_dir)):
        subdir = os.path.join(target_dir, entry)
        if not os.path.isdir(subdir) or entry.startswith("_"):
            continue

        files = [f for f in os.listdir(subdir) if os.path.isfile(os.path.join(subdir, f))]
        print(f"\n[{entry}] {len(files)} 个文件")

        for idx, fname in enumerate(files):
            filepath = os.path.join(subdir, fname)
            total += 1

            if (idx + 1) % 100 == 0:
                print(f"  处理 {idx+1}/{len(files)}...")

            # 记录retag前是否是[未知]
            was_unknown = retag_unknown and fname.startswith("[未知]")

            province, new_path = tag_file(filepath, preview=preview,
                                          retag_unknown=retag_unknown)

            # 统计从[未知]成功识别的
            if was_unknown and province != "未知":
                retag_count += 1
            stats[province] += 1
            if province != "未知":
                tagged += 1

    # 打印统计
    print("\n" + "=" * 60)
    print("  省份分布统计")
    print("=" * 60)

    # 按数量排序
    sorted_provinces = sorted(stats.items(), key=lambda x: x[1], reverse=True)

    print(f"\n{'省份':<10} {'文件数':>8} {'占比':>8}")
    print("-" * 28)
    for province, count in sorted_provinces:
        pct = count / total * 100 if total > 0 else 0
        print(f"{province:<10} {count:>8} {pct:>7.1f}%")
    print("-" * 28)
    print(f"{'合计':<10} {total:>8}")
    print(f"\n识别率: {tagged}/{total} = {tagged/total*100:.1f}%")
    if retag_unknown and retag_count > 0:
        print(f"本次从[未知]成功识别: {retag_count} 个")

    if preview:
        print("\n[预览模式] 去掉 --preview 正式标注。")

    return stats


def main():
    parser = argparse.ArgumentParser(description="省份标注工具")
    parser.add_argument("--dir", default=r"F:\jarvis", help="目标目录")
    parser.add_argument("--preview", action="store_true", help="预览模式")
    parser.add_argument("--retag-unknown", action="store_true",
                        help="对已标[未知]的文件重新尝试识别（含读取二进制内容）")
    args = parser.parse_args()

    tag_all(args.dir, preview=args.preview, retag_unknown=args.retag_unknown)


if __name__ == "__main__":
    main()
