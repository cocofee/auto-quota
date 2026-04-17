"""
本地匹配API服务 — 在你的电脑上运行，提供定额匹配算力

懒猫盒子通过HTTP调用这个服务来执行匹配任务，
这样懒猫只需要轻量镜像（~200MB），算力全在你的电脑上。

启动方式：
    python local_match_server.py
    或者双击「启动匹配服务.bat」

端口：9300（默认，可通过 LOCAL_MATCH_PORT 环境变量修改）
"""

import base64
import json
import os
import re
import shutil
import socket
import time
import uuid
import threading
from pathlib import Path

from dotenv import load_dotenv
from fastapi import FastAPI, Form, Header, HTTPException, Query, UploadFile
from fastapi.responses import FileResponse, JSONResponse
import uvicorn

from loguru import logger
from src.excel_compat import ensure_openpyxl_input, validate_excel_upload
from src.output_writer import safe_excel_text

# 加载项目 .env
load_dotenv()

import config  # 项目全局配置（省份列表等）
import main as auto_quota_main  # 匹配入口

# ============================================================
# 配置
# ============================================================

def _require_api_key() -> str:
    api_key = os.getenv("LOCAL_MATCH_API_KEY", "").strip()
    if not api_key:
        raise RuntimeError(
            "LOCAL_MATCH_API_KEY 未配置，出于安全原因拒绝启动。\n"
            "请先设置一个随机高强度密钥，再启动本地匹配服务。"
        )
    return api_key


def _mask_secret(secret: str) -> str:
    if len(secret) <= 8:
        return "*" * len(secret)
    return f"{secret[:4]}...{secret[-4:]}"


# API密钥（必须显式配置，和懒猫环境变量 LOCAL_MATCH_API_KEY 保持一致）
API_KEY = _require_api_key()

# 最大并发匹配任务数
MAX_CONCURRENT = int(os.getenv("LOCAL_MATCH_MAX_CONCURRENT", "5"))

# 单文件上传上限（MB）
MAX_UPLOAD_MB = int(os.getenv("LOCAL_MATCH_MAX_UPLOAD_MB", "30"))

# 临时文件目录
TEMP_DIR = Path("output/temp/remote_match")
TEMP_DIR.mkdir(parents=True, exist_ok=True)

# 任务保留时间（秒），超时后自动清理
TASK_TTL = 3600  # 1小时

# 服务端口（默认 9300；若默认端口不可绑定且未显式指定 LOCAL_MATCH_PORT，则自动回退到可用端口）
DEFAULT_PORT = 9300
_port_env = os.getenv("LOCAL_MATCH_PORT", "").strip()
REQUESTED_PORT = int(_port_env or str(DEFAULT_PORT))
PORT_WAS_EXPLICITLY_SET = bool(_port_env)
PORT = REQUESTED_PORT

# 默认监听所有网卡，便于懒猫盒子/局域网访问；如需仅本机访问可显式设为 127.0.0.1
HOST = os.getenv("LOCAL_MATCH_HOST", "0.0.0.0").strip() or "0.0.0.0"


def _format_bind_error(exc: OSError) -> str:
    winerror = getattr(exc, "winerror", None)
    if winerror is not None:
        return f"{exc} (WinError {winerror})"
    return str(exc)


def _can_bind(host: str, port: int) -> tuple[bool, str | None]:
    family = socket.AF_INET6 if ":" in host and host != "0.0.0.0" else socket.AF_INET
    bind_target = (host, port, 0, 0) if family == socket.AF_INET6 else (host, port)

    try:
        with socket.socket(family, socket.SOCK_STREAM) as sock:
            if os.name == "nt" and hasattr(socket, "SO_EXCLUSIVEADDRUSE"):
                try:
                    sock.setsockopt(socket.SOL_SOCKET, socket.SO_EXCLUSIVEADDRUSE, 1)
                except OSError:
                    pass
            sock.bind(bind_target)
        return True, None
    except OSError as exc:
        return False, _format_bind_error(exc)


def _resolve_listen_port(host: str, requested_port: int, port_was_explicitly_set: bool) -> tuple[int, str | None]:
    can_bind, bind_error = _can_bind(host, requested_port)
    if can_bind:
        return requested_port, None

    port_hint = (
        f"LOCAL_MATCH_PORT={requested_port} 不可用"
        if port_was_explicitly_set
        else f"默认端口 {requested_port} 不可用"
    )
    raise RuntimeError(
        f"{port_hint}，无法启动本地匹配服务。\n"
        f"绑定失败原因: {bind_error}\n"
        "请手动设置一个可用端口后重试，例如设置 LOCAL_MATCH_PORT=9400，并同步更新所有远程客户端的 LOCAL_MATCH_URL。"
    )

# ============================================================
# 全局状态
# ============================================================

# 任务字典：match_id → 任务状态
# 状态结构：{
#   "status": "running" / "completed" / "failed",
#   "progress": 0~100,
#   "current_idx": 当前第几条,
#   "message": 进度文字,
#   "results": 匹配结果（完成后填入）,
#   "error": 错误信息（失败时填入）,
#   "created_at": 创建时间戳,
#   "work_dir": 工作目录路径,
# }
_tasks: dict[str, dict] = {}
_tasks_lock = threading.Lock()

# 并发控制信号量
_semaphore = threading.Semaphore(MAX_CONCURRENT)

# ============================================================
# FastAPI 应用
# ============================================================

app = FastAPI(
    title="本地匹配API服务",
    description="提供定额匹配算力，供懒猫盒子远程调用",
    version="1.0.0",
)


def _safe_list(raw) -> list[str]:
    if isinstance(raw, list):
        return [str(item) for item in raw if item]
    if not raw:
        return []
    if isinstance(raw, str):
        return [raw]
    return []


def _infer_specialty_code_from_quota_id(quota_id: str | None) -> str:
    if not quota_id:
        return ""
    alpha_match = re.match(r"^([A-Za-z]+\d{0,2})-", str(quota_id).strip())
    if alpha_match:
        return alpha_match.group(1).upper()
    numeric_match = re.match(r"^(\d{1,2})-", str(quota_id).strip())
    if not numeric_match:
        return ""
    num = int(numeric_match.group(1))
    return f"C{num}" if 1 <= num <= 12 else str(num)


def _infer_record_category(record: dict) -> str:
    text_parts = [
        str(record.get("bill_name") or ""),
        str(record.get("bill_text") or ""),
        *_safe_list(record.get("quota_names")),
        *_safe_list(record.get("quota_ids")),
    ]
    text = " ".join(text_parts)
    if re.search(r"光伏|升压站|发电", text, re.IGNORECASE):
        return "光伏"
    if re.search(r"电力|变电|输电|配电装置|电力电缆|变压器|母线|开关站|开闭所|间隔", text, re.IGNORECASE):
        return "电力"

    specialty = ""
    for quota_id in _safe_list(record.get("quota_ids")):
        specialty = _infer_specialty_code_from_quota_id(quota_id)
        if specialty:
            break
    if not specialty:
        specialty = str(record.get("specialty") or "").upper()

    if specialty.startswith("C"):
        return "安装"
    if specialty in {"A", "B"}:
        return "建筑装饰"
    if specialty == "D":
        return "市政"
    if specialty == "E":
        return "园林绿化"
    return "安装"


def _filter_records_by_category(records: list[dict], category: str | None) -> list[dict]:
    if not category or category == "all":
        return records
    return [record for record in records if _infer_record_category(record) == category]


def _extract_province_name(name: str | None) -> str:
    text = str(name or "").strip()
    if not text:
        return ""
    city_to_province = {
        "佛山": "广东",
        "深圳": "广东",
        "广州": "广东",
        "东莞": "广东",
        "珠海": "广东",
        "中山": "广东",
        "惠州": "广东",
    }
    match = re.search(r"(.{2,6}?)(省|市|自治区)", text[:12])
    if match:
        region = match.group(1)
    else:
        fallback = re.match(r"^[^\d(（]{2,6}", text)
        region = fallback.group(0).strip() if fallback else text[:2]
    return city_to_province.get(region, region)


def _infer_specialty_from_province_label(name: str | None) -> str:
    text = str(name or "")
    if not text:
        return ""
    if re.search(r"光伏|发电|升压站", text, re.IGNORECASE):
        return "光伏"
    if re.search(r"电力|输电|变电|配电", text, re.IGNORECASE):
        return "电力"
    if re.search(r"园林|绿化", text):
        return "园林绿化"
    if re.search(r"市政", text):
        return "市政"
    if re.search(r"装饰|装修", text):
        return "建筑装饰"
    if re.search(r"安装", text):
        return "安装"
    return ""


def _filter_records_by_scope(
    records: list[dict],
    province_name: str | None,
    specialty_name: str | None,
) -> list[dict]:
    filtered = records
    if province_name:
        filtered = [
            record for record in filtered
            if _extract_province_name(record.get("province")) == province_name
        ]
    if specialty_name and specialty_name != "all":
        filtered = [
            record for record in filtered
            if _infer_specialty_from_province_label(record.get("province")) == specialty_name
        ]
    return filtered


def _verify_api_key(api_key: str):
    """验证API密钥"""
    if api_key != API_KEY:
        raise HTTPException(status_code=401, detail="API Key不正确")


def _sanitize_client_filename(filename: str | None, default: str = "input.xlsx") -> str:
    raw = (filename or "").replace("\\", "/").split("/")[-1]
    raw = raw.replace("\x00", "").replace("\r", "").replace("\n", "").strip().strip(". ")
    return raw or default


def _decode_params_payload(payload: str | None) -> dict:
    """Decode ASCII-safe match params sent by the web backend."""
    if not payload:
        return {}
    try:
        raw = base64.b64decode(payload).decode("utf-8")
        data = json.loads(raw)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"params_payload 解码失败: {exc}") from exc
    if not isinstance(data, dict):
        raise HTTPException(status_code=400, detail="params_payload 必须是对象")
    return data


def _safe_join_under(base_dir: Path, filename: str) -> Path:
    candidate = (base_dir / filename).resolve()
    base_resolved = base_dir.resolve()
    if candidate.parent != base_resolved:
        raise HTTPException(status_code=400, detail="非法文件名")
    return candidate


async def _read_and_validate_upload(file: UploadFile, label: str) -> tuple[bytes, str, str]:
    filename = _sanitize_client_filename(file.filename)
    max_bytes = MAX_UPLOAD_MB * 1024 * 1024
    content = bytearray()
    while chunk := await file.read(1024 * 1024):
        content.extend(chunk)
        if len(content) > max_bytes:
            raise HTTPException(status_code=400, detail=f"{label}大小超过 {MAX_UPLOAD_MB}MB 限制")
    try:
        info = validate_excel_upload(filename, bytes(content[:8]))
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return bytes(content), filename, info.normalized_suffix


@app.get("/health")
def health_check(x_api_key: str = Header(default="")):
    """健康检查 - 返回版本号和可用省份列表"""
    _verify_api_key(x_api_key)

    meta = _build_province_meta_payload()

    # 统计当前活跃任务数
    with _tasks_lock:
        active = sum(1 for t in _tasks.values() if t["status"] == "running")

    return {
        "status": "ok",
        "version": "1.0.0",
        **meta,
        "active_tasks": active,
        "max_concurrent": MAX_CONCURRENT,
    }


@app.get("/provinces-meta")
def provinces_meta(x_api_key: str = Header(default="")):
    """获取省份列表和分组元数据"""
    _verify_api_key(x_api_key)
    return _build_province_meta_payload()


def _build_province_meta_payload() -> dict[str, object]:
    return {
        "provinces": config.list_db_provinces(),
        "groups": config.get_province_groups(),
        "subgroups": config.get_province_subgroups(),
    }


@app.post("/match")
async def create_match(
    file: UploadFile,
    province: str = Form(...),
    mode: str = Form(default="search"),
    sheet: str = Form(default=None),
    limit: int = Form(default=None),
    no_experience: bool = Form(default=False),
    agent_llm: str = Form(default=None),
    params_payload: str = Form(default=None),
    x_api_key: str = Header(default=""),
):
    """提交匹配任务 — 上传Excel+参数，异步执行，返回match_id"""
    _verify_api_key(x_api_key)

    # 检查并发限制（非阻塞尝试获取信号量）
    if not _semaphore.acquire(blocking=False):
        raise HTTPException(
            status_code=429,
            detail=f"服务繁忙，当前已有{MAX_CONCURRENT}个任务在执行，请稍后重试"
        )

    try:
        # 生成任务ID和工作目录
        match_id = str(uuid.uuid4())
        work_dir = TEMP_DIR / match_id
        work_dir.mkdir(parents=True, exist_ok=True)

        content, filename, normalized_suffix = await _read_and_validate_upload(file, "Excel文件")
        input_name = f"input{normalized_suffix}"

        # 保存上传的Excel文件
        input_path = _safe_join_under(work_dir, input_name)
        input_path.write_bytes(content)

        # 初始化任务状态
        with _tasks_lock:
            _tasks[match_id] = {
                "status": "running",
                "progress": 0,
                "current_idx": 0,
                "message": "任务已创建，等待执行...",
                "results": None,
                "error": None,
                "created_at": time.time(),
                "work_dir": str(work_dir),
            }

        # 组装匹配参数
        decoded_payload = _decode_params_payload(params_payload)
        if decoded_payload:
            province = str(decoded_payload.get("province") or province)
            mode = str(decoded_payload.get("mode") or mode)
            sheet = decoded_payload.get("sheet") or sheet
            limit = decoded_payload.get("limit", limit)
            no_experience = bool(decoded_payload.get("no_experience", no_experience))
            agent_llm = decoded_payload.get("agent_llm") or agent_llm

        params = {
            "input_file": str(input_path),
            "province": province,
            "mode": mode,
            "sheet": sheet,
            "limit": limit,
            "no_experience": no_experience,
            "agent_llm": agent_llm,
        }

        # 启动后台线程执行匹配
        thread = threading.Thread(
            target=_run_match,
            args=(match_id, params),
            daemon=True,
        )
        thread.start()

        return {"match_id": match_id}

    except Exception:
        # 出错时释放信号量
        _semaphore.release()
        raise


@app.get("/match/{match_id}/progress")
def get_progress(match_id: str, x_api_key: str = Header(default="")):
    """查询匹配进度"""
    _verify_api_key(x_api_key)

    with _tasks_lock:
        task = _tasks.get(match_id)

    if not task:
        raise HTTPException(status_code=404, detail="任务不存在")

    return {
        "status": task["status"],
        "progress": task["progress"],
        "current_idx": task["current_idx"],
        "message": task["message"],
        "error": task.get("error"),
    }


@app.get("/match/{match_id}/results")
def get_results(match_id: str, x_api_key: str = Header(default="")):
    """获取匹配结果 — 任务完成后才能调用"""
    _verify_api_key(x_api_key)

    with _tasks_lock:
        task = _tasks.get(match_id)

    if not task:
        raise HTTPException(status_code=404, detail="任务不存在")

    if task["status"] == "running":
        raise HTTPException(status_code=409, detail="任务还在执行中")

    if task["status"] == "failed":
        raise HTTPException(status_code=500, detail=task.get("error", "匹配失败"))

    # 返回结果JSON
    return task["results"]


@app.get("/match/{match_id}/output.xlsx")
def download_excel(match_id: str, x_api_key: str = Header(default="")):
    """下载输出的Excel文件"""
    _verify_api_key(x_api_key)

    with _tasks_lock:
        task = _tasks.get(match_id)

    if not task:
        raise HTTPException(status_code=404, detail="任务不存在")

    excel_path = Path(task["work_dir"]) / "output.xlsx"
    if not excel_path.exists():
        raise HTTPException(status_code=404, detail="Excel文件不存在")

    return FileResponse(
        path=str(excel_path),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="output.xlsx",
    )


# ============================================================
# 编清单接口（本地有清单库数据，直接执行）
# ============================================================

@app.post("/compile-bill/preview")
async def compile_bill_preview(
    file: UploadFile,
    bill_version: str = Form(default="2024"),
    x_api_key: str = Header(default=""),
):
    """编清单预览 — 上传Excel，自动匹配12位清单编码，返回结果"""
    _verify_api_key(x_api_key)

    content, filename, normalized_suffix = await _read_and_validate_upload(file, "工程量文件")

    # 保存临时文件
    work_dir = TEMP_DIR / f"compile_{uuid.uuid4().hex[:8]}"
    work_dir.mkdir(parents=True, exist_ok=True)
    input_path = _safe_join_under(work_dir, f"input{normalized_suffix}")
    input_path.write_bytes(content)

    try:
        from src.bill_reader import BillReader
        from src.bill_compiler import compile_items

        reader = BillReader()
        items = reader.read_file(str(input_path))

        if not items:
            raise HTTPException(status_code=400, detail="未从Excel中读取到清单项，请检查文件格式。")

        compiled = compile_items(items, bill_version=bill_version)

        # 构建返回数据
        result_items = []
        matched_count = 0
        for i, item in enumerate(compiled):
            original_code = item.get("code", "").strip()
            bill_match = item.get("bill_match")

            if bill_match and bill_match.get("code_12"):
                code = bill_match["code_12"]
                code_source = "matched"
                matched_count += 1
            elif original_code and len(original_code) >= 9:
                code = original_code
                code_source = "original"
                matched_count += 1
            else:
                code = original_code
                code_source = "unmatched"

            result_items.append({
                "index": i + 1,
                "name": item.get("name", ""),
                "description": item.get("description", ""),
                "unit": item.get("unit", ""),
                "quantity": item.get("quantity", None),
                "bill_code": safe_excel_text(code),
                "bill_code_source": code_source,
                "matched_name": safe_excel_text(bill_match.get("name", "") if bill_match else ""),
                "standard_name": safe_excel_text(item.get("standard_name", "")),
                "standard_unit": safe_excel_text(item.get("standard_unit", "")),
                "compiled_features": safe_excel_text(item.get("compiled_features", "")),
                "sheet_name": safe_excel_text(item.get("sheet_name", "")),
                "section": safe_excel_text(item.get("section", "")),
            })

        return {
            "total": len(result_items),
            "matched": matched_count,
            "unmatched": len(result_items) - matched_count,
            "bill_version": bill_version,
            "items": result_items,
        }

    finally:
        shutil.rmtree(work_dir, ignore_errors=True)


@app.post("/compile-bill/execute")
async def compile_bill_execute(
    file: UploadFile,
    bill_version: str = Form(default="2024"),
    x_api_key: str = Header(default=""),
):
    """编清单导出 — 上传Excel，返回编好的工程量清单Excel文件"""
    _verify_api_key(x_api_key)

    content, filename, normalized_suffix = await _read_and_validate_upload(file, "工程量文件")

    work_dir = TEMP_DIR / f"compile_{uuid.uuid4().hex[:8]}"
    work_dir.mkdir(parents=True, exist_ok=True)
    input_path = _safe_join_under(work_dir, f"input{normalized_suffix}")
    input_path.write_bytes(content)

    try:
        import openpyxl
        from src.bill_reader import BillReader
        from src.bill_compiler import compile_items

        reader = BillReader()
        items = reader.read_file(str(input_path))
        if not items:
            raise HTTPException(status_code=400, detail="未从Excel中读取到清单项，请检查文件格式。")

        compiled = compile_items(items, bill_version=bill_version)

        # 生成结果Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "工程量清单"

        headers = ["序号", "项目编码", "项目名称", "项目特征描述", "计量单位", "工程量", "编码来源", "原始名称"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = openpyxl.styles.Font(bold=True)

        for i, item in enumerate(compiled):
            original_code = item.get("code", "").strip()
            bill_match = item.get("bill_match")

            if bill_match and bill_match.get("code_12"):
                code = bill_match["code_12"]
                source = "自动匹配"
            elif original_code and len(original_code) >= 9:
                code = original_code
                source = "原始编码"
            else:
                code = original_code
                source = "未匹配"

            # 优先用标准名称和标准单位，没有则降级用原始值
            display_name = item.get("standard_name") or item.get("name", "")
            display_features = item.get("compiled_features") or item.get("description", "")
            display_unit = item.get("standard_unit") or item.get("unit", "")
            original_name = item.get("name", "")

            row = i + 2
            ws.cell(row=row, column=1, value=i + 1)
            ws.cell(row=row, column=2, value=safe_excel_text(code))
            ws.cell(row=row, column=3, value=safe_excel_text(display_name))
            ws.cell(row=row, column=4, value=safe_excel_text(display_features))
            ws.cell(row=row, column=5, value=safe_excel_text(display_unit))
            ws.cell(row=row, column=6, value=item.get("quantity", ""))
            ws.cell(row=row, column=7, value=safe_excel_text(source))
            ws.cell(row=row, column=8, value=safe_excel_text(original_name))

        col_widths = [8, 18, 20, 50, 10, 12, 12, 25]
        for col, width in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

        out_path = work_dir / "output.xlsx"
        wb.save(str(out_path))
        wb.close()

        return FileResponse(
            path=str(out_path),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=f"{Path(_sanitize_client_filename(filename)).stem}_工程量清单.xlsx",
        )

    except HTTPException:
        raise
    except Exception as e:
        shutil.rmtree(work_dir, ignore_errors=True)
        raise HTTPException(status_code=500, detail=f"编清单失败: {e}")


# ============================================================
# 定额搜索接口（供懒猫远程模式转发）
# ============================================================

@app.get("/quota-search")
def search_quotas(
    keyword: str,
    province: str,
    book: str = None,
    chapter: str = None,
    limit: int = 20,
    x_api_key: str = Header(default=""),
):
    """按关键词搜索定额"""
    _verify_api_key(x_api_key)
    from src.quota_db import QuotaDB

    try:
        db = QuotaDB(province)
        results = db.search_by_keywords(keyword, chapter=chapter, book=book, limit=limit)
        items = [
            {
                "quota_id": r.get("quota_id", ""),
                "name": r.get("name", ""),
                "unit": r.get("unit", ""),
                "chapter": r.get("chapter", ""),
                "book": r.get("book", ""),
            }
            for r in results
        ]
        return {"items": items, "total": len(items), "keyword": keyword, "province": province}
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail=f"省份 '{province}' 的定额库不存在")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"搜索失败: {e}")


@app.get("/quota-search/by-id")
def get_quota_by_id(
    quota_id: str,
    province: str,
    x_api_key: str = Header(default=""),
):
    """按定额编号精确查询"""
    _verify_api_key(x_api_key)
    from src.quota_db import QuotaDB

    try:
        db = QuotaDB(province)
        results = db.get_quota_by_id(quota_id)
        if not results:
            return {"items": [], "total": 0}
        items = [
            {
                "quota_id": r.get("quota_id", ""),
                "name": r.get("name", ""),
                "unit": r.get("unit", ""),
                "chapter": r.get("chapter", ""),
                "book": r.get("book", ""),
            }
            for r in results
        ]
        return {"items": items, "total": len(items)}
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail=f"省份 '{province}' 的定额库不存在")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"查询失败: {e}")


@app.get("/quota-search/provinces")
def list_search_provinces(x_api_key: str = Header(default="")):
    """获取可用的省份定额库列表"""
    _verify_api_key(x_api_key)
    provinces = config.list_db_provinces()
    return {"items": provinces}


def _find_sibling_libs(province: str) -> list[str]:
    """找同省的其他定额库（用于跨库搜索）

    例如传入"河南省通用安装工程预算定额(2016)"，
    返回["河南省房屋建筑与装饰工程预算定额(2016)", "河南省市政工程预算定额(2016)"]

    匹配逻辑：提取省份前缀（如"河南省"），找所有同前缀的库，排除自己。
    """
    import re
    # 提取省份前缀：匹配"XX省"或"XX市"或"XX自治区"
    m = re.match(r'^(.{2,6}(?:省|市|自治区))', province)
    if not m:
        return []
    prefix = m.group(1)

    all_libs = config.list_db_provinces()
    siblings = [
        lib for lib in all_libs
        if lib.startswith(prefix) and lib != province
    ]
    return siblings


@app.get("/quota-search/smart")
def smart_search(
    name: str,
    province: str,
    description: str = "",
    specialty: str = "",
    limit: int = 10,
    x_api_key: str = Header(default=""),
):
    """智能搜索定额（清单原文 → 自动清洗+同义词+级联搜索）

    和 /quota-search 的区别：
    - /quota-search 需要调用方自己把"JDG20"转成"紧定式钢导管"
    - /quota-search/smart 直接传清单原文，系统自动做术语转换和级联搜索

    参数:
        name: 清单项目名称（如"JDG20暗配"、"PPR给水管DN25"）
        province: 省份定额库名称（如"北京市建设工程施工消耗量标准(2024)"）
        description: 清单特征描述（可选，如"沟槽连接 镀锌钢管"）
        specialty: 专业册号（可选，如"C10"，不传则自动识别）
        limit: 最大返回条数

    返回:
        items: 候选定额列表（按匹配度排序）
        search_query: 系统构建的搜索词（方便调试）
    """
    _verify_api_key(x_api_key)

    try:
        from src.text_parser import TextParser
        from src.hybrid_searcher import HybridSearcher
        from src.specialty_classifier import classify as classify_specialty

        parser = TextParser()

        # 第1步：自动识别专业（如果没传）
        if not specialty:
            spec_result = classify_specialty(name, description)
            specialty = spec_result.get("primary", "") if isinstance(spec_result, dict) else ""

        # 第2步：构建搜索query（清洗+同义词+规范化）
        search_query = parser.build_quota_query(
            name, description, specialty=specialty
        )

        # 第3步：级联搜索（BM25 + 向量，主专业 → 全库）
        searcher = HybridSearcher(province)
        books = [specialty] if specialty else None
        candidates = searcher.search(search_query, top_k=limit, books=books)

        # 如果主专业搜不到足够结果，扩到全库
        if len(candidates) < 3 and books:
            candidates_all = searcher.search(search_query, top_k=limit, books=None)
            # 合并去重
            seen = {c.get("quota_id") for c in candidates}
            for c in candidates_all:
                if c.get("quota_id") not in seen:
                    candidates.append(c)
                    seen.add(c.get("quota_id"))
            candidates = candidates[:limit]

        # 第4步：同省跨库搜索
        # 安装项目里经常有土建相关清单（拆除/开槽/封堵/支墩/防水等），
        # 这些定额在建筑装饰库里，不在安装库里。
        # 触发条件：主库结果不够 OR 清单含跨库关键词（即使主库有结果也要搜）
        _CROSS_LIB_KEYWORDS = (
            "拆除", "拆卸",           # 拆除定额基本都在建筑装饰库
            "抹灰", "粉刷",           # 装饰工程
            "砌筑", "砌墙", "拆墙",   # 土建工程
            "支墩", "基础",           # 混凝土基础
        )
        need_cross = len(candidates) < 3  # 结果太少，必须跨库
        if not need_cross:
            # 结果够了，但如果清单含跨库关键词，也要搜（防止语义偏离）
            combined_text = name + " " + description
            need_cross = any(kw in combined_text for kw in _CROSS_LIB_KEYWORDS)

        if need_cross:
            sibling_libs = _find_sibling_libs(province)
            for sib_lib in sibling_libs:
                try:
                    sib_searcher = HybridSearcher(sib_lib)
                    sib_results = sib_searcher.search(search_query, top_k=limit, books=None)
                    seen = {c.get("quota_id") for c in candidates}
                    for c in sib_results:
                        if c.get("quota_id") not in seen:
                            # 标记来源库，方便调试
                            c["cross_lib"] = sib_lib
                            candidates.append(c)
                            seen.add(c.get("quota_id"))
                except Exception as e:
                    logger.warning(f"跨库搜索 {sib_lib} 失败（跳过）: {e}")
            candidates = candidates[:limit]

        # 构建返回结果
        items = [
            {
                "quota_id": c.get("quota_id", ""),
                "name": c.get("name", ""),
                "unit": c.get("unit", ""),
                "chapter": c.get("chapter", ""),
                "book": c.get("book", ""),
                "score": round(c.get("hybrid_score", 0), 4),
                "cross_lib": c.get("cross_lib", ""),  # 跨库来源（空=主库）
            }
            for c in candidates
        ]

        return {
            "items": items,
            "total": len(items),
            "search_query": search_query,
            "specialty": specialty,
            "province": province,
        }

    except FileNotFoundError:
        raise HTTPException(status_code=404, detail=f"省份 '{province}' 的定额库不存在")
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"智能搜索失败: {e}")


# ============================================================
# 经验库写入（远程模式下懒猫转发到这里）
# ============================================================

from pydantic import BaseModel as _BaseModel
from typing import Optional as _Optional


class _StoreExperienceRequest(_BaseModel):
    """经验库写入请求"""
    name: str
    desc: str = ""
    quota_ids: list[str]
    quota_names: list[str] = []
    reason: str = ""
    specialty: str = ""
    province: str = ""
    confirmed: bool = False
    feedback_payload: dict | None = None


@app.post("/experience/store")
def store_experience_api(
    req: _StoreExperienceRequest,
    x_api_key: str = Header(default=""),
):
    """单条经验库写入"""
    _verify_api_key(x_api_key)
    try:
        from tools.jarvis_store import store_one
        result = store_one(
            name=req.name,
            desc=req.desc,
            quota_ids=req.quota_ids,
            quota_names=req.quota_names,
            reason=req.reason,
            specialty=req.specialty,
            province=req.province or None,
            confirmed=req.confirmed,
            feedback_payload=req.feedback_payload,
        )
        return {"success": bool(result), "record_id": result if isinstance(result, int) else 0}
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"经验库写入失败: {e}")


class _StoreExperienceBatchRequest(_BaseModel):
    """批量经验库写入请求"""
    records: list[dict]
    province: str
    reason: str = ""
    confirmed: bool = False


@app.post("/experience/store-batch")
def store_experience_batch_api(
    req: _StoreExperienceBatchRequest,
    x_api_key: str = Header(default=""),
):
    """批量经验库写入"""
    _verify_api_key(x_api_key)
    try:
        from tools.jarvis_store import store_one
        count = 0
        for rec in req.records:
            if rec.get("quota_ids"):
                ok = store_one(
                    name=rec["name"],
                    desc=rec.get("desc", ""),
                    quota_ids=rec["quota_ids"],
                    quota_names=rec.get("quota_names", []),
                    reason=req.reason,
                    specialty=rec.get("specialty", ""),
                    province=req.province or None,
                    confirmed=req.confirmed,
                    feedback_payload=rec.get("feedback_payload"),
                )
                if ok:
                    count += 1
        return {"success": True, "count": count}
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"批量经验库写入失败: {e}")


class _FlagDisputedRequest(_BaseModel):
    """争议标记请求"""
    bill_name: str
    province: str
    reason: str = ""


@app.post("/experience/flag-disputed")
def flag_disputed_api(
    req: _FlagDisputedRequest,
    x_api_key: str = Header(default=""),
):
    """标记权威层记录为有争议（纠正经验库直通结果时调用）"""
    _verify_api_key(x_api_key)
    try:
        from src.experience_db import ExperienceDB
        db = ExperienceDB(province=req.province)
        affected = db.flag_disputed(
            bill_name=req.bill_name,
            province=req.province,
            reason=req.reason,
        )
        return {"success": True, "affected": affected}
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"争议标记失败: {e}")


# ============================================================
# 主材价格查询（远程模式下懒猫转发到这里）
# ============================================================

@app.get("/experience/stats")
def experience_stats_api(x_api_key: str = Header(default="")):
    """返回经验库统计。"""
    _verify_api_key(x_api_key)
    try:
        from src.experience_db import ExperienceDB

        db = ExperienceDB()
        return db.get_stats()
    except Exception as e:
        logger.exception("获取经验库统计失败")
        raise HTTPException(status_code=500, detail=f"获取经验库统计失败: {e}")


@app.get("/experience/records")
def experience_records_api(
    layer: str = Query(default="all"),
    province: str | None = Query(default=None),
    province_name: str | None = Query(default=None),
    specialty_name: str | None = Query(default=None),
    category: str | None = Query(default=None),
    page: int = Query(default=1),
    size: int = Query(default=20),
    x_api_key: str = Header(default=""),
):
    """返回经验库记录列表。"""
    _verify_api_key(x_api_key)
    if page < 1:
        page = 1
    if size < 1 or size > 100:
        size = 20

    try:
        from src.experience_db import ExperienceDB

        db = ExperienceDB(province=province)
        if layer == "authority":
            records = db.get_authority_records(province=province, limit=0)
            for record in records:
                record["layer_type"] = "authority"
        elif layer == "candidate":
            records = db.get_candidate_records(province=province, limit=0)
            for record in records:
                record["layer_type"] = "candidate"
        else:
            authority = db.get_authority_records(province=province, limit=0)
            for record in authority:
                record["layer_type"] = "authority"
            candidate = db.get_candidate_records(province=province, limit=0)
            for record in candidate:
                record["layer_type"] = "candidate"
            records = authority + candidate

        records = _filter_records_by_scope(records, province_name, specialty_name)
        records = _filter_records_by_category(records, category)
        total = len(records)
        start = (page - 1) * size
        end = start + size
        return {
            "items": records[start:end],
            "total": total,
            "page": page,
            "size": size,
        }
    except Exception as e:
        logger.exception("获取经验记录失败")
        raise HTTPException(status_code=500, detail=f"获取经验记录失败: {e}")


@app.get("/experience/search")
def experience_search_api(
    q: str = Query(...),
    province: str | None = Query(default=None),
    province_name: str | None = Query(default=None),
    specialty_name: str | None = Query(default=None),
    category: str | None = Query(default=None),
    limit: int = Query(default=20),
    x_api_key: str = Header(default=""),
):
    """搜索经验库记录。"""
    _verify_api_key(x_api_key)
    if not q.strip():
        raise HTTPException(status_code=400, detail="搜索关键词不能为空")
    if limit < 1 or limit > 200:
        limit = 20

    try:
        from src.experience_db import ExperienceDB

        db = ExperienceDB(province=province)
        text = q.strip()
        escaped = text.replace("%", "\\%").replace("_", "\\_")
        like_pattern = f"%{escaped}%"
        conn = db._connect(row_factory=True)
        try:
            cursor = conn.cursor()
            text_match = """(
                bill_text = ? OR COALESCE(bill_name, '') = ?
                OR bill_text LIKE ? ESCAPE '\\' OR COALESCE(bill_name, '') LIKE ? ESCAPE '\\'
            )"""
            rank_order = """
                CASE
                    WHEN bill_text = ? THEN 0
                    WHEN COALESCE(bill_name, '') = ? THEN 1
                    WHEN bill_text LIKE ? ESCAPE '\\' THEN 2
                    WHEN COALESCE(bill_name, '') LIKE ? ESCAPE '\\' THEN 3
                    ELSE 4
                END ASC,
                confidence DESC, id DESC
            """
            if province:
                where = f"province = ? AND {text_match}"
                params = [
                    province, text, text, like_pattern, like_pattern,
                    text, text, like_pattern, like_pattern, limit,
                ]
            else:
                where = text_match
                params = [
                    text, text, like_pattern, like_pattern,
                    text, text, like_pattern, like_pattern, limit,
                ]

            cursor.execute(
                f"""
                SELECT * FROM experiences
                WHERE {where}
                ORDER BY {rank_order}
                LIMIT ?
                """,
                params,
            )
            rows = cursor.fetchall()
            items = [db._normalize_record_quota_fields(dict(row)) for row in rows]
            for item in items:
                item["layer_type"] = item.get("layer", "candidate")
            items = _filter_records_by_scope(items, province_name, specialty_name)
            items = _filter_records_by_category(items, category)
            return {"items": items, "total": len(items)}
        finally:
            conn.close()
    except Exception as e:
        logger.exception("搜索经验库失败")
        raise HTTPException(status_code=500, detail=f"搜索经验库失败: {e}")


@app.post("/experience/{record_id:int}/promote")
def promote_experience_api(
    record_id: int,
    x_api_key: str = Header(default=""),
):
    """晋升经验记录到权威层。"""
    _verify_api_key(x_api_key)
    try:
        from src.experience_db import ExperienceDB

        db = ExperienceDB()
        success = db.promote_to_authority(record_id)
        if not success:
            raise HTTPException(status_code=404, detail="记录不存在或已在权威层")
        return {"message": "晋升成功"}
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("晋升经验记录失败")
        raise HTTPException(status_code=500, detail=f"晋升经验记录失败: {e}")


@app.post("/experience/{record_id:int}/demote")
def demote_experience_api(
    record_id: int,
    x_api_key: str = Header(default=""),
):
    """降级经验记录到候选层。"""
    _verify_api_key(x_api_key)
    try:
        from src.experience_db import ExperienceDB

        db = ExperienceDB()
        success = db.demote_to_candidate(record_id)
        if not success:
            raise HTTPException(status_code=404, detail="记录不存在或已在候选层")
        return {"message": "降级成功"}
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("降级经验记录失败")
        raise HTTPException(status_code=500, detail=f"降级经验记录失败: {e}")


@app.delete("/experience/{record_id:int}")
def delete_experience_api(
    record_id: int,
    x_api_key: str = Header(default=""),
):
    """删除经验记录。"""
    _verify_api_key(x_api_key)
    try:
        from src.experience_db import ExperienceDB

        db = ExperienceDB()
        conn = db._connect()
        try:
            cursor = conn.execute("DELETE FROM experiences WHERE id = ?", (record_id,))
            conn.commit()
            deleted = cursor.rowcount > 0
        finally:
            conn.close()

        if deleted:
            try:
                coll = db.collection
                if coll is not None:
                    coll.delete(ids=[str(record_id)])
            except Exception as e:
                logger.opt(exception=e).warning(f"清理向量索引失败（id={record_id}）")

        if not deleted:
            raise HTTPException(status_code=404, detail="记录不存在")
        return {"message": "删除成功"}
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("删除经验记录失败")
        raise HTTPException(status_code=500, detail=f"删除经验记录失败: {e}")


@app.delete("/experience/by-province")
def delete_experience_by_province_api(
    province: str = Query(...),
    x_api_key: str = Header(default=""),
):
    """按省份删除经验记录。"""
    _verify_api_key(x_api_key)
    if not province or not province.strip():
        raise HTTPException(status_code=400, detail="省份名称不能为空")

    try:
        from src.experience_db import ExperienceDB

        db = ExperienceDB(province=province)
        conn = db._connect()
        try:
            cursor = conn.execute("SELECT id FROM experiences WHERE province = ?", (province,))
            ids_to_delete = [str(row[0]) for row in cursor.fetchall()]
            if not ids_to_delete:
                return {"message": "已删除 0 条记录", "deleted": 0}

            conn.execute("DELETE FROM experiences WHERE province = ?", (province,))
            conn.commit()

            try:
                coll = db.collection
                if coll is not None:
                    coll.delete(ids=ids_to_delete)
            except Exception as e:
                logger.opt(exception=e).warning(f"批量清理向量索引失败（省份={province}）")
        finally:
            conn.close()

        return {"message": f"已删除 {len(ids_to_delete)} 条记录", "deleted": len(ids_to_delete)}
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("按省份删除经验记录失败")
        raise HTTPException(status_code=500, detail=f"按省份删除经验记录失败: {e}")


class _BatchPromoteRequest(_BaseModel):
    """批量晋升请求。"""
    province: str | None = None
    dry_run: bool = True


@app.post("/experience/batch-promote")
def batch_promote_api(
    req: _BatchPromoteRequest,
    x_api_key: str = Header(default=""),
):
    """智能批量晋升候选层记录。"""
    _verify_api_key(x_api_key)
    try:
        from src.experience_db import ExperienceDB

        db = ExperienceDB(province=req.province)
        records = db.get_candidate_records(province=req.province, limit=0)
        records = [record for record in records if record.get("source") != "project_import_suspect"]

        promoted = 0
        skipped = 0
        errors: list[str] = []

        for record in records:
            quota_ids_raw = record.get("quota_ids", "[]")
            if isinstance(quota_ids_raw, str):
                try:
                    quota_ids = json.loads(quota_ids_raw)
                except Exception:
                    quota_ids = []
            else:
                quota_ids = quota_ids_raw

            if not quota_ids:
                skipped += 1
                if len(errors) < 5:
                    bill = record.get("bill_name") or record.get("bill_text", "")[:30]
                    errors.append(f"{bill}: 无定额编号")
                continue

            bill_text = record.get("bill_text", "")
            try:
                validation = db._validate_quota_ids(
                    bill_text,
                    quota_ids,
                    province=record.get("province", ""),
                )
            except Exception:
                skipped += 1
                continue

            if not validation.get("valid", False):
                skipped += 1
                if len(errors) < 5:
                    bill = record.get("bill_name") or bill_text[:30]
                    err_msg = "; ".join(validation.get("errors", []))[:50]
                    errors.append(f"{bill}: {err_msg}")
                continue

            if req.dry_run:
                promoted += 1
                continue

            if db.promote_to_authority(record["id"], reason="智能批量晋升（定额校验通过）"):
                promoted += 1
            else:
                skipped += 1

        return {
            "total": len(records),
            "promoted": promoted,
            "skipped": skipped,
            "errors": errors,
            "dry_run": req.dry_run,
        }
    except Exception as e:
        logger.exception("批量晋升失败")
        raise HTTPException(status_code=500, detail=f"批量晋升失败: {e}")


@app.get("/material-price/provinces")
def material_price_provinces(x_api_key: str = Header(default="")):
    """返回有价格数据的省份列表"""
    _verify_api_key(x_api_key)
    import sqlite3
    from src.material_db import DB_PATH
    conn = sqlite3.connect(str(DB_PATH))
    try:
        rows = conn.execute(
            """SELECT province, COUNT(*) as cnt
               FROM price_fact
               WHERE province != '' AND province != '全国'
               GROUP BY province
               ORDER BY cnt DESC"""
        ).fetchall()
        return {"provinces": [{"name": r[0], "count": r[1]} for r in rows]}
    finally:
        conn.close()


@app.get("/material-price/cities")
def material_price_cities(
    province: str,
    x_api_key: str = Header(default=""),
):
    """返回指定省份下有价格数据的城市列表"""
    _verify_api_key(x_api_key)
    import sqlite3
    from src.material_db import DB_PATH
    conn = sqlite3.connect(str(DB_PATH))
    try:
        rows = conn.execute(
            """SELECT city, COUNT(*) as cnt
               FROM price_fact
               WHERE province = ? AND city != ''
               GROUP BY city
               ORDER BY cnt DESC""",
            (province,)
        ).fetchall()
        return {"cities": [{"name": r[0], "count": r[1]} for r in rows]}
    finally:
        conn.close()


@app.get("/material-price/periods")
def material_price_periods(
    province: str,
    city: str = "",
    x_api_key: str = Header(default=""),
):
    """返回指定省份/城市的信息价期次列表"""
    _verify_api_key(x_api_key)
    import sqlite3
    from src.material_db import DB_PATH
    conn = sqlite3.connect(str(DB_PATH))
    try:
        conditions = ["province = ?", "period_start != ''"]
        params: list = [province]
        if city:
            conditions.append("city = ?")
            params.append(city)
        where = " AND ".join(conditions)
        rows = conn.execute(
            f"""SELECT period_start, period_end, COUNT(*) as cnt
                FROM price_fact
                WHERE {where}
                GROUP BY period_start, period_end
                ORDER BY period_end DESC
                LIMIT 24""",
            params
        ).fetchall()

        def _label(start):
            try:
                parts = start.split("-")
                return f"{int(parts[0])}年{int(parts[1])}月"
            except (IndexError, ValueError):
                return start

        return {
            "periods": [
                {"start": r[0], "end": r[1], "count": r[2], "label": _label(r[0])}
                for r in rows
            ]
        }
    finally:
        conn.close()


class _MaterialLookupRequest(_BaseModel):
    """主材批量查价请求"""
    model_config = {"extra": "ignore"}  # 兼容新旧版本，忽略多余字段
    materials: list[dict]
    province: str
    city: str = ""
    period_end: str = ""
    price_type: str = "info"  # info=信息价, market=市场价；普通查价默认只查信息价


def _build_gldjc_search_url(name: str = "", spec: str = "", keyword: str = "", province_code: str = "1") -> str:
    from urllib.parse import quote

    search_keyword = str(keyword or "").strip()
    material_name = str(name or "").strip()
    material_spec = str(spec or "").strip()

    if not search_keyword:
        if material_spec and material_spec not in material_name:
            search_keyword = f"{material_name} {material_spec}".strip()
        else:
            search_keyword = material_name

    if not search_keyword:
        return ""

    return f"https://www.gldjc.com/scj/so.html?keyword={quote(search_keyword, safe='')}&l={province_code}"


def _load_material_price_cache() -> dict:
    cache_path = Path(__file__).resolve().parent / "data" / "material_prices.json"
    if not cache_path.exists():
        return {}
    try:
        with open(cache_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        return data if isinstance(data, dict) else {}
    except Exception:
        return {}


def _build_lookup_label(
    *,
    name: str = "",
    spec: str = "",
    unit: str = "",
    price: object = None,
    source: str = "",
    match_label: str = "",
) -> str | None:
    explicit_label = str(match_label or "").strip()
    if explicit_label:
        return explicit_label

    material_text = " ".join(part for part in [str(name or "").strip(), str(spec or "").strip()] if part).strip()
    parts: list[str] = []

    if str(source or "").strip():
        parts.append(str(source).strip())
    if material_text:
        parts.append(material_text)
    if str(unit or "").strip():
        parts.append(str(unit).strip())

    try:
        if price is not None and str(price).strip() != "":
            parts.append(f"{float(price):.2f}")
    except (TypeError, ValueError):
        pass

    return " | ".join(parts) if parts else None


def _spec_contains_exact(result_spec: str, target_spec: str) -> bool:
    result_text = re.sub(r"\s+", "", str(result_spec or "").upper())
    target_text = re.sub(r"\s+", "", str(target_spec or "").upper())
    if not result_text or not target_text:
        return False

    pattern = re.escape(target_text)
    if target_text[0].isdigit():
        pattern = rf"(?<!\d){pattern}"
    if target_text[-1].isdigit():
        pattern = rf"{pattern}(?!\d)"
    return re.search(pattern, result_text) is not None


def _is_usable_gldjc_cache_entry(spec: str, unit: str, cached: dict) -> bool:
    if not isinstance(cached, dict):
        return False
    if not cached.get("price_with_tax"):
        return False

    matched_unit = str(cached.get("matched_unit") or "").strip()
    if unit and matched_unit:
        import sys

        tools_dir = str(Path(__file__).resolve().parent / "tools")
        if tools_dir not in sys.path:
            sys.path.insert(0, tools_dir)

        try:
            from gldjc_price import _try_convert_price, check_unit_compatible
        except Exception:
            return False

        if not check_unit_compatible(unit, matched_unit) and _try_convert_price(1.0, matched_unit, unit, "", spec) is None:
            return False

    matched_spec = str(cached.get("matched_spec") or "").strip()
    if spec:
        if not matched_spec:
            return False
        if not _spec_contains_exact(matched_spec, spec):
            return False

    return True


@app.post("/material-price/lookup")
def material_price_lookup(
    req: dict,
    x_api_key: str = Header(default=""),
):
    """批量查价"""
    _verify_api_key(x_api_key)
    import re as _re
    from src.material_db import MaterialDB
    db = MaterialDB()

    province = req.get("province", "")
    city = req.get("city", "")
    period_end = req.get("period_end", "")
    materials = req.get("materials", [])

    # 普通查价默认只查信息价；不再走市场价缓存兜底。
    source_filter = ""
    price_type = req.get("price_type", "info")
    if price_type in {"all", "info"}:
        source_filter = "government"
    elif price_type == "market":
        source_filter = "market"

    results = []
    for mat in materials:
        name = mat.get("name", "").strip()
        spec = mat.get("spec", "").strip()
        unit = mat.get("unit", "").strip()
        if not name:
            results.append({**mat, "lookup_price": None, "lookup_source": "名称为空", "lookup_url": None, "lookup_label": None})
            continue
        price_info = db.search_price_by_name(
            name, province=province, city=city, period_end=period_end,
            spec=spec, target_unit=unit,
            source_type=source_filter
        )
        if price_info:
            results.append({
                **mat,
                "lookup_price": price_info["price"],
                "lookup_source": price_info.get("source", "价格库"),
                "lookup_url": None,
                "lookup_label": _build_lookup_label(
                    name=name,
                    spec=spec,
                    unit=price_info.get("unit") or unit,
                    price=price_info.get("price"),
                    source=price_info.get("source", "价格库"),
                ),
            })
            continue
        # 从名称中提取规格再查一次
        m = _re.search(r'[Dd][Nn]\s*\d+|De\s*\d+|Φ\s*\d+|\d+mm', name)
        if m:
            extracted_spec = m.group(0).replace(" ", "")
            short_name = name[:m.start()].strip()
            if short_name:
                price_info2 = db.search_price_by_name(
                    short_name, province=province, city=city, period_end=period_end,
                    spec=extracted_spec, target_unit=unit,
                    source_type=source_filter
                )
                if price_info2:
                    results.append({
                        **mat,
                        "lookup_price": price_info2["price"],
                        "lookup_source": price_info2.get("source", "价格库"),
                        "lookup_url": None,
                        "lookup_label": _build_lookup_label(
                            name=short_name,
                            spec=extracted_spec,
                            unit=price_info2.get("unit") or unit,
                            price=price_info2.get("price"),
                            source=price_info2.get("source", "价格库"),
                        ),
                    })
                    continue
        results.append({**mat, "lookup_price": None, "lookup_source": "未查到", "lookup_url": None, "lookup_label": None})

    found = sum(1 for r in results if r.get("lookup_price") is not None)
    return {
        "results": results,
        "stats": {"total": len(results), "found": found, "not_found": len(results) - found},
    }


class _MaterialContributeRequest(_BaseModel):
    """用户贡献价格请求"""
    items: list[dict]


@app.post("/material-price/contribute")
def material_price_contribute(
    req: _MaterialContributeRequest,
    x_api_key: str = Header(default=""),
):
    """用户手填价格存入候选层"""
    _verify_api_key(x_api_key)
    from src.material_db import MaterialDB
    db = MaterialDB()
    saved = 0
    for item in req.items:
        name = item.get("name", "").strip()
        spec = item.get("spec", "").strip()
        unit = item.get("unit", "").strip()
        price = item.get("price")
        province = item.get("province", "").strip()
        city = item.get("city", "").strip()
        if not name or price is None:
            continue
        try:
            price_val = float(price)
        except (ValueError, TypeError):
            continue
        if price_val <= 0 or price_val > 1_000_000:
            continue
        material_id = db.add_material(name, spec=spec, unit=unit)
        db.add_price(
            material_id=material_id,
            price_incl_tax=price_val,
            source_type="user_contribute",
            province=province, city=city, unit=unit,
            authority_level="reference",
            source_doc="用户手填",
            dedup=True,
        )
        saved += 1
    return {"saved": saved, "message": f"已保存 {saved} 条价格"}


# ============================================================
# 广材网实时查价（从Docker转发过来，本机执行爬虫）
# ============================================================

class _GldjcLookupRequest(_BaseModel):
    materials: list[dict]
    cookie: str
    province: str = ""
    city: str = ""
    period_end: str = ""


class _GldjcCookieVerifyRequest(_BaseModel):
    cookie: str
    province: str = ""
    city: str = ""


def _build_gldjc_session(cookie: str):
    import requests as _requests

    session = _requests.Session()
    for part in re.split(r";\s*", str(cookie or "").strip()):
        if "=" in part:
            key, value = part.split("=", 1)
            session.cookies.set(key.strip(), value.strip())
    return session


def _analyze_gldjc_verify_response(html: str, result_count: int) -> tuple[str, str]:
    text = str(html or "")
    lower = text.lower()
    if "请登录" in text or "登录后" in text or ("login" in lower and "token" not in lower):
        return "invalid", "Cookie已失效，请重新复制完整Cookie"
    if any(flag in text for flag in ("访问过于频繁", "请求过于频繁", "安全验证", "行为验证", "验证码")):
        return "limited", "广材网当前触发限制，请稍后重试或更换Cookie"
    if result_count > 0:
        return "valid", "Cookie有效"
    return "limited", "未命中测试结果，疑似Cookie受限或页面异常"


def _verify_gldjc_cookie(cookie: str, province: str = "", city: str = "") -> dict:
    import sys
    import requests as _requests

    tools_dir = str(Path(__file__).resolve().parent / "tools")
    if tools_dir not in sys.path:
        sys.path.insert(0, tools_dir)

    from gldjc_price import SEARCH_URL, _get_headers, resolve_gldjc_area_code

    verify_keyword = "焊接钢管 DN80"
    session = _build_gldjc_session(cookie)

    region_plans: list[tuple[str, str]] = []
    province_code = resolve_gldjc_area_code(province=province, city="")
    if province_code and province_code != "1":
        region_plans.append((province or "省级", province_code))
    region_plans.append(("全国", "1"))

    seen_codes: set[str] = set()
    checks: list[dict] = []

    for scope_label, area_code in region_plans:
        if area_code in seen_codes:
            continue
        seen_codes.add(area_code)
        try:
            resp = session.get(
                SEARCH_URL,
                params={"keyword": verify_keyword, "l": area_code},
                headers=_get_headers(),
                timeout=20,
            )
            resp.raise_for_status()
        except _requests.RequestException as e:
            return {
                "ok": False,
                "status": "limited",
                "message": f"请求广材网失败: {e}",
                "keyword": verify_keyword,
                "scope": scope_label,
                "area_code": area_code,
                "url": _build_gldjc_search_url(keyword=verify_keyword, province_code=area_code),
                "checks": checks,
            }

        html = resp.text or ""
        result_count = len(re.findall(r'class="price-block"', html))
        status, message = _analyze_gldjc_verify_response(html, result_count)
        check = {
            "scope": scope_label,
            "area_code": area_code,
            "result_count": result_count,
            "status": status,
            "url": _build_gldjc_search_url(keyword=verify_keyword, province_code=area_code),
        }
        checks.append(check)

        if status == "valid":
            return {
                "ok": True,
                "status": status,
                "message": message,
                "keyword": verify_keyword,
                "scope": scope_label,
                "area_code": area_code,
                "url": check["url"],
                "checks": checks,
            }
        if status == "invalid":
            return {
                "ok": False,
                "status": status,
                "message": message,
                "keyword": verify_keyword,
                "scope": scope_label,
                "area_code": area_code,
                "url": check["url"],
                "checks": checks,
            }

    last_check = checks[-1] if checks else {
        "scope": province or "全国",
        "area_code": province_code or "1",
        "url": _build_gldjc_search_url(keyword=verify_keyword, province_code=province_code or "1"),
    }
    return {
        "ok": False,
        "status": "limited",
        "message": "未命中测试结果，疑似Cookie受限或页面异常",
        "keyword": verify_keyword,
        "scope": last_check["scope"],
        "area_code": last_check["area_code"],
        "url": last_check["url"],
        "checks": checks,
    }

@app.post("/material-price/gldjc-lookup")
def material_price_gldjc_lookup(
    req: _GldjcLookupRequest,
    x_api_key: str = Header(default=""),
):
    """广材网实时查价（本机执行，Cookie绑定本机IP）"""
    _verify_api_key(x_api_key)

    cookie = req.cookie.strip()
    materials = req.materials
    province = req.province.strip()
    city = req.city.strip()
    if not cookie:
        raise HTTPException(400, "请输入广材网Cookie")
    if not materials:
        raise HTTPException(400, "材料列表为空")

    # 直接复用 tools/gldjc_price.py 的核心函数
    import sys
    import time
    import random
    tools_dir = str(Path(__file__).parent / "tools")
    if tools_dir not in sys.path:
        sys.path.insert(0, tools_dir)

    from gldjc_price import (
        parse_material, search_material_web, filter_and_score,
        get_representative_price_result, build_match_label, determine_confidence,
        build_approximate_price_candidates,
        GldjcCookieInvalidError,
        load_cache, save_cache, update_cache, check_cache, build_region_search_plans,
    )
    import requests as _requests
    from datetime import datetime

    # 创建带Cookie的session
    session = _requests.Session()
    for part in re.split(r";\s*", cookie):
        if "=" in part:
            key, value = part.split("=", 1)
            session.cookies.set(key.strip(), value.strip())

    cache = load_cache()
    results = []
    search_dedup: dict[str, list] = {}
    net_requests = 0
    blocked = False
    blocked_reason = "已暂停（疑似被限制）"
    MAX_BATCH = 30

    for i, mat in enumerate(materials):
        name = mat.get("name", "").strip()
        unit = mat.get("unit", "").strip()
        spec = mat.get("spec", "").strip()
        lookup_url = _build_gldjc_search_url(name=name, spec=spec)

        if not name:
            results.append({**mat, "gldjc_price": None, "gldjc_source": "名称为空", "gldjc_url": None})
            continue

        # 实时查价默认绕过旧缓存，避免历史误匹配结果被重复复用

        if net_requests >= MAX_BATCH:
            results.append({**mat, "gldjc_price": None, "gldjc_source": f"已达单次上限{MAX_BATCH}条", "gldjc_url": lookup_url})
            continue

        if blocked:
            results.append({**mat, "gldjc_price": None, "gldjc_source": blocked_reason, "gldjc_url": lookup_url})
            continue

        # 实时搜索广材网
        parsed = parse_material(name, spec)
        base_name = parsed["base_name"]
        specs = parsed["specs"]
        search_plans = build_region_search_plans(parsed["search_keywords"], province=province, city=city)
        link_keyword = parsed["search_keywords"][0] if parsed["search_keywords"] else (f"{base_name} {specs[0]}".strip() if specs else base_name)
        link_plan = search_plans[0] if search_plans else {"keyword": link_keyword, "area_code": "1", "scope": "全国"}
        lookup_url = _build_gldjc_search_url(name=name, spec=spec, keyword=link_plan["keyword"], province_code=link_plan["area_code"])
        all_results = []
        searched_keyword = ""
        searched_scope = ""
        searched_area_code = link_plan["area_code"]

        for plan in search_plans:
            kw = plan["keyword"]
            area_code = plan["area_code"]
            scope = plan["scope"]
            dedup_key = f"{area_code}|{kw}"
            if dedup_key in search_dedup:
                all_results = search_dedup[dedup_key]
                searched_keyword = kw
                searched_scope = scope
                searched_area_code = area_code
                if all_results:
                    break
                continue

            if net_requests > 0:
                time.sleep(random.uniform(5, 8))

            try:
                web_results = search_material_web(session, kw, area_code)
            except GldjcCookieInvalidError as exc:
                logger.warning(f"广材网 Cookie 失效，停止后续查询: {exc}")
                blocked = True
                blocked_reason = str(exc)
                break
            net_requests += 1
            searched_keyword = kw
            searched_scope = scope
            searched_area_code = area_code
            search_dedup[dedup_key] = web_results

            if not web_results and net_requests >= 3:
                recent_empty = sum(1 for r in list(search_dedup.values())[-3:] if not r)
                if recent_empty >= 3:
                    logger.warning("广材网连续3次搜索无结果，疑似Cookie失效或被限制")
                    blocked = True

            if web_results:
                all_results = web_results
                break
            time.sleep(random.uniform(1, 2))

        if blocked and not all_results:
            results.append({**mat, "gldjc_price": None, "gldjc_source": blocked_reason, "gldjc_url": lookup_url, "gldjc_label": None})
            continue

        if not all_results:
            update_cache(cache, name, unit, {
                "price_with_tax": None, "confidence": "低",
                "match_status": "未匹配", "source": "广材网",
                "query_date": datetime.now().strftime("%Y-%m-%d"),
                "result_count": 0, "searched_keyword": searched_keyword,
                "gldjc_url": None,
            }, spec=spec, region=searched_scope or province or "全国")
            results.append({**mat, "gldjc_price": None, "gldjc_source": "未查到", "gldjc_url": None, "gldjc_label": None})
            continue

        scored = filter_and_score(all_results, unit, specs, base_name, request_name=name, request_spec=spec)
        if not scored:
            scored = filter_and_score(
                all_results,
                unit,
                specs,
                base_name,
                request_name=name,
                request_spec=spec,
                allow_relaxed_spec=True,
            )
        confidence = determine_confidence(scored, unit, specs)
        approximate_candidates = []
        price_source = scored
        if not price_source:
            approximate_candidates = build_approximate_price_candidates(
                all_results,
                unit,
                request_name=name,
                request_spec=spec,
            )
            price_source = approximate_candidates
            confidence = "低"
        used_approximate = bool(approximate_candidates)
        selected_result = get_representative_price_result(price_source)
        exact_detail_url = str(
            (selected_result or {}).get("detail_url")
            or (selected_result or {}).get("url")
            or ""
        ).strip()
        selected_price = None
        if selected_result:
            try:
                selected_price = round(float(selected_result.get("market_price")), 2)
            except (TypeError, ValueError):
                selected_price = None
        match_label = build_match_label(
            selected_result,
            fallback_text=(searched_keyword or name),
        )
        if used_approximate:
            match_status = "近似匹配"
            if match_label:
                match_label = f"近似价 | {match_label}"
        else:
            match_status = "精确匹配" if confidence == "高" else ("模糊匹配" if confidence == "中" else "低置信度")

        scope_label = searched_scope or province or "全国"
        lookup_url = exact_detail_url or _build_gldjc_search_url(
            name=name,
            spec=spec,
            keyword=searched_keyword or link_keyword,
            province_code=searched_area_code or "1",
        )

        if selected_price:
            update_cache(cache, name, unit, {
                "price_with_tax": selected_price,
                "price_without_tax": round(selected_price / 1.13, 2),
                "confidence": confidence,
                "match_status": match_status,
                "source": "广材网",
                "query_date": datetime.now().strftime("%Y-%m-%d"),
                "result_count": len(price_source), "searched_keyword": searched_keyword,
                "gldjc_url": lookup_url,
                "match_label": match_label,
            }, spec=spec, region=scope_label)
            results.append({
                **mat,
                "gldjc_price": selected_price,
                "gldjc_source": f"{'广材网近似价' if used_approximate else '广材网市场价'}({scope_label})",
                "gldjc_url": lookup_url,
                "gldjc_label": match_label,
            })
        else:
            update_cache(cache, name, unit, {
                "price_with_tax": selected_price,
                "price_without_tax": round(selected_price / 1.13, 2) if selected_price else None,
                "confidence": "低", "match_status": "低置信度", "source": "广材网",
                "query_date": datetime.now().strftime("%Y-%m-%d"),
                "result_count": len(price_source), "searched_keyword": searched_keyword,
                "gldjc_url": None,
                "match_label": None,
            }, spec=spec, region=scope_label)
            results.append({
                **mat,
                "gldjc_price": None,
                "gldjc_source": "未查到",
                "gldjc_url": None,
                "gldjc_label": None,
            })

        if net_requests % 10 == 0:
            save_cache(cache)

    save_cache(cache)
    logger.info(f"广材网查价完成：{len(materials)}条，{net_requests}次请求，"
                f"{'被限制' if blocked else '正常'}")

    found = sum(1 for r in results if r.get("gldjc_price"))
    return {"results": results, "total": len(results), "found": found}


@app.post("/material-price/gldjc-cookie-verify")
def material_price_gldjc_cookie_verify(
    req: _GldjcCookieVerifyRequest,
    x_api_key: str = Header(default=""),
):
    """广材网Cookie快速校验接口。"""
    _verify_api_key(x_api_key)
    cookie = req.cookie.strip()
    province = req.province.strip()
    city = req.city.strip()

    if not cookie:
        raise HTTPException(400, "请输入广材网Cookie")

    return _verify_gldjc_cookie(cookie, province=province, city=city)


# ============================================================
# 后台匹配执行
# ============================================================

def _run_match(match_id: str, params: dict):
    """在后台线程中执行匹配（调用 main.run()）"""
    try:
        work_dir = Path(_tasks[match_id]["work_dir"])

        # 进度回调：更新内存字典
        def progress_cb(percent, current_idx, message, result=None):
            with _tasks_lock:
                if match_id in _tasks:
                    _tasks[match_id].update(
                        progress=percent,
                        current_idx=current_idx,
                        message=message,
                    )

        # 输出路径
        excel_output = str(work_dir / "output.xlsx")
        json_output = str(work_dir / "results.json")
        processing_input, normalize_result = ensure_openpyxl_input(
            params["input_file"],
            work_dir / "input_importable.xlsx",
        )
        if normalize_result:
            _tasks[match_id]["message"] = "已自动转换为可导入的 .xlsx，开始匹配..."

        # 调用核心匹配函数
        result = auto_quota_main.run(
            input_file=str(processing_input),
            mode=params["mode"],
            output=excel_output,
            province=params["province"],
            sheet=params.get("sheet"),
            limit=params.get("limit"),
            no_experience=params.get("no_experience", False),
            agent_llm=params.get("agent_llm"),
            json_output=json_output,
            interactive=False,  # API调用不能交互
            progress_callback=progress_cb,
            original_file=params["input_file"],
        )

        # 标记完成
        with _tasks_lock:
            if match_id in _tasks:
                _tasks[match_id].update(
                    status="completed",
                    progress=100,
                    message="匹配完成",
                    results=result,
                )

    except Exception as e:
        logger.exception(f"本地匹配任务失败: match_id={match_id}")
        # 标记失败
        with _tasks_lock:
            if match_id in _tasks:
                _tasks[match_id].update(
                    status="failed",
                    error=str(e),
                    message=f"匹配失败: {e}",
                )
    finally:
        # 释放并发信号量
        _semaphore.release()


# ============================================================
# 定时清理过期任务
# ============================================================

def _cleanup_expired_tasks():
    """每10分钟清理一次过期任务（完成超过1小时的）"""
    while True:
        time.sleep(600)  # 10分钟检查一次
        now = time.time()
        expired = []

        with _tasks_lock:
            for mid, task in list(_tasks.items()):
                if task["status"] in ("completed", "failed"):
                    if now - task["created_at"] > TASK_TTL:
                        expired.append(mid)
            for mid in expired:
                del _tasks[mid]

        # 清理临时文件
        for mid in expired:
            work_dir = TEMP_DIR / mid
            if work_dir.exists():
                shutil.rmtree(work_dir, ignore_errors=True)

        if expired:
            print(f"[清理] 已清理 {len(expired)} 个过期任务")


# ============================================================
# 启动入口
# ============================================================

def main():
    """启动本地匹配API服务"""
    global PORT

    selected_port, _ = _resolve_listen_port(
        host=HOST,
        requested_port=REQUESTED_PORT,
        port_was_explicitly_set=PORT_WAS_EXPLICITLY_SET,
    )
    PORT = selected_port

    # 启动清理线程
    cleaner = threading.Thread(target=_cleanup_expired_tasks, daemon=True)
    cleaner.start()

    # 打印启动信息
    print("=" * 60)
    print("  本地匹配API服务")
    print("=" * 60)
    print(f"  监听地址: {HOST}")
    print(f"  端口: {PORT}")
    print(f"  API Key: {_mask_secret(API_KEY)}")
    print(f"  最大并发: {MAX_CONCURRENT}")
    print(f"  上传上限: {MAX_UPLOAD_MB}MB")
    print(f"  临时目录: {TEMP_DIR}")
    print()
    print("  把以下配置填入懒猫盒子的环境变量：")
    print(f"    MATCH_BACKEND=remote")
    print(f"    LOCAL_MATCH_URL=http://你的电脑IP:{PORT}")
    print("    LOCAL_MATCH_API_KEY=<使用你当前已配置的密钥>")
    print()
    print("  按 Ctrl+C 停止服务")
    print("=" * 60)

    # 启动服务
    uvicorn.run(
        app,
        host=HOST,
        port=PORT,
        log_level="info",
    )


if __name__ == "__main__":
    main()
