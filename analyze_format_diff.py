"""
分析两个Excel文件的格式差异
文件1：审核输出的.xlsx文件（openpyxl读取）
文件2：原始广联达的.xls文件（xlrd读取）
"""

import openpyxl
from openpyxl.utils import get_column_letter
import xlrd

# ============================================================
# 第一部分：分析文件1（.xlsx格式，用openpyxl）
# ============================================================

file1 = r"C:\Users\Administrator\Documents\trae_projects\auto-quota\output\匹配结果_20260219_103354_c9aec7_已审核.xlsx"

print("=" * 100)
print("【文件1】审核输出文件分析")
print(f"路径：{file1}")
print("=" * 100)

wb1 = openpyxl.load_workbook(file1)
print(f"\nSheet列表：{wb1.sheetnames}")

for sheet_name in wb1.sheetnames:
    ws = wb1[sheet_name]
    print(f"\n{'─' * 100}")
    print(f"Sheet: {sheet_name}")
    print(f"行数: {ws.max_row}, 列数: {ws.max_column}")
    print(f"{'─' * 100}")

    # 合并单元格信息
    print(f"\n  【合并单元格】共 {len(ws.merged_cells.ranges)} 个:")
    for merged in ws.merged_cells.ranges:
        print(f"    {merged}")

    # 列宽信息
    print(f"\n  【列宽信息】:")
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        col_dim = ws.column_dimensions.get(col_letter)
        if col_dim:
            width = col_dim.width
            hidden = col_dim.hidden
        else:
            width = "默认(~8.43)"
            hidden = False
        print(f"    列{col_idx}({col_letter}): 宽度={width}, 隐藏={hidden}")

    # 逐行分析前20行
    print(f"\n  【前20行详细内容和格式】:")
    max_rows_to_check = min(20, ws.max_row)

    for row_idx in range(1, max_rows_to_check + 1):
        # 行高
        row_dim = ws.row_dimensions.get(row_idx)
        if row_dim:
            row_height = row_dim.height
            row_hidden = row_dim.hidden
        else:
            row_height = "默认(~15)"
            row_hidden = False

        print(f"\n  --- 第{row_idx}行 (行高={row_height}, 隐藏={row_hidden}) ---")

        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            value = cell.value

            # 字体信息
            font = cell.font
            font_name = font.name if font.name else "默认"
            font_size = font.size if font.size else "默认"
            font_bold = font.bold
            font_italic = font.italic
            font_color = font.color
            if font_color and font_color.rgb and font_color.rgb != "00000000":
                font_color_str = f"#{font_color.rgb}"
            else:
                font_color_str = "黑色/默认"

            # 背景色/填充
            fill = cell.fill
            if fill.fgColor and fill.fgColor.rgb and fill.fgColor.rgb not in ("00000000", "0"):
                bg_color = f"#{fill.fgColor.rgb}"
            else:
                bg_color = "无"
            fill_type = fill.fill_type if fill.fill_type else "无"

            # 对齐
            alignment = cell.alignment
            h_align = alignment.horizontal if alignment.horizontal else "默认"
            v_align = alignment.vertical if alignment.vertical else "默认"
            wrap_text = alignment.wrap_text

            # 边框
            border = cell.border
            border_parts = []
            for side_name in ['left', 'right', 'top', 'bottom']:
                side = getattr(border, side_name)
                if side and side.style:
                    border_parts.append(f"{side_name}={side.style}")
            border_str = ", ".join(border_parts) if border_parts else "无边框"

            # 数字格式
            number_format = cell.number_format if cell.number_format != "General" else "常规"

            if value is not None or border_str != "无边框":
                print(f"    列{col_idx}: 值=\"{value}\"")
                print(f"          字体={font_name}, 字号={font_size}, 粗体={font_bold}, 斜体={font_italic}, 字色={font_color_str}")
                print(f"          填充类型={fill_type}, 背景色={bg_color}")
                print(f"          对齐: 水平={h_align}, 垂直={v_align}, 自动换行={wrap_text}")
                print(f"          边框: {border_str}")
                print(f"          数字格式: {number_format}")

    # 检查所有行的行高异常
    print(f"\n  【行高异常检查】（正常范围大约12-30）:")
    abnormal_heights = []
    for row_idx in range(1, ws.max_row + 1):
        row_dim = ws.row_dimensions.get(row_idx)
        if row_dim and row_dim.height is not None:
            h = row_dim.height
            if h < 10 or h > 50:
                abnormal_heights.append((row_idx, h))
    if abnormal_heights:
        for r, h in abnormal_heights:
            cell_val = ws.cell(row=r, column=1).value
            print(f"    第{r}行: 行高={h} (A列值=\"{cell_val}\")")
    else:
        print(f"    无异常行高")


# ============================================================
# 第二部分：分析文件2（.xls格式，用xlrd）
# ============================================================

file2 = r"D:\广联达临时文件\2025\2025.2.27-算客工厂-深汕锐博特创新产业园-弱电工程量计算\20250226-深圳弱电项目清单\深汕锐博特创新产业园主体施工总承包工程-1#\深汕锐博特创新产业园主体施工总承包工程-1#电梯工程(单位工程).xls"

print("\n\n")
print("=" * 100)
print("【文件2】原始广联达清单文件分析")
print(f"路径：{file2}")
print("=" * 100)

wb2 = xlrd.open_workbook(file2, formatting_info=True)
print(f"\nSheet列表：{wb2.sheet_names()}")

for sheet_idx, sheet_name in enumerate(wb2.sheet_names()):
    ws2 = wb2.sheet_by_index(sheet_idx)
    print(f"\n{'─' * 100}")
    print(f"Sheet: {sheet_name}")
    print(f"行数: {ws2.nrows}, 列数: {ws2.ncols}")
    print(f"{'─' * 100}")

    # 合并单元格
    print(f"\n  【合并单元格】共 {len(ws2.merged_cells)} 个:")
    for rlo, rhi, clo, chi in ws2.merged_cells:
        print(f"    行{rlo+1}-{rhi}, 列{clo+1}-{chi} (即 {get_column_letter(clo+1)}{rlo+1}:{get_column_letter(chi)}{rhi})")

    # 列宽信息
    print(f"\n  【列宽信息】:")
    for col_idx in range(ws2.ncols):
        # xlrd中列宽单位是 1/256 字符宽度
        col_width_raw = ws2.colinfo_map.get(col_idx)
        if col_width_raw:
            width_chars = col_width_raw.width / 256.0
            hidden = col_width_raw.hidden
            print(f"    列{col_idx+1}({get_column_letter(col_idx+1)}): 宽度={width_chars:.2f}字符 (原始值={col_width_raw.width}), 隐藏={hidden}")
        else:
            print(f"    列{col_idx+1}({get_column_letter(col_idx+1)}): 宽度=默认")

    # 逐行分析前20行
    print(f"\n  【前20行详细内容和格式】:")
    max_rows_to_check = min(20, ws2.nrows)

    for row_idx in range(max_rows_to_check):
        # 行高 (xlrd中行高单位是 twips, 1 point = 20 twips)
        rowinfo = ws2.rowinfo_map.get(row_idx)
        if rowinfo:
            row_height_pt = rowinfo.height / 20.0
            row_hidden = rowinfo.hidden
        else:
            row_height_pt = "默认"
            row_hidden = False

        print(f"\n  --- 第{row_idx+1}行 (行高={row_height_pt}pt, 隐藏={row_hidden}) ---")

        for col_idx in range(ws2.ncols):
            cell = ws2.cell(row_idx, col_idx)
            value = cell.value
            cell_type = cell.ctype  # 0=空, 1=文本, 2=数字, 3=日期, 4=布尔, 5=错误

            type_names = {0: "空", 1: "文本", 2: "数字", 3: "日期", 4: "布尔", 5: "错误"}
            type_str = type_names.get(cell_type, f"未知({cell_type})")

            # 获取XF格式记录
            xf_index = cell.xf_index
            if xf_index is not None and xf_index < len(wb2.xf_list):
                xf = wb2.xf_list[xf_index]

                # 字体
                font_idx = xf.font_index
                if font_idx < len(wb2.font_list):
                    font = wb2.font_list[font_idx]
                    font_name = font.name
                    # xlrd中字号单位是 1/20 point
                    font_size = font.height / 20.0
                    font_bold = font.bold
                    font_italic = font.italic
                    font_color_idx = font.colour_index
                else:
                    font_name = "?"
                    font_size = "?"
                    font_bold = "?"
                    font_italic = "?"
                    font_color_idx = "?"

                # 对齐
                h_align_map = {0: "常规", 1: "左对齐", 2: "居中", 3: "右对齐", 4: "填充",
                               5: "两端对齐", 6: "跨列居中", 7: "分散对齐"}
                v_align_map = {0: "上对齐", 1: "居中", 2: "下对齐", 3: "两端对齐", 4: "分散对齐"}
                h_align = h_align_map.get(xf.alignment.hor_align, f"未知({xf.alignment.hor_align})")
                v_align = v_align_map.get(xf.alignment.vert_align, f"未知({xf.alignment.vert_align})")
                wrap_text = xf.alignment.text_wrapped

                # 背景色
                bg_pattern = xf.background.pattern_colour_index
                bg_color_idx = xf.background.background_colour_index

                # 边框
                border = xf.border
                border_style_map = {0: "无", 1: "细线", 2: "中线", 3: "虚线", 4: "点线",
                                    5: "粗线", 6: "双线", 7: "细虚线"}
                border_parts = []
                for side_name, style_val in [("top", border.top_line_style),
                                              ("bottom", border.bottom_line_style),
                                              ("left", border.left_line_style),
                                              ("right", border.right_line_style)]:
                    if style_val:
                        border_parts.append(f"{side_name}={border_style_map.get(style_val, style_val)}")
                border_str = ", ".join(border_parts) if border_parts else "无边框"

                if value != "" or border_str != "无边框":
                    print(f"    列{col_idx+1}: 值=\"{value}\" (类型={type_str})")
                    print(f"          字体={font_name}, 字号={font_size}pt, 粗体={font_bold}, 斜体={font_italic}, 字色索引={font_color_idx}")
                    print(f"          背景: 图案色索引={bg_pattern}, 背景色索引={bg_color_idx}")
                    print(f"          对齐: 水平={h_align}, 垂直={v_align}, 自动换行={wrap_text}")
                    print(f"          边框: {border_str}")
            else:
                if value != "":
                    print(f"    列{col_idx+1}: 值=\"{value}\" (类型={type_str}) [无格式信息]")

    # 检查所有行的行高
    print(f"\n  【行高分布统计】:")
    heights = {}
    for row_idx in range(ws2.nrows):
        rowinfo = ws2.rowinfo_map.get(row_idx)
        if rowinfo:
            h = round(rowinfo.height / 20.0, 1)
            if h not in heights:
                heights[h] = 0
            heights[h] += 1
    for h in sorted(heights.keys()):
        print(f"    行高 {h}pt: {heights[h]}行")


# ============================================================
# 第三部分：对比总结
# ============================================================

print("\n\n")
print("=" * 100)
print("【对比总结】")
print("=" * 100)

# 重新读取两个文件做对比
wb1 = openpyxl.load_workbook(file1)
wb2 = xlrd.open_workbook(file2, formatting_info=True)

for sheet_name in wb1.sheetnames:
    ws1 = wb1[sheet_name]

    # 尝试在文件2中找到对应sheet
    if sheet_name in wb2.sheet_names():
        ws2 = wb2.sheet_by_name(sheet_name)
    else:
        print(f"\n  Sheet \"{sheet_name}\" 在文件2中不存在，跳过对比")
        continue

    print(f"\n  Sheet: {sheet_name}")
    print(f"  文件1: {ws1.max_row}行 x {ws1.max_column}列")
    print(f"  文件2: {ws2.nrows}行 x {ws2.ncols}列")

    # 对比列宽
    print(f"\n  列宽对比:")
    max_cols = max(ws1.max_column, ws2.ncols)
    for col_idx in range(1, max_cols + 1):
        col_letter = get_column_letter(col_idx)

        # 文件1的列宽
        col_dim1 = ws1.column_dimensions.get(col_letter)
        w1 = col_dim1.width if col_dim1 and col_dim1.width else 8.43

        # 文件2的列宽
        col_info2 = ws2.colinfo_map.get(col_idx - 1)
        w2 = col_info2.width / 256.0 if col_info2 else 8.43

        diff_pct = abs(w1 - w2) / max(w1, w2, 1) * 100
        flag = " <<<< 差异较大" if diff_pct > 20 else ""
        print(f"    列{col_idx}({col_letter}): 文件1={w1:.2f}, 文件2={w2:.2f} (差异{diff_pct:.0f}%){flag}")

    # 对比行高
    print(f"\n  行高对比 (前20行):")
    for row_idx in range(1, min(21, ws1.max_row + 1, ws2.nrows + 1)):
        # 文件1行高
        rd1 = ws1.row_dimensions.get(row_idx)
        h1 = rd1.height if rd1 and rd1.height is not None else 15.0

        # 文件2行高
        ri2 = ws2.rowinfo_map.get(row_idx - 1)
        h2 = ri2.height / 20.0 if ri2 else 15.0

        diff = abs(h1 - h2)
        flag = " <<<< 差异较大" if diff > 3 else ""

        # 获取第一列的值作为参考
        v1 = ws1.cell(row=row_idx, column=1).value
        v2 = ws2.cell_value(row_idx - 1, 0) if row_idx - 1 < ws2.nrows else ""
        print(f"    第{row_idx}行: 文件1行高={h1}pt, 文件2行高={h2:.1f}pt (差{diff:.1f}pt){flag}  |  内容: \"{v1}\" vs \"{v2}\"")

    # 对比字体
    print(f"\n  字体对比 (前20行首个非空单元格):")
    for row_idx in range(1, min(21, ws1.max_row + 1, ws2.nrows + 1)):
        # 找文件1中第一个非空单元格
        for col_idx in range(1, ws1.max_column + 1):
            cell1 = ws1.cell(row=row_idx, column=col_idx)
            if cell1.value is not None:
                f1_name = cell1.font.name if cell1.font.name else "默认"
                f1_size = cell1.font.size if cell1.font.size else "默认"
                f1_bold = cell1.font.bold

                # 对应文件2
                if row_idx - 1 < ws2.nrows and col_idx - 1 < ws2.ncols:
                    cell2 = ws2.cell(row_idx - 1, col_idx - 1)
                    xf_idx = cell2.xf_index
                    if xf_idx is not None and xf_idx < len(wb2.xf_list):
                        xf = wb2.xf_list[xf_idx]
                        fi = xf.font_index
                        if fi < len(wb2.font_list):
                            f2 = wb2.font_list[fi]
                            f2_name = f2.name
                            f2_size = f2.height / 20.0
                            f2_bold = f2.bold
                        else:
                            f2_name = "?"
                            f2_size = "?"
                            f2_bold = "?"
                    else:
                        f2_name = "?"
                        f2_size = "?"
                        f2_bold = "?"
                else:
                    f2_name = "N/A"
                    f2_size = "N/A"
                    f2_bold = "N/A"

                font_diff = ""
                if f1_name != f2_name:
                    font_diff += f" <<<< 字体名不同"
                if str(f1_size) != str(f2_size):
                    font_diff += f" <<<< 字号不同"
                if f1_bold != f2_bold:
                    font_diff += f" <<<< 粗体不同"

                print(f"    第{row_idx}行列{col_idx}: 文件1=({f1_name},{f1_size}pt,粗体={f1_bold}) vs 文件2=({f2_name},{f2_size}pt,粗体={f2_bold}){font_diff}")
                break

    # 对比合并单元格
    print(f"\n  合并单元格对比:")
    merge1 = set(str(m) for m in ws1.merged_cells.ranges)
    merge2 = set()
    for rlo, rhi, clo, chi in ws2.merged_cells:
        merge2.add(f"{get_column_letter(clo+1)}{rlo+1}:{get_column_letter(chi)}{rhi}")

    only_in_1 = merge1 - merge2
    only_in_2 = merge2 - merge1
    common = merge1 & merge2

    print(f"    共同合并: {len(common)}个")
    if only_in_1:
        print(f"    仅文件1有: {only_in_1}")
    if only_in_2:
        print(f"    仅文件2有: {only_in_2}")
    if not only_in_1 and not only_in_2:
        print(f"    合并单元格完全一致")

print("\n\n===== 分析完成 =====")
