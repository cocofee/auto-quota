"""
反馈学习模块
功能：
1. 记录用户对匹配结果的修正（改了哪条清单的哪个定额）
2. 将修正结果自动存入经验库
3. 统计匹配成功率、各类别准确率
4. 从已完成项目的Excel文件导入经验

工作流程：
- 用户在输出Excel中修改匹配结果
- 系统读取修改后的Excel，对比原始匹配
- 发现有改动的条目 → 存入经验库（来源标记为 user_correction）
- 下次遇到相似清单 → 经验库直接命中 → 不再需要搜索+大模型
"""

import json
import re
import sqlite3
import time
from pathlib import Path

import openpyxl
from loguru import logger

import sys
sys.path.insert(0, str(Path(__file__).parent.parent))
import config
from src.experience_db import ExperienceDB


class FeedbackLearner:
    """处理用户反馈，驱动系统学习"""

    def __init__(self):
        self.experience_db = ExperienceDB()

    def learn_from_corrections(self, original_results: list[dict],
                                corrected_results: list[dict]) -> dict:
        """
        从用户修正中学习

        对比原始匹配结果和用户修正后的结果，把改动过的条目存入经验库

        参数:
            original_results: 系统原始匹配结果列表
            corrected_results: 用户修正后的结果列表

        返回:
            {"total": 总条数, "corrections": 修正条数, "confirmed": 确认条数}
        """
        pair_count = min(len(original_results), len(corrected_results))
        if len(original_results) != len(corrected_results):
            logger.warning(
                f"反馈学习输入长度不一致: original={len(original_results)} "
                f"corrected={len(corrected_results)}，仅处理前{pair_count}条"
            )

        stats = {"total": pair_count, "corrections": 0, "confirmed": 0}

        for idx in range(pair_count):
            orig = original_results[idx]
            corrected = corrected_results[idx]
            bill_item = orig.get("bill_item", {})
            from src.text_parser import normalize_bill_text
            bill_text = normalize_bill_text(bill_item.get('name', ''), bill_item.get('description', ''))

            if not bill_text:
                continue

            # 获取原始和修正后的定额列表
            orig_quotas = orig.get("quotas", [])
            corrected_quotas = corrected.get("quotas", [])

            orig_first_id = orig_quotas[0].get("quota_id") if orig_quotas else None
            corrected_first_id = corrected_quotas[0].get("quota_id") if corrected_quotas else None

            if not corrected_first_id:
                continue  # 修正后也没有定额，跳过

            # 收集所有定额编号和名称
            quota_ids = [q["quota_id"] for q in corrected_quotas if q.get("quota_id")]
            quota_names = [q.get("name", "") for q in corrected_quotas if q.get("quota_id")]

            if orig_first_id != corrected_first_id:
                # 用户改了定额 → 以用户修正为准，高置信度存入
                record_id = self.experience_db.add_experience(
                    bill_text=bill_text,
                    quota_ids=quota_ids,
                    quota_names=quota_names,
                    bill_name=bill_item.get("name"),
                    bill_code=bill_item.get("code"),
                    bill_unit=bill_item.get("unit"),
                    source="user_correction",
                    confidence=95,  # 用户修正置信度最高
                )
                if record_id > 0:
                    stats["corrections"] += 1
                else:
                    logger.warning(
                        f"反馈学习写入被拦截: {bill_text[:60]} -> {quota_ids}"
                    )
            else:
                # 用户没改 → 不等于用户确认了（可能根本没看这条）
                # 不存入经验库，避免错误结果通过"沉默确认"积累置信度
                stats["confirmed"] += 1

        logger.info(f"学习完成: 共{stats['total']}条, "
                    f"修正{stats['corrections']}条, 确认{stats['confirmed']}条")

        return stats

    def learn_from_corrected_excel(self, corrected_excel_path: str,
                                    original_results: list[dict] = None) -> dict:
        """
        从用户修正后的Excel文件中学习

        用户拿到输出的Excel后，可能会在Excel中修改定额编号。
        这个方法读取修改后的Excel，提取清单→定额的对应关系，存入经验库。

        参数:
            corrected_excel_path: 用户修正后的Excel文件路径
            original_results: 原始匹配结果（用于对比，可选）

        返回:
            {"total": 总条数, "learned": 学习条数}
        """
        path = Path(corrected_excel_path)
        if not path.exists():
            logger.error(f"文件不存在: {path}")
            return {"total": 0, "learned": 0}

        wb = openpyxl.load_workbook(str(path), data_only=True)

        stats = {"total": 0, "learned": 0}
        quota_id_pattern = re.compile(r'^[A-Za-z]?\d{1,2}-\d+')

        for ws in wb.worksheets:
            # 跳过辅助页
            if ws.title in {"待审核", "统计汇总"}:
                continue

            current_bill = None   # 当前正在处理的清单项
            current_quotas = []   # 当前清单项对应的定额列表

            for row in ws.iter_rows(min_row=1, values_only=True):
                cells = list(row) if row else []
                a = str(cells[0]).strip() if len(cells) > 0 and cells[0] is not None else ""
                b = str(cells[1]).strip() if len(cells) > 1 and cells[1] is not None else ""
                c = str(cells[2]).strip() if len(cells) > 2 and cells[2] is not None else ""
                d = str(cells[3]).strip() if len(cells) > 3 and cells[3] is not None else ""
                e = str(cells[4]).strip() if len(cells) > 4 and cells[4] is not None else ""

                is_labeled_bill = (a == "清单")
                is_labeled_quota = (a == "定额")
                # 兼容当前导出格式：A列为数字序号
                is_numbered_bill = (a.isdigit() and bool(c))
                # 兼容当前导出格式：A列空且B列为定额编号
                is_quota_row = bool(current_bill) and (
                    is_labeled_quota or ((not a) and bool(quota_id_pattern.match(b)))
                )

                if is_labeled_bill or is_numbered_bill:
                    # 如果之前有未保存的清单+定额对，先保存
                    if current_bill and current_quotas:
                        if self._save_bill_quota_pair(current_bill, current_quotas):
                            stats["learned"] += 1

                    # 开始新的清单项
                    if is_labeled_bill:
                        # 旧格式："清单"标记行（unit在第4列，desc在第5列）
                        current_bill = {
                            "name": c,
                            "code": b,
                            "unit": d,
                            "description": e,
                        }
                    else:
                        # 现格式：序号行（desc在第4列，unit在第5列）
                        current_bill = {
                            "name": c,
                            "code": b,
                            "unit": e,
                            "description": d,
                        }
                    current_quotas = []
                    stats["total"] += 1

                elif is_quota_row:
                    # 定额行，收集定额信息
                    quota_id = b
                    quota_name = c
                    if quota_id:
                        current_quotas.append({
                            "quota_id": quota_id,
                            "name": quota_name,
                        })

            # 保存最后一条
            if current_bill and current_quotas:
                if self._save_bill_quota_pair(current_bill, current_quotas):
                    stats["learned"] += 1

        wb.close()
        logger.info(f"从Excel学习完成: 共{stats['total']}条清单, 学习{stats['learned']}条")

        return stats

    def _save_bill_quota_pair(self, bill: dict, quotas: list[dict]) -> bool:
        """保存一条清单→定额的对应关系到经验库，返回是否写入成功。"""
        from src.text_parser import normalize_bill_text
        bill_text = normalize_bill_text(bill.get('name', ''), bill.get('description', ''))
        if not bill_text:
            return False

        quota_ids = [q["quota_id"] for q in quotas if q.get("quota_id")]
        quota_names = [q.get("name", "") for q in quotas]

        if not quota_ids:
            return False

        record_id = self.experience_db.add_experience(
            bill_text=bill_text,
            quota_ids=quota_ids,
            quota_names=quota_names,
            bill_name=bill.get("name"),
            bill_code=bill.get("code"),
            bill_unit=bill.get("unit"),
            source="user_correction",
            confidence=95,
        )
        if record_id <= 0:
            logger.warning(f"Excel学习写入被拦截: {bill_text[:60]} -> {quota_ids}")
            return False
        return True

    def import_completed_project(self, excel_path: str,
                                  project_name: str = None) -> dict:
        """
        从已完成项目的Excel导入经验

        支持的格式：小栗AI输出格式（清单行+定额行交替排列）
        也支持广联达导出格式（需要清单列和定额列）

        参数:
            excel_path: 已完成项目的Excel文件路径
            project_name: 项目名称

        返回:
            导入统计
        """
        # 用和 learn_from_corrected_excel 相同的逻辑读取
        # 但来源标记为 project_import
        path = Path(excel_path)
        if not path.exists():
            logger.error(f"文件不存在: {path}")
            return {"total": 0, "imported": 0}

        wb = openpyxl.load_workbook(str(path), data_only=True)

        records = []
        quota_id_pattern = re.compile(r'^[A-Za-z]?\d{1,2}-\d+')
        for ws in wb.worksheets:
            if ws.title in {"待审核", "统计汇总"}:
                continue
            current_bill = None
            current_quotas = []

            for row in ws.iter_rows(min_row=1, values_only=True):
                if not row or not any(row):
                    continue

                # 尝试识别行类型
                # 兼容两类格式：
                # 1) 旧标签格式：A列为“清单/定额”
                # 2) 当前导出格式：清单行A列为数字序号，定额行A列为空且B列为定额编号
                first_cell = str(row[0]).strip() if row[0] is not None else ""
                code_cell = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
                name_cell = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""

                is_labeled_bill = (first_cell == "清单")
                is_labeled_quota = (first_cell == "定额")
                is_numbered_bill = first_cell.isdigit() and bool(name_cell)
                is_quota_row = bool(current_bill) and (
                    is_labeled_quota or ((not first_cell) and bool(quota_id_pattern.match(code_cell)))
                )

                if is_labeled_bill or is_numbered_bill:
                    # 清单行
                    if current_bill and current_quotas:
                        records.append({
                            "bill_text": f"{current_bill.get('name', '')} {current_bill.get('description', '')}".strip(),
                            "bill_name": current_bill.get("name"),
                            "bill_code": current_bill.get("code"),
                            "bill_unit": current_bill.get("unit"),
                            "quota_ids": [q["quota_id"] for q in current_quotas],
                            "quota_names": [q.get("name", "") for q in current_quotas],
                        })

                    if is_labeled_bill:
                        # 旧格式："清单"标记行（unit在第4列，desc在第5列）
                        current_bill = {
                            "name": name_cell,
                            "code": code_cell,
                            "unit": str(row[3]).strip() if len(row) > 3 and row[3] is not None else "",
                            "description": str(row[4]).strip() if len(row) > 4 and row[4] is not None else "",
                        }
                    else:
                        # 现格式：序号行（desc在第4列，unit在第5列）
                        current_bill = {
                            "name": name_cell,
                            "code": code_cell,
                            "unit": str(row[4]).strip() if len(row) > 4 and row[4] is not None else "",
                            "description": str(row[3]).strip() if len(row) > 3 and row[3] is not None else "",
                        }
                    current_quotas = []

                elif is_quota_row:
                    # 定额行
                    quota_id = code_cell
                    quota_name = name_cell
                    if quota_id:
                        current_quotas.append({
                            "quota_id": quota_id,
                            "name": quota_name,
                        })

            # 保存最后一条
            if current_bill and current_quotas:
                records.append({
                    "bill_text": f"{current_bill.get('name', '')} {current_bill.get('description', '')}".strip(),
                    "bill_name": current_bill.get("name"),
                    "bill_code": current_bill.get("code"),
                    "bill_unit": current_bill.get("unit"),
                    "quota_ids": [q["quota_id"] for q in current_quotas],
                    "quota_names": [q.get("name", "") for q in current_quotas],
                })

        wb.close()

        if not records:
            logger.warning("未从Excel中识别到清单→定额对应关系")
            return {"total": 0, "imported": 0}

        # 批量导入经验库
        result = self.experience_db.import_from_project(
            records, project_name=project_name or path.stem
        )

        return result

    def get_accuracy_stats(self) -> dict:
        """
        统计匹配准确率

        返回:
            {"total_matches": 总匹配数,
             "user_corrections": 用户修正数,
             "accuracy_rate": 准确率(百分比)}
        """
        conn = sqlite3.connect(str(self.experience_db.db_path), timeout=10)
        conn.execute("PRAGMA busy_timeout=5000")
        try:
            cursor = conn.cursor()

            # 总经验数
            cursor.execute("SELECT COUNT(*) FROM experiences")
            total = cursor.fetchone()[0]

            # 用户修正数（说明系统当时匹配错了）
            cursor.execute("SELECT COUNT(*) FROM experiences WHERE source = 'user_correction'")
            corrections = cursor.fetchone()[0]

            # 用户确认数（说明系统匹配对了）
            # 兼容旧数据：历史上可能写过 auto_match
            cursor.execute("""
                SELECT COUNT(*)
                FROM experiences
                WHERE source IN ('user_confirmed', 'auto_match')
            """)
            user_confirmed = cursor.fetchone()[0]
        finally:
            conn.close()

        # 准确率 = 用户确认数 / (用户确认数 + 用户修正数)
        denominator = user_confirmed + corrections
        accuracy = (user_confirmed / denominator * 100) if denominator > 0 else 0

        return {
            "total_experiences": total,
            "auto_confirmed": user_confirmed,  # 兼容旧字段名
            "user_confirmed": user_confirmed,
            "user_corrections": corrections,
            "accuracy_rate": round(accuracy, 1),
        }


# 模块级单例
learner = FeedbackLearner()


# ================================================================
# 命令行入口
# ================================================================

if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="反馈学习模块")
    parser.add_argument("--import-project", help="从已完成项目Excel导入经验")
    parser.add_argument("--learn-excel", help="从修正后的Excel学习")
    parser.add_argument("--stats", action="store_true", help="查看准确率统计")

    args = parser.parse_args()

    fl = FeedbackLearner()

    if args.import_project:
        result = fl.import_completed_project(args.import_project)
        print(f"导入结果: {result}")
    elif args.learn_excel:
        result = fl.learn_from_corrected_excel(args.learn_excel)
        print(f"学习结果: {result}")
    elif args.stats:
        stats = fl.get_accuracy_stats()
        print("=" * 40)
        print("匹配准确率统计")
        print("=" * 40)
        print(f"  总经验数: {stats['total_experiences']}")
        print(f"  系统确认: {stats['auto_confirmed']}")
        print(f"  用户修正: {stats['user_corrections']}")
        print(f"  准确率: {stats['accuracy_rate']}%")
    else:
        parser.print_help()
