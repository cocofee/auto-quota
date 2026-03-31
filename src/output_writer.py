"""
结果输出模块
功能：
1. 读取原始清单Excel，保留完整结构（表头、分部、小节标题）
2. 在每条清单行下面插入匹配到的子目行（定额行）
3. 跳过原始文件中已有的定额行（用我们的匹配结果替换）
4. 格式匹配广联达标准，可直接导入

广联达识别规则：
  - A列有序号(数字) → 清单行
  - A列为空，B列有C开头编号 → 子目行（定额行）
  - A列和B列都为空，C列有文字 → 分节行
"""

import re
import sys
import shutil
import os
import uuid
import tempfile
from pathlib import Path
from datetime import datetime
from typing import Optional

import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from loguru import logger

import config
from src.bill_reader import _is_material_code
from src.excel_compat import convert_excel_to_xlsx


# ======== 主材查价（懒加载，整个输出过程只初始化一次）========

_material_db_instance = None  # 缓存MaterialDB实例，避免每行都新建


def _get_material_price(name: str, spec: str = "",
                        unit: str = "") -> Optional[float]:
    """查主材单价（从价格库）

    返回含税单价（已按目标单位换算），查不到返回None。
    省份自动从当前运行省份提取。
    unit: 主材行的单位（如'm'），用于单位换算（价格库可能是吨价）。
    """
    global _material_db_instance
    if _material_db_instance is None:
        try:
            from src.material_db import MaterialDB
            _material_db_instance = MaterialDB()
        except Exception:
            return None

    # 从定额库长名称提取短省份名（如"北京市建设工程..."→"北京"）
    province_short = _extract_short_province(config.get_current_province())

    # 先用原始name+spec查
    result = _material_db_instance.search_price_by_name(
        name, province=province_short, spec=spec, target_unit=unit)
    if result:
        return result.get("price")

    # 查不到时，尝试从name里拆出规格（如"不锈钢管 DN100"→name="不锈钢管", spec="DN100"）
    if not spec:
        m = re.search(r'[Dd][Nn]\s*\d+', name)
        if m:
            extracted_spec = m.group(0).replace(" ", "")
            short_name = name[:m.start()].strip()
            if short_name:
                result2 = _material_db_instance.search_price_by_name(
                    short_name, province=province_short, spec=extracted_spec,
                    target_unit=unit)
                if result2:
                    return result2.get("price")

    return None


def _extract_short_province(full_name: str) -> str:
    """从定额库全名提取短省份名（用于匹配价格库的province字段）

    例：'北京市建设工程施工消耗量标准(2024)' → '北京'
        '湖北省房屋建筑与装饰工程...' → '湖北'
        '上海市建筑和装饰工程...' → '上海'
    """
    if not full_name:
        return ""
    # 常见省份名（2-3字），取全名开头匹配
    provinces = [
        "黑龙江", "内蒙古",
        "北京", "上海", "天津", "重庆",
        "河北", "山西", "辽宁", "吉林", "江苏", "浙江", "安徽",
        "福建", "江西", "山东", "河南", "湖北", "湖南", "广东",
        "广西", "海南", "四川", "贵州", "云南", "西藏", "陕西",
        "甘肃", "青海", "宁夏", "新疆",
    ]
    for p in provinces:
        if full_name.startswith(p):
            return p
    return ""


# 颜色定义
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
GRAY_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
LIGHT_BLUE_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

# 字体（和广联达一致：宋体9号，全表统一）
HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
BILL_FONT = Font(name="宋体", size=9)
# 定额行字体：宋体9号，和广联达导出格式一致
GLD_FONT = Font(name="宋体", size=9)

# 边框
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# 标准列宽（参考广联达格式，配合 wrap_text 使用）
STANDARD_COL_WIDTHS = {
    "A": 5,     # 序号
    "B": 13,    # 项目编码/定额编号
    "C": 20,    # 项目名称
    "D": 30,    # 项目特征描述/定额名称
    "E": 6,     # 计量单位
    "F": 10,    # 工程量
    "G": 10,    # 综合单价
    "H": 10,    # 合价
    "I": 10,    # 暂估价
}


# ================================================================
# 单位换算（清单单位 ≠ 定额单位时，自动转换工程量）
# ================================================================

# 单位换算系数表：(清单单位, 定额单位) → 乘以系数
UNIT_CONVERSIONS = {}


def _build_unit_conversions():
    """构建双向单位换算表"""
    # (单位A, 单位B, A→B的系数)
    pairs = [
        ("t", "kg", 1000),           # 吨 → 千克
        ("kg", "t", 0.001),          # 千克 → 吨
        ("km", "m", 1000),           # 千米 → 米
        ("m", "km", 0.001),          # 米 → 千米
        ("100m", "m", 100),          # 百米 → 米
        ("m", "100m", 0.01),         # 米 → 百米
        ("m²", "100m²", 0.01),       # 平方米 → 百平方米
        ("100m²", "m²", 100),        # 百平方米 → 平方米
        ("m³", "10m³", 0.1),         # 立方米 → 10立方米
        ("10m³", "m³", 10),          # 10立方米 → 立方米
        ("m³", "100m³", 0.01),       # 立方米 → 百立方米
        ("100m³", "m³", 100),        # 百立方米 → 立方米
    ]
    for u1, u2, factor in pairs:
        UNIT_CONVERSIONS[(u1, u2)] = factor


_build_unit_conversions()


def _safe_confidence(value, default: int = 0) -> int:
    """把置信度安全转换为0-100整数。"""
    try:
        conf = int(value)
    except (TypeError, ValueError):
        return default
    return max(0, min(100, conf))


def _ensure_list(value) -> list:
    """把任意输入收敛为list，避免脏类型触发导出异常。"""
    return value if isinstance(value, list) else []


def _safe_write_cell(ws, row: int, column: int, value=None):
    """安全写入单元格值，遇到合并单元格时跳过。

    当目标格是MergedCell（合并区域中非左上主格）时，openpyxl禁止写入
    并抛 AttributeError。本函数检测到MergedCell后静默跳过，避免整次输出崩溃。

    返回:
        cell对象（可继续设置样式），或 None（合并单元格，已跳过）
    """
    cell = ws.cell(row=row, column=column)
    if isinstance(cell, MergedCell):
        return None
    if value is not None:
        cell.value = value
    return cell


def _is_bill_serial(a_val) -> bool:
    """判断A列是否为清单序号（兼容int/float/字符串）。"""
    if a_val is None:
        return False
    if isinstance(a_val, int):
        return a_val > 0
    if isinstance(a_val, float):
        return a_val > 0 and a_val.is_integer()
    text = str(a_val).strip()
    if not text:
        return False
    if text.isdigit():
        return True
    return bool(re.fullmatch(r"\d+\.0+", text))


def _resolve_output_materials(result: dict) -> list[dict]:
    """获取要输出的主材行列表

    优先用输入文件中提取的主材（source_materials）；
    没有时从清单特征描述里提取主材名称（兜底），让广联达导入时能带上主材。
    返回: [{code, name, unit, qty}, ...]
    """
    bill_item = result.get("bill_item", {})
    # 优先用原文件里已有的主材
    source_mats = bill_item.get("source_materials")
    if isinstance(source_mats, list) and source_mats:
        return source_mats

    # 措施项不加主材
    if result.get("match_source") == "skip_measure":
        return []

    # 没有匹配定额的也不加主材
    quotas = result.get("quotas")
    if not quotas:
        return []

    # 兜底：从清单特征描述里提取主材名称
    material_name = OutputWriter._extract_material_from_description(bill_item)
    if material_name:
        return [{"code": "", "name": material_name,
                 "unit": bill_item.get("unit", ""),
                 "qty": bill_item.get("quantity")}]
    return []


def _is_quota_code(code: str) -> bool:
    """判断是否是定额编号格式（支持 X-XXX / D00003 / 带'换'后缀）。"""
    if not isinstance(code, str):
        return False
    c = code.strip()
    if not c:
        return False
    core = c[:-1] if c.endswith("换") else c
    # 标准定额编号（C10-1-296）、字母前缀+数字（D00003）、措施费纯数字（991305009）
    return (bool(re.match(r'^[A-Za-z]?\d{1,2}-\d+', core))
            or bool(re.match(r'^[A-Za-z]{1,2}\d{4,}$', core))
            or bool(re.fullmatch(r'\d{9,}', core)))


def convert_quantity(bill_qty, bill_unit: str, quota_unit: str):
    """
    单位换算：当清单单位和定额单位不同时，转换工程量

    例如：清单 5t → 定额 kg → 返回 5000
    大多数情况单位相同，直接返回原值。
    """
    if bill_qty is None:
        # 清单工程量缺失时保持空值，避免在结果里误写 0
        return None

    # 归一化工程量为数值，避免字符串数量在换算时触发类型错误
    qty = bill_qty
    if isinstance(qty, str):
        q = qty.strip().replace(",", "")
        if q == "":
            return None
        try:
            qty = float(q)
        except ValueError:
            return bill_qty
    elif not isinstance(qty, (int, float)):
        try:
            qty = float(qty)
        except Exception:
            return bill_qty

    if not bill_unit or not quota_unit:
        return qty

    # 标准化单位文本（统一小写，处理特殊Unicode字符）
    bu = bill_unit.strip().lower().replace("㎡", "m²").replace("㎥", "m³")
    qu = quota_unit.strip().lower().replace("㎡", "m²").replace("㎥", "m³")

    if bu == qu:
        return qty

    factor = UNIT_CONVERSIONS.get((bu, qu))
    if factor:
        converted = round(qty * factor, 4)
        logger.debug(f"单位换算: {qty}{bill_unit} → {converted}{quota_unit} (×{factor})")
        return converted

    # 没有找到换算关系，原样返回（大多数情况单位相同）
    return qty


# ================================================================
# 置信度显示
# ================================================================

def confidence_to_stars(confidence: int, has_quotas: bool) -> str:
    """
    把置信度百分比转换为星级推荐展示

    规则：
    - ≥85% → ★★★推荐(95%)  绿色，基本可信
    - 60-84% → ★★参考(72%)  黄色，建议人工确认
    - <60% → ★待审(45%)     红色，需要人工处理
    - 无匹配 → —

    参数:
        confidence: 置信度百分比（0-100）
        has_quotas: 是否有匹配到定额

    返回:
        星级文字（如 "★★★推荐(85%)"）
    """
    if not has_quotas:
        return "—"
    if confidence >= config.CONFIDENCE_GREEN:
        return f"★★★推荐({confidence}%)"
    elif confidence >= config.CONFIDENCE_YELLOW:
        return f"★★参考({confidence}%)"
    else:
        return f"★待审({confidence}%)"


def safe_excel_text(value):
    """防止Excel公式注入：文本以 = + - @ 开头时前置单引号。"""
    if value is None:
        return value
    if not isinstance(value, str):
        return value
    text = value.replace("\x00", "")
    if text.lstrip().startswith(("=", "+", "-", "@")):
        return "'" + text
    return text


class OutputWriter:
    """匹配结果Excel输出，保留原始清单结构"""

    @staticmethod
    def _convert_xls_for_output(xls_path: str, output_xlsx_path: str):
        """将 .xls 文件转换为 .xlsx 格式（用于输出保留结构）"""
        result = convert_excel_to_xlsx(
            xls_path,
            output_xlsx_path,
            prefer_preserve_format=True,
        )
        if not result.preserved_formatting:
            logger.warning(
                f".xls 转换已降级为值写入模式: method={result.method}; "
                f"warning={result.warning or 'n/a'}"
            )
        return result

    @staticmethod
    def _save_workbook_atomic(wb, output_path: str):
        """原子写入Excel，避免中断时留下半成品文件。

        Windows下 os.replace() 偶尔被杀毒软件/文件索引锁住，加重试机制。
        """
        import time
        out_path = Path(output_path)
        out_path.parent.mkdir(parents=True, exist_ok=True)
        tmp_path = None
        try:
            with tempfile.NamedTemporaryFile(
                suffix=".xlsx",
                prefix=f"{out_path.stem}_tmp_",
                dir=str(out_path.parent),
                delete=False,
            ) as tf:
                tmp_path = tf.name
            wb.save(tmp_path)
            # Windows下 os.replace 偶尔被杀毒/索引服务短暂锁定，重试3次
            for attempt in range(3):
                try:
                    os.replace(tmp_path, out_path)
                    tmp_path = None  # 成功，不需要清理
                    break
                except PermissionError:
                    if attempt < 2:
                        time.sleep(0.5)
                    else:
                        # 最后一次：放弃原子写入，先删目标再移动
                        logger.warning(
                            f"os.replace 权限错误，尝试先删后移: {tmp_path}")
                        try:
                            out_path.unlink(missing_ok=True)
                        except OSError:
                            pass
                        shutil.copy2(tmp_path, str(out_path))
                        tmp_path = None
        finally:
            if tmp_path and Path(tmp_path).exists():
                try:
                    os.remove(tmp_path)
                except OSError:
                    pass

    @staticmethod
    def _confidence_fill(confidence: float):
        if confidence >= config.CONFIDENCE_GREEN:
            return GREEN_FILL
        if confidence >= config.CONFIDENCE_YELLOW:
            return YELLOW_FILL
        return RED_FILL

    @staticmethod
    def _write_alternative_cells(ws, row_idx: int, start_col: int, alternatives):
        for alt_idx, alt in enumerate(_ensure_list(alternatives)[:3]):
            alt_col = start_col + alt_idx
            alt_text = safe_excel_text(f"{alt.get('quota_id', '')} {alt.get('name', '')}")
            cell_alt = _safe_write_cell(ws, row_idx, alt_col, alt_text)
            if cell_alt:
                cell_alt.font = BILL_FONT  # 和正文统一字体
                cell_alt.border = THIN_BORDER

    @staticmethod
    def _brief_explanation(explanation: str) -> str:
        return safe_excel_text(explanation[:80] if explanation else "")

    @staticmethod
    def _brief_knowledge_evidence(result: dict) -> str:
        evidence = result.get("knowledge_evidence")
        if not isinstance(evidence, dict):
            return ""
        parts = []
        if evidence.get("quota_rules"):
            parts.append(f"规则{len(evidence.get('quota_rules') or [])}")
        if evidence.get("quota_explanations"):
            parts.append(f"解释{len(evidence.get('quota_explanations') or [])}")
        if evidence.get("method_cards"):
            parts.append(f"方法卡{len(evidence.get('method_cards') or [])}")
        if not parts:
            return ""
        return f"[知识]{'/'.join(parts)}"

    @staticmethod
    def _check_review_needed(confidence: int, quotas: list,
                             match_source: str) -> bool:
        """判断该条清单是否需要人工复核

        标记条件（任一命中即标记）：
        - 无匹配结果（定额为空）
        - 置信度低于85%（★★参考 或 ★待审）
        - 降级结果（agent_fallback / agent_error）
        - 无经验库命中的纯搜索结果（首次出现的清单写法）
        """
        if not quotas:
            return True
        if confidence < config.CONFIDENCE_GREEN:
            return True
        if match_source in ("agent_fallback", "agent_error", "agent_circuit_break"):
            return True
        return False

    @staticmethod
    def _brief_materials(result: dict, max_items: int = 4) -> str:
        """把结果中的主材列表压缩成可读短文本。"""
        materials = result.get("materials")
        if not isinstance(materials, list) or not materials:
            return ""

        parts = []
        for mat in materials:
            if not isinstance(mat, dict):
                continue
            name = str(mat.get("name", "")).strip()
            if not name:
                continue
            unit = str(mat.get("unit", "")).strip()
            price = mat.get("price", None)
            if price is None:
                text = f"{name}({unit})" if unit else name
            else:
                try:
                    p = f"{float(price):g}"
                except Exception:
                    p = str(price)
                text = f"{name}({p}元/{unit})" if unit else f"{name}({p}元)"
            parts.append(text)
            if len(parts) >= max_items:
                break
        if not parts:
            return ""
        more = len(materials) - len(parts)
        suffix = f" 等{len(materials)}项" if more > 0 else ""
        return safe_excel_text("; ".join(parts) + suffix)

    # 噪声词黑名单：这些文本出现在描述字段值中时，不是有效的主材名称
    _MATERIAL_NOISE_WORDS = {
        "详见图纸", "同清单", "见附件", "按设计", "按图纸", "见图纸",
        "见设计", "详见设计", "按规范", "见规范", "暂定", "待定",
        "按实际", "详见", "同上", "以上", "以下",
    }

    # 描述字段中可能包含主材名称的标签（按优先级排列）
    _MATERIAL_FIELD_KEYS = ("名称", "主材", "设备名称", "材质、规格",
                            "材质,规格", "规格型号", "材质", "规格", "类型")

    @staticmethod
    def _extract_material_from_description(bill_item: dict) -> str:
        """从清单特征描述中提取主材名称（经验库无主材时的兜底）

        安装定额的主材就是被安装的物品本身，名称藏在清单的项目特征描述里。
        例：清单"开关安装"，特征"1.名称:单联双控开关" → 主材是"单联双控开关"
        例：清单"管道安装"，特征"1.材质、规格:PSP钢塑复合管DN20" → 主材是"PSP钢塑复合管DN20"
        """
        if not isinstance(bill_item, dict):
            return ""
        description = bill_item.get("description", "") or ""
        if not description:
            return ""

        # ---- 第1步：从描述中提取 {标签: 值} 字典 ----
        # 内联简化版，不依赖query_builder的私有函数
        fields = {}
        # 有序号格式：1.名称:XXX / 2.规格:YYY
        for m in re.finditer(r'\d+[.、．]\s*([^:：\n]+)[：:]\s*([^\n]*)', description):
            key = m.group(1).strip()
            val = m.group(2).strip()
            if key and val:
                fields[key] = val
        # 无序号格式兜底
        if not fields:
            for m in re.finditer(r'([^:：\n]{2,8})[：:]\s*([^\n]+)', description):
                key = m.group(1).strip()
                val = m.group(2).strip()
                if key and val:
                    fields[key] = val

        if not fields:
            return ""

        # ---- 第2步：按优先级查找主材名称 ----
        # 容错查找：字段key可能含清单名碎片前缀（如"钢阀门 名称"而非"名称"）
        def _find_field(target: str) -> str:
            if target in fields:
                return fields[target]
            for k, v in fields.items():
                if k.endswith(target) or target in k:
                    return v
            return ""

        result_text = ""
        for field_key in OutputWriter._MATERIAL_FIELD_KEYS:
            val = _find_field(field_key)
            if not val or len(val) < 2:
                continue

            # 噪声词过滤
            if any(noise in val for noise in OutputWriter._MATERIAL_NOISE_WORDS):
                continue

            # 纯型号过滤（中文字符不到1/3→大概率是型号如"APE-Z"、"XZP100"）
            # 但"材质、规格"/"材质"等强信号字段跳过此检查——
            # PPR冷水管DN15、PVC排水管等中英混合名称是有效主材，不能误杀
            if field_key not in ("材质、规格", "材质,规格", "材质"):
                chinese_count = sum(1 for c in val if '\u4e00' <= c <= '\u9fff')
                if chinese_count < len(val.strip()) / 3:
                    continue

            # 名称/主材/设备名称等字段，尝试拼上规格（规格类型/规格型号/规格）
            if field_key in ("名称", "主材", "设备名称", "材质"):
                for spec_key in ("规格类型", "规格型号", "规格"):
                    spec_val = _find_field(spec_key)
                    if spec_val and spec_val not in val:
                        val = f"{val} {spec_val}"
                        break

            result_text = val.strip()
            break

        if not result_text:
            return ""

        # ---- 第3步：清理和截断 ----
        # 截断过长文本（主材名称一般不超过40字）
        if len(result_text) > 40:
            result_text = result_text[:40]

        return safe_excel_text(result_text)

    @staticmethod
    def _get_material_text(result: dict) -> str:
        """获取主材文本：优先用输入文件主材，其次经验库，最后从清单描述提取

        统一入口，三处O列写入点都调用这个方法。
        """
        # 优先用输入文件中提取的主材（source_materials），与主材行口径一致
        bill_item = result.get("bill_item", {})
        source_mats = bill_item.get("source_materials")
        if isinstance(source_mats, list) and source_mats:
            parts = [m.get("name", "") for m in source_mats if m.get("name")]
            if parts:
                return safe_excel_text("; ".join(parts[:4]) +
                                       (f" 等{len(parts)}项" if len(parts) > 4 else ""))
        # 经验库有主材（有编号、单位、价格，更完整）→ 次优先
        text = OutputWriter._brief_materials(result)
        if text:
            return text
        # 措施项不填主材
        if result.get("match_source") == "skip_measure":
            return ""
        # 没有匹配结果的也不填
        quotas = result.get("quotas", [])
        if not quotas:
            return ""
        # 从清单描述提取主材名称
        bill_item = result.get("bill_item", {})
        return OutputWriter._extract_material_from_description(bill_item)

    @staticmethod
    def _write_no_match_row(ws, row_idx: int, no_reason: str, max_col: int):
        _safe_write_cell(ws, row_idx, 3, safe_excel_text(f"未匹配: {no_reason}"))
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col)
            if isinstance(cell, MergedCell):
                continue
            cell.font = Font(name="宋体", size=9, color="FF0000")
            cell.fill = RED_FILL
            cell.border = THIN_BORDER

    @staticmethod
    def _apply_row_style(ws, row_idx: int, start_col: int, end_col: int, wrap_cols: set[int]):
        for col_idx in range(start_col, end_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if isinstance(cell, MergedCell):
                continue
            if cell.font == Font():
                cell.font = BILL_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(
                vertical="center",
                wrap_text=(col_idx in wrap_cols),
            )

    @staticmethod
    def _set_header_cell(ws, row_idx: int, col_idx: int, value, fill):
        cell = _safe_write_cell(ws, row_idx, col_idx, value)
        if cell is None:
            return None  # 合并单元格，跳过
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        return cell

    def write_results(self, results: list[dict], output_path: str = None,
                      original_file: str = None) -> str:
        """
        将匹配结果写入Excel

        策略：
        - 有原始文件 → 保留原始结构（复制原文件，在每个Sheet的清单行下插入定额行）
        - 无原始文件 → 新建工作簿（兜底模式）

        关键约束：
        - 原始Excel有几个Sheet输出就有几个Sheet，不能删减
        - 清单行的顺序不能改变（改变顺序 = 废标）
        - 每个Sheet独立处理，不跨Sheet合并
        """
        if not output_path:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = str(
                config.OUTPUT_DIR / f"匹配结果_{timestamp}_{uuid.uuid4().hex[:6]}.xlsx"
            )

        Path(output_path).parent.mkdir(parents=True, exist_ok=True)

        if original_file and Path(original_file).exists():
            # 主模式：保留原始文件的完整结构
            return self._write_preserve_structure(results, output_path, original_file)
        else:
            # 兜底模式：新建工作簿
            return self._write_new_workbook(results, output_path)

    # ================================================================
    # 主模式：保留原始文件完整结构
    # ================================================================

    def _write_preserve_structure(self, results, output_path, original_file):
        """
        保留原始Excel完整结构，在每个Sheet的清单行下方插入定额行

        逻辑：
        1. 复制原始文件到输出路径（保留所有Sheet、格式、合并单元格）
        2. 按sheet_name分组匹配结果
        3. 对每个有匹配结果的Sheet，在清单行下方插入定额行
        4. 没有匹配结果的Sheet原样保留不动
        5. 追加"待审核"和"统计汇总"Sheet
        """
        # 按sheet_name分组结果（保持组内顺序）
        results_by_sheet = {}
        for r in results:
            sheet = r.get("bill_item", {}).get("sheet_name", "")
            if sheet:
                results_by_sheet.setdefault(sheet, []).append(r)

        # 加载原始文件到内存（用BytesIO避免Windows文件锁问题）
        # Windows下 openpyxl.load_workbook(path) 会锁住文件，导致后续 save 到
        # 同一路径时 PermissionError。所以先读到内存再加载。
        import io
        from src.bill_reader import is_xls_format
        if is_xls_format(original_file):
            # .xls 需要先转换为 .xlsx 格式
            self._convert_xls_for_output(original_file, output_path)
            with open(output_path, "rb") as f:
                file_bytes = io.BytesIO(f.read())
        else:
            with open(original_file, "rb") as f:
                file_bytes = io.BytesIO(f.read())

        wb = openpyxl.load_workbook(file_bytes)
        try:
            # 逐个处理有匹配结果的Sheet
            processed_sheets = 0
            for sheet_name in wb.sheetnames:
                if sheet_name not in results_by_sheet:
                    continue  # 不是清单Sheet，原样保留
                ws = wb[sheet_name]
                sheet_results = results_by_sheet[sheet_name]
                self._process_bill_sheet(ws, sheet_results)
                processed_sheets += 1

            # 追加辅助Sheet（追加到最后，不影响原有Sheet顺序）
            ws_review = wb.create_sheet("待审核")
            self._write_review_sheet(ws_review, results)

            ws_stats = wb.create_sheet("统计汇总")
            self._write_stats_sheet(ws_stats, results)

            # 保存到目标路径（wb从内存加载，不锁任何磁盘文件）
            self._save_workbook_atomic(wb, output_path)
        finally:
            wb.close()
        logger.info(
            f"匹配结果已保存（保留原始{processed_sheets}个清单Sheet，"
            f"共{len(wb.sheetnames)}个Sheet）: {output_path}")
        return output_path

    def _process_bill_sheet(self, ws, results: list[dict]):
        """
        处理单个Sheet：在清单行下方插入定额行

        处理步骤：
        0. 取消数据区域合并单元格（防止insert_rows导致数据丢失）
        1. 找到表头行
        2. 删除已有的定额行（如果有的话）
        3. 重新扫描找到所有清单行
        4. 从下往上插入定额行（避免行号偏移）
        5. 添加推荐度/备选列标题
        """
        # 找表头行
        header_row = self._find_header_row_in_ws(ws)
        layout = self._detect_bill_layout(ws, header_row)
        unit_col = layout["unit_col"]
        qty_col = layout["qty_col"]

        # 第0步：取消数据区域合并单元格（防止分页格式下insert_rows错位）
        # 返回原始合并范围，插入完定额行后恢复
        saved_merges = self._unmerge_data_area(ws, header_row)

        # 计算匹配结果列的起始位置（追加到原表最后一列之后，不覆盖原始数据）
        # 保底不小于10（J列），避免原表只有5列时结果列和数据混在一起
        extra_start = max(ws.max_column + 1, 10)

        # 第1步：删除已有定额行（如果原文件中有旧的定额行）
        self._remove_existing_quota_rows(ws, header_row)

        # 第2步：扫描所有清单行（A列是纯数字序号的行）
        bill_rows = []
        for row_idx in range(header_row + 1, ws.max_row + 1):
            a_val = ws.cell(row=row_idx, column=1).value
            if _is_bill_serial(a_val):
                bill_rows.append(row_idx)

        # Step 3: map each result back to the original bill row.
        # Prefer source_row because it is the absolute Excel row number.
        # sheet_bill_seq is only a sheet-local sequence and may become subset-relative.
        # For filtered exports, source_row is the safest locator when present.
        row_result_pairs = []
        used_rows = set()
        bill_row_set = set(bill_rows)
        can_use_source_row_map = True
        for result in results:
            bill_item = result.get("bill_item", {})
            source_row = bill_item.get("source_row")
            if not isinstance(source_row, int):
                can_use_source_row_map = False
                break
            if source_row not in bill_row_set or source_row in used_rows:
                can_use_source_row_map = False
                break
            used_rows.add(source_row)
            row_result_pairs.append((source_row, result))

        can_use_seq_map = False
        if not can_use_source_row_map:
            row_result_pairs = []
            used_seq = set()
            can_use_seq_map = True
            for result in results:
                bill_item = result.get("bill_item", {})
                seq = bill_item.get("sheet_bill_seq")
                if not isinstance(seq, int):
                    can_use_seq_map = False
                    break
                if seq <= 0 or seq > len(bill_rows) or seq in used_seq:
                    can_use_seq_map = False
                    break
                used_seq.add(seq)
                row_result_pairs.append((bill_rows[seq - 1], result))

        # Fallbacks: bill_code exact match, then positional match as last resort.
        if not can_use_source_row_map and not can_use_seq_map:
            # Fallback A: match by bill_code from column B.
            code_pairs = []
            if results:
                # Build {bill_code -> row_idx} from the Excel sheet.
                row_code_map = {}
                for row_idx in bill_rows:
                    b_val = ws.cell(row=row_idx, column=2).value
                    if b_val:
                        b_str = str(b_val).strip()
                        if b_str not in row_code_map:  # first occurrence wins
                            row_code_map[b_str] = row_idx

                used_rows = set()
                for result in results:
                    bill_code = result.get("bill_item", {}).get("code", "")
                    if bill_code and bill_code in row_code_map:
                        target_row = row_code_map[bill_code]
                        if target_row not in used_rows:
                            code_pairs.append((target_row, result))
                            used_rows.add(target_row)

            if code_pairs:
                row_result_pairs = code_pairs
                unmatched_count = len(results) - len(code_pairs)
                if unmatched_count > 0:
                    logger.warning(
                        f"Sheet [{ws.title}]: bill_code fallback left "
                        f"{unmatched_count}/{len(results)} results unmapped")
                else:
                    logger.info(
                        f"Sheet [{ws.title}]: source_row/sheet_bill_seq unavailable, "
                        f"used bill_code fallback ({len(code_pairs)}/{len(bill_rows)})")
            else:
                # Fallback B: preserve original order and align positionally.
                if len(bill_rows) != len(results):
                    logger.warning(
                        f"Sheet [{ws.title}]: bill row count ({len(bill_rows)}) != "
                        f"result count ({len(results)}), falling back to positional mapping")
                num_to_process = min(len(bill_rows), len(results))
                row_result_pairs = [
                    (bill_rows[i], results[i]) for i in range(num_to_process)
                ]
        elif can_use_source_row_map and len(bill_rows) != len(results):
            logger.info(
                f"Sheet [{ws.title}]: subset export mapped by source_row "
                f"({len(results)}/{len(bill_rows)})")
        elif len(bill_rows) != len(results):
            logger.info(
                f"Sheet [{ws.title}]: subset export mapped by sheet_bill_seq "
                f"({len(results)}/{len(bill_rows)})")

        # 第3.5步：保存所有行的原始行高（包括 None = 自动高度）
        # insert_rows 不会正确移动 row_dimensions 的键，需要手动保存/恢复
        # None 表示自动高度，也要保存，否则恢复时这些行可能继承错误的大行高
        original_row_heights = {}  # {原始行号: 行高或None}
        for row_idx in range(header_row + 1, ws.max_row + 1):
            original_row_heights[row_idx] = ws.row_dimensions[row_idx].height

        # 第3.8步：检测清单行的列合并模式（用于给定额行也加相同合并）
        # 例如原表清单行有 F:G 合并（工程量）、I:J 合并，定额行也要同样合并
        bill_row_merges = []  # [(min_col, max_col), ...] 单行内的列合并
        if bill_rows:
            sample_row = bill_rows[0]
            for mr in saved_merges:
                # 找属于清单行的单行合并（min_row == max_row == 清单行）
                if mr[0] == sample_row and mr[2] == sample_row:
                    bill_row_merges.append((mr[1], mr[3]))  # (min_col, max_col)

        # 第4步：从下往上插入定额行（避免插行导致行号偏移）
        # 记录每次插入的位置和行数，用于恢复合并单元格
        insert_records = []
        for row_idx, result in sorted(row_result_pairs, key=lambda x: x[0], reverse=True):
            quotas = _ensure_list(result.get("quotas", []))

            # 在清单行的额外列写入推荐度、备选和主材（动态列位置）
            self._write_bill_extra_info(ws, row_idx, result, extra_start)

            # 要插入的行数（定额行+主材行，无定额不插入行）
            materials = _resolve_output_materials(result)
            quota_rows = len(quotas) if quotas else 0
            mat_rows = len(materials) if quotas else 0
            num_insert = quota_rows + mat_rows

            if num_insert == 0:
                # 没有定额也没有主材，不插入任何行
                # 只在清单行的J-O列写推荐度等信息（已在上面写过）
                continue

            # 插入空行（在清单行的下一行位置）
            ws.insert_rows(row_idx + 1, amount=num_insert)
            insert_records.append((row_idx + 1, num_insert))

            # 写入定额数据
            bill_unit = result.get("bill_item", {}).get("unit", "")
            bill_qty = result.get("bill_item", {}).get("quantity")

            if quotas:
                for q_idx, quota in enumerate(quotas):
                    q_row = row_idx + 1 + q_idx
                    self._write_single_quota_row(
                        ws, q_row, quota, bill_unit, bill_qty,
                        unit_col=unit_col, qty_col=qty_col, extra_start=extra_start)
                    # 注意：定额行的列合并（F:G, I:J等）不在这里做！
                    # 因为后续的 insert_rows 会错误移位这些合并，导致合并跑到
                    # 页面结构行上产生"竖杠"。改为第4.9步统一处理。
                    # 定额行行高：根据名称长度自适应
                    q_name = quota.get("name", "")
                    ws.row_dimensions[q_row].height = 30 if len(q_name) <= 30 else 45

            # 写入主材行（放在所有定额行之后）
            if quotas and materials:
                mat_start = row_idx + 1 + quota_rows
                for m_idx, mat in enumerate(materials):
                    m_row = mat_start + m_idx
                    self._write_single_material_row(
                        ws, m_row, mat,
                        unit_col=unit_col, qty_col=qty_col, extra_start=extra_start)
                    ws.row_dimensions[m_row].height = 30

        # 第4.5步：恢复所有原始行的行高（按插入偏移量计算新位置）
        # insert_records 已经在第4步记录了所有插入点
        sorted_inserts = sorted(insert_records, key=lambda x: x[0])
        for orig_row, orig_height in original_row_heights.items():
            # 计算该行因插入操作下移了多少
            offset = sum(cnt for ins_row, cnt in sorted_inserts if ins_row <= orig_row)
            new_row = orig_row + offset
            if orig_height is not None:
                ws.row_dimensions[new_row].height = orig_height
            else:
                # 原本是自动高度，清除可能被 insert_rows 错误设置的高度
                if new_row in ws.row_dimensions:
                    ws.row_dimensions[new_row].height = None

        # 第4.8步：恢复之前取消的合并单元格（保持原表格结构不变）
        # insert_records 是从下往上插的，恢复时需要从上往下排序
        insert_records.sort(key=lambda x: x[0])
        self._restore_merges(ws, saved_merges, insert_records)

        # 第4.9步：给定额行统一加列合并（F:G, I:J等）
        # 不能在第4步的插入循环中做，因为后续 insert_rows 会把合并错误移位到
        # 页面结构行上，导致重叠合并产生"竖杠"。在所有插入完成后统一处理。
        if bill_row_merges:
            for row_idx in range(header_row + 1, ws.max_row + 1):
                a_val = ws.cell(row=row_idx, column=1).value
                b_val = ws.cell(row=row_idx, column=2).value
                # 识别定额行或主材行：A列为空，B列有值
                if (a_val is None or str(a_val).strip() == "") and b_val:
                    b_str = str(b_val).strip()
                    if _is_quota_code(b_str) or _is_material_code(b_str):
                        for min_col, max_col in bill_row_merges:
                            try:
                                ws.merge_cells(
                                    start_row=row_idx, start_column=min_col,
                                    end_row=row_idx, end_column=max_col)
                            except Exception:
                                pass  # 合并失败（可能与其他合并重叠）不影响主流程

        # 第5步：在表头行添加额外列标题（动态列位置）
        self._add_extra_headers(ws, header_row, extra_start)

        # 第6步：格式化清单行和定额行（保留原表列宽，不覆盖）
        self._apply_post_format(ws, header_row, quantity_col=qty_col,
                                keep_col_widths=True, extra_start=extra_start)

        logger.info(f"Sheet [{ws.title}]: 处理 {len(row_result_pairs)} 条清单项")

    def _find_header_row_in_ws(self, ws) -> int:
        """识别真实表头行，避免把“工程量清单”等标题行误判为表头。"""
        scan_max_row = min(ws.max_row, 30)
        scan_max_col = min(ws.max_column, 20)

        unit_keywords = ("计量单位", "单位")
        qty_keywords = ("工程量", "工程数量", "数量")
        name_keywords = ("项目名称", "清单项目名称", "名称")
        code_keywords = ("项目编码", "编码")

        for row_idx in range(1, scan_max_row + 1):
            texts = []
            for col_idx in range(1, scan_max_col + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val is None:
                    continue
                text = self._normalize_header_text(val)
                if text:
                    texts.append(text)

            if not texts:
                continue

            row_text = " ".join(texts)
            has_unit = any(k in row_text for k in unit_keywords)
            has_qty = any(k in row_text for k in qty_keywords)
            has_name_or_code = any(k in row_text for k in name_keywords + code_keywords)

            if has_unit and has_qty and has_name_or_code:
                return row_idx

        return 1

    @staticmethod
    def _normalize_header_text(value) -> str:
        """Normalize header text for robust keyword matching."""
        if value is None:
            return ""
        return str(value).strip().replace("\n", "").replace(" ", "")

    def _find_header_col(self, ws, header_row: int, include_keywords: tuple[str, ...],
                         exclude_keywords: tuple[str, ...] = ()) -> int | None:
        """Find a header column by keywords on the detected header row."""
        scan_max_col = min(ws.max_column, 30)
        for col_idx in range(1, scan_max_col + 1):
            text = self._normalize_header_text(ws.cell(row=header_row, column=col_idx).value)
            if not text:
                continue
            if any(kw in text for kw in include_keywords) and not any(
                kw in text for kw in exclude_keywords
            ):
                return col_idx
        return None

    def _detect_bill_layout(self, ws, header_row: int) -> dict:
        """Detect unit/quantity columns from the source sheet header."""
        unit_col = self._find_header_col(
            ws,
            header_row,
            include_keywords=("计量单位", "单位"),
            exclude_keywords=("单价", "合价", "费用", "组成"),
        )
        qty_col = self._find_header_col(
            ws,
            header_row,
            include_keywords=("工程量", "工程数量", "数量"),
            exclude_keywords=("单价", "合价", "税", "费用", "组成"),
        )

        if unit_col is None:
            unit_col = 5
        if qty_col is None:
            qty_col = 6

        logger.debug(
            f"Sheet [{ws.title}] layout detected: unit_col={unit_col}, qty_col={qty_col}, header_row={header_row}"
        )
        return {"unit_col": unit_col, "qty_col": qty_col}

    def _unmerge_data_area(self, ws, header_row: int) -> list[tuple]:
        """取消数据区域的所有合并单元格，防止insert_rows导致数据丢失。

        分页打印格式的Excel每页都有重复表头（"分部分项工程和单价措施项目
        清单与计价表"、"序号"、"项目编码"等），这些表头有大量合并单元格
        （如"本页小计"行的A:H整行合并）。

        当在清单行下方插入定额行时，openpyxl的insert_rows在复杂合并结构下
        可能无法正确移动合并区域，导致清单行的B/C等列变成MergedCell（值丢失）。

        解决方案：插入行之前先取消合并 → 插入定额行 → 按偏移量重新合并回去。

        返回: 被取消的合并范围列表 [(min_row, min_col, max_row, max_col), ...]
              用于插入行后重新恢复合并。
        """
        # 收集需要取消的合并范围（只处理表头行以下的数据区域）
        saved_merges = []
        ranges_to_unmerge = []
        for merge_range in ws.merged_cells.ranges:
            if merge_range.min_row > header_row:
                ranges_to_unmerge.append(str(merge_range))
                # 保存坐标（后续按行偏移量重新合并）
                saved_merges.append((
                    merge_range.min_row, merge_range.min_col,
                    merge_range.max_row, merge_range.max_col,
                ))

        # 逐个取消合并
        for range_str in ranges_to_unmerge:
            ws.unmerge_cells(range_str)

        if ranges_to_unmerge:
            logger.debug(
                f"Sheet [{ws.title}]: 取消 {len(ranges_to_unmerge)} 个"
                f"数据区域合并单元格（防止insert_rows错位）")

        return saved_merges

    @staticmethod
    def _restore_merges(ws, saved_merges: list[tuple], insert_points: list[tuple]):
        """插入定额行后，把之前取消的合并单元格按偏移量恢复回去。

        参数:
            ws: 工作表
            saved_merges: _unmerge_data_area 返回的原始合并坐标
            insert_points: 插入记录列表 [(插入位置行号, 插入行数), ...]
                           按原始行号从小到大排序
        """
        if not saved_merges:
            return

        # 对每个原始合并范围，计算插入行导致的偏移量
        # insert_points 格式: [(row, count), ...] 表示在 row 处插入了 count 行
        # 如果合并范围的起始行 > 插入点，则该范围需要下移 count 行
        for min_row, min_col, max_row, max_col in saved_merges:
            offset = 0
            for ins_row, ins_count in insert_points:
                # 插入点在合并范围起始行或之前，该范围需要下移
                # ws.insert_rows(X) 会把 X 及以下的行全部下移
                if ins_row <= min_row:
                    offset += ins_count
            new_min_row = min_row + offset
            new_max_row = max_row + offset
            try:
                ws.merge_cells(
                    start_row=new_min_row, start_column=min_col,
                    end_row=new_max_row, end_column=max_col,
                )
            except Exception:
                pass  # 合并失败不影响主流程（可能与新插入的行重叠）

    def _remove_existing_quota_rows(self, ws, header_row: int):
        """删除已有的定额行和主材行（从下往上删，避免行号偏移）"""
        rows_to_delete = []

        for row_idx in range(header_row + 1, ws.max_row + 1):
            a_val = ws.cell(row=row_idx, column=1).value
            b_val = ws.cell(row=row_idx, column=2).value

            # A列为空、B列有值 → 可能是定额行或主材行
            if (a_val is None or str(a_val).strip() == "") and b_val:
                b_str = str(b_val).strip()
                # 定额行（如C4-4-31、5-325）
                if _is_quota_code(b_str):
                    rows_to_delete.append(row_idx)
                # 主材行（如01290303@1、13030795、补充主材005）
                elif _is_material_code(b_str):
                    rows_to_delete.append(row_idx)

        # 从下往上删除
        for row_idx in reversed(rows_to_delete):
            ws.delete_rows(row_idx, 1)

        if rows_to_delete:
            logger.info(f"Sheet [{ws.title}]: 删除 {len(rows_to_delete)} 条已有定额/主材行")

    def _write_bill_extra_info(self, ws, row_idx: int, result: dict, extra_start: int = 10):
        """在清单行的额外列写入推荐度、匹配说明、备选定额和主材

        extra_start: 额外列的起始列号（动态计算，追加到原表最后一列之后）
        """
        confidence = _safe_confidence(result.get("confidence", 0), default=0)
        quotas = _ensure_list(result.get("quotas", []))
        explanation = result.get("explanation", "")
        match_source = result.get("match_source", "")

        # 判断是否需要复核（任一条件命中即标记）
        review_needed = self._check_review_needed(confidence, quotas, match_source)

        # 置信度颜色
        conf_fill = self._confidence_fill(confidence)

        # 推荐度列（extra_start）：星级推荐（安全写入，跳过合并单元格）
        conf_text = confidence_to_stars(confidence, bool(quotas))
        cell_j = _safe_write_cell(ws, row_idx, extra_start, conf_text)
        if cell_j:
            cell_j.font = BILL_FONT
            cell_j.border = THIN_BORDER
            if quotas:
                cell_j.fill = conf_fill

        # 匹配说明列（extra_start+1）：L4请教/同类待定/需复核 前缀标记
        brief = self._brief_explanation(explanation)
        if result.get("l4_representative"):
            # L4代表项：用户只需改这条，同类自动学习
            group_label = result.get("l4_group_label", "")
            group_size = result.get("l4_group_size", 0)
            brief = f"[请教] {group_label}类{group_size}条不确定，请修正此条（同类自动学习）"
        elif result.get("l4_follower"):
            brief = f"[同类待定] {brief}" if brief else "[同类待定]"
        elif review_needed:
            brief = f"[需复核] {brief}" if brief else "[需复核]"
        # L6: 追加规则提示（如有）
        rule_hints = result.get("rule_hints", "")
        if rule_hints:
            brief = f"{brief} [规则]{rule_hints}" if brief else f"[规则]{rule_hints}"
        knowledge_brief = self._brief_knowledge_evidence(result)
        if knowledge_brief:
            brief = f"{brief} {knowledge_brief}".strip() if brief else knowledge_brief
        cell_k = _safe_write_cell(ws, row_idx, extra_start + 1, brief)
        if cell_k:
            cell_k.font = BILL_FONT
            cell_k.border = THIN_BORDER

        # 备选定额列（extra_start+2 起，共3列）
        self._write_alternative_cells(
            ws, row_idx, start_col=extra_start + 2, alternatives=result.get("alternatives", [])
        )

        # 主材列已移除（主材跟着定额子目走，清单行旁边显示没有意义）

    def _write_quota_rows(self, ws, current_row: int, result: dict,
                          bill_unit: str, bill_qty, max_col: int) -> int:
        """写入一条清单对应的所有定额行（兜底新建模式用）

        参数:
            ws: worksheet对象
            current_row: 当前写入的行号
            result: 匹配结果字典
            bill_unit: 清单单位
            bill_qty: 清单工程量
            max_col: 最大列数（用于格式化）

        返回:
            写入后的下一个可用行号
        """
        quotas = _ensure_list(result.get("quotas", []))

        if quotas:
            for quota in quotas:
                self._write_single_quota_row(
                    ws, current_row, quota, bill_unit, bill_qty)
                current_row += 1
        else:
            # 未匹配提示行
            no_reason = (
                result.get("no_match_reason")
                or result.get("final_reason")
                or result.get("reason_detail")
                or result.get("explanation")
                or "未找到匹配定额"
            )
            self._write_no_match_row(ws, current_row, no_reason, max_col)
            current_row += 1

        # 写入主材行（定额之后）
        materials = _resolve_output_materials(result)
        for mat in materials:
            self._write_single_material_row(ws, current_row, mat)
            current_row += 1

        return current_row

    def _write_single_quota_row(self, ws, q_row: int, quota: dict,
                                bill_unit: str, bill_qty,
                                unit_col: int = 5, qty_col: int = 6,
                                extra_start: int = 10):
        """写入一行定额数据（广联达标准格式：宋体9号、thin边框、无背景、不合并）"""
        # A列留空（广联达靠这个区分清单行和子目行）

        # B列：定额编号（居中）
        cell_b = ws.cell(
            row=q_row, column=2, value=safe_excel_text(quota.get("quota_id", ""))
        )
        cell_b.font = GLD_FONT
        cell_b.alignment = Alignment(horizontal="center", vertical="center",
                                     wrap_text=True)

        # C列：定额名称（左对齐，自动换行）
        cell_c = ws.cell(
            row=q_row, column=3, value=safe_excel_text(quota.get("name", ""))
        )
        cell_c.font = GLD_FONT
        cell_c.alignment = Alignment(horizontal="left", vertical="center",
                                     wrap_text=True)

        # E列：单位（居中）
        quota_unit = quota.get("unit", "") or bill_unit
        cell_e = ws.cell(row=q_row, column=unit_col, value=quota_unit)
        cell_e.font = GLD_FONT
        cell_e.alignment = Alignment(horizontal="center", vertical="center",
                                     wrap_text=True)

        # F列：工程量（右对齐，自动单位换算）
        converted_qty = convert_quantity(bill_qty, bill_unit, quota_unit)
        cell_f = ws.cell(row=q_row, column=qty_col, value=converted_qty)
        cell_f.font = GLD_FONT
        cell_f.alignment = Alignment(horizontal="right", vertical="center",
                                     wrap_text=True)

        # 所有列统一 thin 边框 + 宋体9号（和清单行一致，覆盖到额外列末尾）
        for col in range(1, extra_start + 6):  # A列到主材列
            cell = ws.cell(row=q_row, column=col)
            cell.border = THIN_BORDER
            if cell.font == Font() or cell.font is None:
                cell.font = GLD_FONT
            if cell.alignment is None or cell.alignment == Alignment():
                cell.alignment = Alignment(vertical="center", wrap_text=True)

    def _write_single_material_row(self, ws, m_row: int, material: dict,
                                    unit_col: int = 5, qty_col: int = 6,
                                    extra_start: int = 10):
        """写入一行主材数据（格式与定额行一致：宋体9号、thin边框）

        主材行特征：A列空，B列=材料编码，C列=材料名称，E列=单位，F列=数量
        """
        # B列：材料编码，无编码时写"主"标识（广联达识别用）
        mat_code = material.get("code", "") or "主"
        cell_b = ws.cell(row=m_row, column=2,
                         value=safe_excel_text(mat_code))
        cell_b.font = GLD_FONT
        cell_b.alignment = Alignment(horizontal="center", vertical="center",
                                     wrap_text=True)

        # C列：材料名称（左对齐）
        cell_c = ws.cell(row=m_row, column=3,
                         value=safe_excel_text(material.get("name", "")))
        cell_c.font = GLD_FONT
        cell_c.alignment = Alignment(horizontal="left", vertical="center",
                                     wrap_text=True)

        # E列：单位（居中）
        mat_unit = material.get("unit", "")
        cell_e = ws.cell(row=m_row, column=unit_col, value=mat_unit)
        cell_e.font = GLD_FONT
        cell_e.alignment = Alignment(horizontal="center", vertical="center",
                                     wrap_text=True)

        # F列：数量（右对齐）
        mat_qty = material.get("qty")
        cell_f = ws.cell(row=m_row, column=qty_col, value=mat_qty)
        cell_f.font = GLD_FONT
        cell_f.alignment = Alignment(horizontal="right", vertical="center",
                                     wrap_text=True)

        # G列(qty_col+1)：单价 — 套定额阶段不填，留给"智能填主材"功能填写

        # 所有列统一 thin 边框 + 宋体9号（覆盖到额外列末尾）
        for col in range(1, extra_start + 6):  # A列到主材列
            cell = ws.cell(row=m_row, column=col)
            cell.border = THIN_BORDER
            if cell.font == Font() or cell.font is None:
                cell.font = GLD_FONT
            if cell.alignment is None or cell.alignment == Alignment():
                cell.alignment = Alignment(vertical="center", wrap_text=True)

    def _add_extra_headers(self, ws, header_row: int, extra_start: int = 10):
        """在表头行添加额外列标题（推荐度/匹配说明/备选/主材）"""
        extra_headers = {
            extra_start: ("推荐度", HEADER_FILL),
            extra_start + 1: ("匹配说明", HEADER_FILL),
            extra_start + 2: ("备选1", LIGHT_BLUE_FILL),
            extra_start + 3: ("备选2", LIGHT_BLUE_FILL),
            extra_start + 4: ("备选3", LIGHT_BLUE_FILL),
        }
        for col, (title, fill) in extra_headers.items():
            self._set_header_cell(ws, header_row, col, title, fill)

        # 设置额外列宽（用动态列号转字母）
        col_widths = [14, 40, 30, 30, 30]  # 推荐度/匹配说明/备选1/备选2/备选3
        for i, width in enumerate(col_widths):
            col_letter = get_column_letter(extra_start + i)
            ws.column_dimensions[col_letter].width = width

    def _apply_post_format(self, ws, header_row: int, quantity_col: int = 6,
                           keep_col_widths: bool = False, extra_start: int = 10):
        """
        格式化清单行和定额行（不动分部/合计等原始行，保留其原始格式）

        解决的问题：
        - openpyxl 加载后清单行可能丢失字体/边框/对齐
        - insert_rows 新插入的定额行没有格式
        - 统一设 wrap_text=True 让文字自动换行
        - 设固定列宽让表格宽度一致（keep_col_widths=True时跳过，保留原表列宽）
        """
        # 1. 设固定列宽（保留原表结构时不动，避免破坏分页打印格式）
        if not keep_col_widths:
            for col, width in STANDARD_COL_WIDTHS.items():
                ws.column_dimensions[col].width = width

        # 2. 只格式化清单行、定额行和主材行，跳过分部/合计等原始行
        # 额外列的列号集合（匹配说明/备选1/备选2/备选3 = extra_start+1 到 extra_start+4）
        extra_left_align_cols = set(range(extra_start + 1, extra_start + 5))
        for row_idx in range(header_row + 1, ws.max_row + 1):
            a_val = ws.cell(row=row_idx, column=1).value
            b_val = ws.cell(row=row_idx, column=2).value
            is_bill = _is_bill_serial(a_val)
            is_quota = (
                (a_val is None or str(a_val).strip() == "")
                and b_val
                and _is_quota_code(str(b_val).strip())
            )
            # 主材行：A列空，B列是材料编码格式
            is_material = (
                (a_val is None or str(a_val).strip() == "")
                and b_val
                and not _is_quota_code(str(b_val).strip())
                and _is_material_code(str(b_val).strip())
            )

            if not is_bill and not is_quota and not is_material:
                continue  # 分部/合计/空行等，保留原始格式不动

            for col_idx in range(1, extra_start + 5):  # A列到备选3列
                cell = ws.cell(row=row_idx, column=col_idx)
                if isinstance(cell, MergedCell):
                    continue  # 合并单元格跳过，避免写样式崩溃

                # 边框：统一设 thin 边框
                cell.border = THIN_BORDER

                # 字体：如果丢失或变成默认 Calibri 就覆盖为宋体9号
                if not cell.font or cell.font.name in (None, "Calibri"):
                    cell.font = GLD_FONT

                # 对齐 + wrap_text：根据列确定对齐方式
                h_align = "center"
                if col_idx in (3, 4) or col_idx in extra_left_align_cols:  # C/D/额外列：左对齐
                    h_align = "left"
                elif col_idx in (quantity_col, 7, 8):  # 数量/单价/合价：右对齐
                    h_align = "right"
                cell.alignment = Alignment(
                    horizontal=h_align, vertical="center", wrap_text=True
                )

    # ================================================================
    # 兜底模式：新建工作簿（无原始文件时使用）
    # ================================================================

    def _write_new_workbook(self, results, output_path):
        """无原始文件时，新建工作簿输出"""
        wb = openpyxl.Workbook()
        try:
            self._write_detail_sheet(wb.active, results)
            wb.active.title = "匹配结果明细"

            ws_review = wb.create_sheet("待审核")
            self._write_review_sheet(ws_review, results)

            ws_stats = wb.create_sheet("统计汇总")
            self._write_stats_sheet(ws_stats, results)

            self._save_workbook_atomic(wb, output_path)
        finally:
            wb.close()
        logger.info(f"匹配结果已保存（新建模式）: {output_path}")
        return output_path

    def _write_detail_sheet(self, ws, results: list[dict]):
        """纯新建模式：不依赖原始文件，直接生成标准格式"""
        col_widths = {"A": 6, "B": 16, "C": 50, "D": 40, "E": 8,
                      "F": 12, "G": 12, "H": 12, "I": 12, "J": 14, "K": 20,
                      "L": 18, "M": 18, "N": 18, "O": 20}
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width

        # 行1表头
        headers = {1: "序号", 2: "项目编码", 3: "项目名称", 4: "项目特征描述",
                   5: "计量单位", 6: "工程量", 7: "金额（元）", 10: "推荐度",
                   11: "匹配说明", 12: "备选1", 13: "备选2", 14: "备选3"}
        for col_idx, header in headers.items():
            fill = LIGHT_BLUE_FILL if col_idx >= 12 else HEADER_FILL
            self._set_header_cell(ws, 1, col_idx, header, fill)
        for col_idx in [8, 9]:
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER
        ws.merge_cells("G1:I1")

        # 行2子表头
        for col_idx, header in {7: "综合单价", 8: "合价", 9: "其中：暂估价"}.items():
            self._set_header_cell(ws, 2, col_idx, header, HEADER_FILL)
        for col_idx in list(range(1, 7)) + [10, 11]:
            cell = ws.cell(row=2, column=col_idx)
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER

        # 数据行
        current_row = 3
        for idx, result in enumerate(results, start=1):
            bill = result.get("bill_item", {})
            confidence = _safe_confidence(result.get("confidence", 0), default=0)
            quotas = _ensure_list(result.get("quotas", []))
            explanation = result.get("explanation", "")
            bill_quantity = bill.get("quantity", "")
            bill_unit = bill.get("unit", "")

            conf_fill = self._confidence_fill(confidence)

            # 清单行
            ws.cell(row=current_row, column=1, value=idx)
            ws.cell(row=current_row, column=2, value=safe_excel_text(bill.get("code", "")))
            ws.cell(row=current_row, column=3, value=safe_excel_text(bill.get("name", "")))
            ws.cell(row=current_row, column=4, value=safe_excel_text(bill.get("description", "")))
            ws.cell(row=current_row, column=5, value=bill_unit)
            ws.cell(row=current_row, column=6, value=bill_quantity)
            # J列：星级推荐
            conf_text = confidence_to_stars(confidence, bool(quotas))
            ws.cell(row=current_row, column=10, value=conf_text)
            ws.cell(
                row=current_row,
                column=11,
                value=self._brief_explanation(explanation)
            )

            # L/M/N列：备选定额
            self._write_alternative_cells(
                ws, current_row, start_col=12, alternatives=result.get("alternatives", [])
            )
            ws.cell(row=current_row, column=15, value=self._get_material_text(result))

            self._apply_row_style(ws, current_row, 1, 15, {4, 15})
            if quotas:
                ws.cell(row=current_row, column=10).fill = conf_fill
            current_row += 1

            # 子目行
            current_row = self._write_quota_rows(
                ws, current_row, result, bill_unit, bill_quantity, 9)

        ws.freeze_panes = "A3"

    def _write_review_sheet(self, ws, results: list[dict]):
        """
        写入"待审核"Sheet：只列出黄色和红色的清单项

        方便用户快速定位需要修改的条目，不需要翻看整个匹配结果。
        包含：序号、清单名称、当前定额、推荐度、问题说明、备选1/2/3
        """
        # 列宽（加了项目特征列，整体后移一列）
        col_widths = {"A": 8, "B": 30, "C": 40, "D": 20, "E": 30, "F": 14,
                      "G": 24, "H": 18, "I": 18, "J": 18}
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width

        # 表头
        headers = ["清单序号", "清单名称", "项目特征", "当前定额编号", "当前定额名称",
                   "推荐度", "问题说明", "备选1", "备选2", "备选3"]
        for col_idx, header in enumerate(headers, start=1):
            self._set_header_cell(ws, 1, col_idx, header, HEADER_FILL)

        # 筛选出置信度 < CONFIDENCE_GREEN（85%）的条目
        current_row = 2
        review_count = 0
        follower_count = 0  # L4从属项计数（不逐条显示，末尾汇总）

        for idx, result in enumerate(results, start=1):
            confidence = _safe_confidence(result.get("confidence", 0), default=0)
            quotas = _ensure_list(result.get("quotas", []))
            match_source = result.get("match_source", "")
            review_needed = self._check_review_needed(confidence, quotas, match_source)

            # 统一与主表逻辑：无匹配 / 低置信度 / 降级来源 都进入待审核
            if not review_needed:
                continue

            # L4从属项不在待审核Sheet逐条显示（用户只看代表项即可）
            # 但会在末尾汇总提示，避免从属项被完全遗漏
            if result.get("l4_follower"):
                follower_count += 1
                continue

            bill = result.get("bill_item", {})
            explanation = result.get("explanation", "")
            alternatives = _ensure_list(result.get("alternatives", []))

            # 当前匹配的定额
            main_quota_id = ""
            main_quota_name = ""
            if quotas:
                main_quota_id = quotas[0].get("quota_id", "")
                main_quota_name = quotas[0].get("name", "")

            # 推荐度颜色
            conf_fill = self._confidence_fill(confidence)

            # 写入数据（加了项目特征列，整体后移一列）
            ws.cell(row=current_row, column=1, value=idx)
            ws.cell(row=current_row, column=2, value=safe_excel_text(bill.get("name", "")))
            ws.cell(row=current_row, column=3, value=safe_excel_text(bill.get("description", "")))
            ws.cell(row=current_row, column=4, value=safe_excel_text(main_quota_id))
            ws.cell(row=current_row, column=5, value=safe_excel_text(main_quota_name))

            conf_text = confidence_to_stars(confidence, bool(quotas))
            cell_conf = ws.cell(row=current_row, column=6, value=conf_text)
            cell_conf.fill = conf_fill

            # 问题说明（代表项加 [请教] 前缀）
            brief_text = self._brief_explanation(explanation)
            if result.get("l4_representative"):
                group_label = result.get("l4_group_label", "")
                group_size = result.get("l4_group_size", 0)
                brief_text = f"[请教] {group_label}类{group_size}条，改此条同类自动学习"
            ws.cell(
                row=current_row,
                column=7,
                value=brief_text
            )

            # 备选定额
            self._write_alternative_cells(
                ws, current_row, start_col=8, alternatives=alternatives
            )

            # 格式（项目特征、定额名称、问题说明列自动换行）
            self._apply_row_style(ws, current_row, 1, 10, {2, 3, 5, 7})

            current_row += 1
            review_count += 1

        # L4从属项汇总提示（不逐条列出，但告知用户数量）
        if follower_count > 0:
            current_row += 1  # 空一行
            summary_text = (
                f"另有 {follower_count} 条同类从属项未逐条列出，"
                f"修正上方[请教]代表项后，下次同类自动生效"
            )
            cell = ws.cell(row=current_row, column=1, value=summary_text)
            cell.font = Font(name="微软雅黑", size=10, italic=True, color="808080")
            ws.merge_cells(
                start_row=current_row, start_column=1,
                end_row=current_row, end_column=11)

        # 如果没有待审核项，写个提示
        if review_count == 0:
            cell = ws.cell(row=2, column=1, value="全部匹配结果均为高置信度，无需审核")
            cell.font = Font(name="微软雅黑", size=11, color="006100")
            ws.merge_cells("A2:K2")

        ws.freeze_panes = "A2"
        logger.info(f"待审核Sheet: {review_count} 条需要人工审核")

    def _write_stats_sheet(self, ws, results: list[dict]):
        """写入统计汇总表"""
        total = len(results)
        matched = sum(1 for r in results if r.get("quotas"))
        high_conf = sum(1 for r in results
                        if _safe_confidence(r.get("confidence", 0), default=0) >= config.CONFIDENCE_GREEN)
        mid_conf = sum(1 for r in results
                       if config.CONFIDENCE_YELLOW
                       <= _safe_confidence(r.get("confidence", 0), default=0) < config.CONFIDENCE_GREEN)
        low_conf = sum(1 for r in results
                       if 0 < _safe_confidence(r.get("confidence", 0), default=0) < config.CONFIDENCE_YELLOW)
        no_match = total - matched

        stats = [
            ["统计项", "数量", "占比"],
            ["清单总数", total, "100%"],
            ["已匹配", matched, f"{matched * 100 // max(total, 1)}%"],
            ["高置信度（绿色）", high_conf,
             f"{high_conf * 100 // max(total, 1)}%"],
            ["中置信度（黄色）", mid_conf,
             f"{mid_conf * 100 // max(total, 1)}%"],
            ["低置信度（红色）", low_conf,
             f"{low_conf * 100 // max(total, 1)}%"],
            ["未匹配", no_match, f"{no_match * 100 // max(total, 1)}%"],
        ]

        for row_idx, row_data in enumerate(stats, start=1):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = THIN_BORDER
                if row_idx == 1:
                    cell.font = HEADER_FONT
                    cell.fill = HEADER_FILL
                else:
                    cell.font = BILL_FONT

        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 10
        ws.column_dimensions["C"].width = 10

