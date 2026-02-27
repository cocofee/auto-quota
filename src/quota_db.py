"""
定额数据库管理模块
功能：
1. 读取定额Excel文件（支持openpyxl和ZIP+XML两种方式）
2. 解析每条定额：编号、名称、特征参数、单位、工作类型
3. 用text_parser提取结构化参数
4. 存入SQLite数据库
"""

import zipfile
import xml.etree.ElementTree as ET
import re
import datetime  # 用于检测Excel误解析的日期类型
from pathlib import Path
from loguru import logger

import config
from src.text_parser import parser as text_parser
from db.sqlite import connect as _db_connect


# XML命名空间（xlsx文件内部用的）
SPREADSHEET_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"


class QuotaDB:
    """定额数据库管理器"""

    def __init__(self, province: str = None):
        """
        参数:
            province: 省份名称，如"北京2024"，默认用config中的配置
        """
        self.province = province or config.get_current_province()
        self.db_path = config.get_quota_db_path(self.province)

        # 确保数据库目录存在
        self.db_path.parent.mkdir(parents=True, exist_ok=True)

    def _connect(self, row_factory: bool = False):
        """统一SQLite连接参数"""
        return _db_connect(self.db_path, row_factory=row_factory)

    def init_db(self):
        """创建数据库表结构"""
        conn = self._connect()
        cursor = conn.cursor()

        cursor.execute("""
            CREATE TABLE IF NOT EXISTS quotas (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                quota_id TEXT NOT NULL,          -- 定额编号（如C1-1-1）
                name TEXT NOT NULL,              -- 完整名称（含特征参数）
                unit TEXT,                       -- 计量单位
                work_type TEXT,                  -- 工作类型（安装/土建/市政）
                specialty TEXT,                  -- 专业大类（如"安装"）
                chapter TEXT,                    -- 章节名称（Excel的Sheet名）

                -- 结构化参数（text_parser提取）
                dn INTEGER,                      -- 管径(mm)
                cable_section REAL,              -- 电缆截面(mm²)
                kva REAL,                        -- 容量(kVA)
                kv REAL,                         -- 电压等级(kV)
                ampere REAL,                     -- 电流(A)
                weight_t REAL,                   -- 重量(吨)
                material TEXT,                   -- 材质
                connection TEXT,                 -- 连接方式
                circuits INTEGER,                -- 回路数
                shape TEXT,                      -- 形状(矩形/圆形)
                perimeter REAL,                  -- 周长
                large_side REAL,                 -- 大边长
                elevator_stops INTEGER,          -- 电梯站数
                elevator_speed REAL,             -- 电梯速度(m/s)

                -- 搜索用文本（清洗后的）
                search_text TEXT,                -- 用于BM25和向量搜索的文本

                -- 大册分类（从定额编号前缀提取，如C10、C4）
                book TEXT,                       -- 所属大册编号

                UNIQUE(quota_id, chapter)        -- 同一章节内编号唯一
            )
        """)

        # 版本元数据表：记录每次导入定额的版本信息
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS db_meta (
                key TEXT PRIMARY KEY,
                value TEXT
            )
        """)

        # 导入历史表：记录每个已导入的Excel文件，用于增量导入时跳过已导入的文件
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS import_history (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                file_path TEXT NOT NULL,          -- 文件完整路径（用于区分同名文件）
                file_name TEXT NOT NULL,          -- 文件名（便于展示）
                file_size INTEGER,                -- 文件大小(字节)
                file_mtime REAL,                  -- 文件最后修改时间(时间戳)
                specialty TEXT,                   -- 专业（安装/土建/市政）
                quota_count INTEGER DEFAULT 0,    -- 导入的定额条数
                imported_at REAL,                 -- 导入时间(时间戳)
                UNIQUE(file_path)                 -- 同一路径文件只保留最新一条记录
            )
        """)
        self._migrate_import_history_schema(cursor)
        self._migrate_import_history_hash(cursor)

        # 兼容旧库：检查book列是否存在，不存在则添加
        # （旧版数据库没有book字段，直接建索引会报错"no such column"）
        existing_cols = {row[1] for row in cursor.execute("PRAGMA table_info(quotas)").fetchall()}
        required_columns = {
            "book": "TEXT",
            "circuits": "INTEGER",
            "shape": "TEXT",
            "perimeter": "REAL",
            "large_side": "REAL",
            "elevator_stops": "INTEGER",
            "elevator_speed": "REAL",
        }
        for col_name, col_type in required_columns.items():
            if col_name not in existing_cols:
                cursor.execute(f"ALTER TABLE quotas ADD COLUMN {col_name} {col_type}")
                existing_cols.add(col_name)
                logger.info(f"旧数据库缺少{col_name}列，已自动添加")

        # 回填book为空的历史数据：从quota_id提取第一段作为册号
        # 兼容各省份编号格式：C10-1-5→C10, SC1-1-1→SC1, GY-1→GY, A-1-1→A, 2003-3-1→2003
        cursor.execute("""
            UPDATE quotas SET book = UPPER(
                CASE
                    -- 多字母前缀（2~3个字母+0~2位数字+横杠）
                    -- SC10-xx, GY1-xx 等
                    WHEN quota_id GLOB '[A-Za-z][A-Za-z][A-Za-z][0-9][0-9]-*' THEN SUBSTR(quota_id, 1, 5)
                    WHEN quota_id GLOB '[A-Za-z][A-Za-z][A-Za-z][0-9]-*' THEN SUBSTR(quota_id, 1, 4)
                    WHEN quota_id GLOB '[A-Za-z][A-Za-z][A-Za-z]-*' THEN SUBSTR(quota_id, 1, 3)
                    WHEN quota_id GLOB '[A-Za-z][A-Za-z][0-9][0-9]-*' THEN SUBSTR(quota_id, 1, 4)
                    WHEN quota_id GLOB '[A-Za-z][A-Za-z][0-9]-*' THEN SUBSTR(quota_id, 1, 3)
                    WHEN quota_id GLOB '[A-Za-z][A-Za-z]-*' THEN SUBSTR(quota_id, 1, 2)
                    -- 单字母前缀：C10-xx, A-xx 等
                    WHEN quota_id GLOB '[A-Za-z][0-9][0-9]-*' THEN SUBSTR(quota_id, 1, 3)
                    WHEN quota_id GLOB '[A-Za-z][0-9]-*' THEN SUBSTR(quota_id, 1, 2)
                    WHEN quota_id GLOB '[A-Za-z]-*' THEN SUBSTR(quota_id, 1, 1)
                    -- 纯数字前缀（1~4位数字+横杠）
                    WHEN quota_id GLOB '[0-9][0-9][0-9][0-9]-*' THEN SUBSTR(quota_id, 1, 4)
                    WHEN quota_id GLOB '[0-9][0-9][0-9]-*' THEN SUBSTR(quota_id, 1, 3)
                    WHEN quota_id GLOB '[0-9][0-9]-*' THEN SUBSTR(quota_id, 1, 2)
                    WHEN quota_id GLOB '[0-9]-*' THEN SUBSTR(quota_id, 1, 1)
                    ELSE ''
                END
            ) WHERE book IS NULL OR book = ''
        """)
        backfilled = cursor.rowcount
        if backfilled > 0:
            logger.info(f"回填book字段{backfilled}条")

        # 创建索引加速查询
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_quota_id ON quotas(quota_id)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_specialty ON quotas(specialty)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_chapter ON quotas(chapter)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_book ON quotas(book)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_dn ON quotas(dn)")

        conn.commit()
        conn.close()
        logger.info(f"数据库初始化完成: {self.db_path}")

    def _migrate_import_history_schema(self, cursor):
        """将旧版 import_history(file_name 唯一) 升级为 file_path 唯一。"""
        columns = {row[1] for row in cursor.execute("PRAGMA table_info(import_history)").fetchall()}
        if "file_path" in columns:
            return

        cursor.execute("DROP TABLE IF EXISTS import_history_new")
        cursor.execute("""
            CREATE TABLE import_history_new (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                file_path TEXT NOT NULL,
                file_name TEXT NOT NULL,
                file_size INTEGER,
                file_mtime REAL,
                specialty TEXT,
                quota_count INTEGER DEFAULT 0,
                imported_at REAL,
                UNIQUE(file_path)
            )
        """)
        cursor.execute("""
            INSERT OR REPLACE INTO import_history_new
                (file_path, file_name, file_size, file_mtime, specialty, quota_count, imported_at)
            SELECT
                file_name AS file_path,
                file_name,
                file_size,
                file_mtime,
                specialty,
                quota_count,
                imported_at
            FROM import_history
        """)
        cursor.execute("DROP TABLE import_history")
        cursor.execute("ALTER TABLE import_history_new RENAME TO import_history")
        logger.info("已将import_history迁移为按file_path去重")

    def _migrate_import_history_hash(self, cursor):
        """为 import_history 表增加 file_hash/status/error_msg 列"""
        columns = {row[1] for row in cursor.execute("PRAGMA table_info(import_history)").fetchall()}
        if "file_hash" not in columns:
            cursor.execute("ALTER TABLE import_history ADD COLUMN file_hash TEXT")
            logger.info("import_history 已新增 file_hash 列")
        if "status" not in columns:
            cursor.execute("ALTER TABLE import_history ADD COLUMN status TEXT DEFAULT 'success'")
            logger.info("import_history 已新增 status 列")
        if "error_msg" not in columns:
            cursor.execute("ALTER TABLE import_history ADD COLUMN error_msg TEXT DEFAULT ''")
            logger.info("import_history 已新增 error_msg 列")

    def import_excel(self, excel_path: str, specialty: str = "安装",
                     clear_existing: bool = True):
        """
        导入定额Excel文件到SQLite数据库

        参数:
            excel_path: Excel文件路径
            specialty: 专业类别（安装/土建/市政）
            clear_existing: 是否清除该specialty的旧数据（默认True）。
                多文件导入同一specialty时，第一个文件设True，后续设False以追加数据。
        """
        excel_path = Path(excel_path)
        if not excel_path.exists():
            logger.error(f"文件不存在: {excel_path}")
            raise FileNotFoundError(f"定额文件不存在: {excel_path}")

        # 初始化数据库
        self.init_db()

        # 尝试用openpyxl读取，失败则用ZIP+XML方式
        try:
            quotas = self._read_with_openpyxl(excel_path, specialty)
        except Exception as e:
            logger.warning(f"openpyxl读取失败({e})，切换到ZIP+XML方式")
            quotas = self._read_with_zip_xml(excel_path, specialty)

        # 写入数据库（按specialty隔离，不影响其他专业）
        self._save_to_db(quotas, specialty=specialty, clear_existing=clear_existing)

        # 记录版本号（用定额数量+导入时间生成，改了定额数据版本号就会变）
        self._update_version(len(quotas))

        logger.info(f"导入完成: {specialty}定额 共{len(quotas)}条 → {self.db_path}")

        # 失效BookClassifier缓存（定额库变了，需要重新学习"词→册"概率）
        try:
            from src.book_classifier import BookClassifier
            BookClassifier.invalidate_cache(self.province)
        except Exception:
            pass  # BookClassifier未初始化时忽略

        return len(quotas)

    def _read_with_openpyxl(self, excel_path: Path, specialty: str) -> list[dict]:
        """用openpyxl读取Excel"""
        import openpyxl

        wb = openpyxl.load_workbook(str(excel_path), read_only=True, data_only=True)
        try:
            quotas = []

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                chapter = sheet_name.strip()
                logger.info(f"  读取Sheet: {chapter}")

                for row in ws.iter_rows(min_row=1, values_only=True):
                    quota = self._parse_row(row, chapter, specialty)
                    if quota:
                        quotas.append(quota)
        finally:
            wb.close()
        return quotas

    def _read_with_zip_xml(self, excel_path: Path, specialty: str) -> list[dict]:
        """
        用ZIP+XML方式读取Excel（备用方案）
        xlsx本质是个ZIP包，里面有XML文件
        """
        quotas = []

        with zipfile.ZipFile(str(excel_path), 'r') as zf:
            # 读取共享字符串表（所有文本内容在这里）
            shared_strings = self._read_shared_strings(zf)

            # 读取工作簿，获取Sheet名称列表
            sheet_names = self._read_sheet_names(zf)

            # 逐个Sheet读取
            for idx, sheet_name in enumerate(sheet_names):
                chapter = sheet_name.strip()
                sheet_file = f"xl/worksheets/sheet{idx + 1}.xml"

                if sheet_file not in zf.namelist():
                    logger.warning(f"  Sheet文件不存在: {sheet_file}")
                    continue

                logger.info(f"  读取Sheet: {chapter}")
                rows = self._parse_sheet_xml(zf, sheet_file, shared_strings)

                for row_values in rows:
                    quota = self._parse_row(row_values, chapter, specialty)
                    if quota:
                        quotas.append(quota)

        return quotas

    def _read_shared_strings(self, zf: zipfile.ZipFile) -> list[str]:
        """读取xlsx的共享字符串表"""
        ss_file = "xl/sharedStrings.xml"
        if ss_file not in zf.namelist():
            return []

        with zf.open(ss_file) as f:
            tree = ET.parse(f)
            root = tree.getroot()

        strings = []
        # 共享字符串可能在<si><t>文本</t></si>或<si><r><t>文本</t></r></si>结构中
        for si in root.findall(f"{{{SPREADSHEET_NS}}}si"):
            texts = []
            for t_elem in si.iter(f"{{{SPREADSHEET_NS}}}t"):
                if t_elem.text:
                    texts.append(t_elem.text)
            strings.append("".join(texts))

        return strings

    def _read_sheet_names(self, zf: zipfile.ZipFile) -> list[str]:
        """读取工作簿中的Sheet名称列表"""
        wb_file = "xl/workbook.xml"
        with zf.open(wb_file) as f:
            tree = ET.parse(f)
            root = tree.getroot()

        names = []
        for sheet in root.iter(f"{{{SPREADSHEET_NS}}}sheet"):
            name = sheet.get("name", "")
            names.append(name)

        return names

    def _parse_sheet_xml(self, zf: zipfile.ZipFile, sheet_file: str,
                         shared_strings: list[str]) -> list[list]:
        """解析单个Sheet的XML，返回行列表"""
        with zf.open(sheet_file) as f:
            tree = ET.parse(f)
            root = tree.getroot()

        rows_data = []
        for row_elem in root.iter(f"{{{SPREADSHEET_NS}}}row"):
            row_values = [None, None, None, None]  # 至少4列（A,B,C,D）

            for cell in row_elem.findall(f"{{{SPREADSHEET_NS}}}c"):
                ref = cell.get("r", "")      # 单元格引用（如A1, B2）
                cell_type = cell.get("t", "")  # 类型（s=共享字符串, 空=数字）

                # 解析列号（A=0, B=1, C=2, D=3）
                col_letter = re.match(r'([A-Z]+)', ref)
                if not col_letter:
                    continue
                col_idx = self._col_letter_to_index(col_letter.group(1))

                # 读取单元格值
                v_elem = cell.find(f"{{{SPREADSHEET_NS}}}v")
                if v_elem is None or v_elem.text is None:
                    continue

                if cell_type == "s":
                    # 共享字符串类型，用索引查找实际文本
                    str_idx = int(v_elem.text)
                    if str_idx < len(shared_strings):
                        value = shared_strings[str_idx]
                    else:
                        value = ""
                else:
                    value = v_elem.text

                # 扩展行列表以容纳这个列
                while len(row_values) <= col_idx:
                    row_values.append(None)
                row_values[col_idx] = value

            rows_data.append(row_values)

        return rows_data

    def _col_letter_to_index(self, letters: str) -> int:
        """列字母转数字索引: A→0, B→1, Z→25, AA→26"""
        result = 0
        for ch in letters:
            result = result * 26 + (ord(ch) - ord('A') + 1)
        return result - 1

    def _parse_row(self, row_values, chapter: str, specialty: str) -> dict | None:
        """
        解析一行定额数据

        参数:
            row_values: 行数据列表 [编号, 名称, 单位, 工作类型, ...]
            chapter: 章节名称（Sheet名）
            specialty: 专业类别

        返回:
            定额字典，或None（如果该行不是有效定额）
        """
        if not row_values or len(row_values) < 2:
            return None

        # 读取各列
        # P0修复：检测Excel误解析为日期的定额编号
        # 场景：编号"2003-03-01"被Excel/openpyxl自动解析为datetime(2003,3,1)
        # 修复：检测datetime对象，转回"年-月-日"数字编号格式
        raw_id = row_values[0]
        if isinstance(raw_id, (datetime.datetime, datetime.date)):
            # 去掉前导零，和普通定额编号格式保持一致（如5-3-32而非05-03-32）
            raw_id = f"{raw_id.year}-{raw_id.month}-{raw_id.day}"
            logger.debug(f"日期→编号修复: {row_values[0]} → {raw_id}")
        quota_id = str(raw_id or "").strip()
        name = str(row_values[1] or "").strip()
        unit = str(row_values[2] or "").strip() if len(row_values) > 2 else ""
        work_type = str(row_values[3] or "").strip() if len(row_values) > 3 else ""

        # 过滤无效行：编号和名称都不能为空
        if not quota_id or not name:
            return None

        # 过滤表头行（编号不像定额编号的跳过）
        # 定额编号通常格式: C10-1-1, A-1-5, SC1-1-1, GY-1, 01-01-001 等
        # 允许：字母+数字(C10-..., SC1-...)、字母+横杠(A-1-..., GY-...)、纯数字开头(01-...)
        if not re.match(r'^[A-Za-z]+[\d-]|^\d', quota_id):
            return None

        # 用text_parser提取结构化参数
        params = text_parser.parse(name)

        # 构建搜索文本
        search_text = text_parser.build_search_text(name)

        # 从定额编号提取所属大册（通用格式，兼容各省份）
        # C10-5-41 → C10, A-1-5 → A, SC1-1-1 → SC1, GY-1 → GY, 2003-3-1 → 2003
        book_match = re.match(r'^([A-Za-z]+\d{0,2})-', quota_id)
        if not book_match:
            # 纯数字前缀：1-2-3 → "1", 2003-3-1 → "2003"
            # 支持1~4位数字前缀（广东市政用4位如2003-xx-xx）
            book_match = re.match(r'^(\d{1,4})-', quota_id)
        book = book_match.group(1).upper() if book_match else ""

        return {
            "quota_id": quota_id,
            "name": name,
            "unit": unit,
            "work_type": work_type,
            "specialty": specialty,
            "chapter": chapter,
            "dn": params.get("dn"),
            "cable_section": params.get("cable_section"),
            "kva": params.get("kva"),
            "kv": params.get("kv"),
            "ampere": params.get("ampere"),
            "weight_t": params.get("weight_t"),
            "material": params.get("material"),
            "connection": params.get("connection"),
            "circuits": params.get("circuits"),
            "shape": params.get("shape"),
            "perimeter": params.get("perimeter"),
            "large_side": params.get("large_side"),
            "elevator_stops": params.get("elevator_stops"),
            "elevator_speed": params.get("elevator_speed"),
            "search_text": search_text,
            "book": book,
        }

    def _save_to_db(self, quotas: list[dict], specialty: str = None,
                    clear_existing: bool = True):
        """批量写入SQLite数据库

        参数:
            quotas: 定额列表
            specialty: 专业名称，指定时只删除该专业的旧数据（多专业共存）
            clear_existing: 是否清除旧数据（默认True）。
                多文件导入同一specialty时，第一个文件设True清除旧数据，
                后续文件设False直接追加，避免覆盖前面已导入的数据。
        """
        conn = self._connect()
        cursor = conn.cursor()

        try:
            # 删除+插入在同一事务中，失败时自动回滚，避免删了旧数据但插入失败导致数据丢失
            # 按专业删除旧数据（不影响其他专业）
            if clear_existing:
                if specialty:
                    cursor.execute("DELETE FROM quotas WHERE specialty = ?", (specialty,))
                    logger.info(f"已清空 specialty='{specialty}' 的旧数据")
                else:
                    cursor.execute("DELETE FROM quotas")
                    logger.info("已清空全部旧数据")
            else:
                logger.info(f"追加模式: 保留 specialty='{specialty}' 的已有数据")

            # 批量插入
            insert_sql = """
                INSERT OR IGNORE INTO quotas
                (quota_id, name, unit, work_type, specialty, chapter,
                 dn, cable_section, kva, kv, ampere, weight_t, material, connection,
                 circuits, shape, perimeter, large_side, elevator_stops, elevator_speed,
                 search_text, book)
                VALUES
                (?, ?, ?, ?, ?, ?,
                 ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?,
                 ?, ?)
            """

            batch = []
            for q in quotas:
                batch.append((
                    q["quota_id"], q["name"], q["unit"], q["work_type"],
                    q["specialty"], q["chapter"],
                    q["dn"], q["cable_section"], q["kva"], q["kv"],
                    q["ampere"], q["weight_t"], q["material"], q["connection"],
                    q["circuits"], q["shape"], q["perimeter"], q["large_side"],
                    q["elevator_stops"], q["elevator_speed"],
                    q["search_text"], q.get("book", ""),
                ))

            cursor.executemany(insert_sql, batch)
            conn.commit()
            logger.info(f"写入数据库: {len(batch)}条记录")
        except Exception:
            conn.rollback()
            logger.error(f"数据库写入失败，已回滚（specialty={specialty}）")
            raise
        finally:
            conn.close()

    def _update_version(self, quota_count: int):
        """更新定额库版本号

        版本号格式："{定额数量}_{时间戳}"
        每次重新导入定额后版本号都会变化，
        用于经验库判断"这条经验是否基于当前版本的定额库"。
        """
        import time
        version = f"{quota_count}_{int(time.time())}"

        conn = self._connect()
        cursor = conn.cursor()
        cursor.execute("""
            INSERT OR REPLACE INTO db_meta (key, value)
            VALUES ('version', ?)
        """, (version,))
        conn.commit()
        conn.close()

        logger.info(f"定额库版本号已更新: {version}")

    def get_version(self) -> str:
        """获取当前定额库的版本号

        返回:
            版本号字符串，如果从未导入过则返回空字符串
        """
        try:
            conn = self._connect()
            try:
                cursor = conn.cursor()
                cursor.execute("SELECT value FROM db_meta WHERE key = 'version'")
                row = cursor.fetchone()
                return row[0] if row else ""
            finally:
                conn.close()
        except Exception as e:
            logger.debug(f"读取定额库版本号失败，按空版本返回: {e}")
            return ""

    # ================================================================
    # 导入历史（用于增量导入）
    # ================================================================

    def record_import(self, file_path: str, specialty: str, quota_count: int,
                      status: str = "success", error_msg: str = ""):
        """记录一个文件的导入信息，用于下次增量导入时判断是否已导入过

        参数:
            file_path: Excel文件的完整路径
            specialty: 专业类别
            quota_count: 本次导入的定额条数
            status: 导入状态（success/error）
            error_msg: 错误信息（status=error时记录）
        """
        self.init_db()
        import time as _time
        import hashlib
        p = Path(file_path)
        full_path = str(p.resolve())

        # 计算文件内容MD5（定额Excel通常几MB，秒级完成）
        file_hash = ""
        file_size = 0
        file_mtime = 0.0
        try:
            stat = p.stat()
            file_size = stat.st_size
            file_mtime = stat.st_mtime
            hasher = hashlib.md5()
            with open(p, "rb") as f:
                for chunk in iter(lambda: f.read(8192), b""):
                    hasher.update(chunk)
            file_hash = hasher.hexdigest()
        except OSError:
            pass  # 文件不可读时仍记录历史（status=error场景）

        conn = self._connect()
        try:
            cursor = conn.cursor()
            cursor.execute("""
                INSERT OR REPLACE INTO import_history
                    (file_path, file_name, file_size, file_mtime, file_hash,
                     specialty, quota_count, imported_at, status, error_msg)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (full_path, p.name, file_size, file_mtime, file_hash,
                  specialty, quota_count, _time.time(), status, error_msg))
            conn.commit()
        finally:
            conn.close()

    def get_import_history(self) -> list[dict]:
        """获取所有已导入文件的记录

        返回:
            [{"file_path": "D:/quota/安装.xlsx", "file_name": "安装.xlsx",
              "file_size": 12345, "file_mtime": 1700000000.0, "file_hash": "abc123...",
              "specialty": "安装", "quota_count": 3000, "imported_at": 1700000000.0}, ...]
        """
        self.init_db()  # 确保表存在（兼容旧数据库）
        conn = self._connect(row_factory=True)
        try:
            cursor = conn.cursor()
            cursor.execute(
                "SELECT file_path, file_name, file_size, file_mtime, file_hash, "
                "specialty, quota_count, imported_at, status, error_msg "
                "FROM import_history"
            )
            return [dict(r) for r in cursor.fetchall()]
        finally:
            conn.close()

    def clear_import_history(self):
        """清空导入历史（全量重导时调用）"""
        conn = self._connect()
        try:
            conn.execute("DELETE FROM import_history")
            conn.commit()
        finally:
            conn.close()

    # ================================================================
    # 查询接口
    # ================================================================

    def get_all_quotas(self) -> list[dict]:
        """获取所有定额记录"""
        conn = self._connect(row_factory=True)
        try:
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM quotas ORDER BY id")
            rows = [dict(row) for row in cursor.fetchall()]
        finally:
            conn.close()
        return rows

    def get_quota_count(self) -> int:
        """获取定额总数"""
        conn = self._connect()
        try:
            cursor = conn.cursor()
            cursor.execute("SELECT COUNT(*) FROM quotas")
            count = cursor.fetchone()[0]
        finally:
            conn.close()
        return count

    def get_quota_by_id(self, quota_id: str) -> list[dict]:
        """根据定额编号查询"""
        conn = self._connect(row_factory=True)
        try:
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM quotas WHERE quota_id = ?", (quota_id,))
            rows = [dict(row) for row in cursor.fetchall()]
        finally:
            conn.close()
        return rows

    def search_by_keyword(self, keyword: str, limit: int = 20) -> list[dict]:
        """简单关键词搜索（SQLite LIKE查询）"""
        conn = self._connect(row_factory=True)
        try:
            cursor = conn.cursor()
            cursor.execute(
                "SELECT * FROM quotas WHERE name LIKE ? OR search_text LIKE ? LIMIT ?",
                (f"%{keyword}%", f"%{keyword}%", limit)
            )
            rows = [dict(row) for row in cursor.fetchall()]
        finally:
            conn.close()
        return rows

    def get_chapters(self) -> list[str]:
        """获取所有章节名称"""
        conn = self._connect()
        try:
            cursor = conn.cursor()
            cursor.execute("SELECT DISTINCT chapter FROM quotas ORDER BY chapter")
            chapters = [row[0] for row in cursor.fetchall()]
        finally:
            conn.close()
        return chapters

    def get_specialties(self) -> list[str]:
        """获取所有专业大类名称（如"安装"、"土建"、"市政"）"""
        conn = self._connect()
        try:
            cursor = conn.cursor()
            cursor.execute("SELECT DISTINCT specialty FROM quotas ORDER BY specialty")
            specialties = [row[0] for row in cursor.fetchall()]
        finally:
            conn.close()
        return specialties

    def get_chapters_by_specialty(self, specialty: str) -> list[str]:
        """获取指定专业下的所有章节名称

        参数:
            specialty: 专业名称，如"安装"
        """
        conn = self._connect()
        try:
            cursor = conn.cursor()
            cursor.execute(
                "SELECT DISTINCT chapter FROM quotas WHERE specialty = ? ORDER BY chapter",
                (specialty,)
            )
            chapters = [row[0] for row in cursor.fetchall()]
        finally:
            conn.close()
        return chapters

    def get_quotas_by_chapter(self, chapter: str, limit: int = 500) -> list[dict]:
        """获取指定章节下的所有定额条目

        参数:
            chapter: 章节名称
            limit: 最大返回数量（默认500，一个章节通常不超过几百条）
        """
        conn = self._connect(row_factory=True)
        try:
            cursor = conn.cursor()
            cursor.execute(
                "SELECT * FROM quotas WHERE chapter = ? ORDER BY quota_id LIMIT ?",
                (chapter, limit)
            )
            rows = [dict(row) for row in cursor.fetchall()]
        finally:
            conn.close()
        return rows

    def search_by_keywords(self, keywords: str, chapter: str = None,
                           book: str = None, limit: int = 50) -> list[dict]:
        """多关键词AND搜索（空格分隔的多个关键词，必须全部包含才返回）

        参数:
            keywords: 搜索文本，多个词用空格分隔（如"室外 镀锌钢管"）
            chapter: 限定在某个章节内搜索（可选）
            book: 限定在某个大册内搜索（可选，如"C10"）
            limit: 最大返回数量
        """
        # 把输入按空格拆分成多个关键词
        word_list = keywords.strip().split()
        if not word_list:
            return []

        # 构建SQL：每个关键词都必须出现在name或search_text中（AND逻辑）
        conditions = []
        params = []
        for word in word_list:
            conditions.append("(name LIKE ? OR search_text LIKE ?)")
            params.extend([f"%{word}%", f"%{word}%"])

        sql = f"SELECT * FROM quotas WHERE {' AND '.join(conditions)}"

        # 如果指定了章节，加上章节过滤
        if chapter:
            sql += " AND chapter = ?"
            params.append(chapter)

        # 如果指定了大册，加上大册过滤
        if book:
            sql += " AND book = ?"
            params.append(book)

        sql += f" ORDER BY quota_id LIMIT ?"
        params.append(limit)

        conn = self._connect(row_factory=True)
        try:
            cursor = conn.cursor()
            cursor.execute(sql, params)
            rows = [dict(row) for row in cursor.fetchall()]
        finally:
            conn.close()
        return rows

    def get_stats(self) -> dict:
        """获取数据库统计信息"""
        conn = self._connect()
        try:
            cursor = conn.cursor()

            cursor.execute("SELECT COUNT(*) FROM quotas")
            total = cursor.fetchone()[0]

            cursor.execute("SELECT COUNT(DISTINCT chapter) FROM quotas")
            chapters = cursor.fetchone()[0]

            cursor.execute("SELECT COUNT(DISTINCT specialty) FROM quotas")
            specialties = cursor.fetchone()[0]

            # 各参数非空的数量
            cursor.execute("SELECT COUNT(*) FROM quotas WHERE dn IS NOT NULL")
            with_dn = cursor.fetchone()[0]

            cursor.execute("SELECT COUNT(*) FROM quotas WHERE cable_section IS NOT NULL")
            with_section = cursor.fetchone()[0]
        finally:
            conn.close()

        return {
            "total": total,
            "chapters": chapters,
            "specialties": specialties,
            "with_dn": with_dn,
            "with_cable_section": with_section,
        }

    # ================================================================
    # 大册（book）相关方法
    # ================================================================

    def upgrade_add_book_field(self):
        """
        升级：给已有数据添加 book 字段（从定额编号前缀提取）

        从 quota_id 提取册号，例如：
          C10-1-5  → book="C10"
          C4-8-3   → book="C4"
          C1-1-100 → book="C1"

        如果 book 列不存在，先 ALTER TABLE 添加。
        """
        from src.specialty_classifier import get_book_from_quota_id

        conn = self._connect()
        try:
            cursor = conn.cursor()

            # 检查 book 列是否已存在
            cursor.execute("PRAGMA table_info(quotas)")
            columns = [col[1] for col in cursor.fetchall()]
            if "book" not in columns:
                cursor.execute("ALTER TABLE quotas ADD COLUMN book TEXT")
                logger.info("已添加 book 列到 quotas 表")

            # 创建索引
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_book ON quotas(book)")

            # 批量更新：从 quota_id 提取 book
            cursor.execute("SELECT id, quota_id FROM quotas")
            rows = cursor.fetchall()

            updated = 0
            for row_id, quota_id in rows:
                book = get_book_from_quota_id(quota_id)
                if book:
                    cursor.execute("UPDATE quotas SET book = ? WHERE id = ?",
                                   (book, row_id))
                    updated += 1

            conn.commit()
        except Exception:
            conn.rollback()
            raise
        finally:
            conn.close()
        logger.info(f"book字段更新完成: {updated}/{len(rows)} 条定额已标记册号")

    def get_books(self) -> list[dict]:
        """
        获取所有大册的列表（含每册定额数量）

        返回:
            [{"code": "C1", "name": "机械设备安装", "count": 3232}, ...]
        """
        from src.specialty_classifier import BOOKS

        conn = self._connect()
        try:
            cursor = conn.cursor()

            # 统计每册的定额数量
            cursor.execute("""
                SELECT book, COUNT(*) as cnt
                FROM quotas
                WHERE book IS NOT NULL
                GROUP BY book
                ORDER BY book
            """)
            counts = {row[0]: row[1] for row in cursor.fetchall()}
        finally:
            conn.close()

        result = []
        for code, info in BOOKS.items():
            result.append({
                "code": code,
                "name": info["name"],
                "count": counts.get(code, 0),
            })

        return result

    def get_chapters_by_book(self, book: str) -> list[dict]:
        """
        获取指定大册下的所有章节（含每章定额数量）

        参数:
            book: 大册编号（如"C10"）

        返回:
            [{"chapter": "081_Ⅰ 室外管道", "count": 138}, ...]
        """
        conn = self._connect()
        try:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT chapter, COUNT(*) as cnt
                FROM quotas
                WHERE book = ?
                GROUP BY chapter
                ORDER BY chapter
            """, (book,))
            result = [{"chapter": row[0], "count": row[1]} for row in cursor.fetchall()]
        finally:
            conn.close()
        return result


# ================================================================
# 工具函数
# ================================================================

def detect_specialty_from_excel(excel_path: str) -> str:
    """
    从Excel的D列（工作类型）自动识别specialty

    读取所有Sheet的前30行D列值，取出现次数最多的作为该文件的specialty。
    这是最可靠的识别方式，不依赖文件名或编号前缀。
    """
    import openpyxl
    try:
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        try:
            type_counts = {}
            for sheet_name in wb.sheetnames:  # 扫描所有Sheet
                ws = wb[sheet_name]
                for j, row in enumerate(ws.iter_rows(min_row=1, max_row=30, values_only=True)):
                    if row and len(row) > 3 and row[3]:
                        val = str(row[3]).strip()
                        if val and val not in ("工作类型", "类型", "类别"):  # 跳过表头
                            type_counts[val] = type_counts.get(val, 0) + 1
        finally:
            wb.close()

        if type_counts:
            return max(type_counts, key=type_counts.get)
    except Exception as e:
        logger.warning(f"自动识别specialty失败({e})，从文件名推断")

    # 兜底：从文件名推断（覆盖更多关键词）
    name = Path(excel_path).stem
    keyword_map = {
        "安装": "安装",
        "建筑": "土建",
        "装饰": "土建",
        "土建": "土建",
        "市政": "市政",
        "园林": "园林",
        "仿古": "仿古",
        "修缮": "修缮",
        "电气": "安装",
        "给排水": "安装",
        "暖通": "安装",
        "消防": "安装",
        "智能化": "安装",
        "通风": "安装",
    }
    for keyword, specialty in keyword_map.items():
        if keyword in name:
            return specialty
    return "未知"


# ================================================================
# 命令行入口：直接运行此文件可导入定额
# ================================================================

if __name__ == "__main__":
    logger.info("=" * 60)
    logger.info("开始导入定额数据")
    logger.info("=" * 60)

    db = QuotaDB()

    # 扫描省份目录下所有xlsx文件，自动识别specialty并导入
    quota_dir = config.get_quota_data_dir(db.province)
    if not quota_dir.exists():
        # 兼容旧目录结构：如果省份目录不存在，回退到QUOTA_DATA_DIR
        quota_dir = config.QUOTA_DATA_DIR
        logger.warning(f"省份目录不存在，使用旧目录: {quota_dir}")

    xlsx_files = list(quota_dir.glob("*.xlsx"))
    if not xlsx_files:
        logger.error(f"目录下没有xlsx文件: {quota_dir}")
        logger.error(f"请将广联达导出的定额Excel放到: {quota_dir}")
    else:
        total = 0
        cleared_specialties = set()  # 记录已清理旧数据的specialty，避免重复清理导致数据丢失
        for xlsx_file in sorted(xlsx_files):
            # 从D列自动识别specialty
            specialty = detect_specialty_from_excel(str(xlsx_file))
            # 同一specialty的第一个文件清除旧数据，后续文件追加
            is_first = specialty not in cleared_specialties
            cleared_specialties.add(specialty)
            logger.info(f"导入: {xlsx_file.name} → specialty='{specialty}' ({'清除旧数据' if is_first else '追加'})")
            count = db.import_excel(str(xlsx_file), specialty=specialty,
                                    clear_existing=is_first)
            total += count
            logger.info(f"  完成: {count}条")

        # 打印统计信息
        stats = db.get_stats()
        logger.info(f"\n数据库统计:")
        logger.info(f"  总记录数: {stats['total']}")
        logger.info(f"  章节数: {stats['chapters']}")
        logger.info(f"  专业数: {stats['specialties']}")

