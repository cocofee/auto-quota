"""
工程量清单读取模块。

目标：
1. 兼容标准清单 Excel。
2. 兼容手工/非标清单 Excel。
3. 兼容工程量汇总表这类“半结构化输入”，把它压成可进入套定额的标准 bill item。
"""

from __future__ import annotations

import os
import re
import tempfile
from pathlib import Path

import openpyxl
from loguru import logger

from src.excel_compat import detect_excel_file_format
from src.text_parser import parser as text_parser


def is_xls_format(file_path) -> bool:
    """检测文件实际是否为旧版 xls。"""
    try:
        return detect_excel_file_format(file_path).actual_format == "xls"
    except Exception:
        return False


def _is_quota_code(code: str) -> bool:
    """判断是否为定额编号。"""
    if not isinstance(code, str):
        return False
    c = code.strip()
    if not c:
        return False
    core = c[:-1] if c.endswith("换") else c
    return bool(re.match(r"^[A-Za-z]?\d{1,2}-\d+", core)) or bool(
        re.match(r"^[A-Za-z]{1,2}\d{4,}$", core)
    )


def _is_material_code(code: str) -> bool:
    """判断是否为材料/主材编码。"""
    if not isinstance(code, str):
        return False
    c = code.strip()
    if not c:
        return False
    if re.match(r"^CL\d", c, re.IGNORECASE):
        return True
    if re.match(r"^ZCGL\d", c, re.IGNORECASE):
        return True
    if "Z@" in c:
        return True
    if c.startswith("补充主材"):
        return True
    if "@" in c:
        return True
    if re.fullmatch(r"\d{7,8}", c):
        return True
    return False


class BillReader:
    """工程量清单读取器。"""

    COLUMN_PATTERNS = {
        "index": ["序号"],
        "code": ["项目编码", "编码", "清单编码", "子目编码"],
        "name": [
            "项目名称",
            "名称",
            "清单名称",
            "货物名称",
            "设备名称",
            "项目内容",
            "子目名称",
        ],
        "description": [
            "项目特征",
            "特征描述",
            "项目特征描述",
            "特征",
            "子目特征描述",
            "子目特征",
            "工作内容及范围",
            "工作内容",
            "规格",
            "规格型号",
            "技术参数",
        ],
        "unit": ["计量单位", "单位", "计量\n单位"],
        "quantity": ["工程量", "工程数量", "暂估数量", "数量"],
    }

    SUMMARY_COLUMN_PATTERNS = {
        "calc_item": ["计算项目"],
        "system_type": ["系统类型"],
        "type": ["类型"],
        "name": ["名称"],
        "material": ["材质"],
        "material_conn": ["材质-连接方式", "刷油/保温材质-厚度/保护层"],
        "spec": ["规格型号-类型", "规格型号", "规格"],
        "laying": ["敷设方式", "连接方式"],
        "qty_name": ["工程量名称"],
        "unit": ["单位"],
        "quantity": ["工程量"],
    }

    SKIP_SHEET_KEYWORDS = [
        "报告",
        "成果",
        "统计",
        "封面",
        "目录",
        "说明",
        "分析",
        "对比",
        "审核",
        "签章",
    ]

    BILL_SHEET_KEYWORDS = [
        "分部分项",
        "清单",
        "工程量汇总表",
        "工程量表",
        "汇总表",
    ]

    @staticmethod
    def _normalize_header_text(text: str) -> str:
        if text is None:
            return ""
        normalized = str(text).strip().replace("\n", "").replace("\r", "")
        normalized = normalized.replace("（", "(").replace("）", ")")
        normalized = re.sub(r"\s+", "", normalized)
        return normalized

    @classmethod
    def _match_header_pattern(cls, cell_text: str, pattern: str) -> int:
        norm_cell = cls._normalize_header_text(cell_text)
        norm_pattern = cls._normalize_header_text(pattern)
        if not norm_cell or not norm_pattern:
            return 0
        # 短词必须精确匹配，避免“名称”误命中“工程量名称”。
        if len(norm_pattern) <= 2:
            return len(norm_pattern) if norm_cell == norm_pattern else 0
        if norm_cell == norm_pattern:
            return len(norm_pattern) + 100
        if norm_pattern in norm_cell:
            return len(norm_pattern)
        return 0

    def read_excel(self, file_path: str, sheet_name: str = None) -> list[dict]:
        """读取清单 Excel。"""
        file_path = Path(file_path)
        if not file_path.exists():
            raise FileNotFoundError(f"清单文件不存在: {file_path}")

        logger.info(f"读取清单文件: {file_path}")

        temp_xlsx_path = None
        actual_path = file_path
        if is_xls_format(file_path):
            if file_path.suffix.lower() != ".xls":
                logger.info("  文件后缀是 xlsx 但实际是 xls，自动转换")
            try:
                temp_xlsx_path = self._convert_xls_to_xlsx(file_path)
                actual_path = Path(temp_xlsx_path)
            except Exception as e:
                logger.error(f"  .xls 转换失败: {e}")
                raise ValueError(f"无法读取 .xls 文件: {file_path}。转换失败: {e}")

        wb = openpyxl.load_workbook(str(actual_path), read_only=True, data_only=True)
        all_items = []
        try:
            sheets_to_read = [sheet_name] if sheet_name else self._filter_bill_sheets(wb.sheetnames)
            for sn in sheets_to_read:
                if sn not in wb.sheetnames:
                    logger.warning(f"Sheet '{sn}' 不存在，跳过")
                    continue
                ws = wb[sn]
                items = self._read_sheet(ws, sn)
                if items:
                    all_items.extend(items)
                    logger.info(f"  Sheet '{sn}': 读取 {len(items)} 条清单项")
        finally:
            wb.close()
            if temp_xlsx_path:
                try:
                    Path(temp_xlsx_path).unlink(missing_ok=True)
                except Exception:
                    pass

        if not all_items:
            logger.warning("未读取到任何清单项目，请检查文件格式")

        logger.info(f"清单读取完成: 共 {len(all_items)} 条项目")
        return all_items

    def _convert_xls_to_xlsx(self, xls_path: Path) -> str:
        """把 xls 转成临时 xlsx。"""
        import xlrd

        xls_wb = xlrd.open_workbook(str(xls_path))
        temp_dir = Path(__file__).parent.parent / "output" / "temp"
        temp_dir.mkdir(parents=True, exist_ok=True)
        fd, temp_path = tempfile.mkstemp(suffix=".xlsx", prefix="xls_convert_", dir=str(temp_dir))
        os.close(fd)

        try:
            xlsx_wb = openpyxl.Workbook()
            xlsx_wb.remove(xlsx_wb.active)
            for sheet_idx in range(xls_wb.nsheets):
                xls_sheet = xls_wb.sheet_by_index(sheet_idx)
                xlsx_sheet = xlsx_wb.create_sheet(title=xls_sheet.name)
                for row_idx in range(xls_sheet.nrows):
                    for col_idx in range(xls_sheet.ncols):
                        cell = xls_sheet.cell(row_idx, col_idx)
                        value = cell.value
                        if cell.ctype == 3:
                            try:
                                value = xlrd.xldate_as_datetime(value, xls_wb.datemode)
                            except Exception:
                                pass
                        if value is not None and value != "":
                            xlsx_sheet.cell(row=row_idx + 1, column=col_idx + 1, value=value)
            xlsx_wb.save(temp_path)
        except Exception:
            Path(temp_path).unlink(missing_ok=True)
            raise
        finally:
            xls_wb.release_resources()

        return temp_path

    def _filter_bill_sheets(self, sheet_names: list[str]) -> list[str]:
        """过滤明显非业务 sheet。"""
        preferred = [
            sn
            for sn in sheet_names
            if any(kw in sn for kw in self.BILL_SHEET_KEYWORDS)
            and not any(kw in sn for kw in self.SKIP_SHEET_KEYWORDS)
        ]
        if preferred:
            logger.info(f"  识别到候选 Sheet: {preferred}")
            return preferred

        filtered = []
        for sn in sheet_names:
            if any(kw in sn for kw in self.SKIP_SHEET_KEYWORDS):
                logger.debug(f"  跳过非清单 Sheet: '{sn}'")
                continue
            filtered.append(sn)
        return filtered if filtered else sheet_names

    def get_sheet_info(self, file_path: str) -> list[dict]:
        """获取每个 sheet 的可读性信息。"""
        file_path = Path(file_path)
        if not file_path.exists():
            raise FileNotFoundError(f"文件不存在: {file_path}")

        actual_path = file_path
        temp_xlsx_path = None
        if is_xls_format(file_path):
            temp_xlsx_path = self._convert_xls_to_xlsx(file_path)
            actual_path = Path(temp_xlsx_path)

        wb = openpyxl.load_workbook(str(actual_path), read_only=True, data_only=True)
        result = []
        try:
            for sn in wb.sheetnames:
                ws = wb[sn]
                header_rows = []
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if i >= 20:
                        break
                    header_rows.append(row)
                col_map, _ = self._detect_columns(header_rows)
                summary_map, _ = self._detect_summary_columns(header_rows)
                matched_count = max(len(col_map), len(summary_map))
                is_bill = bool(summary_map) or ("name" in col_map and len(col_map) >= 2)
                result.append(
                    {
                        "name": sn,
                        "is_bill": is_bill,
                        "matched_headers": matched_count,
                    }
                )
        finally:
            wb.close()
            if temp_xlsx_path:
                try:
                    Path(temp_xlsx_path).unlink(missing_ok=True)
                except Exception:
                    pass
        return result

    def _read_sheet(self, ws, sheet_name: str) -> list[dict]:
        header_rows = []
        all_rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            all_rows.append(row)
            if i < 20:
                header_rows.append(row)

        if not all_rows:
            return []

        summary_col_map, summary_header_row_idx = self._detect_summary_columns(header_rows)
        if summary_col_map:
            return self._read_summary_sheet(all_rows, summary_col_map, summary_header_row_idx, sheet_name)

        col_map, header_row_idx = self._detect_columns(header_rows)
        if not col_map:
            col_map, header_row_idx = self._try_simple_format(header_rows)

        if not col_map or "name" not in col_map:
            logger.debug(f"  Sheet '{sheet_name}': 未检测到清单格式，跳过")
            return []

        logger.debug(f"  列映射: {col_map}, 表头行: {header_row_idx}")

        items = []
        current_section = ""
        sheet_bill_seq = 0

        for i, row in enumerate(all_rows):
            if i <= header_row_idx:
                continue

            item = self._parse_bill_row(row, col_map, sheet_name, current_section, source_row=i + 1)
            if item is None:
                section = self._detect_section_header(row, col_map)
                if section:
                    current_section = section
                continue

            sheet_bill_seq += 1
            item["sheet_bill_seq"] = sheet_bill_seq
            items.append(item)

        self._extract_materials_from_rows(items, all_rows, col_map, header_row_idx)
        return items

    def _extract_materials_from_rows(self, items: list[dict], all_rows: list,
                                     col_map: dict, header_row_idx: int):
        """从原始行中提取主材行，挂到对应清单项。"""
        if not items:
            return

        row_to_item = {item["source_row"]: item for item in items if item.get("source_row")}
        idx_col = col_map.get("index", 0)
        code_col = col_map.get("code", 1)
        name_col = col_map.get("name", 2)
        unit_col = col_map.get("unit")
        qty_col = col_map.get("quantity")
        current_item = None

        for i, row in enumerate(all_rows):
            if i <= header_row_idx or not row:
                continue
            source_row = i + 1

            def _get(col_idx):
                if col_idx is not None and col_idx < len(row):
                    v = row[col_idx]
                    return str(v).strip() if v is not None else ""
                return ""

            a_val = _get(idx_col)
            b_val = _get(code_col)
            c_val = _get(name_col)

            if source_row in row_to_item:
                current_item = row_to_item[source_row]
                current_item.setdefault("source_materials", [])
                continue

            if current_item is None:
                continue

            if a_val and (a_val.isdigit() or re.fullmatch(r"\d+\.0+", a_val)):
                current_item = None
                continue

            if not a_val and b_val:
                if _is_quota_code(b_val):
                    continue

                is_material = _is_material_code(b_val) or bool(re.fullmatch(r"\d{7,8}", b_val))
                if is_material and c_val:
                    mat_unit = _get(unit_col) if unit_col is not None else ""
                    mat_qty_str = _get(qty_col) if qty_col is not None else ""
                    mat_qty = None
                    if mat_qty_str:
                        try:
                            mat_qty = float(mat_qty_str.replace(",", ""))
                        except (ValueError, TypeError):
                            pass
                    current_item["source_materials"].append(
                        {"code": b_val, "name": c_val, "unit": mat_unit, "qty": mat_qty}
                    )

    def _detect_columns(self, header_rows: list) -> tuple[dict, int]:
        """检测标准/非标清单表头。"""
        for row_idx, row in enumerate(header_rows):
            if not row:
                continue

            col_map = {}
            field_scores = {}
            for col_idx, cell_value in enumerate(row):
                if cell_value is None:
                    continue
                cell_text = str(cell_value)
                if len(self._normalize_header_text(cell_text)) > 20:
                    continue

                for field, patterns in self.COLUMN_PATTERNS.items():
                    for pattern in patterns:
                        score = self._match_header_pattern(cell_text, pattern)
                        if score > field_scores.get(field, 0):
                            col_map[field] = col_idx
                            field_scores[field] = score

            if "name" in col_map and len(col_map) >= 2:
                return col_map, row_idx

        return {}, -1

    def _detect_summary_columns(self, header_rows: list) -> tuple[dict, int]:
        """识别工程量汇总表这类半结构化输入。"""
        for row_idx, row in enumerate(header_rows):
            if not row:
                continue

            col_map = {}
            field_scores = {}
            for col_idx, cell_value in enumerate(row):
                if cell_value is None:
                    continue
                cell_text = str(cell_value)
                for field, patterns in self.SUMMARY_COLUMN_PATTERNS.items():
                    for pattern in patterns:
                        score = self._match_header_pattern(cell_text, pattern)
                        if score > field_scores.get(field, 0):
                            col_map[field] = col_idx
                            field_scores[field] = score

            if (
                "quantity" in col_map
                and "unit" in col_map
                and "qty_name" in col_map
                and ("name" in col_map or "type" in col_map or "calc_item" in col_map)
            ):
                return col_map, row_idx

        return {}, -1

    def _read_summary_sheet(self, all_rows: list, col_map: dict,
                            header_row_idx: int, sheet_name: str) -> list[dict]:
        """把工程量汇总表压成标准 bill item。"""
        items = []
        seen_keys = set()

        def get_val(row, field):
            idx = col_map.get(field)
            if idx is None or idx >= len(row) or row[idx] is None:
                return ""
            return str(row[idx]).strip()

        for i, row in enumerate(all_rows):
            if i <= header_row_idx or not row:
                continue

            qty_name = get_val(row, "qty_name")
            unit = get_val(row, "unit")
            quantity_str = get_val(row, "quantity")
            if not unit or not quantity_str:
                continue

            # 汇总表通常是一条明细 + 一条合计，先保留非合计行。
            if "合计" in qty_name:
                continue

            try:
                quantity = float(quantity_str.replace(",", ""))
            except (ValueError, TypeError):
                continue

            calc_item = get_val(row, "calc_item")
            item_type = get_val(row, "type")
            item_name = get_val(row, "name")
            system_type = get_val(row, "system_type")
            material = get_val(row, "material")
            material_conn = get_val(row, "material_conn")
            spec = get_val(row, "spec")
            laying = get_val(row, "laying")

            name = item_name or item_type or calc_item
            if not name:
                continue

            desc_parts = []
            for label, value in [
                ("系统", system_type),
                ("计算项目", calc_item if calc_item and calc_item != name else ""),
                ("类型", item_type if item_type and item_type != name else ""),
                ("材质", material),
                ("材质/连接", material_conn),
                ("规格", spec),
                ("敷设/连接", laying),
                ("工程量名称", qty_name),
            ]:
                clean_value = (value or "").strip()
                if clean_value and clean_value != "<空>":
                    desc_parts.append(f"{label}:{clean_value}")
            description = " / ".join(desc_parts)

            dedupe_key = (sheet_name, name, description, unit, quantity)
            if dedupe_key in seen_keys:
                continue
            seen_keys.add(dedupe_key)

            search_text = text_parser.build_search_text(name, description)
            params = text_parser.parse(f"{name} {description}")
            items.append(
                {
                    "index": "",
                    "code": "",
                    "name": name,
                    "description": description,
                    "unit": unit,
                    "quantity": quantity,
                    "search_text": search_text,
                    "params": params,
                    "sheet_name": sheet_name,
                    "section": system_type or "",
                    "source_row": i + 1,
                }
            )

        return items

    def _try_simple_format(self, header_rows: list) -> tuple[dict, int]:
        """识别没有标准表头的简单表格。"""
        for row_idx, row in enumerate(header_rows):
            if not row or len(row) < 2:
                continue

            first_val = str(row[0] or "").strip()
            second_val = str(row[1] or "").strip()
            if re.match(r"^[A-Za-z]?\d", first_val) and len(second_val) > 1:
                col_map = {"code": 0, "name": 1}
                if len(row) > 2 and row[2]:
                    col_map["unit"] = 2
                if len(row) > 3 and row[3]:
                    try:
                        float(str(row[3]).replace(",", ""))
                        col_map["quantity"] = 3
                    except (ValueError, TypeError):
                        pass
                return col_map, row_idx - 1

        return {}, -1

    def _parse_bill_row(self, row, col_map: dict, sheet_name: str,
                        section: str, source_row: int = None) -> dict | None:
        """解析单行清单。"""
        if not row:
            return None

        def get_val(field):
            idx = col_map.get(field)
            if idx is not None and idx < len(row) and row[idx] is not None:
                return str(row[idx]).strip()
            return ""

        name = get_val("name")
        code = get_val("code")
        description = get_val("description")
        unit = get_val("unit")
        quantity_str = get_val("quantity")

        if not name:
            return None

        skip_keywords = [
            "合计",
            "小计",
            "序号",
            "项目名称",
            "总计",
            "分部分项",
            "措施项目",
            "工程概况",
            "工程名称",
            "总说明",
            "说明",
            "人工小计",
            "材料小计",
            "机械小计",
            "管理费",
            "利润",
            "税金",
            "规费",
            "安全文明",
        ]
        if any(kw in name for kw in skip_keywords):
            return None

        if re.match(r"^(长度|数量|重量|面积|体积|周长|高度|宽度|厚度)\s*[\(（]", name):
            return None

        if len(name) > 100:
            return None

        quantity = None
        if quantity_str:
            try:
                quantity = float(quantity_str.replace(",", "").strip())
            except (ValueError, TypeError):
                pass

        has_code = bool(code and re.search(r"\d", code))
        has_quantity = quantity is not None
        has_unit = bool(unit)
        if not has_code and not has_quantity and not has_unit:
            return None

        if code and re.match(r"^[A-Z]\.\d", code) and not has_quantity and not has_unit and not description:
            return None

        if _is_quota_code(code):
            return None

        if _is_material_code(code):
            return None

        search_text = text_parser.build_search_text(name, description)
        params = text_parser.parse(f"{name} {description}")
        index_str = get_val("index")

        return {
            "index": index_str,
            "code": code,
            "name": name,
            "description": description,
            "unit": unit,
            "quantity": quantity,
            "search_text": search_text,
            "params": params,
            "sheet_name": sheet_name,
            "section": section,
            "source_row": source_row,
        }

    def _detect_section_header(self, row, col_map: dict) -> str | None:
        """检测分部/章节标题行。"""
        if not row:
            return None

        name_text = ""
        if "name" in col_map:
            name_idx = col_map["name"]
            if name_idx < len(row) and row[name_idx]:
                name_text = str(row[name_idx]).strip()

        has_index = False
        if "index" in col_map:
            idx_val = row[col_map["index"]] if col_map["index"] < len(row) else None
            if idx_val is not None and str(idx_val).strip():
                has_index = True

        if not has_index:
            for field in ["code", "name"]:
                idx = col_map.get(field)
                if idx is not None and idx < len(row) and row[idx]:
                    text = str(row[idx]).strip()
                    if len(text) >= 2 and not re.match(r"^\d", text):
                        return text
        else:
            if name_text and self._looks_like_section_title(name_text):
                has_unit = False
                has_qty = False
                if "unit" in col_map:
                    u_idx = col_map["unit"]
                    if u_idx < len(row) and row[u_idx]:
                        has_unit = bool(str(row[u_idx]).strip())
                if "quantity" in col_map:
                    q_idx = col_map["quantity"]
                    if q_idx < len(row) and row[q_idx]:
                        try:
                            float(str(row[q_idx]).replace(",", "").strip())
                            has_qty = True
                        except (ValueError, TypeError):
                            pass
                if not has_unit and not has_qty:
                    return name_text

        return None

    @staticmethod
    def _looks_like_section_title(text: str) -> bool:
        if len(text) < 2:
            return False
        if re.match(r"^[一二三四五六七八九十百]+[、\s]", text):
            return True
        if re.match(r"^第[一二三四五六七八九十百\d]+[部章节]", text):
            return True
        return False


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="读取清单 Excel 并预览前 N 条结果")
    parser.add_argument("input", help="清单 Excel 路径")
    parser.add_argument("--limit", type=int, default=20, help="预览条数，默认 20")
    args = parser.parse_args()

    reader = BillReader()
    items = reader.read_excel(args.input)

    logger.info(f"\n前 {min(len(items), args.limit)} 条清单项:")
    for item in items[: args.limit]:
        section = (item.get("section") or "")[:10]
        logger.info(
            f"  [{item.get('index', '')}] {item.get('code', '')} | {item.get('name', '')[:30]} | "
            f"{item.get('unit', '')} | {item.get('quantity', '')} | 分部:{section}"
        )
        if item.get("description"):
            desc_line1 = item["description"].split("\n")[0][:50]
            logger.info(f"    特征: {desc_line1}")
        if item.get("params"):
            logger.info(f"    参数: {item['params']}")
