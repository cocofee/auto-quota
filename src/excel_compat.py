from __future__ import annotations

import shutil
import subprocess
from dataclasses import dataclass
from pathlib import Path


OLE_HEADER = b"\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1"
ZIP_HEADER = b"PK\x03\x04"


@dataclass(frozen=True)
class ExcelFormatInfo:
    declared_suffix: str
    actual_format: str
    normalized_suffix: str
    is_mislabeled: bool


@dataclass(frozen=True)
class ExcelConversionResult:
    method: str
    preserved_formatting: bool
    warning: str | None = None


def detect_excel_format_from_header(header: bytes, filename: str = "") -> ExcelFormatInfo:
    declared_suffix = Path(filename or "").suffix.lower()
    actual_format = "unknown"
    if header.startswith(ZIP_HEADER):
        actual_format = "xlsx"
    elif header.startswith(OLE_HEADER):
        actual_format = "xls"

    normalized_suffix = declared_suffix
    if actual_format in {"xls", "xlsx"}:
        normalized_suffix = f".{actual_format}"

    return ExcelFormatInfo(
        declared_suffix=declared_suffix,
        actual_format=actual_format,
        normalized_suffix=normalized_suffix,
        is_mislabeled=declared_suffix in {".xls", ".xlsx"} and normalized_suffix != declared_suffix,
    )


def detect_excel_file_format(file_path: str | Path, filename: str | None = None) -> ExcelFormatInfo:
    path = Path(file_path)
    with open(path, "rb") as f:
        header = f.read(8)
    return detect_excel_format_from_header(header, filename or path.name)


def validate_excel_upload(filename: str, header: bytes) -> ExcelFormatInfo:
    info = detect_excel_format_from_header(header, filename)
    if info.declared_suffix not in {".xls", ".xlsx"}:
        raise ValueError(
            f"不支持的文件格式 '{info.declared_suffix}'，请上传 .xlsx 或 .xls 文件"
        )
    if info.actual_format == "unknown":
        raise ValueError("文件内容无法识别为 Excel，请上传真正的 .xlsx 或 .xls 文件")
    return info


def convert_excel_to_xlsx(
    input_path: str | Path,
    output_path: str | Path,
    prefer_preserve_format: bool = True,
) -> ExcelConversionResult:
    src = Path(input_path)
    dst = Path(output_path)
    info = detect_excel_file_format(src)
    if info.actual_format == "unknown":
        raise ValueError(f"无法识别 Excel 格式: {src}")

    if info.actual_format == "xlsx":
        if src.resolve() != dst.resolve():
            shutil.copy2(src, dst)
        return ExcelConversionResult(method="copy", preserved_formatting=True)

    warnings: list[str] = []
    if prefer_preserve_format:
        try:
            _convert_with_excel_com(src, dst)
            return ExcelConversionResult(method="excel_com", preserved_formatting=True)
        except Exception as exc:
            warnings.append(f"Excel COM 不可用: {exc}")

        try:
            _convert_with_soffice(src, dst)
            return ExcelConversionResult(method="soffice", preserved_formatting=True)
        except Exception as exc:
            warnings.append(f"LibreOffice 不可用: {exc}")

    _convert_with_xlrd_value_only(src, dst)
    warning = "; ".join(warnings) if warnings else None
    return ExcelConversionResult(
        method="xlrd_value_only",
        preserved_formatting=False,
        warning=warning,
    )


def ensure_openpyxl_input(
    input_path: str | Path,
    normalized_xlsx_path: str | Path,
) -> tuple[Path, ExcelConversionResult | None]:
    src = Path(input_path)
    info = detect_excel_file_format(src)
    if info.actual_format == "unknown":
        raise ValueError(f"无法识别 Excel 格式: {src}")
    if info.actual_format == "xlsx":
        return src, None

    dst = Path(normalized_xlsx_path)
    result = convert_excel_to_xlsx(src, dst, prefer_preserve_format=False)
    return dst, result


def _convert_with_excel_com(src: Path, dst: Path) -> None:
    import pythoncom
    import win32com.client

    pythoncom.CoInitialize()
    excel = None
    workbook = None
    try:
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        workbook = excel.Workbooks.Open(str(src.resolve()), UpdateLinks=False, ReadOnly=True)
        workbook.SaveAs(str(dst.resolve()), FileFormat=51)
    finally:
        if workbook is not None:
            workbook.Close(False)
        if excel is not None:
            excel.Quit()
        pythoncom.CoUninitialize()


def _convert_with_soffice(src: Path, dst: Path) -> None:
    soffice = shutil.which("soffice")
    if not soffice:
        raise FileNotFoundError("未找到 soffice")

    dst.parent.mkdir(parents=True, exist_ok=True)
    subprocess.run(
        [
            soffice,
            "--headless",
            "--convert-to",
            "xlsx",
            "--outdir",
            str(dst.parent),
            str(src),
        ],
        check=True,
        capture_output=True,
        text=True,
    )

    converted = dst.parent / f"{src.stem}.xlsx"
    if not converted.exists():
        raise FileNotFoundError(f"未找到转换结果: {converted}")
    if converted.resolve() != dst.resolve():
        if dst.exists():
            dst.unlink()
        converted.replace(dst)


def _convert_with_xlrd_value_only(src: Path, dst: Path) -> None:
    import openpyxl
    import xlrd

    xls_wb = xlrd.open_workbook(str(src))
    xlsx_wb = openpyxl.Workbook()
    xlsx_wb.remove(xlsx_wb.active)
    try:
        for sheet_idx in range(xls_wb.nsheets):
            xls_sheet = xls_wb.sheet_by_index(sheet_idx)
            xlsx_sheet = xlsx_wb.create_sheet(title=xls_sheet.name)
            for row_idx in range(xls_sheet.nrows):
                for col_idx in range(xls_sheet.ncols):
                    cell = xls_sheet.cell(row_idx, col_idx)
                    value = cell.value
                    if cell.ctype == 3:
                        try:
                            value = xlrd.xldate_as_datetime(value, xls_wb.datemode)
                        except Exception:
                            pass
                    if value is not None and value != "":
                        xlsx_sheet.cell(row=row_idx + 1, column=col_idx + 1, value=value)
        xlsx_wb.save(str(dst))
    finally:
        xls_wb.release_resources()
        xlsx_wb.close()
