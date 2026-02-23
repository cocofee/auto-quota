"""
两文件对比学习模块
功能：
1. 对比"原始输出Excel"和"用户修正后Excel"
2. 自动检测用户改了哪些定额
3. 将修正结果存入经验库（权威层）

使用场景：
    用户在广联达中打开我们的输出Excel → 修正不正确的定额 → 保存
    然后运行"导入修正.bat"，系统自动对比差异并学习

Excel格式约定：
    清单行：A列有序号（数字），B列是项目编码，C列是项目名称
    定额行：A列为空，B列是定额编号（如C10-1-10），C列是定额名称
    分部行：A列和B列都为空，C列有文字
"""

import re
from pathlib import Path

import openpyxl
from loguru import logger

import config
from src.text_parser import normalize_bill_text


class DiffLearner:
    """两文件对比学习器"""

    def diff_and_learn(self, original_path: str, corrected_path: str,
                       province: str = None) -> dict:
        """
        对比原始输出和修正后文件，自动学习差异

        参数:
            original_path: 原始输出Excel路径（系统生成的）
            corrected_path: 修正后Excel路径（用户在广联达里改过的）
            province: 省份（如"北京2024"），默认使用配置文件中的省份

        返回:
            学习统计结果:
            {
                "total": 总清单条数,
                "confirmed": 未修改的条数（用户确认了系统匹配正确）,
                "corrected": 修改的条数（用户换了定额）,
                "skipped": 跳过的条数（无法对比或原始就没匹配到）,
                "details": 修正详情列表
            }
        """
        province = province or config.get_current_province()

        # 读取两个文件的清单→定额映射
        logger.info(f"读取原始文件: {original_path}")
        original_mapping = self._read_bill_quota_mapping(original_path)
        logger.info(f"  原始文件: {len(original_mapping)} 条清单")

        logger.info(f"读取修正文件: {corrected_path}")
        corrected_mapping = self._read_bill_quota_mapping(corrected_path)
        logger.info(f"  修正文件: {len(corrected_mapping)} 条清单")

        if not original_mapping:
            logger.error("原始文件中未读取到清单数据")
            return {"total": 0, "confirmed": 0, "corrected": 0,
                    "skipped": 0, "details": []}

        # 初始化经验库
        from src.experience_db import ExperienceDB
        experience_db = ExperienceDB()

        # 逐条对比
        confirmed = 0  # 未修改（确认正确）
        corrected = 0  # 已修改（用户纠正）
        skipped = 0    # 跳过
        details = []   # 修正详情

        for idx, orig_item in enumerate(original_mapping):
            bill_name = orig_item["bill_name"]
            bill_desc = orig_item.get("bill_desc", "")
            bill_code = orig_item.get("bill_code", "")
            bill_unit = orig_item.get("bill_unit", "")
            orig_quotas = orig_item["quota_ids"]

            # 在修正文件中找对应的清单项（按序号对应）
            if idx >= len(corrected_mapping):
                skipped += 1
                continue

            corr_item = corrected_mapping[idx]
            corr_quotas = corr_item["quota_ids"]

            # 原始就没匹配到定额的 → 看修正版有没有加
            if not orig_quotas and not corr_quotas:
                skipped += 1
                continue

            # 对比定额编号列表是否相同
            if set(orig_quotas) == set(corr_quotas):
                # 未修改 → 用户确认了系统匹配正确
                confirmed += 1

                # 存入经验库（user_confirmed，权威层）
                if corr_quotas:
                    normalized_text = normalize_bill_text(bill_name, bill_desc)
                    corr_names = corr_item.get("quota_names", [])
                    record_id = experience_db.add_experience(
                        bill_text=normalized_text,
                        quota_ids=corr_quotas,
                        quota_names=corr_names,
                        bill_name=bill_name,
                        bill_code=bill_code,
                        bill_unit=bill_unit,
                        source="user_confirmed",
                        confidence=90,
                        province=province,
                        notes="diff_learner自动对比确认",
                    )
                    if record_id <= 0:
                        logger.warning(
                            f"diff_learner确认写入被拦截: {normalized_text[:60]} -> {corr_quotas}"
                        )
            else:
                # 已修改 → 用户纠正了定额
                corrected += 1

                detail = {
                    "bill_name": bill_name,
                    "original_quotas": orig_quotas,
                    "corrected_quotas": corr_quotas,
                }
                details.append(detail)

                logger.info(
                    f"  修正: {bill_name[:30]} | "
                    f"{','.join(orig_quotas)} → {','.join(corr_quotas)}"
                )

                # 存入经验库（user_correction，权威层，置信度更高）
                if corr_quotas:
                    normalized_text = normalize_bill_text(bill_name, bill_desc)
                    corr_names = corr_item.get("quota_names", [])
                    record_id = experience_db.add_experience(
                        bill_text=normalized_text,
                        quota_ids=corr_quotas,
                        quota_names=corr_names,
                        bill_name=bill_name,
                        bill_code=bill_code,
                        bill_unit=bill_unit,
                        source="user_correction",
                        confidence=95,
                        province=province,
                        notes="diff_learner自动对比学习",
                    )
                    if record_id <= 0:
                        logger.warning(
                            f"diff_learner修正写入被拦截: {normalized_text[:60]} -> {corr_quotas}"
                        )

        total = len(original_mapping)
        result = {
            "total": total,
            "confirmed": confirmed,
            "corrected": corrected,
            "skipped": skipped,
            "details": details,
        }

        logger.info("=" * 50)
        logger.info("对比学习完成")
        logger.info(f"  清单总数: {total}")
        logger.info(f"  确认正确: {confirmed} ({confirmed * 100 // max(total, 1)}%)")
        logger.info(f"  用户纠正: {corrected} ({corrected * 100 // max(total, 1)}%)")
        logger.info(f"  跳过: {skipped}")
        logger.info("=" * 50)

        return result

    def _read_bill_quota_mapping(self, file_path: str) -> list[dict]:
        """
        从Excel中读取清单→定额的映射关系

        返回:
            列表，每项是一条清单及其下面的定额:
            {
                "bill_name": "镀锌钢管DN25",
                "bill_desc": "丝接",
                "bill_code": "030402011001",
                "bill_unit": "m",
                "quota_ids": ["C10-1-10", "C10-7-1"],
                "quota_names": ["管道安装 镀锌钢管DN25", "管卡安装 DN25"],
            }

        读取逻辑：
        - 遍历Excel所有行
        - A列有序号 → 清单行，开始新的一条记录
        - A列为空、B列有定额编号格式 → 定额行，归入上一条清单
        - 其他行跳过（分部标题、空行、表头等）
        """
        file_path = Path(file_path)
        if not file_path.exists():
            raise FileNotFoundError(f"文件不存在: {file_path}")

        wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
        try:
            # 找到有效的Sheet（优先找"匹配结果明细"，否则用第一个）
            ws = None
            for sn in wb.sheetnames:
                if "匹配" in sn or "明细" in sn or "结果" in sn:
                    ws = wb[sn]
                    break
            if ws is None:
                ws = wb[wb.sheetnames[0]]

            # 先检测表头和列映射
            rows = list(ws.iter_rows(values_only=True))
        finally:
            wb.close()

        header_row_idx, col_map = self._detect_header(rows)

        index_col = col_map.get("index", 0)
        code_col = col_map.get("code", 1)
        name_col = col_map.get("name", 2)
        desc_col = col_map.get("description", 3)
        unit_col = col_map.get("unit", 4)

        mapping = []
        current_bill = None  # 当前正在处理的清单项

        for row_idx, row in enumerate(rows):
            if row_idx <= header_row_idx:
                continue  # 跳过表头

            if not row:
                continue

            # 取各列值
            a_val = row[index_col] if index_col < len(row) else None
            b_val = row[code_col] if code_col < len(row) else None
            c_val = row[name_col] if name_col < len(row) else None
            d_val = row[desc_col] if desc_col < len(row) else None
            e_val = row[unit_col] if unit_col < len(row) else None

            # 判断行类型
            a_str = str(a_val).strip() if a_val is not None else ""
            b_str = str(b_val).strip() if b_val is not None else ""

            if a_str.isdigit():
                # 清单行：A列有序号
                # 先保存上一条清单
                if current_bill is not None:
                    mapping.append(current_bill)

                current_bill = {
                    "bill_name": str(c_val).strip() if c_val else "",
                    "bill_desc": str(d_val).strip() if d_val else "",
                    "bill_code": b_str,
                    "bill_unit": str(e_val).strip() if e_val else "",
                    "quota_ids": [],
                    "quota_names": [],
                }

            elif a_str == "" and self._is_quota_id(b_str):
                # 定额行：A列为空，B列是定额编号格式
                if current_bill is not None:
                    current_bill["quota_ids"].append(b_str)
                    quota_name = str(c_val).strip() if c_val else ""
                    current_bill["quota_names"].append(quota_name)

            # 其他行（分部标题、空行等）→ 跳过

        # 别忘了最后一条清单
        if current_bill is not None:
            mapping.append(current_bill)

        return mapping

    def _is_quota_id(self, text: str) -> bool:
        """
        判断一个字符串是否像定额编号

        定额编号格式举例：
        - C10-1-10  （北京安装定额）
        - C4-8-3    （北京电气定额）
        - 5-325     （四川定额）
        - D00003    （某些地区定额）

        关键特征：含字母开头 + 数字 + 横杠分隔
        """
        if not text:
            return False
        # 常见格式：C开头+数字+横杠（如C10-1-10）
        if re.match(r'^[A-Za-z]?\d{1,2}-\d+', text):
            return True
        # D+5位数字格式（如D00003）
        if re.match(r'^[A-Za-z]\d{4,}$', text):
            return True
        return False

    def _detect_header(self, rows):
        """
        检测表头行位置和列映射（和output_writer中的逻辑类似）

        返回: (header_row_idx, col_map)
        """
        patterns = {
            "index": ["序号"],
            "code": ["项目编码", "编码", "清单编码"],
            "name": ["项目名称", "名称", "清单名称"],
            "description": ["项目特征", "特征描述", "项目特征描述"],
            "unit": ["计量单位", "单位"],
            "quantity": ["工程量"],
        }

        for row_idx, row in enumerate(rows[:20]):
            if not row:
                continue
            col_map = {}
            for col_idx, cell in enumerate(row):
                if cell is None:
                    continue
                text = str(cell).strip().replace("\n", "")
                if len(text) > 15:
                    continue
                for field, keywords in patterns.items():
                    for kw in keywords:
                        if kw in text:
                            col_map[field] = col_idx
                            break

            if "name" in col_map and len(col_map) >= 2:
                return row_idx, col_map

        # 没找到表头，用默认值
        return 0, {"index": 0, "code": 1, "name": 2, "description": 3,
                   "unit": 4, "quantity": 5}


# 命令行入口：直接运行测试
if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(
        description="对比原始输出和修正后Excel，自动学习差异",
    )
    parser.add_argument("original", help="原始输出Excel文件路径")
    parser.add_argument("corrected", help="修正后Excel文件路径")
    parser.add_argument("--province", default=None,
                        help=f"省份（默认: {config.CURRENT_PROVINCE}）")

    args = parser.parse_args()

    learner = DiffLearner()
    result = learner.diff_and_learn(args.original, args.corrected,
                                     province=args.province)

    if result["corrected"] > 0:
        logger.info("\n修正详情:")
        for d in result["details"]:
            logger.info(
                f"  {d['bill_name'][:40]}: "
                f"{','.join(d['original_quotas'])} → "
                f"{','.join(d['corrected_quotas'])}"
            )
